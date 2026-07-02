# GROWTH OS 360 ACTION PATCH · MANAGEMENT + FINDER + PALETTE
import streamlit as st
import streamlit.components.v1 as st_components
import pandas as pd
import openpyxl
import math
import os
import glob
import shutil
import calendar
import re
import unicodedata
import html
import uuid
import json
import requests
from datetime import datetime, date, time, timedelta
from urllib.parse import quote_plus

# =========================
# CONFIG
# =========================

def get_app_period():
    today = date.today()
    quarter = (today.month - 1) // 3 + 1
    iso_week = today.isocalendar().week
    return f"Q{quarter} W{iso_week} {today.year}"


APP_PERIOD = get_app_period()
COMMENTS_FILE = "growth_os_comments.csv"
TEMPLATE_QUEUE_FILE = "growth_os_templates.csv"
ACQUISITION_TRACKER_FILE = "growth_os_acquisition_tracker.csv"
CAMPAIGN_WEEKLY_TRACKER_FILE = "growth_os_campaign_weekly_tracker.csv"
CHANGELOG_FILE = "growth_os_brand_changelog.csv"
DAY_QUEUE_CURSOR_FILE = "growth_os_day_queue_cursor.json"
CALL_QUALITY_HISTORY_FILE = "growth_os_call_quality_history.csv"
ROLEPLAY_OBJECTIONS_FILE = "growth_os_roleplay_objections.csv"
ROLEPLAY_HISTORY_FILE = "growth_os_roleplay_history.csv"
BRAND_LINKS_FILE = "growth_os_brand_links.csv"
CALL_HISTORY_FILE = "growth_os_call_history.csv"
BACKUP_FOLDER = "backups"

# ── Defaults de negocio ──────────────────────────────────────────────────────
# Estos valores son SOLO fallback. Si el Excel tiene una hoja "Config" con pares
# clave/valor (col A = clave, col B = valor), esos valores mandan. Así cada
# Farmer adapta Growth OS a su operación editando 6 celdas, sin tocar código.
# Claves reconocidas (case-insensitive):
#   farmer_name · farmer_role · portfolio_country · ars_per_usd · cop_per_usd
#   ads_revenue_target_usd · contacts_start_date (YYYY-MM-DD)
ARS_PER_USD = 1400
COP_PER_USD = 3900
ADS_REVENUE_TARGET_USD = 17574
CONTACTS_START_DATE = date(2026, 6, 1)

# Growth OS es agnóstico de país. Este filtro define qué país del Excel se carga
# en el portafolio activo. Dejar en "" desactiva el filtro y carga todas las filas.
PORTFOLIO_COUNTRY = "Argentina"

# Identidad del Farmer — usada en plantillas de outreach (firmas de email /
# WhatsApp), headers de páginas y detección de presentación en Call Quality.
FARMER_NAME = "Sabas Ramírez"
FARMER_ROLE = "Especialista en crecimiento de marcas digitales"

# Main visual identity requested by Sabas
# Slate / Neon Tangerine / Mint palette · blue background + white surfaces + tangerine as secondary accent
# Core: Space Indigo #1D2659 · Slate Indigo #4E63D9 · Neon Tangerine #FF8A3D · Laser Green #7ED321 · Soft Mint #DDFBCF
PALETTE = {
    # ── Tokens de identidad (light mode) ─────────────────────
    "noble_black":    "#EDEDEB",   # fondo base light (gris PDF)
    "wahoo":          "#FFFFFF",   # cards / superficies blancas
    "blue_estate":    "#1B3F8B",   # azul oscuro primario
    "burning_orange": "#FF7124",   # naranja tangerine — identidad dashboard
    "laser_green":    "#7ED321",   # verde lima — datos positivos
    "cinnamon_ice":   "#6B7280",   # texto secundario sobre fondo claro
    "pale_cashmere":  "#1A1A2E",   # texto principal sobre fondo claro

    # ── Scales derivadas ─────────────────────────────────────
    "wahoo_dark":     "#F5F5F3",   # hover cards (un pelo más oscuro que el fondo)
    "blue_soft":      "#3D64B8",   # blue_estate atenuado
    "blue_glow":      "#1B3F8B",   # intel highlight
    "laser_soft":     "#A8E05F",   # laser_green atenuado
    "laser_dim":      "rgba(126,211,33,0.12)",   # laser bg subtle
    "orange_dark":    "#D95A10",   # burning_orange oscuro
    "orange_soft":    "rgba(255,113,36,0.12)",   # burning_orange bg subtle
    "cashmere_dim":   "rgba(107,114,128,0.5)",   # texto muted
    "negative":       "#FF4D2E",   # rojo-naranja — datos negativos / alertas
    "negative_soft":  "rgba(255,77,46,0.12)",    # fondo alerta negativa

    # ── Sticker tokens (reemplaza glass) ─────────────────────
    "glass_bg":       "#FFFFFF",
    "glass_bg_hover": "#F5F5F3",
    "glass_border":   "rgba(0,0,0,0.08)",
    "glass_border_hover": "rgba(27,63,139,0.25)",

    # ── Alias backward-compat — no eliminar ──────────────────
    "space_indigo":   "#FFFFFF",
    "slate_indigo":   "#1B3F8B",
    "navy":           "#FFFFFF",
    "navy_dark":      "#F5F5F3",
    "navy_soft":      "#EEF2FF",
    "blue":           "#1B3F8B",
    "blue_mist":      "rgba(27,63,139,0.12)",
    "laser":          "#7ED321",
    "mint":           "#1A1A2E",
    "mint_soft":      "#1A1A2E",
    "mint_muted":     "#6B7280",
    "soft_mint":      "#1A1A2E",
    "white":          "#1A1A2E",
    "cream":          "#1A1A2E",
    "cream_soft":     "#1A1A2E",
    "cream_muted":    "#6B7280",
    "coral":          "#FF7124",
    "coral_soft":     "rgba(255,113,36,0.12)",
    "gold":           "#FF7124",
    "gold_soft":      "rgba(255,113,36,0.12)",
    "neon_tangerine": "#FF7124",
    "tangerine_dark": "#D95A10",
    "tangerine_soft": "rgba(255,113,36,0.12)",
    "orange":         "#FF7124",
    "red":            "#FF4D2E",
    "red_soft":       "rgba(255,77,46,0.12)",
    "sky":            "#1B3F8B",
    "sky_soft":       "#3D64B8",
    "emerald_dark":   "#5A9E00",
    "world_red":      "#FF4D2E",
    "world_cyan":     "#1B3F8B",
    "world_yellow":   "#FF7124",
    "world_green":    "#7ED321",
}


GROWTH_SHEET = "Growth OS"
EARNINGS_SHEET = "Earnings"
AGENDA_SHEET = "Agenda"
CURRENT_GMV_SHEET = "Current GMV"
MAY_GMV_SHEET = "MAY GMV"
CURRENT_ADS_SHEET = "Current ADS"
CURRENT_MD_SHEET = "Current MD"
CURRENT_MD_PRO_SHEET = "Current MD pro"
SEASONAL_EVENTS_SHEET = "Seasonal Events"
TOP_PRODUCTS_SHEET = "Top 100 products CABA"
CROSS_SELL_SHEET = "Top 100 Cross Selling CABA"
DEFINITIVE_TOP_PRODUCTS_SHEET = "Definitive Top Products"
STORE_ID_SHEET = "Store ID"
PRIORITY_DATA_SHEET = "Priority Data"
COINVERSION_SHEET = "COINVERSION"
ASIGNACION_JUNIO_SHEET = "Asignacion Junio"
CURRENT_CHURN_SHEET = "Current Churn"
HEADER_ROW = 3

st.set_page_config(page_title="Growth OS", page_icon="📈", layout="wide")


# =========================
# FILE HELPERS
# =========================

@st.cache_data(ttl=600, show_spinner=False)
def _workbook_has_growth_sheet(path, mtime):
    """True si el .xlsx contiene la hoja 'Growth OS'. Cacheado por (path, mtime)
    para no reabrir workbooks en cada rerun. mtime forma parte de la clave."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        found = GROWTH_SHEET in wb.sheetnames
        wb.close()
        return found
    except Exception:
        return False


def find_excel_file():
    """
    Localiza el workbook base del Farmer, en orden de preferencia:
      1. El 'SR Farmer Base AR*.xlsx' más reciente (compatibilidad con el
         nombre histórico — evita cargar una copia vieja tipo '(1)').
      2. Cualquier .xlsx de la carpeta que contenga la hoja 'Growth OS',
         el más reciente primero. Esto hace la app portable: cada Farmer
         piloto usa su propio archivo con el nombre que quiera.
    Se excluyen archivos temporales de Excel (~$) y la carpeta de backups.
    """
    legacy = [
        f for f in glob.glob("SR Farmer Base AR*.xlsx")
        if not os.path.basename(f).startswith("~$")
    ]
    if legacy:
        legacy.sort(key=os.path.getmtime, reverse=True)
        return legacy[0]

    candidates = [
        f for f in glob.glob("*.xlsx")
        if not os.path.basename(f).startswith("~$")
        and BACKUP_FOLDER not in f
    ]
    candidates.sort(key=os.path.getmtime, reverse=True)
    for f in candidates:
        if _workbook_has_growth_sheet(f, os.path.getmtime(f)):
            return f

    # Sin candidatos: devolver el nombre histórico para que los mensajes de
    # "archivo no encontrado" del resto de la app sigan siendo coherentes.
    return "SR Farmer Base AR.xlsx"


EXCEL_FILE = find_excel_file()


# =========================
# CONFIG SHEET — parámetros del Farmer leídos desde el Excel
# =========================

CONFIG_SHEET = "Config"


@st.cache_data(ttl=600, show_spinner=False)
def load_app_config(excel_path, mtime):
    """
    Lee la hoja 'Config' (col A = clave, col B = valor) y devuelve un dict
    con claves normalizadas. Si la hoja no existe, devuelve {} y la app usa
    los defaults del código. Cacheado por mtime del archivo.
    """
    cfg = {}
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        if CONFIG_SHEET in wb.sheetnames:
            for row in wb[CONFIG_SHEET].iter_rows(min_row=1, max_col=2, values_only=True):
                key = str(row[0]).strip().lower().replace(" ", "_") if row and row[0] is not None else ""
                if key and len(row) > 1 and row[1] is not None:
                    cfg[key] = row[1]
        wb.close()
    except Exception:
        pass
    return cfg


def _cfg_str(cfg, key, default):
    v = cfg.get(key)
    return str(v).strip() if v is not None and str(v).strip() else default


def _cfg_num(cfg, key, default):
    try:
        v = float(str(cfg.get(key)).replace(",", "."))
        return v if v > 0 else default
    except Exception:
        return default


def _cfg_date(cfg, key, default):
    v = cfg.get(key)
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return default


if os.path.exists(EXCEL_FILE):
    _app_cfg = load_app_config(EXCEL_FILE, os.path.getmtime(EXCEL_FILE))
else:
    _app_cfg = {}

FARMER_NAME            = _cfg_str(_app_cfg, "farmer_name", FARMER_NAME)
FARMER_ROLE            = _cfg_str(_app_cfg, "farmer_role", FARMER_ROLE)
PORTFOLIO_COUNTRY      = _cfg_str(_app_cfg, "portfolio_country", PORTFOLIO_COUNTRY)
ARS_PER_USD            = _cfg_num(_app_cfg, "ars_per_usd", ARS_PER_USD)
COP_PER_USD            = _cfg_num(_app_cfg, "cop_per_usd", COP_PER_USD)
ADS_REVENUE_TARGET_USD = _cfg_num(_app_cfg, "ads_revenue_target_usd", ADS_REVENUE_TARGET_USD)
CONTACTS_START_DATE    = _cfg_date(_app_cfg, "contacts_start_date", CONTACTS_START_DATE)

# Derivados de identidad (para plantillas y Call Quality)
FARMER_FIRST_NAME  = FARMER_NAME.split()[0] if FARMER_NAME.strip() else "Farmer"
FARMER_ROLE_INLINE = (FARMER_ROLE[0].lower() + FARMER_ROLE[1:]) if FARMER_ROLE else ""


# =========================
# ASIGNACIÓN ACTIVA — detección dinámica de la hoja del mes
# =========================

_SPANISH_MONTHS = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


@st.cache_data(ttl=600, show_spinner=False)
def _detect_asignacion_sheet(excel_path, mtime):
    """
    Encuentra la hoja de asignación vigente sin hardcodear el mes.
    Busca hojas cuyo nombre empiece con 'Asignacion'/'Asignación' y elige la
    del mes más reciente según el nombre (ej: 'Asignacion Julio' > 'Asignacion
    Junio'). Si ninguna trae mes reconocible, usa la última en orden del
    workbook. Así el dashboard rota de mes a mes sin tocar código.
    """
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        names = wb.sheetnames
        wb.close()
    except Exception:
        return ASIGNACION_JUNIO_SHEET

    def _norm(s):
        s = str(s).strip().lower()
        return "".join(ch for ch in unicodedata.normalize("NFKD", s)
                       if not unicodedata.combining(ch))

    matches = [n for n in names if _norm(n).startswith("asignacion")]
    if not matches:
        return ASIGNACION_JUNIO_SHEET

    def _month_score(sheet_name):
        low = _norm(sheet_name)
        month = next((num for name, num in _SPANISH_MONTHS.items() if name in low), 0)
        year_m = re.search(r"(20\d{2})", low)
        year = int(year_m.group(1)) if year_m else date.today().year
        return (year, month)

    scored = [(m, _month_score(m)) for m in matches]
    if all(s[1][1] == 0 for s in scored):
        return matches[-1]
    return max(scored, key=lambda x: x[1])[0]


if os.path.exists(EXCEL_FILE):
    ASIGNACION_SHEET = _detect_asignacion_sheet(EXCEL_FILE, os.path.getmtime(EXCEL_FILE))
else:
    ASIGNACION_SHEET = ASIGNACION_JUNIO_SHEET


# =========================
# DATA ISSUES REGISTRY — avisos visibles cuando una fuente de datos falla
# =========================
# Filosofía: un cero silencioso es peor que un error visible. Cuando un loader
# falla (hoja renombrada, columnas cambiadas, export corrupto), en vez de
# tragarse la excepción y mostrar ceros, registra el problema acá. El sidebar
# muestra "⚠️ N avisos de datos" y el usuario sabe QUÉ se degradó y CÓMO
# arreglarlo — clave para pilotos sin soporte presencial.

def _log_data_issue(context, detail, hint=""):
    """Registra un problema de datos para mostrarlo en el sidebar."""
    try:
        issues = st.session_state.setdefault("_data_issues", {})
        issues[context] = {
            "detail": str(detail)[:300],
            "hint": hint,
            "time": datetime.now().strftime("%H:%M"),
        }
    except Exception:
        pass  # fuera del runtime de Streamlit (tests) no hay session_state


def _resolve_data_issue(context):
    """Limpia un aviso cuando la fuente vuelve a cargar bien."""
    try:
        st.session_state.setdefault("_data_issues", {}).pop(context, None)
    except Exception:
        pass


def normalize(text):
    if text is None:
        return ""
    return str(text).strip().lower()


def normalize_brand_id(value):
    """
    Makes Excel IDs and typed/copied IDs comparable.
    Handles examples like:
    9049, 9049.0, "AR-9049", "AR9049 - Brand", "65184 - Multistorefull" -> "9049" / "65184"
    """
    if value is None:
        return ""

    text = str(value).strip().upper()
    if text.endswith(".0") and text.replace(".0", "", 1).isdigit():
        text = text[:-2]

    # Prefer an AR-prefixed ID anywhere in the string.
    match = re.search(r"\bAR\s*-?\s*(\d+)\b", text)
    if match:
        return str(int(match.group(1)))

    # Otherwise use the first long-ish numeric block, so pasted values like
    # "65184 - Multistorefull" work without manual cleaning.
    match = re.search(r"\b(\d+)\b", text)
    if match:
        return str(int(match.group(1)))

    text = text.replace("AR-", "").replace("AR", "").strip()
    try:
        as_float = float(text)
        if as_float.is_integer():
            return str(int(as_float))
    except Exception:
        pass
    return text


def strip_brand_id_prefix(value):
    """
    Quita el ID de marca de un texto tipo 'AR16026 - Bonafide' y deja solo 'Bonafide'.
    Usar siempre para lo que se muestra al usuario / aliado (títulos, mensajes, etc).
    """
    text = clean(value, "")
    if not text:
        return text
    match = re.match(r"^\s*(?:AR\s*-?\s*\d+|\d+)\s*[-–—:]\s*(.+)$", text, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return text.strip()


def get_id_column_name(df):
    """
    Finds the ID column even if Excel/Pandas reads it with spaces or unusual casing.
    """
    for col in df.columns:
        if normalize(col) == "id":
            return col
    for col in df.columns:
        if "id" == normalize(str(col)).replace(" ", ""):
            return col
    return None


def clean(value, default="-"):
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass
    text = str(value)
    if text.lower() in ["nan", "none"]:
        return default
    return text


def to_number(value, default=0):
    try:
        if value is None or pd.isna(value):
            return default
        if isinstance(value, str):
            value = (
                value.replace("ARS", "")
                .replace("USD", "")
                .replace("COP", "")
                .replace("$", "")
                .replace(".", "")
                .replace(",", ".")
                .replace("%", "")
                .strip()
            )
        return float(value)
    except Exception:
        return default


def fmt_number(value):
    try:
        return f"{float(value):,.0f}".replace(",", ".")
    except Exception:
        return "-"


def fmt_money(value, prefix="$"):
    try:
        return f"{prefix} {float(value):,.0f}".replace(",", ".")
    except Exception:
        return "-"


def fmt_ars(value):
    return f"ARS {fmt_money(value)}"


def fmt_usd(value):
    return f"USD {fmt_money(value)}"


def fmt_cop(value):
    return f"COP {fmt_money(value)}"


def fmt_contact_number(value):
    """Formats phone/contact values without Excel's trailing .0."""
    text = clean(value, "-").strip()
    if text in ["", "-"]:
        return "-"
    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except Exception:
        pass
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def fmt_percent0(value):
    try:
        v = float(value)
        if abs(v) <= 2:
            v *= 100
        return f"{v:.0f}%"
    except Exception:
        return "-"


def fmt_percent2(value):
    try:
        v = float(value)
        if abs(v) <= 2:
            v *= 100
        return f"{v:.2f}%".replace(".", ",")
    except Exception:
        return "-"


def fmt_ratio(value):
    if value in [None, "-", ""]:
        return "-"
    try:
        v = float(value)
        if math.isnan(v):
            return "-"
        return f"{v:.1f}x"
    except Exception:
        return str(value)


def make_backup(excel_path):
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.splitext(os.path.basename(excel_path))[0]
    backup_path = os.path.join(BACKUP_FOLDER, f"{base}_backup_{stamp}.xlsx")
    shutil.copy2(excel_path, backup_path)
    return backup_path


def save_brand_changelog(brand_id, brand_name, updates_dict, old_row):
    """Append one row per changed field to CHANGELOG_FILE."""
    FIELD_LABEL_MAP = {
        "name": "Brand Name", "last_gmv_ars": "Last GMV ARS",
        "last_aov_ars": "AOV ARS", "ltor": "LTOR Tier",
        "churn": "Churn Status", "ads": "Ads Status",
        "ads_bookings": "Ads Bookings ARS", "ads_roi": "Ads ROI",
        "md": "MD Status", "md_bookings": "MD Discount / Promo",
        "md_roi": "MD ROI", "manager": "Manager",
        "assistant": "Assistant", "email": "Email", "comments": "Notes",
        "category": "Category", "contact_number": "Contact Number",
        "commission_rate": "Commission Rate", "pro_users_pct": "PRO Users %",
    }
    FIELD_COL_MAP = {
        "name":         ["name", "brand name", "restaurant name"],
        "last_gmv_ars": ["last gmv ars", "gmv ars"],
        "last_aov_ars": ["last aov ars", "aov ars"],
        "ltor":         ["ltor tier", "ltor"],
        "churn":        ["churn", "churn status"],
        "ads":          ["ads"],
        "ads_bookings": ["ads bookings", "ad bookings"],
        "ads_roi":      ["ads roi", "ad roi"],
        "md":           ["md", "md status"],
        "md_bookings":  ["md discount", "md promo", "markdown discount", "md bookings"],
        "md_roi":       ["md roi"],
        "manager":      ["manager", "restaurant manager", "account manager"],
        "assistant":    ["assistant"],
        "email":        ["email", "mail"],
        "comments":     ["comments", "comment"],
        "category":         ["category"],
        "contact_number":   ["contact number", "phone", "contact"],
        "commission_rate":  ["comm. rate", "commission rate", "commission"],
        "pro_users_pct":    ["pro users %", "pro %"],
    }
    rows_to_append = []
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for field_key, new_value in updates_dict.items():
        col_options = FIELD_COL_MAP.get(field_key, [])
        old_value = get_from_row(old_row, col_options, default="") if old_row is not None else ""
        old_str = str(old_value).strip() if old_value not in [None, "-", ""] else ""
        new_str = str(new_value).strip() if new_value not in [None, ""] else ""
        if old_str != new_str:
            rows_to_append.append({
                "datetime":   stamp,
                "brand_id":   brand_id,
                "brand_name": brand_name,
                "field":      FIELD_LABEL_MAP.get(field_key, field_key),
                "old_value":  old_str,
                "new_value":  new_str,
            })
    if not rows_to_append:
        return
    new_df = pd.DataFrame(rows_to_append)
    if os.path.exists(CHANGELOG_FILE):
        try:
            existing = pd.read_csv(CHANGELOG_FILE)
            combined = pd.concat([existing, new_df], ignore_index=True)
        except Exception:
            combined = new_df
    else:
        combined = new_df
    combined.to_csv(CHANGELOG_FILE, index=False, encoding="utf-8-sig")


# =========================
# LOAD DATA
# =========================

@st.cache_data(ttl=3000, show_spinner=False)
def load_growth_data():
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()

    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=GROWTH_SHEET, header=HEADER_ROW - 1)
    except Exception as e:
        _log_data_issue("Growth OS (hoja maestra)", e,
                        f"Verificá que '{EXCEL_FILE}' tenga la hoja '{GROWTH_SHEET}' con headers en la fila {HEADER_ROW}.")
        return pd.DataFrame()
    _resolve_data_issue("Growth OS (hoja maestra)")
    df.columns = [normalize(c) for c in df.columns]

    if "id" in df.columns:
        df = df[df["id"].notna()].copy()

    # Filtrar por país de operación — configurable vía PORTFOLIO_COUNTRY, no hardcodeado.
    # Growth OS es agnóstico de país: cambiar PORTFOLIO_COUNTRY adapta el dashboard
    # a cualquier portafolio sin tocar lógica.
    if "country" in df.columns and PORTFOLIO_COUNTRY:
        df = df[df["country"].astype(str).str.contains(PORTFOLIO_COUNTRY, case=False, na=False)].copy()

    return df


@st.cache_data(ttl=3000, show_spinner=False)
def load_asignacion_activa():
    """
    Carga la hoja de asignación activa (detectada dinámicamente) y devuelve un DataFrame con columnas normalizadas:
      brand_id   → str   (normalizado con normalize_brand_id)
      brand_name → str
      turbo      → bool  (True SOLO si columna C contiene un valor numérico — el Store Turbo ID)
      is_new     → bool  (True si la fila está en fuente roja — marca nueva del calendario)
    Devuelve DataFrame vacío si el sheet no existe o el archivo no está.
    """
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame(columns=["brand_id", "brand_name", "turbo", "is_new"])

    # ── Leer colores de fuente con openpyxl (pandas no expone esto) ────────────
    red_ids = set()
    turbo_numeric_ids = set()
    try:
        _wb = openpyxl.load_workbook(EXCEL_FILE, read_only=False, data_only=True)
        if ASIGNACION_SHEET in _wb.sheetnames:
            _ws = _wb[ASIGNACION_SHEET]
            for _row in _ws.iter_rows(min_row=2):
                _id_cell    = _row[0] if len(_row) > 0 else None
                _turbo_cell = _row[2] if len(_row) > 2 else None
                if _id_cell is None or _id_cell.value is None:
                    continue
                _bid = normalize_brand_id(_id_cell.value)
                if not _bid:
                    continue
                # Fuente roja → marca nueva
                try:
                    if (
                        _id_cell.font
                        and _id_cell.font.color
                        and _id_cell.font.color.type == "rgb"
                        and _id_cell.font.color.rgb in ("FFFF0000", "FF0000")
                    ):
                        red_ids.add(_bid)
                except Exception:
                    pass
                # Turbo → columna C tiene un valor numérico (el Store Turbo ID), no texto
                try:
                    if _turbo_cell and _turbo_cell.value is not None:
                        _tv = str(_turbo_cell.value).strip()
                        # Es Turbo si el valor es un número (entero o float)
                        if _tv.replace(".", "", 1).isdigit():
                            turbo_numeric_ids.add(_bid)
                except Exception:
                    pass
        _wb.close()
    except Exception:
        pass

    try:
        raw = pd.read_excel(EXCEL_FILE, sheet_name=ASIGNACION_SHEET, header=None)
    except Exception:
        return pd.DataFrame(columns=["brand_id", "brand_name", "turbo", "is_new"])

    # Detectar fila de encabezado buscando una celda que contenga "brand" o "id"
    header_index = None
    for i in range(min(20, len(raw))):
        row_vals = [normalize(x) for x in raw.iloc[i].values]
        if any("brand" in v or v == "id" for v in row_vals):
            header_index = i
            break

    if header_index is None:
        header_index = 0

    raw.columns = [normalize(c) for c in raw.iloc[header_index].values]
    df = raw.iloc[header_index + 1:].copy().reset_index(drop=True)
    df = df.dropna(how="all")

    # Mapeo flexible de columnas
    id_col   = _first_existing_col(df, ["brand id", "id", "store id", "tienda id"])
    name_col = _first_existing_col(df, ["brand name", "name", "nombre", "tienda"])

    if not id_col:
        return pd.DataFrame(columns=["brand_id", "brand_name", "turbo", "is_new"])

    out = pd.DataFrame()
    out["brand_id"]   = df[id_col].apply(normalize_brand_id)
    out["brand_name"] = df[name_col].apply(lambda v: clean(v, "-")) if name_col else "-"

    # Turbo: True solo si el brand_id está en el set de IDs con valor numérico en col C
    out["turbo"]  = out["brand_id"].isin(turbo_numeric_ids)
    # is_new: True si la fila tenía fuente roja
    out["is_new"] = out["brand_id"].isin(red_ids)

    out = out[out["brand_id"] != ""].copy()
    return out.reset_index(drop=True)


@st.cache_data(ttl=3000, show_spinner=False)
def get_turbo_info(brand_id):
    """
    Devuelve True si la marca es STORE TURBO según 'Asignacion Junio'.
    brand_id puede venir como int, float o str — se normaliza antes de comparar.
    """
    bid = normalize_brand_id(brand_id)
    if not bid:
        return False
    try:
        df = load_asignacion_activa()
        if df.empty:
            return False
        return bool((df["brand_id"] == bid).any() and
                    df.loc[df["brand_id"] == bid, "turbo"].values[0])
    except Exception:
        return False


@st.cache_data(ttl=3000, show_spinner=False)
def load_current_churn():
    """
    Carga la hoja 'Current Churn' y devuelve un dict {store_id: churn_status},
    con UNA entrada por fila (una por store).
    Stores que NO aparecen en la hoja se consideran 'On'.
    """
    if not os.path.exists(EXCEL_FILE):
        return {}
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=CURRENT_CHURN_SHEET)
        df.columns = [normalize(c) for c in df.columns]

        id_col  = _first_existing_col(df, ["country_brand_id", "brand id", "brand_id", "id"])
        sta_col = _first_existing_col(df, ["estado actual", "estado", "status", "churn status"])

        if not id_col or not sta_col:
            return {}

        result = {}
        for _, row in df.iterrows():
            bid = normalize_brand_id(row.get(id_col, ""))
            sta = clean(row.get(sta_col, ""), "").strip()
            if not bid or not sta:
                continue
            # One entry per store — last row wins if duplicate store IDs exist
            result[bid] = sta
        return result
    except Exception:
        return {}


@st.cache_data(ttl=3000, show_spinner=False)
def load_current_churn_per_brand():
    """
    Carga la hoja 'Current Churn' y devuelve un dict {brand_id: worst_churn_status}.
    Para marcas con múltiples tiendas, toma el PEOR estado presente:
      prioridad W3 > W2 > W1 > Off
    Usado en Brand Finder y get_churn_status (display de una sola marca).
    """
    # INTENCIONAL: jerarquía de DISPLAY por marca (peor estado ACTIVO primero).
    # Off pesa menos aquí a propósito: una marca multi-tienda con un solo local
    # cerrado no debe mostrarse "Off" entera si sus otras tiendas siguen en W1-W3.
    # La priorización de RETENCIÓN (Opportunity List) usa la jerarquía inversa:
    # Off primero, porque ahí el criterio es rescate, no diagnóstico de estado.
    _churn_order = {"W3": 4, "W2": 3, "W1": 2, "Off": 1}

    if not os.path.exists(EXCEL_FILE):
        return {}
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=CURRENT_CHURN_SHEET)
        df.columns = [normalize(c) for c in df.columns]

        id_col  = _first_existing_col(df, ["country_brand_id", "brand id", "brand_id", "id"])
        sta_col = _first_existing_col(df, ["estado actual", "estado", "status", "churn status"])

        if not id_col or not sta_col:
            return {}

        result = {}
        for _, row in df.iterrows():
            bid = normalize_brand_id(row.get(id_col, ""))
            sta = clean(row.get(sta_col, ""), "").strip()
            if not bid or not sta:
                continue
            current_priority = _churn_order.get(result.get(bid, ""), 0)
            new_priority      = _churn_order.get(sta, 0)
            if new_priority > current_priority:
                result[bid] = sta
        return result
    except Exception as e:
        _log_data_issue('Current Churn', e, 'Verificá columnas COUNTRY_BRAND_ID y Estado Actual.')
        return {}


@st.cache_data(ttl=3000, show_spinner=False)
def get_churn_status(brand_id):
    """
    Devuelve el Churn Status (peor estado) de una marca desde Current Churn, con emoji incluido.
    Usa load_current_churn_per_brand para obtener el peor estado entre todas las stores.
    Si no aparece en la hoja, retorna '✅ On'.
    """
    bid = normalize_brand_id(brand_id)
    if not bid:
        return "✅ On"
    churn_map = load_current_churn_per_brand()
    raw = churn_map.get(bid, "On")
    return _churn_label_with_emoji(raw)


@st.cache_data(ttl=3000, show_spinner=False)
def load_earnings_data():
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()

    return pd.read_excel(EXCEL_FILE, sheet_name=EARNINGS_SHEET, header=None)


@st.cache_data(ttl=300, show_spinner=False)
def load_agenda_data():
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()

    raw = pd.read_excel(EXCEL_FILE, sheet_name=AGENDA_SHEET, header=None)

    header_index = None
    for i in range(min(25, len(raw))):
        row_values = [normalize(x) for x in list(raw.iloc[i].values)]
        if "id" in row_values and "name" in row_values and "notes" in row_values:
            header_index = i
            break

    if header_index is None:
        _log_data_issue("Agenda", "Header no encontrado",
                        "La hoja Agenda necesita una fila con columnas id, name y notes en las primeras 25 filas.")
        return pd.DataFrame()
    _resolve_data_issue("Agenda")

    headers = [normalize(x) for x in list(raw.iloc[header_index].values)]
    data = raw.iloc[header_index + 1:].copy()
    data.columns = headers
    data = data.dropna(how="all")

    # Keep the real Excel row number so we can update a task/event later.
    # Pandas index is zero-based, Excel rows are one-based.
    data["_excel_row"] = data.index + 1

    if "" in data.columns:
        data = data.drop(columns=[""], errors="ignore")

    return data



def strip_accents(value):
    text = clean(value, "")
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def norm_text(value):
    return strip_accents(value).lower().strip()


def normalize_text(value):
    """Alias unificado. Usar esta función en todo código nuevo en lugar de norm_text o strip_accents directo."""
    return norm_text(value)


def extract_brand_id_from_current(value):
    if value is None:
        return ""
    text = str(value).strip()
    
    # Intenta buscar el primer bloque numérico limpio (el ID de la tienda)
    match = re.search(r"\b(\d+)\b", text)
    if match:
        return str(int(match.group(1)))
        
    return normalize_brand_id(value)

@st.cache_data(ttl=3000, show_spinner=False)
def _load_gmv_sheet_data(sheet_name):
    """
    Generic loader for GMV-style sheets with columns:
    Brand, GMV, GMV USD, Ordenes, AOV, AOV USD.
    Used for both 'Current GMV' and 'MAY GMV'.
    """
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()

    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        _resolve_data_issue(f"Hoja {sheet_name}")
    except Exception as e:
        _log_data_issue(f"Hoja {sheet_name}", e,
                        "Verificá que la hoja exista con columnas Brand, GMV y Ordenes.")
        return pd.DataFrame()

    df.columns = [normalize(c).replace("_", " ") for c in df.columns]
    df = df.loc[:, ~df.columns.duplicated()].copy()

    # Buscador flexible de la columna de marca/ID
    brand_col = _first_existing_col(df, ["brand", "brand name", "tienda", "nombre tienda", "code", "id"])
    if not brand_col:
        return pd.DataFrame()

    df["_id"] = df[brand_col].apply(extract_brand_id_from_current)
    # _brand_name_norm: nombre puro sin ID numérico al inicio (ej: "1234 - McDonald's" → "mcdonald's")
    def _extract_pure_name(v):
        s = str(v).strip()
        # Remover prefijo "12345 - " o "12345 - " si existe
        s = re.sub(r"^\d+[\s\-–]+", "", s)
        return normalize(s)
    df["_brand_name_norm"] = df[brand_col].apply(_extract_pure_name)
    # Mantener filas con ID numérico OR con nombre de marca válido (MAY GMV solo tiene nombres)
    df = df[(df["_id"] != "") | (df["_brand_name_norm"].str.len() > 0)].copy()
    df = df[df["_brand_name_norm"].str.strip() != ""].copy()

    if df.empty:
        return df

    # Mapeo elástico de métricas clave de GMV y AOV
    gmv_ars_col = _first_existing_col(df, ["gmv ars", "gmv local", "gmv", "ventas ars"])
    gmv_usd_col = _first_existing_col(df, ["gmv usd", "gmv_usd"])
    orders_col = _first_existing_col(df, ["ordenes", "orders", "pedidos", "ordenes gmv"])
    aov_ars_col = _first_existing_col(df, ["aov ars", "aov local", "ticket promedio ars"])
    aov_usd_col = _first_existing_col(df, ["aov usd", "aov_usd"])

    df["gmv ars"] = _prepare_numeric_col(df, gmv_ars_col)
    df["gmv usd"] = _prepare_numeric_col(df, gmv_usd_col)
    df["ordenes"] = _prepare_numeric_col(df, orders_col)
    df["aov ars"] = _prepare_numeric_col(df, aov_ars_col)
    df["aov usd"] = _prepare_numeric_col(df, aov_usd_col)

    # Recalcular si los valores calculados de Excel vienen vacíos o en cero
    gmv_usd_mask = (df["gmv usd"] == 0) & (df["gmv ars"] > 0)
    df["gmv usd"] = df["gmv usd"].mask(gmv_usd_mask, df["gmv ars"] / ARS_PER_USD)

    aov_ars_mask = (df["aov ars"] == 0) & (df["gmv ars"] > 0) & (df["ordenes"] > 0)
    df["aov ars"] = df["aov ars"].mask(aov_ars_mask, df["gmv ars"] / df["ordenes"])

    aov_usd_mask = (df["aov usd"] == 0) & (df["aov ars"] > 0)
    df["aov usd"] = df["aov usd"].mask(aov_usd_mask, df["aov ars"] / ARS_PER_USD)

    df = df.sort_values(by="gmv ars", ascending=False).reset_index(drop=True)
    df["_caba_rank"] = df.index + 1

    return df


def load_current_gmv_data():
    return _load_gmv_sheet_data(CURRENT_GMV_SHEET)


def load_may_gmv_data():
    return _load_gmv_sheet_data(MAY_GMV_SHEET)


@st.cache_data(ttl=3000, show_spinner=False)
def get_may_brand_metrics(brand_id, brand_name=""):
    """
    Busca la marca en MAY GMV cruzando por:
    1. ID numérico exacto (si la hoja trae IDs)
    2. Nombre de marca (normalizado, parcial) — necesario cuando MAY GMV usa nombres en vez de IDs,
       o cuando el ID de la hoja difiere del ID de Growth OS.
    La columna Brand en MAY GMV suele tener formato "12345 - Nombre" o solo "Nombre".
    """
    if not os.path.exists(EXCEL_FILE):
        return None

    # Leer MAY GMV directo para no depender del filtro de _id
    try:
        raw = pd.read_excel(EXCEL_FILE, sheet_name=MAY_GMV_SHEET)
    except Exception:
        return None

    raw.columns = [normalize(c).replace("_", " ") for c in raw.columns]
    raw = raw.loc[:, ~raw.columns.duplicated()].copy()

    brand_col = _first_existing_col(raw, ["brand", "brand name", "tienda", "nombre tienda", "code", "id"])
    if not brand_col:
        return None

    raw = raw[raw[brand_col].notna()].copy()
    if raw.empty:
        return None

    # Extraer ID numérico y nombre puro de cada fila
    raw["_row_id"]   = raw[brand_col].apply(normalize_brand_id)
    raw["_row_name"] = raw[brand_col].apply(
        lambda v: normalize(re.sub(r"^\d+[\s\-–]+", "", str(v).strip()))
    )

    target_id   = normalize_brand_id(brand_id)
    target_name = normalize(brand_name) if brand_name else ""

    result = pd.DataFrame()

    # Intento 1: ID exacto
    if target_id:
        result = raw[raw["_row_id"] == target_id]

    # Intento 2: nombre exacto
    if result.empty and target_name:
        result = raw[raw["_row_name"] == target_name]

    # Intento 3: nombre parcial — el nombre de Growth OS contiene el de MAY GMV o viceversa
    if result.empty and target_name:
        result = raw[raw["_row_name"].str.contains(re.escape(target_name), na=False)]
    if result.empty and target_name:
        result = raw[raw["_row_name"].apply(lambda v: target_name in v or (len(v) > 3 and v in target_name))]

    if result.empty:
        return None

    row = result.iloc[0]
    # MAY GMV columnas pueden llamarse "gmv ars", "gmv local", "gmv" — buscar flexible
    def _get_num(r, keys):
        for k in keys:
            v = r.get(k)
            if v is not None:
                n = to_number(v, 0)
                if n != 0:
                    return n
        return 0

    gmv_ars = _get_num(row, ["gmv ars", "gmv local", "gmv"])
    gmv_usd = _get_num(row, ["gmv usd", "gmv_usd"]) or (gmv_ars / ARS_PER_USD if gmv_ars else 0)
    orders  = _get_num(row, ["ordenes", "orders", "pedidos"])
    aov_ars = _get_num(row, ["aov ars", "aov local", "aov"]) or ((gmv_ars / orders) if gmv_ars and orders else 0)
    aov_usd = _get_num(row, ["aov usd", "aov_usd"]) or (aov_ars / ARS_PER_USD if aov_ars else 0)

    return {
        "gmv_ars": gmv_ars,
        "gmv_usd": gmv_usd,
        "gmv_cop": gmv_usd * COP_PER_USD,
        "orders": orders,
        "aov_ars": aov_ars,
        "aov_usd": aov_usd,
        "aov_cop": aov_usd * COP_PER_USD,
    }




@st.cache_data(ttl=3000, show_spinner=False)
def load_detalle_caba():
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name="Detalle CABA")
        _resolve_data_issue("Detalle CABA")
        df.columns = [normalize(c) for c in df.columns]
        # Extract brand_id from Brand column (format: "72087 - Ayguacamolee")
        if "brand" in df.columns:
            df["brand_id"] = df["brand"].apply(normalize_brand_id)
        # Calculate AOV
        gmv_col = next((c for c in df.columns if "gmv" in c), None)
        ord_col = next((c for c in df.columns if "orden" in c or "order" in c), None)
        if gmv_col and ord_col:
            df[gmv_col] = pd.to_numeric(df[gmv_col], errors="coerce").fillna(0)
            df[ord_col] = pd.to_numeric(df[ord_col], errors="coerce").fillna(0)
            df["aov"] = df.apply(
                lambda r: r[gmv_col] / r[ord_col] if r[ord_col] > 0 else 0, axis=1
            )
            df["_gmv"] = df[gmv_col]
            df["_ordenes"] = df[ord_col]
        return df
    except Exception as e:
        _log_data_issue("Detalle CABA", e,
                        "El GMV/AOV del portafolio sale de esta hoja. Verificá columnas Brand, GMV y Ordenes.")
        return pd.DataFrame()


@st.cache_data(ttl=3000, show_spinner=False)
def load_cvr_data():
    """Carga CVR% → {brand_name_clean: avg_cvr_ultimas4semanas}."""
    if not os.path.exists(EXCEL_FILE):
        return {}
    try:
        raw = pd.read_excel(EXCEL_FILE, sheet_name="CVR%", header=None)
        raw.columns = list(range(len(raw.columns)))
        mask = raw[0].astype(str).str.strip() == "CVR %"
        cvr = raw[mask].copy()
        if cvr.empty:
            return {}
        # 5 semanas disponibles; últimas 4 = cols 4,6,8,10
        val_cols = [4, 6, 8, 10]
        result = {}
        for _, r in cvr.iterrows():
            brand = str(r[1]).strip().lower()
            if not brand or brand in ["nan", "brand name"]:
                continue
            vals = []
            for c in val_cols:
                try:
                    v = float(r[c])
                    if pd.notna(v) and v > 0:
                        vals.append(v)
                except (TypeError, ValueError):
                    pass
            if vals:
                result[brand] = sum(vals) / len(vals)
        return result
    except Exception as e:
        _log_data_issue('CVR%', e, 'El export semanal de CVR cambió de formato o no está.')
        return {}


def get_portfolio_gmv_aov_from_detalle_caba():
    """
    Calcula GMV total, Ordenes totales y AOV del portafolio
    filtrando Detalle CABA por las marcas de Asignacion Junio (cruce por Brand Name).
    AOV del portafolio = promedio de los AOVs individuales por marca.
    """
    try:
        aj_df     = load_asignacion_activa()
        detalle   = load_detalle_caba()
        if aj_df.empty or detalle.empty:
            return None

        # ── Cruce por ID (primario) con fallback por nombre ──────────────────
        # El cruce por nombre exacto perdía marcas por diferencias de tildes,
        # sufijos y escritura entre hojas (medido: ~$15.7M ARS de GMV, +6.1%).
        # Ambas fuentes traen el ID numérico: Asignación como "AR16516" y
        # Detalle CABA como prefijo "16516 - Marca". El nombre queda solo como
        # red de seguridad para filas sin ID legible.
        gmv_col = "_gmv"     if "_gmv"     in detalle.columns else None
        ord_col = "_ordenes" if "_ordenes" in detalle.columns else None
        if not gmv_col or not ord_col:
            return None

        name_col = _first_existing_col(detalle, ["brand name", "brand", "nombre", "tienda"])

        # Índices agregados de Detalle CABA: por ID y por nombre normalizado
        det = detalle.copy()
        if "brand_id" not in det.columns and name_col:
            det["brand_id"] = det[name_col].apply(normalize_brand_id)
        det["_norm_name"] = det[name_col].apply(
            lambda x: normalize(re.sub(r"^\d+[\s\-–]+", "", str(x).strip()))
        ) if name_col else ""

        by_id = det[det["brand_id"].astype(str) != ""].groupby("brand_id").agg(
            gmv_total=(gmv_col, "sum"), ord_total=(ord_col, "sum")
        )
        by_name = det[det["_norm_name"].astype(str) != ""].groupby("_norm_name").agg(
            gmv_total=(gmv_col, "sum"), ord_total=(ord_col, "sum")
        )

        total_gmv_ars = 0.0
        total_orders  = 0.0
        matched = 0
        no_sales = 0
        for _, aj_row in aj_df.iterrows():
            bid   = normalize_brand_id(aj_row.get("brand_id", ""))
            bname = normalize(str(aj_row.get("brand_name", "")))
            if bid and bid in by_id.index:
                total_gmv_ars += float(by_id.loc[bid, "gmv_total"])
                total_orders  += float(by_id.loc[bid, "ord_total"])
                matched += 1
            elif bname and bname in by_name.index:
                total_gmv_ars += float(by_name.loc[bname, "gmv_total"])
                total_orders  += float(by_name.loc[bname, "ord_total"])
                matched += 1
            else:
                # Marca asignada sin filas de GMV en el export del período:
                # facturó cero → candidata natural a activación / rescate.
                no_sales += 1

        if matched == 0:
            return None

        # ── AOV ponderado: GMV total ÷ órdenes totales ────────────────────────
        # El promedio simple de AOVs individuales daba el mismo peso a una marca
        # de 5 órdenes que a una de 200, distorsionando la métrica frente a las
        # tablas oficiales. El ponderado es el estándar y cuadra en auditoría.
        aov_ars = (total_gmv_ars / total_orders) if total_orders > 0 else 0
        gmv_usd = total_gmv_ars / ARS_PER_USD
        aov_usd = aov_ars / ARS_PER_USD

        return {
            "gmv_ars": total_gmv_ars,
            "gmv_usd": gmv_usd,
            "gmv_cop": gmv_usd * COP_PER_USD,
            "orders":  total_orders,
            "aov_ars": aov_ars,
            "aov_usd": aov_usd,
            "aov_cop": aov_usd * COP_PER_USD,
            # Cobertura del cruce — consumido por el panel Data Health
            "brands_total":    int(len(aj_df)),
            "brands_matched":  int(matched),
            "brands_no_sales": int(no_sales),
        }
    except Exception:
        return None


def get_cvr_for_brand(brand_name, cr_fallback=None):
    """Devuelve (cvr_avg, source). Fallback: cr% - 6pp."""
    cvr_map = load_cvr_data()
    key = str(brand_name).strip().lower()
    if key in cvr_map:
        return cvr_map[key], "CVR%"
    if cr_fallback is not None:
        try:
            fb = float(cr_fallback)
            if fb > 0:
                fb_norm = fb if fb <= 1 else fb / 100
                return max(0.0, fb_norm - 0.06), "Growth OS -6%"
        except (TypeError, ValueError):
            pass
    return None, "Sin datos"


def get_cvr_category_benchmark(categoria):
    """CVR promedio de marcas de la misma categoría cruzando Detalle CABA × CVR%."""
    try:
        detalle = load_detalle_caba()
        cvr_map = load_cvr_data()
        if detalle.empty or not cvr_map:
            return None
        cat_col = next((c for c in detalle.columns if "categor" in c), None)
        brand_col = next((c for c in detalle.columns if "brand" in c), None)
        if not cat_col or not brand_col:
            return None
        cat_norm = normalize(categoria)
        sub = detalle[detalle[cat_col].apply(lambda v: cat_norm in normalize(str(v)))]
        if sub.empty:
            return None
        bench_vals = []
        for brand_raw in sub[brand_col].unique():
            bkey = str(brand_raw).strip().lower()
            if " - " in bkey:
                bkey = bkey.split(" - ", 1)[1].strip()
            if bkey in cvr_map:
                bench_vals.append(cvr_map[bkey])
        if not bench_vals:
            return None
        return sum(bench_vals) / len(bench_vals)
    except Exception:
        return None


@st.cache_data(ttl=3000, show_spinner=False)
def load_traffic_data():
    """Carga Traffic # -> {brand_name_clean: avg_traffic_semanal (promedio 5 semanas)}."""
    if not os.path.exists(EXCEL_FILE):
        return {}
    try:
        raw = pd.read_excel(EXCEL_FILE, sheet_name="Traffic #", header=None)
        raw.columns = list(range(len(raw.columns)))
        mask = raw[0].astype(str).str.strip() == "Trafico"
        if not mask.any():
            mask = raw[0].astype(str).str.strip().str.contains("fico", na=False)
        traffic = raw[mask].copy()
        if traffic.empty:
            return {}
        val_cols = [2, 4, 6, 8, 10]
        result = {}
        for _, r in traffic.iterrows():
            brand = str(r[1]).strip().lower()
            if not brand or brand in ["nan", "brand name", "--"]:
                continue
            vals = []
            for c in val_cols:
                try:
                    v = float(r[c])
                    if pd.notna(v) and v > 0:
                        vals.append(v)
                except (TypeError, ValueError):
                    pass
            if vals:
                result[brand] = sum(vals) / len(vals)
        return result
    except Exception as e:
        _log_data_issue('Traffic #', e, 'El export semanal de Traffic cambió de formato o no está.')
        return {}


def get_traffic_for_brand(brand_name):
    """Devuelve avg traffic semanal para una marca o None si no hay dato."""
    traffic_map = load_traffic_data()
    key = str(brand_name).strip().lower()
    return traffic_map.get(key, None)


def get_traffic_category_benchmark(categoria):
    """Traffic semanal promedio de marcas de la misma categoria (Detalle CABA x Traffic #)."""
    try:
        detalle = load_detalle_caba()
        traffic_map = load_traffic_data()
        if detalle.empty or not traffic_map:
            return None
        cat_col   = next((c for c in detalle.columns if "categor" in c), None)
        brand_col = next((c for c in detalle.columns if "brand" in c), None)
        if not cat_col or not brand_col:
            return None
        cat_norm = normalize(categoria)
        sub = detalle[detalle[cat_col].apply(lambda v: cat_norm in normalize(str(v)))]
        if sub.empty:
            return None
        bench_vals = []
        for brand_raw in sub[brand_col].unique():
            bkey = str(brand_raw).strip().lower()
            if " - " in bkey:
                bkey = bkey.split(" - ", 1)[1].strip()
            if bkey in traffic_map:
                bench_vals.append(traffic_map[bkey])
        if not bench_vals:
            return None
        return sum(bench_vals) / len(bench_vals)
    except Exception:
        return None


@st.cache_data(ttl=3000, show_spinner=False)
def get_market_context(categoria, lever, brand_gmv=None, brand_orders=None):
    df = load_detalle_caba()
    if df.empty:
        return {}
    # Find category column
    cat_col = next((c for c in df.columns if "categor" in c), None)
    if not cat_col:
        return {}
    # Filter by category (case-insensitive partial match)
    cat_norm = normalize(categoria)
    mask = df[cat_col].apply(lambda v: cat_norm in normalize(str(v)))
    sub = df[mask]
    if sub.empty:
        return {}
    gmv_vals = sub["_gmv"] if "_gmv" in sub.columns else pd.Series([], dtype=float)
    aov_vals = sub["aov"] if "aov" in sub.columns else pd.Series([], dtype=float)
    gmv_vals = gmv_vals[gmv_vals > 0]
    aov_vals = aov_vals[aov_vals > 0]
    total_brands = len(sub)
    gmv_avg = float(gmv_vals.mean()) if len(gmv_vals) else 0
    aov_avg = float(aov_vals.mean()) if len(aov_vals) else 0
    gmv_total = float(gmv_vals.sum()) if len(gmv_vals) else 0
    top_row = sub.loc[sub["_gmv"].idxmax()] if "_gmv" in sub.columns and not sub.empty else None
    top_brand_name = clean(top_row.get("brand", "-")) if top_row is not None else "-"
    top_gmv = float(top_row["_gmv"]) if top_row is not None else 0
    # Percentile
    brand_percentile = None
    if brand_gmv is not None and len(gmv_vals) > 0:
        below = (gmv_vals < brand_gmv).sum()
        brand_percentile = round(below / len(gmv_vals) * 100, 1)
    return {
        "market_gmv_total": fmt_ars(gmv_total),
        "market_brand_count": total_brands,
        "market_gmv_avg": fmt_ars(gmv_avg),
        "market_aov_avg": fmt_ars(aov_avg),
        "market_top_brand": top_brand_name,
        "market_top_gmv": fmt_ars(top_gmv),
        "brand_percentile": f"{brand_percentile}%" if brand_percentile is not None else "N/D",
    }


def get_orders_from_detalle_caba(brand_id, brand_name=""):
    """
    Suma las Ordenes de una marca específica desde Detalle CABA (puede haber
    múltiples filas por marca — distintas tiendas/stores, se suman todas).
    Cruce por Brand Name (misma lógica acordada para el portafolio), con
    fallback a Brand ID si el nombre no matchea.
    No cacheada directamente: depende de load_detalle_caba(), que ya está
    cacheada — evita doble nivel de caché para una función liviana.
    """
    try:
        detalle = load_detalle_caba()
        if detalle.empty or "_ordenes" not in detalle.columns:
            return 0

        target_name = normalize(str(brand_name)) if brand_name else ""
        if target_name and "brand" in detalle.columns:
            name_col = next((c for c in detalle.columns if c == "brand"), None)
            if name_col:
                # El campo Brand viene como "72087 - Ayguacamolee" → extraer la parte del nombre
                mask = detalle[name_col].astype(str).apply(
                    lambda x: normalize(re.sub(r"^\d+\s*-\s*", "", str(x).strip())) == target_name
                )
                filtered = detalle[mask]
                if not filtered.empty:
                    return float(filtered["_ordenes"].sum())

        # Fallback: cruce por brand_id
        if "brand_id" in detalle.columns:
            target_id = normalize_brand_id(brand_id)
            filtered = detalle[detalle["brand_id"] == target_id]
            if not filtered.empty:
                return float(filtered["_ordenes"].sum())

        return 0
    except Exception:
        return 0


@st.cache_data(ttl=3000, show_spinner=False)
def get_current_brand_metrics(brand_id):
    df = load_current_gmv_data()

    if df.empty:
        return None

    target = normalize_brand_id(brand_id)
    result = df[df["_id"] == target]

    if result.empty:
        return None

    row = result.iloc[0]
    current_gmv_ars = to_number(row.get("gmv ars"), 0)
    current_gmv_usd = to_number(row.get("gmv usd"), 0) or (current_gmv_ars / ARS_PER_USD if current_gmv_ars else 0)
    current_orders = to_number(row.get("ordenes"), 0)
    current_aov_ars = to_number(row.get("aov ars"), 0) or ((current_gmv_ars / current_orders) if current_gmv_ars and current_orders else 0)
    current_aov_usd = to_number(row.get("aov usd"), 0) or (current_aov_ars / ARS_PER_USD if current_aov_ars else 0)

    return {
        "gmv_ars": current_gmv_ars,
        "gmv_usd": current_gmv_usd,
        "gmv_cop": current_gmv_usd * COP_PER_USD,
        "orders": current_orders,
        "aov_ars": current_aov_ars,
        "aov_usd": current_aov_usd,
        "aov_cop": current_aov_usd * COP_PER_USD,
        "caba_rank": _format_rank(row.get("_caba_rank", "-")) if "_format_rank" in globals() else f"#{int(row.get('_caba_rank'))}",
    }


@st.cache_data(ttl=3000, show_spinner=False)
def get_pareto_tiers_map():
    """
    Calcula el Tier de Pareto de cada marca según su % acumulado de GMV total
    (Current GMV, ya viene ordenado desc por gmv ars con _caba_rank).
      Tier A → marcas que en conjunto representan el primer 80% del GMV total
      Tier B → el siguiente 15% acumulado (80%-95%)
      Tier C → el 5% final (95%-100%), e incluye también las marcas del
               portafolio vigente sin GMV registrado todavía (recién asignadas)
    El universo de GMV se restringe al portafolio vigente (Asignacion Junio) antes
    de calcular el acumulado, así el corte 80/15/5 es el Pareto del mes actual y no
    arrastra GMV de marcas ya reasignadas a otro Farmer.
    Devuelve dict {brand_id: "A"|"B"|"C"} — toda marca del portafolio vigente
    queda clasificada, igual que en el Pareto de referencia (Excel).
    """
    df = load_current_gmv_data()
    if df.empty or "gmv ars" not in df.columns:
        return {}

    d = df[["_id", "gmv ars"]].copy()

    aj_ids = set()
    try:
        _aj_tiers = load_asignacion_activa()
        if not _aj_tiers.empty:
            aj_ids = set(_aj_tiers["brand_id"].dropna().astype(str))
            aj_ids.discard("")
            if aj_ids:
                d = d[d["_id"].isin(aj_ids)].copy()
    except Exception:
        pass

    d["gmv ars"] = pd.to_numeric(d["gmv ars"], errors="coerce").fillna(0)
    d = d[d["gmv ars"] > 0].sort_values("gmv ars", ascending=False).reset_index(drop=True)

    total_gmv = d["gmv ars"].sum()
    if total_gmv <= 0:
        return {}

    d["_cum_pct"] = d["gmv ars"].cumsum() / total_gmv

    def _tier_for(cum_pct):
        if cum_pct <= 0.80:
            return "A"
        elif cum_pct <= 0.95:
            return "B"
        return "C"

    d["_tier"] = d["_cum_pct"].apply(_tier_for)
    tiers_map = dict(zip(d["_id"], d["_tier"]))

    # Marcas del portafolio vigente sin GMV en Current GMV (recién asignadas,
    # sin historial todavía) -> Tier C por defecto, igual criterio que el Excel.
    for bid in aj_ids:
        if bid and bid not in tiers_map:
            tiers_map[bid] = "C"

    return tiers_map


def get_pareto_tier(brand_id):
    """Devuelve 'A', 'B', 'C' o None si la marca no tiene GMV registrado en Current GMV."""
    bid = normalize_brand_id(brand_id)
    return get_pareto_tiers_map().get(bid)


PARETO_TIER_STYLE = {
    "A": {"color": "linear-gradient(145deg,#FF7124,#D95A10)", "label": "Pareto Tier A", "icon": "🥇"},
    "B": {"color": "linear-gradient(145deg,#1B3F8B,#3D64B8)", "label": "Pareto Tier B", "icon": "🥈"},
    "C": {"color": "linear-gradient(145deg,#6B7280,#4B5563)", "label": "Pareto Tier C", "icon": "🥉"},
}


def render_pareto_badge_html(brand_id):
    """Sticker circular con el Tier de Pareto, mismo estilo visual que ocupaba el badge Mundialista."""
    tier = get_pareto_tier(brand_id)
    if not tier:
        return ""
    style = PARETO_TIER_STYLE.get(tier)
    if not style:
        return ""
    return (
        f"<span class='hero-mundialista-badge' style='background:{style['color']};'>"
        f"<span class='badge-cup'>{style['icon']}</span> {style['label']}</span>"
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_current_gmv_totals():
    """
    Reads totals directly from the 'Total' summary row in the Current GMV sheet.
    This is the authoritative figure from the exported report — matches exactly
    what the Excel shows (GMV, Orders, AOV already computed by the export).
    Falls back to row-sum if the Total row is not found.
    """
    if not os.path.exists(EXCEL_FILE):
        return None
    try:
        raw = pd.read_excel(EXCEL_FILE, sheet_name=CURRENT_GMV_SHEET, engine="openpyxl")
        # Normalise column names for flexible lookup
        raw.columns = [str(c).strip() for c in raw.columns]

        # Find the Total row — Brand column contains exactly "Total"
        brand_col = next((c for c in raw.columns if c.lower() in ["brand", "brand name", "tienda", "nombre"]), raw.columns[0])
        total_mask = raw[brand_col].astype(str).str.strip().str.lower() == "total"
        total_row = raw[total_mask]

        if not total_row.empty:
            row = total_row.iloc[0]
            # Column name mapping: GMV (ARS), GMV USD, Ordenes, AOV (ARS), AOV USD
            gmv_ars_col  = next((c for c in raw.columns if c.upper() in ["GMV"] or c.lower() in ["gmv", "gmv ars", "gmv local"]), None)
            gmv_usd_col  = next((c for c in raw.columns if "usd" in c.lower() and "gmv" in c.lower()), None)
            orders_col   = next((c for c in raw.columns if c.lower() in ["ordenes", "orders", "pedidos"]), None)
            aov_ars_col  = next((c for c in raw.columns if c.upper() in ["AOV"] or c.lower() in ["aov", "aov ars", "aov local"]), None)
            aov_usd_col  = next((c for c in raw.columns if "usd" in c.lower() and "aov" in c.lower()), None)

            gmv_ars  = to_number(row.get(gmv_ars_col),  0) if gmv_ars_col  else 0
            gmv_usd  = to_number(row.get(gmv_usd_col),  0) if gmv_usd_col  else 0
            orders   = to_number(row.get(orders_col),   0) if orders_col   else 0
            aov_ars  = to_number(row.get(aov_ars_col),  0) if aov_ars_col  else 0
            aov_usd  = to_number(row.get(aov_usd_col),  0) if aov_usd_col  else 0

            # Derive missing values if needed
            if gmv_usd == 0 and gmv_ars > 0:
                gmv_usd = gmv_ars / ARS_PER_USD
            if aov_ars == 0 and gmv_ars > 0 and orders > 0:
                aov_ars = gmv_ars / orders
            if aov_usd == 0 and gmv_usd > 0 and orders > 0:
                aov_usd = gmv_usd / orders

            return {
                "gmv_ars": gmv_ars,
                "gmv_usd": gmv_usd,
                "gmv_cop": gmv_usd * COP_PER_USD,
                "orders":  orders,
                "aov_ars": aov_ars,
                "aov_usd": aov_usd,
                "aov_cop": aov_usd * COP_PER_USD,
            }
    except Exception:
        pass

    # ── Fallback: sum all data rows (no portfolio filter) ────────────────────
    current_df = load_current_gmv_data()
    if current_df is None or current_df.empty:
        return None

    total_gmv_ars = pd.to_numeric(current_df["gmv ars"], errors="coerce").fillna(0).sum()
    total_gmv_usd = pd.to_numeric(current_df["gmv usd"], errors="coerce").fillna(0).sum()
    total_orders  = pd.to_numeric(current_df["ordenes"],  errors="coerce").fillna(0).sum()
    aov_ars = total_gmv_ars / total_orders if total_orders else 0
    aov_usd = total_gmv_usd / total_orders if total_orders else 0

    return {
        "gmv_ars": total_gmv_ars,
        "gmv_usd": total_gmv_usd,
        "gmv_cop": total_gmv_usd * COP_PER_USD,
        "orders":  total_orders,
        "aov_ars": aov_ars,
        "aov_usd": aov_usd,
        "aov_cop": aov_usd * COP_PER_USD,
    }




@st.cache_data(ttl=3000, show_spinner=False)
def get_portfolio_ids():
    """
    Returns Brand IDs from Growth OS + Asignacion Junio combined.
    Used to filter Current sheets to the full active portfolio.
    """
    ids = set()
    try:
        growth_df = load_growth_data()
        id_col = get_id_column_name(growth_df) if not growth_df.empty else None
        if id_col:
            ids.update(growth_df[id_col].apply(normalize_brand_id).dropna().astype(str))
    except Exception:
        pass
    try:
        aj = load_asignacion_activa()
        if not aj.empty:
            ids.update(aj["brand_id"].dropna().astype(str))
    except Exception:
        pass
    ids.discard("")
    return ids


def money_from_usd(usd_value):
    usd = to_number(usd_value, 0)
    return {
        "usd": usd,
        "ars": usd * ARS_PER_USD,
        "cop": usd * COP_PER_USD,
    }


@st.cache_data(ttl=3000, show_spinner=False)
def _read_current_sheet(sheet_name):
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        df.columns = [normalize(c) for c in df.columns]
        _resolve_data_issue(f"Hoja {sheet_name}")
        return df
    except Exception as e:
        _log_data_issue(f"Hoja {sheet_name}", e,
                        f"Verificá que la hoja '{sheet_name}' exista en '{EXCEL_FILE}' con su formato de export original.")
        return pd.DataFrame()



def _elapsed_days_for_current_month(value=None):
    """Returns elapsed days for a MTD export row. Used only to normalize accumulated Ads booking to weekly booking."""
    today = date.today()
    try:
        dt = pd.to_datetime(value, errors="coerce")
    except Exception:
        dt = pd.NaT

    if pd.isna(dt):
        # Fallback to calendar day. If the export has no month, assume MTD through today.
        return max(today.day, 1)

    try:
        year = int(dt.year)
        month = int(dt.month)
        days_in_month = calendar.monthrange(year, month)[1]
        if today.year == year and today.month == month:
            return max(min(today.day, days_in_month), 1)
        # If export month is already closed, use the full month.
        if (today.year, today.month) > (year, month):
            return days_in_month
        return max(min(today.day, days_in_month), 1)
    except Exception:
        return max(today.day, 1)


@st.cache_data(ttl=3000, show_spinner=False)
def load_current_ads_data(portfolio_only=False):
    df = _read_current_sheet(CURRENT_ADS_SHEET)
    if df.empty:
        return pd.DataFrame()

    if "code" not in df.columns:
        return pd.DataFrame()

    df["_id"] = df["code"].apply(normalize_brand_id)

    for col in ["bookings net", "revenue net", "sales ads usd", "roi"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            df[col] = 0

    # Current ADS exports Booking as accumulated MTD. We keep accumulated values for totals,
    # but create a normalized weekly field for Brand Finder and Campaign Designer.
    if "month" in df.columns:
        df["_elapsed_days"] = df["month"].apply(_elapsed_days_for_current_month)
    else:
        df["_elapsed_days"] = _elapsed_days_for_current_month(None)
    df["_elapsed_days"] = pd.to_numeric(df["_elapsed_days"], errors="coerce").fillna(max(date.today().day, 1)).clip(lower=1)
    df["bookings weekly net"] = df["bookings net"] / df["_elapsed_days"] * 7

    if portfolio_only:
        ids = get_portfolio_ids()
        if ids:
            df = df[df["_id"].astype(str).isin(ids)].copy()
        else:
            df = pd.DataFrame()

    if df.empty:
        return df

    grouped = df.groupby("_id", as_index=False).agg({
        "bookings net": "sum",          # accumulated MTD booking
        "bookings weekly net": "sum",   # normalized weekly booking estimate
        "revenue net": "sum",           # accumulated MTD revenue; do NOT normalize
        "sales ads usd": "sum",
        "roi": "mean",
        "_elapsed_days": "max",
    })

    # Weighted ROI when possible: sales generated / revenue.
    grouped["_roi_calc"] = grouped.apply(
        lambda r: (r["sales ads usd"] / r["revenue net"]) if r["revenue net"] else r["roi"],
        axis=1
    )
    grouped["roi"] = grouped["_roi_calc"].fillna(grouped["roi"]).fillna(0)

    return grouped


@st.cache_data(ttl=3000, show_spinner=False)
def get_current_ads_metrics(brand_id):
    df = load_current_ads_data(portfolio_only=False)
    if df.empty:
        return {
            "active": False,
            "bookings_usd": 0,              # weekly booking used in Finder/Designer
            "bookings_accum_usd": 0,        # accumulated MTD booking from Current ADS
            "bookings_source": "none",
            "bookings_is_approx": False,
            "revenue_usd": 0,               # accumulated MTD revenue; do NOT normalize
            "sales_usd": 0,
            "roi": 0,
        }

    target = normalize_brand_id(brand_id)
    result = df[df["_id"].astype(str) == target]

    if result.empty:
        return {
            "active": False,
            "bookings_usd": 0,
            "bookings_accum_usd": 0,
            "bookings_source": "none",
            "bookings_is_approx": False,
            "revenue_usd": 0,
            "sales_usd": 0,
            "roi": 0,
        }

    row = result.iloc[0]
    bookings_accum = to_number(row.get("bookings net"), 0)
    bookings_weekly = to_number(row.get("bookings weekly net"), 0)
    revenue = to_number(row.get("revenue net"), 0)
    sales = to_number(row.get("sales ads usd"), 0)
    roi = to_number(row.get("roi"), 0)

    return {
        "active": any(v > 0 for v in [bookings_accum, revenue, sales]),
        "bookings_usd": bookings_weekly,
        "bookings_accum_usd": bookings_accum,
        "bookings_source": "current_ads_approx" if bookings_weekly else "none",
        "bookings_is_approx": bool(bookings_weekly),
        "revenue_usd": revenue,
        "sales_usd": sales,
        "roi": roi,
    }


def get_current_ads_totals():
    df = load_current_ads_data(portfolio_only=True)
    if df.empty:
        return {
            "bookings_usd": 0,
            "revenue_usd": 0,
            "sales_usd": 0,
            "roi": 0,
        }

    bookings = pd.to_numeric(df["bookings net"], errors="coerce").fillna(0).sum()
    revenue = pd.to_numeric(df["revenue net"], errors="coerce").fillna(0).sum()
    sales = pd.to_numeric(df["sales ads usd"], errors="coerce").fillna(0).sum()
    roi = sales / revenue if revenue else 0
    # Proyección de revenue al cierre del mes: BOOKINGS NET × 90% (umbral de comisión)
    projected_revenue_usd = bookings * 0.90

    return {
        "bookings_usd": bookings,
        "revenue_usd": revenue,
        "projected_revenue_usd": projected_revenue_usd,
        "sales_usd": sales,
        "roi": roi,
    }


def _first_existing_col(df, candidates):
    for candidate in candidates:
        key = normalize(candidate)
        if key in df.columns:
            return key
    return None


def _prepare_numeric_col(df, col_name):
    if col_name and col_name in df.columns:
        return pd.to_numeric(df[col_name], errors="coerce").fillna(0)
    return pd.Series([0] * len(df), index=df.index)


@st.cache_data(ttl=3000, show_spinner=False)
def load_current_md_data(portfolio_only=False, pro=False):
    """
    Loads Current MD or Current MD pro sheet.
    Sales  → column E (index 4): MARKDOWN $ / MARKDOWN PRO USR $   (already in USD)
    ROI    → column J (index 9): ROI / ROI MD PRIME
    Orders → used only to filter active rows.
    Conversion: USD values are converted to ARS (* ARS_PER_USD) and COP (* COP_PER_USD)
    at display time via money_from_usd() in the dashboard cards.
    """
    sheet = CURRENT_MD_PRO_SHEET if pro else CURRENT_MD_SHEET

    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception:
        sheet_names = []

    if pro and CURRENT_MD_PRO_SHEET not in sheet_names:
        sheet = CURRENT_MD_SHEET

    # ── Read raw with positional columns so we can pin col E and col J ──────
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()
    try:
        raw = pd.read_excel(EXCEL_FILE, sheet_name=sheet, header=0)
    except Exception:
        return pd.DataFrame()

    if raw.empty:
        if pro:
            try:
                raw = pd.read_excel(EXCEL_FILE, sheet_name=CURRENT_MD_SHEET, header=0)
            except Exception:
                return pd.DataFrame()
        if raw.empty:
            return pd.DataFrame()

    df = raw.copy()

    # ── Positional anchors (0-based): col E = idx 4, col I = idx 8, col J = idx 9 ──
    COL_E_IDX = 4   # MARKDOWN $ / MARKDOWN PRO USR $  → sales (USD)
    COL_I_IDX = 8   # TOTAL GMV MD / TOTAL GMV MD PRO  → gmv (col I)
    COL_J_IDX = 9   # ROI / ROI MD PRIME               → roi

    # Normalise headers for name-based lookups (ID, orders, campaigns)
    df.columns = [normalize(c).replace("_", " ").strip() for c in df.columns]

    id_col = _first_existing_col(df, ["brand id", "brand_id", "code", "id", "tienda id"])
    if not id_col:
        return pd.DataFrame()

    df["_id"] = df[id_col].apply(normalize_brand_id)

    # ── Sales: col E (idx 4) → already USD ───────────────────────────────────
    col_e_name = df.columns[COL_E_IDX] if COL_E_IDX < len(df.columns) else None
    df["_sales_usd"] = _prepare_numeric_col(df, col_e_name) if col_e_name else pd.Series([0]*len(df), index=df.index)

    # ── GMV total: col I (idx 8) → positional, always correct ───────────────
    col_i_name = df.columns[COL_I_IDX] if COL_I_IDX < len(df.columns) else None
    df["_gmv_usd"] = _prepare_numeric_col(df, col_i_name) if col_i_name else pd.Series([0]*len(df), index=df.index)

    # ── ROI: col J (idx 9) ────────────────────────────────────────────────────
    col_j_name = df.columns[COL_J_IDX] if COL_J_IDX < len(df.columns) else None
    df["_roi_raw"] = _prepare_numeric_col(df, col_j_name) if col_j_name else pd.Series([0]*len(df), index=df.index)

    # ── campaigns, orders: name-based ────────────────────────────────────────
    if pro:
        campaigns_col = _first_existing_col(df, ["campaigns pro #", "campaings pro #", "campaigns pro", "campaigns_pro", "campaigns"])
        orders_col    = _first_existing_col(df, ["orders md pro usr", "orders md pro usr #", "orders md pro", "orders_md_pro", "orders pro", "orders", "ordenes pro"])
    else:
        campaigns_col = _first_existing_col(df, ["campaings #", "campaigns #", "campaigns", "campaigns md"])
        orders_col    = _first_existing_col(df, ["orders md #", "orders md", "orders_md", "ordenes md", "pedidos md", "orders", "ordenes", "pedidos"])

    df["_campaigns"]  = _prepare_numeric_col(df, campaigns_col)
    df["_orders"]     = _prepare_numeric_col(df, orders_col)

    # Keep only rows with real orders
    df = df[df["_orders"] > 0].copy()

    if portfolio_only:
        ids = get_portfolio_ids()
        if ids:
            df = df[df["_id"].astype(str).isin(ids)].copy()
        else:
            df = pd.DataFrame()

    return df

@st.cache_data(ttl=3000, show_spinner=False)
def get_current_md_metrics(brand_id, pro=False):
    df = load_current_md_data(portfolio_only=False, pro=pro)

    if df.empty:
        return {
            "active": False,
            "sales_usd": 0,
            "roi": 0,
            "gmv_usd": 0,
            "campaigns": 0,
            "orders": 0,
        }

    target = normalize_brand_id(brand_id)
    result = df[df["_id"].astype(str) == target]

    if result.empty:
        return {
            "active": False,
            "sales_usd": 0,
            "roi": 0,
            "gmv_usd": 0,
            "campaigns": 0,
            "orders": 0,
        }

    sales = pd.to_numeric(result["_sales_usd"], errors="coerce").fillna(0).sum()
    gmv = pd.to_numeric(result["_gmv_usd"], errors="coerce").fillna(0).sum()
    campaigns = pd.to_numeric(result["_campaigns"], errors="coerce").fillna(0).sum()
    orders = pd.to_numeric(result["_orders"], errors="coerce").fillna(0).sum()
    # ROI comes directly from col J of the sheet
    roi = pd.to_numeric(result["_roi_raw"], errors="coerce").replace(0, float("nan")).mean()
    roi = roi if not (roi != roi) else 0  # NaN guard

    return {
        "active": orders > 0,
        "sales_usd": sales,
        "roi": roi,
        "gmv_usd": gmv,
        "campaigns": campaigns,
        "orders": orders,
    }


@st.cache_data(ttl=300, show_spinner=False)
def _read_md_totals_from_sheet(pro=False):
    """
    Reads the Total row directly from Current MD or Current MD pro.
    Col D = GMV TOTAL $, Col E = MARKDOWN $ (or MARKDOWN PRO USR $),
    Col F = MARKDOWN % (E/D) — only counted when E > 0.
    """
    sheet = CURRENT_MD_PRO_SHEET if pro else CURRENT_MD_SHEET
    if not os.path.exists(EXCEL_FILE):
        return {"gmv_total_usd": 0, "markdown_usd": 0, "markdown_pct": 0}
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            sheet = CURRENT_MD_SHEET
        ws = wb[sheet]
        total_row_idx = None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val is not None and str(val).strip().lower() == "total":
                total_row_idx = r
        result = {"gmv_total_usd": 0, "markdown_usd": 0, "markdown_pct": 0}
        if total_row_idx:
            d_val = ws.cell(total_row_idx, 4).value  # GMV TOTAL $
            e_val = ws.cell(total_row_idx, 5).value  # MARKDOWN $ / MARKDOWN PRO USR $
            f_val = ws.cell(total_row_idx, 6).value  # MARKDOWN % — E/D ratio
            result["gmv_total_usd"] = to_number(d_val, 0)
            result["markdown_usd"]  = to_number(e_val, 0) if e_val else 0
            # F = E/D — only valid when E > 0
            if result["markdown_usd"] > 0:
                result["markdown_pct"] = to_number(f_val, 0) if f_val else (
                    result["markdown_usd"] / result["gmv_total_usd"]
                    if result["gmv_total_usd"] > 0 else 0
                )
        wb.close()
        return result
    except Exception:
        return {"gmv_total_usd": 0, "markdown_usd": 0, "markdown_pct": 0}


def _read_md_targets_from_earnings():
    """
    Reads MD and MD Pro penetration targets from the Earnings sheet.
    Earnings row 3 (pandas index 2, header=None):
      col index 5 (Excel col F) = MD target %      (e.g. 0.0667 = 6.67%)
      col index 6 (Excel col G) = MD Pro target %  (e.g. 0.0727 = 7.27%)
    """
    defaults = {"md_target_pct": 0.0675, "md_pro_target_pct": 0.0722}
    if not os.path.exists(EXCEL_FILE):
        return defaults
    try:
        raw = load_earnings_data()
        if raw.empty:
            return defaults
        md_pct     = to_number(raw.iloc[2, 5], 0)   # col F
        md_pro_pct = to_number(raw.iloc[2, 6], 0)   # col G
        return {
            "md_target_pct":     md_pct     if md_pct     > 0 else defaults["md_target_pct"],
            "md_pro_target_pct": md_pro_pct if md_pro_pct > 0 else defaults["md_pro_target_pct"],
        }
    except Exception:
        return defaults


def get_markdown_dollar_total():
    """
    Returns the active MD GMV (MARKDOWN $) from the Total row of Current MD.
    Source: col E of the Total row — only counted when E > 0 (E/D = col F).
    """
    return _read_md_totals_from_sheet(pro=False)["markdown_usd"]


@st.cache_data(ttl=300, show_spinner=False)
def _read_md_roi_from_j290(pro=False):
    """
    Reads the ROI value directly from cell J290 of Current MD or Current MD pro.
    This is the authoritative portfolio-level ROI figure as computed by the sheet.
    """
    sheet = CURRENT_MD_PRO_SHEET if pro else CURRENT_MD_SHEET
    if not os.path.exists(EXCEL_FILE):
        return 0
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            # Fallback to non-pro sheet if pro not found
            sheet = CURRENT_MD_SHEET
        ws = wb[sheet]
        roi_val = ws["J290"].value
        wb.close()
        return to_number(roi_val, 0)
    except Exception:
        return 0


def get_current_md_totals(pro=False):
    df = load_current_md_data(portfolio_only=True, pro=pro)

    if df.empty:
        return {
            "sales_usd": 0,
            "gmv_usd": 0,
            "campaigns": 0,
            "orders": 0,
            "roi": 0,
        }

    sales = pd.to_numeric(df["_sales_usd"], errors="coerce").fillna(0).sum()
    gmv = pd.to_numeric(df["_gmv_usd"], errors="coerce").fillna(0).sum()
    campaigns = pd.to_numeric(df["_campaigns"], errors="coerce").fillna(0).sum()
    orders = pd.to_numeric(df["_orders"], errors="coerce").fillna(0).sum()
    # ROI read directly from cell J290 of the sheet — authoritative portfolio-level figure
    roi = _read_md_roi_from_j290(pro=pro)

    return {
        "sales_usd": sales,
        "gmv_usd": gmv,
        "campaigns": campaigns,
        "orders": orders,
        "roi": roi,
    }



def status_from_active(active):
    return "Active 🚀" if active else "Inactive 💤"


def fmt_roi(value):
    try:
        return f"{float(value):.2f}x"
    except Exception:
        return "-"


@st.cache_data(ttl=300, show_spinner=False)
def load_seasonal_events_data():
    df = _read_current_sheet(SEASONAL_EVENTS_SHEET)
    if df.empty:
        return pd.DataFrame()
    return df
@st.cache_data(ttl=300, show_spinner=False)
def load_coinversion_data():
    """Loads COINVERSION sheet. No header row — data starts at row 0."""
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=COINVERSION_SHEET, header=None, engine="openpyxl")
    except Exception:
        return pd.DataFrame()
    if df.empty:
        return pd.DataFrame()
    # Drop fully empty rows (row 0 is blank per inspection)
    df = df.dropna(how="all").reset_index(drop=True)
    # Assign known column positions: 0=ID, 1=Name, 2=Country, 3=Tier, 4=Status
    df.columns = ["brand_id", "brand_name", "country", "tier", "status"] + [f"extra_{i}" for i in range(max(0, len(df.columns) - 5))]
    df["_id"] = df["brand_id"].apply(normalize_brand_id)
    df["_name_norm"] = df["brand_name"].apply(lambda x: norm_text(clean(x, "")))
    return df


@st.cache_data(ttl=600, show_spinner=False)
def get_coinversion_info(brand_id, name=""):
    """Returns Coinversion sticker info for a brand.
    
    Tiers: GOLDEN → 🤝 Coinv. Golden | HIDDEN_GEM → 💎 Coinv. Hidden Gem
    """
    df = load_coinversion_data()
    if df.empty:
        return {"found": False, "sticker": None, "tier": None}

    target = normalize_brand_id(brand_id)
    result = df[df["_id"].astype(str) == target] if target else pd.DataFrame()

    # Fallback by name
    if result.empty and name:
        name_norm = norm_text(name)
        result = df[df["_name_norm"].apply(lambda x: x == name_norm or name_norm in x or x in name_norm)]

    if result.empty:
        return {"found": False, "sticker": None, "tier": None}

    row = result.iloc[0]
    tier = clean(row.get("tier"), "").strip().upper()

    if tier == "GOLDEN":
        sticker = "🤝 Coinv. Golden"
    elif "HIDDEN" in tier or "GEM" in tier:
        sticker = "💎 Coinv. Hidden Gem"
    else:
        sticker = "🤝 Coinversión"

    return {
        "found": True,
        "sticker": sticker,
        "tier": tier,
        "brand_name": clean(row.get("brand_name"), ""),
        "status": clean(row.get("status"), ""),
    }


def recommend_booster_for_brand(category, current_gmv_ars, current_aov_ars, cr, pro, ads_metrics, md_metrics):
    events = load_seasonal_events_data()

    if events.empty:
        return {
            "event": "-",
            "reason": "No Seasonal Events sheet found.",
            "period": "-",
            "event_type": "-",
        }

    category_text = norm_text(category)
    keywords = get_category_keywords(category)
    cr_value = to_number(cr, 0)
    pro_value = to_number(pro, 0)
    if cr_value > 2:
        cr_value = cr_value / 100
    if pro_value > 2:
        pro_value = pro_value / 100

    best = None
    best_score = -999
    best_reason = "-"

    for _, event in events.iterrows():
        event_name = clean(event.get("event_name"), "-")
        search_blob = " ".join([
            clean(event.get("category_fit"), ""),
            clean(event.get("product_focus"), ""),
            clean(event.get("commercial_criteria"), ""),
            clean(event.get("priority_logic"), ""),
            event_name,
        ])
        search_text = norm_text(search_blob)

        score = 0
        reasons = []

        if any(keyword in search_text for keyword in keywords):
            score += 40
            reasons.append("category fit")

        if "low cr" in search_text or "cr gap" in search_text or "conversion" in search_text:
            if cr_value and cr_value < 0.15:
                score += 25
                reasons.append("low CR")

        if "low/medium aov" in search_text or "ticket" in search_text or "reasonable" in search_text:
            if current_aov_ars and current_aov_ars <= 30000:
                score += 15
                reasons.append("AOV fit")

        if "gmv" in search_text:
            if current_gmv_ars and current_gmv_ars > 0:
                score += 10
                reasons.append("current GMV")

        if "inactive md" in search_text or "markdown" in search_text:
            if not md_metrics.get("active", False):
                score += 15
                reasons.append("MD opportunity")

        if "traffic" in search_text or "positioning" in search_text or "visibility" in search_text:
            if not ads_metrics.get("active", False):
                score += 8
                reasons.append("visibility opportunity")

        # Small manual boosts by event family.
        event_id = clean(event.get("event_id"), "").upper()
        if "BURGER" in event_id and any(k in category_text for k in ["burger", "hamburg"]):
            score += 50
        if "DESAYUNA" in event_id and any(k in category_text for k in ["cafe", "panaderia", "desayuno", "pasteleria"]):
            score += 50
        if "CHAMPIONS" in event_id and any(k in category_text for k in ["pizza", "empanada", "burger", "hamburg", "pollo", "sushi"]):
            score += 35
        if "CRAZY" in event_id and cr_value and cr_value < 0.15:
            score += 30

        if score > best_score:
            best_score = score
            best = event
            best_reason = " + ".join(dict.fromkeys(reasons)) if reasons else "market/event fit"

    if best is None:
        return {
            "event": "-",
            "reason": "-",
            "period": "-",
            "event_type": "-",
        }

    return {
        "event": clean(best.get("event_name"), "-"),
        "reason": best_reason,
        "period": clean(best.get("period"), "-"),
        "event_type": clean(best.get("event_type"), "-"),
    }



def fmt_signed_percent(value):
    try:
        v = float(value)
        if abs(v) <= 2:
            v *= 100
        sign = "+" if v > 0 else ""
        return f"{sign}{v:.0f}%"
    except Exception:
        return "-"


def safe_ratio(numerator, denominator):
    try:
        denominator = float(denominator)
        if denominator == 0:
            return None
        return float(numerator) / denominator
    except Exception:
        return None


def get_category_keywords(category):
    text = norm_text(category)

    keyword_rules = [
        (["pizza", "empanada"], ["pizza", "empanada", "muzzarella", "mozzarella", "napolitana", "fugazzeta"]),
        (["helado", "helados", "heladeria"], ["helado", "kilo", "kg", "mcflurry", "postre"]),
        (["panaderia", "cafe", "cafeteria", "desayuno", "pasteleria"], ["cafe", "medialuna", "tostado", "donut", "factura", "croissant", "desayuno"]),
        (["hamburg", "burger"], ["hamburg", "burger", "papas", "cheddar", "doble carne"]),
        (["sushi"], ["sushi", "roll", "salmon", "philadelphia", "nigiri"]),
        (["pollo", "chicken"], ["pollo", "chicken", "alitas", "nuggets"]),
        (["vegetariana", "veggie", "vegano"], ["ensalada", "veggie", "vegetariano", "verdura", "wrap", "bowl"]),
        (["argentina", "parrilla", "milanesa", "minutas"], ["milanesa", "suprema", "napolitana", "bife", "empanada", "asado", "lomito"]),
        (["arabe"], ["shawarma", "kebab", "falafel", "hummus", "arab"]),
        (["mexicana"], ["taco", "burrito", "quesadilla", "nachos", "guacamole"]),
        (["peruana"], ["lomo", "salteado", "ceviche", "aji"]),
        (["pasta", "pastas", "italiana"], ["pasta", "ravioles", "noquis", "ñoquis", "sorrentinos", "lasagna", "lasana"]),
        (["china", "comida china"], ["arroz", "salteado", "chow", "pollo", "cerdo", "wok"]),
    ]

    keywords = []
    for category_tokens, product_tokens in keyword_rules:
        if any(token in text for token in category_tokens):
            keywords.extend(product_tokens)

    if not keywords:
        keywords = [token for token in re.split(r"\W+", text) if len(token) >= 4]

    return list(dict.fromkeys(keywords))


def product_matches_keywords(product, keywords):
    product_text = norm_text(product)
    return any(keyword in product_text for keyword in keywords)


def clean_product_name(value):
    text = clean(value, "-")
    text = re.sub(r"^\s*\d+\.\s*", "", text)
    return text.strip()


@st.cache_data(ttl=300, show_spinner=False)
def get_caba_category_trends(category):
    keywords = get_category_keywords(category)

    product_trends = []
    cross_sell_trend = "-"

    # Top products trend
    try:
        products = pd.read_excel(EXCEL_FILE, sheet_name=TOP_PRODUCTS_SHEET)
        products.columns = [normalize(c) for c in products.columns]

        if "product" in products.columns:
            products = products[products["product"].notna()].copy()

            if "rank" in products.columns:
                products = products.sort_values(by="rank", ascending=True)
            elif "piezas" in products.columns:
                products = products.sort_values(by="piezas", ascending=False)

            products["_match"] = products["product"].apply(lambda x: product_matches_keywords(x, keywords))
            matched = products[products["_match"]].copy()

            if matched.empty:
                matched = products.copy()

            for product in matched["product"].tolist():
                product_text = clean_product_name(product)
                if product_text not in product_trends and product_text != "-":
                    product_trends.append(product_text)
                if len(product_trends) == 3:
                    break
    except Exception:
        product_trends = []

    # Cross-sell trend
    try:
        cross = pd.read_excel(EXCEL_FILE, sheet_name=CROSS_SELL_SHEET)
        cross.columns = [normalize(c) for c in cross.columns]

        if "pprincipal" in cross.columns and "psecundario" in cross.columns:
            if "%" in cross.columns:
                cross["_pct"] = pd.to_numeric(cross["%"], errors="coerce").fillna(0)
            else:
                cross["_pct"] = 0

            cross["_principal_clean"] = cross["pprincipal"].apply(clean_product_name)
            cross["_secondary_clean"] = cross["psecundario"].apply(clean_product_name)
            cross["_match"] = cross.apply(
                lambda r: product_matches_keywords(r["_principal_clean"], keywords) or product_matches_keywords(r["_secondary_clean"], keywords),
                axis=1,
            )

            matched = cross[cross["_match"]].copy()
            if matched.empty:
                matched = cross.copy()

            matched = matched.sort_values(by="_pct", ascending=False)
            if not matched.empty:
                best = matched.iloc[0]
                cross_sell_trend = f"{best['_principal_clean']} → {best['_secondary_clean']}"
    except Exception:
        cross_sell_trend = "-"

    while len(product_trends) < 3:
        product_trends.append("-")

    return {
        "products": product_trends[:3],
        "cross_sell": cross_sell_trend,
    }



def get_col(df, options, default=None):
    for option in options:
        col = normalize(option)
        if col in df.columns:
            return df[col]
    return pd.Series([default] * len(df), index=df.index)


def get_from_row(row, options, default="-"):
    for option in options:
        col = normalize(option)
        if col in row.index and not pd.isna(row.get(col)):
            return row.get(col)
    return default


def cell(df, r, c, default="-"):
    try:
        value = df.iloc[r, c]
        if pd.isna(value):
            return default
        return value
    except Exception:
        return default


# =========================
# EXCEL WRITE HELPERS
# =========================

def get_sheet_headers(ws, header_row=HEADER_ROW):
    headers = {}
    for c in range(1, ws.max_column + 1):
        value = ws.cell(header_row, c).value
        if value is not None:
            headers[normalize(value)] = c
    return headers


def find_column(headers, candidates):
    for candidate in candidates:
        key = normalize(candidate)
        if key in headers:
            return headers[key]
    return None


def find_brand_row(ws, headers, brand_id):
    id_col = find_column(headers, ["id"])
    if not id_col:
        return None

    target = normalize_brand_id(brand_id)

    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        value = ws.cell(r, id_col).value
        if value is not None and normalize_brand_id(value) == target:
            return r

    return None


@st.cache_data(ttl=3000, show_spinner=False)
def get_brand_ranking_from_excel(brand_id):
    """
    Reads ranking from the already-cached Growth OS dataframe.
    First tries the Ranking column. If blank, calculates rank from Last GMV ARS.
    Avoids opening openpyxl on every call.
    """
    try:
        df = load_growth_data()
        if df.empty:
            return "-"

        id_col   = get_id_column_name(df)
        rank_col = _first_existing_col(df, ["ranking", "rank", "top gmv rank", "top gmv ranking", "gmv ranking"])
        gmv_col  = _first_existing_col(df, ["last gmv ars", "gmv ars"])

        if not id_col:
            return "-"

        target = normalize_brand_id(brand_id)
        match = df[df[id_col].apply(normalize_brand_id) == target]

        if match.empty:
            return "-"

        # 1) Try direct Ranking value
        if rank_col:
            rank_text = clean(match.iloc[0].get(rank_col, "-"), "-")
            if rank_text not in ["", "-", "nan", "None"]:
                rank_text = str(rank_text).strip()
                if rank_text.endswith(".0"):
                    rank_text = rank_text[:-2]
                if not rank_text.startswith("#") and rank_text.replace(".", "", 1).isdigit():
                    return "#" + str(int(float(rank_text)))
                return rank_text

        # 2) Fallback: calculate rank from Last GMV ARS
        if not gmv_col:
            return "-"

        target_gmv = to_number(match.iloc[0].get(gmv_col, None), None)
        if target_gmv is None:
            return "-"

        all_gmv = df[gmv_col].apply(lambda x: to_number(x, None))
        higher_count = int((all_gmv > target_gmv).sum())
        return f"#{higher_count + 1}"

    except Exception:
        return "-"





def is_formula_cell(cell_obj):
    value = cell_obj.value
    return isinstance(value, str) and value.startswith("=")


def write_if_editable(ws, row_number, headers, candidates, value, updated, locked, missing):
    col = find_column(headers, candidates)
    label = candidates[0]

    if not col:
        missing.append(label)
        return

    cell_obj = ws.cell(row_number, col)

    if is_formula_cell(cell_obj):
        locked.append(label)
        return

    cell_obj.value = value
    updated.append(label)


def _update_agenda_notes_inner(wb, brand_id, notes_value, append=False):
    """Misma lógica de update_agenda_notes pero opera sobre un wb ya abierto, sin guardar ni cerrar."""
    if AGENDA_SHEET not in wb.sheetnames:
        return False, "Agenda sheet not found."

    ws = wb[AGENDA_SHEET]

    header_row = None
    headers = {}

    for r in range(1, min(25, ws.max_row) + 1):
        row_headers = {}
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            if value is not None:
                row_headers[normalize(value)] = c

        if "id" in row_headers and "notes" in row_headers:
            header_row = r
            headers = row_headers
            break

    if not header_row:
        return False, "Agenda headers not found."

    id_col = headers.get("id")
    notes_col = headers.get("notes")

    target = normalize_brand_id(brand_id)
    matched_rows = []

    for r in range(header_row + 1, ws.max_row + 1):
        value = ws.cell(r, id_col).value
        if value is not None and normalize_brand_id(value) == target:
            matched_rows.append(r)

    if not matched_rows:
        return False, "Brand ID not found in Agenda."

    row_number = matched_rows[0]
    notes_cell = ws.cell(row_number, notes_col)

    if append:
        previous = clean(notes_cell.value, default="")
        stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        new_note = f"[{stamp}] {notes_value}"
        notes_cell.value = f"{previous}\n\n{new_note}".strip() if previous else new_note
    else:
        notes_cell.value = notes_value

    return True, "Agenda notes updated."


def update_agenda_notes(excel_path, brand_id, notes_value, append=False):
    """Wrapper público sin cambios de comportamiento: abre, aplica, guarda, cierra."""
    wb = openpyxl.load_workbook(excel_path)
    ok, msg = _update_agenda_notes_inner(wb, brand_id, notes_value, append=append)
    if not ok:
        wb.close()
        return ok, msg
    try:
        wb.save(excel_path)
        wb.close()
        return True, "Agenda notes updated."
    except PermissionError:
        wb.close()
        return False, f"'{os.path.basename(EXCEL_FILE)}' está abierto en Excel. Cerralo para poder guardar las notas de Agenda."



def _add_event_to_agenda_inner(wb, event_data):
    """Misma lógica de add_event_to_agenda pero opera sobre un wb ya abierto, sin guardar ni cerrar."""
    if AGENDA_SHEET not in wb.sheetnames:
        return False, "Agenda sheet not found."

    ws = wb[AGENDA_SHEET]

    header_row = None
    headers = {}

    for r in range(1, min(25, ws.max_row) + 1):
        row_headers = {}
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            if value is not None:
                row_headers[normalize(value)] = c

        if "id" in row_headers and "name" in row_headers and "notes" in row_headers:
            header_row = r
            headers = row_headers
            break

    if not header_row:
        return False, "Agenda headers not found."

    def agenda_col(candidates):
        for candidate in candidates:
            key = normalize(candidate)
            if key in headers:
                return headers[key]
        return None

    col_map = {
        "date": agenda_col(["date", "data"]),
        "time": agenda_col(["time"]),
        "id": agenda_col(["id"]),
        "name": agenda_col(["name"]),
        "task": agenda_col(["task"]),
        "channel": agenda_col(["channel"]),
        "priority": agenda_col(["priority"]),
        "status": agenda_col(["status"]),
        "notes": agenda_col(["notes"]),
    }

    missing = [k for k, v in col_map.items() if v is None]
    if missing:
        return False, "Missing Agenda columns: " + ", ".join(missing)

    next_row = ws.max_row + 1
    for r in range(header_row + 1, ws.max_row + 2):
        is_empty = True
        for col in col_map.values():
            if ws.cell(r, col).value not in [None, ""]:
                is_empty = False
                break
        if is_empty:
            next_row = r
            break

    ws.cell(next_row, col_map["date"]).value = event_data.get("date")
    ws.cell(next_row, col_map["time"]).value = event_data.get("time")
    ws.cell(next_row, col_map["id"]).value = event_data.get("id")
    ws.cell(next_row, col_map["name"]).value = event_data.get("name")
    ws.cell(next_row, col_map["task"]).value = event_data.get("task")
    ws.cell(next_row, col_map["channel"]).value = event_data.get("channel")
    ws.cell(next_row, col_map["priority"]).value = event_data.get("priority")
    ws.cell(next_row, col_map["status"]).value = event_data.get("status")
    ws.cell(next_row, col_map["notes"]).value = event_data.get("notes")

    return True, "Event added to Weekly Calendar."


def add_event_to_agenda(excel_path, event_data):
    """Wrapper público sin cambios de comportamiento: abre, aplica, guarda, cierra."""
    wb = openpyxl.load_workbook(excel_path)
    ok, msg = _add_event_to_agenda_inner(wb, event_data)
    if not ok:
        wb.close()
        return ok, msg
    try:
        wb.save(excel_path)
        wb.close()
        return True, "Event added to Weekly Calendar."
    except PermissionError:
        wb.close()
        return False, f"'{os.path.basename(EXCEL_FILE)}' está abierto en Excel. Cerralo para poder guardar el evento."




def _update_contact_followup_fields_inner(wb, brand_id, contact_channel=None, opportunity_status=None, comment_text=""):
    """Misma lógica de update_contact_followup_fields pero opera sobre un wb ya abierto, sin guardar ni cerrar."""
    if GROWTH_SHEET not in wb.sheetnames:
        return False, "Growth OS sheet not found."

    ws = wb[GROWTH_SHEET]
    headers = get_sheet_headers(ws)
    row_number = find_brand_row(ws, headers, brand_id)

    if not row_number:
        return False, "Brand ID not found."

    updated_any = False
    effective_contact = _is_effective_comment(comment_text, opportunity_status)

    def set_if_found(candidates, value):
        nonlocal updated_any
        col = find_column(headers, candidates)
        if col:
            ws.cell(row_number, col).value = value
            updated_any = True
            return True
        return False

    # Last positive contact drives Follow-Up List temperature.
    # Only refresh it when the interaction was real/effective.
    if effective_contact:
        set_if_found(["last positive contact", "last contact", "last contacted"], date.today())

    channel_map = {
        "Call": "📞",
        "WhatsApp": "💬",
        "Chat": "💬",
        "Email": "✉️",
        "Meet": "🖥️",
        "Other": "📝",
    }
    if contact_channel:
        set_if_found(["last contact via", "contact via", "last via"], channel_map.get(contact_channel, contact_channel))

    status_map = {
        "Ghost 👻": "👻",
        "Not Contacted 👻": "👻",
        "Follow-up ✅": "✅",
        "Negotiation ⏳": "⏳",
        "Deal Closed 🏆": "🏆",
        "Activated 🚀": "🚀",
        "OFF 😴": "💤",
        "Rejected ❌": "❌",
    }
    if opportunity_status:
        set_if_found(
            ["opp ", "opp", "opportunity status", "commercial status", "status"],
            status_map.get(opportunity_status, opportunity_status)
        )

    # Always append the comment text to the Growth OS comments column (if it exists).
    if comment_text:
        comments_col = find_column(headers, ["comments", "comment"])
        if comments_col:
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
            new_note = f"[{stamp}] {comment_text}"
            cell_obj = ws.cell(row_number, comments_col)
            if not is_formula_cell(cell_obj):
                existing = clean(cell_obj.value, default="")
                cell_obj.value = f"{existing}\n\n{new_note}".strip() if existing else new_note
                updated_any = True

    if not updated_any:
        return False, "No matching follow-up fields found."

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    return True, "Follow-up fields updated."


def update_contact_followup_fields(excel_path, brand_id, contact_channel=None, opportunity_status=None, comment_text=""):
    """Wrapper público sin cambios de comportamiento: abre, aplica, guarda, cierra."""
    if not os.path.exists(excel_path):
        return False, "Excel file not found."
    wb = openpyxl.load_workbook(excel_path)
    ok, msg = _update_contact_followup_fields_inner(wb, brand_id, contact_channel=contact_channel, opportunity_status=opportunity_status, comment_text=comment_text)
    if not ok:
        wb.close()
        return ok, msg
    try:
        wb.save(excel_path)
        wb.close()
        return True, "Follow-up fields updated."
    except PermissionError:
        wb.close()
        return False, f"'{os.path.basename(EXCEL_FILE)}' está abierto en Excel. Cerralo para poder guardar el follow-up."


def _update_brand_in_excel_inner(wb, brand_id, updates):
    """Misma lógica de update_brand_in_excel pero opera sobre un wb ya abierto, sin guardar ni cerrar."""
    if GROWTH_SHEET not in wb.sheetnames:
        return False, "Growth OS sheet not found.", [], [], []

    ws = wb[GROWTH_SHEET]
    headers = get_sheet_headers(ws)
    row_number = find_brand_row(ws, headers, brand_id)

    if not row_number:
        return False, "Brand ID not found.", [], [], []

    updated = []
    locked = []
    missing = []

    field_map = {
        "name": ["name", "brand name", "restaurant name"],
        "last_gmv_ars": ["last gmv ars", "gmv ars", "last gmv local"],
        "last_aov_ars": ["last aov ars", "aov ars", "last aov local"],
        "ltor": ["ltor tier", "ltor"],
        "ads": ["ads"],
        "ads_bookings": ["ads bookings", "ads bookings ars", "ad bookings"],
        "ads_roi": ["ads roi", "ad roi"],
        "md": ["md", "md status"],
        "md_bookings": ["md discount", "md promo", "markdown discount", "markdown promo", "md bookings", "md bookings ars", "md booking"],
        "md_roi": ["md roi"],
        "manager": ["manager", "restaurant manager", "account manager"],
        "assistant": ["assistant"],
        "email": ["email", "mail", "contact mail"],
        "churn": ["churn", "churn status"],
        "comments": ["comments", "comment"],
        "category": ["category"],
        "contact_number": ["contact number", "phone", "contact"],
        "commission_rate": ["comm. rate", "commission rate", "commission"],
        "pro_users_pct": ["pro users %", "pro %", "pro users", "prime users %"],
    }

    for key, candidates in field_map.items():
        if key in updates:
            write_if_editable(ws, row_number, headers, candidates, updates[key], updated, locked, missing)

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    return True, "Changes saved successfully.", updated, locked, missing


def update_brand_in_excel(brand_id, updates):
    """Wrapper público: abre una sola vez, aplica todos los cambios (incluido
    Agenda notes si corresponde), guarda, cierra. Backup corre en background
    para no bloquear el guardado. Invalida solo el caché de Growth OS/Agenda,
    no el archivo entero (CVR%, Traffic, Detalle CABA, etc. no cambiaron aquí)."""
    if not os.path.exists(EXCEL_FILE):
        return False, "Excel file not found.", [], [], [], None

    import threading as _threading_ub
    _threading_ub.Thread(target=make_backup, args=(EXCEL_FILE,), daemon=True).start()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ok, msg, updated, locked, missing = _update_brand_in_excel_inner(wb, brand_id, updates)

    if not ok:
        wb.close()
        return False, msg, updated, locked, missing, None

    if "comments" in updates:
        try:
            _update_agenda_notes_inner(wb, brand_id, updates["comments"], append=False)
        except Exception:
            pass

    try:
        wb.save(EXCEL_FILE)
        wb.close()
    except PermissionError:
        wb.close()
        return False, f"'{os.path.basename(EXCEL_FILE)}' está abierto en Excel. Cerralo y volvé a intentar el guardado.", updated, locked, missing, None

    # Invalidar solo lo que pudo haber cambiado — Growth OS y Agenda.
    # st.cache_data.clear() completo forzaba releer las ~40 funciones cacheadas
    # del archivo (CVR%, Traffic, Detalle CABA, Priority Data, etc.) que no
    # tienen nada que ver con un cambio de manager/email/categoría de una marca.
    try:
        load_growth_data.clear()
        load_agenda_data.clear()
    except Exception:
        st.cache_data.clear()  # fallback de seguridad si los nombres cambiaron

    return True, "Changes saved successfully.", updated, locked, missing, None


# =========================
# COMMENT HELPERS
# =========================

def _is_not_contacted_text(value):
    text = norm_text(value)
    if not text:
        return False
    patterns = [
        "no contesta", "no responde", "sin respuesta", "ghosted", "not contacted",
        "buzon", "buzón", "no answer", "wrong number", "numero errado", "número errado"
    ]
    return any(p in text for p in patterns)


def _is_effective_comment(comment, status=""):
    return bool(clean(comment, "").strip()) and not _is_not_contacted_text(comment) and not _is_not_contacted_text(status)


def save_comment_csv(brand_id, brand_name, comment, contact_channel="", opportunity_status="", commercial_action=""):
    new_row = pd.DataFrame([{
        "datetime": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "brand_id": str(normalize_brand_id(brand_id)),
        "brand_name": brand_name,
        "comment": comment,
        "contact_channel": contact_channel,
        "opportunity_status": opportunity_status,
        "commercial_action": commercial_action,
    }])

    if os.path.exists(COMMENTS_FILE):
        old = pd.read_csv(COMMENTS_FILE, encoding="utf-8-sig")
        for col in new_row.columns:
            if col not in old.columns:
                old[col] = ""
        final = pd.concat([old[new_row.columns], new_row], ignore_index=True)
    else:
        final = new_row

    final.to_csv(COMMENTS_FILE, index=False, encoding="utf-8-sig")


def _parse_claude_note_fields(note_text):
    """
    Parsea una nota [Auto] con análisis completo (generada por Claude en chat)
    y devuelve sus campos estructurados para guardar en el histórico y para
    prellenar el cuadro de accionable en el Follow-up.
    Si la nota no tiene ese formato (nota manual, vieja, o status de No Answer),
    devuelve todos los campos vacíos — la fila igual se guarda, solo sin
    estructura de análisis.
    """
    empty = {
        "sentiment": "", "palancas": "", "status_sugerido": "",
        "resumen": "", "proximos_pasos": "", "retomar": "",
        "cal_aplica": False, "cal_fecha": "", "cal_canal": "",
        "cal_prioridad": "", "cal_tema": "",
    }
    if not note_text or not note_text.strip().startswith("[Auto]"):
        return empty

    header_match = re.search(
        r"\[Auto\]\s*([^\n·]+)·\s*Palancas:\s*([^\n·]+)·\s*Status sugerido:\s*([^\n]+)",
        note_text
    )
    sentiment  = header_match.group(1).strip() if header_match else ""
    palancas   = header_match.group(2).strip() if header_match else ""
    status_sug = header_match.group(3).strip() if header_match else ""

    resumen_match = re.search(r"(?i)resumen:\s*(.+?)(?:\n\s*\n|\Z)", note_text, re.DOTALL)
    resumen = resumen_match.group(1).strip() if resumen_match else ""

    pasos_match = re.search(r"(?i)próximos pasos:\s*(.+?)(?:\n\s*\n|\Z)", note_text, re.DOTALL)
    proximos_pasos = pasos_match.group(1).strip() if pasos_match else ""

    retomar_matches = re.findall(r"(?im)^\s*retomar:\s*(.+?)\s*$", note_text)
    retomar = retomar_matches[-1].strip() if retomar_matches else ""

    # ── Bloque de accionable (Calendario) ─────────────────────────────────────
    cal_aplica, cal_fecha, cal_canal, cal_prioridad, cal_tema = False, "", "", "", ""
    cal_match = re.search(r"(?is)Calendario:\s*\n?(.+?)(?:\n\s*\n|\Z)", note_text)
    if cal_match:
        cal_block = cal_match.group(1)
        if not re.search(r"(?i)no aplica", cal_block):
            cal_aplica = True

            def _cal_field(key):
                fm = re.search(rf"(?im)^\s*{key}:\s*(.+?)\s*$", cal_block)
                return fm.group(1).strip() if fm else ""

            cal_fecha = _cal_field("Fecha")
            cal_canal = _cal_field("Canal")
            cal_prioridad = _cal_field("Prioridad")
            cal_tema = _cal_field("Tema")

    return {
        "sentiment": sentiment, "palancas": palancas, "status_sugerido": status_sug,
        "resumen": resumen, "proximos_pasos": proximos_pasos, "retomar": retomar,
        "cal_aplica": cal_aplica, "cal_fecha": cal_fecha, "cal_canal": cal_canal,
        "cal_prioridad": cal_prioridad, "cal_tema": cal_tema,
    }


def save_call_history_row(brand_id, brand_name, note_text, contact_channel="", opportunity_status=""):
    """
    Guarda una fila en el histórico de contactos (growth_os_call_history.csv).
    Se llama junto a save_comment_csv en cada Save Follow-up — esta tabla es la
    que alimenta análisis de tendencia por marca: sentimiento a lo largo del
    tiempo, palancas que se repiten sin cerrarse, promesas incumplidas, etc.
    No reemplaza growth_os_comments.csv (que sigue siendo la fuente de "última
    nota" en vivo); esta es la serie histórica completa, una fila por contacto.
    """
    parsed = _parse_claude_note_fields(note_text)
    now = datetime.now()
    new_row = pd.DataFrame([{
        "datetime":          now.strftime("%Y-%m-%d %H:%M"),
        "date":              now.strftime("%Y-%m-%d"),
        "week":              now.strftime("%G-W%V"),
        "brand_id":          str(normalize_brand_id(brand_id)),
        "brand_name":        brand_name,
        "contact_channel":   contact_channel,
        "opportunity_status": opportunity_status,
        "sentiment":         parsed["sentiment"],
        "palancas":          parsed["palancas"],
        "status_sugerido":   parsed["status_sugerido"],
        "resumen":           parsed["resumen"],
        "proximos_pasos":    parsed["proximos_pasos"],
        "retomar":           parsed["retomar"],
        "proximo_accionable_fecha":     parsed["cal_fecha"],
        "proximo_accionable_canal":     parsed["cal_canal"],
        "proximo_accionable_prioridad": parsed["cal_prioridad"],
        "proximo_accionable_tema":      parsed["cal_tema"],
        "source":            "claude" if note_text and note_text.strip().startswith("[Auto]") else "manual",
        "raw_note":          note_text or "",
    }])

    if os.path.exists(CALL_HISTORY_FILE):
        old = pd.read_csv(CALL_HISTORY_FILE, encoding="utf-8-sig")
        for col in new_row.columns:
            if col not in old.columns:
                old[col] = ""
        final = pd.concat([old[new_row.columns], new_row], ignore_index=True)
    else:
        final = new_row

    final.to_csv(CALL_HISTORY_FILE, index=False, encoding="utf-8-sig")


def _extract_customer_text(transcript):
    """
    Separa la transcripción por hablante (formato Amazon Connect: 'Agent:' / 'Customer:'
    o 'Agente:' / 'Cliente:'). Devuelve (customer_text, full_text_lower).
    Si no detecta ningún prefijo de hablante, devuelve el texto completo como customer_text
    (fallback seguro para transcripciones sin formato).
    """
    speaker_pattern = re.compile(
        r'(?im)^\s*(agent|agente|customer|cliente)\s*:\s*(.*)$'
    )
    matches = speaker_pattern.findall(transcript)
    if not matches:
        return transcript.lower(), transcript.lower()

    customer_lines = []
    for speaker, line in matches:
        if speaker.lower() in ("customer", "cliente"):
            customer_lines.append(line)
    customer_text = " ".join(customer_lines).lower()
    full_text = transcript.lower()
    return (customer_text if customer_text.strip() else full_text), full_text


def _has_negated(text, keyword, window=4):
    """
    Busca `keyword` en `text` y revisa si dentro de las `window` palabras anteriores
    aparece una negación (no, nunca, jamás, para nada, imposible, no creo, no puedo).
    Devuelve True si la mención está negada (y por tanto debe invertirse o descartarse).
    """
    negators = {"no", "nunca", "jamás", "tampoco", "ni"}
    negator_phrases = ["para nada", "imposible", "no creo", "no puedo", "no podemos",
                        "no queremos", "no tenemos", "ni de", "que va"]
    idx = text.find(keyword)
    if idx == -1:
        return False
    if any(phrase in text[max(0, idx - 40):idx] for phrase in negator_phrases):
        return True
    preceding = text[max(0, idx - 40):idx].split()
    preceding_window = preceding[-window:] if len(preceding) > window else preceding
    return any(w.strip(",.;") in negators for w in preceding_window)


def _detect_sentiment_weighted(customer_text):
    """
    Detecta sentimiento solo sobre el texto del cliente, con:
      1. Rechazo explícito de alto peso -> Negative inmediato, sin importar nada más.
      2. Negación de señales positivas -> esas señales se descartan, no cuentan a favor.
      3. Frases completas de rechazo a palancas específicas (ej. "no me interesan los descuentos").
    """
    low = customer_text

    high_weight_rejection = [
        "no me llames más", "no me llame más", "no estoy interesado", "no estoy interesada",
        "estamos bien así", "no es el momento", "tengo contrato con otro", "ya tengo otro proveedor",
        "no quiero seguir", "dejame en paz", "déjeme en paz", "no quiero hablar de esto",
        "no me interesa para nada", "ya dije que no"
    ]
    if any(phrase in low for phrase in high_weight_rejection):
        return "Negative", True  # True = rechazo explícito, máxima prioridad

    positive_signals = ["perfecto", "de acuerdo", "vamos", "sí claro", "me interesa",
                        "lo hacemos", "dale", "genial", "excelente", "cerramos", "activamos",
                        "okay", "ok", "buenísimo", "bárbaro", "confirmado", "listo"]
    negative_signals = ["no me interesa", "no quiero", "estoy conforme", "no gracias",
                        "no puedo", "no tengo presupuesto", "ya no", "cancel",
                        "cortamos", "no voy a", "lo dejo", "me retiro"]

    pos = 0
    for k in positive_signals:
        if k in low and not _has_negated(low, k):
            pos += 1
    neg = sum(1 for k in negative_signals if k in low)

    if neg >= 2 or (neg > pos and neg > 0):
        return "Negative", False
    elif pos >= 2 or pos > neg:
        return "Positive", False
    else:
        return "Neutral", False


def analyze_transcript_locally(transcript):
    """
    Analyzes a call transcript using pure Python keyword/regex logic.
    No external API calls. Returns a dict with:
      - summary: str (auto-generated note to save as comment)
      - action_items: list[str]
      - suggested_status: str (one of the follow-up statuses)
      - levers: list[str] (palancas detected)
      - sentiment: str (Positive / Neutral / Negative)
    """
    if not transcript or not transcript.strip():
        return None

    customer_text, low = _extract_customer_text(transcript)
    result = {}

    # ── Sentiment (solo sobre lo que dice el cliente, con negaciones y rechazo explícito) ──
    sentiment_value, explicit_rejection = _detect_sentiment_weighted(customer_text)
    result["sentiment"] = sentiment_value
    result["explicit_rejection"] = explicit_rejection

    # ── Levers / palancas detectadas ─────────────────────────────────────────
    levers = []
    if any(k in low for k in ["rappi ads", "ads", "publicidad", "banner", "sponsored", "visibilidad paga", "campaña"]):
        levers.append("ADS")
    md_rejection_phrases = ["no me interesan los descuentos", "no tenemos margen para descuentos",
                             "no queremos hacer promoción", "no podemos bajar el precio",
                             "ya tenemos promoción", "no me sirve el descuento"]
    md_keywords_present = any(k in low for k in ["descuento", "promo", "markdown", "porcentaje", "20%", "25%", "30%", "oferta", "promoción"])
    md_explicitly_rejected = any(phrase in low for phrase in md_rejection_phrases)
    if md_keywords_present and not md_explicitly_rejected:
        levers.append("MD")
    if any(k in low for k in ["top restaurant", "destacado", "posicionamiento", "ranking", "visibilidad orgánica"]):
        levers.append("Top Restaurant")
    if any(k in low for k in ["menú", "menu", "catálogo", "fotos", "productos", "carta", "assortment"]):
        levers.append("Assortment")
    if any(k in low for k in ["cancelar", "dar de baja", "no quiero seguir", "churn", "me voy", "me retiro", "cerrar cuenta"]):
        levers.append("Churn Risk")
    result["levers"] = levers

    # ── Suggested status ─────────────────────────────────────────────────────
    if explicit_rejection or any(k in low for k in ["no me interesa", "no quiero", "no voy a", "lo dejo", "me retiro", "no tengo presupuesto", "no quiero invertir"]):
        result["suggested_status"] = "Rejected ❌"
    elif any(k in low for k in ["activamos", "cerramos", "lo hacemos", "confirmado", "arrancamos", "activar"]):
        result["suggested_status"] = "Deal Closed 🏆"
    elif any(k in low for k in ["lo pienso", "lo consulto", "te llamo", "negociando", "pendiente", "voy a ver", "la próxima semana"]):
        result["suggested_status"] = "Negotiation ⏳"
    elif any(k in low for k in ["no contestó", "no atendió", "no disponible", "buzón", "voicemail", "no answer", "cortó"]):
        result["suggested_status"] = "Ghost 👻"
    else:
        result["suggested_status"] = "Follow-up ✅"

    # ── Action items ──────────────────────────────────────────────────────────
    action_patterns = [
        r"(enviar[á]?[a-záéíóúñü ]{3,40}(?:propuesta|plantilla|mail|información|info|cotización))",
        r"(llamar[a-záéíóúñü ]{0,20}(?:la próxima semana|mañana|el [a-záéíóúñü]+))",
        r"(agendar[a-záéíóúñü ]{0,30})",
        r"(revisar[a-záéíóúñü ]{3,40})",
        r"(confirmar[a-záéíóúñü ]{3,40})",
        r"(activar[a-záéíóúñü ]{3,40})",
        r"(subir[a-záéíóúñü ]{3,30}(?:fotos|menú|productos|catálogo))",
    ]
    actions = []
    for pat in action_patterns:
        matches = re.findall(pat, low)
        for m in matches:
            clean_action = m.strip().capitalize()
            if len(clean_action) > 8 and clean_action not in actions:
                actions.append(clean_action)
    result["action_items"] = actions[:5]  # max 5

    # ── Auto-summary (note that replaces New Comment) ─────────────────────────
    lever_str = ", ".join(levers) if levers else "sin palanca específica"
    action_str = " · ".join(actions[:2]) if actions else "sin próximos pasos detectados"
    status_emoji = result["suggested_status"]
    sentiment_map = {"Positive": "✅ Positivo", "Neutral": "➡️ Neutral", "Negative": "⚠️ Negativo"}
    sentiment_str = sentiment_map.get(result["sentiment"], "➡️ Neutral")

    result["summary"] = (
        f"[Auto] {sentiment_str} · Palancas: {lever_str} · "
        f"Status sugerido: {status_emoji} · Próximos pasos: {action_str}"
    )

    return result


# ── Detección automática de objeciones para el Role Play Trainer ────────────
_OBJECTION_LEVER_KEYWORDS = {
    "Ads":   ["ads", "publicidad", "banner", "sponsored", "visibilidad paga", "campaña", "presupuesto", "inversión", "invertir"],
    "MD":    ["descuento", "promo", "markdown", "porcentaje", "oferta", "promoción", "precio"],
    "Churn": ["cancelar", "dar de baja", "no quiero seguir", "churn", "me voy", "me retiro", "cerrar cuenta"],
}

_OBJECTION_REJECTION_PHRASES = [
    "no me interesa", "no quiero", "no puedo", "no tengo presupuesto", "no tenemos margen",
    "no me sirve", "ya tengo otro proveedor", "no es el momento", "no estoy interesado",
    "no estoy interesada", "estamos bien así", "no voy a", "lo dejo", "me retiro",
    "no me llames más", "no me llame más", "no quiero seguir", "ya dije que no",
    "no quiero hablar de esto",
]


def _extract_objection_sentence(customer_text, phrase):
    """
    Devuelve la oración completa del cliente que contiene `phrase`, recortando
    por los delimitadores de oración más cercanos (. ! ? o salto de línea).
    Si no encuentra límites claros, devuelve una ventana de ~140 caracteres
    centrada en la frase, para que la objeción guardada tenga contexto legible.
    """
    idx = customer_text.find(phrase)
    if idx == -1:
        return phrase
    start = idx
    for sep in [".", "!", "?", "\n"]:
        p = customer_text.rfind(sep, 0, idx)
        if p != -1:
            start = max(start if start != idx else 0, p + 1)
    start = max(0, min(start, idx))
    end = idx + len(phrase)
    for sep in [".", "!", "?", "\n"]:
        p = customer_text.find(sep, end)
        if p != -1:
            end = min(end if end != idx + len(phrase) else len(customer_text), p + 1)
            break
    else:
        end = min(len(customer_text), idx + len(phrase) + 100)
    sentence = customer_text[start:end].strip()
    return sentence if sentence else phrase


def detect_and_save_objection_from_transcript(transcript, ideal_response_hint=""):
    """
    Analiza una transcripción y, si detecta un rechazo explícito a una palanca
    comercial (Ads / MD / Churn), guarda automáticamente la objeción en el banco
    del Role Play Trainer (ROLEPLAY_OBJECTIONS_FILE). No requiere ninguna API
    externa — usa el mismo set de frases de rechazo que ya usa el motor de
    sentimiento (_OBJECTION_REJECTION_PHRASES).

    Evita duplicados: si ya existe una objeción muy similar (mismo texto
    normalizado) en el banco, no vuelve a guardarla.

    Devuelve (saved: bool, objection_text: str | None).
    """
    if not transcript or not transcript.strip():
        return False, None

    customer_text, _ = _extract_customer_text(transcript)

    found_phrase = next((p for p in _OBJECTION_REJECTION_PHRASES if p in customer_text), None)
    if not found_phrase:
        return False, None

    objection_text = _extract_objection_sentence(customer_text, found_phrase)
    if len(objection_text.strip()) < 8:
        return False, None

    # ── Detectar palanca involucrada por las keywords presentes cerca del rechazo ──
    lever = "General"
    for lv, kws in _OBJECTION_LEVER_KEYWORDS.items():
        if any(k in customer_text for k in kws):
            lever = lv
            break

    # ── Evitar duplicados: comparar texto normalizado contra el banco existente ──
    try:
        if os.path.exists(ROLEPLAY_OBJECTIONS_FILE):
            existing = pd.read_csv(ROLEPLAY_OBJECTIONS_FILE, dtype=str).fillna("")
            existing_norm = existing.get("objection_text", pd.Series([], dtype=str)).apply(norm_text)
            if norm_text(objection_text) in set(existing_norm):
                return False, objection_text  # ya existe, no duplicar
        else:
            existing = pd.DataFrame(columns=["objection_id", "datetime", "objection_text", "lever", "category", "ideal_response", "tags"])
    except Exception:
        existing = pd.DataFrame(columns=["objection_id", "datetime", "objection_text", "lever", "category", "ideal_response", "tags"])

    new_row = pd.DataFrame([{
        "objection_id": str(uuid.uuid4()),
        "datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "objection_text": objection_text.strip().capitalize(),
        "lever": lever,
        "category": "General",
        "ideal_response": ideal_response_hint.strip(),
        "tags": "auto-detectada",
    }])
    combined = pd.concat([existing, new_row], ignore_index=True)
    combined.to_csv(ROLEPLAY_OBJECTIONS_FILE, index=False, encoding="utf-8-sig")
    return True, objection_text


def evaluate_and_save_call_detail(transcript, brand_id, brand_name, farmer_email, call_date, wb=None):
    """
    Evaluates a call transcript using pure Python keyword/regex logic (no API),
    then writes the full Call Detail matrix row to the Excel file.
    Runs silently — no UI output.

    If `wb` (an already-open openpyxl Workbook) is provided, reuses it and does
    NOT save/close — the caller owns the save/close lifecycle. This avoids
    opening the entire Excel file a second time when called right before the
    main Save Follow-up flow, which already opens its own workbook.
    """
    if not transcript or not transcript.strip():
        return

    customer_text, low = _extract_customer_text(transcript)

    # ── Helper: check any keyword present (sobre texto completo, para detectar
    # palancas mencionadas por cualquiera de los dos hablantes) ────────────────
    def has(keywords):
        return 1 if any(k in low for k in keywords) else 0

    def has_pattern(patterns):
        return 1 if any(re.search(p, low) for p in patterns) else 0

    def avg(*vals):
        return round(sum(vals) / len(vals), 2)

    # ── SENTIMENT (solo sobre lo que dice el cliente) ──────────────────────────
    sentiment, _explicit_rejection_detail = _detect_sentiment_weighted(customer_text)
    pos = sum(1 for k in ["perfecto", "de acuerdo", "vamos", "sí claro", "me interesa",
                           "lo hacemos", "dale", "genial", "excelente", "cerramos",
                           "activamos", "okay", "ok", "buenísimo", "bárbaro",
                           "confirmado", "listo"] if k in customer_text)

    # ── INTRODUCCIÓN ──────────────────────────────────────────────────────────
    identified = has([f"soy {FARMER_FIRST_NAME.lower()}", "soy de rappi", "te llamo de rappi", "mi nombre es",
                       f"habla {FARMER_FIRST_NAME.lower()}", "habla el farmer", "soy el farmer"])
    named_brand = has(["rappi", "rappi ads", "la plataforma"])
    said_hello  = has(["hola", "buenos días", "buenas tardes", "buenas noches", "buen día", "hey"])
    pct_intro   = avg(identified, named_brand, said_hello)
    intro_comment = (
        "Introducción completa: se identificó, nombró Rappi y saludó." if pct_intro == 1.0
        else "Introducción parcial — faltó alguno de: identificarse, nombrar Rappi o saludar."
        if pct_intro > 0 else "No se detectó introducción clara."
    )

    # ── MANEJO DE LLAMADA (In Control) ────────────────────────────────────────
    in_control = has(["vamos a ver", "te propongo", "la idea es", "lo que hicimos",
                       "lo que sugiero", "entonces lo que hacemos", "lo que vamos a hacer",
                       "mi propuesta", "te recomiendo", "arranquemos"])
    # Penalize if agent got sidetracked by irrelevant topics
    lost_control = has(["no sé cómo funciona", "no tengo esa info", "no puedo decirte",
                         "no tengo acceso", "tendría que consultar"])
    in_control_final = 1 if (in_control and not lost_control) else (0 if lost_control else in_control)
    pct_handling = in_control_final
    handling_comment = (
        "El farmer mantuvo el control de la conversación y guió hacia los objetivos."
        if in_control_final else
        "No se evidenció claramente que el farmer llevara la conversación hacia sus metas."
    )

    # ── INTERACCIÓN CON ALIADO ────────────────────────────────────────────────
    effective_comm   = has(["entendido", "claro", "sí, entiendo", "ya veo", "de acuerdo",
                             "tiene sentido", "me parece", "sí sí"])
    handle_unknown   = has(["te averiguo", "te consulto", "lo escalo", "te mando la info",
                             "te envío", "pregunto y te confirmo", "no tengo ese dato pero"])
    no_interruption  = 1 if not has(["perdón, te corto", "te interrumpo", "un momento",
                                      "espera", "para, para"]) else 0
    pct_partner      = avg(effective_comm, handle_unknown, no_interruption)
    partner_comment  = (
        "Buena interacción con el aliado: comunicación fluida, sin interrupciones."
        if pct_partner >= 0.66
        else "Interacción con el aliado mejorable — revisar manejo de respuestas y posibles interrupciones."
    )

    # ── CIERRE / EXECUTIVE SUMMARY ────────────────────────────────────────────
    exec_summary = has(["entonces quedamos en", "como acordamos", "para resumir",
                         "los próximos pasos son", "lo que vamos a hacer es",
                         "te confirmo que", "cerramos con", "el plan es"])
    dm_confirm   = has(["sos el dueño", "sos quien decide", "hablo con el encargado",
                         "sos el que maneja", "el titular", "el responsable",
                         "sos quien toma las decisiones"])
    respond_all  = has(["respondido", "te expliqué", "ya te conté", "como te dije",
                         "te aclaré", "ya te respondí"])
    self_svc     = has(["en la app", "desde el portal", "podés hacerlo vos",
                         "en el panel", "self service", "autogestión", "lo hacés solo"])
    exec_comment = (
        "El farmer cerró con resumen de próximos pasos y confirmó el plan."
        if exec_summary else
        "No se detectó un cierre con resumen claro de próximos pasos."
    )
    call_summary_text = (
        "Llamada cerrada con acuerdo y próximos pasos definidos." if exec_summary and (pos >= 1)
        else "Llamada cerrada sin acuerdo claro o próximos pasos pendientes."
    )

    # ── PALANCAS COMERCIALES ──────────────────────────────────────────────────
    # TOP RESTAURANT
    top_rest_present = has(["top restaurant", "destacado", "posicionamiento",
                             "visibilidad orgánica", "ranking de restaurantes",
                             "aparecer primero", "mejor posición"])
    top_rest_action  = has(["activamos top", "vamos con top restaurant",
                             "subimos el posicionamiento", "arrancamos con el destacado"]) if top_rest_present else 0
    top_rest_comment = (
        "Se trabajó Top Restaurant con plan de acción." if top_rest_action
        else "Se mencionó Top Restaurant pero sin plan concreto." if top_rest_present
        else "No se trabajó la palanca Top Restaurant."
    )

    # ADS / INVESTMENT
    ads_present      = has(["rappi ads", "ads", "publicidad", "banner",
                             "sponsored", "visibilidad paga", "campaña paga",
                             "inversión en publicidad"])
    ads_action       = has_pattern([
        r"(presupuesto|budget).{0,30}(ads|publicidad)",
        r"(ads|publicidad).{0,30}(presupuesto|budget|pesos|ars|\$)",
        r"(arranc|activ|empez).{0,20}(ads|campaña|publicidad)",
        r"\$\s*\d+.{0,10}(ads|publicidad|campaña)",
    ]) if ads_present else 0
    investment_present = ads_present
    pct_investment   = ads_present
    invest_comment   = (
        "Se trabajó ADS con propuesta de presupuesto o plan concreto." if ads_action
        else "Se mencionó ADS pero sin plan ni presupuesto concreto." if ads_present
        else "No se trabajó la palanca ADS / Inversión."
    )

    # MARKDOWN / DESCUENTO
    md_present = has(["descuento", "promo", "markdown", "20%", "25%", "30%",
                       "oferta", "promoción", "deal", "porcentaje de descuento",
                       "precio especial"])

    # CHURN
    churn_present = has(["cancelar", "dar de baja", "no quiero seguir", "churn",
                          "me voy", "me retiro", "cerrar la cuenta",
                          "dejar de usar rappi", "no vale la pena"])
    churn_action  = has(["te propongo retener", "vamos a ayudarte", "podemos ofrecerte",
                          "no te vayas", "qué necesitás para quedarte",
                          "hagamos algo para que continúes"]) if churn_present else 0
    churn_comment = (
        "Se detectó riesgo de churn y se trabajó con plan de retención." if churn_action
        else "Se detectó riesgo de churn pero sin plan de retención claro." if churn_present
        else "No se detectó ni trabajó riesgo de churn."
    )

    # ASSORTMENT / MENÚ
    assortment_present = has(["menú", "menu", "catálogo", "fotos", "productos",
                                "carta", "assortment", "platos", "opciones del menú"])
    assortment_action  = has(["subimos fotos", "actualizamos el menú", "agregamos productos",
                                "mejoramos el catálogo", "activamos el menú",
                                "completamos el catálogo"]) if assortment_present else 0
    assortment_comment = (
        "Se trabajó el catálogo/menú con acción concreta." if assortment_action
        else "Se mencionó el catálogo/menú pero sin acción concreta." if assortment_present
        else "No se trabajó la palanca de catálogo/menú."
    )

    # ── %EC / %ENC ────────────────────────────────────────────────────────────
    # EC = efectividad de contacto: cuántas dimensiones clave se cumplieron
    ec_items = [identified, named_brand, said_hello, in_control_final,
                 effective_comm, exec_summary, dm_confirm,
                 ads_present, md_present, top_rest_present]
    pct_ec = round(sum(ec_items) / len(ec_items), 2)

    # ENC = efectividad de no contacto: aplica si no se llegó al decision maker
    enc_items = [handle_unknown, self_svc, no_interruption]
    pct_enc = round(sum(enc_items) / len(enc_items), 2) if not dm_confirm else 0.0

    # ── ACTION ITEMS (texto) ──────────────────────────────────────────────────
    action_patterns = [
        r"(enviar[á]?[a-záéíóúñü ]{3,40}(?:propuesta|plantilla|mail|información|info|cotización))",
        r"(llamar[a-záéíóúñü ]{0,20}(?:la próxima semana|mañana|el [a-záéíóúñü]+))",
        r"(agendar[a-záéíóúñü ]{0,30})",
        r"(activar[a-záéíóúñü ]{3,40})",
        r"(confirmar[a-záéíóúñü ]{3,40})",
        r"(subir[a-záéíóúñü ]{3,30}(?:fotos|menú|productos|catálogo))",
    ]
    actions_found = []
    for pat in action_patterns:
        for m in re.findall(pat, low):
            clean_a = m.strip().capitalize()
            if len(clean_a) > 8 and clean_a not in actions_found:
                actions_found.append(clean_a)
    action_items_text = "\n".join(actions_found[:5]) if actions_found else ""

    # ── SUMMARY auto-text ─────────────────────────────────────────────────────
    levers_found = []
    if ads_present:       levers_found.append("ADS")
    if md_present:        levers_found.append("MD")
    if top_rest_present:  levers_found.append("Top Restaurant")
    if assortment_present: levers_found.append("Assortment")
    if churn_present:     levers_found.append("Churn")
    summary_text = (
        f"Llamada {sentiment.lower()}. "
        f"Palancas trabajadas: {', '.join(levers_found) if levers_found else 'ninguna detectada'}. "
        f"{'Cerró con próximos pasos.' if exec_summary else 'Sin cierre claro.'}"
    )

    # ── BUILD RESULT DICT ─────────────────────────────────────────────────────
    result = {
        "Call Sentiment":              sentiment,
        "Summary":                     summary_text,
        "Action Items":                action_items_text,
        "Feature Requests":            "",
        "Introduction Comment":        intro_comment,
        "%Introduction":               pct_intro,
        "%Identified Himself":         identified,
        "%Named Brand":                named_brand,
        "%Said Hello":                 said_hello,
        "Handling Comment":            handling_comment,
        "%Call Handling":              pct_handling,
        "In Control":                  in_control_final,
        "Partner Interaction Comment": partner_comment,
        "Effective Communication":     effective_comm,
        "Handle Unknown Responses":    handle_unknown,
        "No Interruption":             no_interruption,
        "Executive Summary Comment":   exec_comment,
        "Respond to All Questions":    respond_all,
        "Self Service Info":           self_svc,
        "Call Summary":                call_summary_text,
        "%Partner Interaction":        pct_partner,
        "%Exec. Summary Provided":     exec_summary,
        "%Decision Maker Confirm.":    dm_confirm,
        "Top Rest Comment":            top_rest_comment,
        "Top Rest Subject Present":    top_rest_present,
        "%Top Rest Action Plan":       top_rest_action,
        "Investment Comment":          invest_comment,
        "%Investment":                 pct_investment,
        "Investment Subject Present":  investment_present,
        "ADS Subject Present":         ads_present,
        "%ADS Action Plan":            ads_action,
        "MD Subject Present":          md_present,
        "Churn Comment":               churn_comment,
        "%Churn Subject Present":      churn_present,
        "%Churn Action Plan":          churn_action,
        "%Self Service Info":          self_svc,
        "Assortment Comment":          assortment_comment,
        "Assortment Subject Present":  assortment_present,
        "%Assortment":                 assortment_action,
        "%EC":                         pct_ec,
        "%ENC":                        pct_enc,
    }

    # ── WRITE TO EXCEL ────────────────────────────────────────────────────────
    _wb_was_injected = wb is not None
    try:
        if wb is None:
            wb = openpyxl.load_workbook(EXCEL_FILE)
        if "Call Detail" not in wb.sheetnames:
            return
        ws = wb["Call Detail"]
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]

        new_row_data = {
            "Call Sentiment":              result["Call Sentiment"],
            "Líder":                       "",
            "Country":                     "AR",
            "Farmer":                      farmer_email,
            "Rol":                         "Farmer Jr",
            "Call Date":                   call_date,
            "Start":                       datetime.now(),
            "End":                         datetime.now(),
            "Duration":                    "",
            "Duration (s)":                0,
            "Destination Number":          "",
            "Country Brand ID":            str(brand_id),
            "%Interaction Success":        result["%EC"],
            **result,
            "Caller ID":                   "",
            "Disconnected By":             "agent",
            "Log ID":                      str(uuid.uuid4()),
        }

        ws.append([new_row_data.get(h, None) for h in headers])
        if not _wb_was_injected:
            wb.save(EXCEL_FILE)
    except Exception:
        return  # Falla silenciosamente



def _load_comments_df():
    if not os.path.exists(COMMENTS_FILE):
        return pd.DataFrame()
    try:
        df = pd.read_csv(COMMENTS_FILE, encoding="utf-8-sig")
    except Exception:
        return pd.DataFrame()
    if df.empty:
        return df
    for col in ["brand_id", "brand_name", "comment", "contact_channel", "opportunity_status", "commercial_action"]:
        if col not in df.columns:
            df[col] = ""
    df["brand_id"] = df["brand_id"].apply(normalize_brand_id)
    df["_dt"] = pd.to_datetime(df.get("datetime"), errors="coerce")
    return df


def get_saved_comments(brand_id):
    comments = _load_comments_df()
    if comments.empty or "brand_id" not in comments.columns:
        return ""

    brand_comments = comments[comments["brand_id"] == normalize_brand_id(brand_id)].copy()

    if brand_comments.empty:
        return ""

    brand_comments = brand_comments.sort_values(by="_dt", ascending=True, na_position="last")
    return "\n\n".join(
        f"[{clean(r.get('datetime'), '-')}] {clean(r.get('comment'), '')}"
        for _, r in brand_comments.iterrows()
        if clean(r.get('comment'), '').strip()
    )


def get_last_comments_map(limit=2):
    comments = _load_comments_df()
    if comments.empty:
        return {}
    comments = comments.sort_values(by="_dt", ascending=True, na_position="last")
    result = {}
    for bid, group in comments.groupby("brand_id"):
        notes = [clean(x, "").strip() for x in group["comment"].tolist() if clean(x, "").strip()]
        result[bid] = " | ".join(notes[-limit:])
    return result


@st.cache_data(ttl=300, show_spinner=False)
def _load_productivity_sheet_raw(excel_path):
    """
    Loader único y cacheado de la hoja 'Productivity' completa.
    Las 5 funciones que antes leían esta hoja por separado (contact stats,
    last contact map, levers por marca, productivity heatmap) ahora reutilizan
    este DataFrame en vez de volver a golpear el disco cada vez. TTL de 5 min:
    suficiente para reflejar cambios recientes sin releer en cada rerun de
    Streamlit (que ocurre en cada click, filtro o tipificación).
    """
    if not os.path.exists(excel_path):
        return pd.DataFrame()
    try:
        raw = pd.read_excel(excel_path, sheet_name="Productivity", header=0)
    except Exception as e:
        _log_data_issue('Productivity', e, 'Contactos y palancas salen de esta hoja.')
        return pd.DataFrame()
    raw.columns = [str(c).strip() for c in raw.columns]
    return raw


def _load_productivity_contact_stats(excel_path, start_date=CONTACTS_START_DATE):
    """
    Reads the Productivity sheet and counts contacts by channel:
      - Amazon Connect  -> calls
      - WhatsApp        -> chats
      - Videoconferencia -> meets
    No Contactado comes from columna 'Contactado?' == 'NO'.
    Filters by Date column >= start_date (June 1 2026).
    """
    if not os.path.exists(excel_path):
        return None
    raw = _load_productivity_sheet_raw(excel_path)
    if raw.empty:
        return None
    raw = raw.copy()  # evita mutar el DataFrame cacheado compartido

    # Keep original column names (stripped) for positional access
    raw.columns = [str(c).strip() for c in raw.columns]
    cols_lower = [c.lower() for c in raw.columns]

    # Resolve columns by NAME first (most reliable), fall back to position
    # Col C (idx 2)  = Medio de Contacto
    # Col F (idx 5)  = Fase  → "Aliado no contactado" = No Answer
    # Col I (idx 8)  = Month
    # Col J (idx 9)  = Week
    # Col K (idx 10) = Date
    medio_col = next((raw.columns[i] for i, c in enumerate(cols_lower) if "medio de contacto" in c), None)
    # Fallback by position if name search failed
    if not medio_col and len(raw.columns) > 2:
        medio_col = raw.columns[2]

    # Columna F (índice 5) = Fase — "Aliado no contactado" identifica No Answer
    fase_col = next((raw.columns[i] for i, c in enumerate(cols_lower) if c == "fase"), None)
    if not fase_col and len(raw.columns) > 5:
        fase_col = raw.columns[5]

    date_col = next((raw.columns[i] for i, c in enumerate(cols_lower) if c == "date"), None)
    week_col = next((raw.columns[i] for i, c in enumerate(cols_lower) if c == "week"), None)

    if not medio_col:
        return None

    df = raw.copy()

    # Filter from start_date (June 1) using Date column, fall back to Week
    if date_col:
        df["_date_dt"] = pd.to_datetime(df[date_col], errors="coerce")
        df = df[df["_date_dt"].notna() & (df["_date_dt"] >= pd.Timestamp(start_date))].copy()
    elif week_col:
        df["_week_dt"] = pd.to_datetime(df[week_col], errors="coerce")
        df = df[df["_week_dt"].notna() & (df["_week_dt"] >= pd.Timestamp(start_date))].copy()

    # Count No Answer: col F (Fase) == "Aliado no contactado" — ALL rows after date filter
    if fase_col and fase_col in df.columns:
        df["_fase"] = df[fase_col].astype(str).str.strip().str.lower()
        not_cont = int(df["_fase"].str.contains("aliado no contactado", case=False, na=False).sum())
        # Effective rows = those NOT marked as "Aliado no contactado"
        df_effective = df[~df["_fase"].str.contains("aliado no contactado", case=False, na=False)].copy()
    else:
        not_cont = 0
        df_effective = df.copy()

    # Channel counting on effective (SI) rows only
    df_effective = df_effective[df_effective[medio_col].notna()].copy()
    df_effective["_medio"] = df_effective[medio_col].astype(str).str.strip().str.lower()

    total_effective = len(df_effective)
    calls = int(df_effective["_medio"].str.contains("amazon connect|amazon", case=False, na=False).sum())
    chats = int(df_effective["_medio"].str.contains("whatsapp", case=False, na=False).sum())
    meets = int(df_effective["_medio"].str.contains("videoconferencia|videoconf|video", case=False, na=False).sum())

    return {
        "total_effective": total_effective,
        "calls": calls,
        "chats": chats,
        "meets": meets,
        "not_contacted": not_cont,
        "source": "productivity_sheet",
    }


def get_comment_contact_stats(start_date=CONTACTS_START_DATE, fallback_total=0):
    """
    Contact performance stats.
    Primary source: Productivity sheet · column "medio de contacto"
      Amazon Connect → calls, WhatsApp → chats, Videoconferencia → meets.
    Fallback: comments CSV (legacy channel tagging).
    not_contacted is always read from the comments CSV.
    """
    # ── Primary: Productivity sheet ───────────────────────────────────────────
    prod_stats = _load_productivity_contact_stats(EXCEL_FILE, start_date)

    # ── Primary: use not_contacted from Productivity sheet (col F = Fase) ────
    # When Productivity data is available, not_contacted comes directly from it
    # (rows where Fase == "Aliado no contactado"). The comments CSV is only a
    # fallback for when the Productivity sheet is unavailable.
    if prod_stats is not None:
        return prod_stats

    # ── Fallback: comments CSV (only when Productivity sheet is unavailable) ──
    comments = _load_comments_df()
    not_contacted = 0
    if not comments.empty and "_dt" in comments.columns:
        start_ts = pd.Timestamp(start_date)
        recent = comments[comments["_dt"].notna() & (comments["_dt"] >= start_ts)].copy()
        if not recent.empty:
            recent["_not_contacted"] = recent.apply(
                lambda r: _is_not_contacted_text(r.get("comment", "")) or _is_not_contacted_text(r.get("opportunity_status", "")),
                axis=1,
            )
            not_contacted = int(recent["_not_contacted"].sum())

    # ── Fallback: comments CSV ────────────────────────────────────────────────
    empty_stats = {
        "total_effective": int(fallback_total or 0),
        "calls": 0,
        "chats": 0,
        "meets": 0,
        "not_contacted": not_contacted,
        "source": "baseline" if fallback_total else "empty",
    }
    if comments.empty or "_dt" not in comments.columns:
        return empty_stats

    start_ts = pd.Timestamp(start_date)
    recent = comments[comments["_dt"].notna() & (comments["_dt"] >= start_ts)].copy()
    if recent.empty:
        return empty_stats

    recent["_effective"] = recent.apply(
        lambda r: _is_effective_comment(r.get("comment", ""), r.get("opportunity_status", "")),
        axis=1,
    )
    effective = recent[recent["_effective"]].copy()
    channel   = effective.get("contact_channel", pd.Series([], dtype=str)).astype(str).str.lower()

    return {
        "total_effective": int(len(effective)),
        "calls": int(channel.str.contains("call|llamada|zoho", regex=True, na=False).sum()),
        "chats": int(channel.str.contains("whatsapp|chat|treble|msg|message", regex=True, na=False).sum()),
        "meets": int(channel.str.contains("meet|reunion|reunión|video", regex=True, na=False).sum()),
        "not_contacted": not_contacted,
        "source": "comments_csv",
    }



# =========================
# ACQUISITION TRACKER HELPERS
# =========================

TEST_TRACKER_BRANDS = ["chipacala", "conos baikatsu", "tutotano"]
ACQUISITION_TRACKER_RESET_TAG = "2026-05-21"
ACQUISITION_TRACKER_RESET_MARKER = f".growth_os_acquisition_tracker_reset_{ACQUISITION_TRACKER_RESET_TAG}.done"
ACQUISITION_TRACKER_COLUMNS = [
    "datetime", "date", "week_start", "brand_id", "brand_name", "pipeline_stage",
    "type", "movement", "commercial_action", "negotiation_type", "ads_booking_ars",
    "md_discount", "rejection_reason", "total_ars", "total_usd", "opportunity_status", "comment",
]


def reset_acquisition_tracker_once():
    """
    Resets the Acquisition Tracker one time for the new clean cycle.
    The old CSV is backed up before the tracker is emptied.
    """
    if os.path.exists(ACQUISITION_TRACKER_RESET_MARKER):
        return False

    try:
        os.makedirs(BACKUP_FOLDER, exist_ok=True)
        if os.path.exists(ACQUISITION_TRACKER_FILE):
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"growth_os_acquisition_tracker_backup_before_reset_{stamp}.csv"
            backup_path = os.path.join(BACKUP_FOLDER, backup_name)
            shutil.copy2(ACQUISITION_TRACKER_FILE, backup_path)

        pd.DataFrame(columns=ACQUISITION_TRACKER_COLUMNS).to_csv(
            ACQUISITION_TRACKER_FILE,
            index=False,
            encoding="utf-8-sig",
        )

        with open(ACQUISITION_TRACKER_RESET_MARKER, "w", encoding="utf-8") as marker:
            marker.write(f"Reset completed on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        return True
    except Exception:
        return False


def _commercial_action_type(commercial_action):
    text = clean(commercial_action, "")
    low = norm_text(text)
    if "ads" in low and ("markdown" in low or "md" in low):
        return "ADS + MD"
    if "ads" in low:
        return "ADS"
    if "markdown" in low or "md" in low or "mardan" in low:
        return "MD"
    return "Commercial"


def _commercial_action_movement(commercial_action):
    low = norm_text(commercial_action)
    if "reject" in low or "rechaz" in low or "❌" in commercial_action:
        return "Rejected"
    if "negotiation" in low or "negoci" in low or "⏳" in commercial_action:
        return "Negotiation"
    if "deactivate" in low or "desactivar" in low:
        return "Deactivate"
    if "upsell" in low or "upselling" in low or "🧉" in commercial_action:
        return "Upselling"
    if "activate" in low or "activar" in low:
        return "Acquisition"
    return "Update"


def _pipeline_stage_from_values(opportunity_status="", commercial_action="", pipeline_stage=""):
    stage = clean(pipeline_stage, "").strip()
    if stage and stage != "-":
        return stage

    status_low = norm_text(opportunity_status)
    action_low = norm_text(commercial_action)
    if "reject" in status_low or "rechaz" in status_low or "reject" in action_low or "rechaz" in action_low or "❌" in str(commercial_action):
        return "Rejected"
    if "negotiation" in status_low or "negoci" in status_low or "negotiation" in action_low or "negoci" in action_low or "⏳" in str(commercial_action):
        return "Negotiation"
    return "Closed"


def _tracker_is_test_brand(value):
    brand = norm_text(value)
    return any(test_brand in brand for test_brand in TEST_TRACKER_BRANDS)


def cleanup_test_tracker_records():
    """
    Removes old test rows from the CSV tracker.
    This is intentionally limited to the three test brands Sabas asked to delete.
    """
    if not os.path.exists(ACQUISITION_TRACKER_FILE):
        return 0

    try:
        df = pd.read_csv(ACQUISITION_TRACKER_FILE)
    except Exception:
        return 0

    if df.empty or "brand_name" not in df.columns:
        return 0

    mask = df["brand_name"].apply(_tracker_is_test_brand)
    removed = int(mask.sum())

    if removed:
        df = df[~mask].copy()
        df.to_csv(ACQUISITION_TRACKER_FILE, index=False, encoding="utf-8-sig")

    return removed


def normalize_markdown_discount(value):
    """Keeps Markdown as discount/promo label, never as money.
    Examples: 15 -> 15%, "15" -> 15%, "15%" -> 15%, "2x1" -> 2x1.
    """
    text = clean(value, "").strip()
    if text in ["", "-", "0", "0.0"]:
        return ""
    low = norm_text(text)
    if "2x1" in low or "2 x 1" in low or "two for one" in low:
        return "2x1"
    if "%" in text:
        return text
    try:
        number = float(str(text).replace(",", "."))
        if number.is_integer():
            number = int(number)
        if number > 0 and number <= 100:
            return f"{number}%"
    except Exception:
        pass
    return text


def save_acquisition_tracker_event(
    brand_id,
    brand_name,
    commercial_action,
    ads_budget_ars=0,
    md_discount="",
    opportunity_status="",
    comment="",
    pipeline_stage="",
    negotiation_type="",
    rejection_reason="",
):
    reset_acquisition_tracker_once()

    if clean(commercial_action, "") in ["", "No commercial change"]:
        return False, "No commercial action to track."

    ads_budget_ars = to_number(ads_budget_ars, 0)
    md_discount = normalize_markdown_discount(md_discount)
    stage = _pipeline_stage_from_values(opportunity_status, commercial_action, pipeline_stage)

    # Ads is measured in money. Markdown is measured as discount/promo label.
    # Therefore total_ars only includes Ads booking / Ads budget in negotiation, never MD percentage.
    total_ars = ads_budget_ars

    new_row = pd.DataFrame([{
        "datetime": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "date": date.today().strftime("%Y-%m-%d"),
        "week_start": (date.today() - timedelta(days=date.today().weekday())).strftime("%Y-%m-%d"),
        "brand_id": normalize_brand_id(brand_id),
        "brand_name": brand_name,
        "pipeline_stage": stage,
        "type": _commercial_action_type(commercial_action),
        "movement": _commercial_action_movement(commercial_action),
        "commercial_action": commercial_action,
        "negotiation_type": clean(negotiation_type, ""),
        "ads_booking_ars": ads_budget_ars,
        "md_discount": md_discount,
        "rejection_reason": clean(rejection_reason, ""),
        "total_ars": total_ars,
        "total_usd": total_ars / ARS_PER_USD if total_ars else 0,
        "opportunity_status": opportunity_status,
        "comment": comment,
    }])

    if os.path.exists(ACQUISITION_TRACKER_FILE):
        old = pd.read_csv(ACQUISITION_TRACKER_FILE)

        # Backward compatibility: old files may have md_booking_ars; from now on,
        # MD is tracked as md_discount instead of money.
        if "md_discount" not in old.columns:
            old["md_discount"] = ""
        if "pipeline_stage" not in old.columns:
            old["pipeline_stage"] = old.apply(
                lambda r: _pipeline_stage_from_values(r.get("opportunity_status", ""), r.get("commercial_action", "")),
                axis=1
            )
        if "negotiation_type" not in old.columns:
            old["negotiation_type"] = ""
        if "rejection_reason" not in old.columns:
            old["rejection_reason"] = ""
        if "week_start" not in old.columns:
            old["_dt_tmp"] = pd.to_datetime(old.get("datetime"), errors="coerce")
            old["week_start"] = old["_dt_tmp"].apply(
                lambda x: (x.date() - timedelta(days=x.date().weekday())).strftime("%Y-%m-%d") if not pd.isna(x) else ""
            )
            old = old.drop(columns=["_dt_tmp"], errors="ignore")

        for col in new_row.columns:
            if col not in old.columns:
                old[col] = ""
        final = pd.concat([old[new_row.columns], new_row], ignore_index=True)
    else:
        final = new_row

    final.to_csv(ACQUISITION_TRACKER_FILE, index=False, encoding="utf-8-sig")
    return True, "Acquisition Tracker updated."


def _load_acquisition_tracker_df():
    reset_acquisition_tracker_once()
    cleanup_test_tracker_records()

    if not os.path.exists(ACQUISITION_TRACKER_FILE):
        return pd.DataFrame()
    try:
        df = pd.read_csv(ACQUISITION_TRACKER_FILE)
    except Exception:
        return pd.DataFrame()
    if df.empty:
        return df

    df["_dt"] = pd.to_datetime(df.get("datetime"), errors="coerce")

    if "md_discount" not in df.columns:
        df["md_discount"] = ""
    if "negotiation_type" not in df.columns:
        df["negotiation_type"] = ""
    if "rejection_reason" not in df.columns:
        df["rejection_reason"] = ""
    if "pipeline_stage" not in df.columns:
        df["pipeline_stage"] = df.apply(
            lambda r: _pipeline_stage_from_values(r.get("opportunity_status", ""), r.get("commercial_action", "")),
            axis=1
        )
    if "week_start" not in df.columns:
        df["week_start"] = df["_dt"].apply(
            lambda x: (x.date() - timedelta(days=x.date().weekday())).strftime("%Y-%m-%d") if not pd.isna(x) else ""
        )

    df["pipeline_stage"] = df.apply(
        lambda r: _pipeline_stage_from_values(r.get("opportunity_status", ""), r.get("commercial_action", ""), r.get("pipeline_stage", "")),
        axis=1
    )

    for col in ["ads_booking_ars", "total_ars", "total_usd"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        else:
            df[col] = 0

    if "total_ars" not in df.columns and "ads_booking_ars" in df.columns:
        df["total_ars"] = df["ads_booking_ars"]

    return df


def _tracker_metric_count(df, col, pattern):
    if df.empty or col not in df.columns:
        return 0
    return int(df[col].astype(str).str.contains(pattern, case=False, na=False, regex=True).sum())


def page_acquisition_tracker():
    render_header("Acquisition Tracker", "Closed actions and negotiation pipeline")

    df = _load_acquisition_tracker_df()
    tracker_is_empty = df.empty
    if tracker_is_empty:
        # Keep the tracker visually complete even after reset: metrics and tables stay visible at zero.
        df = pd.DataFrame(columns=ACQUISITION_TRACKER_COLUMNS)
        df["_dt"] = pd.to_datetime(pd.Series([], dtype="datetime64[ns]"))

    today_ts = pd.Timestamp(date.today())
    month_start = pd.Timestamp(date.today().replace(day=1))
    week_start_date = date.today() - timedelta(days=date.today().weekday())
    week_start = pd.Timestamp(week_start_date)

    if tracker_is_empty:
        st.info("Acquisition Tracker reset: no records yet. Metrics and tables below start at 0 and will fill from new Brand Finder actions.")

    month_df = df[df["_dt"].notna() & (df["_dt"] >= month_start)].copy()
    if month_df.empty:
        month_df = df.copy()

    week_df = df[df["_dt"].notna() & (df["_dt"] >= week_start) & (df["_dt"] <= today_ts + pd.Timedelta(days=1))].copy()

    stage_series = month_df["pipeline_stage"].astype(str).str.lower() if "pipeline_stage" in month_df.columns else pd.Series([], dtype=str)
    closed_df = month_df[~stage_series.isin(["negotiation", "rejected"])].copy()
    negotiation_df = month_df[stage_series == "negotiation"].copy()
    rejected_df = month_df[stage_series == "rejected"].copy()
    negotiation_week_df = week_df[week_df["pipeline_stage"].astype(str).str.lower() == "negotiation"].copy() if "pipeline_stage" in week_df.columns else pd.DataFrame(columns=month_df.columns)

    total_actions = len(closed_df)
    total_ads = closed_df["ads_booking_ars"].sum() if "ads_booking_ars" in closed_df.columns else 0

    md_actions = _tracker_metric_count(closed_df, "type", "MD")
    ads_actions = _tracker_metric_count(closed_df, "type", "ADS")

    acquisitions_count = _tracker_metric_count(closed_df, "movement", "Acquisition")
    upsellings_count = _tracker_metric_count(closed_df, "movement", "Upselling")

    negotiations_wtd = len(negotiation_week_df)
    negotiation_ads_budget_wtd = negotiation_week_df["ads_booking_ars"].sum() if "ads_booking_ars" in negotiation_week_df.columns else 0
    negotiation_ads_count = _tracker_metric_count(negotiation_week_df, "type", "ADS")
    negotiation_md_count = _tracker_metric_count(negotiation_week_df, "type", "MD")

    st.markdown("### Closed Actions")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.metric("Closed MTD", fmt_number(total_actions), help="Tracked closed commercial moves this month")
    with c2:
        st.metric("ADS Booked MTD", fmt_ars(total_ads), help="Ads is measured in ARS")
    with c3:
        st.metric("ADS / MD Activations", f"{fmt_number(ads_actions)} / {fmt_number(md_actions)}", help="Commercial actions by type")
    with c4:
        st.metric(
            "Acq / Upsell",
            f"{fmt_number(acquisitions_count)} / {fmt_number(upsellings_count)}",
            help="Acquisitions vs Upsellings tracked this month. Deactivations are excluded from this calculator."
        )

    st.markdown("### Negotiation Pipeline")
    n1, n2, n3, n4 = st.columns(4)
    with n1:
        st.metric("Negotiations WTD", fmt_number(negotiations_wtd), help="Negotiations opened this week")
    with n2:
        st.metric("ADS in Negotiation", fmt_ars(negotiation_ads_budget_wtd), help="Ads budget currently negotiated this week")
    with n3:
        st.metric("ADS / MD Negotiations", f"{fmt_number(negotiation_ads_count)} / {fmt_number(negotiation_md_count)}", help="Negotiation mix by lever")
    with n4:
        st.metric("Rejected MTD", fmt_number(len(rejected_df)), help="Rejected opportunities tracked this month")

    view_cols = [
        "datetime", "brand_id", "brand_name", "pipeline_stage", "type", "movement", "commercial_action",
        "negotiation_type", "ads_booking_ars", "md_discount", "rejection_reason", "opportunity_status", "comment"
    ]

    st.markdown("#### Closed Actions Detail")
    closed_view = closed_df.copy()
    for col in view_cols:
        if col not in closed_view.columns:
            closed_view[col] = ""
    closed_view = closed_view[view_cols].sort_values(by="datetime", ascending=False)
    _render_html_table(closed_view)

    st.markdown("#### Negotiation Pipeline Detail")
    negotiation_view = negotiation_df.copy()
    for col in view_cols:
        if col not in negotiation_view.columns:
            negotiation_view[col] = ""
    negotiation_view = negotiation_view[view_cols].sort_values(by="datetime", ascending=False)
    _render_html_table(negotiation_view)

    st.markdown("#### Rejected Opportunities")
    rejected_view = rejected_df.copy()
    for col in view_cols:
        if col not in rejected_view.columns:
            rejected_view[col] = ""
    rejected_view = rejected_view[view_cols].sort_values(by="datetime", ascending=False)
    _render_html_table(rejected_view)



# =========================
# TEMPLATE QUEUE HELPERS
# =========================

TEMPLATE_TYPES = [
    "None",
    "Presentación inicial",
    "Seguimiento",
    "Activar campañas",
    "No Contactado",
    "Churn / Chon",
]


def _template_status_tone(status):
    text = norm_text(status)
    if "not contacted" in text or "ghost" in text or "no contesta" in text:
        return {
            "label": "apertura",
            "line": "Estoy intentando ubicar el canal correcto para acompañar la gestión de la marca de forma ordenada.",
            "cta": "¿Me confirman por favor el contacto efectivo y una disponibilidad para revisar la marca?",
        }
    if "negotiation" in text or "negoci" in text:
        return {
            "label": "cierre",
            "line": "Retomo lo que venimos revisando para avanzar con una decisión concreta y no dejar la oportunidad abierta.",
            "cta": "¿Lo cerramos y dejamos definido el siguiente paso de implementación?",
        }
    if "activated" in text or "deal closed" in text or "activ" in text:
        return {
            "label": "ejecución",
            "line": "Con la activación encaminada, la idea es cuidar ejecución, seguimiento y próximos ajustes de performance.",
            "cta": "¿Validamos el arranque y dejamos definido el seguimiento de resultados?",
        }
    if "off" in text or "sleep" in text or "💤" in str(status) or "😴" in str(status):
        return {
            "label": "reactivación",
            "line": "Veo oportunidad de reactivar la marca con una gestión más enfocada y cuidando las palancas correctas.",
            "cta": "¿Revisamos una alternativa de reactivación para esta semana?",
        }
    if "rejected" in text or "rechaz" in text:
        return {
            "label": "puerta abierta",
            "line": "Entiendo que por ahora no se avance, pero quiero dejar identificadas las oportunidades para retomarlas cuando tenga sentido.",
            "cta": "¿Te parece si lo dejamos mapeado y lo retomamos más adelante?",
        }
    return {
        "label": "seguimiento",
        "line": "Estoy revisando la marca con enfoque 360 para mantener la gestión alineada entre operación, menú, promociones y visibilidad.",
        "cta": "¿Lo revisamos y definimos el siguiente paso?",
    }


def _action_dict(actions):
    result = {}
    for item in actions or []:
        area = clean(item.get("area"), "")
        if "OPS" in area:
            key = "ops"
        elif "Menu" in area:
            key = "menu"
        elif "MD" in area:
            key = "md"
        elif "Ads" in area:
            key = "ads"
        else:
            key = area.lower()
        result[key] = {
            "action": clean(item.get("action"), "Following"),
            "reason": clean(item.get("reason"), ""),
        }
    for key in ["ops", "menu", "md", "ads"]:
        result.setdefault(key, {"action": "Following", "reason": "Stable"})
    return result


def _short_360_summary(action_map):
    labels = {
        "ops": "OPS",
        "menu": "Menú",
        "md": "MD",
        "ads": "Ads",
    }
    parts = []
    for key in ["ops", "menu", "md", "ads"]:
        action = clean(action_map.get(key, {}).get("action"), "Following")
        if action == "Following":
            parts.append(f"{labels[key]}: Following")
        else:
            parts.append(f"{labels[key]}: {action}")
    return " · ".join(parts)


def _first_priority_action(action_map):
    # Order is intentional: don't scale commercial traffic if OPS/Menu has a blocking issue.
    for key in ["ops", "menu", "md", "ads"]:
        action = clean(action_map.get(key, {}).get("action"), "Following")
        if action and action != "Following":
            return key, action, clean(action_map.get(key, {}).get("reason"), "")
    return "all", "Following", "Las palancas principales se ven estables."


def _brand_template_context(row, brand_id, opportunity_status=""):
    name = strip_brand_id_prefix(clean(get_from_row(row, ["name", "brand name", "restaurant name"])))
    category_raw = clean(get_from_row(row, ["category"]))
    category, stickers = _split_category_and_stickers(category_raw)

    current = get_current_brand_metrics(brand_id) or {
        "gmv_ars": 0, "gmv_usd": 0, "aov_ars": 0, "aov_usd": 0, "orders": 0
    }
    ads_raw = get_current_ads_metrics(brand_id)
    md_raw = get_current_md_metrics(brand_id, pro=False)
    mdp_raw = get_current_md_metrics(brand_id, pro=True)
    ads_current, md_current, md_pro_current = _merge_growth_manual_status(row, ads_raw, md_raw, mdp_raw)
    booster = recommend_booster_for_brand(
        category,
        current.get("gmv_ars", 0),
        current.get("aov_ars", 0),
        get_from_row(row, ["cr %", "conversion rate", "conversion"], 0),
        get_from_row(row, ["pro users %", "pro %"], 0),
        ads_current,
        md_current,
    )
    actions = build_360_actions(name, category, ads_current, md_current, md_pro_current, booster)
    action_map = _action_dict(actions)

    return {
        "brand_id": normalize_brand_id(brand_id),
        "brand_name": name,
        "category": category,
        "stickers": stickers,
        "manager": clean(get_from_row(row, ["manager", "restaurant manager", "account manager"])),
        "assistant": clean(get_from_row(row, ["assistant"])),
        "email": clean(get_from_row(row, ["email"])),
        "contact": fmt_contact_number(get_from_row(row, ["contact number", "phone", "contact"])),
        "gmv_ars": current.get("gmv_ars", 0),
        "aov_ars": current.get("aov_ars", 0),
        "ads_active": bool(ads_current.get("active", False)),
        "ads_roi": to_number(ads_current.get("roi"), 0),
        "md_active": bool(md_current.get("active", False)),
        "md_roi": to_number(md_current.get("roi"), 0),
        "md_pro_active": bool(md_pro_current.get("active", False)),
        "md_pro_roi": to_number(md_pro_current.get("roi"), 0),
        "booster": booster,
        "actions": actions,
        "action_map": action_map,
        "churn": get_churn_status(normalize_brand_id(brand_id)) if brand_id else clean(get_from_row(row, ["churn", "churn status"], "-")),
        "status": opportunity_status,
        "tone": _template_status_tone(opportunity_status),
    }


def _template_subject(template_type, ctx):
    brand = ctx["brand_name"].upper()
    if template_type == "Presentación inicial":
        return f"PRESENTACIÓN COMERCIAL RAPPI | {brand}"
    if template_type == "Seguimiento":
        return f"SEGUIMIENTO COMERCIAL RAPPI | {brand}"
    if template_type == "Activar campañas":
        return f"ACTIVACIÓN DE CAMPAÑAS RAPPI | {brand}"
    if template_type == "No Contactado":
        return f"CONTACTO COMERCIAL RAPPI | {brand}"
    if template_type == "Churn / Chon":
        return f"REVISIÓN CHURN / CHON | {brand}"
    return f"GESTIÓN COMERCIAL RAPPI | {brand}"


def _campaign_angle(ctx):
    action_map = ctx["action_map"]
    _, first_action, first_reason = _first_priority_action(action_map)
    booster = ctx.get("booster", {})
    event = clean(booster.get("event"), "-")
    event_line = f" También veo una oportunidad de temporada con {event}." if event not in ["", "-"] else ""
    return first_action, first_reason, event_line


def _churn_reading(ctx):
    churn = clean(ctx.get("churn"), "-")
    status = clean(ctx.get("status"), "-")
    churn_text = norm_text(churn)

    if churn in ["", "-"]:
        risk = "sin estado Chon visible"
        action = "validar el estado actual de la marca"
    elif "w3" in churn_text or "🆘" in churn or "☠" in churn:
        risk = "riesgo alto / deterioro avanzado"
        action = "definir acción de recuperación prioritaria"
    elif "w2" in churn_text or "🚨" in churn:
        risk = "riesgo medio con alerta activa"
        action = "revisar causa del deterioro y acordar corrección"
    elif "w1" in churn_text or "⚠" in churn:
        risk = "alerta temprana"
        action = "prevenir que la marca siga deteriorándose"
    elif "off" in churn_text or "😴" in churn:
        risk = "marca apagada / inactiva"
        action = "evaluar reactivación"
    elif "on" in churn_text or "✅" in churn:
        risk = "marca activa"
        action = "mantener seguimiento para evitar caída"
    else:
        risk = churn
        action = "revisar el estado Chon y definir próximos pasos"

    if status not in ["", "-"]:
        second_data = f"Estado de gestión: {status}"
    else:
        second_data = f"Lectura Chon: {risk}"

    return churn, risk, action, second_data


def generate_template_messages(template_type, ctx, source_comment=""):
    brand = ctx["brand_name"]
    category = ctx["category"]
    tone = ctx["tone"]
    action_map = ctx["action_map"]
    summary = _short_360_summary(action_map)
    priority_key, priority_action, priority_reason = _first_priority_action(action_map)
    subject = _template_subject(template_type, ctx)

    greeting = f"Hola equipo de {brand},"

    if template_type == "Presentación inicial":
        email_body = f"""{greeting}

Mucho gusto. Soy {FARMER_NAME}, {FARMER_ROLE_INLINE} en Rappi.

A partir de ahora estaré acompañando la gestión comercial de {brand}. La idea es trabajar la marca con una mirada 360, revisando operación, menú, promociones, visibilidad y oportunidades de crecimiento dentro de la app.

{tone['line']}

Lectura inicial 360: {summary}.

{tone['cta']}

Quedo atento.

{FARMER_NAME}
{FARMER_ROLE}
Rappi"""
        whatsapp_body = f"""{greeting} Soy {FARMER_NAME}, {FARMER_ROLE_INLINE} en Rappi. A partir de ahora estaré acompañando la gestión comercial de {brand}. Ya tengo la marca mapeada con enfoque 360: {summary}. {tone['cta']}"""

    elif template_type == "Seguimiento":
        source_line = f"\n\nNota de seguimiento: {source_comment}" if clean(source_comment, "").strip() else ""
        email_body = f"""{greeting}

Retomo la gestión comercial de {brand} para avanzar con el siguiente paso.

{tone['line']}

La lectura 360 marca como prioridad: {priority_action}. Motivo: {priority_reason}.

Resumen 360: {summary}.{source_line}

{tone['cta']}

Quedo atento.

{FARMER_NAME}
Rappi"""
        whatsapp_body = f"""{greeting} Retomo la gestión de {brand}. Según la lectura 360, la prioridad sería: {priority_action}. Motivo: {priority_reason}. {tone['cta']}"""

    elif template_type == "Activar campañas":
        campaign_action, campaign_reason, event_line = _campaign_angle(ctx)
        email_body = f"""{greeting}

Revisando {brand}, veo oportunidad de trabajar una activación comercial con foco en crecimiento y performance.

La lectura 360 sugiere empezar por: {campaign_action}. Motivo: {campaign_reason}.{event_line}

Para cuidar la ejecución, la idea es alinear las palancas completas: OPS, menú, MD y Ads. Resumen actual: {summary}.

¿Validamos la campaña y definimos la estructura para activarla esta semana?

Quedo atento.

{FARMER_NAME}
Rappi"""
        whatsapp_body = f"""{greeting} Revisando {brand}, veo oportunidad de activar campaña. La lectura 360 sugiere empezar por {campaign_action}: {campaign_reason}. Resumen: {summary}. ¿Lo validamos para avanzar esta semana?"""

    elif template_type == "No Contactado":
        email_body = f"""{greeting}

Te estuve llamando para revisar algunos puntos importantes de la marca en Rappi y coordinar el canal correcto de gestión.

La idea es poder revisar {brand} con enfoque 360: operación, menú, promociones, visibilidad y próximos pasos comerciales. Lectura rápida actual: {summary}.

El punto principal a revisar sería: {priority_action}. Motivo: {priority_reason}.

¿Me confirman por favor un contacto efectivo o una disponibilidad breve para revisarlo?

Quedo atento.

{FARMER_NAME}
Rappi"""
        whatsapp_body = f"""{greeting} te estuve llamando para revisar algunos puntos importantes de {brand} en Rappi. La idea es coordinar una gestión 360: operación, menú, promociones y visibilidad. Lectura rápida: {summary}. El punto principal sería {priority_action}. ¿Me confirman un contacto efectivo o una disponibilidad breve?"""

    elif template_type == "Churn / Chon":
        churn, churn_risk, churn_action, second_data = _churn_reading(ctx)
        email_body = f"""{greeting}

Les escribo para revisar un tema puntual de Chon/Churn de {brand}.

Tenemos dos datos específicos para validar:
1. Estado Chon: {churn}
2. {second_data}

La recomendación puntual es: {churn_action}.

La idea es revisarlo con ustedes para evitar deterioro adicional y definir una acción concreta de recuperación o seguimiento.

¿Lo podemos revisar en una llamada breve?

Quedo atento.

{FARMER_NAME}
Rappi"""
        whatsapp_body = f"""{greeting} les escribo por un tema puntual de Chon/Churn de {brand}. Datos a revisar: Estado Chon: {churn}. {second_data}. Recomendación: {churn_action}. ¿Lo revisamos en una llamada breve para definir acción?"""

    else:
        email_body = f"""{greeting}

Comparto seguimiento comercial de {brand} con lectura 360: {summary}.

{tone['cta']}

{FARMER_NAME}
Rappi"""
        whatsapp_body = f"""{greeting} Comparto seguimiento de {brand}. Lectura 360: {summary}. {tone['cta']}"""

    return {
        "subject": subject,
        "email_body": email_body.strip(),
        "whatsapp_body": whatsapp_body.strip(),
    }


def _load_templates_df():
    if not os.path.exists(TEMPLATE_QUEUE_FILE):
        return pd.DataFrame()
    try:
        df = pd.read_csv(TEMPLATE_QUEUE_FILE)
    except Exception:
        return pd.DataFrame()
    required = [
        "template_id", "datetime", "brand_id", "brand_name", "template_type", "opportunity_status",
        "email", "contact", "subject", "email_body", "whatsapp_body", "source_comment", "status"
    ]
    for col in required:
        if col not in df.columns:
            df[col] = ""
    return df


def save_template_to_queue(ctx, template_type, source_comment=""):
    if template_type in ["", "None", None]:
        return False, "No template selected."

    messages = generate_template_messages(template_type, ctx, source_comment=source_comment)
    new_row = pd.DataFrame([{
        "template_id": str(uuid.uuid4())[:8],
        "datetime": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "brand_id": ctx.get("brand_id", ""),
        "brand_name": ctx.get("brand_name", ""),
        "template_type": template_type,
        "opportunity_status": ctx.get("status", ""),
        "email": ctx.get("email", ""),
        "contact": ctx.get("contact", ""),
        "subject": messages["subject"],
        "email_body": messages["email_body"],
        "whatsapp_body": messages["whatsapp_body"],
        "source_comment": source_comment,
        "status": "Pending",
    }])

    old = _load_templates_df()
    if old.empty:
        final = new_row
    else:
        for col in new_row.columns:
            if col not in old.columns:
                old[col] = ""
        final = pd.concat([old[new_row.columns], new_row], ignore_index=True)
    final.to_csv(TEMPLATE_QUEUE_FILE, index=False, encoding="utf-8-sig")
    return True, "Template saved to Pending Templates."


def update_template_status(template_id, new_status):
    df = _load_templates_df()
    if df.empty or "template_id" not in df.columns:
        return False
    mask = df["template_id"].astype(str) == str(template_id)
    if not mask.any():
        return False
    df.loc[mask, "status"] = new_status
    df.to_csv(TEMPLATE_QUEUE_FILE, index=False, encoding="utf-8-sig")
    return True


def _load_day_queue_cursor():
    """Load saved cursor position {brand_id, brand_name, saved_at}."""
    if os.path.exists(DAY_QUEUE_CURSOR_FILE):
        try:
            with open(DAY_QUEUE_CURSOR_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_day_queue_cursor(brand_id, brand_name):
    """Persist cursor position to disk."""
    data = {
        "brand_id": str(brand_id),
        "brand_name": str(brand_name),
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        with open(DAY_QUEUE_CURSOR_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass


# ── Búsqueda en Google del restaurante (link guardado, sin API) ────────────

def _build_google_search_url(brand_name, category="", contact=""):
    """
    Arma una URL de búsqueda de Google con nombre + categoría + país del portafolio
    (PORTFOLIO_COUNTRY) y teléfono si está disponible, para encontrar el local:
    nombre, dirección, teléfono y mapa.
    """
    parts = [strip_brand_id_prefix(brand_name)]
    if category and category not in ["", "-"]:
        parts.append(category)
    if PORTFOLIO_COUNTRY:
        parts.append(PORTFOLIO_COUNTRY)
    if contact and contact not in ["", "-"]:
        parts.append(contact)
    query = " ".join(str(p) for p in parts if p)
    return "https://www.google.com/search?q=" + quote_plus(query)


@st.cache_data(ttl=60, show_spinner=False)
def _load_brand_links_df():
    if not os.path.exists(BRAND_LINKS_FILE):
        return pd.DataFrame(columns=["brand_id", "brand_name", "google_link", "saved_at"])
    try:
        df = pd.read_csv(BRAND_LINKS_FILE, dtype=str).fillna("")
        for col in ["brand_id", "brand_name", "google_link", "saved_at"]:
            if col not in df.columns:
                df[col] = ""
        return df
    except Exception:
        return pd.DataFrame(columns=["brand_id", "brand_name", "google_link", "saved_at"])


def _get_saved_brand_link(brand_id):
    df = _load_brand_links_df()
    if df.empty:
        return ""
    bid = normalize_brand_id(brand_id)
    match = df[df["brand_id"].apply(normalize_brand_id) == bid]
    if match.empty:
        return ""
    return clean(match.iloc[-1].get("google_link"), "")


def _save_brand_link(brand_id, brand_name, link):
    """Guarda (o actualiza) el link de Google encontrado para una marca. No vuelve a buscar después."""
    link = clean(link, "").strip()
    if not link:
        return False
    bid = normalize_brand_id(brand_id)
    df = _load_brand_links_df()
    df = df[df["brand_id"].apply(normalize_brand_id) != bid].copy()
    new_row = pd.DataFrame([{
        "brand_id": str(bid),
        "brand_name": str(brand_name),
        "google_link": link,
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }])
    out = pd.concat([df, new_row], ignore_index=True)
    out.to_csv(BRAND_LINKS_FILE, index=False, encoding="utf-8-sig")
    _load_brand_links_df.clear()
    return True


def _collect_priority_topics(brand_id, name, ads_current, md_current, ads_roi):
    """
    Detecta los 'temas importantes' que deben mencionarse en los mensajes de Day Queue:
    - Upselling de campaña (Ads con ROI >= 3, espacio para escalar)
    - Fotos del menú por debajo del 90%
    - Chasing/Purchasing Experience por debajo del 90%
    - Solicitud de Catálogo PDF pendiente
    - Cancelaciones
    - Reclamaciones / reclamos

    Devuelve una lista de strings cortos, en orden de prioridad, listos para
    insertarse en los templates de mensaje.
    """
    topics = []

    # Upselling: Ads activo con buen retorno -> hay espacio para escalar inversión.
    ads_active = bool(ads_current.get("active", False))
    if ads_active and to_number(ads_roi, 0) >= 3:
        topics.append("visibilidad pagada con buen retorno — hay espacio para aumentar la inversión")

    # Señales de Smart Priorities (Fotos, Chasing Experience, PDF, Cancelaciones, Reclamaciones)
    sp_signals = get_priority_signals_for_brand(brand_id, name)
    seen_kinds = set()
    if sp_signals.get("found"):
        for lv in sp_signals.get("levers", []):
            kind = _classify_priority_lever(lv.get("metric", ""))
            if kind in seen_kinds:
                continue
            if kind == "ops_claims":
                topics.append("reclamos de clientes que están abiertos")
                seen_kinds.add(kind)
            elif kind == "ops_cancellations":
                topics.append("pedidos cancelados por encima del promedio")
                seen_kinds.add(kind)
            elif kind == "menu_photos":
                topics.append("fotos del menú por debajo del 90%")
                seen_kinds.add(kind)
            elif kind == "menu_purchase_experience":
                topics.append("experiencia de compra por debajo del umbral esperado")
                seen_kinds.add(kind)
            elif kind == "menu_pdf":
                topics.append("catálogo en PDF pendiente de enviar")
                seen_kinds.add(kind)
            elif kind == "ops_availability":
                topics.append("tienda con horario o disponibilidad reducida")
                seen_kinds.add(kind)

    # Chequeo directo de métricas de menú por si Smart Priorities no las trae.
    try:
        _menu = get_menu_health_for_brand(brand_id, name) if "get_menu_health_for_brand" in dir() else {}
    except Exception:
        _menu = {}
    if isinstance(_menu, dict):
        _photos_val = to_number(_menu.get("photos"), 0)
        _purch_val = to_number(_menu.get("purchasing_experience"), 0)
        if _photos_val and _photos_val < 0.90 and "menu_photos" not in seen_kinds:
            topics.append("fotos del menú por debajo del 90%")
            seen_kinds.add("menu_photos")
        if _purch_val and _purch_val < 0.90 and "menu_purchase_experience" not in seen_kinds:
            topics.append("experiencia de compra por debajo del umbral esperado")
            seen_kinds.add("menu_purchase_experience")

    return topics


def _format_priority_topics_line(topics):
    """Convierte la lista de temas en una frase lista para insertar en el mensaje."""
    if not topics:
        return ""
    if len(topics) == 1:
        items = topics[0]
    elif len(topics) == 2:
        items = f"{topics[0]} y {topics[1]}"
    else:
        items = ", ".join(topics[:-1]) + f" y {topics[-1]}"
    return f"Además, debemos revisar todos los puntos que tenemos en este momento: {items}."


def _churn_risk_reading(churn_status):
    """
    Lectura corta de riesgo/acción a partir del label de churn con emoji
    (✅ On · ⚠️ W1 · 🚨 W2 · 🆘 W3 · 😴 Off).
    Devuelve (risk_text, action_text, urgent_bool).
    'Off' se trata como caso urgente: la marca está apagada y hay que reconectar ya.
    """
    churn_text = norm_text(clean(churn_status, ""))
    if "off" in churn_text or "😴" in str(churn_status):
        return "marca apagada / desconectada", "reconectar la marca de forma urgente y entender qué pasó", True
    if "w3" in churn_text or "🆘" in str(churn_status) or "☠" in str(churn_status):
        return "riesgo alto / deterioro avanzado", "definir acción de recuperación prioritaria", True
    if "w2" in churn_text or "🚨" in str(churn_status):
        return "riesgo medio con alerta activa", "revisar la causa del deterioro y acordar una corrección", False
    if "w1" in churn_text or "⚠" in str(churn_status):
        return "alerta temprana", "prevenir que la marca siga deteriorándose", False
    return "marca activa", "mantener el seguimiento para evitar una caída", False


def _build_churn_day_queue_message(name, category, churn_status, priority_topics=None):
    """
    Mensaje pre-llamada específico para marcas con deterioro de ventas.
    Vocabulario coloquial: sin "churn", "W1/W2/W3", ni tecnicismos de plataforma.
    Returns (subject, whatsapp_body, email_body) — sin API, fully offline.
    """
    risk, action, urgent = _churn_risk_reading(churn_status)

    if urgent:
        subject = f"🚨 URGENTE — {name} — recuperemos las ventas"
        pain_wa = (f"⚠️ Urgente: {name} tiene una caída importante en ventas dentro de tu categoría {category}. "
                   f"La tienda está desconectada o con actividad muy baja en este momento. "
                   f"Quiero llamarte hoy para entender qué pasó y ver cómo lo resolvemos juntos cuanto antes.")
        pain_email = (f"Te escribo con carácter urgente: {name} está mostrando una caída importante en ventas "
                       f"dentro de {category}. En este momento la tienda no está generando pedidos con normalidad.\n\n"
                       f"Necesito entender qué está pasando de tu lado para ayudarte a recuperar esa actividad. "
                       f"¿Podemos hablar hoy? Es prioritario antes de que se pierda más venta.")
    else:
        subject = f"Revisemos las ventas de {name} — {category}"
        pain_wa = (f"Vi que {name} tuvo una baja en ventas últimamente en {category}. "
                   f"Te llamo hoy para revisarlo juntos y ver cómo lo corregimos.")
        pain_email = (f"Revisando el desempeño de {category}, noté que {name} tuvo una baja en ventas reciente. "
                       f"Te llamo hoy para revisarlo juntos y definir qué palanca activamos para recuperarlo.")

    topics_line = _format_priority_topics_line(priority_topics or [])

    greeting_wa = f"Hola, soy Sabas de Rappi 👋 {pain_wa}"
    if topics_line:
        greeting_wa += f" {topics_line}"

    greeting_email = (f"Hola,\n\nSoy {FARMER_NAME}, tu farmer de Rappi.\n\n"
                      f"{pain_email}\n\n"
                      + (f"{topics_line}\n\n" if topics_line else "")
                      + (f"¿Tenés un momento hoy para una llamada urgente?\n\nSaludos,\n{FARMER_NAME}\nRappi"
                         if urgent else
                         f"¿Tenés un momento hoy para una llamada breve?\n\nSaludos,\n{FARMER_NAME}\nRappi"))

    return subject, greeting_wa, greeting_email


def _build_day_queue_message(name, category, lever, ads_current, md_current, cr, gmv_ars, aov_ars, campaign_design, priority_topics=None):
    """
    Rule-based pre-call message generator.
    Returns (subject, whatsapp_body, email_body) — no API, fully offline.
    Vocabulario coloquial:
      - GMV          → venta total / facturación
      - AOV          → ticket promedio
      - Markdown/MD  → promoción / promo / descuento activo
      - Ads          → visibilidad pagada / publicidad / tienda pagada
      - CR           → de cada 100 visitas cuántos compran / conversión
      - ROI/Retorno  → por cada peso invertido ganás X / retorno
      - Bookings     → inversión en pauta
    """
    lever = lever or "Ads"
    ads_active = bool(ads_current.get("active", False))
    md_active  = bool(md_current.get("active", False))
    ads_roi    = to_number(ads_current.get("roi"), 0)
    discount   = campaign_design.get("discount", 20)
    hero       = campaign_design.get("hero_product", "tu producto principal")
    impact_low = campaign_design.get("impact_low", 5)
    impact_high = campaign_design.get("impact_high", 12)
    cr_val     = _normalize_rate_value(cr) or 0

    if lever == "MD" and not md_active:
        pain_wa    = (f"Vi que {name} aún no tiene una promoción activa en Rappi. En {category}, "
                      f"las marcas con descuento activo están creciendo hasta un {impact_high}% más en ventas. "
                      f"Te llamo hoy para mostrarte cómo activarla con {discount}% en {hero} — sin complicaciones.")
        pain_email = (f"Analizando el portafolio de {category}, noté que {name} todavía no tiene una promoción activa. "
                      f"Eso significa tráfico orgánico que hoy va a la competencia.\n\n"
                      f"Te propongo activar una campaña con {discount}% de descuento en {hero}, "
                      f"con impacto proyectado de +{impact_low}%–+{impact_high}% en ventas totales. "
                      f"Te llamo hoy para revisarlo juntos.")
        subject = f"Activemos una promoción para {name} — {discount}% en {hero}"

    elif lever == "Ads" and not ads_active:
        pain_wa    = (f"Vi que {name} no tiene visibilidad pagada activa en Rappi. En {category}, "
                      f"el tráfico va a quienes tienen pauta activa — hoy ese tráfico va a tu competencia. "
                      f"Te llamo hoy para mostrarte el presupuesto de entrada y el impacto esperado en tu venta.")
        pain_email = (f"Revisando el desempeño de marcas en {category}, noté que {name} no tiene visibilidad pagada activa. "
                      f"Eso representa visitas que hoy está capturando la competencia.\n\n"
                      f"En tu categoría, la inversión en pauta tiene un retorno promedio muy sano. "
                      f"Te llamo hoy para mostrarte los números concretos y un presupuesto de entrada.")
        subject = f"Sumemos visibilidad pagada para {name} — arrancamos esta semana"

    elif lever == "Ads" and ads_active and ads_roi >= 3:
        pain_wa    = (f"Vi que la tienda pagada de {name} está generando un retorno de {ads_roi:.1f}x sobre lo que invertís — eso está muy bien. "
                      f"Lo que significa es que hay espacio real para escalar la inversión y multiplicar ese resultado. "
                      f"Te llamo hoy para mostrarte los números.")
        pain_email = (f"Revisando el rendimiento de {name}, la inversión en pauta está generando un retorno de {ads_roi:.1f}x. "
                      f"Por cada peso que invertís, estás recuperando {ads_roi:.1f}. Eso es una señal clara de que el canal funciona.\n\n"
                      f"El paso lógico es escalar la inversión, no mantenerla. "
                      f"Te llamo hoy para mostrarte la proyección concreta.")
        subject = f"Escalemos la inversión de {name} — retorno actual {ads_roi:.1f}x"

    elif cr_val and cr_val < 0.12:
        pain_wa    = (f"Revisando los datos de {name}, vi que de cada 100 personas que ven tu tienda, "
                      f"menos de las esperadas terminan comprando — está por debajo del promedio de {category}. "
                      f"Una promoción puede ser la palanca más directa para mover eso. "
                      f"Te llamo hoy con una propuesta concreta.")
        pain_email = (f"Analizando el rendimiento de {name} en {category}, vi que la conversión está por debajo del promedio. "
                      f"Dicho de otra forma: el tráfico llega, pero no termina comprando tanto como debería.\n\n"
                      f"Una promoción con {discount}% de descuento en {hero} es la palanca más directa para mover eso. "
                      f"Te llamo hoy para revisarlo.")
        subject = f"Mejoremos la conversión de {name} — {category}"

    else:
        pain_wa    = (f"Te llamo hoy para revisar una oportunidad concreta de crecimiento para {name} en Rappi. "
                      f"Tengo los datos de {category} y una propuesta lista.")
        pain_email = (f"Tengo una propuesta de crecimiento para {name} basada en el análisis actual de {category} en Rappi. "
                      f"Te llamo hoy para revisarla juntos.")
        subject = f"Oportunidad de crecimiento para {name} — Rappi"

    topics_line = _format_priority_topics_line(priority_topics or [])

    greeting_wa    = f"Hola, soy Sabas de Rappi 👋 {pain_wa}"
    if topics_line:
        greeting_wa += f" {topics_line}"

    greeting_email = (f"Hola,\n\nSoy {FARMER_NAME}, tu farmer de Rappi.\n\n"
                      f"{pain_email}\n\n"
                      + (f"{topics_line}\n\n" if topics_line else "")
                      + f"¿Tenés un momento hoy para una llamada breve?\n\nSaludos,\n{FARMER_NAME}\nRappi")

    return subject, greeting_wa, greeting_email


def page_day_queue():
    render_header("Day Queue", "Cola diaria de pre-llamada · plantillas listas por marca")

    # ── Load scored portfolio from Priority Data sheet ────────────────────────
    priority_df = load_priority_data()
    use_priority = not priority_df.empty

    if use_priority:
        # Build a unique brand list from Priority Data (total rows only)
        total_rows = priority_df[priority_df["_metric_norm"] == "total"].copy()
        if total_rows.empty:
            total_rows = priority_df.drop_duplicates(subset=["_id"]).copy()
        total_rows = total_rows.sort_values("_score", ascending=False).reset_index(drop=True)
        queue_source = total_rows
    else:
        # Fallback to Growth OS scored data
        data = _prepare_growth_scored_data()
        if data.empty:
            st.error("No se encontró data del portafolio (Priority Data ni Growth OS).")
            return
        id_col = get_id_column_name(data)
        if not id_col:
            st.error("Columna ID no encontrada.")
            return
        data = data.sort_values("_opportunity_score", ascending=False).reset_index(drop=True)
        queue_source = data

    # ── Load last contact maps ────────────────────────────────────────────────
    prod_map = get_productivity_last_contact_map(EXCEL_FILE)
    meta_map = get_last_comment_meta_map(limit=1)

    # ── Cursor logic ──────────────────────────────────────────────────────────
    cursor          = _load_day_queue_cursor()
    cursor_brand_id = cursor.get("brand_id", "")
    cursor_saved_at = cursor.get("saved_at", "")

    start_idx = 0
    if cursor_brand_id:
        cid = normalize_brand_id(cursor_brand_id)
        matches = queue_source[queue_source["_id"].apply(normalize_brand_id) == cid].index.tolist()
        if matches:
            start_idx = matches[0] + 1

    # ── Header: cursor info + controls ───────────────────────────────────────
    col_info, col_reset = st.columns([3, 1])
    with col_info:
        if cursor_brand_id and cursor_saved_at:
            cursor_name = cursor.get("brand_name", cursor_brand_id)
            st.caption(f"📍 Posición guardada: **{cursor_name}** · marcada el {cursor_saved_at} · mostrando desde la #{start_idx + 1}")
        else:
            st.caption("📍 Sin posición guardada — mostrando desde el inicio del portafolio")
    with col_reset:
        if st.button("↺ Reiniciar posición"):
            if os.path.exists(DAY_QUEUE_CURSOR_FILE):
                os.remove(DAY_QUEUE_CURSOR_FILE)
            st.rerun()

    # ── Slice: next 40 brands from cursor ────────────────────────────────────
    queue_slice = queue_source.iloc[start_idx:start_idx + 40].copy()

    if queue_slice.empty:
        st.info("Llegaste al final del portafolio. Reiniciá la posición para volver al inicio.")
        return

    st.markdown(f"### {len(queue_slice)} marcas · hoy")

    # ── Helper: sticker badge ─────────────────────────────────────────────────
    def _sticker(label, value, color_val="#1A1A2E", bg="rgba(255,255,255,0.92)", border="rgba(0,0,0,0.07)"):
        return (
            f"<div style='background:{bg};border:1px solid {border};border-radius:10px;"
            f"padding:8px 14px;display:inline-block;margin-right:8px;margin-bottom:6px;'>"
            f"<div style='font-size:10px;font-weight:700;letter-spacing:.06em;"
            f"color:#6B7280;text-transform:uppercase;margin-bottom:2px;'>{label}</div>"
            f"<div style='font-size:15px;font-weight:700;color:{color_val};'>{value}</div>"
            f"</div>"
        )

    # ── Accordion state: only one card open at a time ────────────────────────
    _dq_open_key = "dq_open_card_idx"
    if _dq_open_key not in st.session_state:
        st.session_state[_dq_open_key] = None  # None = all closed by default

    # ── Render each brand card ────────────────────────────────────────────────
    for idx, (_, row) in enumerate(queue_slice.iterrows()):
        brand_id = normalize_brand_id(row.get("_id", ""))
        if not brand_id:
            continue

        # Name & category — try Priority Data cols first, fall back to Growth OS
        name = clean(row.get("_brand_col") or row.get("_name"), f"Marca {brand_id}")
        name = strip_brand_id_prefix(name)
        opp_score = round(float(to_number(row.get("_score") or row.get("_opportunity_score"), 0)), 1)

        # Pull full brand metrics from Growth OS
        growth_df = load_growth_data()
        id_col_g  = get_id_column_name(growth_df) if not growth_df.empty else None
        brand_row = None
        if id_col_g and not growth_df.empty:
            match = growth_df[growth_df[id_col_g].apply(normalize_brand_id) == brand_id]
            if not match.empty:
                brand_row = match.iloc[0]

        category_raw = clean(get_from_row(brand_row, ["category", "categoria", "cat"], ""), "") if brand_row is not None else ""
        category     = category_raw.split("·")[0].strip() if "·" in category_raw else category_raw.strip()
        gmv_ars      = to_number(get_from_row(brand_row, ["last gmv ars", "gmv ars"], 0), 0) if brand_row is not None else 0
        aov_ars      = to_number(get_from_row(brand_row, ["last aov ars", "aov ars"], 0), 0) if brand_row is not None else 0
        cr_raw       = to_number(get_from_row(brand_row, ["cr %", "conversion rate", "cr"], 0), 0) if brand_row is not None else 0

        # Last contact badge
        last_dt = get_last_contact_dt(brand_id, name, prod_map, meta_map)
        if pd.isna(pd.to_datetime(last_dt, errors="coerce")):
            days_ago      = None
            contact_badge = "🔴 Sin contacto"
            badge_bg      = "rgba(229,51,42,0.12)"
            badge_txt     = "#FF4D2E"
        else:
            days_ago = (datetime.now() - pd.to_datetime(last_dt)).days
            contact_badge = (
                "🟢 Hoy" if days_ago == 0
                else f"🟢 Hace {days_ago}d" if days_ago <= 7
                else f"🟡 Hace {days_ago}d" if days_ago <= 14
                else f"🟠 Hace {days_ago}d" if days_ago <= 21
                else f"🔴 Hace {days_ago}d"
            )
            badge_bg  = ("rgba(111,242,75,0.08)" if (days_ago or 99) <= 7
                         else "rgba(255,113,36,0.10)" if (days_ago or 99) <= 21
                         else "rgba(229,51,42,0.10)")
            badge_txt = ("#7ED321" if (days_ago or 99) <= 7
                         else "#D95A10" if (days_ago or 99) <= 21
                         else "#FF4D2E")

        # Lever status
        ads_current_raw   = get_current_ads_metrics(brand_id)
        md_current_raw    = get_current_md_metrics(brand_id, pro=False)
        md_pro_current_raw = get_current_md_metrics(brand_id, pro=True)
        ads_current, md_current, md_pro_current = _merge_growth_manual_status(
            brand_row if brand_row is not None else row,
            ads_current_raw, md_current_raw, md_pro_current_raw
        )

        ads_active = bool(ads_current.get("active", False))
        md_active  = bool(md_current.get("active", False))
        ads_roi    = to_number(ads_current.get("roi"), 0)
        ads_budget = to_number(ads_current.get("budget") or ads_current.get("bookings"), 0)
        md_campaign_name = clean(md_current.get("campaign_name") or md_current.get("name") or md_current.get("promo_name"), "")

        lever = "Ads" if ads_active else "MD"

        commission_raw = _normalize_rate_value(get_from_row(brand_row if brand_row is not None else row,
                                                             ["comm. rate", "commission rate", "commission"], 0))
        booster = recommend_booster_for_brand(category, gmv_ars, aov_ars, cr_raw, 0, ads_current, md_current)
        campaign_design = design_campaign_for_brand(
            name, category, gmv_ars, aov_ars, cr_raw, 0,
            commission_raw, ads_current, md_current, md_pro_current,
            booster, {}, brand_id=brand_id
        )

        # ── Temas importantes para el mensaje (revisión 360) ───────────────
        priority_topics = _collect_priority_topics(
            brand_id, name, ads_current, md_current, ads_roi
        )

        contact_number = fmt_contact_number(
            get_from_row(brand_row if brand_row is not None else row, ["contact number", "phone", "contact"], "")
        ) if (brand_row is not None or row is not None) else ""

        # ── Estado Chon/Churn y presencia en Priority Data ──────────────────
        churn_status = get_churn_status(brand_id)  # "✅ On" si no figura en Current Churn
        churn_text_norm = norm_text(clean(churn_status, ""))
        in_churn_risk = any(k in churn_text_norm for k in ["w1", "w2", "w3", "off"]) or \
                        any(c in str(churn_status) for c in ["⚠", "🚨", "🆘", "😴", "☠"])
        sp_signals_for_churn = get_priority_signals_for_brand(brand_id, name)
        in_priority_data = bool(sp_signals_for_churn.get("found"))
        show_churn_sticker = in_priority_data or in_churn_risk

        if in_churn_risk:
            subject, wa_body, email_body = _build_churn_day_queue_message(
                strip_brand_id_prefix(name), category, churn_status, priority_topics
            )
        else:
            subject, wa_body, email_body = _build_day_queue_message(
                strip_brand_id_prefix(name), category, lever, ads_current, md_current,
                cr_raw, gmv_ars, aov_ars, campaign_design, priority_topics
            )

        # ── Card — accordion: clicking a card closes all others ─────────────
        card_key = f"dq_{brand_id}_{idx}"
        _is_open = (st.session_state[_dq_open_key] == idx)
        _expander_label = f"#{start_idx + idx + 1} · {name} · {category} · Score {opp_score}"
        # Toggle button that acts as the accordion header
        _toggle_col, _ = st.columns([1, 0.001])
        with _toggle_col:
            if st.button(
                ("▼ " if _is_open else "▶ ") + _expander_label,
                key=f"toggle_{card_key}",
                use_container_width=True,
            ):
                st.session_state[_dq_open_key] = None if _is_open else idx
                st.rerun()
        if not _is_open:
            continue
        with st.container():
            # Top row: badges + mark position button
            top_l, top_r = st.columns([3, 1])
            with top_l:
                ads_badge_bg  = "rgba(59,72,131,0.20)" if ads_active else "rgba(255,255,255,0.90)"
                ads_badge_txt = "#1B3F8B" if ads_active else "#6B7280"
                md_badge_bg   = "rgba(111,242,75,0.10)" if md_active else "rgba(255,255,255,0.90)"
                md_badge_txt  = "#7ED321" if md_active else "#6B7280"
                churn_badge_html = ""
                if show_churn_sticker:
                    _is_urgent_churn = any(c in str(churn_status) for c in ["🆘", "🚨", "☠", "😴"])
                    _churn_bg = ("rgba(229,51,42,0.15)" if _is_urgent_churn
                                  else "rgba(255,113,36,0.10)" if "⚠" in str(churn_status)
                                  else "rgba(255,255,255,0.92)")
                    _churn_txt = ("#FF4D2E" if _is_urgent_churn
                                   else "#FF7124" if "⚠" in str(churn_status)
                                   else "#6B7280")
                    churn_badge_html = (
                        f"<span style='background:{_churn_bg};color:{_churn_txt};font-size:12px;font-weight:600;"
                        f"padding:3px 10px;border-radius:20px;margin-right:6px;'>"
                        f"📊 Churn — {html.escape(clean(churn_status, '✅ On'))}</span>"
                    )
                    if _is_urgent_churn:
                        churn_badge_html += (
                            f"<span style='background:rgba(229,51,42,0.20);color:#FF4D2E;font-size:12px;font-weight:700;"
                            f"padding:3px 10px;border-radius:20px;margin-right:6px;'>"
                            f"🚨 Urgente — reconectar</span>"
                        )
                st.markdown(
                    f"<span style='background:{badge_bg};color:{badge_txt};font-size:12px;"
                    f"font-weight:600;padding:3px 10px;border-radius:20px;margin-right:6px;'>"
                    f"{contact_badge}</span>"
                    f"<span style='background:{ads_badge_bg};color:{ads_badge_txt};font-size:12px;font-weight:600;"
                    f"padding:3px 10px;border-radius:20px;margin-right:6px;'>"
                    f"{'✅ Ads' if ads_active else '⬜ Sin Ads'}</span>"
                    f"<span style='background:{md_badge_bg};color:{md_badge_txt};font-size:12px;font-weight:600;"
                    f"padding:3px 10px;border-radius:20px;margin-right:6px;'>"
                    f"{'✅ Markdown' if md_active else '⬜ Sin Markdown'}</span>"
                    f"{churn_badge_html}",
                    unsafe_allow_html=True
                )
            with top_r:
                if st.button("📍 Marcar posición", key=f"cursor_{card_key}"):
                    _save_day_queue_cursor(brand_id, name)
                    st.success(f"Posición guardada en {name}")
                    st.rerun()

            # ── Ir al Brand Finder de esta marca ──────────────────────────────
            if st.button(
                f"🔍 Ver perfil completo de {html.escape(strip_brand_id_prefix(name))} en Brand Finder",
                key=f"goto_bf_{card_key}",
                use_container_width=True,
            ):
                st.session_state["_bf_goto_brand_id"] = brand_id
                st.session_state["active_page"] = "Brand Finder"
                st.rerun()


            # ── Stickers de métricas ───────────────────────────────────────────
            stickers_html = ""
            # Venta total (GMV)
            if gmv_ars and gmv_ars > 0:
                gmv_fmt = f"ARS {fmt_money(gmv_ars)}" if gmv_ars < 1_000_000 else f"ARS {gmv_ars/1_000_000:.1f}M"
                stickers_html += _sticker("Venta total", gmv_fmt, "#1A1A2E")
            # Ticket promedio (AOV)
            if aov_ars and aov_ars > 0:
                stickers_html += _sticker("Ticket promedio", f"ARS {fmt_money(aov_ars)}", "#1A1A2E")
            # Tasa de conversión
            if cr_raw and to_number(cr_raw, 0) > 0:
                stickers_html += _sticker("Conversión", fmt_percent0(cr_raw), "#6B7280")
            # Budget de publicidad (si tiene)
            if ads_active and ads_budget > 0:
                stickers_html += _sticker("Budget publicidad", f"ARS {fmt_money(ads_budget)}", "#1B3F8B",
                                           "rgba(59,72,131,0.15)", "rgba(27,63,139,0.12)")
            # Retorno publicidad (si tiene)
            if ads_active and ads_roi and ads_roi > 0:
                roi_color = "#7ED321" if ads_roi >= 3 else ("#FF7124" if ads_roi >= 1.5 else "#FF4D2E")
                stickers_html += _sticker("Retorno ads", f"{ads_roi:.1f}x", roi_color,
                                           "rgba(111,242,75,0.08)", "rgba(111,242,75,0.20)")
            # Nombre campaña promo (si tiene)
            if md_active and md_campaign_name and md_campaign_name not in ["-", ""]:
                short_camp = md_campaign_name[:28] + "…" if len(md_campaign_name) > 28 else md_campaign_name
                stickers_html += _sticker("Campaña promo", short_camp, "#7ED321",
                                           "rgba(111,242,75,0.08)", "rgba(111,242,75,0.20)")

            if stickers_html:
                st.markdown(
                    f"<div style='display:flex;flex-wrap:wrap;gap:0;margin:12px 0 8px 0;'>{stickers_html}</div>",
                    unsafe_allow_html=True
                )

            # ── Alertas Smart Priorities ───────────────────────────────────────
            sp_signals = get_priority_signals_for_brand(brand_id, name)
            sp_alerts_html = ""
            if sp_signals.get("found"):
                for lv in sp_signals.get("levers", []):
                    kind = _classify_priority_lever(lv.get("metric", ""))
                    metric_label = clean(lv.get("metric"), "")
                    if kind == "ops_claims":
                        sp_alerts_html += (
                            f"<span style='background:rgba(229,51,42,0.15);color:#FF4D2E;font-size:12px;"
                            f"font-weight:600;padding:3px 10px;border-radius:20px;margin-right:6px;margin-bottom:4px;display:inline-block;'>"
                            f"⚠️ Reclamaciones — {html.escape(metric_label)}</span>"
                        )
                    elif kind == "ops_cancellations":
                        sp_alerts_html += (
                            f"<span style='background:rgba(229,51,42,0.15);color:#FF4D2E;font-size:12px;"
                            f"font-weight:600;padding:3px 10px;border-radius:20px;margin-right:6px;margin-bottom:4px;display:inline-block;'>"
                            f"⚠️ Cancelaciones — {html.escape(metric_label)}</span>"
                        )
                    elif kind == "menu_photos":
                        sp_alerts_html += (
                            f"<span style='background:rgba(255,113,36,0.10);color:#FF7124;font-size:12px;"
                            f"font-weight:600;padding:3px 10px;border-radius:20px;margin-right:6px;margin-bottom:4px;display:inline-block;'>"
                            f"📸 Fotos — {html.escape(metric_label)}</span>"
                        )
                    elif kind == "ops_availability":
                        sp_alerts_html += (
                            f"<span style='background:rgba(255,113,36,0.10);color:#FF7124;font-size:12px;"
                            f"font-weight:600;padding:3px 10px;border-radius:20px;margin-right:6px;margin-bottom:4px;display:inline-block;'>"
                            f"🔌 Availability — {html.escape(metric_label)}</span>"
                        )
                    elif kind == "menu_purchase_experience":
                        sp_alerts_html += (
                            f"<span style='background:rgba(255,113,36,0.10);color:#FF7124;font-size:12px;"
                            f"font-weight:600;padding:3px 10px;border-radius:20px;margin-right:6px;margin-bottom:4px;display:inline-block;'>"
                            f"🛒 Chasing Experience — {html.escape(metric_label)}</span>"
                        )
            # Also check menu metrics for photos/chasing_experience thresholds directly
            _menu_for_alert = get_menu_health_for_brand(brand_id, name) if "get_menu_health_for_brand" in dir() else {}
            if isinstance(_menu_for_alert, dict):
                _photos_val = to_number(_menu_for_alert.get("photos"), 0)
                _purch_val  = to_number(_menu_for_alert.get("purchasing_experience"), 0)
                if _photos_val and _photos_val < 0.90 and "📸 Fotos" not in sp_alerts_html:
                    sp_alerts_html += (
                        f"<span style='background:rgba(255,113,36,0.10);color:#FF7124;font-size:12px;"
                        f"font-weight:600;padding:3px 10px;border-radius:20px;margin-right:6px;margin-bottom:4px;display:inline-block;'>"
                        f"📸 Fotos {fmt_percent0(_photos_val)} — por debajo del 90%</span>"
                    )
                if _purch_val and _purch_val < 0.90 and "🛒 Chasing Experience" not in sp_alerts_html:
                    sp_alerts_html += (
                        f"<span style='background:rgba(255,113,36,0.10);color:#FF7124;font-size:12px;"
                        f"font-weight:600;padding:3px 10px;border-radius:20px;margin-right:6px;margin-bottom:4px;display:inline-block;'>"
                        f"🛒 Chasing Experience {fmt_percent0(_purch_val)} — por debajo del 90%</span>"
                    )
            if sp_alerts_html:
                st.markdown(
                    f"<div style='display:flex;flex-wrap:wrap;gap:0;margin:8px 0 8px 0;'>{sp_alerts_html}</div>",
                    unsafe_allow_html=True
                )

            # ── Palanca recomendada ────────────────────────────────────────────
            if lever == "Ads" and ads_active:
                _lever_label = "📢 Ads — escalar retorno existente"
            elif lever == "Ads":
                _lever_label = "📢 Ads — activar primera campaña"
            else:
                _disc = campaign_design.get("discount", 20)
                _hero = campaign_design.get("hero_product", "producto principal")
                _lever_label = f"💸 Promo — activar {_disc}% en {_hero}"

            st.markdown(
                f"<div style='font-size:12px;font-weight:700;color:#6B7280;text-transform:uppercase;"
                f"margin:12px 0 4px;'>Palanca recomendada</div>"
                f"<div style='font-size:14px;font-weight:600;color:#1A1A2E;margin-bottom:12px;'>"
                f"{_lever_label}</div>",
                unsafe_allow_html=True
            )

            # ── Templates ─────────────────────────────────────────────────────
            _msg_variant = "churn" if in_churn_risk else "std"
            wa_key    = f"wa_{card_key}_{_msg_variant}"
            subj_key  = f"subj_{card_key}_{_msg_variant}"
            email_key = f"email_{card_key}_{_msg_variant}"

            t1, t2 = st.tabs(["WhatsApp / Treble", "Email"])
            with t1:
                st.text_area(
                    "Mensaje WhatsApp",
                    value=wa_body,
                    height=130,
                    key=wa_key,
                    label_visibility="collapsed"
                )
            with t2:
                st.text_input("Asunto", value=subject, key=subj_key)
                st.text_area(
                    "Cuerpo email",
                    value=email_body,
                    height=200,
                    key=email_key,
                    label_visibility="collapsed"
                )


# =========================
# THEME
# =========================

# ── Sidebar collapsible nav via session_state ─────────────────────────────
if "nav_collapsed" not in st.session_state:
    st.session_state["nav_collapsed"] = False
if "active_page" not in st.session_state:
    st.session_state["active_page"] = "Management Dashboard"

NAV_GROUPS = [
    ("Principal", [
        ("Management Dashboard", "📊"),
        ("Opportunity List",     "🎯"),
        ("Follow-Up List",       "🔁"),
        ("Brand Finder",         "🔍"),
        ("Day Queue",            "📋"),
        ("Pareto Hub",           "🧭"),
    ]),
    ("Tracking", [
        ("Acquisition Tracker",      "🚀"),
        ("Campaign Weekly Tracker",  "📣"),
        ("Weekly Calendar",          "📅"),
        ("Brand Update",             "✏️"),
    ]),
    ("Análisis", [
        ("Earnings Calculator",   "💰"),
        ("Productivity HeatMap",  "🌡️"),
        ("Call Quality Trainer",  "🎙️"),
        ("Role Play Trainer",     "🥊"),
    ]),
]

collapsed = st.session_state["nav_collapsed"]

# ── CSS for collapsible sidebar ───────────────────────────────────────────
st.markdown("""
<style>
/* Hide native Streamlit sidebar toggle button on small screens to avoid conflict */
[data-testid="collapsedControl"] { display: none !important; }

/* Force sidebar width based on collapse state */
section[data-testid="stSidebar"] > div:first-child {
    transition: width 0.25s cubic-bezier(.4,0,.2,1) !important;
}

.nav-toggle-btn {
    background: rgba(255,255,255,0.95);
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 10px;
    color: #1A1A2E;
    cursor: pointer;
    font-size: 16px;
    padding: 7px 10px;
    transition: background 0.18s, border-color 0.18s;
    width: 100%;
    text-align: center;
    margin-bottom: 4px;
}
.nav-toggle-btn:hover {
    background: rgba(0,0,0,0.08);
    border-color: rgba(27,63,139,0.20);
}

.nav-section-label {
    font-size: 9px;
    font-weight: 800;
    letter-spacing: .12em;
    text-transform: uppercase;
    color: rgba(107,114,128,0.45);
    padding: 10px 4px 4px 4px;
}

.nav-item {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 8px 10px;
    border-radius: 10px;
    font-size: 13px;
    font-weight: 600;
    color: #4B5563;
    cursor: pointer;
    transition: background 0.15s, color 0.15s;
    white-space: nowrap;
    overflow: hidden;
    border: 1px solid transparent;
    margin-bottom: 2px;
    width: 100%;
    text-align: left;
    background: transparent;
}
.nav-item:hover {
    background: rgba(255,255,255,0.95);
    color: #1A1A2E;
}
.nav-item.active {
    background: rgba(27,63,139,0.12);
    color: #1A1A2E;
    border-color: rgba(126,211,33,0.22);
}
.nav-item .nav-icon {
    font-size: 16px;
    flex-shrink: 0;
    width: 22px;
    text-align: center;
}
.nav-item .nav-label {
    overflow: hidden;
    text-overflow: ellipsis;
}

.nav-logo-full {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 4px 2px 12px 2px;
    border-bottom: 1px solid rgba(255,255,255,0.95);
    margin-bottom: 8px;
}
.nav-logo-icon {
    font-size: 24px;
    flex-shrink: 0;
}
.nav-logo-text {
    font-size: 15px;
    font-weight: 800;
    color: #1A1A2E;
    letter-spacing: -.01em;
    line-height: 1.2;
}
.nav-logo-sub {
    font-size: 10px;
    color: rgba(107,114,128,0.60);
    font-weight: 600;
}
.nav-divider {
    height: 1px;
    background: rgba(255,255,255,0.95);
    margin: 6px 0;
}
</style>
""", unsafe_allow_html=True)

# ── Sidebar content ───────────────────────────────────────────────────────
with st.sidebar:
    # Toggle button
    toggle_label = "◀" if not collapsed else "▶"
    if st.button(toggle_label, key="nav_toggle", help="Colapsar / expandir navegación",
                 use_container_width=True):
        st.session_state["nav_collapsed"] = not st.session_state["nav_collapsed"]
        st.rerun()

    collapsed = st.session_state["nav_collapsed"]

    if not collapsed:
        # Logo expanded
        st.markdown(f"""
        <div class="nav-logo-full">
            <div class="nav-logo-icon">📈</div>
            <div>
                <div class="nav-logo-text">Growth OS</div>
                <div class="nav-logo-sub">Commercial Excellence</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown('<div style="text-align:center;font-size:22px;padding-bottom:10px;border-bottom:1px solid rgba(255,255,255,0.95);margin-bottom:8px;">📈</div>', unsafe_allow_html=True)

    # Nav groups
    current_page = st.session_state["active_page"]
    for group_label, items in NAV_GROUPS:
        if not collapsed:
            st.markdown(f'<div class="nav-section-label">{group_label}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="nav-divider"></div>', unsafe_allow_html=True)

        for page_name, icon in items:
            is_active = (current_page == page_name)
            active_class = "active" if is_active else ""
            label_html = f'<span class="nav-label">{page_name}</span>' if not collapsed else ""
            btn_key = f"nav_{page_name.replace(' ', '_').lower()}"
            if st.button(
                f"{icon}  {page_name}" if not collapsed else icon,
                key=btn_key,
                use_container_width=True,
            ):
                st.session_state["active_page"] = page_name
                st.rerun()

    if not collapsed:
        st.markdown('<div class="nav-divider" style="margin-top:12px;"></div>', unsafe_allow_html=True)

        # ── Avisos de datos: visibles, no invasivos ──────────────────────
        _issues = st.session_state.get("_data_issues", {})
        if _issues:
            with st.expander(f"⚠️ {len(_issues)} aviso{'s' if len(_issues) > 1 else ''} de datos"):
                for _ctx, _info in _issues.items():
                    st.markdown(f"**{_ctx}** · {_info.get('time', '')}")
                    if _info.get("hint"):
                        st.caption(_info["hint"])

        if "dark_mode" not in st.session_state:
            st.session_state["dark_mode"] = False
        _dm_label = "🌙 Dark mode" if not st.session_state["dark_mode"] else "☀️ Light mode"
        if st.button(_dm_label, key="nav_dark_mode_toggle", use_container_width=True):
            st.session_state["dark_mode"] = not st.session_state["dark_mode"]
            st.rerun()
        st.caption(f"📁 {EXCEL_FILE}")

page = st.session_state["active_page"]
DARK_MODE = st.session_state.get("dark_mode", False)

LIGHT = not DARK_MODE


COLORS = {
    # Backgrounds
    "bg":              "#EDEDEB",
    "card":            "#FFFFFF",
    "card2":           "#F5F5F3",
    "sidebar":         "#FFFFFF",

    # Text
    "text":            "#1A1A2E",
    "muted":           "#6B7280",
    "label":           "#1A1A2E",
    "sidebar_text":    "#1A1A2E",

    # Borders
    "border":          "rgba(0,0,0,0.08)",
    "border_hover":    "rgba(27,63,139,0.25)",

    # Accents
    "active":          "#7ED321",
    "active_text":     "#FFFFFF",
    "blue":            "#1B3F8B",
    "blue_text":       "#FFFFFF",
    "accent":          "#FF7124",
    "accent_dark":     "#D95A10",
    "accent_soft":     "rgba(255,113,36,0.12)",
    "commercial":      "#FF7124",
    "commercial_dark": "#D95A10",
    "commercial_soft": "rgba(255,113,36,0.12)",
    "intel":           "#1B3F8B",
    "intel_soft":      "#7ED321",
    "success":         "#7ED321",
    "warning":         "#FF7124",
    "danger":          "#FF4D2E",
    "negative":        "#FF4D2E",
    "negative_soft":   "rgba(255,77,46,0.12)",
    "input_bg":        "#F5F5F3",
    "info_bg":         "rgba(27,63,139,0.08)",
    "info_text":       "#1A1A2E",
}

flag_html = '<img class="flag-img" src="https://flagcdn.com/w80/ar.png">'


# =========================
# CSS
# =========================

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700;800&display=swap');

* {{ font-family: 'DM Sans', sans-serif; }}

/* ── APP BACKGROUND — glassmorphism light ── */
.stApp {{
    background: #EDEDEB !important;
    background-image:
        radial-gradient(ellipse 70% 60% at 15% 20%, rgba(27,63,139,0.18) 0%, transparent 65%),
        radial-gradient(ellipse 60% 50% at 85% 75%, rgba(255,113,36,0.12) 0%, transparent 60%),
        radial-gradient(ellipse 50% 40% at 50% 50%, rgba(126,211,33,0.07) 0%, transparent 55%),
        radial-gradient(ellipse 40% 35% at 70% 10%, rgba(27,63,139,0.10) 0%, transparent 50%) !important;
    color: {COLORS["text"]};
}}

/* Glass overlay on main content area */
.stApp > div:first-child {{
    backdrop-filter: blur(0px);
}}

/* ── SIDEBAR ── */
section[data-testid="stSidebar"] {{
    background: rgba(255,255,255,0.55) !important;
    backdrop-filter: blur(32px) saturate(180%) !important;
    -webkit-backdrop-filter: blur(32px) saturate(180%) !important;
    border-right: 1px solid rgba(255,255,255,0.7) !important;
    box-shadow: 4px 0 24px rgba(27,63,139,0.06) !important;
}}

section[data-testid="stSidebar"] * {{
    color: {COLORS["sidebar_text"]} !important;
}}

section[data-testid="stSidebar"] .stRadio label {{
    display: flex;
    align-items: center;
    padding: 8px 10px;
    border-radius: 12px;
    font-size: 13px;
    font-weight: 600;
    color: rgba(107,114,128,0.8) !important;
    transition: background 0.18s, color 0.18s;
    cursor: pointer;
}}
section[data-testid="stSidebar"] .stRadio label:hover {{
    background: rgba(27,63,139,0.08);
    color: #1A1A2E !important;
}}
section[data-testid="stSidebar"] .stRadio [data-checked="true"] label,
section[data-testid="stSidebar"] .stRadio input:checked + label {{
    background: rgba(27,63,139,0.12);
    color: #1B3F8B !important;
    border: 1px solid rgba(27,63,139,0.25);
    border-radius: 12px;
    font-weight: 700;
}}

/* ── BLOCK CONTAINER ── */
.block-container {{
    max-width: 1400px;
    padding-top: 2rem;
    padding-bottom: 5rem;
}}

/* ── VERTICAL SPACING between Streamlit elements ── */
div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlockBorderWrapper"],
div[data-testid="stVerticalBlock"] > div.element-container {{
    margin-bottom: 8px !important;
}}
div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlockBorderWrapper"]:has(.stack-card),
div[data-testid="stVerticalBlock"] > div.element-container:has(.stack-card) {{
    margin-bottom: 16px !important;
}}

/* ── TEXT GLOBAL ── */
[data-testid="stMarkdownContainer"] h1,
[data-testid="stMarkdownContainer"] h2,
[data-testid="stMarkdownContainer"] h3,
[data-testid="stMarkdownContainer"] h4,
[data-testid="stMarkdownContainer"] h5,
[data-testid="stMarkdownContainer"] h6,
[data-testid="stMarkdownContainer"] p,
[data-testid="stCaptionContainer"],
[data-testid="stCaptionContainer"] * {{
    color: {COLORS["text"]} !important;
}}

[data-testid="stDataFrame"] *,
[data-testid="stTable"] *,
[data-testid="stAlert"] *,
.stTextInput *,
.stTextArea textarea,
.stSelectbox *,
.stNumberInput * {{
    color: {COLORS["text"]};
}}

/* ── INPUTS ── */
input, textarea, select {{
    background: rgba(255,255,255,0.85) !important;
    color: #1A1A2E !important;
    border-radius: 12px !important;
    border: 1px solid rgba(0,0,0,0.10) !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04) !important;
}}
input::placeholder, textarea::placeholder {{
    color: rgba(107,114,128,0.5) !important;
}}
div[data-testid="stTextInput"] label,
div[data-testid="stSelectbox"] label,
div[data-testid="stNumberInput"] label,
div[data-testid="stTextArea"] label {{
    color: #6B7280 !important;
    font-weight: 700 !important;
    font-size: 12px !important;
    text-transform: uppercase;
    letter-spacing: .06em;
}}

div[data-testid="column"] {{ padding: 0 10px !important; }}
div[data-testid="stHorizontalBlock"] {{ gap: 20px !important; }}

/* ── FLAG ── */
.flag-img {{
    width: 44px; height: 29px; border-radius: 6px;
    object-fit: cover; margin-right: 12px; vertical-align: middle;
}}

/* ── APP HEADER — sticker style ── */
.app-header {{
    background: #FFFFFF;
    border: none;
    border-radius: 24px;
    padding: 24px 32px;
    margin-bottom: 28px;
    display: flex;
    justify-content: space-between;
    align-items: center;
    box-shadow: 0 4px 24px rgba(27,63,139,0.10), 0 1px 4px rgba(0,0,0,0.06);
    transition: box-shadow .2s, transform .2s;
}}
.app-header:hover {{
    box-shadow: 0 8px 32px rgba(27,63,139,0.15), 0 2px 8px rgba(0,0,0,0.08);
    transform: translateY(-1px);
}}

.header-title {{
    font-size: 36px;
    font-weight: 800;
    color: #1A1A2E;
    display: flex;
    align-items: center;
    line-height: 1;
}}

.header-subtitle {{
    margin-top: 8px;
    font-size: 14px;
    font-weight: 600;
    color: #6B7280;
}}

.period-pill {{
    font-size: 12px;
    font-weight: 700;
    color: #7ED321;
    background: rgba(126,211,33,0.10);
    border: 1px solid rgba(126,211,33,0.30);
    border-radius: 20px;
    padding: 5px 14px;
    letter-spacing: .06em;
}}

.section-title {{
    font-size: 13px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: .1em;
    color: rgba(107,114,128,0.6);
    margin: 28px 0 14px;
}}

/* ── STICKER CARD BASE (reemplaza glassmorphism) ── */
.glass-card,
.metric-card,
.mgmt-card,
.mgmt-section,
.panel,
.kpi-card,
.update-card,
.agenda-card,
.brand-card,
.status-card,
.hero-card,
.stack-card,
.wide-info-card,
.info-card,
.comments-card,
.salary-card,
.bucket-card {{
    background: #FFFFFF !important;
    backdrop-filter: none !important;
    -webkit-backdrop-filter: none !important;
    border: none !important;
    border-radius: 20px !important;
    box-shadow: 0 4px 20px rgba(27,63,139,0.08), 0 1px 4px rgba(0,0,0,0.05) !important;
    transition: transform .25s cubic-bezier(.34,1.56,.64,1), box-shadow .2s ease !important;
    padding: 20px 22px 18px !important;
    margin-bottom: 14px;
    display: flex;
    flex-direction: column;
    gap: 6px;
}}

/* ── WIDE INFO CARD: padding for section containers ── */
.wide-info-card {{
    padding: 24px 26px 22px !important;
    margin-bottom: 20px;
}}
.wide-info-card:hover {{
    transform: none !important;
    box-shadow: 0 4px 20px rgba(27,63,139,0.08), 0 1px 4px rgba(0,0,0,0.05) !important;
}}

.glass-card:hover,
.metric-card:hover,
.mgmt-card:hover,
.mgmt-section:hover,
.panel:hover,
.kpi-card:hover,
.update-card:hover,
.agenda-card:hover,
.brand-card:hover,
.status-card:hover,
.hero-card:hover,
.stack-card:hover,
.info-card:hover,
.comments-card:hover,
.salary-card:hover,
.bucket-card:hover {{
    transform: translateY(-3px) scale(1.01) !important;
    box-shadow: 0 12px 36px rgba(27,63,139,0.14), 0 3px 10px rgba(0,0,0,0.07) !important;
}}

/* ── BUSINESS INFORMATION CARDS (Brand Finder) ── */
.business-card-grid {{
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 16px;
    margin-top: 14px;
}}
@media (max-width: 1100px) {{
    .business-card-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
}}

.business-mini-card {{
    background: #FFFFFF !important;
    border: 1px solid rgba(27,63,139,0.09) !important;
    border-radius: 16px !important;
    padding: 16px 18px 14px !important;
    box-shadow: 0 2px 10px rgba(27,63,139,0.07), 0 1px 3px rgba(0,0,0,0.04) !important;
    display: flex;
    flex-direction: column;
    gap: 2px;
    min-height: 80px;
    position: relative;
    transition: transform .2s, box-shadow .2s;
}}
.business-mini-card:hover {{
    transform: translateY(-2px);
    box-shadow: 0 8px 24px rgba(27,63,139,0.12), 0 2px 6px rgba(0,0,0,0.06) !important;
}}
.business-mini-card.lever-ads   {{ border-left: 3px solid #FF7124 !important; }}
.business-mini-card.lever-md    {{ border-left: 3px solid #1B3F8B !important; }}
.business-mini-card.lever-pro   {{ border-left: 3px solid #7ED321 !important; }}
.business-mini-card.lever-menu  {{ border-left: 3px solid rgba(107,114,128,0.3) !important; }}

.card-label {{
    font-size: 10px !important;
    font-weight: 800 !important;
    text-transform: uppercase !important;
    letter-spacing: .08em !important;
    color: rgba(107,114,128,0.6) !important;
    margin-bottom: 2px;
}}
.card-value {{
    font-size: 18px !important;
    font-weight: 800 !important;
    color: #1A1A2E !important;
    line-height: 1.15;
    margin-bottom: 2px;
}}
.card-copy {{
    font-size: 11px !important;
    font-weight: 500 !important;
    color: rgba(107,114,128,0.65) !important;
    line-height: 1.4;
}}
.card-chip {{
    display: inline-block;
    font-size: 10px;
    font-weight: 700;
    padding: 2px 8px;
    border-radius: 999px;
    background: rgba(27,63,139,0.07);
    color: #1B3F8B;
    margin-top: 4px;
}}
.card-chipline {{ margin-top: 6px; }}

/* ── 360° ACTION PRIORITY HEADER GRID ── */
.priority-top-grid {{
    display: grid;
    grid-template-columns: repeat(5, minmax(0, 1fr));
    gap: 10px 16px;
    background: rgba(27,63,139,0.03);
    border-radius: 12px;
    padding: 12px 16px;
    margin-bottom: 14px;
}}
@media (max-width: 1100px) {{
    .priority-top-grid {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
}}
.priority-levers {{
    padding-top: 8px;
    display: flex;
    flex-wrap: wrap;
    gap: 6px;
}}
.priority-chip {{
    display: inline-block;
    font-size: 11px;
    font-weight: 700;
    padding: 4px 10px;
    border-radius: 999px;
    background: rgba(27,63,139,0.08);
    color: #1B3F8B;
    border: 1px solid rgba(27,63,139,0.15);
}}

/* ── TEXT COLORS INSIDE CARDS ── */
.metric-label, .mgmt-label, .kpi-label,
.stack-label, .sticker-label, .hero-info-label, .info-mini-label,
.update-title, .small-muted, .wide-info-title {{
    color: rgba(107,114,128,0.7) !important;
}}

.metric-value, .mgmt-value, .kpi-value,
.stack-main, .brand-name, .hero-name, .status-value,
.metric-title, .mgmt-section-title {{
    color: #1A1A2E !important;
}}

/* ── SALARY / BUCKET CARDS: jerarquía tipográfica explícita ── */
.salary-label, .bucket-title {{
    color: rgba(107,114,128,0.75) !important;
    font-size: 11px !important;
    font-weight: 700 !important;
    text-transform: uppercase;
    letter-spacing: .04em;
    margin-bottom: 8px;
    line-height: 1.3;
}}
.salary-value {{
    color: #1A1A2E !important;
    font-size: 22px !important;
    font-weight: 800 !important;
    line-height: 1.15;
}}
.bucket-card .kpi-label {{
    font-size: 10px !important;
    font-weight: 700 !important;
    text-transform: uppercase;
    letter-spacing: .04em;
    color: rgba(107,114,128,0.6) !important;
    margin-bottom: 4px;
}}
.bucket-card .kpi-value {{
    font-size: 22px !important;
    font-weight: 800 !important;
    color: #1A1A2E !important;
}}

/* ── MANAGEMENT DASHBOARD TYPOGRAPHY HIERARCHY ── */
.mgmt-ars {{
    font-size: 32px !important;
    font-weight: 900 !important;
    color: #1A1A2E !important;
    letter-spacing: -0.02em;
    line-height: 1.1;
}}
.mgmt-conv {{
    font-size: 14px !important;
    font-weight: 500 !important;
    color: #6B7280 !important;
    margin-top: 4px !important;
    line-height: 1.4;
}}
.mgmt-section-title-card {{
    background: transparent !important;
    box-shadow: none !important;
    padding: 0 !important;
    margin-bottom: 8px !important;
}}
.mgmt-section-title-copy {{
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: .12em;
    color: rgba(107,114,128,0.55);
}}

.metric-foot, .mgmt-foot, .stack-foot, .stack-sub,
.hero-id, .hero-info-value, .info-mini-value {{
    color: rgba(107,114,128,0.7) !important;
}}

/* ── COLOR ACCENT OVERRIDES ── */
.up, .usd {{ color: #7ED321 !important; }}
.ars       {{ color: #FF7124 !important; }}
.cop       {{ color: #1B3F8B !important; }}
.blue-val  {{ color: #1B3F8B !important; }}
.negative  {{ color: #FF4D2E !important; }}

/* ── BADGES / PILLS ── */
.brand-badge {{
    color: #FF7124 !important;
    background: rgba(255,113,36,0.08) !important;
    border: 1.5px solid rgba(255,113,36,0.3) !important;
    border-radius: 999px;
    padding: 6px 12px;
    font-size: 12px;
    font-weight: 800;
    white-space: nowrap;
    transition: background .2s, color .2s;
}}
.brand-badge:hover {{
    background: #FF7124 !important;
    color: #FFFFFF !important;
}}

.category-chip {{
    background: rgba(255,113,36,0.07) !important;
    border: 1.5px solid rgba(255,113,36,0.25) !important;
    color: #1A1A2E !important;
    border-radius: 999px; padding: 5px 10px;
    font-weight: 800; font-size: 12px;
}}
.category-chip:hover {{ background: #FF7124 !important; color: #FFFFFF !important; }}

.priority-chip {{
    background: rgba(27,63,139,0.08) !important;
    border: 1.5px solid rgba(27,63,139,0.25) !important;
    border-radius: 999px; padding: 7px 11px;
    font-weight: 800; font-size: 12px; line-height: 1.15;
    color: #1B3F8B !important;
}}
.priority-chip:hover {{ background: #1B3F8B !important; color: #FFFFFF !important; }}

/* ── CAMPAIGN DESIGNER MINI CARDS ── */
.campaign-mini-card {{
    background: #FFFFFF !important;
    border: 1px solid rgba(27,63,139,0.10) !important;
    border-radius: 14px !important;
    box-shadow: 0 2px 10px rgba(27,63,139,0.07), 0 1px 3px rgba(0,0,0,0.04) !important;
    transition: transform .2s cubic-bezier(.34,1.56,.64,1), box-shadow .2s ease !important;
}}
.campaign-mini-card:hover {{
    transform: translateY(-3px) scale(1.01) !important;
    box-shadow: 0 10px 28px rgba(27,63,139,0.14), 0 3px 8px rgba(0,0,0,0.07) !important;
}}
.campaign-mini-card.lever-ads   {{ border-left: 3px solid #FF7124 !important; }}
.campaign-mini-card.lever-md    {{ border-left: 3px solid #1B3F8B !important; }}
.campaign-mini-card.lever-pro   {{ border-left: 3px solid #7ED321 !important; }}
.campaign-mini-card.lever-menu  {{ border-left: 3px solid rgba(107,114,128,0.3) !important; }}
.campaign-mini-card.lever-ops   {{ border-left: 3px solid rgba(27,63,139,0.25) !important; }}


.signal-pill {{
    background: rgba(255,113,36,0.07) !important;
    border: 1.5px solid rgba(255,113,36,0.25) !important;
    color: #1A1A2E !important;
    border-radius: 999px; padding: 9px 13px;
    font-size: 13px; font-weight: 800;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}}

/* ── PRIORITY PILLS ── */
.priority-pill {{ padding: 6px 13px; border-radius: 999px; font-weight: 800; display: inline-block; }}
.high {{ color: #FF4D2E; border: 1.5px solid #FF4D2E; }}
.mid  {{ color: #FF7124; border: 1.5px solid #FF7124; }}
.low  {{ color: rgba(27,63,139,0.6); border: 1.5px solid rgba(27,63,139,0.4); background: rgba(27,63,139,0.06); }}
.done {{ color: rgba(107,114,128,0.4); border: 1.5px solid rgba(107,114,128,0.2); }}

/* ── BOXES ── */
.target-box, .result-box, .variable-box, .zero-box, .lime-box,
.success-box, .legend-box, .net-card {{
    background: #FFFFFF !important;
    border-radius: 18px;
    padding: 16px 12px;
    text-align: center;
    color: #1A1A2E;
    box-shadow: 0 2px 12px rgba(0,0,0,0.06);
}}
.target-box  {{ border: 1.5px solid rgba(126,211,33,0.4) !important; color: #7ED321 !important; }}
.result-box  {{ border: 1.5px solid rgba(27,63,139,0.3) !important;  color: #1B3F8B !important; }}
.variable-box {{ border: 1.5px solid rgba(27,63,139,0.3) !important; color: #1B3F8B !important;
    min-height: 150px; display: flex; flex-direction: column; justify-content: center; }}
.zero-box    {{ border: 1.5px solid rgba(255,77,46,0.4) !important;  color: #FF4D2E !important; }}
.lime-box    {{ border: 1.5px solid rgba(126,211,33,0.4) !important; color: #1B3F8B !important; font-size: 24px; font-weight: 800; }}
.success-box {{ border: 2px solid #7ED321 !important; color: #7ED321 !important; font-size: 18px; font-weight: 800; margin: 16px 0; }}
.net-card    {{ border: 2px solid rgba(126,211,33,0.4) !important; }}
.net-card .salary-label, .net-card .salary-value {{ color: #7ED321 !important; }}
.legend-box  {{ border: 1.5px solid rgba(27,63,139,0.3) !important; color: #1B3F8B !important; font-size: 16px; font-weight: 700; margin-top: 18px; }}

/* ── STICKERS (mini) ── */
.sticker {{
    background: #FFFFFF !important;
    border: none !important;
    border-radius: 16px; padding: 12px 14px; min-height: 70px;
    box-shadow: 0 2px 12px rgba(27,63,139,0.08), 0 1px 3px rgba(0,0,0,0.04);
    transition: box-shadow .2s, transform .2s;
}}
.sticker:hover {{
    box-shadow: 0 8px 24px rgba(27,63,139,0.14) !important;
    transform: translateY(-2px);
}}
.sticker-value {{ font-size: 16px; font-weight: 800; margin-top: 6px; color: #1A1A2E !important; }}

/* ── ACTION CARDS ── */
.action-grid {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 24px; }}
.action-card {{
    background: #FFFFFF !important;
    border: none !important;
    border-radius: 18px; padding: 20px 18px; min-height: 180px;
    box-shadow: 0 4px 16px rgba(27,63,139,0.08), 0 1px 4px rgba(0,0,0,0.05);
    transition: box-shadow .2s, transform .25s cubic-bezier(.34,1.56,.64,1);
    display: flex; flex-direction: column;
}}
.action-card:hover {{
    transform: translateY(-3px) scale(1.01) !important;
    box-shadow: 0 12px 32px rgba(27,63,139,0.14), 0 3px 8px rgba(0,0,0,0.07) !important;
}}

/* ── TABLE CONTAINER ── */
.table-glass {{
    background: #FFFFFF;
    border: none;
    border-radius: 16px;
    overflow: hidden;
    margin-top: 16px;
    box-shadow: 0 4px 20px rgba(27,63,139,0.08), 0 1px 4px rgba(0,0,0,0.05);
}}

.table-glass table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 13px;
}}
.table-glass thead tr {{
    background: rgba(27,63,139,0.04);
    border-bottom: 1px solid rgba(0,0,0,0.06);
}}
.table-glass thead th {{
    text-align: left;
    font-size: 10px;
    font-weight: 700;
    color: rgba(107,114,128,0.6);
    letter-spacing: .1em;
    text-transform: uppercase;
    padding: 10px 16px;
}}
.table-glass tbody tr {{
    border-bottom: 1px solid rgba(0,0,0,0.04);
    transition: background .15s, border-left .15s;
    border-left: 3px solid transparent;
}}
.table-glass tbody tr:nth-child(even) {{ background: rgba(27,63,139,0.02); }}
.table-glass tbody tr:hover {{ background: rgba(27,63,139,0.04); }}
.table-glass tbody tr.row-green:hover  {{ border-left: 3px solid #7ED321; }}
.table-glass tbody tr.row-orange:hover {{ border-left: 3px solid #FF7124; }}
.table-glass tbody tr.row-blue:hover   {{ border-left: 3px solid #1B3F8B; }}
.table-glass td {{
    padding: 10px 16px;
    color: rgba(107,114,128,0.8);
}}
.table-glass td.td-name   {{ color: #1A1A2E; font-weight: 600; }}
.table-glass td.td-green  {{ color: #7ED321; font-weight: 600; }}
.table-glass td.td-orange {{ color: #FF7124; font-weight: 600; }}
.table-glass td.td-red    {{ color: #FF4D2E; font-weight: 600; }}
.table-glass td.td-blue   {{ color: #1B3F8B; font-weight: 600; }}
.table-glass td.td-muted  {{ color: rgba(107,114,128,0.4); }}

.table-header-bar {{
    display: flex; align-items: center; justify-content: space-between;
    padding: 12px 16px; border-bottom: 1px solid rgba(0,0,0,0.05);
}}
.table-title {{ font-size: 13px; font-weight: 600; color: #1A1A2E; }}
.table-footer {{
    padding: 10px 16px; border-top: 1px solid rgba(0,0,0,0.04);
    display: flex; justify-content: space-between; align-items: center;
    font-size: 10px; color: rgba(107,114,128,0.4);
}}

/* ── MGMT LAYOUT GRIDS ── */
.mgmt-row {{ display: grid; gap: 24px; margin-bottom: 24px; align-items: stretch; }}
.mgmt-row.three {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
.mgmt-row.two   {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
.mgmt-row.line4 {{ grid-template-columns: 1fr 1fr 2.35fr; }}
.mgmt-subgrid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 24px; }}
.mgmt-subgrid.three {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
.mgmt-icon {{
    width: 44px; height: 44px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    color: white; font-size: 22px; font-weight: 800; margin-bottom: 14px;
}}
.mgmt-clean-card {{ display: flex; flex-direction: column; justify-content: center; }}

@media (max-width: 1100px) {{
    .mgmt-row.three, .mgmt-row.two, .mgmt-row.line4,
    .mgmt-subgrid, .mgmt-subgrid.three {{ grid-template-columns: 1fr; }}
}}

/* ── HERO CARD (Brand Finder) ── */
.hero-card {{ padding: 28px 30px 24px !important; margin-bottom: 20px; }}
.hero-grid {{ display: grid; grid-template-columns: 1.25fr 1.05fr 1.25fr; gap: 28px; align-items: center; }}
.hero-name {{ font-size: clamp(28px, 3vw, 46px); font-weight: 900; line-height: 1.02; color: #1A1A2E; letter-spacing: -0.04em; }}
.hero-name-row {{ display: flex; align-items: center; gap: 12px; flex-wrap: wrap; }}
.hero-mundialista-badge {{
    display: inline-flex; align-items: center; gap: 7px; border-radius: 999px;
    padding: 6px 10px; font-size: 11px; font-weight: 900; text-transform: uppercase;
    letter-spacing: .02em; color: #FFFFFF;
    background: #FF7124; border: none; white-space: nowrap;
}}
.hero-id {{ color: rgba(107,114,128,0.6); font-size: 18px; font-weight: 800; margin-top: 10px; }}
.hero-divider {{ border: none; border-top: 1px solid rgba(0,0,0,0.07); margin: 18px 0 14px; }}
.hero-info-grid {{ display: grid; grid-template-columns: repeat(4, minmax(0,1fr)); gap: 20px 24px; }}
.hero-info-item {{ display: flex; flex-direction: column; gap: 4px; }}
.hero-info-label {{ font-size: 10px; text-transform: uppercase; font-weight: 800; color: rgba(107,114,128,0.5); }}

/* ── BUTTONS (incluye sidebar — el naranja es el estilo de marca en ambos modos) ── */
.stButton > button {{
    background: #FF7124 !important;
    color: #FFFFFF !important;
    border: none !important;
    border-radius: 12px !important;
    font-weight: 700 !important;
    font-size: 13px !important;
    padding: 10px 20px !important;
    box-shadow: 0 4px 12px rgba(255,113,36,0.25) !important;
    transition: box-shadow .2s, transform .15s !important;
}}
.stButton > button:hover {{
    background: #D95A10 !important;
    box-shadow: 0 6px 18px rgba(255,113,36,0.35) !important;
    transform: translateY(-1px) !important;
}}
.stButton > button:active {{
    transform: translateY(0) !important;
    box-shadow: 0 2px 8px rgba(255,113,36,0.2) !important;
}}

/* ── PROGRESS BARS ── */
.stProgress > div > div > div > div {{
    background: linear-gradient(90deg, #FF7124, #FF4D2E) !important;
    border-radius: 999px !important;
}}
.stProgress > div > div > div {{
    background: rgba(0,0,0,0.07) !important;
    border-radius: 999px !important;
}}

/* ── SELECTBOX / DROPDOWN ── */
div[data-testid="stSelectbox"] > div > div {{
    background: #FFFFFF !important;
    border: 1px solid rgba(0,0,0,0.10) !important;
    border-radius: 12px !important;
    color: #1A1A2E !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06) !important;
}}

/* ── EXPANDER ── */
details summary {{
    background: #FFFFFF !important;
    border-radius: 12px !important;
    padding: 10px 16px !important;
    color: #1A1A2E !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 8px rgba(27,63,139,0.06) !important;
    border: none !important;
}}

/* ── SCROLLBAR ── */
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: rgba(0,0,0,0.04); border-radius: 999px; }}
::-webkit-scrollbar-thumb {{ background: rgba(27,63,139,0.2); border-radius: 999px; }}
::-webkit-scrollbar-thumb:hover {{ background: rgba(27,63,139,0.35); }}

/* ── CALENDAR (Agenda) ── */
.cal-grid {{ background: #FFFFFF; border-radius: 16px; box-shadow: 0 4px 20px rgba(27,63,139,0.08); overflow: hidden; }}
.cal-header-cell {{
    background: rgba(27,63,139,0.04);
    color: #6B7280 !important;
    font-size: 11px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: .06em;
    padding: 10px 8px;
    text-align: center;
    border-bottom: 1px solid rgba(0,0,0,0.06);
}}
.cal-day-today {{
    background: #1B3F8B !important;
    color: #FFFFFF !important;
    border-radius: 50%;
    width: 28px; height: 28px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-weight: 700;
}}
.cal-time-label {{
    color: #6B7280 !important;
    font-size: 11px !important;
    font-weight: 500 !important;
    padding: 4px 8px;
    white-space: nowrap;
}}
.cal-cell {{
    border: 0.5px solid rgba(0,0,0,0.05);
    min-height: 80px;
    padding: 4px;
    background: #FFFFFF;
    color: #1A1A2E !important;
    font-size: 12px;
}}
.cal-cell:hover {{ background: rgba(27,63,139,0.03); }}
.cal-event {{
    border-radius: 6px;
    padding: 3px 7px;
    font-size: 11px;
    font-weight: 600;
    margin-bottom: 2px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}}
/* Fix Streamlit calendar/agenda table text */
[data-testid="stMarkdownContainer"] table td,
[data-testid="stMarkdownContainer"] table th {{
    color: #1A1A2E !important;
    border-color: rgba(0,0,0,0.06) !important;
}}
/* Generic fix for any white/near-white text on light bg */
span[style*="color: white"],
span[style*="color:#FFFFFF"],
span[style*="color: #FFFFFF"] {{
    color: #1A1A2E !important;
}}

/* ── MGMT STACK CARDS ── */
.mgmt-stack-card {{
    background: #FFFFFF !important;
    border: none !important;
    border-radius: 20px !important;
    box-shadow: 0 4px 20px rgba(27,63,139,0.08), 0 1px 4px rgba(0,0,0,0.05) !important;
}}

/* ── ALERT / INFO BOXES ── */
div[data-testid="stAlert"] {{
    background: rgba(27,63,139,0.06) !important;
    border: 1px solid rgba(27,63,139,0.15) !important;
    border-radius: 12px !important;
    color: #1A1A2E !important;
}}

/* ── NEGATIVE DATA INDICATORS ── */
.negative-val {{ color: #FF4D2E !important; font-weight: 700; }}
.outline-negative {{
    color: #FF4D2E !important;
    border: 1.5px solid #FF4D2E !important;
    background: transparent !important;
    border-radius: 8px;
    padding: 2px 8px;
    font-weight: 700;
    font-size: 12px;
}}

/* ── DATAFRAME OVERRIDES ── */
[data-testid="stDataFrame"] {{
    border-radius: 16px !important;
    overflow: hidden !important;
    box-shadow: 0 4px 20px rgba(27,63,139,0.08) !important;
}}
[data-testid="stDataFrame"] th {{
    background: rgba(27,63,139,0.04) !important;
    color: #6B7280 !important;
    font-size: 11px !important;
    font-weight: 700 !important;
    text-transform: uppercase !important;
    letter-spacing: .06em !important;
}}
[data-testid="stDataFrame"] td {{
    color: #1A1A2E !important;
    font-size: 13px !important;
}}
[data-testid="stDataFrame"] tr:nth-child(even) {{
    background: rgba(27,63,139,0.02) !important;
}}

</style>
""", unsafe_allow_html=True)

# ── DARK MODE OVERRIDE ────────────────────────────────────────────────────
# Mantiene la misma estructura visual, reemplaza el gris claro de fondo por
# un tono entre negro y azul medianoche, y ajusta texto/bordes/superficies
# para que sigan siendo legibles sobre ese fondo. Se activa con el toggle
# del sidebar (st.session_state["dark_mode"]).
if DARK_MODE:
    st.markdown("""
    <style>
    /* ── Fondo base: negro azulado medianoche ── */
    .stApp, [data-testid="stAppViewContainer"], [data-testid="stHeader"] {
        background: #0B0E1A !important;
    }
    section[data-testid="stSidebar"] > div:first-child {
        background: #101424 !important;
    }
    section[data-testid="stSidebar"] * {
        color: #C7CCDC !important;
    }
    .nav-item { color: #9AA3BD !important; }
    .nav-item:hover { background: rgba(255,255,255,0.06) !important; color: #FFFFFF !important; }
    .nav-item.active { background: rgba(126,211,33,0.14) !important; color: #FFFFFF !important; }
    .nav-logo-text, .nav-section-label { color: #C7CCDC !important; }
    .nav-toggle-btn { background: #161B2E !important; color: #C7CCDC !important; border-color: rgba(255,255,255,0.10) !important; }

    /* ── Vencer la regla base de mayor especificidad que fuerza texto oscuro
       en th/td de tablas HTML (Opportunity List, ADS/MD tables, etc.) — debe
       tener especificidad igual (mismo selector compuesto) para ganar. ── */
    [data-testid="stMarkdownContainer"] table td,
    [data-testid="stMarkdownContainer"] table th {
        color: #E4E7F1 !important;
        border-color: rgba(255,255,255,0.10) !important;
    }
    span[style*="color: white"],
    span[style*="color:#FFFFFF"],
    span[style*="color: #FFFFFF"] {
        color: #F2F4F9 !important;
    }

    /* ── TODAS las superficies blancas conocidas → azul-noche oscuro ──
       Cubre cards, headers, boxes, stickers, tablas y widgets nativos.
       Sin esto, cualquier clase fuera de esta lista se queda blanca. */
    .glass-card, .metric-card, .mgmt-card, .mgmt-section, .panel, .kpi-card,
    .update-card, .agenda-card, .brand-card, .status-card, .hero-card,
    .stack-card, .wide-info-card, .info-card, .comments-card, .salary-card,
    .bucket-card, .campaign-mini-card, .multibrand-box, .business-card-grid > div,
    .app-header, .business-mini-card, .target-box, .result-box, .variable-box,
    .zero-box, .lime-box, .success-box, .legend-box, .net-card, .sticker,
    .action-card, .table-glass, .cal-grid, .mgmt-stack-card, .hm-card, .hm-zone,
    .pareto-card, .qt-score-card, .qt-trainer-card, .qt-card-header, .qt-example,
    .hm-insight, .day-col, .hour-col, .evt,
    div[data-testid="stSelectbox"] > div > div, details summary {
        background: #141A2E !important;
        box-shadow: 0 4px 20px rgba(0,0,0,0.35), 0 1px 4px rgba(0,0,0,0.25) !important;
    }

    /* ── Texto: neutralizar los colores oscuros hardcodeados (#1A1A2E / #6B7280 / negros)
       SIN tocar los colores semánticos (verde/rojo/naranja/azul) que ya contrastan bien
       sobre fondo oscuro. Se apunta directo a los valores de color usados en HTML inline. ──*/
    .stApp, .stApp p, .stApp span, .stApp div, .stApp label, .stApp li, .stApp td, .stApp th {
        color: #E4E7F1 !important;
    }
    /* Las clases de "label"/título secundario quedan en gris claro, no gris oscuro */
    .stack-label, .sticker-label, .hero-info-label, .info-mini-label,
    .card-label, .kpi-title, .salary-label, .bucket-title, .small-muted,
    .pareto-meta, .pareto-row-label, .header-subtitle, .nav-logo-sub,
    .hero-id, .update-title, .wide-info-title {
        color: #8C93AC !important;
    }
    /* Los valores principales quedan en blanco/casi-blanco, no negro */
    .hero-name, .hero-info-value, .kpi-value, .salary-value,
    .stack-main, .brand-name, .status-value, .metric-title,
    .mgmt-section-title, .pareto-name, .pareto-row-value,
    .header-title, .bucket-card .kpi-value {
        color: #F2F4F9 !important;
    }

    /* ── Neutralizar SOLO los colores de texto neutros inline (no los semánticos) ──
       Cualquier span/div con color inline #1A1A2E (texto oscuro estándar) o
       rgba(107,114,128,*) (gris secundario) se reescribe a claro. Verde/rojo/
       naranja/azul semánticos quedan intactos porque no se tocan sus selectores. */
    [style*="color:#1A1A2E"], [style*="color: #1A1A2E"],
    [style*="color:#1a1a2e"], [style*="color: #1a1a2e"] {
        color: #F2F4F9 !important;
    }
    [style*="color:rgba(107,114,128"], [style*="color: rgba(107,114,128"],
    [style*="color:#6B7280"], [style*="color: #6B7280"],
    [style*="color:#aaa"], [style*="color: #aaa"] {
        color: #8C93AC !important;
    }

    /* ── Superficies con background blanco translúcido INLINE (no clase CSS) ──
       Decenas de cards en Analytics, Day Queue, Opportunity List, Pitch Facts,
       Brand vs Brand, etc. usan style="background:rgba(255,255,255,0.9X)" directo
       en el HTML en vez de una clase — sin esto se quedan blancas en dark mode. */
    [style*="background:rgba(255,255,255,0.9"],
    [style*="background: rgba(255,255,255,0.9"],
    [style*="background:rgba(255, 255, 255, 0.9"],
    [style*="background: rgba(255, 255, 255, 0.9"] {
        background: #141A2E !important;
    }
    [style*="background:rgba(255,255,255,0.92)"],
    [style*="background:rgba(255,255,255,0.95)"],
    [style*="background:rgba(255,255,255,0.97)"] {
        background: #1B2238 !important;
    }
    /* Hex blanco sólido inline — viene de COLORS["card"] insertado vía f-string
       (Opportunity List ADS/MD/Churn target bars, entre otras). */
    [style*="background:#FFFFFF;border:1px solid"],
    [style*="background: #FFFFFF;border:1px solid"],
    [style*="background:#FFFFFF; border:1px solid"] {
        background: #141A2E !important;
    }

    /* ── Inputs / widgets nativos de Streamlit ── */
    [data-testid="stTextInput"] input, [data-testid="stNumberInput"] input,
    [data-testid="stTextArea"] textarea, [data-testid="stSelectbox"] div[data-baseweb="select"],
    [data-testid="stDateInput"] input, [data-testid="stTimeInput"] input,
    [data-testid="stDateInput"] div[data-baseweb], [data-testid="stTimeInput"] div[data-baseweb] {
        background: #1B2238 !important;
        color: #E4E7F1 !important;
        border-color: rgba(255,255,255,0.10) !important;
    }
    [data-testid="stCheckbox"] label {
        color: #E4E7F1 !important;
    }
    [data-testid="stExpander"] {
        background: #141A2E !important;
        border-color: rgba(255,255,255,0.10) !important;
    }
    [data-testid="stDataFrame"] {
        background: #141A2E !important;
    }
    [data-testid="stDataFrame"] table, [data-testid="stDataFrame"] th, [data-testid="stDataFrame"] td {
        background: #141A2E !important;
        color: #E4E7F1 !important;
        border-color: rgba(255,255,255,0.08) !important;
    }
    [data-testid="stDataFrame"] tr:nth-child(even) td {
        background: rgba(255,255,255,0.03) !important;
    }

    /* ── Botones: el naranja de marca (#FF7124) se mantiene igual en dark mode —
       no se sobreescribe aquí. Solo se ajusta el resto del chrome alrededor. ── */

    /* ── Dividers / bordes sutiles ── */
    .nav-divider, .hero-divider, hr {
        background: rgba(255,255,255,0.08) !important;
        border-color: rgba(255,255,255,0.08) !important;
    }

    /* ── st_components.html (iframes): Weekly Calendar, Brand vs Brand, etc. ──
       Estos renderizan en un iframe con su propio <body>, fuera del alcance
       del CSS de .stApp — se oscurecen aparte. */
    iframe {
        background: #141A2E !important;
    }
    </style>
    """, unsafe_allow_html=True)

# =========================
# UI HELPERS
# =========================

def render_header(title="Growth OS", subtitle="Commercial Management System · Rappi"):
    """Renderiza el header glass de cada página."""
    from datetime import date as _date
    today = _date.today()
    quarter = (today.month - 1) // 3 + 1
    iso_week = today.isocalendar().week
    period = f"Q{quarter} · W{iso_week} · {today.year}"

    subtitle_html = f'<div class="header-subtitle">{subtitle}</div>' if subtitle else ""
    st.markdown(f"""
    <div class="app-header">
        <div>
            <div class="header-title">{title}</div>
            {subtitle_html}
        </div>
        <div class="period-pill">{period}</div>
    </div>
    """, unsafe_allow_html=True)


def render_metric_card(label, value, color_class, icon="↗", foot="vs last month", delta="▲"):
    icon_bg = {
        "ars": COLORS["accent"],
        "usd": COLORS["success"],
        "cop": COLORS["intel_soft"],
        "blue-val": COLORS["intel_soft"],
    }.get(color_class, COLORS["intel"])

    if foot or delta:
        foot_html = f'<div class="metric-foot">{foot} &nbsp; <span class="up">{delta}</span></div>'
    else:
        foot_html = ''

    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-icon" style="background:{icon_bg};">{icon}</div>
        <div class="metric-label">{label}</div>
        <div class="metric-value {color_class}">{value}</div>
        {foot_html}
    </div>
    """, unsafe_allow_html=True)


def option_index(options, current):
    current = clean(current, default="")
    if current in options:
        return options.index(current)
    return 0


# (sidebar handled in THEME block above)




def _status_col_active_ids(sheet_df, status_candidates):
    """Returns active IDs from Growth OS manual status columns.
    Important: do not count "Inactive" as active just because it contains the word "active".
    """
    if sheet_df.empty:
        return set()
    id_col = get_id_column_name(sheet_df)
    if not id_col:
        return set()
    status_col = _first_existing_col(sheet_df, status_candidates)
    if not status_col:
        return set()

    active = set()
    inactive_markers = ["inactive", "inactivo", "off", "sleep", "sleeping", "paused", "pause", "💤", "😴", "false", "no"]
    active_exact = ["active", "activated", "activo", "activa", "on", "true", "yes", "si", "sí"]

    for _, r in sheet_df.iterrows():
        raw_status = clean(r.get(status_col), "")
        status_text = norm_text(raw_status)

        if any(marker in status_text for marker in inactive_markers):
            continue

        is_active = (
            "🚀" in raw_status
            or status_text in active_exact
            or status_text.startswith("active ")
            or status_text.startswith("activated ")
            or status_text.startswith("activo ")
            or status_text.startswith("activa ")
        )

        if is_active:
            bid = normalize_brand_id(r.get(id_col))
            if bid:
                active.add(bid)

    return active


@st.cache_data(ttl=3000, show_spinner=False)
def get_live_campaign_coverage_counts():
    """Unified calculator for Management Dashboard and Campaign Weekly Tracker.
    Base total = Asignacion Junio (fuente de verdad del portafolio).
    Counts active campaigns from Current exports filtered by Asignacion Junio brands.
    """
    # Total desde Asignacion Junio (fuente de verdad)
    try:
        _aj_df = load_asignacion_activa()
        aj_ids = set(_aj_df["brand_id"].tolist()) if not _aj_df.empty else set()
        total  = len(_aj_df) if not _aj_df.empty else 0
    except Exception:
        aj_ids = set()
        total  = 0

    portfolio_ids = aj_ids if aj_ids else get_portfolio_ids()

    # ADS activas: Current ADS, BOOKINGS NET > 0, filtradas por Asignacion Junio
    ads_ids = set()
    try:
        _raw_ads = _read_current_sheet(CURRENT_ADS_SHEET)
        if not _raw_ads.empty:
            _raw_ads.columns = [normalize(c) for c in _raw_ads.columns]
            _bid_col  = _first_existing_col(_raw_ads, ["code", "brand id", "brand_id", "id"])
            _book_col = _first_existing_col(_raw_ads, ["bookings net", "bookings_net"])
            if _bid_col and _book_col:
                _raw_ads["_book_num"] = pd.to_numeric(_raw_ads[_book_col], errors="coerce").fillna(0)
                _raw_ads["_bid_norm"] = _raw_ads[_bid_col].apply(normalize_brand_id)
                for _, r in _raw_ads[_raw_ads["_book_num"] > 0].iterrows():
                    bid = r["_bid_norm"]
                    if bid and (bid in portfolio_ids):
                        ads_ids.add(bid)
    except Exception:
        ads_df = load_current_ads_data(portfolio_only=True)
        if not ads_df.empty:
            for _, r in ads_df.iterrows():
                if to_number(r.get("bookings net"), 0) > 0:
                    bid = normalize_brand_id(r.get("_id"))
                    if bid and (bid in portfolio_ids):
                        ads_ids.add(bid)

    # MD activas: Current MD, BRANDS MD # > 0, filtradas por Asignacion Junio
    md_ids = set()
    try:
        _raw_md = _read_current_sheet(CURRENT_MD_SHEET)
        if not _raw_md.empty:
            _raw_md.columns = [normalize(c) for c in _raw_md.columns]
            _md_bid_col  = _first_existing_col(_raw_md, ["brand id", "brand_id", "id"])
            _md_flag_col = _first_existing_col(_raw_md, ["brands md #", "brands md#", "brands md"])
            if _md_bid_col and _md_flag_col:
                _raw_md["_flag_num"] = pd.to_numeric(_raw_md[_md_flag_col], errors="coerce").fillna(0)
                _raw_md["_bid_norm"] = _raw_md[_md_bid_col].apply(normalize_brand_id)
                for _, r in _raw_md[_raw_md["_flag_num"] > 0].iterrows():
                    bid = r["_bid_norm"]
                    if bid and (bid in portfolio_ids):
                        md_ids.add(bid)
    except Exception:
        pass

    # MD Pro activas: Current MD pro, BRANDS MD # PRO > 0, filtradas por Asignacion Junio
    md_pro_ids = set()
    try:
        _raw_mdp = _read_current_sheet(CURRENT_MD_PRO_SHEET)
        if not _raw_mdp.empty:
            _raw_mdp.columns = [normalize(c) for c in _raw_mdp.columns]
            _mdp_bid_col  = _first_existing_col(_raw_mdp, ["brand id", "brand_id", "id"])
            _mdp_flag_col = _first_existing_col(_raw_mdp, ["brands md # pro", "brands md# pro", "brands md #pro", "brands md pro"])
            if _mdp_bid_col and _mdp_flag_col:
                _raw_mdp["_flag_num"] = pd.to_numeric(_raw_mdp[_mdp_flag_col], errors="coerce").fillna(0)
                _raw_mdp["_bid_norm"] = _raw_mdp[_mdp_bid_col].apply(normalize_brand_id)
                for _, r in _raw_mdp[_raw_mdp["_flag_num"] > 0].iterrows():
                    bid = r["_bid_norm"]
                    if bid and (bid in portfolio_ids):
                        md_pro_ids.add(bid)
    except Exception:
        pass

    return {
        "total":       total,
        "ads":         len(ads_ids),
        "md":          len(md_ids),
        "md_pro":      len(md_pro_ids),
        "pct_ads":     (len(ads_ids)    / total) if total else 0,
        "pct_md":      (len(md_ids)     / total) if total else 0,
        "pct_md_pro":  (len(md_pro_ids) / total) if total else 0,
    }



def fmt_roi2(value):
    try:
        return f"{float(value):.2f}"
    except Exception:
        return "-"

# =========================
# MANAGEMENT DASHBOARD
# =========================

@st.cache_data(ttl=300, show_spinner=False)
def _read_growth_summary_values():
    """Reads the Growth OS General KPIs block. This is the last-month/baseline reference.
    Cacheado: esta página es la de aterrizaje del dashboard, se visita en cada
    sesión — sin caché abría el workbook completo en cada rerun (cada click,
    filtro o navegación de vuelta a esta página)."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    ws = wb[GROWTH_SHEET]
    values = {
        "gmv_ars": to_number(ws["AU4"].value, 0),
        "gmv_usd": to_number(ws["AU6"].value, 0),
        "gmv_cop": to_number(ws["AU8"].value, 0),
        "aov_ars": to_number(ws["AV4"].value, 0),
        "aov_usd": to_number(ws["AV6"].value, 0),
        "aov_cop": to_number(ws["AV8"].value, 0),
        "brands_ads": to_number(ws["AW4"].value, 0),
        "pct_brands_ads": to_number(ws["AX4"].value, 0),
        "brands_md": to_number(ws["AW6"].value, 0),
        "pct_brands_md": to_number(ws["AX6"].value, 0),
        "total_pro": to_number(ws["AY6"].value, 0),
        "total_cr": to_number(ws["AX8"].value, 0),
        "gross_bookings_ars": to_number(ws["AU10"].value, 0),
        "gross_bookings_usd": to_number(ws["AV10"].value, 0),
        "gross_bookings_cop": to_number(ws["AW10"].value, 0),
        "effective_contacts": to_number(ws["AX10"].value, 0),
        "total_comm": to_number(ws["AY10"].value, 0),
    }
    wb.close()
    return values


@st.cache_data(ttl=300, show_spinner=False)
def _compute_growth_summary_fallback():
    raw = pd.read_excel(EXCEL_FILE, sheet_name=GROWTH_SHEET, header=None)
    portfolio = raw.iloc[3:253].copy()

    # Total de marcas desde Asignacion Junio (fuente de verdad del portafolio)
    try:
        _aj_df = load_asignacion_activa()
        total_brands = len(_aj_df) if not _aj_df.empty else 250
    except Exception:
        total_brands = 250

    def col_sum(idx):
        if idx >= portfolio.shape[1]:
            return 0
        return pd.to_numeric(portfolio.iloc[:, idx], errors="coerce").fillna(0).sum()

    def count_exact(idx, expected):
        if idx >= portfolio.shape[1]:
            return 0
        s = portfolio.iloc[:, idx].astype(str).str.strip().str.lower()
        return int((s == str(expected).strip().lower()).sum())

    gmv_ars = col_sum(20)
    gmv_usd = gmv_ars / ARS_PER_USD
    gmv_cop = gmv_usd * COP_PER_USD
    aov_ars = col_sum(23) / total_brands
    aov_usd = aov_ars / ARS_PER_USD
    aov_cop = aov_usd * COP_PER_USD
    brands_ads = count_exact(29, "Active 🚀")
    brands_md = count_exact(34, "Active 🚀")
    gross_bookings_ars = col_sum(30) * 4
    gross_bookings_usd = gross_bookings_ars / ARS_PER_USD
    gross_bookings_cop = gross_bookings_usd * COP_PER_USD
    return {
        "gmv_ars": gmv_ars,
        "gmv_usd": gmv_usd,
        "gmv_cop": gmv_cop,
        "aov_ars": aov_ars,
        "aov_usd": aov_usd,
        "aov_cop": aov_cop,
        "brands_ads": brands_ads,
        "pct_brands_ads": brands_ads / total_brands,
        "brands_md": brands_md,
        "pct_brands_md": brands_md / total_brands,
        "total_pro": col_sum(27) / total_brands,
        "total_cr": col_sum(28) / total_brands,
        "gross_bookings_ars": gross_bookings_ars,
        "gross_bookings_usd": gross_bookings_usd,
        "gross_bookings_cop": gross_bookings_cop,
        "effective_contacts": 0,
        "total_comm": col_sum(17) / total_brands,
    }


def page_management_dashboard():
    render_header("Management Dashboard", "General Overview of Commercial Performance · Rappi")

    if not os.path.exists(EXCEL_FILE):
        st.error(f"No encontré el archivo de datos '{EXCEL_FILE}'. Poné tu workbook (.xlsx con la hoja Growth OS) en la misma carpeta que app_glass.py.")
        return

    # Wrappers locales — preservan los nombres usados más abajo en esta función
    # sin duplicar la lógica; la lectura real ya está cacheada a nivel de módulo.
    read_summary_values = _read_growth_summary_values
    compute_summary_fallback = _compute_growth_summary_fallback

    try:
        baseline_vals = read_summary_values()
        if baseline_vals["gmv_ars"] == 0 or baseline_vals["aov_ars"] == 0 or baseline_vals["gross_bookings_ars"] == 0:
            fallback_vals = compute_summary_fallback()
            for key, value in fallback_vals.items():
                if baseline_vals.get(key, 0) == 0:
                    baseline_vals[key] = value
    except Exception as e:
        try:
            baseline_vals = compute_summary_fallback()
        except Exception as e2:
            st.error(f"Could not read Management Dashboard values from Excel: {e2}")
            st.caption(f"First attempt error: {e}")
            return

    # GMV y AOV desde Detalle CABA filtrado por Asignacion Junio (fuente de verdad)
    detalle_vals = get_portfolio_gmv_aov_from_detalle_caba()
    current_vals = detalle_vals or get_current_gmv_totals() or {}
    vals = baseline_vals.copy()
    if current_vals:
        for key in ["gmv_ars", "gmv_usd", "gmv_cop", "aov_ars", "aov_usd", "aov_cop"]:
            vals[key] = current_vals.get(key, baseline_vals.get(key, 0))

    current_ads_totals = get_current_ads_totals()
    ads_bookings = money_from_usd(current_ads_totals["bookings_usd"])
    ads_revenue = money_from_usd(current_ads_totals["revenue_usd"])
    revenue_goal = money_from_usd(ADS_REVENUE_TARGET_USD)
    current_md_totals = get_current_md_totals(pro=False)
    current_md_pro_totals = get_current_md_totals(pro=True)
    md_sales = money_from_usd(current_md_totals["sales_usd"])
    md_pro_sales = money_from_usd(current_md_pro_totals["sales_usd"])
    _md_ars       = fmt_ars(md_sales["ars"])
    _md_usd       = fmt_usd(md_sales["usd"])
    _md_cop       = fmt_cop(md_sales["cop"])
    _mdp_ars      = fmt_ars(md_pro_sales["ars"])
    _mdp_usd      = fmt_usd(md_pro_sales["usd"])
    _mdp_cop      = fmt_cop(md_pro_sales["cop"])
    _md_camp      = int(current_md_totals.get("campaigns", 0))
    _mdp_camp     = int(current_md_pro_totals.get("campaigns", 0))
    contact_stats = get_comment_contact_stats(fallback_total=baseline_vals.get("effective_contacts", 0))

    def panel_title(title):
        st.markdown(f"""
<div class="mgmt-section-title-card">
    <div class="mgmt-section-title-copy">{title}</div>
</div>
""", unsafe_allow_html=True)

    def stack_money(label, ars, usd, cop, foot=""):
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:26px 28px;">
    <div class="stack-label">{label}</div>
    <div style="margin-top:10px;">
        <div style="font-size:36px;font-weight:900;color:#1A1A2E;letter-spacing:-0.02em;line-height:1.1;">{fmt_usd(usd)}</div>
        <div style="font-size:13px;font-weight:500;color:#6B7280;margin-top:6px;">{fmt_ars(ars)} &middot; {fmt_cop(cop)}</div>
    </div>
    <div class="stack-foot" style="margin-top:12px;">{foot if foot else '&nbsp;'}</div>
</div>
""", unsafe_allow_html=True)

    def simple_card(label, value, foot="", tone="blue"):
        # Management Dashboard cards stay visually uniform; KPI meaning comes from the label, not card color.
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:26px 28px;">
    <div class="stack-label">{label}</div>
    <div class="stack-main">{value}</div>
    <div class="stack-foot">{foot if foot else '&nbsp;'}</div>
</div>
""", unsafe_allow_html=True)

    def progress_card(label, current, goal, value_formatter, kind="progress"):
        ratio = safe_ratio(current, goal)
        remaining = max(goal - current, 0)
        ratio_text = fmt_percent0(ratio) if ratio is not None else "-"
        main = ratio_text if kind == "progress" else fmt_signed_percent(ratio)
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:26px 28px;">
    <div class="stack-label">{label}</div>
    <div class="stack-main">{main}</div>
    <div class="stack-foot">Reference: {value_formatter(goal)}</div>
    <div class="stack-foot">Remaining: {value_formatter(remaining)}</div>
</div>
""", unsafe_allow_html=True)

    # ── ROW 1: GMV + AOV ──────────────────────────────────────────────────────
    gmv_pct   = max(0, min(1, safe_ratio(vals["gmv_ars"], baseline_vals.get("gmv_ars", 1)) or 0))
    gmv_bar   = round(gmv_pct * 100)
    gmv_color = "#7ED321" if gmv_pct >= 1 else ("#FF7124" if gmv_pct >= 0.7 else "#FF4D2E")

    aov_change = safe_ratio(vals["aov_ars"] - baseline_vals.get("aov_ars", 0), baseline_vals.get("aov_ars", 1)) or 0
    aov_sign   = "+" if aov_change >= 0 else ""
    aov_arrow  = "▲" if aov_change >= 0 else "▼"
    aov_color  = "#7ED321" if aov_change >= 0 else "#FF4D2E"

    gmv_surplus     = vals["gmv_ars"] - baseline_vals.get("gmv_ars", 0)
    gmv_surplus_usd = vals["gmv_usd"] - baseline_vals.get("gmv_usd", 0)
    gmv_surplus_cop = vals["gmv_cop"] - baseline_vals.get("gmv_cop", 0)
    surplus_sign    = "+" if gmv_surplus >= 0 else ""
    surplus_color   = "#7ED321" if gmv_surplus >= 0 else "#FF4D2E"
    surplus_label   = "▲ Excedente sobre mes anterior" if gmv_surplus >= 0 else "▼ Por debajo del mes anterior"

    _v_gmv_ars   = fmt_ars(vals["gmv_ars"])
    _v_gmv_usd   = fmt_usd(vals["gmv_usd"])
    _v_gmv_cop   = fmt_cop(vals["gmv_cop"])
    _v_aov_ars   = fmt_ars(vals["aov_ars"])
    _v_aov_usd   = fmt_usd(vals["aov_usd"])
    _v_aov_cop   = fmt_cop(vals["aov_cop"])
    _v_ref_gmv   = fmt_ars(baseline_vals.get("gmv_ars", 0))
    _v_ref_aov   = fmt_ars(baseline_vals.get("aov_ars", 0))
    _v_sur_gmv   = fmt_ars(gmv_surplus)
    _v_sur_usd   = fmt_usd(gmv_surplus_usd)
    _v_sur_cop   = fmt_cop(gmv_surplus_cop)

    gmv_change    = safe_ratio(vals["gmv_ars"] - baseline_vals.get("gmv_ars", 0), baseline_vals.get("gmv_ars", 1)) or 0
    gmv_sign_ch   = "+" if gmv_change >= 0 else ""
    gmv_arrow_ch  = "&#9650;" if gmv_change >= 0 else "&#9660;"
    gmv_color_ch  = "#7ED321" if gmv_change >= 0 else "#FF4D2E"

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:26px 28px;">
  <div class="stack-label">CURRENT GMV &middot; MTD</div>
  <div style="margin-top:10px;">
    <div style="font-size:36px;font-weight:900;color:#1A1A2E;letter-spacing:-0.02em;line-height:1.1;">{_v_gmv_usd}</div>
    <div style="font-size:13px;font-weight:500;color:#6B7280;margin-top:6px;">{_v_gmv_ars} &middot; {_v_gmv_cop}</div>
  </div>
  <div style="margin-top:16px;display:flex;align-items:center;gap:14px;">
    <div style="font-size:38px;font-weight:900;color:{gmv_color_ch};line-height:1;">{gmv_arrow_ch}</div>
    <div>
      <div style="font-size:28px;font-weight:900;color:{gmv_color_ch};">{gmv_sign_ch}{fmt_percent0(gmv_change)}</div>
      <div style="font-size:12px;color:#6B7280;font-weight:700;">vs last month &middot; {fmt_usd(baseline_vals.get("gmv_usd", 0))}</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:26px 28px;">
  <div class="stack-label">CURRENT AOV</div>
  <div style="margin-top:10px;">
    <div style="font-size:36px;font-weight:900;color:#1A1A2E;letter-spacing:-0.02em;line-height:1.1;">{_v_aov_usd}</div>
    <div style="font-size:13px;font-weight:500;color:#6B7280;margin-top:6px;">{_v_aov_ars} &middot; {_v_aov_cop}</div>
  </div>
  <div style="margin-top:16px;display:flex;align-items:center;gap:14px;">
    <div style="font-size:38px;font-weight:900;color:{aov_color};line-height:1;">{'&#9650;' if aov_change >= 0 else '&#9660;'}</div>
    <div>
      <div style="font-size:28px;font-weight:900;color:{aov_color};">{aov_sign}{fmt_percent0(aov_change)}</div>
      <div style="font-size:12px;color:#6B7280;font-weight:700;">vs last month &middot; {fmt_usd(baseline_vals.get("aov_usd", 0))}</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

   # ── ROW 2: COVERAGE + PRO/CR ───────────────────────────────────────────────
    live_coverage  = get_live_campaign_coverage_counts()
    total_brands   = max(live_coverage["total"], 1)
    ads_pct_bar    = round(live_coverage["pct_ads"]    * 100)
    md_pct_bar     = round(live_coverage["pct_md"]     * 100)
    md_pro_pct_bar = round(live_coverage["pct_md_pro"] * 100)
    pro_pct_bar    = round(min(to_number(baseline_vals.get("total_pro", 0), 0) * 100, 100))

    # CR desde CVR% sheet — promedio del portafolio (Asignacion Junio)
    try:
        _cvr_map = load_cvr_data()
        _aj_cvr  = load_asignacion_activa()
        _cvr_vals = []
        for _, _aj_row in _aj_cvr.iterrows():
            _bname = normalize(str(_aj_row.get("brand_name", "")))
            if _bname in _cvr_map and _cvr_map[_bname] > 0:
                _cvr_vals.append(_cvr_map[_bname])
        cr_pct_bar = round(min((sum(_cvr_vals) / len(_cvr_vals)) * 100, 100)) if _cvr_vals else round(min(to_number(baseline_vals.get("total_cr", 0), 0) * 100, 100))
    except Exception:
        cr_pct_bar = round(min(to_number(baseline_vals.get("total_cr", 0), 0) * 100, 100))

    def _svg_donut(pct, color, size=110, stroke=14):
        r = (size - stroke) / 2
        circ = 2 * 3.14159 * r
        filled = round(circ * pct / 100, 1)
        gap = round(circ - filled, 1)
        cx = cy = size / 2
        return (
            f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}" style="display:block;">'
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="rgba(255,255,255,0.1)" stroke-width="{stroke}"/>'
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" stroke-width="{stroke}" '
            f'stroke-dasharray="{filled} {gap}" stroke-dashoffset="{circ * 0.25}" stroke-linecap="round"/>'
            f'</svg>'
        )

    ads_donut_color    = "#FF7124" if ads_pct_bar    < 60 else ("#7ED321" if ads_pct_bar    >= 80 else "#FF7124")
    md_donut_color     = "#1B3F8B" if md_pct_bar     < 60 else ("#7ED321" if md_pct_bar     >= 80 else "#1B3F8B")
    md_pro_donut_color = "#C084FC" if md_pro_pct_bar < 60 else ("#7ED321" if md_pro_pct_bar >= 80 else "#C084FC")

    coverage_html = (
        '<div class="stack-card mgmt-stack-card" style="padding:26px 28px;">'
        '<div class="stack-label">BRAND COVERAGE &middot; LIVE</div>'
        '<div style="margin-top:18px;display:flex;justify-content:space-around;align-items:center;gap:10px;">'
        # ADS donut
        '<div style="display:flex;flex-direction:column;align-items:center;gap:8px;">'
        f'<div style="position:relative;width:90px;height:90px;">'
        f'{_svg_donut(ads_pct_bar, "#FF7124", size=90, stroke=12)}'
        f'<div style="position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);text-align:center;">'
        f'<div style="font-size:17px;font-weight:900;color:#FF7124;line-height:1;">{ads_pct_bar}%</div>'
        f'</div></div>'
        f'<div style="text-align:center;">'
        f'<div style="font-size:11px;font-weight:900;color:#FF7124;">ADS</div>'
        f'<div style="font-size:10px;color:#6B7280;font-weight:700;">{live_coverage["ads"]} marcas</div>'
        f'</div></div>'
        # MD donut
        '<div style="display:flex;flex-direction:column;align-items:center;gap:8px;">'
        f'<div style="position:relative;width:90px;height:90px;">'
        f'{_svg_donut(md_pct_bar, "#1B3F8B", size=90, stroke=12)}'
        f'<div style="position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);text-align:center;">'
        f'<div style="font-size:17px;font-weight:900;color:#1B3F8B;line-height:1;">{md_pct_bar}%</div>'
        f'</div></div>'
        f'<div style="text-align:center;">'
        f'<div style="font-size:11px;font-weight:900;color:#1B3F8B;">MD</div>'
        f'<div style="font-size:10px;color:#6B7280;font-weight:700;">{live_coverage["md"]} marcas</div>'
        f'</div></div>'
        # MD PRO donut (nuevo)
        '<div style="display:flex;flex-direction:column;align-items:center;gap:8px;">'
        f'<div style="position:relative;width:90px;height:90px;">'
        f'{_svg_donut(md_pro_pct_bar, "#C084FC", size=90, stroke=12)}'
        f'<div style="position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);text-align:center;">'
        f'<div style="font-size:17px;font-weight:900;color:#C084FC;line-height:1;">{md_pro_pct_bar}%</div>'
        f'</div></div>'
        f'<div style="text-align:center;">'
        f'<div style="font-size:11px;font-weight:900;color:#C084FC;">MD PRO</div>'
        f'<div style="font-size:10px;color:#6B7280;font-weight:700;">{live_coverage["md_pro"]} marcas</div>'
        f'</div></div>'
        '</div>'
        f'<div style="font-size:11px;color:#6B7280;margin-top:14px;text-align:center;">Portfolio: {total_brands} marcas &middot; live tracker</div>'
        '</div>'
    )

    def _waffle_icons(filled_count, total=10, filled_color="#7ED321", empty_color="rgba(255,255,255,0.15)"):
        person_path = '<circle cx="12" cy="7" r="4"/><path d="M4 21c0-4.418 3.582-8 8-8s8 3.582 8 8"/>'
        icons = []
        for i in range(total):
            color = filled_color if i < filled_count else empty_color
            icons.append(
                f'<svg width="22" height="22" viewBox="0 0 24 24" fill="{color}" '
                f'xmlns="http://www.w3.org/2000/svg" style="flex-shrink:0;">'
                f'{person_path}</svg>'
            )
        return '<div style="display:flex;gap:3px;flex-wrap:wrap;margin-top:6px;">' + "".join(icons) + '</div>'

    pro_icons_count = max(0, min(10, round(pro_pct_bar / 10)))
    cr_icons_count  = max(0, min(10, round(cr_pct_bar  / 10)))

    portfolio_html = (
        '<div class="stack-card mgmt-stack-card" style="padding:26px 28px;">'
        '<div class="stack-label">PORTFOLIO PERFORMANCE</div>'
        '<div style="margin-top:16px;display:flex;flex-direction:column;gap:18px;">'
        '<div>'
        '<div style="display:flex;justify-content:space-between;font-size:12px;font-weight:800;margin-bottom:4px;">'
        '<span style="color:#7ED321;">🟢 PRO % &nbsp;<span style="font-weight:600;color:#6B7280;">usuarios con plan Pro</span></span>'
        f'<span style="color:#7ED321;">{pro_pct_bar}%</span>'
        '</div>'
        f'<div style="font-size:11px;color:#6B7280;margin-bottom:2px;">{pro_icons_count} de 10 = {pro_pct_bar}% PRO</div>'
        + _waffle_icons(pro_icons_count, filled_color="#7ED321") +
        '</div>'
        '<div>'
        '<div style="display:flex;justify-content:space-between;font-size:12px;font-weight:800;margin-bottom:4px;">'
        '<span style="color:#6B7280;">⚪ CR % &nbsp;<span style="font-weight:600;">compran vs no compran</span></span>'
        f'<span style="color:#6B7280;">{cr_pct_bar}%</span>'
        '</div>'
        f'<div style="font-size:11px;color:#6B7280;margin-bottom:2px;">{cr_icons_count} de 10 = {cr_pct_bar}% convierten</div>'
        + _waffle_icons(cr_icons_count, filled_color="#1B3F8B", empty_color="rgba(255,255,255,0.15)") +
        '</div>'
        '</div>'
        '<div style="font-size:11px;color:#6B7280;margin-top:12px;">Baseline reference · last month snapshot · 1 muñequito = 10%</div>'
        '</div>'
    )

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(coverage_html, unsafe_allow_html=True)
    with c2:
        st.markdown(portfolio_html, unsafe_allow_html=True)

    # ── ROW 3: ADS BOOKINGS + ADS REVENUE (DONUT) ─────────────────────────────
    rev_pct      = max(0, safe_ratio(ads_revenue["usd"], ADS_REVENUE_TARGET_USD) or 0)
    rev_done_pct = round(min(rev_pct, 1) * 100)
    rev_over_pct = round((rev_pct - 1) * 100) if rev_pct > 1 else 0
    donut_color  = "#7ED321" if rev_pct >= 1 else ("#FF7124" if rev_pct >= 0.5 else "#1B3F8B")

    book_pct   = max(0, safe_ratio(ads_bookings["ars"], baseline_vals.get("gross_bookings_ars", 1)) or 0)
    book_bar   = round(min(book_pct, 1) * 100)
    book_over  = round((book_pct - 1) * 100) if book_pct > 1 else 0
    book_color = "#7ED321" if book_pct >= 1 else ("#FF7124" if book_pct >= 0.7 else "#FF4D2E")

    # ── Barra delgada lateral derecha (thin right-side bar) ──────────────────
    def _thin_bar_right_svg(current_val, goal_val, label_goal,
                             base_color="#1B3F8B", over_color="#7ED321", under_color="#FF4D2E",
                             width=52, height=110):
        """
        Barra vertical delgada, alineada a la derecha del sticker.
        - Segmento azul   : ejecución real (hasta la meta)
        - Segmento verde  : sobre ejecución (por encima de la meta)
        - Segmento rojo   : no cumplimiento (encima del azul, hasta la meta)
        - Línea de meta   : punteada
        """
        pad_t, pad_b = 14, 14
        chart_h = height - pad_t - pad_b
        bar_w   = 14
        bar_x   = (width - bar_w) / 2

        max_val = max(current_val, goal_val, 1) * 1.18

        def to_y(val):
            ratio = min(max(val / max_val, 0), 1)
            return pad_t + chart_h * (1 - ratio)

        bottom_y = pad_t + chart_h
        goal_y   = to_y(goal_val)
        exec_top = to_y(min(current_val, goal_val))
        exec_h   = bottom_y - exec_top

        # Barra base azul (ejecución real hasta meta)
        rects = (
            f'<rect x="{bar_x:.1f}" y="{exec_top:.1f}" width="{bar_w}" height="{exec_h:.1f}" '
            f'rx="3" fill="{base_color}" opacity="0.9"/>'
        )

        if current_val > goal_val:
            over_top = to_y(current_val)
            over_h   = exec_top - over_top
            rects += (
                f'<rect x="{bar_x:.1f}" y="{over_top:.1f}" width="{bar_w}" height="{over_h:.1f}" '
                f'rx="3" fill="{over_color}"/>'
            )
        elif current_val < goal_val:
            gap_h = exec_top - goal_y
            if gap_h > 0:
                rects += (
                    f'<rect x="{bar_x:.1f}" y="{goal_y:.1f}" width="{bar_w}" height="{gap_h:.1f}" '
                    f'rx="3" fill="{under_color}" opacity="0.85"/>'
                )

        # Línea de meta punteada
        meta_line = (
            f'<line x1="{bar_x - 4:.1f}" y1="{goal_y:.1f}" x2="{bar_x + bar_w + 4:.1f}" y2="{goal_y:.1f}" '
            f'stroke="rgba(255,255,255,0.3)" stroke-width="1.5" stroke-dasharray="3 2"/>'
        )

        # Porcentaje de ejecución encima de la barra
        pct_val = round(min(current_val / max(goal_val, 1), 9.99) * 100)
        pct_color = over_color if current_val >= goal_val else under_color
        pct_label = (
            f'<text x="{width/2:.1f}" y="{pad_t - 3:.1f}" text-anchor="middle" '
            f'font-size="8" font-weight="900" fill="{pct_color}">{pct_val}%</text>'
        )
        # "Meta" debajo
        meta_label = (
            f'<text x="{width/2:.1f}" y="{height - 2:.1f}" text-anchor="middle" '
            f'font-size="7" fill="rgba(107,114,128,0.60)" font-weight="700">Meta</text>'
        )

        return (
            f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" '
            f'xmlns="http://www.w3.org/2000/svg">'
            f'{rects}{meta_line}{pct_label}{meta_label}'
            f'</svg>'
        )

    # ── Leyenda compacta ──────────────────────────────────────────────────────
    legend_html = (
        '<div style="display:flex;gap:10px;align-items:center;margin-bottom:8px;flex-wrap:wrap;">'
        '<span style="display:flex;align-items:center;gap:4px;font-size:10px;font-weight:700;color:#6B7280;">'
        '<span style="display:inline-block;width:10px;height:10px;background:#1B3F8B;border-radius:2px;"></span>'
        'Ejecución</span>'
        '<span style="display:flex;align-items:center;gap:4px;font-size:10px;font-weight:700;color:#6B7280;">'
        '<span style="display:inline-block;width:10px;height:10px;background:#7ED321;border-radius:2px;"></span>'
        'Sobre meta</span>'
        '<span style="display:flex;align-items:center;gap:4px;font-size:10px;font-weight:700;color:#6B7280;">'
        '<span style="display:inline-block;width:10px;height:10px;background:#FF4D2E;border-radius:2px;"></span>'
        'No cumplimiento</span>'
        '</div>'
    )

    _b_usd  = fmt_usd(ads_bookings["usd"])
    _b_ars  = fmt_ars(ads_bookings["ars"])
    _b_cop  = fmt_cop(ads_bookings["cop"])
    _b_goal_ars = baseline_vals.get("gross_bookings_ars", 0)
    _b_goal_usd = baseline_vals.get("gross_bookings_usd", _b_goal_ars / ARS_PER_USD if _b_goal_ars else 0)
    _b_goal = fmt_usd(_b_goal_usd)

    _r_usd  = ads_revenue["usd"]
    _r_fmt  = fmt_usd(_r_usd)
    _r_rem  = fmt_usd(max(ADS_REVENUE_TARGET_USD - _r_usd, 0))
    _r_goal = fmt_usd(ADS_REVENUE_TARGET_USD)

    _book_num = ads_bookings["usd"]
    _rev_num  = ads_revenue["usd"]

    book_bar_svg = _thin_bar_right_svg(
        current_val=_book_num,
        goal_val=_b_goal_usd,
        label_goal=_b_goal,
    )
    rev_bar_svg = _thin_bar_right_svg(
        current_val=_rev_num,
        goal_val=ADS_REVENUE_TARGET_USD,
        label_goal=_r_goal,
    )

    c1, c2 = st.columns(2)
    with c1:
        _over_book  = _book_num - _b_goal_usd
        _pct_book   = round(min(_book_num / max(_b_goal_usd, 1), 9.99) * 100)
        _sc_book    = "#7ED321" if _book_num >= _b_goal_usd else "#FF4D2E"
        _sl_book    = f"▲ +{fmt_usd(_over_book)} sobre meta" if _book_num >= _b_goal_usd else f"▼ Faltan {fmt_usd(abs(_over_book))}"
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:22px 24px;position:relative;overflow:hidden;">
  <div class="stack-label">ADS GROSS BOOKINGS &middot; MTD</div>
  {legend_html}
  <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-top:4px;">
    <div>
      <div class="stack-main mgmt-ars" style="margin-top:0;">{_b_usd}</div>
      <div class="stack-sub mgmt-conv">{_b_ars} &middot; {_b_cop}</div>
      <div style="font-size:12px;font-weight:800;color:{_sc_book};margin-top:8px;">{_sl_book}</div>
      <div style="font-size:11px;color:#6B7280;margin-top:2px;">Meta: {_b_goal} &middot; {_pct_book}% ejec.</div>
    </div>
    <div style="flex-shrink:0;">{book_bar_svg}</div>
  </div>
</div>
""", unsafe_allow_html=True)

    with c2:
        _over_rev  = _rev_num - ADS_REVENUE_TARGET_USD
        _pct_rev   = round(min(_rev_num / max(ADS_REVENUE_TARGET_USD, 1), 9.99) * 100)
        _sc_rev    = "#7ED321" if _rev_num >= ADS_REVENUE_TARGET_USD else "#FF4D2E"
        _sl_rev    = f"▲ +{fmt_usd(_over_rev)} sobre meta" if _rev_num >= ADS_REVENUE_TARGET_USD else f"▼ Remaining: {_r_rem}"
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:22px 24px;position:relative;overflow:hidden;">
  <div class="stack-label">ADS REVENUE &middot; MTD vs GOAL</div>
  {legend_html}
  <div style="display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-top:4px;">
    <div>
      <div class="stack-main mgmt-ars" style="margin-top:0;">{_r_fmt}</div>
      <div style="font-size:12px;color:#6B7280;font-weight:700;margin-top:4px;">Revenue generado</div>
      <div style="font-size:12px;font-weight:800;color:{_sc_rev};margin-top:8px;">{_sl_rev}</div>
      <div style="font-size:11px;color:#6B7280;margin-top:2px;">Meta: {_r_goal} &middot; {_pct_rev}% ejec.</div>
    </div>
    <div style="flex-shrink:0;">{rev_bar_svg}</div>
  </div>
</div>
""", unsafe_allow_html=True)

    # ── ROW 4: MD + MD PRO ────────────────────────────────────────────────────
    md_roi_val    = to_number(current_md_totals.get("roi"), 0)
    mdpro_roi_val = to_number(current_md_pro_totals.get("roi"), 0)
    benchmark     = 3.2
    import math as _math_md

    def _needle_gauge_svg(roi_val, bmark=3.2, max_val=8.0, width=110, height=68):
        """Bullet chart SVG for ROI — horizontal segmented bar with benchmark tick."""
        W, H = width, height
        # Three zones: red (0→60% of bmark), orange (60%→100% of bmark), green (bmark→max)
        red_end    = bmark * 0.6 / max_val
        orange_end = bmark       / max_val
        bar_y      = H - 30
        bar_h      = 10
        pad_l, pad_r = 4, 4
        bar_w      = W - pad_l - pad_r
        rx         = 4  # corner radius

        def seg_x(ratio):
            return pad_l + ratio * bar_w

        red_w    = seg_x(red_end)   - pad_l
        orange_w = seg_x(orange_end) - seg_x(red_end)
        green_w  = bar_w - red_w - orange_w

        # Clamp filled bar at max_val, never negative
        fill_ratio = min(max(roi_val / max_val, 0), 1)
        fill_w     = fill_ratio * bar_w

        nc = "#7ED321" if roi_val >= bmark else ("#FF7124" if roi_val >= bmark * 0.6 else "#FF4D2E")

        # Benchmark tick position
        bm_x = seg_x(bmark / max_val)

        # Labels row above bar
        label_y = bar_y - 5

        svg_parts = [
            f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" style="overflow:visible;">',
            # Background zones
            f'<rect x="{pad_l:.1f}" y="{bar_y}" width="{bar_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(255,255,255,0.1)"/>',
            f'<rect x="{pad_l:.1f}" y="{bar_y}" width="{red_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(255,92,122,0.25)"/>',
            f'<rect x="{seg_x(red_end):.1f}" y="{bar_y}" width="{orange_w:.1f}" height="{bar_h}" fill="rgba(255,138,61,0.20)"/>',
            f'<rect x="{seg_x(orange_end):.1f}" y="{bar_y}" width="{green_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(191,255,0,0.22)"/>',
            # Filled performance bar (thinner, centered vertically)
            f'<rect x="{pad_l:.1f}" y="{bar_y + 2}" width="{fill_w:.1f}" height="{bar_h - 4}" rx="{rx - 1}" fill="{nc}" opacity="0.92"/>',
            # Benchmark tick
            f'<rect x="{bm_x - 1:.1f}" y="{bar_y - 3}" width="2" height="{bar_h + 6}" rx="1" fill="rgba(255,255,255,0.3)" opacity="0.45"/>',
            # Value text (left-aligned, above bar)
            f'<text x="{pad_l}" y="{label_y}" font-size="11" font-weight="900" fill="{nc}">{roi_val:.1f}x</text>',
            # Benchmark label (right side, above bar)
            f'<text x="{W - pad_r}" y="{label_y}" text-anchor="end" font-size="8" fill="rgba(107,114,128,0.60)" font-weight="700">bm {bmark}x</text>',
            # Zone labels below bar
            f'<text x="{pad_l}" y="{bar_y + bar_h + 10}" font-size="6.5" fill="#FF4D2E" font-weight="700">low</text>',
            f'<text x="{seg_x(orange_end):.1f}" y="{bar_y + bar_h + 10}" font-size="6.5" fill="#7ED321" font-weight="700">target</text>',
            '</svg>',
        ]
        return "".join(svg_parts)

    md_gauge    = _needle_gauge_svg(md_roi_val,    benchmark, max_val=max(benchmark * 2.5, md_roi_val * 1.2 + 0.1))
    mdpro_gauge = _needle_gauge_svg(mdpro_roi_val, benchmark, max_val=max(benchmark * 2.5, mdpro_roi_val * 1.2 + 0.1))

    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:26px 28px;position:relative;overflow:hidden;">
  <div class="stack-label">MD SALES &middot; MTD</div>
  <div class="stack-main mgmt-ars" style="margin-top:8px;">{_md_usd}</div>
  <div class="stack-sub mgmt-conv">{_md_ars} &middot; {_md_cop}</div>
  <div style="margin-top:6px;font-size:12px;font-weight:800;color:#1B3F8B;">📊 {_md_camp} campaigns</div>
  <div style="position:absolute;bottom:10px;right:14px;opacity:0.90;">
    {md_gauge}
  </div>
</div>
""", unsafe_allow_html=True)

    with c2:
        st.markdown(f"""
<div class="stack-card mgmt-stack-card" style="padding:26px 28px;position:relative;overflow:hidden;">
  <div class="stack-label">MD PRO SALES &middot; MTD</div>
  <div class="stack-main mgmt-ars" style="margin-top:8px;">{_mdp_usd}</div>
  <div class="stack-sub mgmt-conv">{_mdp_ars} &middot; {_mdp_cop}</div>
  <div style="margin-top:6px;font-size:12px;font-weight:800;color:#7ED321;">📊 {_mdp_camp} campaigns</div>
  <div style="position:absolute;bottom:10px;right:14px;opacity:0.90;">
    {mdpro_gauge}
  </div>
</div>
""", unsafe_allow_html=True)

    # ── ROW 5: CONTACT PERFORMANCE ────────────────────────────────────────────
    _cs_eff   = contact_stats["total_effective"]
    _cs_calls = contact_stats["calls"]
    _cs_chats = contact_stats["chats"]
    _cs_meets = contact_stats["meets"]
    _cs_ghost = contact_stats["not_contacted"]
    total_c   = max(_cs_eff, 1)

    _cs_total_fmt = fmt_number(_cs_eff)
    _cs_since     = CONTACTS_START_DATE.strftime("%d %b")

    # Recalculate percentages over the full universe (effective + no answer)
    _total_universe = max(_cs_eff + _cs_ghost, 1)
    calls_pct  = round(_cs_calls / _total_universe * 100)
    chats_pct  = round(_cs_chats / _total_universe * 100)
    meets_pct  = round(_cs_meets / _total_universe * 100)
    ghost_pct  = round(_cs_ghost / _total_universe * 100)

    # Source badge: show where data came from
    _cs_source = contact_stats.get("source", "")
    _src_badge = (
        "Productivity" if _cs_source == "productivity_sheet"
        else "Comments CSV" if _cs_source == "comments_csv"
        else ""
    )

    _contact_html = (
        '<div class="stack-card mgmt-stack-card" style="padding:26px 28px;margin-top:4px;">'
        '<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:18px;">'
        f'<div>'
        f'<div class="stack-label">CONTACT PERFORMANCE · since {_cs_since}</div>'
        f'{"<div style=\"font-size:10px;color:#aaa;margin-top:2px;\">fuente: " + _src_badge + "</div>" if _src_badge else ""}'
        f'</div>'
        f'<div style="font-size:32px;font-weight:900;color:#FF7124;">{_cs_total_fmt} <span style="font-size:14px;color:#6B7280;font-weight:700;">contactos efectivos</span></div>'
        '</div>'
        # ── Stacked horizontal bar ──────────────────────────────────────────
        '<div style="display:flex;height:28px;border-radius:999px;overflow:hidden;width:100%;margin-bottom:14px;">'
        f'<div style="width:{calls_pct}%;background:#FF7124;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:800;color:#fff;white-space:nowrap;overflow:hidden;" title="📞 Amazon Connect {calls_pct}%">'
        f'{"📞 " + str(calls_pct) + "%" if calls_pct >= 7 else ""}'
        '</div>'
        f'<div style="width:{chats_pct}%;background:#1B3F8B;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:800;color:#fff;white-space:nowrap;overflow:hidden;" title="💬 WhatsApp {chats_pct}%">'
        f'{"💬 " + str(chats_pct) + "%" if chats_pct >= 7 else ""}'
        '</div>'
        f'<div style="width:{meets_pct}%;background:#7ED321;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:800;color:#7ED321;white-space:nowrap;overflow:hidden;" title="🖥️ Meet {meets_pct}%">'
        f'{"🖥️ " + str(meets_pct) + "%" if meets_pct >= 7 else ""}'
        '</div>'
        f'<div style="width:{ghost_pct}%;background:#FF4D2E;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:800;color:#fff;white-space:nowrap;overflow:hidden;" title="👻 No Contactado {ghost_pct}%">'
        f'{"👻 " + str(ghost_pct) + "%" if ghost_pct >= 7 else ""}'
        '</div>'
        '</div>'
        # ── Legend row ──────────────────────────────────────────────────────
        '<div style="display:flex;gap:20px;flex-wrap:wrap;">'
        f'<div style="display:flex;align-items:center;gap:6px;font-size:12px;font-weight:800;">'
        f'<div style="width:12px;height:12px;border-radius:3px;background:#FF7124;flex-shrink:0;"></div>'
        f'<span style="color:#FF7124;">📞 Amazon Connect</span>'
        f'<span style="color:#6B7280;font-weight:600;">{_cs_calls} &middot; {calls_pct}%</span>'
        '</div>'
        f'<div style="display:flex;align-items:center;gap:6px;font-size:12px;font-weight:800;">'
        f'<div style="width:12px;height:12px;border-radius:3px;background:#1B3F8B;flex-shrink:0;"></div>'
        f'<span style="color:#1B3F8B;">💬 WhatsApp</span>'
        f'<span style="color:#6B7280;font-weight:600;">{_cs_chats} &middot; {chats_pct}%</span>'
        '</div>'
        f'<div style="display:flex;align-items:center;gap:6px;font-size:12px;font-weight:800;">'
        f'<div style="width:12px;height:12px;border-radius:3px;background:#7ED321;flex-shrink:0;"></div>'
        f'<span style="color:#7ED321;">🖥️ Meet</span>'
        f'<span style="color:#6B7280;font-weight:600;">{_cs_meets} &middot; {meets_pct}%</span>'
        '</div>'
        f'<div style="display:flex;align-items:center;gap:6px;font-size:12px;font-weight:800;">'
        f'<div style="width:12px;height:12px;border-radius:3px;background:#FF4D2E;flex-shrink:0;"></div>'
        f'<span style="color:#FF4D2E;">👻 No Answer</span>'
        f'<span style="color:#6B7280;font-weight:600;">{_cs_ghost} &middot; {ghost_pct}%</span>'
        '</div>'
        '</div>'
        '</div>'
    )
    st.markdown(_contact_html, unsafe_allow_html=True)

    # ── ROW 6: PROYECCIÓN DE CIERRE DE MES ────────────────────────────────────
    # Current sheets are updated weekly (every Sunday).
    # The export always reflects last Sunday's cutoff, NOT today.
    # Using today.day would inflate the denominator → artificially low daily rate.
    _today         = date.today()
    _days_in_month = calendar.monthrange(_today.year, _today.month)[1]

    # Last Sunday (if today IS Sunday, _days_since_sunday == 0 → use today)
    _days_since_sunday = (_today.weekday() + 1) % 7   # Sun=0, Mon=1 … Sat=6
    _last_sunday       = _today - timedelta(days=_days_since_sunday)
    # Clamp to day-1 if last Sunday fell in the prior month
    if _last_sunday.month != _today.month:
        _last_sunday = _today.replace(day=1)

    _elapsed_days  = _last_sunday.day              # days covered by the weekly export
    _remain_days   = _days_in_month - _today.day   # days still left this month

    # GMV projection
    _current_gmv_usd = vals.get("gmv_usd", 0)
    _daily_gmv_usd   = (_current_gmv_usd / _elapsed_days) if _elapsed_days > 0 else 0
    _projected_gmv   = _daily_gmv_usd * _days_in_month
    _ref_gmv_usd     = baseline_vals.get("gmv_usd", 0)
    _proj_vs_ref     = ((_projected_gmv / _ref_gmv_usd) - 1) if _ref_gmv_usd > 0 else 0
    _proj_sign       = "+" if _proj_vs_ref >= 0 else ""
    _proj_color      = "#7ED321" if _proj_vs_ref >= 0 else "#FF4D2E"
    _proj_arrow      = "&#9650;" if _proj_vs_ref >= 0 else "&#9660;"

    # ADS Revenue projection
    _current_ads_rev_usd = ads_revenue["usd"]
    _daily_ads_usd       = (_current_ads_rev_usd / _elapsed_days) if _elapsed_days > 0 else 0
    _projected_ads       = _daily_ads_usd * _days_in_month
    _ads_proj_pct        = (_projected_ads / ADS_REVENUE_TARGET_USD * 100) if ADS_REVENUE_TARGET_USD > 0 else 0
    _ads_needed_daily    = max(ADS_REVENUE_TARGET_USD - _current_ads_rev_usd, 0) / _remain_days if _remain_days > 0 else 0

    # Status label for GMV
    if _proj_vs_ref >= 0.05:
        _proj_status = "🟢 En camino de superar el mes anterior"
    elif _proj_vs_ref >= -0.05:
        _proj_status = "🟡 En línea con el mes anterior"
    else:
        _proj_status = "🔴 Proyección por debajo del mes anterior"

    # ADS status
    if _ads_proj_pct >= 100:
        _ads_status_label = "✅ En camino a cumplir target ADS"
        _ads_status_color = "#7ED321"
    elif _ads_proj_pct >= 75:
        _ads_status_label = "⚡ Necesita acelerar ADS"
        _ads_status_color = "#FF7124"
    else:
        _ads_status_label = "🚨 Gap crítico ADS"
        _ads_status_color = "#FF4D2E"

    st.markdown(
        f'<div class="stack-label" style="margin-bottom:10px;margin-top:8px;">📅 PROYECCIÓN DE CIERRE DE MES · corte dom. {_last_sunday.strftime("%d/%m")} · quedan {_remain_days} días</div>',
        unsafe_allow_html=True,
    )
    _pcol1, _pcol2, _pcol3 = st.columns(3)
    with _pcol1:
        st.markdown(f"""
        <div style="background:rgba(255,255,255,0.90);border:1px solid rgba(0,0,0,0.08);border-radius:16px;padding:16px 18px;height:100%;">
          <div style="font-size:11px;font-weight:900;text-transform:uppercase;color:#6B7280;margin-bottom:8px;">Proyección GMV al ritmo actual</div>
          <div style="font-size:24px;font-weight:900;color:#1A1A2E;line-height:1;">{fmt_usd(_projected_gmv)}</div>
          <div style="font-size:12px;color:#6B7280;margin-top:6px;">{fmt_ars(_projected_gmv * ARS_PER_USD)} estimado</div>
          <div style="margin-top:10px;display:flex;align-items:center;gap:8px;">
            <span style="font-size:22px;color:{_proj_color};font-weight:900;">{_proj_arrow}</span>
            <span style="font-size:15px;font-weight:900;color:{_proj_color};">{_proj_sign}{fmt_percent0(_proj_vs_ref)} vs mes anterior</span>
          </div>
          <div style="margin-top:8px;font-size:12px;color:#6B7280;">{_proj_status}</div>
        </div>
        """, unsafe_allow_html=True)
    with _pcol2:
        st.markdown(f"""
        <div style="background:rgba(255,255,255,0.90);border:1px solid rgba(0,0,0,0.08);border-radius:16px;padding:16px 18px;height:100%;">
          <div style="font-size:11px;font-weight:900;text-transform:uppercase;color:#6B7280;margin-bottom:8px;">Proyección ADS Revenue</div>
          <div style="font-size:24px;font-weight:900;color:#1A1A2E;line-height:1;">{fmt_usd(_projected_ads)}</div>
          <div style="font-size:12px;color:#6B7280;margin-top:6px;">Target: {fmt_usd(ADS_REVENUE_TARGET_USD)} &middot; {_ads_proj_pct:.0f}% cubierto</div>
          <div style="margin-top:10px;font-size:13px;font-weight:800;color:{_ads_status_color};">{_ads_status_label}</div>
          <div style="margin-top:6px;font-size:12px;color:#6B7280;">Necesitás generar {fmt_usd(_ads_needed_daily)}/día los próximos {_remain_days}d para llegar</div>
        </div>
        """, unsafe_allow_html=True)
    with _pcol3:
        st.markdown(f"""
        <div style="background:rgba(255,255,255,0.90);border:1px solid rgba(0,0,0,0.08);border-radius:16px;padding:16px 18px;height:100%;">
          <div style="font-size:11px;font-weight:900;text-transform:uppercase;color:#6B7280;margin-bottom:8px;">Ritmo diario actual</div>
          <div style="font-size:24px;font-weight:900;color:#1A1A2E;line-height:1;">{fmt_usd(_daily_gmv_usd)} <span style="font-size:14px;color:#6B7280;font-weight:700;">/día GMV</span></div>
          <div style="font-size:14px;font-weight:900;color:#1A1A2E;margin-top:8px;">{fmt_usd(_daily_ads_usd)} <span style="font-size:12px;color:#6B7280;font-weight:700;">/día ADS Rev.</span></div>
          <div style="margin-top:10px;font-size:12px;color:#6B7280;">Corte semanal: dom. {_last_sunday.strftime("%d/%m")} · {_elapsed_days} días acumulados</div>
          <div style="font-size:12px;color:#6B7280;margin-top:4px;">Datos actualizados desde Current sheets.</div>
        </div>
        """, unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════
    # 🩺 DATA HEALTH — el sistema se audita a sí mismo
    # Reconciliación entre las tres listas maestras (Asignación, Growth OS,
    # Detalle CABA) + frescura del archivo + fuente de configuración.
    # Convierte desalineaciones silenciosas en un control visible.
    # ══════════════════════════════════════════════════════════════════════
    try:
        _dh_cov = get_portfolio_gmv_aov_from_detalle_caba() or {}
        _dh_total   = int(_dh_cov.get("brands_total", 0))
        _dh_matched = int(_dh_cov.get("brands_matched", 0))
        _dh_nosales = int(_dh_cov.get("brands_no_sales", 0))
        _dh_cov_pct = (_dh_matched / _dh_total * 100) if _dh_total else 0

        # Marcas asignadas sin ficha en la hoja maestra Growth OS
        _dh_gos = load_growth_data()
        _dh_aj  = load_asignacion_activa()
        _dh_sin_ficha = 0
        if not _dh_gos.empty and not _dh_aj.empty and "id" in _dh_gos.columns:
            _gos_ids = set(_dh_gos["id"].apply(normalize_brand_id))
            _aj_ids  = set(_dh_aj["brand_id"].apply(normalize_brand_id))
            _dh_sin_ficha = len(_aj_ids - _gos_ids)

        # Frescura del archivo
        _dh_age_h = (datetime.now() - datetime.fromtimestamp(os.path.getmtime(EXCEL_FILE))).total_seconds() / 3600
        _dh_age_txt = (f"hace {_dh_age_h:.0f} h" if _dh_age_h < 48 else f"hace {_dh_age_h/24:.0f} días")
        _dh_age_color = "#7ED321" if _dh_age_h < 48 else ("#FF7124" if _dh_age_h < 24 * 7 else "#FF4D2E")

        _dh_cfg_src = "Hoja Config del Excel" if _app_cfg else "Defaults del código"
        _dh_issues  = len(st.session_state.get("_data_issues", {}))
        _dh_cov_color = "#7ED321" if _dh_cov_pct >= 70 else ("#FF7124" if _dh_cov_pct >= 50 else "#FF4D2E")

        st.markdown('<div style="font-size:13px;font-weight:900;text-transform:uppercase;color:#6B7280;margin:18px 0 10px 2px;">🩺 Data Health · reconciliación de fuentes</div>', unsafe_allow_html=True)
        _dhc1, _dhc2, _dhc3, _dhc4 = st.columns(4)
        _dh_card = (
            '<div style="background:rgba(255,255,255,0.90);border:1px solid rgba(0,0,0,0.08);'
            'border-radius:16px;padding:14px 16px;height:100%;">'
            '<div style="font-size:11px;font-weight:900;text-transform:uppercase;color:#6B7280;margin-bottom:6px;">{label}</div>'
            '<div style="font-size:22px;font-weight:900;color:{color};line-height:1;">{value}</div>'
            '<div style="font-size:11px;color:#6B7280;margin-top:6px;">{sub}</div></div>'
        )
        with _dhc1:
            st.markdown(_dh_card.format(
                label="Archivo de datos", color=_dh_age_color,
                value=_dh_age_txt,
                sub=f"{os.path.basename(EXCEL_FILE)}"), unsafe_allow_html=True)
        with _dhc2:
            st.markdown(_dh_card.format(
                label="Cobertura GMV del cruce", color=_dh_cov_color,
                value=f"{_dh_matched}/{_dh_total} · {_dh_cov_pct:.0f}%",
                sub="Marcas asignadas con GMV matcheado por ID"), unsafe_allow_html=True)
        with _dhc3:
            st.markdown(_dh_card.format(
                label="Radar de activación", color="#FF7124" if _dh_nosales else "#7ED321",
                value=f"{_dh_nosales} marcas",
                sub="Asignadas sin ventas en el período — candidatas a rescate"), unsafe_allow_html=True)
        with _dhc4:
            _dh4_color = "#FF4D2E" if (_dh_issues or _dh_sin_ficha) else "#7ED321"
            _dh4_val = "OK ✓" if not (_dh_issues or _dh_sin_ficha) else f"{_dh_issues + _dh_sin_ficha} pendientes"
            _dh4_sub = f"Avisos de datos: {_dh_issues} · Sin ficha en Growth OS: {_dh_sin_ficha} · Config: {_dh_cfg_src}"
            st.markdown(_dh_card.format(
                label="Integridad del sistema", color=_dh4_color,
                value=_dh4_val, sub=_dh4_sub), unsafe_allow_html=True)
    except Exception as e:
        _log_data_issue("Panel Data Health", e, "El panel de salud no pudo calcularse; el resto del dashboard no se ve afectado.")

    st.markdown(f"""
    <div class="legend-box">
        Source: {EXCEL_FILE} · Current sheets filtered to portfolio IDs. Effective Contacts from {CONTACTS_START_DATE.strftime('%Y-%m-%d')} onward, excluding No Contesta / Not Contacted.
    </div>
    """, unsafe_allow_html=True)


# =========================
# OPPORTUNITY LIST + FOLLOW-UP LIST
# =========================

def _to_numeric_series(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)


def _normalize_series(series):
    numeric = _to_numeric_series(series)
    max_value = numeric.max()
    min_value = numeric.min()

    if max_value == min_value:
        return pd.Series([0] * len(numeric), index=numeric.index)

    return (numeric - min_value) / (max_value - min_value)


def _is_active_status(value):
    text = clean(value, "").strip().lower()
    return text.startswith("active") or "🚀" in clean(value, "") or text.startswith("activated")


def _is_on_status(value):
    text = clean(value, "").strip().lower()
    return "on" in text and "off" not in text and "w1" not in text and "w2" not in text and "w3" not in text


def _churn_label_with_emoji(raw_status):
    """
    Convierte el valor crudo de churn (W1, W2, W3, Off, On) al label con emoji
    según los emojis definidos en la hoja Growth OS:
      ✅ On · ⚠️ W1 · 🚨 W2 · 🆘 W3 · 😴 Off
    Si el valor ya contiene emoji se devuelve tal cual.
    """
    text = clean(raw_status, "").strip()
    if not text:
        return "✅ On"
    low = text.lower()
    # Already has emoji — return as-is
    if any(c in text for c in ["✅", "⚠️", "🚨", "🆘", "😴", "☠️"]):
        return text
    if "w3" in low:
        return "🆘 W3"
    if "w2" in low:
        return "🚨 W2"
    if "w1" in low:
        return "⚠️ W1"
    if "off" in low:
        return "😴 Off"
    # On (or anything else)
    return "✅ On"


def _normalize_commercial_status(value):
    text = clean(value, "").strip()
    low = text.lower()

    if "❌" in text or "reject" in low or "rechaz" in low:
        return "❌"
    if "🚀" in text or "activ" in low:
        return "🚀"
    if "🏆" in text or "deal" in low or "cerrado" in low:
        return "🏆"
    if "⏳" in text or "⌛" in text or "negoti" in low or "negoci" in low:
        return "⏳"
    if "✅" in text or "sin novedad" in low or "no news" in low:
        return "✅"
    if "👻" in text or "💤" in text or "not contacted" in low or "no contesta" in low:
        return "👻"

    return "👻"




def _status_label_from_value(value, default="OFF 😴"):
    """Returns the visible pipeline status label used in Follow-Up List."""
    text = clean(value, "").strip()
    low = norm_text(text)

    if not text:
        return default
    if "reject" in low or "rechaz" in low or "❌" in text:
        return "Rejected ❌"
    if "deal" in low or "cerrado" in low or "closed" in low or "🏆" in text:
        return "Deal Closed 🏆"
    if "activated" in low or "activate" in low or "activado" in low or "🚀" in text:
        return "Activated 🚀"
    if "negoti" in low or "negoci" in low or "⏳" in text or "⌛" in text:
        return "Negotiation ⏳"
    if "follow-up" in low or "follow up" in low or "no news" in low or "sin novedad" in low or "✅" in text:
        return "Follow-up ✅"
    if "not contacted" in low or "no contesta" in low or "ghost" in low or "👻" in text:
        return "Not Contacted 👻"
    if "off" in low or "sleep" in low or "😴" in text or "💤" in text:
        return "OFF 😴"
    return default


def get_latest_status_map(default="OFF 😴"):
    """Latest manually selected opportunity status by brand from growth_os_comments.csv.
    If a brand has never been touched/saved with status, Follow-Up List should show OFF.
    """
    comments = _load_comments_df()
    if comments.empty or "brand_id" not in comments.columns:
        return {}
    comments = comments.sort_values(by="_dt", ascending=True, na_position="last")
    result = {}
    for bid, group in comments.groupby("brand_id"):
        statuses = [clean(x, "").strip() for x in group.get("opportunity_status", []) if clean(x, "").strip()]
        if statuses:
            result[normalize_brand_id(bid)] = _status_label_from_value(statuses[-1], default=default)
    return result

def _format_rank(number):
    try:
        return f"#{int(number)}"
    except Exception:
        return "-"


def _format_id(value):
    return f"AR-{normalize_brand_id(value)}"


def _prepare_growth_scored_data():
    df = load_growth_data()

    # ── Merge marcas de Asignacion Junio que no están en Growth OS ─────────────
    # Para cada una se crea una fila sintética con "?" en campos desconocidos,
    # de modo que aparezcan en Opportunity List, Follow-Up, Day Queue, etc.
    # También construimos lookup de turbo e is_new para enriquecer TODAS las filas.
    _aj_turbo_ids = set()
    _aj_new_ids   = set()
    _aj_all_ids   = set()
    try:
        aj = load_asignacion_activa()
        if not aj.empty:
            _aj_all_ids = set(aj["brand_id"].dropna().astype(str))
            _aj_all_ids.discard("")
            if "turbo" in aj.columns:
                _aj_turbo_ids = set(aj.loc[aj["turbo"] == True, "brand_id"].astype(str))
            if "is_new" in aj.columns:
                _aj_new_ids = set(aj.loc[aj["is_new"] == True, "brand_id"].astype(str))

            existing_ids = set()
            if not df.empty:
                id_col_check = get_id_column_name(df)
                if id_col_check:
                    existing_ids = set(df[id_col_check].apply(normalize_brand_id).dropna().astype(str))

            # ── Pre-build lookup tables for synthetic row enrichment ────────────
            # 1. Detalle CABA: category + GMV by brand_id
            _detalle_cat_map = {}   # brand_id -> categoria
            _detalle_gmv_map = {}   # brand_id -> gmv total
            try:
                _dc = load_detalle_caba()
                if not _dc.empty and "brand_id" in _dc.columns:
                    cat_col_dc = next((c for c in _dc.columns if "categor" in c), None)
                    gmv_col_dc = "_gmv" if "_gmv" in _dc.columns else next((c for c in _dc.columns if "gmv" in c), None)
                    if cat_col_dc:
                        for _bid_dc, _grp in _dc.groupby("brand_id"):
                            _cats = _grp[cat_col_dc].dropna().astype(str)
                            if not _cats.empty:
                                _detalle_cat_map[str(_bid_dc)] = _cats.mode().iloc[0] if not _cats.mode().empty else _cats.iloc[0]
                    if gmv_col_dc:
                        for _bid_dc, _grp in _dc.groupby("brand_id"):
                            _gmv_val = pd.to_numeric(_grp[gmv_col_dc], errors="coerce").fillna(0).sum()
                            _detalle_gmv_map[str(_bid_dc)] = float(_gmv_val)
            except Exception:
                pass

            # 2. CVR% map (reuse existing function)
            _cvr_map_syn = {}
            try:
                _cvr_map_syn = load_cvr_data()  # {brand_name_lower: avg_cvr}
            except Exception:
                pass

            # 3. Portfolio GMV map (Growth OS existing brands) for ranking reference
            _portfolio_gmv_list = []
            try:
                if not df.empty:
                    _gos_id_col = get_id_column_name(df)
                    _gos_gmv_col = next((c for c in df.columns if c in ["last gmv ars", "gmv ars"]), None)
                    if _gos_id_col and _gos_gmv_col:
                        for _, _gr in df.iterrows():
                            _g_bid = normalize_brand_id(_gr.get(_gos_id_col, ""))
                            _g_gmv = to_number(_gr.get(_gos_gmv_col, 0), 0)
                            if _g_bid and _g_gmv > 0:
                                _portfolio_gmv_list.append((_g_bid, _g_gmv))
            except Exception:
                pass

            new_rows = []
            for _, aj_row in aj.iterrows():
                bid = str(aj_row["brand_id"])
                if bid in existing_ids or not bid:
                    continue

                bname = str(aj_row["brand_name"])

                # Category from Detalle CABA (by brand_id)
                cat_syn = _detalle_cat_map.get(bid, "")

                # CVR from CVR% sheet (by brand name lower)
                cvr_syn = _cvr_map_syn.get(bname.strip().lower(), "")
                cvr_display = str(round(cvr_syn * 100, 1)) if cvr_syn and isinstance(cvr_syn, float) and cvr_syn > 0 else ""

                # GMV from Detalle CABA for ranking calculation
                gmv_syn = _detalle_gmv_map.get(bid, 0.0)

                # Portfolio ranking: how many existing brands have higher GMV
                ranking_syn = ""
                if gmv_syn > 0:
                    _higher = sum(1 for _, g in _portfolio_gmv_list if g > gmv_syn)
                    ranking_syn = f"#{_higher + 1}"

                synthetic = {
                    "id":              bid,
                    "name":            bname,
                    "country":         PORTFOLIO_COUNTRY or "-",
                    "ltor tier":       "No Priorizado",
                    "churn":           "",
                    "churn status":    "",
                    "category":        cat_syn,
                    "ads":             "",
                    "ads bookings":    "",
                    "ads roi":         "",
                    "md":              "",
                    "md status":       "",
                    "md bookings":     "",
                    "md roi":          "",
                    "last gmv ars":    "",
                    "gmv ars":         "",
                    "last aov ars":    "",
                    "aov ars":         "",
                    "cr %":            cvr_display,
                    "conversion rate": cvr_display,
                    "pro users %":     "",
                    "pro %":           "",
                    "comm. rate":      "",
                    "manager":         "",
                    "assistant":       "",
                    "email":           "",
                    "contact number":  "",
                    "ranking":         ranking_syn,
                    "comments":        "",
                }
                new_rows.append(synthetic)

            if new_rows:
                synthetic_df = pd.DataFrame(new_rows)
                synthetic_df.columns = [normalize(c) for c in synthetic_df.columns]
                if df.empty:
                    df = synthetic_df
                else:
                    df = pd.concat([df, synthetic_df], ignore_index=True)
    except Exception:
        pass  # Si falla el merge, seguir con lo que hay

    if df.empty:
        return df

    id_col = get_id_column_name(df)
    if not id_col:
        return pd.DataFrame()

    data = df.copy()

    data["_id"] = data[id_col].apply(normalize_brand_id)
    # ── Portafolio vigente = Asignacion Junio (fuente de verdad) ───────────────
    # Si la asignación cambió de mes, Growth OS puede conservar filas de marcas
    # que ya no son del portafolio actual (reasignadas a otro Farmer). Se filtran
    # acá para que Day Queue, Opportunity List, Follow-Up List y Brand Finder
    # reflejen únicamente el portafolio vigente, igual que el resto del dashboard.
    if _aj_all_ids:
        data = data[data["_id"].isin(_aj_all_ids)].copy()
    # ── Flags de Asignacion Junio disponibles en toda la app ──────────────────
    data["_is_new"]   = data["_id"].isin(_aj_new_ids)    # marca nueva (rojo en Excel)
    data["_is_turbo"] = data["_id"].isin(_aj_turbo_ids)  # tiene Store Turbo asignado
    data["_name"] = get_col(data, ["name", "brand name", "restaurant name"], "").apply(lambda x: clean(x, ""))
    data["_gmv"] = _to_numeric_series(get_col(data, ["last gmv ars", "gmv ars"], 0))
    data["_aov"] = _to_numeric_series(get_col(data, ["last aov ars", "aov ars"], 0))
    data["_cr"] = _to_numeric_series(get_col(data, ["cr %", "conversion rate", "conversion"], 0))
    data["_pro"] = _to_numeric_series(get_col(data, ["pro users %", "pro %"], 0))
    data["_ads"] = get_col(data, ["ads"], "").apply(lambda x: clean(x, ""))
    data["_ads_bookings"] = _to_numeric_series(get_col(data, ["ads bookings", "ad bookings"], 0))
    data["_md"] = get_col(data, ["md", "md status", "markdown"], "").apply(lambda x: clean(x, ""))
    data["_churn"] = get_col(data, ["churn", "churn status"], "").apply(lambda x: clean(x, ""))
    # Override with Current Churn sheet — source of truth for churn status.
    # Brands not in the sheet are considered 'On'.
    try:
        _churn_map = load_current_churn()
        if _churn_map:
            data["_churn"] = data["_id"].apply(lambda bid: _churn_map.get(normalize_brand_id(bid), "On"))
    except Exception:
        pass
    data["_commercial_status_raw"] = get_col(
        data,
        ["opp ", "opp", "opportunity status", "commercial status", "status"],
        "",
    ).apply(lambda x: clean(x, ""))

    gmv_norm = _normalize_series(data["_gmv"])
    cr_norm = _normalize_series(data["_cr"])
    pro_norm = _normalize_series(data["_pro"])
    aov_norm = _normalize_series(data["_aov"])

    data["_opportunity_score"] = (
        gmv_norm * 0.45
        + cr_norm * 0.25
        + pro_norm * 0.20
        + aov_norm * 0.10
    )

    data["_follow_up_score"] = (
        gmv_norm * 0.35
        + cr_norm * 0.25
        + pro_norm * 0.25
        + aov_norm * 0.15
    )

    return data


def _render_html_table(df, max_rows=200, visible_rows=10):
    """
    SaaS-style scrollable table — 10 visible rows, sticky header, hover lift effect.
    Pills/stickers on Name, Status, Brand, Restaurant, brand_name, Revenue Proj 80%,
    commercial_action, movement, pipeline_stage, opportunity_status columns.
    """
    if df is None or df.empty:
        st.info("No data to display.")
        return

    import uuid as _uuid
    tid = "t" + _uuid.uuid4().hex[:8]

    display_df = df.head(max_rows).copy()

    ROW_H   = 49
    THEAD_H = 42
    scroll_h = ROW_H * visible_rows + THEAD_H

    # ── Revenue multiplier (x4) helper ───────────────────────────────────────
    def _revenue_x4(text):
        """Parse a USD value and return x4, formatted."""
        try:
            clean_val = text.replace("$","").replace(",","").replace(" ","").strip()
            val = float(clean_val)
            mult = val * 4
            if mult >= 1000:
                return f"${mult:,.0f}"
            return f"${mult:.2f}"
        except Exception:
            return text

    # ── pill/sticker factory ─────────────────────────────────────────────────
    def _make_pill(text, bg, fg):
        return (
            f'<span style="display:inline-block;padding:3px 11px;border-radius:20px;'
            f'font-size:11px;font-weight:700;background:{bg};color:{fg};'
            f'white-space:nowrap;letter-spacing:0.01em;">{html.escape(text)}</span>'
        )

    # Name / Brand / Restaurant / brand_name sticker — palette blue
    def _name_pill(text):
        return _make_pill(text, "rgba(245,197,24,0.15)", "#F5C518")

    # Status pill — color driven by content
    def _status_pill(text):
        low = text.lower()
        if any(k in low for k in ["✅on", "✅ on", "active 🚀", "estable", "✅"]):
            return _make_pill(text, "rgba(111,242,75,0.12)", "#7ED321")
        elif any(k in low for k in ["⚠️w1", "⚠️ w1", "⚠️ revisar", "revisar"]):
            return _make_pill(text, "rgba(255,113,36,0.08)", "#D95A10")
        elif any(k in low for k in ["🚨w2", "🚨 w2"]):
            return _make_pill(text, "rgba(255,113,36,0.12)", "#D95A10")
        elif any(k in low for k in ["🆘w3", "🆘 w3"]):
            return _make_pill(text, "rgba(229,51,42,0.12)", "#FF4D2E")
        elif any(k in low for k in ["☠️off", "☠️ off", "😴off", "😴 off", "inactive 💤", "off"]):
            return _make_pill(text, "rgba(255,255,255,0.92)", "#6B7280")
        elif any(k in low for k in ["frío", "frio", "❄️"]):
            return _make_pill(text, "rgba(59,72,131,0.10)", "#1B3F8B")
        elif any(k in low for k in ["renegociación", "renegociacion", "🔄"]):
            return _make_pill(text, "rgba(59,72,131,0.10)", "#1B3F8B")
        elif any(k in low for k in ["prioritized 🔥"]):
            return _make_pill(text, "rgba(255,113,36,0.08)", "#D95A10")
        elif any(k in low for k in ["non prioritized"]):
            return _make_pill(text, "rgba(255,255,255,0.92)", "#6B7280")
        else:
            return _make_pill(text, "rgba(59,72,131,0.08)", "#6B7280")

    # Opp pill (already existed)
    def _opp_pill(text):
        low = text.lower()
        if any(k in low for k in ["🏆 acquire", "🏆acquire"]):
            return _make_pill(text, "rgba(255,113,36,0.08)", "#D95A10")
        elif any(k in low for k in ["🔧 upsell urgente", "🔧upsell urgente", "upsell urgente"]):
            return _make_pill(text, "rgba(229,51,42,0.08)", "#FF4D2E")
        elif any(k in low for k in ["⚡ upselling", "⚡upselling"]):
            return _make_pill(text, "rgba(59,72,131,0.10)", "#1B3F8B")
        elif "adquisición" in low or "adquisicion" in low:
            return _make_pill(text, "rgba(59,72,131,0.10)", "#1B3F8B")
        elif "upselling" in low:
            return _make_pill(text, "rgba(59,72,131,0.10)", "#1B3F8B")
        return _make_pill(text, "rgba(59,72,131,0.08)", "#6B7280")

    # Revenue Proj 80% pill — green money sticker, value x4
    def _revenue_pill(text):
        if text in ("-", "", "—"):
            return f'<span style="font-size:12px;color:rgba(255,255,255,0.15);">—</span>'
        val4 = _revenue_x4(text)
        return _make_pill(f"↑ {val4}", "rgba(111,242,75,0.12)", "#7ED321")

    # commercial_action / movement pill — indigo
    def _action_pill(text):
        low = text.lower()
        if any(k in low for k in ["closed", "cerrado", "won", "ganado"]):
            return _make_pill(text, "rgba(111,242,75,0.12)", "#7ED321")
        elif any(k in low for k in ["rejected", "rechazado", "lost"]):
            return _make_pill(text, "rgba(229,51,42,0.12)", "#FF4D2E")
        elif any(k in low for k in ["negotiation", "negociación", "pipeline"]):
            return _make_pill(text, "rgba(255,113,36,0.08)", "#D95A10")
        return _make_pill(text, "rgba(59,72,131,0.15)", "#1B3F8B")

    # pipeline_stage / opportunity_status pill — amber
    def _stage_pill(text):
        low = text.lower()
        if any(k in low for k in ["closed", "won", "cerrado"]):
            return _make_pill(text, "rgba(111,242,75,0.12)", "#7ED321")
        elif any(k in low for k in ["rejected", "lost", "rechazado"]):
            return _make_pill(text, "rgba(229,51,42,0.12)", "#FF4D2E")
        return _make_pill(text, "rgba(255,113,36,0.08)", "#D95A10")

    # ── column routing map ───────────────────────────────────────────────────
    # key = col name lowercased, value = pill function

    def _pene_pill(text):
        low = text.lower()
        if "✅" in low or "en rango" in low:
            return _make_pill(text, "rgba(111,242,75,0.12)", "#7ED321")
        elif "⚠️" in low or "sobre techo" in low:
            return _make_pill(text, "rgba(255,113,36,0.12)", "#D95A10")
        elif "bajo" in low:
            return _make_pill(text, "rgba(255,113,36,0.08)", "#D95A10")
        elif "0%" in low or "sin promo" in low or "objetivo" in low:
            return _make_pill(text, "rgba(255,113,36,0.08)", "#D95A10")
        elif "sin gmv" in low:
            return _make_pill(text, "rgba(255,255,255,0.92)", "#6B7280")
        return _make_pill(text, "rgba(59,72,131,0.08)", "#6B7280")

    COL_PILL_MAP = {
        "name":               _name_pill,
        "restaurant":         _name_pill,
        "brand":              _name_pill,
        "brand_name":         _name_pill,
        "brand name":         _name_pill,
        "status":             _status_pill,
        "churn status":       _status_pill,
        "opp":                _opp_pill,
        "md strategy":        _opp_pill,
        "opportunity":        _opp_pill,
        "ltor":               _status_pill,
        "revenue proj 80%":   _revenue_pill,
        "commercial_action":  _action_pill,
        "commercial action":  _action_pill,
        "movement":           _action_pill,
        "pipeline_stage":     _stage_pill,
        "pipeline stage":     _stage_pill,
        "opportunity_status": _stage_pill,
        "opportunity status": _stage_pill,
        "strategic note":     _status_pill,
        "notes mtd":          _status_pill,
        "false roi check":    _status_pill,
        "pressure stability": _status_pill,
        "renegotiation":      _status_pill,
        "penetración md":     _pene_pill,
        "penetracion md":     _pene_pill,
        "cierre":             _status_pill,
        "días":               lambda t: t,   # raw HTML badge — no escaping
        "próximo contacto":   lambda t: t,   # raw HTML badge — no escaping
        "roi trend":          lambda t: t,   # raw HTML SVG sparkline — no escaping
    }

    # ── sticky header ────────────────────────────────────────────────────────
    th_base = (
        'position:sticky;top:0;z-index:2;padding:10px 14px;'
        'text-align:left;font-size:11px;font-weight:700;'
        'letter-spacing:0.05em;text-transform:uppercase;color:rgba(107,114,128,0.60);'
        'background:rgba(255,255,255,0.92);border-bottom:2px solid rgba(255,255,255,0.92);'
        'white-space:nowrap;box-shadow:0 1px 0 rgba(255,255,255,0.92);'
    )
    header_cells = (
        f'<th style="position:sticky;top:0;z-index:2;padding:10px 8px 10px 20px;width:36px;'
        f'text-align:center;font-size:11px;font-weight:700;letter-spacing:0.05em;'
        f'text-transform:uppercase;color:rgba(107,114,128,0.60);background:rgba(255,255,255,0.92);'
        f'border-bottom:2px solid rgba(255,255,255,0.92);box-shadow:0 1px 0 rgba(255,255,255,0.92);">N.</th>'
    )
    for col in display_df.columns:
        header_cells += f'<th style="{th_base}">{html.escape(str(col))}</th>'
    header = f'<thead><tr style="background:rgba(255,255,255,0.92);">{header_cells}</tr></thead>'

    # ── rows ─────────────────────────────────────────────────────────────────
    rows_html = []
    for i, (_, row) in enumerate(display_df.iterrows()):
        base_bg  = "rgba(255,255,255,0.92)" if i % 2 == 0 else "rgba(27,63,139,0.03)"
        hover_in = (
            "this.style.background='rgba(59,72,131,0.15)';"
            "this.style.boxShadow='0 4px 18px rgba(78,99,217,0.13)';"
            "this.style.transform='scale(1.003) translateY(-1px)';"
            "this.style.zIndex='1';"
            "this.style.position='relative';"
        )
        hover_out = (
            f"this.style.background='{base_bg}';"
            "this.style.boxShadow='none';"
            "this.style.transform='none';"
            "this.style.zIndex='0';"
        )
        row_num = (
            f'<td style="padding:12px 8px 12px 20px;text-align:center;'
            f'font-size:12px;font-weight:600;color:rgba(255,255,255,0.15);'
            f'border-bottom:1px solid rgba(255,255,255,0.92);">{i+1}</td>'
        )
        cells = row_num
        for col, val in zip(display_df.columns, row):
            try:
                is_nan = isinstance(val, float) and math.isnan(val)
            except Exception:
                is_nan = False
            text = "-" if (val is None or is_nan) else str(val)
            col_key = str(col).lower().strip()

            pill_fn = COL_PILL_MAP.get(col_key)
            if pill_fn and text not in ("-", "", "—"):
                cell_inner = pill_fn(text)
            elif text in ("-", "", "—"):
                cell_inner = '<span style="font-size:12px;color:rgba(255,255,255,0.15);">—</span>'
            else:
                stripped = (text.replace("ARS","").replace("USD","").replace("$","")
                               .replace(".","").replace(",","").replace("%","").replace("x","").strip())
                is_numeric = stripped.lstrip("-").isdigit()
                if is_numeric:
                    cell_inner = f'<span style="font-size:13px;font-weight:700;color:#1A1A2E;">{html.escape(text)}</span>'
                else:
                    cell_inner = f'<span style="font-size:12px;color:#6B7280;">{html.escape(text)}</span>'

            # For raw-HTML columns (días, próximo contacto, roi trend) strip tags for the tooltip
            import re as _re
            tooltip_text = _re.sub(r"<[^>]+>", "", text) if col_key in ("días", "próximo contacto", "roi trend") else text
            cells += (
                f'<td style="padding:12px 14px;border-bottom:1px solid rgba(255,255,255,0.92);'
                f'white-space:nowrap;max-width:260px;overflow:hidden;text-overflow:ellipsis;'
                f'transition:background 0.15s;" title="{html.escape(tooltip_text)}">'
                f'{cell_inner}</td>'
            )
        rows_html.append(
            f'<tr style="background:{base_bg};cursor:default;'
            f'transition:box-shadow 0.18s,transform 0.18s,background 0.15s;"'
            f' onmouseover="{hover_in}" onmouseout="{hover_out}">'
            f'{cells}</tr>'
        )

    body = f'<tbody>{"".join(rows_html)}</tbody>'

    count_note = ""
    if len(df) > max_rows:
        count_note = (
            f'<div style="font-size:11px;color:rgba(107,114,128,0.60);padding:8px 20px 12px;">'
            f'Showing first {max_rows} of {len(df)} rows.</div>'
        )

    table_html = (
        f'<div id="{tid}" style="border-radius:16px;border:1px solid rgba(0,0,0,0.08);'
        f'box-shadow:0 2px 16px rgba(0,0,0,0.07);margin:8px 0 20px 0;'
        f'background:rgba(255,255,255,0.92);overflow:hidden;">'
        f'<div style="overflow-x:auto;overflow-y:auto;max-height:{scroll_h}px;">'
        f'<table style="width:100%;border-collapse:collapse;'
        f'font-family:Inter,-apple-system,sans-serif;">'
        f'{header}{body}</table>'
        f'</div>'
        f'{count_note}'
        f'</div>'
    )

    st.markdown(table_html, unsafe_allow_html=True)
def _render_light_table(df, height=420):
    _render_html_table(df)


def _render_target_progress_bar(label, active_usd, pipeline_usd, target_usd, color_active="#7ED321", color_pipeline="#FF7124", projected_usd=0):
    """
    Barra de 4 segmentos:
      - Verde:   Activo hoy (REVENUE NET MTD)
      - Azul:    Proyectado restante (BOOKINGS x 90% - REVENUE NET)
      - Naranja: Pipeline (Opp List) — hacia 100% o hacia 120% si ya se cubrió
      - Gris:    Gap — distancia a 100% o a 120% según el caso
    """
    if target_usd <= 0:
        return

    color_projected = "#4B9CF2"
    color_muted     = "#8C93AC" if DARK_MODE else COLORS["muted"]
    color_card      = "#141A2E" if DARK_MODE else COLORS["card"]
    color_border    = "rgba(255,255,255,0.10)" if DARK_MODE else COLORS["border"]
    target_120      = target_usd * 1.2
    base_covered    = active_usd + projected_usd

    target_140      = target_usd * 1.4
    at_100          = base_covered >= target_usd
    at_120          = base_covered >= target_120
    at_140          = base_covered >= target_140
    pct_covered     = base_covered / target_usd if target_usd > 0 else 0

    # Bar ceiling escalates: <100% → 100%, 100–120% → 120%, 120–140% → 140%, 140%+ → 140%
    if at_140:
        bar_ceiling = target_140
    elif at_120:
        bar_ceiling = target_140
    elif at_100:
        bar_ceiling = target_120
    else:
        bar_ceiling = target_usd

    pct_active    = min(active_usd / bar_ceiling, 1.0) * 100
    pct_projected = min(projected_usd / bar_ceiling, max(0.0, 1.0 - pct_active / 100)) * 100

    if at_120:
        # Ya superó 120% — peleando por 140%
        tramo_next          = target_usd * 0.2          # tramo de 120% a 140%
        pipeline_capped     = min(pipeline_usd, tramo_next)
        gap_usd             = max(tramo_next - pipeline_usd, 0)
        pct_pipeline        = (pipeline_capped / bar_ceiling) * 100
        pct_gap             = (gap_usd / bar_ceiling) * 100
        overall_label       = "{:.0f}% cubierto".format(pct_covered * 100)
        gap_label           = "Gap a 140%"
        marker_pct_100      = target_usd / bar_ceiling * 100
        marker_pct_120      = target_120 / bar_ceiling * 100
        marker_html         = (
            '<div style="position:absolute;top:-2px;left:{:.1f}%;width:2px;height:18px;background:rgba(255,255,255,0.35);border-radius:1px;"></div>'.format(marker_pct_100) +
            '<div style="position:absolute;top:-2px;left:{:.1f}%;width:2px;height:18px;background:rgba(255,255,255,0.65);border-radius:1px;"></div>'.format(marker_pct_120)
        )
        scale_html          = '<div style="display:flex;justify-content:space-between;font-size:10px;color:{};margin-bottom:8px;padding:0 2px;"><span>0</span><span>100%</span><span>120%</span><span>140%</span></div>'.format(color_muted)
        if at_140:
            status_label    = "🏆 +140% — Modo bestia"
            status_color    = "#7ED321"
        else:
            status_label    = "🔥 Peleando 140%"
            status_color    = "#7ED321"
    elif at_100:
        # Superó 100% — peleando por 120%
        tramo_120       = target_usd * 0.2
        pipeline_capped = min(pipeline_usd, tramo_120)
        gap_usd         = max(tramo_120 - pipeline_usd, 0)
        pct_pipeline    = (pipeline_capped / bar_ceiling) * 100
        pct_gap         = (gap_usd / bar_ceiling) * 100
        overall_label   = "{:.0f}% cubierto".format(pct_covered * 100)
        status_color    = "#4B9CF2"
        status_label    = "⚡ Peleando 120%"
        gap_label       = "Gap a 120%"
        marker_html     = '<div style="position:absolute;top:-2px;left:{:.1f}%;width:2px;height:18px;background:rgba(255,255,255,0.45);border-radius:1px;"></div>'.format(target_usd / bar_ceiling * 100)
        scale_html      = '<div style="display:flex;justify-content:space-between;font-size:10px;color:{};margin-bottom:8px;padding:0 2px;"><span>0</span><span>100%</span><span>120%</span></div>'.format(color_muted)
    else:
        remaining       = target_usd - base_covered
        pipeline_capped = min(pipeline_usd, remaining)
        gap_usd         = max(remaining - pipeline_usd, 0)
        pct_pipeline    = (pipeline_capped / bar_ceiling) * 100
        pct_gap         = (gap_usd / bar_ceiling) * 100
        overall_pct     = min((base_covered + pipeline_usd) / target_usd * 100, 100)
        overall_label   = "{:.0f}% cubierto".format(overall_pct)
        status_color    = "#FF7124" if pct_covered >= 0.70 else "#FF4D2E"
        status_label    = "Needs focus" if pct_covered >= 0.70 else "Gap critical"
        gap_label       = "Gap"
        marker_html     = ""
        scale_html      = '<div style="margin-bottom:8px;"></div>'

    color_track     = "rgba(255,255,255,0.08)" if DARK_MODE else "rgba(255,255,255,0.95)"
    color_badge_bg  = "rgba(255,255,255,0.06)" if DARK_MODE else "rgba(255,255,255,.06)"
    color_rest      = "rgba(255,255,255,0.10)" if DARK_MODE else "rgba(255,255,255,.04)"
    color_gap_seg   = "rgba(255,255,255,0.14)" if DARK_MODE else "rgba(255,255,255,.07)"
    color_gap_dot   = "rgba(255,255,255,0.30)" if DARK_MODE else "rgba(255,255,255,.25)"

    html = """
<div style="background:{card};border:1px solid {border};border-radius:16px;padding:18px 22px 16px;margin-bottom:18px;">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">
    <div style="font-size:13px;font-weight:800;text-transform:uppercase;letter-spacing:.05em;color:{muted};">{label}</div>
    <div style="font-size:12px;font-weight:700;color:{sc};background:{badge_bg};border-radius:20px;padding:3px 12px;border:1px solid {sc}40;">{sl}</div>
  </div>
  <div style="position:relative;margin-bottom:4px;">
    <div style="display:flex;height:14px;border-radius:8px;overflow:hidden;background:{track};">
      <div style="width:{pa:.1f}%;background:{ca};border-radius:8px 0 0 8px;transition:width .4s;"></div>
      <div style="width:{pp:.1f}%;background:{cp};transition:width .4s;"></div>
      <div style="width:{ppl:.1f}%;background:{cpl};transition:width .4s;"></div>
      <div style="width:{pg:.1f}%;background:{gap_seg};transition:width .4s;"></div>
      <div style="flex:1;background:{rest};border-radius:0 8px 8px 0;"></div>
    </div>
    {marker}
  </div>
  {scale}
  <div style="display:flex;gap:24px;flex-wrap:wrap;">
    <div style="font-size:12px;color:{muted};">
      <span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:{ca};margin-right:5px;"></span>
      <b style="color:{ca};">Activo hoy</b>&nbsp; {v_active}
    </div>
    <div style="font-size:12px;color:{muted};">
      <span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:{cp};margin-right:5px;"></span>
      <b style="color:{cp};">Proyectado restante</b>&nbsp; {v_proj}
    </div>
    <div style="font-size:12px;color:{muted};">
      <span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:{cpl};margin-right:5px;"></span>
      <b style="color:{cpl};">Pipeline (Opp List)</b>&nbsp; {v_pipe}
    </div>
    <div style="font-size:12px;color:{muted};">
      <span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:{gap_dot};margin-right:5px;"></span>
      <b>{gl}</b>&nbsp; {v_gap}
    </div>
    <div style="font-size:12px;color:{muted};margin-left:auto;">
      <b>Target:</b>&nbsp; {v_target}&nbsp;&nbsp;
      <b style="color:{sc};">{ol}</b>
    </div>
  </div>
</div>""".format(
        card=color_card, border=color_border, muted=color_muted,
        track=color_track, badge_bg=color_badge_bg, rest=color_rest,
        gap_seg=color_gap_seg, gap_dot=color_gap_dot,
        label=label,
        sc=status_color, sl=status_label,
        pa=pct_active, ca=color_active,
        pp=pct_projected, cp=color_projected,
        ppl=pct_pipeline, cpl=color_pipeline,
        pg=pct_gap,
        marker=marker_html, scale=scale_html,
        v_active=fmt_usd(active_usd),
        v_proj=fmt_usd(projected_usd),
        v_pipe=fmt_usd(pipeline_usd),
        gl=gap_label, v_gap=fmt_usd(gap_usd),
        v_target=fmt_usd(target_usd), ol=overall_label,
    )
    st.markdown(html, unsafe_allow_html=True)


def _render_churn_distribution_bar(counts, total):
    """
    Barra de distribución de severidad de churn sobre el total de stores.
    counts: dict con claves 'On', 'W1', 'W2', 'W3', 'Off' → cantidad de stores.
    total: número total de stores (Current Churn).
    """
    if total <= 0:
        return

    color_border = "rgba(255,255,255,0.10)" if DARK_MODE else COLORS["border"]
    color_muted  = "#8C93AC" if DARK_MODE else COLORS["muted"]
    color_card   = "#141A2E" if DARK_MODE else COLORS["card"]
    color_track  = "rgba(255,255,255,0.08)" if DARK_MODE else "rgba(255,255,255,0.95)"
    color_badge_bg = "rgba(255,255,255,0.06)"

    segments = [
        ("On",  "✅ On",  "#7ED321"),
        ("W1",  "⚠️ W1",  "#FFD166"),
        ("W2",  "🚨 W2",  "#FF7124"),
        ("W3",  "🆘 W3",  "#FF4D2E"),
        ("Off", "😴 Off", "rgba(255,255,255,.25)"),
    ]

    bars_html = ""
    legend_html = ""
    for key, lbl, color in segments:
        n = counts.get(key, 0)
        pct = (n / total) * 100 if total > 0 else 0
        if pct > 0:
            bars_html += (
                f'<div style="width:{pct:.1f}%;background:{color};transition:width .4s;" '
                f'title="{lbl}: {n}"></div>'
            )
        legend_html += (
            f'<div style="font-size:12px;color:{color_muted};">'
            f'<span style="display:inline-block;width:10px;height:10px;border-radius:50%;'
            f'background:{color};margin-right:5px;"></span>'
            f'<b style="color:{color};">{lbl}</b>&nbsp; {n} '
            f'<span style="opacity:.6;">({pct:.1f}%)</span></div>'
        )

    n_churned = total - counts.get("On", 0)
    pct_churned = (n_churned / total) * 100 if total > 0 else 0

    html = f"""
<div style="background:{color_card};border:1px solid {color_border};border-radius:16px;padding:18px 22px 16px;margin-bottom:18px;">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">
    <div style="font-size:13px;font-weight:800;text-transform:uppercase;letter-spacing:.05em;color:{color_muted};">
      Distribución de Churn · {total} stores
    </div>
    <div style="font-size:12px;font-weight:700;color:{COLORS['danger']};background:rgba(255,255,255,.06);border-radius:20px;padding:3px 12px;border:1px solid {COLORS['danger']}40;">
      {n_churned} en riesgo ({pct_churned:.1f}%)
    </div>
  </div>
  <div style="display:flex;height:14px;border-radius:8px;overflow:hidden;background:{color_track};margin-bottom:12px;">
    {bars_html}
  </div>
  <div style="display:flex;gap:24px;flex-wrap:wrap;">
    {legend_html}
  </div>
</div>"""
    st.markdown(html, unsafe_allow_html=True)



def _render_target_input_block(ads_target_default, md_target_default):
    """
    Compact target input row. Returns (ads_target_usd, md_target_usd, weeks_left).
    Tries to pre-fill from the Earnings sheet; user can override inline.
    """
    today = date.today()
    days_in_month   = calendar.monthrange(today.year, today.month)[1]
    elapsed_days    = today.day
    remaining_days  = days_in_month - elapsed_days
    weeks_left      = max(remaining_days / 7, 0.5)   # at least half a week
    week_of_month   = math.ceil(elapsed_days / 7)

    _tib_card   = "#141A2E" if DARK_MODE else COLORS['card']
    _tib_border = "rgba(255,255,255,0.10)" if DARK_MODE else COLORS['border']
    _tib_muted  = "#8C93AC" if DARK_MODE else COLORS['muted']

    st.markdown(f"""
    <div style="
        background: {_tib_card};
        border: 1px solid {_tib_border};
        border-radius: 14px;
        padding: 14px 20px 12px;
        margin-bottom: 16px;
        display:flex; align-items:center; gap:12px; flex-wrap:wrap;
    ">
        <div style="font-size:12px; font-weight:800; text-transform:uppercase; letter-spacing:.04em; color:{_tib_muted};">
            🎯 Target Engine
        </div>
        <div style="font-size:12px; color:{_tib_muted}; margin-left:auto;">
            Semana <b>{week_of_month}</b> del mes &nbsp;·&nbsp;
            Quedan <b>{remaining_days} días</b> &nbsp;(<b>{weeks_left:.1f} semanas</b>)
        </div>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns([2, 2, 1])
    with c1:
        ads_target_input = st.number_input(
            "🟠 ADS Revenue Target (USD)",
            min_value=0.0,
            value=float(ads_target_default) if ads_target_default else 16038.0,
            step=500.0,
            key="opp_ads_target",
            help="Target mensual de ADS Revenue en USD. Se pre-llena desde tu Earnings sheet.",
        )
    with c2:
        md_target_input = st.number_input(
            "🔵 MD GMV Target (USD)",
            min_value=0.0,
            value=float(md_target_default) if md_target_default else 0.0,
            step=500.0,
            key="opp_md_target",
            help="Target mensual de MD GMV en USD.",
        )
    with c3:
        st.metric("Semanas restantes", f"{weeks_left:.1f}")

    return ads_target_input, md_target_input, weeks_left


@st.cache_data(ttl=3000, show_spinner=False)
def load_current_churn_raw_df():
    """
    Lee la hoja Current Churn completa y devuelve el DataFrame normalizado.
    Usado por el Churn Segment de Opportunity List (una fila por store).
    Cacheado para no releer el Excel en cada render.
    """
    if not os.path.exists(EXCEL_FILE):
        return pd.DataFrame()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=CURRENT_CHURN_SHEET)
        df.columns = [normalize(c) for c in df.columns]
        return df
    except Exception as e:
        _log_data_issue('Current Churn', e, 'Verificá columnas COUNTRY_BRAND_ID y Estado Actual.')
        return pd.DataFrame()


def page_opportunity_list():
    render_header("Opportunity List", "Target-driven commercial radar · Ads, Markdown & Churn")

    data = _prepare_growth_scored_data()

    if data.empty:
        st.error("Could not load Growth OS data or ID column.")
        return

    # ── Load targets ─────────────────────────────────────────────────────────
    # ADS target: hardcoded constant (17,574 USD)
    ads_target_from_sheet = ADS_REVENUE_TARGET_USD
    # ADS result: sum of REVENUE NET from Current ADS sheet (computed below after get_current_ads_totals)
    raw_earnings = load_earnings_data()
    # ads_result_from_sheet will be set from Current ADS totals below

    # MD targets: read % from Earnings sheet (col F=MD, col G=MD Pro), row 3
    _md_targets = _read_md_targets_from_earnings()
    md_target_pct     = _md_targets["md_target_pct"]      # e.g. 0.0667
    md_pro_target_pct = _md_targets["md_pro_target_pct"]  # e.g. 0.0727

    # MD activo: read directly from Total row of Current MD sheets (col E, only when E > 0)
    _md_totals     = _read_md_totals_from_sheet(pro=False)
    _md_pro_totals = _read_md_totals_from_sheet(pro=True)
    active_md_gmv_usd     = _md_totals["markdown_usd"]      # col E total row, Current MD
    active_md_pro_gmv_usd = _md_pro_totals["markdown_usd"]  # col E total row, Current MD pro
    md_gmv_total_usd      = _md_totals["gmv_total_usd"]     # col D total row (same in both)

    # MD GMV targets = GMV Total (col D) × target % from Earnings
    md_gmv_target_usd     = md_gmv_total_usd * md_target_pct
    md_pro_gmv_target_usd = md_gmv_total_usd * md_pro_target_pct
    # Combined target for the progress bar (MD + MD Pro together)
    active_md_combined_usd = active_md_gmv_usd + active_md_pro_gmv_usd
    md_combined_target_usd = md_gmv_target_usd + md_pro_gmv_target_usd

    # ── Umbral de comisión MD ────────────────────────────────────────────────
    # La comisión de MD se paga solo si la penetración MD actual (MD activo /
    # GMV base) alcanza al menos el 90% del target MD (no del target combinado,
    # ni de MD Pro). Por debajo de ese piso, no hay comisión aunque el target
    # combinado esté parcialmente cubierto.
    MD_COMMISSION_THRESHOLD_PCT = 0.90
    md_pene_actual_pct  = (active_md_gmv_usd / md_gmv_total_usd) if md_gmv_total_usd > 0 else 0
    md_commission_floor_usd = md_gmv_target_usd * MD_COMMISSION_THRESHOLD_PCT
    md_commission_pct_of_target = (md_pene_actual_pct / md_target_pct) if md_target_pct > 0 else 0
    md_commission_paid = md_commission_pct_of_target >= MD_COMMISSION_THRESHOLD_PCT
    # Gap real para pagar comisión = lo que falta de MD activo para llegar al 90% del target MD
    md_commission_gap_usd = max(md_commission_floor_usd - active_md_gmv_usd, 0)

    # For backward compat with portfolio_gmv_usd references below
    portfolio_gmv_totals = get_current_gmv_totals()
    portfolio_gmv_usd = to_number(portfolio_gmv_totals.get("gmv_usd"), 0) if portfolio_gmv_totals else 0

    # ── ADS target (hardcoded, Target Engine eliminado) ─────────────────────
    ads_target_usd = ADS_REVENUE_TARGET_USD  # 17,574 USD

    # ── Build current-active revenue totals (suma REVENUE NET de Current ADS) ─
    ads_totals = get_current_ads_totals()
    # Activo hoy = REVENUE NET acumulado MTD (lo que ya se consumió)
    ads_result_from_sheet = to_number(ads_totals.get("revenue_usd"), 0)
    active_ads_revenue_usd = ads_result_from_sheet
    # Proyectado RESTANTE = (BOOKINGS NET × 80%) - REVENUE NET ya consumido
    # Es el tramo azul: lo que todavía falta llegar de las campañas que ya están corriendo
    _projected_total = to_number(ads_totals.get("projected_revenue_usd"), 0)
    ads_projected_usd = max(_projected_total - ads_result_from_sheet, 0)

    # weeks_left para Rev Proj de la tabla
    import calendar as _cal
    _today = date.today()
    _days_in_month = _cal.monthrange(_today.year, _today.month)[1]
    _remaining_days = _days_in_month - _today.day
    _elapsed_days_opp = _today.day
    weeks_left = max(_remaining_days / 7, 0.5)

    # ── Semana del mes actual (para proyectar GMV de Current GMV al mes completo) ──
    # week_of_month: 1 = primera semana, 2 = segunda, etc.
    _week_of_month_opp = max(math.ceil(_elapsed_days_opp / 7), 1)
    # Factor de proyección: si estamos en semana 2 de 4, el GMV acumulado
    # representa ~50% del mes → proyectamos × (4 / semana_actual)
    _gmv_projection_factor = min(4.0 / _week_of_month_opp, 4.0)

    # ── Mapa de GMV actual por brand (Current GMV sheet) ─────────────────────
    # Usamos esto para el booking sugerido en la Opp List.
    _current_gmv_df = load_current_gmv_data()
    _current_gmv_map: dict = {}  # {brand_id: gmv_ars_proyectado_al_mes}
    if not _current_gmv_df.empty and "gmv ars" in _current_gmv_df.columns:
        for _, _cgrow in _current_gmv_df.iterrows():
            _cgid = normalize_brand_id(_cgrow.get("_id", ""))
            _cg_gmv = to_number(_cgrow.get("gmv ars"), 0)
            if _cgid and _cg_gmv > 0:
                # Proyectar GMV acumulado MTD al mes completo
                _current_gmv_map[_cgid] = _cg_gmv * _gmv_projection_factor

    # ── Build ADS and MD maps ─────────────────────────────────────────────────
    def build_ads_map():
        current_ads = load_current_ads_data(portfolio_only=False)
        if current_ads.empty:
            return {}
        metrics = {}
        for _, row in current_ads.iterrows():
            brand_id = normalize_brand_id(row.get("_id"))
            bookings = to_number(row.get("bookings net"), 0)
            revenue  = to_number(row.get("revenue net"), 0)
            sales    = to_number(row.get("sales ads usd"), 0)
            roi      = to_number(row.get("roi"), 0)
            active   = any(v > 0 for v in [bookings, revenue, sales])
            metrics[brand_id] = {"active": active, "roi": roi}
        return metrics

    def build_md_map(pro=False):
        current_md = load_current_md_data(portfolio_only=False, pro=pro)
        if current_md.empty:
            return {}
        grouped = current_md.groupby("_id", as_index=False).agg({
            "_sales_usd": "sum",
            "_gmv_usd": "sum",
            "_campaigns": "sum",
            "_orders": "sum",
            "_roi_raw": "mean",
        })
        metrics = {}
        for _, row in grouped.iterrows():
            brand_id = normalize_brand_id(row.get("_id"))
            sales  = to_number(row.get("_sales_usd"), 0)
            gmv    = to_number(row.get("_gmv_usd"), 0)
            orders = to_number(row.get("_orders"), 0)
            roi    = (gmv / sales) if sales else to_number(row.get("_roi_raw"), 0)
            metrics[brand_id] = {"active": orders > 0, "roi": roi, "gmv_usd": gmv}
        return metrics

    ads_metrics = build_ads_map()
    md_metrics  = build_md_map(pro=False)

    # ═══════════════════════════════════════════════════════════════════════════
    # ADS SEGMENT
    # ═══════════════════════════════════════════════════════════════════════════
    data["_ads_current_active"] = data["_id"].apply(
        lambda x: ads_metrics.get(normalize_brand_id(x), {}).get("active", False)
    )
    data["_ads_current_roi"] = data["_id"].apply(
        lambda x: ads_metrics.get(normalize_brand_id(x), {}).get("roi", 0)
    )

    ads_acquire   = ~data["_ads_current_active"]
    # Upselling (ROI > 4.5x) ya está contabilizado en ads_result_from_sheet (BOOKINGS × 90%),
    # por lo que el pipeline solo incluye Acquire — marcas que aún no están activas.
    ads_upselling = data["_ads_current_active"] & (data["_ads_current_roi"] > 4.5)

    ads_df = data[ads_acquire | ads_upselling].copy()
    ads_df["_opp_group"] = ads_df.apply(
        lambda r: 0 if not r["_ads_current_active"] else 1, axis=1
    )
    ads_df["Opp"] = ads_df["_opp_group"].map({0: "🏆 Acquire", 1: "⚡ Upselling"})
    # Status = cadencia de contacto (🟢/🟡/🟠/🔴) — mismo lenguaje que Salud de Cartera
    _opp_prod_map  = get_productivity_last_contact_map(EXCEL_FILE)
    _opp_meta_map  = get_last_comment_meta_map(limit=1)
    ads_df["_last_contact_dt"] = ads_df.apply(
        lambda r: get_last_contact_dt(r.get("_id", ""), r.get("_name", ""), _opp_prod_map, _opp_meta_map), axis=1
    )
    ads_df["Status"] = ads_df["_last_contact_dt"].apply(_cadencia_status)

    def _ads_suggested_booking(row):
        """
        Weekly budget estimate: 8% of projected monthly GMV / 4 weeks.
        GMV source priority:
          1. Current GMV sheet (GMV acumulado MTD × factor proyección al mes)
          2. Fallback: Last GMV ARS del Growth OS (columna _gmv)
        Projection factor = 4 / semana_del_mes (ej: semana 2 → ×2, semana 3 → ×1.33)
        """
        brand_id = normalize_brand_id(row.get("_id", ""))
        # 1) Current GMV proyectado
        gmv = _current_gmv_map.get(brand_id, 0)
        # 2) Fallback: Last GMV ARS del Growth OS (sin proyección, ya es mensual)
        if gmv <= 0:
            gmv = to_number(row.get("_gmv"), 0)
        if gmv <= 0:
            return 0
        return _round_budget_up_ars(gmv * 0.08 / 4, step=1000)

    ads_df["_suggested_booking_ars"] = ads_df.apply(_ads_suggested_booking, axis=1)
    ads_df["_suggested_booking_usd"] = ads_df["_suggested_booking_ars"] / ARS_PER_USD
    # Rev Proj = booking semanal estimado × semanas restantes del mes × 90% (umbral comisión).
    # Refleja lo que puede generar este brand si entra hoy, hasta el cierre del mes.
    ads_df["_revenue_proj_weekly_usd"] = ads_df["_suggested_booking_usd"] * 0.90
    ads_df["_revenue_proj_monthly_usd"] = ads_df["_revenue_proj_weekly_usd"] * weeks_left

    # ── Opportunity score comercial (Opción C) ──────────────────────────────
    # Mide qué tan rentable/probable es cerrar el brand HOY, no qué tan
    # "sana" está su operación. Reemplaza el _opportunity_score genérico
    # (que pondera GMV/CR/Pro/AOV de salud) por uno orientado a revenue real.
    #   - Rev Proj mensual (60%): revenue que entra al target si se cierra hoy.
    #   - Probabilidad de cierre por status comercial (30%): un brand
    #     "En negociación"/"Interesado" cierra más rápido que uno sin contacto.
    #   - GMV actual proyectado (10%): tamaño del aliado, como tie-breaker.
    _ads_status_prob_map = {
        "🏆": 1.0,   # Deal cerrado / acuerdo alcanzado
        "⏳": 0.9,   # En negociación
        "✅": 0.6,   # Sin novedad (contacto activo, sin avance)
        "👻": 0.2,   # Sin contacto / no contesta
        "🚀": 0.5,   # Activo (ya corriendo, baja prioridad de "cierre")
        "❌": 0.0,   # Rechazado
    }
    ads_df["_ads_status_norm"] = ads_df["_commercial_status_raw"].apply(_normalize_commercial_status)
    ads_df["_ads_status_prob"] = ads_df["_ads_status_norm"].map(_ads_status_prob_map).fillna(0.2)

    ads_df["_ads_current_gmv_ars"] = ads_df["_id"].apply(
        lambda x: _current_gmv_map.get(normalize_brand_id(x), 0)
    )

    _rev_proj_norm = _normalize_series(ads_df["_revenue_proj_monthly_usd"])
    _status_prob_norm = _normalize_series(ads_df["_ads_status_prob"])
    _gmv_current_norm = _normalize_series(ads_df["_ads_current_gmv_ars"])

    ads_df["_opportunity_score"] = (
        _rev_proj_norm * 0.60
        + _status_prob_norm * 0.30
        + _gmv_current_norm * 0.10
    )

    ads_df = ads_df.sort_values(
        by=["_opp_group", "_opportunity_score"],
        ascending=[True, False],
    ).reset_index(drop=True)

    # ── Cumulative target coverage ────────────────────────────────────────────
    # Gap real = lo que falta después de activo + proyectado (booking × 90%)
    ads_gap_usd = max(ads_target_usd - ads_result_from_sheet - ads_projected_usd, 0) if ads_target_usd > 0 else 0
    ads_df["_cumrev_usd"] = ads_df["_revenue_proj_monthly_usd"].cumsum()

    def _ads_target_pct(rev_proj_this_brand):
        # Puntos porcentuales que aportaría este brand individualmente sobre el target
        if ads_target_usd <= 0 or rev_proj_this_brand <= 0:
            return "-"
        pp = (rev_proj_this_brand / ads_target_usd) * 100
        return f"+{pp:.1f} pp"

    def _ads_closes_at(idx):
        """Returns a label if this brand crosses the target threshold."""
        if ads_target_usd <= 0:
            return ""
        cum = ads_df.loc[:idx, "_revenue_proj_monthly_usd"].sum()
        prev = ads_df.loc[:idx - 1, "_revenue_proj_monthly_usd"].sum() if idx > 0 else 0
        if prev < ads_gap_usd <= cum:
            return "🎯 cierra aquí"
        return ""

    ads_df["Rank"] = ads_df.index + 1

    # Pipeline total = sum of all projected monthly revenue from Opp List
    ads_pipeline_usd = ads_df["_revenue_proj_monthly_usd"].sum()

    # ── Progress bar ADS ─────────────────────────────────────────────────────
    st.markdown("## 🟠 ADS")
    if ads_target_usd > 0:
        _render_target_progress_bar(
            label=f"ADS Revenue Target · USD {fmt_number(ads_target_usd)}",
            active_usd=ads_result_from_sheet,
            projected_usd=ads_projected_usd,
            pipeline_usd=min(ads_pipeline_usd, max(ads_gap_usd, 0)),
            target_usd=ads_target_usd,
        )
        # "Close-out" line
        brands_needed = 0
        running = 0
        for _, r in ads_df.iterrows():
            running += r["_revenue_proj_monthly_usd"]
            brands_needed += 1
            if running >= ads_gap_usd:
                break
        if ads_gap_usd > 0:
            st.markdown(
                f"<div style='font-size:13px; color:{COLORS['accent']}; font-weight:700; margin-bottom:12px;'>"
                f"⚡ Cierra los top <b>{brands_needed}</b> brands de esta lista para cubrir el gap de {fmt_usd(ads_gap_usd)} "
                f"({'ya cubierto por activos 🎉' if ads_result_from_sheet >= ads_target_usd else ''})</div>",
                unsafe_allow_html=True,
            )

    st.caption(
        f"Acquire = inactive en Current ADS · Upselling = activo con ROI > 4.5x · "
        f"GMV base: Current GMV × proyección semana {_week_of_month_opp}/4 (factor ×{_gmv_projection_factor:.2f}) "
        f"— fallback a Last GMV ARS si la marca no figura en Current GMV · "
        f"Rev Proj = 90% del booking estimado × semanas restantes del mes · "
        f"% Target = cobertura acumulada sobre el target mensual ADS."
    )

    def _gmv_source_label(row):
        """Indica si el GMV usado viene de Current GMV (proyectado) o del Growth OS (histórico)."""
        brand_id = normalize_brand_id(row.get("_id", ""))
        if brand_id in _current_gmv_map and _current_gmv_map[brand_id] > 0:
            return f"📡 ARS {fmt_number(_current_gmv_map[brand_id])} (W{_week_of_month_opp}→mes)"
        gmv_fallback = to_number(row.get("_gmv"), 0)
        if gmv_fallback > 0:
            return f"📁 ARS {fmt_number(gmv_fallback)} (histórico)"
        return "-"

    ads_view = pd.DataFrame({
        "Rank":               ads_df["Rank"].apply(_format_rank),
        "Opp":                ads_df["Opp"],
        "ID":                 ads_df["_id"].apply(_format_id),
        "Name":               ads_df["_name"],
        "Status":             ads_df["Status"],
        "GMV base (fuente)":  ads_df.apply(_gmv_source_label, axis=1),
        "Booking/sem (est)":  ads_df["_suggested_booking_ars"].apply(
            lambda x: f"ARS {fmt_number(x)}" if x > 0 else "-"
        ),
        "Rev Proj (mes)":     ads_df["_revenue_proj_monthly_usd"].apply(
            lambda x: fmt_usd(x) if x > 0 else "-"
        ),
        "% Target acum":      [
            _ads_target_pct(ads_df.loc[i, "_revenue_proj_monthly_usd"]) for i in ads_df.index
        ],
        "Cierre":             [_ads_closes_at(i) for i in ads_df.index],
    })
    _render_light_table(ads_view, height=380)

    # ═══════════════════════════════════════════════════════════════════════════
    # MARKDOWN SEGMENT
    # ═══════════════════════════════════════════════════════════════════════════
    data["_md_current_active"] = data["_id"].apply(
        lambda x: md_metrics.get(normalize_brand_id(x), {}).get("active", False)
    )
    data["_md_current_roi"] = data["_id"].apply(
        lambda x: md_metrics.get(normalize_brand_id(x), {}).get("roi", 0)
    )
    data["_md_current_gmv_usd"] = data["_id"].apply(
        lambda x: md_metrics.get(normalize_brand_id(x), {}).get("gmv_usd", 0)
    )

    # ── Penetración sana MD ──────────────────────────────────────────────────
    # Definición: % del GMV mensual del brand que pasa por MD (promo activa).
    # Rango sano: 10–20% (conservador).
    # Penetración actual  = gmv_md_corriente / gmv_mensual_brand
    # Penetración objetivo = 10% (floor sano) — 20% (techo sano)
    # Gap = max(0, 10% - actual) → cuánto le falta para entrar al rango

    MD_PENE_LOW  = 0.10   # 10% floor
    MD_PENE_HIGH = 0.20   # 20% ceiling

    def _md_penetration(row):
        gmv_brand_ars = to_number(row.get("_gmv"), 0)
        gmv_brand_usd = gmv_brand_ars / ARS_PER_USD if gmv_brand_ars > 0 else 0
        gmv_md_usd    = to_number(row.get("_md_current_gmv_usd"), 0)

        if gmv_brand_usd <= 0:
            return {
                "actual_pct": None,
                "target_low_usd":  0,
                "target_high_usd": 0,
                "gap_usd":         0,
                "label":           "Sin GMV",
                "status":          "unknown",
            }

        actual_pct    = gmv_md_usd / gmv_brand_usd if gmv_md_usd > 0 else 0
        target_low    = gmv_brand_usd * MD_PENE_LOW
        target_high   = gmv_brand_usd * MD_PENE_HIGH
        gap_usd       = max(target_low - gmv_md_usd, 0)

        if actual_pct == 0:
            status = "sin_promo"
            label  = f"0% · objetivo {MD_PENE_LOW*100:.0f}–{MD_PENE_HIGH*100:.0f}%"
        elif actual_pct < MD_PENE_LOW:
            status = "bajo"
            label  = f"{actual_pct*100:.1f}% · bajo ({MD_PENE_LOW*100:.0f}% mín)"
        elif actual_pct <= MD_PENE_HIGH:
            status = "sano"
            label  = f"{actual_pct*100:.1f}% ✅ en rango"
        else:
            status = "alto"
            label  = f"{actual_pct*100:.1f}% ⚠️ sobre techo"

        return {
            "actual_pct":      actual_pct,
            "target_low_usd":  target_low,
            "target_high_usd": target_high,
            "gap_usd":         gap_usd,
            "label":           label,
            "status":          status,
        }

    # ── Pre-calcular penetración sobre todos los brands activos ─────────────
    # Necesaria antes de definir los grupos para que Upsell Urgente pueda
    # identificarse por penetración, no solo por ROI.
    data["_pene_data"] = data.apply(_md_penetration, axis=1)
    data["_pene_status"] = data["_pene_data"].apply(lambda x: x["status"])

    # ── Grupos MD ────────────────────────────────────────────────────────────
    # Orden de prioridad en tabla (sort ASC por _opp_group):
    #   0 = 🔧 Upsell Urgente  → activo + penetración < 10%  (problema ya presente, hay que llamar)
    #   1 = ⚡ Upselling        → activo + ROI > 3.2x + penetración ≥ 10% (escalar lo que funciona)
    #   2 = 🏆 Acquire          → inactivo (nuevo negocio)

    md_upsell_urgente = (
        data["_md_current_active"] &
        data["_pene_status"].isin(["bajo", "sin_promo", "cero"])
    )
    md_upselling = (
        data["_md_current_active"] &
        (data["_md_current_roi"] > 3.2) &
        ~md_upsell_urgente
    )
    md_acquire = ~data["_md_current_active"]

    md_df = data[md_upsell_urgente | md_upselling | md_acquire].copy()

    def _md_opp_group(row):
        if not row["_md_current_active"]:
            return 2  # Acquire
        if row["_pene_status"] in ("bajo", "sin_promo", "cero", "unknown"):
            return 0  # Upsell Urgente — penetración < 10%
        return 1      # Upselling — activo, penetración sana, ROI > 3.2x

    md_df["_opp_group"] = md_df.apply(_md_opp_group, axis=1)
    md_df["Opp"] = md_df["_opp_group"].map({
        0: "🔧 Upsell Urgente",
        1: "⚡ Upselling",
        2: "🏆 Acquire",
    })
    # Status = cadencia de contacto (🟢/🟡/🟠/🔴) — mismo lenguaje que Salud de Cartera
    md_df["_last_contact_dt"] = md_df.apply(
        lambda r: get_last_contact_dt(r.get("_id", ""), r.get("_name", ""), _opp_prod_map, _opp_meta_map), axis=1
    )
    md_df["Status"] = md_df["_last_contact_dt"].apply(_cadencia_status)

    def _md_opp_type(row):
        if row["_opp_group"] == 0:
            pene_status = row.get("_pene_status", "")
            if pene_status == "sin_promo":
                return "Urgente · 0% penetración"
            return "Urgente · penetración < 10%"
        if row["_opp_group"] == 1:
            roi = to_number(row.get("_md_current_roi", 0), 0)
            if roi > 5.0:
                return "Upselling · Arquitectura"
            return "Upselling · Profundidad"
        return "Adquisición promocional"

    md_df["MD Strategy"] = md_df.apply(_md_opp_type, axis=1)

    pene_data = md_df.apply(_md_penetration, axis=1)
    md_df["_pene_label"]     = pene_data.apply(lambda x: x["label"])
    md_df["_pene_status"]    = pene_data.apply(lambda x: x["status"])
    md_df["_pene_gap_usd"]   = pene_data.apply(lambda x: x["gap_usd"])
    md_df["_pene_low_usd"]   = pene_data.apply(lambda x: x["target_low_usd"])
    md_df["_pene_high_usd"]  = pene_data.apply(lambda x: x["target_high_usd"])

    # Recomendación de paso siguiente por brand — aparece en columna nueva
    def _md_next_step(row):
        group  = row.get("_opp_group", 2)
        pstatus = row.get("_pene_status", "")
        roi    = to_number(row.get("_md_current_roi", 0), 0)
        gap    = to_number(row.get("_pene_gap_usd", 0), 0)
        if group != 0:
            return "—"
        if pstatus == "sin_promo":
            return f"Promo activa sin penetración · Revisar config · Gap: {fmt_usd(gap)}"
        if roi < 2.0:
            return f"ROI bajo ({roi:.1f}x) + pene < 10% · Renegociar descuento · Gap: {fmt_usd(gap)}"
        if roi >= 3.5:
            return f"ROI OK ({roi:.1f}x) pero alcance bajo · Ampliar promo · Gap: {fmt_usd(gap)}"
        return f"Penetración baja · Revisar promo con brand · Gap: {fmt_usd(gap)}"

    md_df["Próximo Paso"] = md_df.apply(_md_next_step, axis=1)

    # Mantenemos _gmv_proj_monthly_usd como el GMV objetivo al 15% (mid-range)
    # para que la barra de progreso del target siga funcionando si está configurada.
    md_df["_gmv_proj_monthly_usd"] = md_df["_gmv"].apply(
        lambda x: (to_number(x, 0) * 0.15) / ARS_PER_USD
    )

    md_df = md_df.sort_values(
        by=["_opp_group", "_opportunity_score"],
        ascending=[True, False],
    ).reset_index(drop=True)

    # ── MD barra de progreso: activo vs target ───────────────────────────────
    # Activo  = col E fila Total de Current MD + Current MD Pro (solo cuando E > 0)
    # Target  = col D fila Total × % target de Earnings (MD=col F, MD Pro=col G)
    md_gap_usd = max(md_combined_target_usd - active_md_combined_usd, 0) if md_combined_target_usd > 0 else 0
    # Pipeline desde Opp List: suma de rango medio (15%) de cada brand recomendado
    md_df["_gmv_proj_monthly_usd"] = md_df["_gmv"].apply(
        lambda x: (to_number(x, 0) * 0.15) / ARS_PER_USD
    )
    md_df["_cum_gmv_usd"] = md_df["_gmv_proj_monthly_usd"].cumsum()

    def _md_closes_at(idx):
        if md_combined_target_usd <= 0:
            return ""
        cum  = md_df.loc[:idx, "_gmv_proj_monthly_usd"].sum()
        prev = md_df.loc[:idx - 1, "_gmv_proj_monthly_usd"].sum() if idx > 0 else 0
        if prev < md_gap_usd <= cum:
            return "🎯 cierra aquí"
        return ""

    md_df["Rank"] = md_df.index + 1
    md_pipeline_usd = md_df["_gmv_proj_monthly_usd"].sum()

    # ── Progress bar MD ──────────────────────────────────────────────────────
    if md_combined_target_usd > 0:
        st.markdown("## 🔵 MARKDOWN")
        pene_md_pct     = _md_totals["markdown_pct"] * 100      # col F total row Current MD
        pene_mdpro_pct  = _md_pro_totals["markdown_pct"] * 100  # col F total row Current MD pro
        _label_md = (
            f"MD GMV · MD {pene_md_pct:.2f}% (target {md_target_pct*100:.2f}%) + "
            f"MD Pro {pene_mdpro_pct:.2f}% (target {md_pro_target_pct*100:.2f}%) · "
            f"Target combinado {fmt_usd(md_combined_target_usd)}"
        )
        _render_target_progress_bar(
            label=_label_md,
            active_usd=active_md_combined_usd,
            pipeline_usd=min(md_pipeline_usd, md_gap_usd),
            target_usd=md_combined_target_usd,
            color_active="#1B3F8B",
            color_pipeline="#FF7124",
        )
        st.markdown(
            f"<div style='font-size:12px; color:{COLORS['muted']}; margin-bottom:10px;'>"
            f"📊 MD activo: <b style='color:{COLORS['intel']};'>{fmt_usd(active_md_gmv_usd)}</b> ({pene_md_pct:.2f}%)"
            f" &nbsp;·&nbsp; MD Pro activo: <b style='color:{COLORS['intel']};'>{fmt_usd(active_md_pro_gmv_usd)}</b> ({pene_mdpro_pct:.2f}%)"
            f" &nbsp;·&nbsp; GMV base (col D): <b>{fmt_usd(md_gmv_total_usd)}</b></div>",
            unsafe_allow_html=True,
        )

        # ── Aviso umbral de comisión MD (90% del target MD) ──────────────────
        # La comisión NO depende del target combinado (MD+MD Pro) sino de que
        # la penetración MD sola alcance al menos el 90% del target MD.
        if md_commission_paid:
            st.markdown(
                f"<div style='font-size:12px; color:#7ED321; font-weight:700; margin-bottom:10px;'>"
                f"✅ Comisión MD habilitada · penetración MD {pene_md_pct:.2f}% es "
                f"{md_commission_pct_of_target*100:.1f}% del target MD ({md_target_pct*100:.2f}%) "
                f"— ≥ {MD_COMMISSION_THRESHOLD_PCT*100:.0f}% requerido</div>",
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f"<div style='font-size:12px; color:#FF7124; font-weight:700; margin-bottom:10px;'>"
                f"⚠️ Comisión MD NO habilitada · penetración MD {pene_md_pct:.2f}% es "
                f"{md_commission_pct_of_target*100:.1f}% del target MD ({md_target_pct*100:.2f}%) "
                f"— falta {fmt_usd(md_commission_gap_usd)} de MD activo para llegar al "
                f"{MD_COMMISSION_THRESHOLD_PCT*100:.0f}% del target</div>",
                unsafe_allow_html=True,
            )
        if md_gap_usd > 0:
            md_brands_needed = 0
            running_md = 0
            for _, r in md_df.iterrows():
                running_md += r["_gmv_proj_monthly_usd"]
                md_brands_needed += 1
                if running_md >= md_gap_usd:
                    break
            st.markdown(
                f"<div style='font-size:13px; color:{COLORS['intel']}; font-weight:700; margin-bottom:12px;'>"
                f"⚡ Cierra los top <b>{md_brands_needed}</b> brands para cubrir el gap de {fmt_usd(md_gap_usd)}</div>",
                unsafe_allow_html=True,
            )
    else:
        st.markdown("## 🔵 MARKDOWN")
        st.info("No se pudo leer la fila Total de Current MD. Verificá que el sheet esté cargado.")

    st.caption(
        "Upsell Urgente = activo con penetración < 10% · "
        "Upselling = activo + ROI > 3.2x + penetración ≥ 10% · "
        "Acquire = inactivo en Current MD · "
        "Penetración = GMV en promo ÷ GMV mensual del brand · "
        "Gap al 10% = cuánto GMV falta para entrar al rango mínimo."
    )

    md_view = pd.DataFrame({
        "Rank":            md_df["Rank"].apply(_format_rank),
        "Opp":             md_df["Opp"],
        "MD Strategy":     md_df["MD Strategy"],
        "ID":              md_df["_id"].apply(_format_id),
        "Name":            md_df["_name"],
        "Status":          md_df["Status"],
        "Penetración MD":  [md_df.loc[i, "_pene_label"] for i in md_df.index],
        "Objetivo (USD)":  md_df.apply(
            lambda r: f"{fmt_usd(r['_pene_low_usd'])} – {fmt_usd(r['_pene_high_usd'])}"
            if r["_pene_low_usd"] > 0 else "-", axis=1
        ),
        "Gap al 10%":      md_df["_pene_gap_usd"].apply(
            lambda x: fmt_usd(x) if x > 0 else "—"
        ),
        "Próximo Paso":    md_df["Próximo Paso"],
    })
    _render_light_table(md_view, height=380)

    # ═══════════════════════════════════════════════════════════════════════════
    # CHURN SEGMENT — una fila por STORE (no por brand)
    # Fuente: hoja Current Churn leída directamente, una entrada por store.
    # ═══════════════════════════════════════════════════════════════════════════
    _raw_churn_df = load_current_churn_raw_df()
    if not _raw_churn_df.empty:
        _cc_id_col   = _first_existing_col(_raw_churn_df, ["country_brand_id", "brand id", "brand_id", "id"])
        _cc_sta_col  = _first_existing_col(_raw_churn_df, ["estado actual", "estado", "status", "churn status"])
        _cc_name_col = _first_existing_col(_raw_churn_df, ["name", "brand name", "store name", "nombre"])
        _cc_gmv_col  = _first_existing_col(_raw_churn_df, ["gmv", "gmv ars", "last gmv ars", "gmv usd"])
    else:
        _cc_id_col = _cc_sta_col = _cc_name_col = _cc_gmv_col = None

    if not _raw_churn_df.empty and _cc_id_col and _cc_sta_col:
        # Build a GMV map from data (Growth OS) for enrichment
        _gmv_map = {normalize_brand_id(r["_id"]): r["_gmv"] for _, r in data.iterrows() if r.get("_id")}
        _status_map = {normalize_brand_id(r["_id"]): r["_commercial_status_raw"] for _, r in data.iterrows() if r.get("_id")}
        _name_map = {normalize_brand_id(r["_id"]): r["_name"] for _, r in data.iterrows() if r.get("_id")}

        churn_rows = []
        for _, crow in _raw_churn_df.iterrows():
            bid = normalize_brand_id(crow.get(_cc_id_col, ""))
            sta = clean(crow.get(_cc_sta_col, ""), "").strip()
            if not bid or not sta:
                continue
            # Nombre: primero de la hoja Current Churn, fallback Growth OS
            if _cc_name_col:
                name_val = clean(crow.get(_cc_name_col, ""), "")
            else:
                name_val = ""
            if not name_val:
                name_val = _name_map.get(bid, bid)
            # GMV: primero de la hoja Current Churn, fallback Growth OS
            if _cc_gmv_col:
                gmv_val = to_number(crow.get(_cc_gmv_col, 0), 0)
            else:
                gmv_val = 0
            if gmv_val <= 0:
                gmv_val = to_number(_gmv_map.get(bid, 0), 0)
            gmv_usd = gmv_val / ARS_PER_USD if gmv_val > 0 else 0
            comm_status_raw = _status_map.get(bid, "")
            churn_rows.append({
                "_bid": bid,
                "_name": name_val,
                "_status_raw": comm_status_raw,
                "_churn_raw": sta,
                "_gmv_usd": gmv_usd,
            })

        churn_df = pd.DataFrame(churn_rows)

        # ── Orden de retención: Off > W3 > W2 > W1, luego GMV descendente ────
        # Regla de gestión: una marca Off ya dejó de facturar — es rescate
        # inmediato y encabeza la lista (ordenada por GMV histórico: cuánta
        # plata se está yendo). Después la escalera de riesgo W3 → W1.
        # Nota: get_brand_churn_map usa la jerarquía inversa a propósito —
        # allá se elige el "peor estado ACTIVO" para mostrar el status de una
        # marca multi-tienda sin marcarla Off entera por un local cerrado.
        _sev = {"Off": 5, "W3": 4, "W2": 3, "W1": 2}
        churn_df["_sev"] = churn_df["_churn_raw"].apply(lambda x: _sev.get(x, 0))
        churn_df = churn_df.sort_values(by=["_sev", "_gmv_usd"], ascending=[False, False]).reset_index(drop=True)
        churn_df["Rank"] = churn_df.index + 1

        churn_view = pd.DataFrame({
            "Rank":          churn_df["Rank"].apply(_format_rank),
            "Churn Status":  churn_df["_churn_raw"].apply(_churn_label_with_emoji),
            "ID":            churn_df["_bid"].apply(_format_id),
            "Name":          churn_df["_name"],
            "Status":        churn_df["_bid"].apply(
                lambda bid: _cadencia_status(get_last_contact_dt(bid, _name_map.get(bid, ""), _opp_prod_map, _opp_meta_map))
            ),
            "GMV at Risk":   churn_df["_gmv_usd"].apply(lambda x: fmt_usd(x) if x > 0 else "-"),
        })
        total_gmv_at_risk = churn_df["_gmv_usd"].sum()
        n_stores = len(churn_df)
    else:
        # Fallback: usar data como antes
        churn_mask = ~data["_churn"].apply(_is_on_status)
        churn_df = data[churn_mask].copy()
        churn_df["_gmv_at_risk_usd"] = churn_df["_gmv"].apply(lambda x: to_number(x, 0) / ARS_PER_USD)
        churn_df = churn_df.sort_values(by="_opportunity_score", ascending=False).reset_index(drop=True)
        churn_df["Rank"] = churn_df.index + 1
        churn_view = pd.DataFrame({
            "Rank":          churn_df["Rank"].apply(_format_rank),
            "Churn Status":  churn_df["_churn"].apply(_churn_label_with_emoji),
            "ID":            churn_df["_id"].apply(_format_id),
            "Name":          churn_df["_name"],
            "Status":        churn_df["_id"].apply(
                lambda bid: _cadencia_status(get_last_contact_dt(bid, _name_map.get(normalize_brand_id(bid), ""), _opp_prod_map, _opp_meta_map))
            ),
            "GMV at Risk":   churn_df["_gmv_at_risk_usd"].apply(lambda x: fmt_usd(x) if x > 0 else "-"),
        })
        total_gmv_at_risk = churn_df["_gmv_at_risk_usd"].sum()
        n_stores = len(churn_df)

    st.markdown("## 🔴 CHURN")

    # ── Barra de distribución de severidad (On/W1/W2/W3/Off) ─────────────────
    # "On" = total de stores del portfolio (Asignacion Junio) que NO aparecen
    # en Current Churn con un estado W1/W2/W3/Off. El total de la barra es el
    # universo completo del portfolio, no solo las filas de Current Churn.
    if _cc_sta_col is not None:
        _all_statuses = _raw_churn_df[_cc_sta_col].apply(lambda x: clean(x, "").strip()) if not _raw_churn_df.empty else pd.Series([], dtype=str)
        _all_statuses = _all_statuses[_all_statuses != ""]
        _dist_counts = _all_statuses.value_counts().to_dict()
        # Normalizar claves a W1/W2/W3/Off (estados "en churn")
        _dist_counts_norm = {}
        for k, v in _dist_counts.items():
            kn = k.strip()
            if kn in ("W1", "W2", "W3", "Off"):
                _dist_counts_norm[kn] = _dist_counts_norm.get(kn, 0) + v

        _n_churned = sum(_dist_counts_norm.values())
        _asignacion_df = load_asignacion_activa()
        _portfolio_total = len(_asignacion_df)

        if _portfolio_total > 0:
            _dist_counts_norm["On"] = max(_portfolio_total - _n_churned, 0)
            _dist_total = _portfolio_total
        else:
            # Fallback: sin Asignacion Junio, usar solo lo que hay en Current Churn
            _dist_total = sum(_dist_counts_norm.values())

        _render_churn_distribution_bar(_dist_counts_norm, _dist_total)

    if total_gmv_at_risk > 0:
        st.markdown(
            f"<div style='font-size:13px; color:{COLORS['danger']}; font-weight:700; margin-bottom:12px;'>"
            f"⚠️ GMV total en riesgo por churn: <b>{fmt_usd(total_gmv_at_risk)}</b> — "
            f"{n_stores} store{'s' if n_stores != 1 else ''} fuera</div>",
            unsafe_allow_html=True,
        )
    st.caption("Ordenado por severidad (W3 → W2 → W1 → Off) y luego GMV. Cada fila = una store individual de Current Churn.")
    _render_light_table(churn_view, height=380)



def _days_since_timestamp(value):
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return int((pd.Timestamp(date.today()).normalize() - parsed.normalize()).days)


def _followup_visual_status(manual_status, last_comment_dt):
    """Keeps Sabas-selected emoji/status until the relationship is cold again after 22 days."""
    days = _days_since_timestamp(last_comment_dt)
    if days is not None and days > 22:
        return "Frío ❄️"
    return _status_label_from_value(manual_status, default="OFF 😴")


def _cadencia_status(last_contact_dt):
    """
    Returns cadence-based status matching the Salud de Cartera / 21-day heat map.
    Used as the universal Status column in Opportunity List and Follow-Up List.
    Source: last contact date (from Productivity sheet or comments CSV).
      🟢 Activo    → 0–10 days
      🟡 Cadencia  → 11–15 days
      🟠 Alerta    → 16–21 days
      🔴 Fría      → > 21 days or never contacted
    """
    days = _days_since_timestamp(last_contact_dt)
    if days is None:
        return "🔴 Sin contacto"
    if days <= 10:
        return "🟢 Activo"
    if days <= 15:
        return "🟡 Cadencia"
    if days <= 21:
        return "🟠 Alerta"
    return "🔴 Fría"


def _format_followup_last_update(value):
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return "Sin comentario"
    return parsed.strftime("%Y-%m-%d %H:%M")


def get_last_comment_meta_map(limit=2):
    comments = _load_comments_df()
    if comments.empty:
        return {}
    comments = comments.sort_values(by="_dt", ascending=True, na_position="last")
    result = {}
    for bid, group in comments.groupby("brand_id"):
        group = group.copy().sort_values(by="_dt", ascending=True, na_position="last")
        last_dt = group["_dt"].dropna().iloc[-1] if not group["_dt"].dropna().empty else pd.NaT
        statuses = [clean(x, "").strip() for x in group.get("opportunity_status", pd.Series([], dtype=str)).tolist() if clean(x, "").strip()]

        # Notes: include ALL non-ghost comments including [Auto] AI-generated ones.
        # Ghost = opportunity_status contains "ghost" OR comment starts with 👻 emoji.
        # [Auto] AI-generated notes are real notes and must always be included.
        def _is_ghost_comment(row):
            txt = clean(row.get("comment", ""), "").strip().lower()
            st_val = clean(row.get("opportunity_status", ""), "").strip().lower()
            return "ghost" in st_val or txt.startswith("👻")

        all_comments = [clean(r.get("comment", ""), "").strip() for _, r in group.iterrows() if clean(r.get("comment", ""), "").strip()]
        # Include [Auto] comments — they are AI-generated transcription notes, not ghosts.
        real_comments = [clean(r.get("comment", ""), "").strip() for _, r in group.iterrows()
                         if clean(r.get("comment", ""), "").strip() and not _is_ghost_comment(r)]

        notes_source = real_comments if real_comments else all_comments

        # Store brand_name for cross-reference by name (avoids brand_id normalization mismatches)
        brand_name_raw = ""
        if not group.empty and "brand_name" in group.columns:
            _bnames = group["brand_name"].dropna().astype(str).str.strip()
            _bnames = _bnames[_bnames != ""]
            if not _bnames.empty:
                brand_name_raw = _bnames.iloc[-1]

        _meta_entry = {
            "last_dt": last_dt,
            "last_update": _format_followup_last_update(last_dt),
            "notes": " | ".join(notes_source[-limit:]) if notes_source else "-",
            "status": statuses[-1] if statuses else "OFF 😴",
            "brand_name": brand_name_raw,
        }
        result[normalize_brand_id(bid)] = _meta_entry
        # Also index by normalized brand name for fallback resolution
        if brand_name_raw:
            _bname_key = norm_text(brand_name_raw)
            if _bname_key and _bname_key not in result:
                result[_bname_key] = _meta_entry
    return result


@st.cache_data(ttl=3000, show_spinner=False)
def get_productivity_effective_rows(excel_path):
    """
    Reads the Productivity sheet and returns a DataFrame with:
      _date_k -> datetime from column K (Date)
      _week_j -> datetime from column J (Week)
      _effective -> bool, True unless col F (Fase) == "Aliado no contactado"
    Used for HOY / SEMANA / MES contact counters in Follow-Up List.
    Col F (idx 5)  = Fase
    Col J (idx 9)  = Week
    Col K (idx 10) = Date
    """
    if not os.path.exists(excel_path):
        return pd.DataFrame(columns=["_date_k", "_week_j", "_effective"])
    raw = _load_productivity_sheet_raw(excel_path)
    if raw.empty:
        return pd.DataFrame(columns=["_date_k", "_week_j", "_effective"])
    raw = raw.copy()

    raw.columns = [str(c).strip() for c in raw.columns]
    cols_lower = [c.lower() for c in raw.columns]

    fase_col = next((raw.columns[i] for i, c in enumerate(cols_lower) if c == "fase"), None)
    if not fase_col and len(raw.columns) > 5:
        fase_col = raw.columns[5]

    if len(raw.columns) < 11:
        return pd.DataFrame(columns=["_date_k", "_week_j", "_effective"])

    week_col = raw.columns[9]   # J
    date_col = raw.columns[10]  # K

    out = pd.DataFrame()
    out["_date_k"] = pd.to_datetime(raw[date_col], errors="coerce")
    out["_week_j"] = pd.to_datetime(raw[week_col], errors="coerce")

    if fase_col and fase_col in raw.columns:
        _fase = raw[fase_col].astype(str).str.strip().str.lower()
        out["_effective"] = ~_fase.str.contains("aliado no contactado", case=False, na=False)
    else:
        out["_effective"] = True

    return out



def get_productivity_last_contact_map(excel_path):
    """
    Reads the Productivity sheet and returns a dict:
        { normalized_brand_name -> most_recent_contact_date (pd.Timestamp) }
    Column K (index 10) = contact date
    Column Q (index 16) = brand name
    Groups by brand name and keeps the most recent date.
    """
    if not os.path.exists(excel_path):
        return {}
    raw = _load_productivity_sheet_raw(excel_path)
    if raw.empty:
        return {}
    raw = raw.copy()

    cols = list(raw.columns)
    if len(cols) < 17:
        return {}

    date_col  = cols[10]   # K
    brand_col = cols[16]   # Q

    sub = raw[[date_col, brand_col]].copy()
    sub.columns = ["_date", "_brand"]
    sub["_date"]  = pd.to_datetime(sub["_date"],  errors="coerce")
    sub["_brand"] = sub["_brand"].apply(lambda x: str(x).strip().lower() if pd.notna(x) else "")
    sub = sub[sub["_date"].notna() & (sub["_brand"] != "")]

    result = {}
    for brand, group in sub.groupby("_brand"):
        latest = group["_date"].max()
        if pd.notna(latest):
            if hasattr(latest, "tzinfo") and latest.tzinfo is not None:
                latest = latest.tz_localize(None)
            result[brand] = latest
    return result


def get_productivity_levers_for_brand(excel_path, brand_name, month=None):
    """
    Reads the Productivity sheet and returns a summary of levers worked
    for the given brand in the current month (or specified month).

    Columns used:
      K (idx 10) = Date  |  Q (idx 16) = Brand name
      Lever cols (binary SI/NO): Markdown, Ads, Conectividad, Catálogo,
          Cancelaciones, DR, Tiempos, Pains del aliado, Churn, On Hold
      Multi-value: Ajustes Catálogo
      Extra: Tipo Ads, ¿Se aceptó lo ofrecido?, Fase

    Priority logic (caller side):
      1. [Auto] transcript comment → always wins, don't call this
      2. This function → if productivity rows exist for the brand this month
      3. Regular CSV / Excel / meta → fallback

    Returns dict or None if brand not found / sheet unavailable.
    """
    if not os.path.exists(excel_path):
        return None
    raw = _load_productivity_sheet_raw(excel_path)
    if raw.empty:
        return None
    raw = raw.copy()

    raw.columns = [str(c).strip() for c in raw.columns]
    cols = list(raw.columns)
    if len(cols) < 17:
        return None

    date_col  = cols[10]   # K = Date
    brand_col = cols[16]   # Q = Brand name

    brand_key = str(brand_name).strip().lower() if brand_name else ""
    if not brand_key:
        return None

    raw["_brand_norm"] = raw[brand_col].apply(
        lambda x: str(x).strip().lower() if pd.notna(x) else ""
    )
    brand_rows = raw[raw["_brand_norm"] == brand_key].copy()
    if brand_rows.empty:
        return None

    # Filter to current month (fallback: all rows for this brand)
    if month is None:
        today = date.today()
        month = (today.year, today.month)

    brand_rows["_date"] = pd.to_datetime(brand_rows[date_col], errors="coerce")
    month_rows = brand_rows[
        (brand_rows["_date"].dt.year == month[0]) &
        (brand_rows["_date"].dt.month == month[1])
    ].copy()

    if month_rows.empty:
        return None  # Nothing logged this month for this brand

    # ── Helpers ────────────────────────────────────────────────────────────────
    def col_si(col_name):
        if col_name not in month_rows.columns:
            return False
        return (month_rows[col_name].astype(str).str.upper() == "SI").any()

    def col_multi_has(col_name, value):
        if col_name not in month_rows.columns:
            return False
        return month_rows[col_name].fillna("").apply(
            lambda x: value.lower() in str(x).lower()
        ).any()

    # ── Detect active levers ───────────────────────────────────────────────────
    levers = []
    for lv in ["Markdown", "Ads", "Conectividad", "Catálogo",
               "Cancelaciones", "DR", "Tiempos", "Pains del aliado",
               "Churn", "On Hold"]:
        if col_si(lv):
            levers.append(lv)

    ajustes = []
    for aj in ["Catálogo al 100%", "Igualdad en precios", "Fotos",
               "Disponibilidad del producto", "Catálogo PDF",
               "Generación de combos", "Purchasing Experience"]:
        if col_multi_has("Ajustes Catálogo", aj):
            ajustes.append(aj)
    if ajustes and "Ajustes Catálogo" not in levers:
        levers.append("Ajustes Catálogo")

    # Extra detail
    ads_tipo = ""
    if col_si("Ads") and "Tipo Ads" in month_rows.columns:
        _tipos = (
            month_rows[month_rows["Ads"].astype(str).str.upper() == "SI"]["Tipo Ads"]
            .dropna().astype(str).str.strip()
        )
        ads_tipo = ", ".join(sorted({t for t in _tipos if t.lower() not in ["nan", ""]}))

    accepted_md = False
    if col_si("Markdown") and "¿Se aceptó lo ofrecido?" in month_rows.columns:
        accepted_md = (
            month_rows["¿Se aceptó lo ofrecido?"]
            .astype(str).str.strip().str.lower() == "si"
        ).any()

    churn   = col_si("Churn")
    on_hold = col_si("On Hold")
    row_count = len(month_rows)

    # Most recent date this month
    latest_dt = month_rows["_date"].max()
    latest_str = latest_dt.strftime("%-d/%b") if pd.notna(latest_dt) else ""

    # Fase info (most recent non-null)
    fase_str = ""
    fase_col_name = next((c for c in month_rows.columns if c.lower() == "fase"), None)
    if fase_col_name:
        fases = month_rows[fase_col_name].dropna().astype(str).str.strip()
        fases = fases[~fases.str.lower().isin(["nan", ""])]
        if not fases.empty:
            fase_str = fases.iloc[-1]

    # ── Generate human-readable nota ──────────────────────────────────────────
    month_name = date(month[0], month[1], 1).strftime("%B")

    if not levers:
        nota = (
            f"{month_name}: {row_count} contacto(s) registrado(s) sin palancas "
            f"específicas. Última entrada: {latest_str}."
        )
    else:
        comercial_lvs = [l for l in levers if l in ["Markdown", "Ads", "Conectividad", "Ajustes Catálogo"]]
        operativo_lvs = [l for l in levers if l in ["Catálogo", "Cancelaciones", "DR", "Tiempos"]]
        incendio_lvs  = [l for l in levers if l in ["Pains del aliado", "Churn", "On Hold"]]

        partes = []
        if comercial_lvs:
            partes.append(f"Comercial → {', '.join(comercial_lvs)}")
        if operativo_lvs:
            partes.append(f"Operativo → {', '.join(operativo_lvs)}")
        if incendio_lvs:
            partes.append(f"⚠️ {', '.join(incendio_lvs)}")

        extras = []
        if accepted_md:
            extras.append("MD aceptado ✓")
        if ads_tipo:
            extras.append(f"Ads: {ads_tipo}")
        if ajustes:
            extras.append(f"Catálogo: {', '.join(ajustes)}")
        if fase_str and fase_str.lower() not in ["nan", ""]:
            extras.append(f"Fase: {fase_str}")

        nota = f"{month_name} ({row_count} contacto{'s' if row_count != 1 else ''} · último {latest_str}) — {' | '.join(partes)}."
        if extras:
            nota += f" [{'; '.join(extras)}]"

    return {
        "levers":       levers,
        "ajustes":      ajustes,
        "ads_tipo":     ads_tipo,
        "churn":        churn,
        "on_hold":      on_hold,
        "accepted_md":  accepted_md,
        "row_count":    row_count,
        "latest_str":   latest_str,
        "fase":         fase_str,
        "nota_generada": nota,
    }


def get_last_contact_dt(brand_id, name, prod_map=None, meta_map=None):
    """
    Unified last-contact resolver. Always returns the most recent date
    across both sources: Productivity sheet (by brand name) and Comments CSV (by brand_id).
    If both exist, returns the more recent one.
    prod_map and meta_map can be passed in to avoid reloading on every call.
    """
    if prod_map is None:
        prod_map = get_productivity_last_contact_map(EXCEL_FILE)
    if meta_map is None:
        meta_map = get_last_comment_meta_map(limit=1)

    brand_key = str(name).strip().lower() if name else ""
    prod_dt = prod_map.get(brand_key)
    comment_dt = meta_map.get(normalize_brand_id(brand_id), {}).get("last_dt")

    prod_ts = pd.to_datetime(prod_dt, errors="coerce") if prod_dt else pd.NaT
    comment_ts = pd.to_datetime(comment_dt, errors="coerce") if comment_dt else pd.NaT

    if pd.isna(prod_ts) and pd.isna(comment_ts):
        return pd.NaT
    if pd.isna(prod_ts):
        return comment_ts
    if pd.isna(comment_ts):
        return prod_ts
    return max(prod_ts, comment_ts)



def page_follow_up_list():
    render_header("Follow-Up List", "Strategic farming agenda for active commercial accounts")

    data = _prepare_growth_scored_data()

    if data.empty:
        st.error("Could not load Growth OS data or ID column.")
        return

    on_mask = data["_churn"].apply(_is_on_status)

    # Show ALL portfolio brands — not just those with active levers.
    # Brands without active Ads/MD are actually the most important to call.
    follow_df = data.copy()

    # Flag lever status for context columns
    follow_df["_has_ads"]  = follow_df["_ads"].apply(_is_active_status)
    follow_df["_has_md"]   = follow_df["_md"].apply(_is_active_status)
    follow_df["_is_on"]    = on_mask
    meta_map = get_last_comment_meta_map(limit=2)

    def meta_value(brand_id, key, default=""):
        return meta_map.get(normalize_brand_id(brand_id), {}).get(key, default)

    # ── Last contact date: Productivity sheet (col K=date, col Q=brand name) ──
    # Fallback to comments CSV if brand not found in Productivity.
    prod_map = get_productivity_last_contact_map(EXCEL_FILE)

    def _resolve_last_contact_dt(row):
        """Returns most recent contact date: Productivity first, comments fallback."""
        brand_key = str(row.get("_name", "")).strip().lower()
        if brand_key and brand_key in prod_map:
            return prod_map[brand_key]
        # fallback: comments CSV
        return meta_value(row["_id"], "last_dt", pd.NaT)

    follow_df["_last_comment_dt"] = follow_df.apply(_resolve_last_contact_dt, axis=1)

    def _fmt_last_update(dt):
        ts = pd.to_datetime(dt, errors="coerce")
        if pd.isna(ts):
            return "Sin contacto"
        return ts.strftime("%Y-%m-%d %H:%M")

    follow_df["Last Update"]    = follow_df["_last_comment_dt"].apply(_fmt_last_update)
    def _shorten_note_preview(notes, max_chars=90):
        """
        Recorta la nota completa a un preview corto para la tabla. Si es una
        nota [Auto] de Claude, usa el párrafo Resumen: (ya condensado); si no,
        toma la primera línea. El texto completo sigue disponible en Brand
        Finder (Última Nota) y en growth_os_call_history.csv — acá solo
        necesitamos una referencia rápida de una línea.
        """
        if not notes or notes == "-":
            return "-"
        text = notes.strip()
        if text.startswith("[Auto]"):
            parsed = _parse_claude_note_fields(text)
            text = parsed["resumen"] if parsed["resumen"] else text.split("\n", 1)[0].replace("[Auto]", "").strip()
        else:
            text = text.split("\n", 1)[0]
        text = " ".join(text.split())  # colapsar saltos de línea y espacios sobrantes
        return text if len(text) <= max_chars else text[:max_chars].rstrip() + "…"

    def _resolve_last_notes(row):
        """Resolves Last Notes by brand_id first, then by normalized brand name as fallback."""
        bid = normalize_brand_id(row.get("_id", ""))
        # Primary: by brand_id
        entry = meta_map.get(bid, {})
        notes = entry.get("notes", "-")
        if not (notes and notes != "-"):
            # Fallback: by normalized brand name (catches brand_id normalization mismatches)
            bname_key = norm_text(str(row.get("_name", "")))
            entry_by_name = meta_map.get(bname_key, {})
            notes = entry_by_name.get("notes", "-")
        return _shorten_note_preview(notes)

    follow_df["Last Notes"] = follow_df.apply(_resolve_last_notes, axis=1)
    follow_df["_manual_status"] = follow_df["_id"].apply(lambda x: meta_value(x, "status", "OFF 😴"))
    # Status = cadencia de contacto (mismo lenguaje que Salud de Cartera)
    # 🟢 Activo (0-10d) · 🟡 Cadencia (11-15d) · 🟠 Alerta (16-21d) · 🔴 Fría (>21d)
    follow_df["Status"] = follow_df["_last_comment_dt"].apply(_cadencia_status)

    follow_df["_has_comment"] = follow_df["_last_comment_dt"].apply(lambda x: not pd.isna(pd.to_datetime(x, errors="coerce")))
    follow_df["_sort_dt"] = pd.to_datetime(follow_df["_last_comment_dt"], errors="coerce")

    # Sort by Smart Priorities first — brands with SP score rank at the top in SP order.
    # Within each SP group, untouched first then oldest comment.
    sp_df = load_priority_data()
    sp_rank_map = {}
    if not sp_df.empty:
        total_rows = sp_df[sp_df["_metric_norm"] == "total"].copy()
        for _, r in total_rows.iterrows():
            bid = normalize_brand_id(r.get("_id", ""))
            rank = r.get("_total_rank")
            if bid and rank is not None:
                sp_rank_map[bid] = int(rank)

    follow_df["_sp_rank"] = follow_df["_id"].apply(lambda x: sp_rank_map.get(normalize_brand_id(x), 9999))

    follow_df = follow_df.sort_values(
        by=["_sp_rank", "_has_comment", "_sort_dt"],
        ascending=[True, True, True],
        na_position="first",
    ).reset_index(drop=True)
    follow_df["Rank"] = follow_df.index + 1

    def _lever_status_label(row):
        parts = []
        if row.get("_has_ads"):
            parts.append("Ads ✅")
        if row.get("_has_md"):
            parts.append("MD ✅")
        if not parts:
            parts.append("Sin palanca 🔴")
        return " · ".join(parts)

    follow_df["Palancas"] = follow_df.apply(_lever_status_label, axis=1)
    follow_df["Churn"]    = follow_df["_churn"].apply(lambda v: _churn_label_with_emoji(v) if v else "✅ On")

    total = len(follow_df)
    sin_palanca = (follow_df["Palancas"] == "Sin palanca 🔴").sum()

    # ─────────────────────────────────────────────────────────────────────────
    # PORTFOLIO HEALTH METER — cadencia 21 días
    # Zonas: Activo (0-10d) · Cadencia (11-15d) · Alerta (16-21d) · Fría (>21d)
    # Denominador = portafolio completo (marcas sin historial cuentan como Frías)
    # Score = % marcas dentro del ciclo de 21 días
    # ─────────────────────────────────────────────────────────────────────────
    today_dt = pd.Timestamp(date.today())

    def _days_since(last_dt):
        """Returns int days since last contact, or None if never contacted."""
        ts = pd.to_datetime(last_dt, errors="coerce")
        if pd.isna(ts):
            return None
        # Strip timezone so comparison with today_dt (tz-naive) never errors
        if ts.tzinfo is not None:
            ts = ts.tz_localize(None)
        delta = (today_dt - ts).days
        return max(0, delta)

    follow_df["_days_since"] = follow_df["_last_comment_dt"].apply(_days_since)

    n_activo   = ((follow_df["_days_since"] >= 0)  & (follow_df["_days_since"] <= 10)).sum()
    n_cadencia = ((follow_df["_days_since"] > 10)  & (follow_df["_days_since"] <= 15)).sum()
    n_alerta   = ((follow_df["_days_since"] > 15)  & (follow_df["_days_since"] <= 21)).sum()
    n_fria     = (follow_df["_days_since"] > 21).sum()
    n_sin_data = follow_df["_days_since"].isna().sum()

    # Denominador = portafolio completo; sin historial cuenta como Fría
    total_portfolio = len(follow_df)
    n_fria_total = n_fria + n_sin_data  # frías reales + sin contacto registrado

    def _pct(n):
        return round(100 * n / total_portfolio) if total_portfolio > 0 else 0

    pct_activo   = _pct(n_activo)
    pct_cadencia = _pct(n_cadencia)
    pct_alerta   = _pct(n_alerta)
    pct_fria     = _pct(n_fria_total)

    # Score = % de marcas dentro del ciclo (≤21 días)
    n_dentro_ciclo = n_activo + n_cadencia + n_alerta
    health_score = round(100 * n_dentro_ciclo / total_portfolio) if total_portfolio > 0 else 0

    if health_score >= 70:
        score_color = "#7ED321"
        score_label = "SANA"
    elif health_score >= 45:
        score_color = "#FF7124"
        score_label = "MODERADA"
    elif health_score >= 25:
        score_color = "#FF7124"
        score_label = "EN ALERTA"
    else:
        score_color = "#FF4D2E"
        score_label = "CRÍTICA"

    bw_activo   = pct_activo
    bw_cadencia = pct_cadencia
    bw_alerta   = pct_alerta
    bw_fria     = 100 - bw_activo - bw_cadencia - bw_alerta  # fills remainder

    # ── Columnas nuevas: Días y Próximo Contacto ──────────────────────────────
    def _fmt_dias(days_val):
        """Colored badge showing days since last contact."""
        try:
            if days_val is None:
                return "—"
            d = int(float(days_val))  # handles both int, float, and float NaN via ValueError
        except (ValueError, TypeError):
            return "—"
        if d <= 10:
            color = "#7ED321"
        elif d <= 15:
            color = "#7ED321"
        elif d <= 21:
            color = "#FF7124"
        else:
            color = "#FF4D2E"
        return (
            f'<span style="background:{color}22; color:{color}; '
            f'font-weight:700; padding:2px 8px; border-radius:6px; '
            f'font-size:12px; border:1px solid {color}66;">{d}d</span>'
        )

    def _fmt_proximo(last_dt):
        """
        Próximo contacto ideal = last_dt + 15d (alerta amarilla).
        Vencido en 21d (alerta roja).
        Si nunca contactado → Pendiente (rojo).
        """
        ts = pd.to_datetime(last_dt, errors="coerce")
        if pd.isna(ts):
            return '<span style="color:#FF4D2E; font-weight:700;">Sin contacto</span>'
        ideal_dt  = ts + timedelta(days=15)
        limite_dt = ts + timedelta(days=21)
        days_to_ideal  = (ideal_dt.date()  - date.today()).days
        days_to_limite = (limite_dt.date() - date.today()).days

        if days_to_limite < 0:
            # Ya venció el límite de 21 días
            overdue = abs(days_to_limite)
            return (
                f'<span style="color:#FF4D2E; font-weight:700;">'
                f'Vencido hace {overdue}d</span>'
            )
        elif days_to_ideal <= 0:
            # Pasó los 15d, todavía dentro del límite → amarillo
            return (
                f'<span style="color:#FF7124; font-weight:700;">'
                f'Llamar ya · vence en {days_to_limite}d</span>'
            )
        else:
            # Dentro del ciclo ideal
            return (
                f'<span style="color:#1B3F8B;">'
                f'en {days_to_ideal}d · {ideal_dt.strftime("%d/%m")}</span>'
            )

    follow_df["_dias_fmt"]    = follow_df["_days_since"].apply(_fmt_dias)
    follow_df["_proximo_fmt"] = follow_df["_last_comment_dt"].apply(_fmt_proximo)

    st.markdown("## Active Account Follow-Ups")
    # ── Contador efectivo: Hoy / Semana / Mes ────────────────────────────────
    # Fuente: CSV de comentarios (growth_os_comments.csv) — última tipificación
    # registrada desde el Brand Finder, filtrando por opportunity_status positivo
    # (Follow-up, Deal Closed, Negotiation). Reemplaza la fuente anterior (Productivity),
    # que no reflejaba directamente la tipificación hecha en el dashboard.
    #   HOY    → datetime del comentario == hoy
    #   SEMANA → datetime del comentario dentro de la semana actual (lunes-domingo)
    #   MES    → datetime del comentario dentro del mes actual
    _today      = date.today()
    _week_start = _today - timedelta(days=_today.weekday())
    _week_end   = _week_start + timedelta(days=6)

    _POSITIVE_STATUSES = {"Follow-up ✅", "Deal Closed 🏆", "Negotiation ⏳"}

    _comments_df = _load_comments_df()
    _contacts_today = 0
    _contacts_this_week = 0
    _contacts_this_month = 0

    if not _comments_df.empty and "_dt" in _comments_df.columns:
        _positive = _comments_df[
            _comments_df["opportunity_status"].isin(_POSITIVE_STATUSES)
            & _comments_df["_dt"].notna()
        ].copy()

        if not _positive.empty:
            _positive["_date_only"] = _positive["_dt"].dt.date

            _mask_today = _positive["_date_only"] == _today
            _mask_week  = _positive["_date_only"].apply(lambda d: _week_start <= d <= _week_end)
            _mask_month = _positive["_date_only"].apply(lambda d: d.year == _today.year and d.month == _today.month)

            _contacts_today      = int(_mask_today.sum())
            _contacts_this_week  = int(_mask_week.sum())
            _contacts_this_month = int(_mask_month.sum())

    _week_label  = f"{_week_start.strftime('%d %b')} – {_week_end.strftime('%d %b')}"
    _month_label = _today.strftime("%B %Y")

    # ── Filtro por zona de temperatura ────────────────────────────────────────
    _temp_filter = st.selectbox(
        "🌡️ Filtrar por zona de temperatura",
        ["Todas", "🟢 Activo (0–10d)", "🟡 Cadencia (11–15d)", "🟠 Alerta (16–21d)", "🔴 Fría / Sin contacto (>21d)"],
        index=0,
        key="followup_temp_filter",
    )

    follow_df_filtered = follow_df.copy()
    if _temp_filter == "🟢 Activo (0–10d)":
        follow_df_filtered = follow_df_filtered[follow_df_filtered["_days_since"].between(0, 10, inclusive="both")]
    elif _temp_filter == "🟡 Cadencia (11–15d)":
        follow_df_filtered = follow_df_filtered[follow_df_filtered["_days_since"].between(11, 15, inclusive="both")]
    elif _temp_filter == "🟠 Alerta (16–21d)":
        follow_df_filtered = follow_df_filtered[follow_df_filtered["_days_since"].between(16, 21, inclusive="both")]
    elif _temp_filter == "🔴 Fría / Sin contacto (>21d)":
        follow_df_filtered = follow_df_filtered[
            follow_df_filtered["_days_since"].isna() | (follow_df_filtered["_days_since"] > 21)
        ]

    follow_view = pd.DataFrame({
        "Rank":             follow_df_filtered["Rank"].apply(_format_rank),
        "Status":           follow_df_filtered["Status"],
        "Días":             follow_df_filtered["_dias_fmt"],
        "Próximo Contacto": follow_df_filtered["_proximo_fmt"],
        "Last Update":      follow_df_filtered["Last Update"],
        "ID":               follow_df_filtered["_id"].apply(_format_id),
        "Restaurant":       follow_df_filtered["_name"],
        "Palancas":         follow_df_filtered["Palancas"],
        "Churn":            follow_df_filtered["Churn"],
        "Last Notes":       follow_df_filtered["Last Notes"],
    })
    _hm_track_color = "#141A2E" if DARK_MODE else "#FFFFFF"
    _hm_card_bg     = "#141A2E" if DARK_MODE else "#FFFFFF"
    st.markdown(f"""
    <style>
    .hm-card {{
        background: {_hm_card_bg};
        border: 1px solid {_hm_card_bg};
        border-radius: 16px;
        padding: 20px 24px 16px 24px;
        margin-bottom: 18px;
    }}
    .hm-title {{
        font-size: 11px;
        font-weight: 700;
        letter-spacing: 2px;
        text-transform: uppercase;
        color: #1B3F8B;
        margin-bottom: 14px;
    }}
    .hm-score-ring {{
        display: flex;
        align-items: center;
        gap: 24px;
        margin-bottom: 16px;
    }}
    .hm-ring {{
        width: 80px;
        height: 80px;
        border-radius: 50%;
        background: conic-gradient({score_color} {health_score}%, {_hm_track_color} {health_score}%);
        display: flex;
        align-items: center;
        justify-content: center;
        flex-shrink: 0;
        position: relative;
    }}
    .hm-ring-inner {{
        width: 58px;
        height: 58px;
        border-radius: 50%;
        background: {_hm_card_bg};
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
    }}
    .hm-ring-num {{
        font-size: 20px;
        font-weight: 900;
        color: {score_color};
        line-height: 1;
    }}
    .hm-ring-sub {{
        font-size: 8px;
        font-weight: 700;
        color: {score_color};
        letter-spacing: 1px;
        opacity: 0.8;
    }}
    .hm-score-label {{
        font-size: 22px;
        font-weight: 900;
        color: {score_color};
        letter-spacing: 1px;
    }}
    .hm-score-desc {{
        font-size: 13px;
        color: #1B3F8B;
        margin-top: 2px;
    }}
    .hm-zones {{
        display: flex;
        gap: 10px;
        margin-bottom: 14px;
        flex-wrap: wrap;
    }}
    .hm-zone {{
        flex: 1;
        min-width: 100px;
        background: #FFFFFF;
        border-radius: 10px;
        padding: 10px 12px;
        position: relative;
        overflow: hidden;
    }}
    .hm-zone-accent {{
        position: absolute;
        left: 0; top: 0; bottom: 0;
        width: 4px;
        border-radius: 10px 0 0 10px;
    }}
    .hm-zone-emoji {{ font-size: 16px; }}
    .hm-zone-name {{ font-size: 10px; font-weight: 700; letter-spacing: 1px; color: #1B3F8B; text-transform: uppercase; margin-top: 4px; }}
    .hm-zone-count {{ font-size: 24px; font-weight: 900; color: rgba(255,255,255,0.9); line-height: 1.1; }}
    .hm-zone-pct {{ font-size: 11px; color: #1B3F8B; }}
    .hm-zone-range {{ font-size: 10px; color: #3D64B8; margin-top: 2px; }}
    .hm-bar-wrap {{ border-radius: 8px; overflow: hidden; height: 10px; display: flex; margin-top: 4px; }}
    .hm-bar-seg {{ height: 10px; transition: width 0.3s; }}
    .hm-legend {{ display: flex; gap: 16px; margin-top: 6px; flex-wrap: wrap; }}
    .hm-legend-item {{ display: flex; align-items: center; gap: 5px; font-size: 11px; color: #1B3F8B; }}
    .hm-legend-dot {{ width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }}
    </style>
    <div class="hm-card">
        <div class="hm-title">🌡️ Salud de Cartera · Cadencia 21 Días</div>
        <div class="hm-score-ring">
            <div class="hm-ring">
                <div class="hm-ring-inner">
                    <span class="hm-ring-num">{health_score}</span>
                    <span class="hm-ring-sub">/100</span>
                </div>
            </div>
            <div>
                <div class="hm-score-label">Cartera {score_label}</div>
                <div class="hm-score-desc">{total_portfolio} marcas en portafolio · {n_dentro_ciclo} dentro del ciclo · {n_fria_total} frías o sin contacto</div>
            </div>
        </div>
        <div class="hm-zones">
            <div class="hm-zone">
                <div class="hm-zone-accent" style="background:#7ED321;"></div>
                <div class="hm-zone-emoji">🟢</div>
                <div class="hm-zone-name">Activo</div>
                <div class="hm-zone-count">{n_activo}</div>
                <div class="hm-zone-pct">{pct_activo}% del portafolio</div>
                <div class="hm-zone-range">0 – 10 días</div>
            </div>
            <div class="hm-zone">
                <div class="hm-zone-accent" style="background:#7ED321;"></div>
                <div class="hm-zone-emoji">🟡</div>
                <div class="hm-zone-name">Cadencia ideal</div>
                <div class="hm-zone-count">{n_cadencia}</div>
                <div class="hm-zone-pct">{pct_cadencia}% del portafolio</div>
                <div class="hm-zone-range">11 – 15 días</div>
            </div>
            <div class="hm-zone">
                <div class="hm-zone-accent" style="background:#FF7124;"></div>
                <div class="hm-zone-emoji">🟠</div>
                <div class="hm-zone-name">Alerta</div>
                <div class="hm-zone-count">{n_alerta}</div>
                <div class="hm-zone-pct">{pct_alerta}% del portafolio</div>
                <div class="hm-zone-range">16 – 21 días</div>
            </div>
            <div class="hm-zone">
                <div class="hm-zone-accent" style="background:#FF4D2E;"></div>
                <div class="hm-zone-emoji">🔴</div>
                <div class="hm-zone-name">Fría</div>
                <div class="hm-zone-count">{n_fria_total}</div>
                <div class="hm-zone-pct">{pct_fria}% del portafolio</div>
                <div class="hm-zone-range">&gt; 21 días · sin contacto</div>
            </div>
        </div>
        <div class="hm-bar-wrap">
            <div class="hm-bar-seg" style="width:{bw_activo}%; background:#7ED321;"></div>
            <div class="hm-bar-seg" style="width:{bw_cadencia}%; background:#7ED321;"></div>
            <div class="hm-bar-seg" style="width:{bw_alerta}%; background:#FF7124;"></div>
            <div class="hm-bar-seg" style="width:{bw_fria}%; background:#FF4D2E;"></div>
        </div>
        <div class="hm-legend">
            <div class="hm-legend-item"><div class="hm-legend-dot" style="background:#7ED321;"></div> Activo (0–10d)</div>
            <div class="hm-legend-item"><div class="hm-legend-dot" style="background:#7ED321;"></div> Cadencia (11–15d)</div>
            <div class="hm-legend-item"><div class="hm-legend-dot" style="background:#FF7124;"></div> Alerta (16–21d)</div>
            <div class="hm-legend-item"><div class="hm-legend-dot" style="background:#FF4D2E;"></div> Fría (&gt;21d · sin contacto)</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    _showing = len(follow_df_filtered)

    # Color thresholds: today ≥5 green, ≥2 orange, else red
    #                   week  ≥15 green, ≥8 orange, else red
    #                   month ≥60 green, ≥30 orange, else red
    _today_color = "#7ED321" if _contacts_today >= 5 else ("#FF7124" if _contacts_today >= 2 else "#FF4D2E")
    _wk_color    = "#7ED321" if _contacts_this_week >= 15 else ("#FF7124" if _contacts_this_week >= 8 else "#FF4D2E")
    _mo_color    = "#7ED321" if _contacts_this_month >= 60 else ("#FF7124" if _contacts_this_month >= 30 else "#FF4D2E")

    st.markdown(f"""
<div style="display:flex;flex-wrap:wrap;gap:10px;margin-bottom:14px;">
    <div style="display:inline-flex;align-items:center;gap:10px;background:rgba(111,242,75,.06);
        border:1.5px solid rgba(111,242,75,.22);border-radius:12px;padding:10px 18px;">
        <div style="font-size:10px;font-weight:900;text-transform:uppercase;color:#6B7280;letter-spacing:1px;">📞 HOY</div>
        <div style="font-size:30px;font-weight:900;color:{_today_color};line-height:1;">{_contacts_today}</div>
        <div style="font-size:11px;color:#1B3F8B;">contactos efectivos</div>
    </div>
    <div style="display:inline-flex;align-items:center;gap:10px;background:rgba(78,99,217,.07);
        border:1.5px solid rgba(78,99,217,.25);border-radius:12px;padding:10px 18px;">
        <div style="font-size:10px;font-weight:900;text-transform:uppercase;color:#6B7280;letter-spacing:1px;">📅 SEMANA</div>
        <div style="font-size:30px;font-weight:900;color:{_wk_color};line-height:1;">{_contacts_this_week}</div>
        <div style="font-size:11px;color:#1B3F8B;">{_week_label}</div>
    </div>
    <div style="display:inline-flex;align-items:center;gap:10px;background:rgba(255,113,36,.06);
        border:1.5px solid rgba(255,113,36,.22);border-radius:12px;padding:10px 18px;">
        <div style="font-size:10px;font-weight:900;text-transform:uppercase;color:#6B7280;letter-spacing:1px;">🗓️ MES</div>
        <div style="font-size:30px;font-weight:900;color:{_mo_color};line-height:1;">{_contacts_this_month}</div>
        <div style="font-size:11px;color:#1B3F8B;">{_month_label}</div>
    </div>
    <div style="display:inline-flex;align-items:center;gap:8px;background:rgba(255,255,255,.04);
        border:1px solid rgba(255,255,255,.10);border-radius:12px;padding:10px 14px;font-size:11px;color:#1B3F8B;">
        <span style="color:#7ED321;">🟢 {n_activo}</span> &nbsp;·&nbsp;
        <span style="color:#7ED321;">🟡 {n_cadencia}</span> &nbsp;·&nbsp;
        <span style="color:#FF7124;">🟠 {n_alerta}</span> &nbsp;·&nbsp;
        <span style="color:#FF4D2E;">🔴 {n_fria_total}</span>
    </div>
</div>
    """, unsafe_allow_html=True)
    st.caption(
        f"**{total_portfolio} marcas** en portafolio · "
        f"**{total - sin_palanca}** con palanca activa · "
        f"**{sin_palanca}** sin palanca · "
        f"Mostrando **{_showing}** marca{'s' if _showing != 1 else ''} {('(filtro: ' + _temp_filter + ')') if _temp_filter != 'Todas' else '(todas)'}"
    )
    _render_light_table(follow_view, height=680)


# =========================
# =========================
# PRODUCTIVITY HEATMAP
# =========================

def page_call_quality_trainer():
    render_header("Call Quality Trainer", f"Análisis de calidad de llamadas · Coach personal · {FARMER_NAME}")


    # ── Load from sheets inside EXCEL_FILE ───────────────────────────────────
    @st.cache_data(ttl=3000, show_spinner=False)
    def _load_quality(excel_path):
        if not os.path.exists(excel_path):
            return pd.DataFrame()
        try:
            df = pd.read_excel(excel_path, sheet_name="Call Detail", header=0)
            df = df[df["Farmer"].notna()].copy()
            return df
        except Exception:
            return pd.DataFrame()

    @st.cache_data(ttl=3000, show_spinner=False)
    def _load_weekly(excel_path):
        if not os.path.exists(excel_path):
            return pd.DataFrame()
        try:
            raw = pd.read_excel(excel_path, sheet_name="Call Quality", header=None)
            header_idx = None
            for i, row in raw.iterrows():
                if any(str(v).strip() == "WEEK" for v in row):
                    header_idx = i
                    break
            if header_idx is None:
                return pd.DataFrame()
            raw.columns = raw.iloc[header_idx]
            raw = raw.iloc[header_idx + 1:].reset_index(drop=True)
            raw.columns.name = None
            raw = raw[raw["WEEK"].notna()].copy()
            for c in raw.columns[1:]:
                raw[c] = pd.to_numeric(raw[c], errors="coerce")
            return raw
        except Exception:
            return pd.DataFrame()

    df   = _load_quality(EXCEL_FILE)
    wdf  = _load_weekly(EXCEL_FILE)

    if df.empty:
        st.info(
            "No se encontr\u00f3 la hoja **Call Detail** en el Excel principal. "
            "Verific\u00e1 que la hoja existe y tiene datos desde A1."
        )
        return
    # Each: (col_name, display_label, comment_col, impact_en_cartera, tip_farmer_sr)
    DIMS = [
        (
            "%Exec. Summary Provided",
            "Exec Summary",
            "Executive Summary Comment",
            "Sin resumen al cierre el aliado no recuerda lo acordado. El MD y los Ads que prometiste en la llamada mueren ahí — la semana siguiente tenés que volver a vender desde cero y el % de conversión se desploma.",
            "Antes de cortar siempre decí: “Te resumo lo que quedamos hoy: [palanca], [acción], [fecha]. ¿Estamos alineados?” Son 20 segundos que valen semanas de seguimiento.",
        ),
        (
            "%Decision Maker Confirm.",
            "Decision Maker",
            None,
            "Si no confirmás que hablas con quien decide, el 100% del pitch comercial puede estar cayendo en alguien que no tiene poder para aceptar un MD o activar Ads. Tus tasas de conversión bajan aunque tu argumento sea perfecto.",
            "En los primeros 30 segundos preguntá: “¿Sos vos quien maneja las promociones en Rappi?” Si no, pedí que te pasen. No pierdas el pitch con el cajero.",
        ),
        (
            "%Investment",
            "Investment",
            "Investment Comment",
            "Investment es el puente entre la llamada y el cierre. Sin un plan concreto de inversión (monto, palanca, fecha), el aliado trata la llamada como informativa y no como un compromiso. Directo impacto en GMV de la cartera.",
            "No alcanza con mencionar el porcentaje. Necesitás proponer: “Arrancamos con 15% de MD esta semana, ¿te parece bien el martes para activarlo?” Anclá fecha y acción específica.",
        ),
        (
            "%Top Rest Action Plan",
            "Top Rest Action Plan",
            "Top Rest Comment",
            "Tus Top Restaurants son los que mueven el 80% del GMV. Si salís de la llamada sin un plan de acción concreto para ellos, estás dejando la palanca más poderosa sin accionar. Un Top Rest sin plan = GMV congelado.",
            "Preparate antes de llamar. Abrí el perfil del aliado y mirá su posición en el ranking. Si es Top Rest, la llamada no puede terminar sin un “el próximo paso es X para el Y de esta semana”.",
        ),
        (
            "%ADS Action Plan",
            "ADS Action Plan",
            None,
            "Ads sin plan de acción es una conversación de scouting, no de cierre. Cada llamada donde mencionaste Ads pero no quedó definido el tipo, el monto y la fecha de activación es una oportunidad de ingreso perdida directamente.",
            "El cierre de Ads tiene que incluir: tipo (Never Ads / Sponsored / Campaign), presupuesto aproximado y fecha de inicio. Si el aliado dice “lo pienso”, dejá un follow-up en 48 horas máximo.",
        ),
        (
            "%Churn Action Plan",
            "Churn Action Plan",
            "Churn Comment",
            "Identificar churn sin plan de acción es el peor escenario: sabés que el aliado se va y no hacés nada concreto. Cada aliado que churna te baja el GMV base y obliga a recuperar volumen con nuevos aliados, que cuestan 3x más.",
            "Cuando identificás riesgo de churn preguntá directamente: “¿Qué necesitaría Rappi hacer diferente para que te quedes activo?” Luego proponé una acción concreta en esa misma llamada.",
        ),
        (
            "%Introduction",
            "Introducción",
            "Introduction Comment",
            "Una introducción incompleta rompe la confianza desde el inicio. Si el aliado no sabe quién sos ni de dónde llamás, su guardia sube y el pitch comercial que viene después tiene menos chance de llegar.",
            "Siempre: nombre + rol + marca + motivo. “Hola, soy Sabas de Rappi, te llamo porque vi que tenés oportunidad de mejorar tu visibilidad esta semana.” Son 10 segundos que bajan la resistencia del aliado.",
        ),
        (
            "%Call Handling",
            "Call Handling",
            "Handling Comment",
            "Sin control de la llamada el aliado dirige la conversación hacia sus problemas operativos y perdés el tiempo del pitch. Una llamada donde manejás solo incendios tiene 0% de probabilidad de cerrar una palanca comercial.",
            "Cuando el aliado desvía hacia un problema operativo, reconocé y redirigí: “Entendido, lo anotamos para dar seguimiento — pero antes de cortar quiero contarte algo que puede mejorar tus ventas esta semana.” Volvé al pitch.",
        ),
        (
            "%Self Service Info",
            "Self Service Info",
            None,
            "No darle al aliado herramientas de autogestión genera dependencia: cada pequeño problema te va a llamar a vos. Eso te roba tiempo de llamadas comerciales y baja tu productividad neta por hora.",
            "Al final de cada llamada mencioná al menos un recurso: “Recordá que podés ver tus métricas en el portal de aliados en tiempo real.” Son 15 segundos que reducen re-llamadas operativas.",
        ),
        (
            "%Assortment",
            "Assortment",
            "Assortment Comment",
            "El catálogo incompleto es la razón número uno de baja conversión en Rappi. Si no trabajás el assortment en la llamada, el aliado puede tener Ads activos y MD configurado pero perder ventas porque la mitad de su menú está sin foto o sin precio.",
            "Revisá el catálogo del aliado antes de llamar. Si tiene menos del 80% de productos activos, eso va al inicio del pitch — es el quick win más fácil de mostrar con impacto inmediato en su CR.",
        ),
    ]

    THRESHOLD = 0.80  # 80%

    # ── Compute per-dim averages ──────────────────────────────────────────────
    dim_scores = {}
    for col, label, *_ in DIMS:
        if col in df.columns:
            vals = pd.to_numeric(df[col], errors="coerce").dropna()
            dim_scores[col] = float(vals.mean()) if len(vals) else None

    below = [(col, label, cmt, impact, tip)
             for col, label, cmt, impact, tip in DIMS
             if dim_scores.get(col) is not None and dim_scores[col] < THRESHOLD]
    above = [(col, label, cmt, impact, tip)
             for col, label, cmt, impact, tip in DIMS
             if dim_scores.get(col) is not None and dim_scores[col] >= THRESHOLD]

    # ── Sort below by score ascending (worst first) ───────────────────────────
    below.sort(key=lambda x: dim_scores[x[0]])

    # ── Weekly trend helper ───────────────────────────────────────────────────
    WEEK_LABEL_MAP = {
        "%Exec. Summary Provided": "Exec Summary Provided",
        "%Decision Maker Confirm.": "Decision Maker Confirm",
        "%Investment":              "Investment",
        "%Top Rest Action Plan":    "Top Rest",
        "%ADS Action Plan":         "Investment",  # fallback
        "%Churn Action Plan":       "Churn",
        "%Introduction":            "Introduction",
        "%Call Handling":           "Call Handling",
        "%Self Service Info":       "Self Service Info",
        "%Assortment":              "Assortment",
    }

    def sparkline_data(col):
        if wdf.empty:
            return []
        mapped = WEEK_LABEL_MAP.get(col, "")
        if mapped in wdf.columns:
            return [round(float(v) * 100, 1) if pd.notna(v) else None for v in wdf[mapped]]
        return []

    # ── Recent failing comment examples ──────────────────────────────────────
    def get_examples(col, cmt_col, n=2):
        if cmt_col not in df.columns:
            return []
        sub = df[(pd.to_numeric(df[col], errors="coerce") < 0.5) & df[cmt_col].notna()]
        return sub[cmt_col].dropna().head(n).tolist()

    # ── CSS ───────────────────────────────────────────────────────────────────
    st.markdown("""
    <style>
    .qt-score-grid { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; margin-bottom:1.5rem; }
    .qt-score-card { background:rgba(255,255,255,0.90); border-radius:8px; padding:12px 14px; }
    .qt-score-val  { font-size:22px; font-weight:600; line-height:1.1; }
    .qt-score-lbl  { font-size:11px; color:#888; margin-top:3px; }
    .qt-score-trend{ font-size:11px; margin-top:2px; }
    .qt-ok  { color:#7ED321; }
    .qt-warn{ color:#D95A10; }
    .qt-bad { color:#FF4D2E; }

    .qt-trainer-card {
        border: 1px solid rgba(0,0,0,0.08);
        border-radius:10px;
        margin-bottom: 1.2rem;
        overflow: hidden;
    }
    .qt-card-header {
        display: flex; align-items: center; gap: 12px;
        padding: 14px 18px 10px;
        border-bottom: 1px solid rgba(255,255,255,0.95);
    }
    .qt-rank-badge {
        width:30px; height:30px; border-radius:50%;
        display:flex; align-items:center; justify-content:center;
        font-size:13px; font-weight:700; flex-shrink:0;
    }
    .qt-dim-label { font-size:14px; font-weight:600; color:#1A1A2E; flex:1; }
    .qt-score-pill {
        font-size:13px; font-weight:600; padding:3px 10px;
        border-radius:20px; flex-shrink:0;
    }
    .qt-spark { font-size:10px; color:#999; margin-left:6px; }
    .qt-card-body { padding: 14px 18px; }
    .qt-section-title {
        font-size:10px; font-weight:700; letter-spacing:.07em;
        text-transform:uppercase; color:#aaa; margin-bottom:6px; margin-top:12px;
    }
    .qt-section-title:first-child { margin-top:0; }
    .qt-impact-box {
        background:rgba(255,113,36,0.06); border-left:3px solid #D95A10;
        border-radius:0 6px 6px 0; padding:8px 12px;
        font-size:12px; color:#D95A10; line-height:1.6;
    }
    .qt-tip-box {
        background:rgba(59,72,131,0.10); border-left:3px solid #1B3F8B;
        border-radius:0 6px 6px 0; padding:8px 12px;
        font-size:12px; color:#6B7280; line-height:1.6;
    }
    .qt-example {
        background:rgba(27,63,139,0.03); border:1px solid rgba(255,255,255,0.95); border-radius:6px;
        padding:8px 10px; font-size:11px; color:#666; line-height:1.55;
        margin-bottom:6px; font-style:italic;
    }
    .qt-sparkline-row { display:flex; align-items:center; gap:4px; margin-top:6px; }
    .qt-spark-dot {
        width:28px; height:28px; border-radius:50%;
        display:flex; align-items:center; justify-content:center;
        font-size:9px; font-weight:600; flex-shrink:0;
    }
    .qt-above-section {
        margin-top:1.5rem; padding:12px 16px; border-radius:8px;
        border:1px solid rgba(111,242,75,0.08); background:rgba(111,242,75,0.08);
        font-size:12px; color:#7ED321; line-height:1.7;
    }
    .qt-above-section b { color:#7ED321; }
    .qt-separator {
        font-size:11px; font-weight:700; letter-spacing:.08em;
        text-transform:uppercase; color:#bbb;
        margin:1.8rem 0 1rem; display:flex; align-items:center; gap:8px;
    }
    .qt-separator::after { content:""; flex:1; height:1px; background:#eee; }
    </style>
    """, unsafe_allow_html=True)

    # ── Summary scorecards ────────────────────────────────────────────────────
    enc_avg = df["%ENC"].mean() * 100 if "%ENC" in df.columns else None
    ec_avg  = df["%EC"].mean()  * 100 if "%EC"  in df.columns else None
    n_calls = len(df)
    n_below = len(below)

    # ── Call Quality Score compuesto (promedio de las 10 dimensiones) ─────────
    _dim_scores_valid = [v for v in dim_scores.values() if v is not None]
    _composite_score  = round(sum(_dim_scores_valid) / len(_dim_scores_valid) * 100, 1) if _dim_scores_valid else None
    _composite_color  = "qt-ok" if _composite_score and _composite_score >= 80 else ("qt-warn" if _composite_score and _composite_score >= 65 else "qt-bad")
    _composite_label  = "✓ Buena calidad" if _composite_score and _composite_score >= 80 else ("↗ En desarrollo" if _composite_score and _composite_score >= 65 else "↓ Necesita trabajo")

    # ── Save weekly score snapshot to history CSV ────────────────────────────
    if _composite_score is not None:
        _cq_week = APP_PERIOD
        _cq_row  = pd.DataFrame([{
            "week":            _cq_week,
            "composite_score": _composite_score,
            "n_calls":         len(df),
            "n_below_80":      len(below),
            "saved_at":        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }])
        if os.path.exists(CALL_QUALITY_HISTORY_FILE):
            try:
                _cq_hist = pd.read_csv(CALL_QUALITY_HISTORY_FILE)
                # Only upsert — don't duplicate rows for the same week
                _cq_hist = _cq_hist[_cq_hist["week"].astype(str) != str(_cq_week)]
                _cq_hist = pd.concat([_cq_hist, _cq_row], ignore_index=True)
            except Exception:
                _cq_hist = _cq_row
        else:
            _cq_hist = _cq_row
        _cq_hist.to_csv(CALL_QUALITY_HISTORY_FILE, index=False, encoding="utf-8-sig")

    # ── Historical trend of composite score ──────────────────────────────────
    if os.path.exists(CALL_QUALITY_HISTORY_FILE):
        try:
            _cq_hist_view = pd.read_csv(CALL_QUALITY_HISTORY_FILE).sort_values("week")
            if len(_cq_hist_view) >= 2:
                _hist_scores = _cq_hist_view["composite_score"].tolist()
                _hist_weeks  = _cq_hist_view["week"].tolist()
                # Build inline SVG sparkline (200x40)
                _sx_min, _sx_max = min(_hist_scores), max(_hist_scores)
                _sx_range = max(_sx_max - _sx_min, 1)
                W, H, PAD = 200, 36, 6
                def _sx(i): return PAD + i * (W - 2*PAD) / max(len(_hist_scores)-1, 1)
                def _sy(v): return H - PAD - (v - _sx_min) / _sx_range * (H - 2*PAD)
                _pts = " ".join(f"{_sx(i):.1f},{_sy(v):.1f}" for i, v in enumerate(_hist_scores))
                _last  = _hist_scores[-1]
                _first = _hist_scores[0]
                _line_color = "#7ED321" if _last >= _first else "#FF4D2E"
                _dot_x = _sx(len(_hist_scores)-1)
                _dot_y = _sy(_last)
                _spark_svg = (
                    f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" style="vertical-align:middle;">' 
                    f'<polyline points="{_pts}" fill="none" stroke="{_line_color}" stroke-width="2" stroke-linejoin="round"/>' 
                    f'<circle cx="{_dot_x:.1f}" cy="{_dot_y:.1f}" r="3.5" fill="{_line_color}"/>' 
                    f'</svg>'
                )
                _trend_label = f"+{_last-_first:.1f}pp" if _last >= _first else f"{_last-_first:.1f}pp"
                _trend_color = "#7ED321" if _last >= _first else "#FF4D2E"
                st.markdown(
                    f'''<div style="display:flex;align-items:center;gap:14px;background:rgba(27,63,139,0.03);
                    border:1px solid rgba(0,0,0,0.08);border-radius:8px;padding:10px 16px;margin-bottom:12px;font-size:12px;color:#555;">
                    <span style="font-weight:700;color:#333;">Historial Call Quality Score</span>
                    {_spark_svg}
                    <span style="font-weight:700;color:{_trend_color};">{_trend_label} ({_hist_weeks[0]} → {_hist_weeks[-1]})</span>
                    <span style="color:#aaa;">({len(_hist_scores)} semanas)</span>
                    </div>''',
                    unsafe_allow_html=True,
                )
        except Exception:
            pass

    enc_color = "qt-ok" if enc_avg and enc_avg >= 80 else ("qt-warn" if enc_avg and enc_avg >= 65 else "qt-bad")
    ec_color  = "qt-ok" if ec_avg  and ec_avg  >= 80 else ("qt-warn" if ec_avg  and ec_avg  >= 65 else "qt-bad")

    st.markdown(f"""
    <div class="qt-score-grid" style="grid-template-columns:repeat(5,minmax(0,1fr));">
      <div class="qt-score-card" style="border:2px solid {"#7ED321" if _composite_score and _composite_score >= 80 else ("#D95A10" if _composite_score and _composite_score >= 65 else "#FF4D2E")};background:{"rgba(111,242,75,0.08)" if _composite_score and _composite_score >= 80 else ("rgba(255,113,36,0.06)" if _composite_score and _composite_score >= 65 else "rgba(229,51,42,0.10)")};">
        <div class="qt-score-val {_composite_color}" style="font-size:28px;">{f"{_composite_score:.1f}%" if _composite_score is not None else "—"}</div>
        <div class="qt-score-lbl" style="font-weight:800;">⭐ Call Quality Score</div>
        <div class="qt-score-trend {_composite_color}">{_composite_label}</div>
      </div>
      <div class="qt-score-card">
        <div class="qt-score-val {enc_color}">{f"{enc_avg:.1f}%" if enc_avg is not None else "—"}</div>
        <div class="qt-score-lbl">%ENC — Efectividad</div>
        <div class="qt-score-trend {"qt-ok" if enc_avg and enc_avg >= 80 else "qt-bad"}">
          {"✓ En objetivo" if enc_avg and enc_avg >= 80 else "↓ Por debajo de meta 80%"}
        </div>
      </div>
      <div class="qt-score-card">
        <div class="qt-score-val {ec_color}">{f"{ec_avg:.1f}%" if ec_avg is not None else "—"}</div>
        <div class="qt-score-lbl">%EC — Calidad efectiva</div>
        <div class="qt-score-trend {"qt-ok" if ec_avg and ec_avg >= 80 else "qt-bad"}">
          {"✓ En objetivo" if ec_avg and ec_avg >= 80 else "↓ Por debajo de meta 80%"}
        </div>
      </div>
      <div class="qt-score-card">
        <div class="qt-score-val" style="color:#1A1A2E">{n_calls}</div>
        <div class="qt-score-lbl">Llamadas analizadas</div>
        <div class="qt-score-trend" style="color:#888">Período cargado</div>
      </div>
      <div class="qt-score-card">
        <div class="qt-score-val {"qt-bad" if n_below >= 5 else "qt-warn"}">{n_below}</div>
        <div class="qt-score-lbl">Dimensiones bajo 80%</div>
        <div class="qt-score-trend {"qt-bad" if n_below >= 5 else "qt-warn"}">
          {"Atención: zona roja" if n_below >= 5 else "Trabajar en progreso"}
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    if not below:
        st.success("Todas las dimensiones superan el 80%. Seguí así — el objetivo ahora es mantener y buscar el 90%.")
        return

    st.markdown(f'<div class="qt-separator">Dimensiones a trabajar — {len(below)} por debajo del 80%</div>', unsafe_allow_html=True)

    # ── Trainer cards — 2-column card grid ───────────────────────────────────
    def _render_trainer_card(rank, col, label, cmt_col, impact, tip):
        score = dim_scores[col]
        pct   = round(score * 100, 1)

        if pct < 40:
            rank_bg, rank_fg, pill_bg, pill_fg = "rgba(229,51,42,0.10)", "#FF4D2E", "rgba(229,51,42,0.10)", "#FF4D2E"
        elif pct < 60:
            rank_bg, rank_fg, pill_bg, pill_fg = "rgba(255,113,36,0.10)", "#D95A10", "rgba(255,113,36,0.10)", "#D95A10"
        else:
            rank_bg, rank_fg, pill_bg, pill_fg = "rgba(255,113,36,0.08)", "#D95A10", "rgba(255,113,36,0.08)", "#D95A10"

        spark = sparkline_data(col)
        spark_html = ""
        if spark:
            for i, v in enumerate(spark):
                if v is None:
                    dot_bg, dot_txt = "rgba(255,255,255,0.95)", "#bbb"
                elif v >= 80:
                    dot_bg, dot_txt = "rgba(111,242,75,0.08)", "#7ED321"
                elif v >= 60:
                    dot_bg, dot_txt = "rgba(255,113,36,0.10)", "#D95A10"
                else:
                    dot_bg, dot_txt = "rgba(229,51,42,0.10)", "#FF4D2E"
                wlabel = f"W{i+1}"
                val_str = f"{round(v)}%" if v is not None else "—"
                spark_html += f'<div class="qt-spark-dot" style="background:{dot_bg};color:{dot_txt}" title="{wlabel}: {val_str}">{val_str}</div>'

        examples = get_examples(col, cmt_col) if cmt_col else []
        ex_html = ""
        for ex in examples:
            ex_html += f'<div class="qt-example">&#8220;{ex[:200]}{"…" if len(ex) > 200 else ""}&#8221;</div>'

        st.markdown(f"""
        <div class="qt-trainer-card" style="height:100%;">
          <div class="qt-card-header">
            <div class="qt-rank-badge" style="background:{rank_bg};color:{rank_fg}">#{rank}</div>
            <div class="qt-dim-label">{label}</div>
            <div class="qt-score-pill" style="background:{pill_bg};color:{pill_fg}">{pct}% <span style="font-weight:400;font-size:11px">/ 80% meta</span></div>
          </div>
          <div class="qt-card-body">
            {"<div class='qt-sparkline-row'>" + spark_html + "<span class='qt-spark'>tendencia semanal</span></div>" if spark_html else ""}
            <div class="qt-section-title">Por qué afecta tu cartera</div>
            <div class="qt-impact-box">{impact}</div>
            <div class="qt-section-title">Cómo mejorarlo — táctica concreta</div>
            <div class="qt-tip-box">{tip}</div>
            {"<div class='qt-section-title'>Ejemplos reales de tus llamadas donde falló</div>" + ex_html if ex_html else ""}
          </div>
        </div>
        """, unsafe_allow_html=True)

    # Render cards in pairs — 2 per row
    for i in range(0, len(below), 2):
        left_col, right_col = st.columns(2)
        rank_l = i + 1
        col_l, label_l, cmt_col_l, impact_l, tip_l = below[i]
        with left_col:
            _render_trainer_card(rank_l, col_l, label_l, cmt_col_l, impact_l, tip_l)
        if i + 1 < len(below):
            rank_r = i + 2
            col_r, label_r, cmt_col_r, impact_r, tip_r = below[i + 1]
            with right_col:
                _render_trainer_card(rank_r, col_r, label_r, cmt_col_r, impact_r, tip_r)
        st.markdown("<div style='margin-bottom:12px'></div>", unsafe_allow_html=True)

    # ── Strengths ─────────────────────────────────────────────────────────────
    if above:
        above_labels = ", ".join(
            f"<b>{label}</b> ({round(dim_scores[col]*100)}%)"
            for col, label, *_ in above
        )
        st.markdown(f"""
        <div class="qt-above-section">
          ✓ <b>Dimensiones en objetivo (≥80%):</b> {above_labels}<br>
          Estos puntos son tu base sólida. No los descuides aunque el foco esté en lo rojo.
        </div>
        """, unsafe_allow_html=True)


def page_productivity_heatmap():
    render_header("Productivity HeatMap", f"Frecuencia de palancas y conversión semanal · {FARMER_NAME}")

    # ── Load Productivity sheet from main Excel ───────────────────────────────
    @st.cache_data(ttl=3000, show_spinner=False)
    def _load_hm_data(excel_path):
        if not os.path.exists(excel_path):
            return pd.DataFrame(), {}
        try:
            df = pd.read_excel(excel_path, sheet_name="Productivity", header=0)
        except Exception:
            return pd.DataFrame(), {}

        df.columns = [str(c).strip() for c in df.columns]

        if "Code" not in df.columns or "Week" not in df.columns:
            return pd.DataFrame(), {}

        df["_week"] = pd.to_datetime(df["Week"], errors="coerce").dt.strftime("%Y-%m-%d")
        mondays = sorted(df["_week"].dropna().unique())
        wmap = {d: f"W{i+1}" for i, d in enumerate(mondays)}
        df["_wl"] = df["_week"].map(wmap)
        return df[df["_wl"].notna()].copy(), wmap

    df, wmap = _load_hm_data(EXCEL_FILE)

    if df.empty:
        st.info(
            "No se encontró la hoja **Productivity** en el archivo base. "
            "Verificá que la hoja existe y tiene datos desde A1."
        )
        return

    WEEKS      = sorted(wmap.values(), key=lambda w: int(w[1:]))
    WEEK_DATES = [
        pd.to_datetime(d).strftime("%d-%b").lstrip("0")
        for d in sorted(wmap.keys())
    ]
    TOTALS = {wl: len(df[df["_wl"] == wl]) for wl in WEEKS}

    # ── Helper: count palanca per week ────────────────────────────────────────
    def freq(col, value="SI", multi=False):
        """Returns list of (count, total) per week."""
        out = []
        for wl in WEEKS:
            w = df[df["_wl"] == wl]
            t = TOTALS[wl]
            if multi:
                c = w[col].fillna("").apply(lambda x: value in str(x)).sum()
            else:
                c = (w[col].astype(str).str.upper() == value.upper()).sum()
            out.append((int(c), t))
        return out

    # ── Palanca groups ────────────────────────────────────────────────────────
    GROUPS = [
        {
            "label": "🟢 Comercial",
            "color_header": PALETTE["slate_indigo"],
            "color_cell_lo": "rgba(59,72,131,0.15)",
            "color_cell_hi": PALETTE["slate_indigo"],
            "color_text_lo": PALETTE["slate_indigo"],
            "color_text_hi": "rgba(255,255,255,0.9)",
            "palancas": [
                ("Markdown",          freq("Markdown")),
                ("Ads",               freq("Ads")),
                ("Conectividad",      freq("Conectividad")),
                ("Catálogo al 100%",  freq("Ajustes Catálogo", "Catálogo al 100%",         multi=True)),
                ("Price parity",      freq("Ajustes Catálogo", "Igualdad en precios",       multi=True)),
                ("Fotos",             freq("Ajustes Catálogo", "Fotos",                     multi=True)),
                ("Disponibilidad",    freq("Ajustes Catálogo", "Disponibilidad del producto",multi=True)),
                ("Catálogo PDF",      freq("Ajustes Catálogo", "Catálogo PDF",              multi=True)),
                ("Combos",            freq("Ajustes Catálogo", "Generación de combos",      multi=True)),
                ("Purch. Experience", freq("Ajustes Catálogo", "Purchasing Experience",     multi=True)),
            ],
        },
        {
            "label": "🟡 Operativo",
            "color_header": PALETTE["emerald_dark"],
            "color_cell_lo": "rgba(111,242,75,0.06)",
            "color_cell_hi": PALETTE["emerald_dark"],
            "color_text_lo": PALETTE["emerald_dark"],
            "color_text_hi": "rgba(255,255,255,0.9)",
            "palancas": [
                ("Catálogo general", freq("Catálogo")),
                ("Cancelaciones",    freq("Cancelaciones")),
                ("DR",               freq("DR")),
                ("Tiempos",          freq("Tiempos")),
            ],
        },
        {
            "label": "🔴 Incendios",
            "color_header": PALETTE["tangerine_dark"],
            "color_cell_lo": "rgba(255,113,36,0.10)",
            "color_cell_hi": PALETTE["tangerine_dark"],
            "color_text_lo": PALETTE["tangerine_dark"],
            "color_text_hi": "rgba(255,255,255,0.9)",
            "palancas": [
                ("Pains del aliado", freq("Pains del aliado")),
                ("Churn",            freq("Churn")),
                ("On Hold",          freq("On Hold")),
            ],
        },
    ]

    html = ""  # acumulador reutilizado por el bloque de Benchmark personal debajo


    # ── Benchmark personal summary — top 3 palancas con mejor récord ─────────
    _bm_records = []
    for g in GROUPS:
        for name, freqs in g["palancas"]:
            _pcts = [round(c / t * 100) if t else 0 for c, t in freqs]
            _best = max(_pcts) if _pcts else 0
            _best_wk = WEEKS[_pcts.index(_best)] if _pcts and _best > 0 else "-"
            if _best > 0:
                _bm_records.append((name, _best, _best_wk))
    _bm_records.sort(key=lambda x: x[1], reverse=True)
    if _bm_records:
        _bm_items_html = "".join(
            f'<div style="display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid rgba(255,255,255,0.95);">' 
            f'<span style="font-size:12px;color:#1A1A2E;min-width:130px;">{bm_name}</span>' 
            f'<span style="font-size:13px;font-weight:700;color:#7ED321;">{bm_pct}%</span>' 
            f'<span style="font-size:11px;color:#6B7280;">mejor en {bm_wk}</span>' 
            f'🏆</div>'
            for bm_name, bm_pct, bm_wk in _bm_records[:6]
        )
        html += f'''
        <div style="margin-top:1.2rem;background:rgba(111,242,75,0.06);border:1px solid #7ED321;border-radius:8px;padding:12px 16px;">
        <div style="font-size:11px;font-weight:700;color:#7ED321;letter-spacing:.06em;text-transform:uppercase;margin-bottom:8px;">
            📈 Benchmark personal — tus récords por palanca
        </div>
        {_bm_items_html}
        </div>
        '''

    if html:
        st.markdown(html, unsafe_allow_html=True)

    # ── Heatmap visual real: eje Y = semanas, eje X = palancas, color = intensidad ──
    # (Complementa la tabla numérica de arriba con una lectura visual rápida
    # de qué palancas concentran más uso a lo largo del tiempo.)
    # NOTA: la variable local 'html' de esta función (acumulador de strings HTML,
    # más arriba) shadowea el módulo html importado globalmente — por eso acá
    # se usa un import explícito con alias seguro en vez de html.escape().
    import html as _html_mod
    _all_palancas = [(name, freqs) for g in GROUPS for name, freqs in g["palancas"]]
    _hm_n_cols = len(_all_palancas)
    _hm_n_rows = len(WEEKS)

    if _hm_n_rows > 0 and _hm_n_cols > 0:
        _HM_CELL_W, _HM_CELL_H = 78, 28
        _HM_LABEL_W = 56
        _HM_TOP_PAD = 70
        _hm_svg_w = _HM_LABEL_W + _hm_n_cols * _HM_CELL_W
        _hm_svg_h = _HM_TOP_PAD + _hm_n_rows * _HM_CELL_H

        _hm_parts = [f'<svg width="{_hm_svg_w}" height="{_hm_svg_h}" viewBox="0 0 {_hm_svg_w} {_hm_svg_h}" xmlns="http://www.w3.org/2000/svg">']

        # Headers de palanca — rotados 45° para que entren nombres largos en columnas angostas
        for ci, (name, freqs) in enumerate(_all_palancas):
            cx = _HM_LABEL_W + ci * _HM_CELL_W + _HM_CELL_W / 2
            _hm_parts.append(
                f'<text x="{cx:.0f}" y="{_HM_TOP_PAD - 8}" text-anchor="start" font-size="9" font-weight="700" '
                f'fill="#6B7280" transform="rotate(-40 {cx:.0f} {_HM_TOP_PAD - 8})">{_html_mod.escape(name)}</text>'
            )

        # Pre-computar % por palanca (columna) — mismo orden que freqs por semana
        _hm_pcts_by_col = [
            [round(c / t * 100) if t else 0 for c, t in freqs]
            for _, freqs in _all_palancas
        ]

        for ri, wl in enumerate(WEEKS):
            row_y = _HM_TOP_PAD + ri * _HM_CELL_H
            _hm_parts.append(
                f'<text x="{_HM_LABEL_W - 8}" y="{row_y + _HM_CELL_H/2 + 3:.0f}" text-anchor="end" '
                f'font-size="10" fill="#1A1A2E" font-weight="600">{wl}</text>'
            )
            for ci in range(_hm_n_cols):
                pct = _hm_pcts_by_col[ci][ri] if ri < len(_hm_pcts_by_col[ci]) else 0
                cx = _HM_LABEL_W + ci * _HM_CELL_W
                t = min(pct / 100, 1.0)
                # Intensidad: de blanco/gris (0%) a azul oscuro (100%)
                _r = int(241 - t * (241 - 27))
                _g = int(242 - t * (242 - 63))
                _b = int(245 - t * (245 - 139))
                _fill = f"#{_r:02x}{_g:02x}{_b:02x}"
                _text_color = "#FFFFFF" if t > 0.5 else "#6B7280"
                _hm_parts.append(
                    f'<rect x="{cx+2}" y="{row_y+2}" width="{_HM_CELL_W-4}" height="{_HM_CELL_H-4}" '
                    f'rx="4" fill="{_fill}"/>'
                    f'<text x="{cx + _HM_CELL_W/2:.0f}" y="{row_y + _HM_CELL_H/2 + 3:.0f}" '
                    f'text-anchor="middle" font-size="9" font-weight="700" fill="{_text_color}">{pct}%</text>'
                )
        _hm_parts.append('</svg>')
        _heatmap_svg = "".join(_hm_parts)

        st.markdown(
            f'<div style="margin-top:1.5rem;">'
            f'<div style="font-size:11px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;'
            f'color:{PALETTE["slate_indigo"]};margin-bottom:10px;">🌡️ Mapa de calor · intensidad de uso por palanca</div>'
            f'<div style="overflow-x:auto;background:rgba(255,255,255,0.92);border-radius:10px;padding:10px;">{_heatmap_svg}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── Gráfica: semana de mayor uso por palanca ───────────────────────────
        _peak_rows = []
        for name, freqs in _all_palancas:
            pcts = [round(c / t * 100) if t else 0 for c, t in freqs]
            if pcts and max(pcts) > 0:
                _bw = pcts.index(max(pcts))
                _peak_rows.append((name, WEEKS[_bw], max(pcts)))
        _peak_rows.sort(key=lambda x: x[2], reverse=True)

        if _peak_rows:
            _peak_max = max(p[2] for p in _peak_rows)
            _peak_bars_html = "".join(
                f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">'
                f'<span style="font-size:11px;color:#1A1A2E;min-width:150px;font-weight:600;">{_html_mod.escape(p_name)}</span>'
                f'<div style="flex:1;background:rgba(0,0,0,0.05);border-radius:6px;height:18px;position:relative;overflow:hidden;">'
                f'<div style="position:absolute;left:0;top:0;height:100%;width:{(p_pct/_peak_max*100) if _peak_max else 0:.0f}%;'
                f'background:{PALETTE["slate_indigo"]};border-radius:6px;"></div>'
                f'</div>'
                f'<span style="font-size:11px;font-weight:700;color:{PALETTE["slate_indigo"]};min-width:36px;">{p_pct}%</span>'
                f'<span style="font-size:10px;color:#6B7280;min-width:32px;">{p_week}</span>'
                f'</div>'
                for p_name, p_week, p_pct in _peak_rows
            )
            st.markdown(
                f'<div style="margin-top:1.2rem;background:rgba(255,255,255,0.92);border-radius:10px;padding:14px 16px;">'
                f'<div style="font-size:11px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;'
                f'color:{PALETTE["slate_indigo"]};margin-bottom:10px;">📅 Semana de mayor uso por palanca</div>'
                f'{_peak_bars_html}'
                f'</div>',
                unsafe_allow_html=True,
            )



# =========================
# EARNINGS CALCULATOR
# =========================

def page_earnings_calculator():
    render_header("Earnings Calculator", "Personal KPI Calculator · Rappi")

    raw = load_earnings_data()

    # ── Valores iniciales desde Excel (si existe), editables desde el dash ────
    _ads_target_xl    = to_number(cell(raw, 2, 1)) if not raw.empty else 0
    _ads_result_xl    = to_number(cell(raw, 2, 2)) if not raw.empty else 0
    _churn_target_xl  = to_number(cell(raw, 2, 3)) if not raw.empty else 0
    _churn_result_xl  = to_number(cell(raw, 2, 4)) if not raw.empty else 0
    _md_target_xl     = to_number(cell(raw, 2, 5)) if not raw.empty else 0
    _md_result_xl     = to_number(cell(raw, 2, 6)) if not raw.empty else 0
    _mdpro_target_xl  = to_number(cell(raw, 2, 7)) if not raw.empty else 0
    _mdpro_result_xl  = to_number(cell(raw, 2, 8)) if not raw.empty else 0
    _prod_target_xl   = to_number(cell(raw, 2, 9)) if not raw.empty else 0
    _prod_result_xl   = to_number(cell(raw, 2, 10)) if not raw.empty else 0
    _transport_xl     = to_number(cell(raw, 8, 2)) if not raw.empty else 0

    with st.expander("✏️ Editar resultados del mes", expanded=False):
        ec1, ec2 = st.columns(2)
        with ec1:
            ads_target = st.number_input("ADS Target (USD)", value=float(_ads_target_xl), step=100.0, key="ec_ads_target")
            ads_result = st.number_input("ADS Result (USD)", value=float(_ads_result_xl), step=100.0, key="ec_ads_result")
            md_target  = st.number_input("MD Target (%)", value=float(_md_target_xl), step=0.001, format="%.4f", key="ec_md_target")
            md_result  = st.number_input("MD Result (%)", value=float(_md_result_xl), step=0.001, format="%.4f", key="ec_md_result")
            mdpro_target = st.number_input("MD PRO Target (%)", value=float(_mdpro_target_xl), step=0.001, format="%.4f", key="ec_mdpro_target")
            mdpro_result = st.number_input("MD PRO Result (%)", value=float(_mdpro_result_xl), step=0.001, format="%.4f", key="ec_mdpro_result")
        with ec2:
            churn_target = st.number_input("Churn Target (tasa)", value=float(_churn_target_xl), step=0.001, format="%.4f", key="ec_churn_target")
            churn_result = st.number_input("Churn Result (tasa)", value=float(_churn_result_xl), step=0.001, format="%.4f", key="ec_churn_result")
            prod_target  = st.number_input("Productividad Target", value=float(_prod_target_xl), step=1.0, key="ec_prod_target")
            prod_result  = st.number_input("Productividad Result", value=float(_prod_result_xl), step=1.0, key="ec_prod_result")
            transport    = st.number_input("Transporte + Conexión (COP)", value=float(_transport_xl), step=1000.0, key="ec_transport")

    ads_ach   = ads_result / ads_target if ads_target else 0
    churn_ach = churn_target / churn_result if churn_result else 0   # tasa de churn: menos es mejor
    md_ach    = md_result / md_target if md_target else 0
    mdpro_ach = mdpro_result / mdpro_target if mdpro_target else 0
    prod_ach  = prod_result / prod_target if prod_target else 0

    # ── Caps por KPI según plan de incentivos: ADS tope 100%, MD/MD PRO/Churn tope 150% ──
    _ads_ach_capped   = min(ads_ach, 1.0)
    _md_ach_capped    = min(md_ach, 1.5)
    _mdpro_ach_capped = min(mdpro_ach, 1.5)
    _churn_ach_capped = min(churn_ach, 1.5)

    variable_percent = 0
    _prod_qualifies = prod_ach >= 0.9
    if _prod_qualifies:
        variable_percent = (
            _ads_ach_capped   * 0.35
            + _churn_ach_capped * 0.25
            + _md_ach_capped    * 0.20
            + _mdpro_ach_capped * 0.20
        )

    def render_kpi_card(name, target, result, achievement, target_formatter, result_formatter):
        import math as _m
        pct       = min(achievement, 1.0)
        over      = achievement - 1.0 if achievement > 1.0 else 0.0
        size      = 148
        stroke    = 18
        r         = (size - stroke) / 2
        circ      = 2 * _m.pi * r
        filled    = round(circ * pct, 1)
        gap       = round(circ - filled, 1)
        offset    = round(circ * 0.25, 1)
        cx = cy   = size / 2

        if achievement >= 1.0:
            arc_color = "#7ED321"
        elif achievement >= 0.7:
            arc_color = "#FF7124"
        else:
            arc_color = "#FF4D2E"

        pct_label  = f"+{round(over*100)}%" if over > 0 else f"{round(achievement*100)}%"
        over_label = "over goal!" if over > 0 else "achieved"

        # Color accent per KPI
        accent_map = {
            "Ads":         "#FF7124",
            "MD":          "#1B3F8B",
            "MD PRO":      "#7ED321",
            "Churn":       "#FF4D2E",
            "Productivity":"#7ED321",
        }
        accent = accent_map.get(name, "#1B3F8B")

        donut_svg = (
            f'<svg width="{size}" height="{size}" viewBox="0 0 {size} {size}">'
            # track
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="rgba(255,255,255,0.95)" stroke-width="{stroke}"/>'
            # progress arc
            f'<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{arc_color}" stroke-width="{stroke}" '
            f'stroke-dasharray="{filled} {gap}" stroke-dashoffset="{offset}" stroke-linecap="round"/>'
            # center % text
            f'<text x="{cx}" y="{cy - 6}" text-anchor="middle" font-size="20" font-weight="900" fill="{arc_color}">{pct_label}</text>'
            f'<text x="{cx}" y="{cy + 12}" text-anchor="middle" font-size="9" fill="rgba(107,114,128,0.60)">{over_label}</text>'
            f'</svg>'
        )

        st.markdown(f"""
<div class="kpi-card" style="
    background:rgba(255,255,255,0.9);
    border:1px solid rgba(0,0,0,0.07);
    box-shadow:0 10px 30px rgba(0,0,0,0.05);
    transition:box-shadow .22s,transform .22s;
    min-height:0;
    padding:22px 20px 18px;
">
  <div style="font-size:13px;font-weight:900;text-transform:uppercase;color:#6B7280;margin-bottom:14px;letter-spacing:.04em;">{name}</div>
  <div style="display:flex;align-items:center;gap:16px;">
    <div style="flex-shrink:0;">{donut_svg}</div>
    <div style="flex:1;min-width:0;">
      <div style="background:rgba(191,255,0,.07);border:1.5px solid rgba(191,255,0,.35);border-radius:14px;padding:10px 12px;margin-bottom:10px;">
        <div style="font-size:10px;font-weight:900;text-transform:uppercase;color:#6B7280;">Target</div>
        <div style="font-size:15px;font-weight:900;color:#1A1A2E;margin-top:4px;line-height:1.15;overflow-wrap:anywhere;">{target_formatter(target)}</div>
      </div>
      <div style="background:rgba(139,157,255,.07);border:1.5px solid rgba(139,157,255,.35);border-radius:14px;padding:10px 12px;">
        <div style="font-size:10px;font-weight:900;text-transform:uppercase;color:#6B7280;">Result</div>
        <div style="font-size:15px;font-weight:900;color:{accent};margin-top:4px;line-height:1.15;overflow-wrap:anywhere;">{result_formatter(result)}</div>
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

    def render_variable_card(value):
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">Variable %</div>
            <div class="variable-box">
                <div class="kpi-label">Compiled Result</div>
                <div class="variable-value">{fmt_percent0(value)}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    top1, top2, top3 = st.columns(3)
    with top1:
        render_kpi_card("Ads", ads_target, ads_result, ads_ach, fmt_usd, fmt_usd)
    with top2:
        render_kpi_card("MD", md_target, md_result, md_ach, fmt_percent2, fmt_percent2)
    with top3:
        render_kpi_card("MD PRO", mdpro_target, mdpro_result, mdpro_ach, fmt_percent2, fmt_percent2)

    bottom1, bottom2, bottom3 = st.columns(3)
    with bottom1:
        render_kpi_card("Churn", churn_target, churn_result, churn_ach, fmt_percent2, fmt_percent2)
    with bottom2:
        render_kpi_card("Productivity", prod_target, prod_result, prod_ach, fmt_number, fmt_number)
    with bottom3:
        render_variable_card(variable_percent)

    if not _prod_qualifies:
        st.warning(f"⚠️ Productividad en {fmt_percent0(prod_ach)} — por debajo del mínimo de 90%. No se gana variable este mes (Variable % = 0).")

    st.markdown("## Revenue Share ADS")
    st.caption("Bono adicional a tu variable · se gana cuando tu cumplimiento de ADS supera 90% Y tu cumplimiento de Markdown es ≥ 90%. Cap: 2.000 USD mensuales.")

    _rs_md_qualifies = md_ach >= 0.90
    bucket1 = max(min(ads_result, ads_target) - ads_target * 0.9, 0) * 0.10
    bucket2 = max(min(ads_result, ads_target * 1.2) - ads_target, 0) * 0.20
    bucket3 = max(ads_result - ads_target * 1.2, 0) * 0.30
    total_comm_usd_uncapped = bucket1 + bucket2 + bucket3
    total_comm_usd = min(total_comm_usd_uncapped, 2000.0) if _rs_md_qualifies else 0.0
    _rs_capped_by_md   = not _rs_md_qualifies and total_comm_usd_uncapped > 0
    _rs_capped_by_cap  = _rs_md_qualifies and total_comm_usd_uncapped > 2000.0

    if not _rs_md_qualifies:
        st.warning(f"⚠️ Markdown en {fmt_percent0(md_ach)} — por debajo del 90% requerido. Revenue Share ADS no se desbloquea este mes aunque ADS esté en {fmt_percent0(ads_ach)}.")
    elif _rs_capped_by_cap:
        st.info(f"ℹ️ Revenue Share ADS calculado: {fmt_usd(total_comm_usd_uncapped)} — topeado al máximo mensual de {fmt_usd(2000)}.")

    # ── Proyección: cuánto falta para el siguiente bucket ─────────────────────
    _b1_threshold   = ads_target * 0.90  # entrada bucket 1
    _b2_threshold   = ads_target * 1.00  # entrada bucket 2
    _b3_threshold   = ads_target * 1.20  # entrada bucket 3

    if ads_result < _b1_threshold:
        _next_bucket_name  = "Bucket 1 (90% del target)"
        _gap_to_next        = _b1_threshold - ads_result
        _next_color         = "#FF4D2E"
        _next_note          = f"Todavía no entraste al rango comisionado. Necesitás cerrar {fmt_usd(_gap_to_next)} más."
    elif ads_result < _b2_threshold:
        _next_bucket_name  = "Bucket 2 (100% del target)"
        _gap_to_next        = _b2_threshold - ads_result
        _next_color         = "#FF7124"
        _next_note          = f"Con {fmt_usd(_gap_to_next)} más entrás al 20% de comisión. Bucket 1 acumulado: {fmt_usd(bucket1)}."
    elif ads_result < _b3_threshold:
        _next_bucket_name  = "Bucket 3 (120% del target)"
        _gap_to_next        = _b3_threshold - ads_result
        _next_color         = "#FF7124"
        _next_note          = f"Con {fmt_usd(_gap_to_next)} más entrás al 30% de comisión. Buckets 1+2 acumulados: {fmt_usd(bucket1 + bucket2)}."
    else:
        _next_bucket_name  = "Superaste los 3 buckets 🏆"
        _gap_to_next        = 0
        _next_color         = "#7ED321"
        _next_note          = f"Estás en el máximo tier de ADS. Total Revenue Share (antes de cap): {fmt_usd(total_comm_usd_uncapped)}."

    # Barra de progreso hacia el siguiente bucket
    if _gap_to_next > 0:
        _current_in_range = ads_result - (_b1_threshold if ads_result >= _b1_threshold else 0)
        _range_size = _gap_to_next + max(_current_in_range, 0)
        _bucket_pct = round(min(max(_current_in_range, 0) / _range_size * 100 if _range_size > 0 else 0, 100))
    else:
        _bucket_pct = 100

    st.markdown(f"""
<div style="background:rgba(255,255,255,0.9);border:1px solid rgba(0,0,0,0.07);border-radius:20px;padding:20px 24px;margin-bottom:16px;
    box-shadow:0 4px 14px rgba(0,0,0,0.05);">
    <div style="font-size:11px;font-weight:900;text-transform:uppercase;color:#6B7280;margin-bottom:10px;">
        🎯 Próximo hito de comisión
    </div>
    <div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap;">
        <div>
            <div style="font-size:22px;font-weight:900;color:{_next_color};">{_next_bucket_name}</div>
            <div style="font-size:15px;font-weight:800;color:#1A1A2E;margin-top:4px;">
                {"Faltan " + fmt_usd(_gap_to_next) + " de ADS Revenue" if _gap_to_next > 0 else "¡Máximo nivel alcanzado!"}
            </div>
        </div>
        <div style="flex:1;min-width:200px;">
            <div style="height:10px;border-radius:8px;background:rgba(255,255,255,0.95);overflow:hidden;margin-bottom:6px;">
                <div style="width:{_bucket_pct}%;height:100%;background:{_next_color};border-radius:8px;transition:width .4s;"></div>
            </div>
            <div style="font-size:11px;color:#6B7280;">{_next_note}</div>
        </div>
    </div>
</div>
    """, unsafe_allow_html=True)

    buckets = [
        ("Bucket 1: 90% to 100% (10%)", bucket1 if _rs_md_qualifies else 0),
        ("Bucket 2: 100% to 120% (20%)", bucket2 if _rs_md_qualifies else 0),
        ("Bucket 3: More than 120% (30%)", bucket3 if _rs_md_qualifies else 0),
        ("Total Revenue Share USD (cap 2k)", total_comm_usd),
    ]

    bcols = st.columns(4)

    for col, (label, value) in zip(bcols, buckets):
        box_class = "zero-box" if value == 0 else "result-box"
        with col:
            st.markdown(f"""
            <div class="bucket-card">
                <div class="bucket-title">{label}</div>
                <div class="{box_class}">
                    <div class="kpi-label">Result</div>
                    <div class="kpi-value">{fmt_usd(value)}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("## Compiled Result")

    _PERFORMANCE_TIERS = [
        (1.0,  "LEGEND",       "#7ED321", "Top Performer, superando las metas y destacándose por su excelencia."),
        (0.8,  "PROFESSIONAL", "#3D64B8", "Cumple con las expectativas y mantiene un rendimiento constante."),
        (0.6,  "ROOKIE",       "#FF7124", "Resultados por debajo del nivel óptimo con un potencial claro de mejora."),
        (0.0,  "RED FLAG",     "#FF4D2E", "Resultados por debajo de las expectativas que necesitan atención inmediata."),
    ]
    _tier_name, _tier_color, _tier_desc = next(
        (name, color, desc) for threshold, name, color, desc in _PERFORMANCE_TIERS
        if variable_percent >= threshold
    )

    st.markdown(f"""
<div style="background:rgba(255,255,255,0.9);border:1px solid rgba(0,0,0,0.07);border-radius:20px;padding:22px 26px;margin-bottom:16px;
    box-shadow:0 4px 14px rgba(0,0,0,0.05);display:flex;align-items:center;gap:20px;flex-wrap:wrap;">
    <div style="background:{_tier_color};color:#FFFFFF;font-weight:900;font-size:13px;letter-spacing:.06em;
        padding:8px 18px;border-radius:30px;text-transform:uppercase;">{_tier_name}</div>
    <div style="flex:1;min-width:200px;">
        <div style="font-size:26px;font-weight:900;color:{_tier_color};">{fmt_percent0(variable_percent)}</div>
        <div style="font-size:12px;color:#6B7280;margin-top:2px;">{_tier_desc}</div>
    </div>
</div>
    """, unsafe_allow_html=True)

    st.markdown("## Salary Summary")

    _BASE_SALARY_COP = 2_000_000
    _HEALTH_PENSION_DEDUCTION_COP = 244_000
    _VARIABLE_BASE_COP = 510_000
    salary = _BASE_SALARY_COP - _HEALTH_PENSION_DEDUCTION_COP
    variable_cop = variable_percent * _VARIABLE_BASE_COP
    commission_cop = total_comm_usd * COP_PER_USD
    gross_total = salary + transport + variable_cop + commission_cop
    net_total = gross_total

    salary_items = [
        ("Salary (neto de salud/pensión)", salary, False),
        ("Transport + Connection", transport, False),
        ("Variable", variable_cop, False),
        ("Revenue Share ADS", commission_cop, False),
        ("Gross Total", gross_total, False),
        ("Net Total", net_total, True),
    ]

    scols = st.columns(3)

    for i, (label, value, is_net) in enumerate(salary_items):
        card_class = "salary-card net-card" if is_net else "salary-card"
        with scols[i % 3]:
            st.markdown(f"""
            <div class="{card_class}">
                <div class="salary-label">{label}</div>
                <div class="salary-value">{fmt_cop(value)}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="legend-box">
        Salario base: {fmt_cop(_BASE_SALARY_COP)} − {fmt_cop(_HEALTH_PENSION_DEDUCTION_COP)} (salud + pensión) = {fmt_cop(salary)} neto ·
        Variable: {fmt_percent0(variable_percent)} sobre base de {fmt_cop(_VARIABLE_BASE_COP)} ·
        Variable cap: ADS 100% / MD, MD PRO y Churn 150% · Qualifier productividad mínimo 90% ·
        Revenue Share ADS requiere MD ≥ 90%, cap {fmt_usd(2000)}/mes
    </div>
    """, unsafe_allow_html=True)


# =========================
# BRAND FINDER
# =========================

def _split_category_and_stickers(category_value):
    raw = clean(category_value, "-")
    if raw == "-":
        return "-", []

    # Category values can arrive as:
    # "Hamburguesa y Americana", "Empanadas y Pizza", "Main | Sticker", "Main, Sticker".
    # The first piece is the main category; the rest are category stickers.
    normalized_raw = re.sub(r"\s+", " ", raw).strip()
    parts = []
    if "|" in normalized_raw:
        parts = [p.strip() for p in normalized_raw.split("|") if p.strip()]
    elif "," in normalized_raw:
        parts = [p.strip() for p in normalized_raw.split(",") if p.strip()]
    elif re.search(r"\s+y\s+", normalized_raw, flags=re.IGNORECASE):
        parts = [p.strip() for p in re.split(r"\s+y\s+", normalized_raw, flags=re.IGNORECASE) if p.strip()]
    else:
        parts = [normalized_raw]

    if not parts:
        return normalized_raw, []

    return parts[0], parts[1:]



def _normalize_email_for_match(value):
    text = clean(value, "").strip().lower()
    if text in ["", "-", "nan", "none"] or "@" not in text:
        return ""
    return text


def _normalize_phone_for_match(value):
    text = fmt_contact_number(value)
    digits = re.sub(r"\D", "", clean(text, ""))
    # Avoid false positives with short extensions or invalid placeholders.
    if len(digits) < 7:
        return ""
    return digits


def _normalize_manager_for_match(value):
    text = _clean_name_for_match(value)
    if text in ["", "-", "nan", "none", "sin manager", "no manager", "na", "n a"]:
        return ""
    # One very short token is too weak and creates noise.
    if len(text) < 4:
        return ""
    return text


def get_multibrand_matches(row, brand_id, max_items=6):
    """
    Detects likely multibrand accounts inside Growth OS using repeated contact signals.
    Confidence logic:
    - Email or phone match = strong signal.
    - Manager-only match = possible signal.
    - 2+ matching fields = high confidence.
    """
    df = load_growth_data()
    if df.empty:
        return []

    id_col = get_id_column_name(df)
    if not id_col:
        return []

    name_col = _first_existing_col(df, ["name", "brand name", "restaurant name"])
    email_col = _first_existing_col(df, ["email", "mail", "contact mail"])
    phone_col = _first_existing_col(df, ["contact number", "phone", "contact"])
    manager_col = _first_existing_col(df, ["manager", "restaurant manager", "account manager"])

    target_id = normalize_brand_id(brand_id)
    target_email = _normalize_email_for_match(get_from_row(row, ["email", "mail", "contact mail"], ""))
    target_phone = _normalize_phone_for_match(get_from_row(row, ["contact number", "phone", "contact"], ""))
    target_manager = _normalize_manager_for_match(get_from_row(row, ["manager", "restaurant manager", "account manager"], ""))

    matches = []
    seen = set()

    for _, other in df.iterrows():
        other_id = normalize_brand_id(other.get(id_col, ""))
        if not other_id or other_id == target_id or other_id in seen:
            continue

        reasons = []
        strong = 0

        if email_col and target_email and _normalize_email_for_match(other.get(email_col, "")) == target_email:
            reasons.append("Email")
            strong += 1
        if phone_col and target_phone and _normalize_phone_for_match(other.get(phone_col, "")) == target_phone:
            reasons.append("Phone")
            strong += 1
        if manager_col and target_manager and _normalize_manager_for_match(other.get(manager_col, "")) == target_manager:
            reasons.append("Manager")

        if not reasons:
            continue

        confidence = "High" if strong >= 1 or len(reasons) >= 2 else "Possible"
        # Manager-only matches are useful, but less reliable.
        if reasons == ["Manager"]:
            confidence = "Possible"

        other_name = clean(other.get(name_col, "-"), "-") if name_col else "-"
        matches.append({
            "id": other_id,
            "name": other_name,
            "confidence": confidence,
            "reasons": reasons,
            "score": (3 if confidence == "High" else 1) + len(reasons),
        })
        seen.add(other_id)

    matches = sorted(matches, key=lambda x: (-x["score"], x["name"]))
    return matches[:max_items]


def render_multibrand_html(row, brand_id):
    matches = get_multibrand_matches(row, brand_id)
    if not matches:
        return ""

    total = len(matches)
    high_count = sum(1 for m in matches if m["confidence"] == "High")
    title = "🏢 Multibrand detected" if high_count else "🏢 Possible multibrand"
    summary = f"{total} linked account{'s' if total != 1 else ''} · {high_count} high confidence"

    chips = []
    for match in matches:
        conf_icon = "✅" if match["confidence"] == "High" else "⚠️"
        reason = "/".join(match["reasons"])
        chip = f"{conf_icon} AR-{html.escape(str(match['id']))} · {html.escape(clean(match['name'], '-'))} · {html.escape(reason)}"
        chips.append(f"<span class='category-chip'>{chip}</span>")

    chips_html = "".join(chips)
    # Keep HTML at column 0. Leading indentation makes Streamlit/Markdown
    # treat it as a code block and display raw <div> text.
    return (
        f"<div class='multibrand-box'>"
        f"<div class='info-mini-label'>{title}</div>"
        f"<div class='chip-line'>{chips_html}</div>"
        f"<div class='multibrand-summary'>{summary}</div>"
        f"</div>"
    )

def _clean_name_for_match(value):
    text = norm_text(value)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _match_rows_by_name(df, name, candidates):
    """Returns all likely rows that match a Growth OS brand name.
    This intentionally matches by Brand/Store Name, not by Store ID.
    """
    if df.empty:
        return pd.DataFrame()

    name_col = _first_existing_col(df, candidates)
    if not name_col:
        return pd.DataFrame()

    target = _clean_name_for_match(name)
    if not target:
        return pd.DataFrame()

    work = df.copy()
    work["_name_norm"] = work[name_col].apply(_clean_name_for_match)
    work = work[work["_name_norm"] != ""].copy()
    if work.empty:
        return work

    exact = work[work["_name_norm"] == target]
    if not exact.empty:
        return exact

    contains = work[work["_name_norm"].apply(lambda x: target in x or x in target)]
    if not contains.empty:
        return contains

    target_tokens = {t for t in target.split() if len(t) >= 3}
    if not target_tokens:
        return pd.DataFrame()

    def token_score(text):
        tokens = {t for t in text.split() if len(t) >= 3}
        if not tokens:
            return 0
        return len(target_tokens & tokens) / max(len(target_tokens), 1)

    work["_name_score"] = work["_name_norm"].apply(token_score)
    fuzzy = work[work["_name_score"] >= 0.67].copy()
    if fuzzy.empty:
        return pd.DataFrame()
    return fuzzy.sort_values(by="_name_score", ascending=False)


def _match_row_by_name(df, name, candidates):
    rows = _match_rows_by_name(df, name, candidates)
    if rows.empty:
        return None
    return rows.iloc[0]


@st.cache_data(ttl=3000, show_spinner=False)
def _load_sheet_safe(sheet_name):
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
        df.columns = [normalize(c) for c in df.columns]
        return df
    except Exception:
        return pd.DataFrame()


def _weighted_average(values, weights=None):
    try:
        v = pd.to_numeric(pd.Series(values), errors="coerce").fillna(0)
        if weights is None:
            return float(v.mean()) if len(v) else 0
        w = pd.to_numeric(pd.Series(weights), errors="coerce").fillna(0)
        if w.sum() <= 0:
            return float(v.mean()) if len(v) else 0
        return float((v * w).sum() / w.sum())
    except Exception:
        return 0


@st.cache_data(ttl=600, show_spinner=False)
def get_menu_metrics_for_brand(name):
    df = _load_sheet_safe("Perfect Store Data")
    matches = _match_rows_by_name(df, name, ["brand", "brand name", "store_name", "store name", "name"])
    if matches.empty:
        return {
            "found": False,
            "traffic": 0,
            "perfect_store": "-",
            "missing_products": 0,
            "purchasing_experience": 0,
            "photos": 0,
            "health_score": 100,
            "secondary": ["No Perfect Store data matched by Brand Name"],
        }

    traffic = pd.to_numeric(matches.get("traffic", pd.Series([0] * len(matches))), errors="coerce").fillna(0)
    missing = pd.to_numeric(matches.get("missing products", pd.Series([0] * len(matches))), errors="coerce").fillna(0)
    purchasing = pd.to_numeric(matches.get("purchasing experience", pd.Series([0] * len(matches))), errors="coerce").fillna(0)
    photos = pd.to_numeric(matches.get("photos", pd.Series([0] * len(matches))), errors="coerce").fillna(0)

    perfect_col = matches.get("perfect store", pd.Series(["-"] * len(matches))).astype(str).str.lower()
    perfect_status = "True" if (not perfect_col.empty and perfect_col.isin(["true", "1", "yes"]).all()) else "False"

    traffic_total = float(traffic.sum())
    missing_max = float(missing.max()) if len(missing) else 0
    purchasing_avg = _weighted_average(purchasing, traffic)
    photos_avg = _weighted_average(photos, traffic)
    perfect_bonus = 1 if perfect_status == "True" else min(photos_avg, purchasing_avg)
    menu_health = max(0, min(100, (photos_avg * 42) + (purchasing_avg * 42) + (perfect_bonus * 16) - min(missing_max * 7, 35)))

    secondary = [
        f"Photos {fmt_percent0(photos_avg)}",
        f"Purchasing Experience {fmt_percent0(purchasing_avg)}",
        f"Traffic {fmt_number(traffic_total)}",
    ]
    if missing_max > 0:
        secondary.append(f"Missing Products {fmt_number(missing_max)}")
    elif photos_avg >= 0.80 and purchasing_avg >= 0.85:
        secondary.append("Rest OK")

    return {
        "found": True,
        "traffic": traffic_total,
        "perfect_store": perfect_status,
        # Use max for missing products because one broken high-traffic/menu signal should be visible.
        "missing_products": missing_max,
        # Use traffic-weighted average for quality metrics.
        "purchasing_experience": purchasing_avg,
        "photos": photos_avg,
        "health_score": menu_health,
        "secondary": secondary,
    }


def _get_ops_rows(sheet_name, name):
    df = _load_sheet_safe(sheet_name)
    # Include both spaced and underscored name headers because exports may come as
    # BRAND_NAME while WO sheets may come as Brand Name / Store Name.
    return _match_rows_by_name(
        df,
        name,
        ["brand", "brand name", "brand_name", "store_name", "store name", "name"],
    )


def _availability_candidate_from_rows(rows):
    """
    Supports the new Availability Data export with:
    BRAND_NAME | Configured Hours | Available Hours | Horas Perdidas
    while keeping the old WO availability format working.
    """
    if rows.empty:
        return None

    configured_col = _first_existing_col(rows, ["configured hours", "configured_hours", "hours configured"])
    available_col = _first_existing_col(rows, ["available hours", "available_hours", "hours available"])
    lost_col = _first_existing_col(rows, ["horas perdidas", "lost hours", "lost_hours", "unavailable hours"])

    if configured_col or available_col or lost_col:
        configured = pd.to_numeric(rows.get(configured_col, pd.Series([0] * len(rows))), errors="coerce").fillna(0)
        available = pd.to_numeric(rows.get(available_col, pd.Series([0] * len(rows))), errors="coerce").fillna(0)

        if lost_col:
            lost = pd.to_numeric(rows.get(lost_col, pd.Series([0] * len(rows))), errors="coerce").fillna(0)
        else:
            lost = (configured - available).clip(lower=0)

        configured_sum = float(configured.sum())
        available_sum = float(available.sum())
        lost_sum = float(lost.sum())
        availability_rate = (available_sum / configured_sum) if configured_sum else 1

        # Score combines absolute business leakage and rate severity.
        rate_gap = max(0, 0.90 - availability_rate)
        score = lost_sum + (rate_gap * 25)

        if lost_sum <= 0.25 and availability_rate >= 0.90:
            score = 0

        return {
            "metric": "Availability",
            "action": "Fix Availability",
            "impact": lost_sum,
            "delta": 0,
            "orders": 0,
            "value": availability_rate,
            "score": score,
            "reason": f"Availability gap detected · Lost {lost_sum:.1f}h · Ava {fmt_percent0(availability_rate)}",
        }

    return None


@st.cache_data(ttl=600, show_spinner=False)
def get_ops_metrics_for_brand(name):
    specs = [
        ("Availability Data", "Availability", "Fix Availability", ["w1 availability", "w2 availability"]),
        ("RTWT Data", "RTWT", "Reduce RTWT", ["w1 rtwt", "w2 rtwt"]),
        ("Cancel Data", "Cancellations", "Reduce Cancellations", ["w1 cancel partner", "w2 cancel partner"]),
        ("DR Data", "Defect Rate", "Improve Defect Rate", ["w1 defect rate", "w2 defect rate"]),
    ]
    candidates = []
    for sheet, metric, action, value_cols in specs:
        rows = _get_ops_rows(sheet, name)
        if rows.empty:
            continue

        # New availability export path: BRAND_NAME + hours columns.
        if sheet == "Availability Data":
            availability_candidate = _availability_candidate_from_rows(rows)
            if availability_candidate is not None:
                candidates.append(availability_candidate)
                continue

        impacts = pd.to_numeric(rows.get("impact (bps)", pd.Series([0] * len(rows))), errors="coerce").fillna(0)
        deltas = pd.to_numeric(rows.get("delta (bps)", pd.Series([0] * len(rows))), errors="coerce").fillna(0)
        orders = pd.to_numeric(rows.get("w1 orders", pd.Series([0] * len(rows))), errors="coerce").fillna(0)

        value = 0
        for col in value_cols:
            if col in rows.columns:
                value = _weighted_average(rows[col], orders)
                break

        impact = float(impacts[impacts > 0].sum())
        delta = float(deltas[deltas > 0].mean()) if (deltas > 0).any() else 0
        orders_sum = float(orders.sum())
        score = max(impact, 0) + max(delta, 0) * 0.01 + orders_sum * 0.0001
        candidates.append({
            "metric": metric,
            "action": action,
            "impact": impact,
            "delta": delta,
            "orders": orders_sum,
            "value": value,
            "score": score,
            "reason": f"{metric} is the strongest OPS signal · Impact {impact:.2f} BPs · Orders {fmt_number(orders_sum)}",
        })
    if not candidates:
        return {
            "action": "Following",
            "reason": "No OPS issue detected by Brand Name in OPS sheets.",
            "metric": "OPS General",
            "health_score": 100,
            "secondary": ["OPS sheets clean / no matched blocker"],
            "candidates": [],
        }

    def candidate_health(c):
        metric = clean(c.get("metric"), "")
        if metric == "Availability":
            value = to_number(c.get("value"), 1)
            return max(0, min(100, value * 100))
        impact = max(0, to_number(c.get("impact"), 0))
        delta = max(0, to_number(c.get("delta"), 0))
        return max(25, min(100, 100 - min(70, impact * 9 + delta * 0.04)))

    best = sorted(candidates, key=lambda x: x["score"], reverse=True)[0]
    best_health = candidate_health(best)
    active_candidates = [c for c in sorted(candidates, key=lambda x: x["score"], reverse=True) if to_number(c.get("score"), 0) > 0.05]

    secondary = []
    for c in active_candidates[:3]:
        metric = clean(c.get("metric"), "OPS")
        if metric == "Availability":
            secondary.append(f"Availability {fmt_percent0(c.get('value', 0))} · Lost {to_number(c.get('impact'), 0):.1f}h")
        else:
            secondary.append(f"{metric}: Impact {to_number(c.get('impact'), 0):.2f} BPs")
    if not secondary:
        secondary = ["Rest OK"]

    if best["score"] <= 0.05:
        return {
            "action": "Following",
            "reason": "OPS metrics look stable by Brand Name.",
            "metric": "OPS General",
            "health_score": 100,
            "secondary": ["Rest OK"],
            "candidates": candidates,
        }
    return {
        "action": best["action"],
        "reason": best.get("reason", f"{best['metric']} is the strongest OPS signal."),
        "metric": best["metric"],
        "health_score": best_health,
        "secondary": secondary,
        "candidates": candidates,
    }


def get_md_campaign_names_for_brand(name):
    df = _load_sheet_safe("MD Names")
    if df.empty:
        return {"md": "-", "md_pro": "-"}

    name_col = _first_existing_col(df, ["name", "campaign name", "promo", "promotion"])
    if not name_col:
        return {"md": "-", "md_pro": "-"}

    matches = _match_rows_by_name(df, name, ["brand name", "brand", "store_name", "store name"])
    if matches.empty:
        return {"md": "-", "md_pro": "-"}

    names = []
    for value in matches[name_col].tolist():
        text = clean(value, "-").strip()
        if text not in ["", "-"] and text not in names:
            names.append(text)

    def is_pro_campaign(text):
        low = norm_text(text)
        return (
            bool(re.search(r"\bpro\b", low))
            or "exclusivo pro" in low
            or "primec" in low
            or bool(re.search(r"\bprime\b", low))
        )

    pro_names = [n for n in names if is_pro_campaign(n)]
    md_names = [n for n in names if not is_pro_campaign(n)]

    # If everything is tagged as Pro but MD is active, still show the best available name instead of blank.
    md_display = " | ".join(md_names[:2]) if md_names else (names[0] if names else "-")
    pro_display = " | ".join(pro_names[:2]) if pro_names else "-"

    return {"md": md_display, "md_pro": pro_display}



def _priority_score_display(value):
    try:
        v = float(value)
        if abs(v) >= 100:
            return f"{v:.1f}"
        if abs(v) >= 10:
            return f"{v:.1f}"
        if abs(v) >= 1:
            return f"{v:.2f}"
        return f"{v:.3f}"
    except Exception:
        return "-"


def _format_priority_date(value):
    """Formats Excel serial dates / pandas timestamps / text dates for Priority Data."""
    if value is None:
        return "-"
    try:
        if pd.isna(value):
            return "-"
    except Exception:
        pass

    # Excel serial number, e.g. 46148 -> 06-May-2026.
    try:
        if isinstance(value, (int, float)) and value > 1000:
            dt = pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce")
            if not pd.isna(dt):
                return dt.strftime("%d-%b")
    except Exception:
        pass

    try:
        dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
        if not pd.isna(dt):
            return dt.strftime("%d-%b")
    except Exception:
        pass

    return clean(value, "-")


def _count_unique_priority_promos(values):
    ids = set()
    for value in values:
        text = clean(value, "").strip()
        if not text or text == "-":
            continue
        parts = re.split(r"[\n,;/|\s]+", text)
        for part in parts:
            part = part.strip()
            if part and part.lower() not in ["nan", "none", "-"]:
                ids.add(part)
    return len(ids)


@st.cache_data(ttl=3000, show_spinner=False)
def load_priority_data():
    df = _load_sheet_safe(PRIORITY_DATA_SHEET)
    if df.empty:
        return pd.DataFrame()

    brand_col = _first_existing_col(df, ["brand", "brand name", "name"])
    metric_col = _first_existing_col(df, ["metric", "métrica", "metrica"])
    score_col = _first_existing_col(df, ["prioridad bd", "priority bd", "prioridad", "priority score"])

    if not brand_col or not metric_col:
        return pd.DataFrame()

    df = df.copy()
    df["_row_order"] = range(len(df))
    # Priority Data exports brand only on the Total row; lever rows below are blank.
    # Forward-fill keeps Ads/Promos/OPS levers attached to the correct brand.
    df["_brand_col"] = df[brand_col].ffill()
    df["_id"] = df["_brand_col"].apply(normalize_brand_id)
    df["_metric"] = df[metric_col].apply(clean)
    df["_metric_norm"] = df["_metric"].apply(norm_text)
    df["_score"] = pd.to_numeric(df[score_col], errors="coerce").fillna(0) if score_col else 0

    total_rows = df[df["_metric_norm"] == "total"].copy().sort_values(by="_row_order")
    rank_map = {}
    for idx, (_, row) in enumerate(total_rows.iterrows(), start=1):
        bid = normalize_brand_id(row.get("_id"))
        if bid and bid not in rank_map:
            rank_map[bid] = idx
    df["_total_rank"] = df["_id"].apply(lambda x: rank_map.get(normalize_brand_id(x)))
    return df


@st.cache_data(ttl=600, show_spinner=False)
def get_priority_signals_for_brand(brand_id, name=""):
    df = load_priority_data()
    if df.empty:
        return {"found": False}

    target = normalize_brand_id(brand_id)
    rows = df[df["_id"].astype(str) == target].copy() if target else pd.DataFrame()

    # Fallback by brand name in case a future export changes ID formatting.
    if rows.empty and name:
        rows = _match_rows_by_name(df, name, ["_brand_col", "brand", "brand name", "name"]).copy()

    if rows.empty:
        return {"found": False}

    rows = rows.sort_values(by="_row_order")
    total_rows = rows[rows["_metric_norm"] == "total"].copy()
    total_row = total_rows.iloc[0] if not total_rows.empty else rows.iloc[0]

    contact_col = _first_existing_col(rows, ["último contacto", "ultimo contacto", "last contact"])
    coinv_col = _first_existing_col(rows, ["coinversion md", "coinversión md", "coinversion", "coinversión"])
    vencida_col = _first_existing_col(rows, ["promo vencida", "promos vencidas", "expired promo"])
    vencer_col = _first_existing_col(rows, ["promo por vencer", "promos por vencer", "expiring promo"])

    last_contact = "-"
    if contact_col:
        non_empty = rows[contact_col].dropna()
        last_contact = _format_priority_date(total_row.get(contact_col) if contact_col in total_row.index else (non_empty.iloc[0] if not non_empty.empty else None))

    coinversion = "No"
    if coinv_col:
        coin_values = rows[coinv_col].dropna().astype(str).tolist()
        if any(norm_text(v) in ["si", "sí", "yes", "true", "1"] or norm_text(v).startswith("si") for v in coin_values):
            coinversion = "Sí"

    promo_vencida = _count_unique_priority_promos(rows[vencida_col].tolist()) if vencida_col else 0
    promo_por_vencer = _count_unique_priority_promos(rows[vencer_col].tolist()) if vencer_col else 0

    lever_rows = rows[rows["_metric_norm"] != "total"].copy().sort_values(by="_row_order")
    levers = []
    desc_col = _first_existing_col(rows, ["descripción", "descripcion", "description"])
    for _, r in lever_rows.iterrows():
        metric = clean(r.get("_metric"), "-")
        if metric in ["", "-"]:
            continue
        descr = clean(r.get(desc_col), "") if desc_col else ""
        levers.append({
            "metric": metric,
            "score": float(to_number(r.get("_score"), 0)),
            "description": descr,
        })

    return {
        "found": True,
        "score": float(to_number(total_row.get("_score"), 0)),
        "rank": total_row.get("_total_rank"),
        "last_contact": last_contact,
        "levers": levers,
        "coinversion": coinversion,
        "promo_vencida": promo_vencida,
        "promo_por_vencer": promo_por_vencer,
    }


def render_priority_signals_html(brand_id, name=""):
    signals = get_priority_signals_for_brand(brand_id, name)
    if not signals.get("found"):
        return ""

    rank = signals.get("rank")
    rank_text = f"#{int(rank)}" if rank not in [None, "", "-"] and not pd.isna(rank) else "-"
    lever_chips = "".join([
        f"<span class='priority-chip'>{html.escape(clean(l.get('metric'), '-'))} · <strong>{_priority_score_display(l.get('score'))}</strong></span>"
        for l in signals.get("levers", [])
    ]) or "<span class='priority-chip'>No levers listed</span>"

    return (
        f"<div class='wide-info-card'>"
        f"<div class='wide-info-title'>Priority Signals</div>"
        f"<div class='priority-top-grid'>"
        f"<div><div class='info-mini-label'>🔥 Priority Score</div><div class='info-mini-value'>{_priority_score_display(signals.get('score'))}</div></div>"
        f"<div><div class='info-mini-label'># Orden de marcado</div><div class='info-mini-value'>{rank_text}</div></div>"
        f"<div><div class='info-mini-label'>🕒 Last Contact</div><div class='info-mini-value'>{html.escape(clean(signals.get('last_contact'), '-'))}</div></div>"
        f"<div><div class='info-mini-label'>% Coinversión MD</div><div class='info-mini-value'>{html.escape(clean(signals.get('coinversion'), 'No'))}</div></div>"
        f"<div><div class='info-mini-label'>⚠️ Promo vencida</div><div class='info-mini-value'>{int(signals.get('promo_vencida') or 0)}</div></div>"
        f"<div><div class='info-mini-label'>⏰ Promo por vencer</div><div class='info-mini-value'>{int(signals.get('promo_por_vencer') or 0)}</div></div>"
        f"<div class='priority-levers'>{lever_chips}</div>"
        f"<div class='priority-note'>Guide only for Smart Priority adherence. 360° Action remains based on the Growth OS reasoning engine.</div>"
        f"</div>"
        f"</div>"
    )




def _normalize_rate_value(value):
    """Converts 18 / 18% / 0.18 into 0.18 for internal calculations."""
    v = to_number(value, 0)
    if abs(v) > 2:
        v = v / 100
    if v < 0:
        v = 0
    return v


def _rate_label(value):
    try:
        return fmt_percent0(_normalize_rate_value(value))
    except Exception:
        return "-"


def _round_budget_ars(value, step=5000):
    try:
        v = float(value)
        if v <= 0:
            return 0
        return int(round(v / step) * step)
    except Exception:
        return 0


def _safe_action_map_from_actions(actions):
    try:
        return _action_dict(actions)
    except Exception:
        return {
            "ops": {"action": "Following", "reason": "Stable"},
            "menu": {"action": "Following", "reason": "Stable"},
            "md": {"action": "Following", "reason": "Stable"},
            "ads": {"action": "Following", "reason": "Stable"},
        }




def get_availability_readiness_for_brand(name):
    """Returns availability readiness using Availability Data by Brand Name.
    New export: BRAND_NAME + Configured Hours + Available Hours + Horas Perdidas.
    Falls back to old WO availability value columns when present.
    """
    rows = _get_ops_rows("Availability Data", name)
    if rows.empty:
        return {"found": False, "rate": None, "lost_hours": 0, "ready": False}

    candidate = _availability_candidate_from_rows(rows)
    if candidate is not None:
        rate = candidate.get("value", None)
        lost = candidate.get("impact", 0)
        try:
            rate_f = float(rate)
        except Exception:
            rate_f = None
        return {
            "found": True,
            "rate": rate_f,
            "lost_hours": float(to_number(lost, 0)),
            "ready": bool(rate_f is not None and rate_f >= 0.90),
        }

    # Legacy WO path: try to read a weighted/mean availability value.
    for col in ["w2 availability", "w1 availability", "availability"]:
        if col in rows.columns:
            vals = pd.to_numeric(rows[col], errors="coerce").dropna()
            if not vals.empty:
                rate = float(vals.mean())
                if rate > 2:
                    rate = rate / 100
                return {"found": True, "rate": rate, "lost_hours": 0, "ready": rate >= 0.90}

    return {"found": True, "rate": None, "lost_hours": 0, "ready": False}


def _format_ars_compact(value):
    try:
        v = float(value)
        if abs(v) >= 1000000:
            return f"{v/1000000:.1f}M".replace(".0", "")
        if abs(v) >= 1000:
            return f"{v/1000:.0f}k"
        return fmt_number(v)
    except Exception:
        return "-"


def _round_budget_up_ars(value, step=1000):
    try:
        v = float(value)
        if v <= 0:
            return 0
        return int(math.ceil(v / step) * step)
    except Exception:
        return 0


def _budget_range_ars(base_value, width_pct=0.18, min_width=3000, step=1000):
    base = max(float(base_value or 0), 0)
    if base <= 0:
        return (0, 0)
    low = _round_budget_ars(base * (1 - width_pct), step=step)
    high = _round_budget_up_ars(base * (1 + width_pct), step=step)
    if high - low < min_width:
        high = _round_budget_up_ars(low + min_width, step=step)
    return max(low, step), max(high, max(low, step))


def _format_budget_range(low, high):
    low = to_number(low, 0)
    high = to_number(high, 0)
    if not low and not high:
        return "—"
    if low == high:
        return f"ARS {_format_ars_compact(low)}"
    return f"ARS {_format_ars_compact(low)}–{_format_ars_compact(high)}"


@st.cache_data(ttl=3000, show_spinner=False)
def _load_store_id_map():
    df = _load_sheet_safe(STORE_ID_SHEET)
    if df.empty:
        return {}
    brand_col = _first_existing_col(df, ["brand id", "brand_id", "id"])
    store_col = _first_existing_col(df, ["store id", "store_id", "store"])
    if not brand_col or not store_col:
        return {}
    result = {}
    for _, row in df.iterrows():
        bid = normalize_brand_id(row.get(brand_col))
        sid = normalize_brand_id(row.get(store_col))
        if bid and sid:
            result[bid] = sid
    return result


@st.cache_data(ttl=3000, show_spinner=False)
def get_definitive_top_products_for_brand(brand_id):
    """Returns the 3 products from Definitive Top Products using Store ID mapping."""
    store_map = _load_store_id_map()
    store_id = store_map.get(normalize_brand_id(brand_id), "")
    if not store_id:
        return []

    df = _load_sheet_safe(DEFINITIVE_TOP_PRODUCTS_SHEET)
    if df.empty:
        return []

    cs_col = _first_existing_col(df, ["country_store_id", "country store id"])
    name_col = _first_existing_col(df, ["name", "product", "product name"])
    rank_col = _first_existing_col(df, ["ranking", "rank"])
    atc_col = _first_existing_col(df, ["atc"])
    vpd_col = _first_existing_col(df, ["vpd"])
    cvr_col = _first_existing_col(df, ["cvr"])
    quality_col = _first_existing_col(df, ["prod. quality", "prod quality", "quality"])

    if not cs_col or not name_col:
        return []

    target = f"AR{store_id}"
    work = df[df[cs_col].astype(str).str.upper().str.replace("-", "", regex=False).str.strip() == target.upper()].copy()
    if work.empty:
        return []

    if rank_col:
        work["_rank"] = pd.to_numeric(work[rank_col], errors="coerce").fillna(999)
        work = work.sort_values(by="_rank", ascending=True)
    else:
        work["_rank"] = range(1, len(work) + 1)

    products = []
    for _, r in work.head(3).iterrows():
        products.append({
            "rank": int(to_number(r.get("_rank"), len(products) + 1)),
            "name": clean(r.get(name_col), "-"),
            "atc": clean(r.get(atc_col), "-") if atc_col else "-",
            "vpd": clean(r.get(vpd_col), "-") if vpd_col else "-",
            "cvr": clean(r.get(cvr_col), "-") if cvr_col else "-",
            "quality": clean(r.get(quality_col), "-") if quality_col else "-",
        })
    return products


def _tokens_for_similarity(value):
    text = norm_text(clean(value, ""))
    tokens = [t for t in re.split(r"\W+", text) if len(t) >= 3]
    stop = {"con", "para", "combo", "mediano", "grande", "unidad", "unidades", "por", "los", "las", "una", "uno", "del"}
    return {t for t in tokens if t not in stop}


def _similarity_score(a, b):
    ta = _tokens_for_similarity(a)
    tb = _tokens_for_similarity(b)
    if not ta or not tb:
        return 0
    return len(ta & tb) / max(len(ta), 1)


def recommend_cross_sell_for_products(category, top_products):
    hero = top_products[0]["name"] if top_products else "Top Product"
    keywords = get_category_keywords(category)
    fallback_map = [
        (["hamburg", "burger"], "Bebida + papas"),
        (["pizza", "empanada"], "Gaseosa / bebida"),
        (["pollo", "chicken"], "Papas + bebida"),
        (["sushi", "roll"], "Bebida / entrada"),
        (["cafe", "panaderia", "desayuno", "pasteleria"], "Café / bebida + producto complementario"),
        (["helado", "heladeria"], "Mayor tamaño / segundo sabor"),
        (["milanesa", "argentina", "parrilla"], "Guarnición + bebida"),
    ]
    fallback = "Bebida / complemento de ticket"
    cat_text = norm_text(category)
    for keys, value in fallback_map:
        if any(k in cat_text for k in keys):
            fallback = value
            break

    cross = _load_sheet_safe(CROSS_SELL_SHEET)
    if cross.empty:
        return {"base_product": hero, "pairing": f"{hero} + {fallback}", "reason": "fallback by category"}

    principal_col = _first_existing_col(cross, ["pprincipal", "principal", "producto principal"])
    secondary_col = _first_existing_col(cross, ["psecundario", "secondary", "producto secundario"])
    pct_col = _first_existing_col(cross, ["%", "pct", "share"])
    if not principal_col or not secondary_col:
        return {"base_product": hero, "pairing": f"{hero} + {fallback}", "reason": "fallback by category"}

    rows = []
    product_names = [p.get("name", "") for p in top_products]
    for _, r in cross.iterrows():
        principal = clean_product_name(r.get(principal_col))
        secondary = clean_product_name(r.get(secondary_col))
        pct = to_number(r.get(pct_col), 0) if pct_col else 0
        if pct > 1:
            pct = pct / 100
        best_product = hero
        best_score = 0
        for product in product_names:
            s = max(_similarity_score(product, principal), _similarity_score(product, secondary))
            if s > best_score:
                best_score = s
                best_product = product
        category_score = 0.15 if (product_matches_keywords(principal, keywords) or product_matches_keywords(secondary, keywords)) else 0
        total = best_score * 100 + category_score * 100 + pct * 15
        rows.append((total, best_score, pct, best_product, principal, secondary))

    rows = sorted(rows, key=lambda x: x[0], reverse=True)
    if rows and rows[0][0] > 0:
        _, match_score, pct, best_product, principal, secondary = rows[0]
        if _similarity_score(best_product, secondary) > _similarity_score(best_product, principal):
            addon = principal
        else:
            addon = secondary
        if not addon or addon == "-":
            addon = fallback
        reason = "CABA cross-selling trend" if match_score > 0 else "category trend"
        return {"base_product": best_product, "pairing": f"{best_product} + {addon}", "reason": reason}

    return {"base_product": hero, "pairing": f"{hero} + {fallback}", "reason": "fallback by category"}


def _recommended_pro_extra(pro, cr=0):
    if not pro:
        return 5, "PRO sin dato claro: usar +5% como incentivo seguro"
    if pro < 0.40:
        return 10, "PRO bajo: usar +10% para captar usuarios PRO"
    if pro < 0.65:
        if cr and cr < 0.12:
            return 10, "PRO medio + CR bajo: +10% ayuda a mover conversión"
        return 5, "PRO medio: +5% protege margen y mantiene incentivo"
    return 5, "PRO alto: +5% es suficiente para no regalar margen a usuarios que ya compran"


def _ads_booking_display_parts(ads_current):
    weekly_ars = to_number(ads_current.get("bookings_usd"), 0) * ARS_PER_USD
    accum_ars = to_number(ads_current.get("bookings_accum_usd"), 0) * ARS_PER_USD
    source = clean(ads_current.get("bookings_source"), "none")
    if weekly_ars <= 0:
        return "—", "No weekly booking detected"
    if source == "growth_os_manual":
        return f"{fmt_ars(weekly_ars)} ✅", "Weekly booking confirmed from Growth OS"
    if source == "current_ads_approx":
        return f"≈ {fmt_ars(weekly_ars)}", f"Approx weekly from Current ADS accumulated ({fmt_ars(accum_ars)} MTD)"
    return fmt_ars(weekly_ars), "Weekly booking"



def _recommended_cross_sell_discount(aov, cr, commission=0, blocking_issue=False, ads_roi=0):
    """Discount logic for cross-selling offers using Sabas guardrails.
    Cross-selling should move AOV/value without breaking margin.
    """
    aov = to_number(aov, 0)
    cr = _normalize_rate_value(cr)
    commission = _normalize_rate_value(commission)
    ads_roi = to_number(ads_roi, 0)

    if blocking_issue:
        return 10, "Cross-sell controlado: OPS/Menu con alerta, incentivo bajo mientras se corrige la base"
    if commission >= 0.25:
        return 10, "Comisión alta: proteger margen en cross-selling"
    if cr and cr < 0.10:
        return 20, "CR bajo: cross-selling necesita incentivo más fuerte para mover conversión"
    if aov and aov < 18000:
        return 15, "AOV bajo: combo/cross-selling con 15% ayuda a subir ticket"
    if ads_roi >= 3:
        return 15, "ROI sano: 15% permite empujar ticket sin ser demasiado agresivo"
    return 10, "Cross-selling de mantenimiento: 10% como incentivo liviano"


def design_campaign_for_brand(name, category, current_gmv_ars, current_aov_ars, cr_value, pro_value, commission_value, ads_current, md_current, md_pro_current, booster, actions, brand_id=None):
    """Rule-based Campaign Designer based on Sabas criteria: Top 3 products, dynamic PRO, booking ranges, cross-selling and ROI>3 upselling."""
    cr = _normalize_rate_value(cr_value)
    pro = _normalize_rate_value(pro_value)
    commission = _normalize_rate_value(commission_value)
    gmv = to_number(current_gmv_ars, 0)
    aov = to_number(current_aov_ars, 0)
    ads_roi = to_number(ads_current.get("roi"), 0)
    ads_bookings_weekly_usd = to_number(ads_current.get("bookings_usd"), 0)
    ads_bookings_weekly_ars = ads_bookings_weekly_usd * ARS_PER_USD
    ads_active = bool(ads_current.get("active", False))
    md_roi = to_number(md_current.get("roi"), 0)
    md_active = bool(md_current.get("active", False))
    action_map = _safe_action_map_from_actions(actions)
    ops_action = clean(action_map.get("ops", {}).get("action"), "Following")
    menu_action = clean(action_map.get("menu", {}).get("action"), "Following")
    blocking_issue = ops_action != "Following" or menu_action != "Following"

    event = clean(booster.get("event"), "-") if isinstance(booster, dict) else "-"
    event_type = clean(booster.get("event_type"), "-") if isinstance(booster, dict) else "-"

    top_products = get_definitive_top_products_for_brand(brand_id) if brand_id else []
    while len(top_products) < 3:
        top_products.append({"rank": len(top_products) + 1, "name": "-", "atc": "-", "vpd": "-", "cvr": "-", "quality": "-"})
    hero_product = top_products[0]["name"] if top_products else "Top Product"
    backup_product = top_products[1]["name"] if len(top_products) > 1 else "-"
    tactical_product = top_products[2]["name"] if len(top_products) > 2 else "-"
    cross_sell = recommend_cross_sell_for_products(category, [p for p in top_products if p.get("name") not in ["", "-"]])
    cross_sell_discount, cross_sell_discount_reason = _recommended_cross_sell_discount(aov, cr, commission, blocking_issue, ads_roi)

    reasons = []
    if cr and cr < 0.12:
        reasons.append("CR bajo: priorizar conversión")
    elif cr and cr >= 0.18:
        reasons.append("CR sano: puede escalar tráfico con más seguridad")
    pro_extra, pro_reason = _recommended_pro_extra(pro, cr)
    reasons.append(pro_reason)
    if blocking_issue:
        reasons.append("OPS/Menu con alerta: escalar controlado, no bloquear venta")
    if event not in ["", "-"]:
        reasons.append(f"Booster/evento sugerido: {event}")
    reasons.append(cross_sell_discount_reason)

    upsell_delta_ars = 0
    ads_action = "Validate GMV before budget"
    aggression = "Base"

    # Commercial pressure model: weekly Ads budget = current monthly GMV × pressure % / 4.
    # Pressure ranges from 4% to 10%, depending on traffic absorption capacity.
    if gmv <= 0:
        pressure_pct = 0
        pressure_label = "No GMV"
        pressure_reason = "No current GMV base to calculate Ads pressure."
    elif blocking_issue or (cr and cr < 0.10) or (ads_active and ads_roi and ads_roi < 1.5):
        pressure_pct = 0.04
        pressure_label = "Low Pressure"
        pressure_reason = "Base fragile: keep Ads pressure controlled while fixing conversion/OPS."
    elif ads_active and ads_roi >= 6 and (not cr or cr >= 0.12):
        pressure_pct = 0.10
        pressure_label = "Aggressive Scaling"
        pressure_reason = "Elite response: strong ROI and enough base health to absorb traffic."
    elif ads_active and ads_roi >= 4.5:
        pressure_pct = 0.085
        pressure_label = "Growth Push"
        pressure_reason = "Strong ROI: increase pressure without jumping straight to the ceiling."
    elif ads_active and ads_roi >= 3:
        pressure_pct = 0.075
        pressure_label = "Growth Push"
        pressure_reason = "ROI >3: room to push controlled upselling."
    elif cr and cr >= 0.18 and not blocking_issue:
        pressure_pct = 0.07
        pressure_label = "Controlled Scaling"
        pressure_reason = "Healthy conversion: Ads pressure can be above starter level."
    elif not ads_active:
        pressure_pct = 0.06
        pressure_label = "Controlled Acquisition"
        pressure_reason = "Starter acquisition, but aligned to the 4%–10% pressure model."
    else:
        pressure_pct = 0.05
        pressure_label = "Controlled Scaling"
        pressure_reason = "Maintain a controlled pressure floor while monitoring ROI."

    target_budget = _round_budget_ars(gmv * pressure_pct / 4, step=1000) if gmv > 0 else 0
    if target_budget > 0:
        target_budget = max(15000, target_budget)
    reasons.append(f"Ads pressure model: {pressure_label} · {int(pressure_pct*100)}% GMV monthly / 4")
    reasons.append(pressure_reason)

    if gmv <= 0:
        base_budget = 0
        budget_low, budget_high = (0, 0)
    elif not ads_active:
        base_budget = target_budget
        budget_low, budget_high = _budget_range_ars(base_budget, width_pct=0.20, min_width=3000, step=1000)
        ads_action = "Ads Acquisition · pressure model" if not blocking_issue else "Small Ads test while fixing base"
        aggression = pressure_label
    elif ads_roi > 3:
        if blocking_issue:
            upsell_delta_ars = max(2000, _round_budget_up_ars(ads_bookings_weekly_ars * 0.05, step=1000))
            upsell_delta_ars = min(upsell_delta_ars, 10000)
            aggression = "Symbolic"
            reasons.append("ROI >3: sí hay upselling, pero controlado por OPS/Menu")
        elif ads_roi >= 6 and cr >= 0.12:
            upsell_delta_ars = max(5000, _round_budget_up_ars(ads_bookings_weekly_ars * 0.50, step=1000))
            aggression = "Aggressive"
            reasons.append("ROI muy alto: espacio para upselling agresivo")
        elif ads_roi >= 4.5:
            upsell_delta_ars = max(5000, _round_budget_up_ars(ads_bookings_weekly_ars * 0.30, step=1000))
            aggression = "Strong"
            reasons.append("ROI alto: escalar inversión")
        else:
            upsell_delta_ars = max(3000, _round_budget_up_ars(ads_bookings_weekly_ars * 0.15, step=1000))
            aggression = "Moderate"
            reasons.append("ROI >3: upselling moderado")

        if ads_bookings_weekly_ars <= 0:
            base_budget = target_budget
        else:
            base_budget = max(target_budget, ads_bookings_weekly_ars + upsell_delta_ars)
        budget_low, budget_high = _budget_range_ars(base_budget, width_pct=0.08 if aggression == "Symbolic" else 0.15, min_width=2000, step=1000)
        ads_action = f"Ads Upselling · {aggression}"
    else:
        base_budget = ads_bookings_weekly_ars if ads_bookings_weekly_ars > 0 else target_budget
        budget_low, budget_high = _budget_range_ars(base_budget, width_pct=0.10, min_width=2000, step=1000) if base_budget else (0, 0)
        ads_action = "Maintain / optimize before upselling"
        reasons.append("ROI ≤3: no upselling; sostener u optimizar antes de subir")

    discount = 20
    product_for_promo = hero_product
    if blocking_issue:
        promo_type = "Low-risk Top Product"
        discount = 15
        product_for_promo = tactical_product if tactical_product not in ["", "-"] else hero_product
        promo_action = f"15% on {product_for_promo}"
    elif aov and aov < 18000:
        promo_type = "Combo / Bundle"
        discount = 15
        product_for_promo = backup_product if backup_product not in ["", "-"] else hero_product
        promo_action = f"15% combo around {product_for_promo}"
    elif cr and cr < 0.10:
        promo_type = "Strong Top Product"
        discount = 25
        product_for_promo = hero_product
        promo_action = f"25% on {product_for_promo}"
    elif cr and cr < 0.15:
        promo_type = "Top Product"
        discount = 25
        product_for_promo = hero_product
        promo_action = f"25% on {product_for_promo}"
    elif not md_active:
        promo_type = "MD Acquisition"
        discount = 20
        product_for_promo = hero_product
        promo_action = f"20% on {product_for_promo}"
    elif md_roi and md_roi < 3.2:
        promo_type = "Promo Optimization"
        discount = 20
        product_for_promo = backup_product if backup_product not in ["", "-"] else hero_product
        promo_action = f"Optimize promo with {product_for_promo}"
    else:
        promo_type = "Maintenance / Seasonal"
        discount = 15
        product_for_promo = tactical_product if tactical_product not in ["", "-"] else hero_product
        promo_action = f"15% selective on {product_for_promo}"

    if commission >= 0.25 and discount >= 25:
        # Con comisión alta, el máximo descuento MD es 20% para proteger margen
        discount = 20
        promo_action = promo_action.replace("25%", "20%").replace("15%", "15%") + " · margin guardrail"
        md_reco = f"{discount}% + {pro_extra}% PRO · {product_for_promo}"
        reasons.append("Comisión alta: descuento reducido a 20% para proteger margen del aliado")
    elif commission and commission < 0.15 and not blocking_issue and cr and cr < 0.12:
        reasons.append("Comisión baja: hay más espacio para incentivo competitivo")

    # Promotional architecture package.
    # Regular promo / booster recommendations work for every brand.
    hq_reco = f"HQ {discount}% + {pro_extra}% PRO · {hero_product}"
    booster_reco = f"{event} · {product_for_promo}" if event not in ["", "-"] else f"No seasonal booster · {product_for_promo}"
    combo_reco = f"Combo {cross_sell.get('pairing', '-')} · {cross_sell_discount}% OFF"

    # Storewide is the regular optional recommendation when needed.
    storewide_discount = 10 if commission >= 0.20 else 15
    min_purchase_ars = _round_budget_up_ars((aov or 0) * 1.25, step=1000) if aov else 0
    storewide_reco = f"Storewide {storewide_discount}% · Min {fmt_ars(min_purchase_ars)}" if min_purchase_ars else f"Storewide {storewide_discount}% · Define min purchase"
    shipping_reco = "Free Shipping capacity" if (pro and pro <= 0.75) or pro == 0 else "Free Shipping optional: PRO mix already strong"

    partner_pressure = commission + (discount / 100.0) + (pro_extra / 100.0) + (cross_sell_discount / 100.0)
    md_reco = f"{discount}% + {pro_extra}% PRO · {product_for_promo}"

    if ads_action.startswith("Ads Upselling"):
        strategy = "Scale performing Ads"
        focus = f"Upselling {aggression.lower()}"
    elif blocking_issue:
        strategy = "Controlled campaign while fixing base"
        focus = "Protect experience"
    elif event not in ["", "-"]:
        strategy = f"{event} + {promo_type}"
        focus = "Seasonal conversion"
    elif not ads_active and not md_active:
        strategy = "Ads + MD Acquisition"
        focus = "Visibility + conversion"
    elif "Combo" in promo_type:
        strategy = "Combo value strategy"
        focus = "Raise AOV / value perception"
    else:
        strategy = f"{promo_type} campaign"
        focus = "Conversion"

    if blocking_issue:
        # Base conservadora cuando la operación está rota — no acumular bonuses de tráfico
        low, high = 1, 4
        if event not in ["", "-"]:
            low += 1; high += 2  # El evento puede sumar un poco igual
        if ads_roi > 3:
            low += 1; high += 1  # ROI sano suma marginalmente incluso con OPS rota
    else:
        low, high = 4, 8
        if not ads_active:
            low += 2; high += 5
        if ads_roi > 3:
            low += 1; high += 4
        if ads_roi >= 6:
            low += 2; high += 4
        if cr and cr < 0.12:
            low += 3; high += 6
        if pro and pro < 0.40 and pro_extra >= 10:
            low += 1; high += 3
        if event not in ["", "-"]:
            low += 2; high += 4
        if ads_roi < 2 and ads_active:
            low = max(1, low - 2); high = max(low + 2, high - 3)
    high = min(high, 25)
    low = min(low, high)

    if blocking_issue or partner_pressure >= 0.60 or (ads_active and ads_roi < 1.5):
        risk = "High"
    elif partner_pressure >= 0.48 or (cr and cr < 0.10):
        risk = "Medium"
    else:
        risk = "Low"

    if not reasons:
        reasons.append("Marca estable: campaña moderada de mantenimiento")

    booking_display, booking_note = _ads_booking_display_parts(ads_current)

    return {
        "following_mode": False,
        "strategy": strategy,
        "focus": focus,
        "event": event,
        "event_type": event_type,
        "ads_action": ads_action,
        "weekly_budget_ars": base_budget,
        "budget_low_ars": budget_low,
        "budget_high_ars": budget_high,
        "current_weekly_booking_ars": ads_bookings_weekly_ars,
        "upsell_delta_ars": upsell_delta_ars,
        "booking_display": booking_display,
        "booking_source_note": booking_note,
        "promo_action": promo_action,
        "md_reco": md_reco,
        "discount": discount,
        "pro_extra": pro_extra,
        "pro_reason": pro_reason,
        "commission": commission,
        "partner_pressure": partner_pressure,
        "impact_low": low,
        "impact_high": high,
        "risk": risk,
        "hero_product": hero_product,
        "backup_product": backup_product,
        "tactical_product": tactical_product,
        "product_for_promo": product_for_promo,
        "top_products": top_products[:3],
        "cross_sell_pairing": cross_sell.get("pairing", "-"),
        "cross_sell_discount": cross_sell_discount,
        "cross_sell_reco": f"{cross_sell.get('pairing', '-')} · {cross_sell_discount}% OFF",
        "cross_sell_reason": cross_sell.get("reason", "-"),
        "cross_sell_discount_reason": cross_sell_discount_reason,
        "hq_reco": hq_reco,
        "booster_reco": booster_reco,
        "combo_reco": combo_reco,
        "storewide_reco": storewide_reco,
        "shipping_reco": shipping_reco,
        "reasons": reasons[:7],
        # Pass-through for GMV estimate in ads plan card
        "_cvr_norm": cr if cr and cr > 0 else 0,
        "_aov_ars": aov,
        "_ads_active": ads_active,
    }


def render_campaign_designer_html(design):
    if not design:
        return ""

    event_value = clean(design.get("event"), "-")
    if event_value in ["", "-"]:
        event_value = "No seasonal event priority"
    risk = clean(design.get("risk"), "Medium")
    pressure = design.get("partner_pressure", 0)
    commission = design.get("commission", 0)
    budget_text = _format_budget_range(design.get("budget_low_ars", 0), design.get("budget_high_ars", 0))
    impact_low = int(design.get("impact_low", 0))
    impact_high = int(design.get("impact_high", 0))
    impact = f"+{impact_low}% to +{impact_high}% GMV"
    upsell = to_number(design.get("upsell_delta_ars"), 0)
    upsell_text = f"+{fmt_ars(upsell)}" if upsell > 0 else "—"

    # GMV extra estimado para la tarjeta Ads Plan
    # Adquisición: GMV esperado del budget total recomendado
    # Upselling:   GMV incremental de la sobreinversión del 15% sobre el budget actual
    _CPC_PROMEDIO = 650
    _ads_active_for_est      = bool(design.get("_ads_active", False))
    _current_booking_for_est = to_number(design.get("current_weekly_booking_ars"), 0)
    if _ads_active_for_est and _current_booking_for_est > 0:
        # Sobreinversión del 15% sobre lo que ya tiene activo
        _ads_budget_for_est = _current_booking_for_est * 0.15
    else:
        _ads_budget_for_est = to_number(design.get("weekly_budget_ars"), 0)
    _cvr_for_est = to_number(design.get("_cvr_norm"), 0)
    _aov_for_est = to_number(design.get("_aov_ars"), 0)
    if _ads_budget_for_est > 0 and _cvr_for_est > 0 and _aov_for_est > 0:
        _est_visits   = _ads_budget_for_est / _CPC_PROMEDIO
        _est_orders   = _est_visits * _cvr_for_est
        _est_gmv      = _est_orders * _aov_for_est
        _est_label    = "upselling +15%" if _ads_active_for_est else "adquisición"
        _gmv_est_text = f"GMV est. {_est_label}: +{fmt_ars(round(_est_gmv / 1000) * 1000)}/sem"
    else:
        _gmv_est_text = ""

    top_products = design.get("top_products", [])[:3]
    while len(top_products) < 3:
        top_products.append({"rank": len(top_products) + 1, "name": "-", "atc": "-"})

    reasons = list(design.get("reasons", []))
    if design.get("cross_sell_reason") not in ["", "-"]:
        reasons.append(f"Cross-selling: {design.get('cross_sell_reason')}")
    if design.get("pro_reason") not in ["", "-"]:
        reasons.append(design.get("pro_reason"))
    reasons_html = "".join([
        f"<span class='card-chip'>{html.escape(clean(reason, '-'))}</span>"
        for reason in reasons[:8]
    ]) or "<span class='card-chip'>No reasons available</span>"

    # ── helpers ──────────────────────────────────────────────────
    def chip(text):
        return f"<span class='card-chip'>{html.escape(text)}</span>"

    def gauge_bar(label, value_pct, color, width_pct=100):
        """Horizontal bar gauge. value_pct is 0–1."""
        pct = min(max(float(value_pct or 0), 0), 1)
        fill = int(pct * width_pct)
        return (
            f"<div style='margin-bottom:10px'>"
            f"<div style='display:flex;justify-content:space-between;font-size:11px;font-weight:700;color:#6B7280;margin-bottom:3px'>"
            f"<span style='text-transform:uppercase;letter-spacing:.04em'>{html.escape(label)}</span>"
            f"<span style='color:#1A1A2E'>{int(pct*100)}%</span></div>"
            f"<div style='background:rgba(255,255,255,0.92);border-radius:999px;height:8px;overflow:hidden'>"
            f"<div style='width:{fill}%;height:100%;background:{color};border-radius:999px;transition:width .4s'></div>"
            f"</div></div>"
        )

    def risk_badge(risk_label):
        colors = {"Low": ("#7ED321", "rgba(111,242,75,0.08)"), "Medium": ("#FF7124", "rgba(255,113,36,0.08)"), "High": ("#FF4D2E", "rgba(229,51,42,0.10)")}
        fg, bg = colors.get(risk_label, ("#FF7124", "rgba(255,113,36,0.08)"))
        return (
            f"<span style='background:{bg};color:{fg};border:1px solid {fg};border-radius:999px;"
            f"padding:4px 12px;font-size:12px;font-weight:900'>{html.escape(risk_label)} Risk</span>"
        )

    def check_row(label, ok):
        icon = "✅" if ok else "⬜"
        color = "#7ED321" if ok else "rgba(107,114,128,0.60)"
        return (
            f"<div style='display:flex;align-items:center;gap:8px;padding:5px 0;"
            f"border-bottom:1px solid rgba(0,0,0,.05);font-size:13px;color:{color};font-weight:700'>"
            f"<span style='font-size:15px'>{icon}</span>{html.escape(label)}</div>"
        )

    # ── 1. STRATEGY MEGA-CARD ─────────────────────────────────────
    # lever class and text color per sub-item
    strategy_sub_items = [
        ("🎯 Strategy",       clean(design.get("strategy"), "-"), clean(design.get("focus"), "-"),                                                          "lever-ops"),
        ("📢 Ads Plan",       clean(design.get("ads_action"), "-"), (f"Current {clean(design.get('booking_display'), '—')} · Suggested {budget_text} · Delta {upsell_text}" + (f" · {_gmv_est_text}" if _gmv_est_text else "")), "lever-ads"),
        ("🏷️ Promo Plan",    clean(design.get("md_reco"), "-"), f"Promo: {clean(design.get('promo_action'), '-')} · PRO +{int(to_number(design.get('pro_extra'),0))}%",      "lever-md"),
        ("🔗 Cross-Selling",  clean(design.get("cross_sell_reco"), "-"), clean(design.get("cross_sell_discount_reason"), "-"),                               "lever-menu"),
        ("📅 Seasonal Booster", event_value, clean(design.get("event_type"), "-"),                                                                           "lever-ops"),
    ]
    sub_cards_html = ""
    for (sub_label, sub_val, sub_copy, sub_cls) in strategy_sub_items:
        sub_cards_html += (
            f"<div class='campaign-mini-card {sub_cls}' style='flex:1 1 170px;min-width:155px;min-height:0;padding:14px 16px;border-radius:14px;background:#F7F8FC !important;'>"
            f"<div style='font-size:10px;text-transform:uppercase;font-weight:900;letter-spacing:.05em;color:rgba(107,114,128,0.60);margin-bottom:4px'>{html.escape(sub_label)}</div>"
            f"<div style='font-size:14px;font-weight:900;color:#1A1A2E;line-height:1.2;overflow-wrap:anywhere'>{html.escape(sub_val)}</div>"
            f"<div style='font-size:11px;color:#6B7280;margin-top:4px;line-height:1.3'>{html.escape(sub_copy)}</div>"
            f"</div>"
        )

    strategy_mega = (
        f"<div class='campaign-mini-card lever-ops' style='grid-column:1/-1;padding:22px 24px;background:rgba(27,63,139,0.03) !important;border:1px solid rgba(27,63,139,0.08) !important;'>"
        f"<div class='card-label'>🚀 Strategy · Full Campaign Plan</div>"
        f"<div class='card-value' style='font-size:20px;margin-bottom:14px'>{html.escape(clean(design.get('strategy'), '-'))}</div>"
        f"<div style='display:flex;flex-wrap:wrap;gap:10px'>{sub_cards_html}</div>"
        f"</div>"
    )

    # ── 1b. TOP PRODUCTS PYRAMID CARD ────────────────────────────
    p1_name = html.escape(clean(top_products[0].get("name", "-"), "-")) if top_products else "-"
    p2_name = html.escape(clean(top_products[1].get("name", "-"), "-")) if len(top_products) > 1 else "-"
    p3_name = html.escape(clean(top_products[2].get("name", "-"), "-")) if len(top_products) > 2 else "-"

    pyramid_card = (
        f"<div class='campaign-mini-card lever-menu' style='padding:20px 22px;display:flex;flex-direction:column;justify-content:space-between'>"
        f"<div class='card-label'>🏅 Top Products Pyramid</div>"
        f"<div style='margin-top:16px;display:flex;flex-direction:column;align-items:center;gap:10px;flex:1'>"
        # Level 1 — Hero
        f"<div style='background:rgba(255,255,255,0.95);border:1px solid rgba(255,255,255,0.15);backdrop-filter:blur(12px);color:#1A1A2E;border-radius:16px;"
        f"padding:18px 22px;font-size:16px;font-weight:900;text-align:center;width:62%;"
        f"line-height:1.3'>"
        f"🥇 Hero<br><span style='font-size:13px;font-weight:700;opacity:.75'>{p1_name}</span>"
        f"</div>"
        # Level 2 — Backup
        f"<div style='background:rgba(255,255,255,0.90);border:1px solid rgba(0,0,0,0.07);backdrop-filter:blur(12px);color:#1A1A2E;border-radius:16px;"
        f"padding:16px 22px;font-size:15px;font-weight:900;text-align:center;width:81%;"
        f"line-height:1.3'>"
        f"🥈 Backup<br><span style='font-size:12px;font-weight:700;opacity:.75'>{p2_name}</span>"
        f"</div>"
        # Level 3 — Tactical
        f"<div style='background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.95);backdrop-filter:blur(12px);color:#1A1A2E;border-radius:16px;"
        f"padding:14px 22px;font-size:14px;font-weight:900;text-align:center;width:100%;"
        f"line-height:1.3'>"
        f"🥉 Tactical<br><span style='font-size:12px;font-weight:700;opacity:.75'>{p3_name}</span>"
        f"</div>"
        f"</div>"
        f"<div style='font-size:11px;color:rgba(107,114,128,0.60);margin-top:14px;text-align:center'>Use as album stickers for the pitch</div>"
        f"</div>"
    )

    # ── 3. GUARDRAILS CARD (rediseñado) ──────────────────────────
    pressure_pct = to_number(pressure, 0)

    risk_color_map = {"Low": "#7ED321", "Medium": "#FF7124", "High": "#FF4D2E"}
    risk_bg_map   = {"Low": "rgba(111,242,75,0.08)",  "Medium": "rgba(255,113,36,0.08)",  "High": "rgba(229,51,42,0.10)"}
    pressure_color = risk_color_map.get(risk, "#FF7124")

    # ── Termómetro SVG para Risk Level (Low / Mid / High) ─────────
    thermo_levels = {"Low": 1, "Medium": 2, "High": 3}
    thermo_level  = thermo_levels.get(risk, 2)
    thermo_colors = {
        1: ("#7ED321", "rgba(111,242,75,0.08)"),   # Low  → green
        2: ("#FF7124", "rgba(255,113,36,0.08)"),   # Mid  → tangerine
        3: ("#FF4D2E", "rgba(229,51,42,0.10)"),   # High → red
    }
    thermo_fg, thermo_bg_light = thermo_colors[thermo_level]
    # Altura del relleno: 33% / 66% / 100%
    thermo_fill_h = thermo_level * 33

    thermometer_svg = (
        f"<div style='display:flex;flex-direction:column;align-items:center;gap:4px;min-width:52px'>"
        f"<svg width='24' height='64' viewBox='0 0 24 64' fill='none' xmlns='http://www.w3.org/2000/svg'>"
        # tubo externo
        f"<rect x='9' y='4' width='6' height='42' rx='3' fill='rgba(0,0,0,0.08)'/>"
        # relleno dinámico (crece de abajo hacia arriba)
        f"<rect x='9' y='{46 - thermo_fill_h // 3 * 14}' width='6' height='{thermo_fill_h // 3 * 14}' rx='3' fill='{thermo_fg}'/>"
        # bulbo
        f"<circle cx='12' cy='52' r='7' fill='{thermo_fg}'/>"
        f"<circle cx='12' cy='52' r='4' fill='white' opacity='.4'/>"
        f"</svg>"
        # etiquetas L / M / H al lado
        f"<div style='display:flex;flex-direction:column;align-items:center;gap:2px;margin-top:-60px;margin-left:28px'>"
        + "".join([
            f"<span style='font-size:9px;font-weight:900;color:{'#FF4D2E' if i==3 else '#FF7124' if i==2 else '#7ED321'};opacity:{'1' if thermo_level==i else '0.3'}'>"
            f"{'H' if i==3 else 'M' if i==2 else 'L'}</span>"
            for i in [3, 2, 1]
        ]) +
        f"</div>"
        f"</div>"
    )

    # ── Gráfico de puntos (dot chart) para Partner Pressure ───────
    pressure_dots_count = min(max(round(pressure_pct * 10), 0), 10)
    pressure_dots_html = "".join([
        f"<span style='display:inline-block;width:14px;height:14px;border-radius:50%;"
        f"background:{'#FF4D2E' if idx >= 6 else '#FF7124' if idx >= 4 else '#7ED321'};"
        f"opacity:{'1' if idx < pressure_dots_count else '0.18'};margin:2px'></span>"
        for idx in range(10)
    ])
    pressure_dot_chart = (
        f"<div style='margin-bottom:10px'>"
        f"<div style='display:flex;justify-content:space-between;font-size:11px;font-weight:700;color:#6B7280;margin-bottom:6px'>"
        f"<span style='text-transform:uppercase;letter-spacing:.04em'>Partner Pressure</span>"
        f"<span style='color:#1A1A2E'>{int(pressure_pct * 100)}%</span></div>"
        f"<div style='display:flex;flex-wrap:wrap;gap:3px'>{pressure_dots_html}</div>"
        f"<div style='display:flex;justify-content:space-between;font-size:9px;color:rgba(107,114,128,0.60);margin-top:4px'>"
        f"<span>LOW</span><span>MID</span><span>HIGH</span></div>"
        f"</div>"
    )

    # ── Gráfico de proyección (barras históricas + proyección) para Impact ──
    impact_mid_val = (impact_low + impact_high) / 2
    impact_norm_val = min(max(impact_mid_val / 25.0, 0), 1)
    impact_color_val = "#7ED321" if impact_norm_val >= 0.5 else "#FF7124" if impact_norm_val >= 0.25 else "#1B3F8B"

    # Barras: 4 históricas (valores simulados crecientes) + 1 proyección destacada
    # Las barras históricas representan semanas/períodos previos normalizados al GMV actual
    _hist_ratios = [0.55, 0.65, 0.72, 0.82]   # tendencia ascendente (relativo al techo)
    _proj_ratio  = min(0.82 + impact_norm_val * 0.18, 1.0)  # proyección = histórico + uplift esperado
    _bar_w = 14   # ancho de cada barra SVG
    _bar_gap = 6
    _chart_h = 48
    _chart_w = 100

    _bars_svg = ""
    for _bi, _ratio in enumerate(_hist_ratios):
        _bx = 4 + _bi * (_bar_w + _bar_gap)
        _bh = max(4, int(_ratio * (_chart_h - 6)))
        _by = _chart_h - _bh
        _bars_svg += (
            f"<rect x='{_bx}' y='{_by}' width='{_bar_w}' height='{_bh}' rx='3' "
            f"fill='rgba(107,114,128,0.45)' opacity='.7'/>"
        )
    # Barra proyectada (última, destacada con el color del impact)
    _proj_bx = 4 + 4 * (_bar_w + _bar_gap)
    _proj_bh = max(4, int(_proj_ratio * (_chart_h - 6)))
    _proj_by = _chart_h - _proj_bh
    _bars_svg += (
        f"<rect x='{_proj_bx}' y='{_proj_by}' width='{_bar_w}' height='{_proj_bh}' rx='3' "
        f"fill='{impact_color_val}'/>"
        # flecha de proyección encima de la última barra
        f"<polygon points='{_proj_bx + _bar_w//2 - 4},{_proj_by - 10} "
        f"{_proj_bx + _bar_w//2 + 4},{_proj_by - 10} "
        f"{_proj_bx + _bar_w//2},{_proj_by - 2}' fill='{impact_color_val}'/>"
        # etiqueta "+ X%" sobre la proyección
        f"<text x='{_proj_bx + _bar_w//2}' y='{_proj_by - 13}' text-anchor='middle' "
        f"font-size='7' font-weight='900' fill='{impact_color_val}'>+{impact_low}%</text>"
    )
    # Línea de tendencia sobre las barras históricas
    _trend_pts = " ".join([
        f"{4 + bi * (_bar_w + _bar_gap) + _bar_w // 2},{_chart_h - max(4, int(r * (_chart_h - 6)))}"
        for bi, r in enumerate(_hist_ratios)
    ])
    _bars_svg += (
        f"<polyline points='{_trend_pts}' fill='none' stroke='#6B7280' "
        f"stroke-width='1.5' stroke-dasharray='3 2' stroke-linecap='round'/>"
    )

    impact_forecast_svg = (
        f"<div style='margin-bottom:10px'>"
        f"<div style='display:flex;justify-content:space-between;font-size:11px;font-weight:700;color:#6B7280;margin-bottom:4px'>"
        f"<span style='text-transform:uppercase;letter-spacing:.04em'>Proyección de Impacto</span>"
        f"<span style='color:{impact_color_val};font-size:12px'>+{impact_low}% – +{impact_high}% GMV</span></div>"
        f"<svg viewBox='0 0 {_chart_w} {_chart_h}' width='100%' height='{_chart_h}' "
        f"xmlns='http://www.w3.org/2000/svg' preserveAspectRatio='none'>"
        f"{_bars_svg}"
        # etiqueta "Proyectado" bajo la última barra
        f"<text x='{_proj_bx + _bar_w//2}' y='{_chart_h - 1}' text-anchor='middle' "
        f"font-size='6' fill='{impact_color_val}' font-weight='700'>Proy.</text>"
        f"</svg>"
        f"</div>"
    )

    guardrails_card = (
        f"<div class='campaign-mini-card lever-pro' style='padding:20px 22px'>"
        f"<div class='card-label'>🛡️ Guardrails</div>"
        # High Risk con termómetro
        f"<div style='display:flex;align-items:center;gap:12px;margin:10px 0 14px;"
        f"background:{thermo_bg_light};border-radius:12px;padding:10px 14px'>"
        f"{thermometer_svg}"
        f"<div style='margin-left:18px'>"
        f"<div style='font-size:10px;text-transform:uppercase;font-weight:900;letter-spacing:.06em;color:rgba(107,114,128,0.60)'>High Risk</div>"
        f"<div style='font-size:18px;font-weight:900;color:{thermo_fg}'>{html.escape(risk)}</div>"
        f"<div style='font-size:10px;color:rgba(107,114,128,0.60);margin-top:2px'>Low · Mid · High</div>"
        f"</div>"
        f"</div>"
        # Partner Pressure con dot chart
        f"{pressure_dot_chart}"
        # Impact con forecast chart
        f"{impact_forecast_svg}"
        f"<div style='margin-top:6px;background:rgba(255,255,255,0.92);border-radius:12px;padding:8px 12px;"
        f"font-size:11px;color:#6B7280;line-height:1.4'>"
        f"⚠️ Pressure ≥60% = High Risk · ROI target Ads &gt;4.5x · MD &gt;3.2x"
        f"</div>"
        f"</div>"
    )

    # ── 5. REASONING CARD — párrafo analítico en español (resumido) ──
    raw_reasons = list(design.get("reasons", []))
    if design.get("cross_sell_reason") not in ["", "-"]:
        raw_reasons.append(design.get("cross_sell_reason"))
    if design.get("pro_reason") not in ["", "-"]:
        raw_reasons.append(design.get("pro_reason"))

    strategy_val   = clean(design.get("strategy"), "")
    focus_val      = clean(design.get("focus"), "")
    ads_action_val = clean(design.get("ads_action"), "")
    md_reco_val    = clean(design.get("md_reco"), "")
    promo_val      = clean(design.get("promo_action"), "")
    event_val      = clean(design.get("event"), "")
    cross_val      = clean(design.get("cross_sell_reco"), "")
    pro_extra_val  = int(to_number(design.get("pro_extra"), 0))

    # Párrafo compacto en español
    _parts = []
    if strategy_val and strategy_val != "-":
        _lead = f"Estrategia <strong>{html.escape(strategy_val)}</strong>"
        if focus_val and focus_val != "-":
            _lead += f" con foco en <em>{html.escape(focus_val)}</em>"
        _parts.append(_lead)
    _levers = []
    if ads_action_val and ads_action_val != "-":
        _levers.append(f"Ads → {html.escape(ads_action_val)} ({budget_text})")
    if md_reco_val and md_reco_val != "-":
        _md_detail = f"{html.escape(promo_val)}" if promo_val and promo_val != "-" else ""
        _pro_detail = f" +{pro_extra_val}% PRO" if pro_extra_val > 0 else ""
        _levers.append(f"MD → {html.escape(md_reco_val)}{(' · ' + _md_detail) if _md_detail else ''}{_pro_detail}")
    if cross_val and cross_val not in ["-", ""]:
        _levers.append(f"Cross-sell: {html.escape(cross_val)}")
    if _levers:
        _parts.append(". ".join(_levers))
    if event_val and event_val not in ["-", "", "No seasonal event priority"]:
        _parts.append(f"Booster estacional disponible: <strong>{html.escape(event_val)}</strong>")
    if raw_reasons:
        _signals = "; ".join([clean(r, "") for r in raw_reasons[:3] if clean(r, "")])
        if _signals:
            _parts.append(f"Señales clave: {html.escape(_signals)}")
    _parts.append(
        f"Impacto proyectado <strong>+{impact_low}%–+{impact_high}% GMV</strong> · "
        f"Riesgo <strong>{html.escape(risk)}</strong> · Presión aliado {int(pressure_pct * 100)}%."
    )

    reasoning_paragraph = ". ".join(_parts) + "." if _parts else "Sin datos de reasoning disponibles."

    reasoning_card = (
        f"<div class='campaign-mini-card lever-ops' style='padding:20px 22px'>"
        f"<div class='card-label'>🧠 Por qué esta estrategia</div>"
        f"<div style='font-size:13px;color:#6B7280;line-height:1.65;margin-top:10px'>"
        f"{reasoning_paragraph}"
        f"</div>"
        f"</div>"
    )

    # ── GRID LAYOUT ───────────────────────────────────────────────
    # Row 1: Strategy mega card (full width)
    # Row 2 (2 cols side by side): Top Products Pyramid | Guardrails
    bottom_grid = (
        f"<div style='display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:14px;margin-top:14px'>"
        f"{pyramid_card}{guardrails_card}"
        f"</div>"
    )

    return (
        f"<div class='wide-info-card campaign-designer-card'>"
        f"<div class='wide-info-title'>🚀 Campaign Designer</div>"
        f"<div style='margin-top:0'>{strategy_mega}</div>"
        f"{bottom_grid}"
        f"<div class='priority-note' style='margin-top:12px'>{html.escape(clean(design.get('booking_source_note'), ''))} · Products from Definitive Top Products · All campaign logic follows Sabas-defined criteria.</div>"
        f"</div>"
    )


def _health_meta(score, action="Following"):
    score = max(0, min(100, to_number(score, 100)))
    if clean(action, "Following") == "Following" and score >= 85:
        return {"score": score, "badge": f"✅ Following · {score:.0f}%", "class": "health-green", "label": "Following"}
    if score >= 85:
        return {"score": score, "badge": f"🟢 {score:.0f}% Healthy", "class": "health-green", "label": "Healthy"}
    if score >= 70:
        return {"score": score, "badge": f"🟡 {score:.0f}% Watch", "class": "health-yellow", "label": "Watch"}
    if score >= 50:
        return {"score": score, "badge": f"🟠 {score:.0f}% Alert", "class": "health-orange", "label": "Alert"}
    return {"score": score, "badge": f"🔴 {score:.0f}% Critical", "class": "health-red", "label": "Critical"}


def _make_360_action(area, action, reason, health_score=100, secondary=None):
    meta = _health_meta(health_score, action)
    return {
        "area": area,
        "action": clean(action, "Following"),
        "reason": clean(reason, "Stable."),
        "health_score": meta["score"],
        "health_badge": meta["badge"],
        "health_class": meta["class"],
        "health_label": meta["label"],
        "secondary": [clean(x, "") for x in (secondary or []) if clean(x, "")],
    }


def _md_health(md_current, action, roi_benchmark=3.2):
    roi = to_number(md_current.get("roi"), 0)
    active = bool(md_current.get("active", False))
    if active and roi >= roi_benchmark:
        return min(98, 86 + min((roi - roi_benchmark) * 4, 12))
    if active and roi > 0:
        return max(42, min(78, 50 + (roi / roi_benchmark) * 28))
    if action == "Seasonal Event":
        return 72
    if not active:
        return 64
    return 75


def _ads_health(ads_current, action, blocking_issue=False):
    roi = to_number(ads_current.get("roi"), 0)
    active = bool(ads_current.get("active", False))
    bookings = to_number(ads_current.get("bookings_usd"), 0)
    revenue = to_number(ads_current.get("revenue_usd"), 0)
    consumption = revenue / bookings if bookings else 0

    if blocking_issue:
        return 60 if action not in ["Following", "Ads Acquisition"] else 70
    if not active:
        return 72
    if roi >= 3 and consumption <= 1.05:
        return min(98, 86 + min((roi - 3) * 3, 12))
    if roi >= 2:
        return 78
    if roi > 0:
        return max(42, 52 + roi * 10)
    return 55


def build_360_actions(name, category, ads_current, md_current, md_pro_current, booster):
    ops = get_ops_metrics_for_brand(name)
    menu = get_menu_metrics_for_brand(name)

    # Menu action
    if not menu.get("found"):
        menu_action = "Following"
        menu_reason = "No menu issue found in Perfect Store sheet."
    elif menu.get("photos", 0) < 0.80:
        menu_action = "Improve Photos"
        menu_reason = f"Photos at {fmt_percent0(menu.get('photos', 0))} · Traffic {fmt_number(menu.get('traffic', 0))}"
    elif menu.get("purchasing_experience", 0) < 0.85:
        menu_action = "Optimize Purchasing Experience"
        menu_reason = f"Purchasing Experience at {fmt_percent0(menu.get('purchasing_experience', 0))}"
    elif menu.get("missing_products", 0) > 0:
        menu_action = "Recover Missing Products"
        menu_reason = f"Missing Products signal: {fmt_number(menu.get('missing_products', 0))} · Traffic {fmt_number(menu.get('traffic', 0))}"
    else:
        menu_action = "Following"
        menu_reason = "Menu health looks stable."

    # MD action
    md_roi = to_number(md_current.get("roi"), 0)
    if not md_current.get("active", False):
        if booster.get("event", "-") not in ["", "-"]:
            md_action = "Seasonal Event"
            md_reason = f"Recommended event: {booster.get('event', '-')}"
        else:
            md_action = "MD Acquisition"
            md_reason = "No active MD detected."
    elif md_roi < 3.2:
        md_action = "Optimize Promo"
        md_reason = f"MD ROI {fmt_roi(md_roi)} below 3.2x benchmark."
    else:
        md_action = "Following"
        md_reason = f"MD ROI {fmt_roi(md_roi)} is healthy."

    # Ads action
    ads_roi = to_number(ads_current.get("roi"), 0)
    ads_bookings = to_number(ads_current.get("bookings_usd"), 0)
    ads_revenue = to_number(ads_current.get("revenue_usd"), 0)
    consumption = ads_revenue / ads_bookings if ads_bookings else 0
    blocking_issue = ops.get("action") not in ["Following", ""] or menu_action not in ["Following", ""]

    if not ads_current.get("active", False):
        if blocking_issue:
            ads_action = "Prepare Ads"
            ads_reason = "Fix OPS/Menu before scaling traffic."
        else:
            ads_action = "Ads Acquisition"
            ads_reason = "Healthy base with no active Ads detected."
    elif ads_roi > 2.0 and consumption < 0.80 and not blocking_issue:
        ads_action = "Increase CPC"
        ads_reason = f"ROI {fmt_roi(ads_roi)} with consumption {fmt_percent0(consumption)}."
    elif ads_roi < 2.0 and consumption > 0.80:
        ads_action = "Reduce CPC"
        ads_reason = f"ROI {fmt_roi(ads_roi)} with high consumption {fmt_percent0(consumption)}."
    elif blocking_issue:
        ads_action = "Hold Scaling"
        ads_reason = "Traffic should not be scaled until OPS/Menu is clean."
    else:
        ads_action = "Following"
        ads_reason = f"Ads ROI {fmt_roi(ads_roi)} stable."

    ops_secondary = ops.get("secondary", []) or ["Rest OK"]
    # Add non-availability OPS levers detected by Smart Priorities when no official OPS metric exists in the dashboard.
    sp_signals_for_ops = get_priority_signals_for_brand("", name)
    sp_ops = []
    for lever in sp_signals_for_ops.get("levers", []) if sp_signals_for_ops.get("found") else []:
        metric = clean(lever.get("metric"), "")
        kind = _classify_priority_lever(metric)
        if kind.startswith("ops") and kind != "ops_availability":
            sp_ops.append(metric)
    if sp_ops:
        ops_secondary.append("SP: " + " • ".join(list(dict.fromkeys(sp_ops))[:4]))

    menu_secondary = menu.get("secondary", []) or ["Rest OK"]
    md_secondary = [
        f"Status {status_from_active(md_current.get('active', False))}",
        f"ROI {fmt_roi(md_roi)}" if md_current.get("active", False) else "No active MD orders",
    ]
    if booster.get("event", "-") not in ["", "-"]:
        md_secondary.append(f"Event {booster.get('event', '-')}")

    ads_secondary = [
        f"Status {status_from_active(ads_current.get('active', False))}",
        f"ROI {fmt_roi(ads_roi)}" if ads_current.get("active", False) else "No active Ads booking",
    ]
    if ads_current.get("active", False):
        ads_secondary.append(f"Consumption {fmt_percent0(consumption)}")

    return [
        _make_360_action("⚙️ OPS General", ops.get("action", "Following"), ops.get("reason", "OPS stable."), ops.get("health_score", 100), ops_secondary),
        _make_360_action("🍔 Menu", menu_action, menu_reason, menu.get("health_score", 100), menu_secondary),
        _make_360_action("🏷️ MD", md_action, md_reason, _md_health(md_current, md_action), md_secondary),
        _make_360_action("🚀 Ads", ads_action, ads_reason, _ads_health(ads_current, ads_action, blocking_issue), ads_secondary),
    ]



def _priority_lever_texts(signals):
    if not signals or not signals.get("found"):
        return []
    return [clean(l.get("metric"), "") for l in signals.get("levers", []) if clean(l.get("metric"), "")]


def _priority_lever_records(signals):
    if not signals or not signals.get("found"):
        return []
    records = []
    for i, lever in enumerate(signals.get("levers", []), start=1):
        metric = clean(lever.get("metric"), "")
        if metric:
            records.append({
                "metric": metric,
                "score": to_number(lever.get("score"), 0),
                "description": clean(lever.get("description"), ""),
                "source_order": i,
                "kind": _classify_priority_lever(metric),
            })
    return records


def _has_priority_lever(lever_texts, *tokens):
    blob = " | ".join([norm_text(x) for x in lever_texts])
    return any(norm_text(t) in blob for t in tokens)


def _action_for_key(actions, key):
    action_map = _safe_action_map_from_actions(actions)
    return action_map.get(key, {"action": "Following", "reason": "Stable"})


def _tactical_class_for_action(action_text, default_score=92):
    text = norm_text(action_text)
    if "critical" in text or "reduce" in text or "fix" in text or "hold" in text:
        return "health-orange"
    if "optimize" in text or "improve" in text or "prepare" in text or "watch" in text:
        return "health-yellow"
    if "acquisition" in text or "seasonal" in text or "increase" in text or "push" in text:
        return "health-yellow"
    return _health_meta(default_score, "Following").get("class", "health-green")


def _lever_class_from_kind(kind):
    kind = clean(kind, "general")
    if kind.startswith("ops"):
        return "lever-ops"
    if kind.startswith("menu"):
        return "lever-menu"
    if kind == "md":
        return "lever-md"
    if kind == "pro":
        return "lever-pro"
    if kind == "ads":
        return "lever-ads"
    return ""


def _make_tactical_item(area, main, argument, cue="", class_name="health-green", order=1, metric="", lever_class=""):
    return {
        "area": clean(area, "-"),
        "main": clean(main, "✅ Following"),
        "argument": clean(argument, "Stable."),
        "cue": clean(cue, ""),
        "class": clean(class_name, "health-green"),
        "lever_class": clean(lever_class, ""),
        "order": order,
        "metric": clean(metric, ""),
    }


def _classify_priority_lever(metric):
    text = norm_text(metric)

    # OPS / operational signals. These often come from Smart Priorities even when
    # Growth OS only has direct availability data, so they need tactical guidance.
    if any(t in text for t in ["rtwt", "wait", "espera", "repartidor", "rider", "courier", "driver", "tiempo"]):
        return "ops_wait_time"
    if any(t in text for t in ["reclamo", "reclam", "claim", "claims", "complaint", "queja"]):
        return "ops_claims"
    if any(t in text for t in ["cancel", "cancelacion", "cancelación"]):
        return "ops_cancellations"
    if any(t in text for t in ["defect", "defecto", "dr ", " defect rate"]):
        return "ops_defects"
    if any(t in text for t in ["availability", "disponibilidad", "ava", "offline", "conexion", "conexión", "online"]):
        return "ops_availability"
    if any(t in text for t in ["ops", "operacion", "operación", "operativo", "operativa", "churn"]):
        return "ops_other"

    # Menu / catalog / conversion surface.
    if any(t in text for t in ["foto", "photo", "imagen", "image"]):
        return "menu_photos"
    if any(t in text for t in ["purchasing", "experiencia", "purchase", "compra"]):
        return "menu_purchase_experience"
    if "pdf" in text:
        return "menu_pdf"
    if any(t in text for t in ["missing", "faltante", "producto", "product", "catalog", "catalogo", "catálogo", "menu", "menú"]):
        return "menu_catalog"

    # Commercial levers.
    if any(t in text for t in ["promo", "markdown", "md", "descuento", "discount", "coinversion", "coinversión"]):
        return "md"
    if any(t in text for t in ["pro", "prime"]):
        return "pro"
    if any(t in text for t in ["ads", "cpc", "booking", "revenue", "pauta", "visibilidad", "visibility", "inversion", "inversión", "hat", "hats"]):
        return "ads"

    return "general"


def _priority_order_rank(kind):
    # Productive call order: frictions first, then conversion surface, then promo/pro,
    # and Ads last because traffic amplifies whatever base exists underneath.
    if kind.startswith("ops"):
        return 10
    if kind.startswith("menu"):
        return 20
    if kind == "md":
        return 30
    if kind == "pro":
        return 35
    if kind == "ads":
        return 40
    return 50


def _ordered_priority_records(records):
    return sorted(records, key=lambda r: (_priority_order_rank(r.get("kind", "general")), -to_number(r.get("score"), 0), r.get("source_order", 999)))


def _conversation_order_text(records):
    if not records:
        return []
    return [f"{idx}. {clean(r.get('metric'), '-')}" for idx, r in enumerate(records, start=1)]


def _menu_inference_context(menu_metrics):
    if not menu_metrics or not menu_metrics.get("found"):
        return "No hay lectura de Perfect Store cruzada; valida con el aliado dónde puede estar la fricción."
    hints = []
    photos = to_number(menu_metrics.get("photos"), 0)
    purchasing = to_number(menu_metrics.get("purchasing_experience"), 0)
    missing = to_number(menu_metrics.get("missing_products"), 0)
    if photos and photos < 0.80:
        hints.append(f"fotos en {fmt_percent0(photos)}")
    if purchasing and purchasing < 0.85:
        hints.append(f"purchasing experience en {fmt_percent0(purchasing)}")
    if missing > 0:
        hints.append(f"{fmt_number(missing)} productos faltantes")
    if hints:
        return "Posible explicación cruzada: " + " · ".join(hints) + "."
    return "Menú se ve estable; usa esta palanca como validación operativa directa, no como acusación."


def _ops_cross_context(ops_metrics, menu_metrics, ads_current, md_current):
    pieces = []
    if ops_metrics and clean(ops_metrics.get("metric"), "") == "Availability":
        pieces.append(f"Ava {fmt_percent0(ops_metrics.get('candidates', [{}])[0].get('value', 0))}" if ops_metrics.get("candidates") else clean(ops_metrics.get("reason"), ""))
    elif ops_metrics and clean(ops_metrics.get("action"), "Following") != "Following":
        pieces.append(clean(ops_metrics.get("reason"), ""))
    menu_hint = _menu_inference_context(menu_metrics)
    if "Posible explicación" in menu_hint:
        pieces.append(menu_hint)
    if ads_current.get("active"):
        pieces.append("Ads activo puede amplificar la fricción si la operación no sostiene la demanda.")
    if md_current.get("active"):
        pieces.append("Promo activa puede traer volumen adicional y volver más visible cualquier falla operativa.")
    return " ".join([p for p in pieces if p]) or "Sin dato numérico directo visible; conviértelo en pregunta de validación y freno comercial."


def _extract_priority_pct_rate(descr):
    """
    Extrae el % de cancelación/reclamación desde la descripción de Priority Data.
    Formato real: ' 219453: 25%' o con múltiples stores ' 261576: 3,79%; 259327: 6,1%'.
    Si hay varios stores, promedia los porcentajes encontrados.
    Devuelve la tasa como float 0-1 (ej: 0.25), o 0 si no encuentra nada.
    """
    if not descr:
        return 0.0
    matches = re.findall(r"(\d+(?:[.,]\d+)?)\s*%", str(descr))
    if not matches:
        return 0.0
    vals = []
    for m in matches:
        try:
            vals.append(float(m.replace(",", ".")))
        except (TypeError, ValueError):
            continue
    if not vals:
        return 0.0
    avg_pct = sum(vals) / len(vals)
    return avg_pct / 100.0


def _build_ops_tactical_card(record, name, ops_metrics, menu_metrics, ads_current, md_current, aov_ars=0, orders_monthly=0, current_gmv_ars=0):
    metric = clean(record.get("metric"), "OPS signal")
    kind = record.get("kind", "ops_other")
    descr = clean(record.get("description"), "")

    if kind == "ops_wait_time":
        main = "⏱️ Reduce RTWT antes de escalar"
        cue_parts = []
        if descr:
            cue_parts.append(f"SP: {descr}.")
        cue_parts.append("Valida horarios pico y producto que demora.")
        cue = " ".join(cue_parts)
        cls = "health-orange"
    elif kind == "ops_claims":
        main = "⚠️ Valida claims antes de escalar tráfico"
        cue_parts = []
        if descr:
            cue_parts.append(f"SP: {descr}.")
        _claim_rate = _extract_priority_pct_rate(descr)
        if _claim_rate > 0 and orders_monthly > 0 and aov_ars > 0:
            _claim_orders_mes = round(orders_monthly * _claim_rate)
            _claim_gmv_riesgo = round(_claim_orders_mes * aov_ars * 0.50 / 1000) * 1000
            cue_parts.append(f"GMV en riesgo: ~{fmt_ars(_claim_gmv_riesgo)}/mes ({_claim_orders_mes} reclamos · {fmt_percent0(_claim_rate)} tasa · aliado absorbe 50%).")
        elif aov_ars > 0:
            cue_parts.append(f"GMV en riesgo: ~{fmt_ars(round(aov_ars * 0.50))} por reclamo (aliado absorbe ~50%).")
        cue = " ".join(cue_parts)
        cls = "health-orange"
    elif kind == "ops_cancellations":
        main = "🛑 Cancelaciones bloquean eficiencia"
        cue_parts = []
        if descr:
            cue_parts.append(f"SP: {descr}.")
        _cancel_rate = _extract_priority_pct_rate(descr)
        if _cancel_rate > 0 and orders_monthly > 0 and aov_ars > 0:
            _cancel_orders_mes = round(orders_monthly * _cancel_rate)
            _cancel_gmv_perdido = round(_cancel_orders_mes * aov_ars / 1000) * 1000
            cue_parts.append(f"GMV perdido: ~{fmt_ars(_cancel_gmv_perdido)}/mes ({_cancel_orders_mes} cancelaciones · {fmt_percent0(_cancel_rate)} tasa · aliado absorbe el 100%).")
        elif aov_ars > 0:
            cue_parts.append(f"GMV en riesgo: ~{fmt_ars(round(aov_ars))} por cancelación (aliado absorbe el 100%).")
        cue = " ".join(cue_parts)
        cls = "health-orange"
    elif kind == "ops_defects":
        main = "🧩 Defect rate como fricción de confianza"
        cue_parts = []
        if descr:
            cue_parts.append(f"SP: {descr}.")
        cue_parts.append("Corregir producto/descripción antes de subir presión comercial.")
        cue = " ".join(cue_parts)
        cls = "health-yellow"
    elif kind == "ops_availability":
        ava_cand = ops_metrics.get("candidates", []) if ops_metrics else []
        ava_val = ava_cand[0].get("value", 0) if ava_cand else 0
        main = f"🟡 Availability {fmt_percent0(ava_val) if ava_val else ''} — freno comercial"
        cue_parts = []
        if ava_val and ava_val > 0 and current_gmv_ars > 0:
            _ava_gap = max(0, 1.0 - ava_val)
            # Upside proporcional sobre el GMV real del mes
            _ava_upside = round(current_gmv_ars * (_ava_gap / ava_val) / 1000) * 1000
            if _ava_upside > 0:
                cue_parts.append(f"Upside estimado: ~{fmt_ars(_ava_upside)}/mes si availability sube de {fmt_percent0(ava_val)} a 100%.")
        else:
            cue_parts.append("Escalar tráfico sobre baja disponibilidad quema budget.")
        cue = " ".join(cue_parts)
        cls = "health-yellow"
    else:
        main = "🟡 Valida fricción OPS"
        cue_parts = []
        if descr:
            cue_parts.append(f"SP: {descr}.")
        else:
            cue_parts.append("Pregunta de validación: ¿qué parte de la operación está frenando hoy?")
        cue = " ".join(cue_parts)
        cls = "health-yellow"

    return _make_tactical_item("⚙️ " + metric, main, "", cue, cls, metric=metric)


def _build_menu_tactical_card(record, name, menu_metrics, campaign_design):
    metric = clean(record.get("metric"), "Menu signal")
    kind = record.get("kind", "menu_catalog")
    # Solo leer métricas reales si Perfect Store encontró datos para este brand
    has_real_data = bool(menu_metrics and menu_metrics.get("found"))
    photos = to_number(menu_metrics.get("photos"), 0) if has_real_data else None
    purchasing = to_number(menu_metrics.get("purchasing_experience"), 0) if has_real_data else None
    missing = to_number(menu_metrics.get("missing_products"), 0) if has_real_data else None

    if kind == "menu_photos":
        main = "📸 Photos = conversion surface"
        argument = f"No lo vendas como 'faltan fotos'; véndelo como conversión. El usuario decide visualmente y las fotos actuales marcan {fmt_percent0(photos) if photos is not None else '-'}; si no se ve comprable, Ads y MD rinden peor."
        cue = "Pide priorizar fotos de productos top y combos antes de empujar tráfico fuerte."
        cls = "health-yellow" if (photos or 0) >= 0.75 else "health-orange"

    elif kind == "menu_purchase_experience":
        main = "🛒 Purchase experience = less friction"
        argument = f"La experiencia de compra está en {fmt_percent0(purchasing) if purchasing is not None else '-'}. Si el usuario no entiende rápido qué compra, baja conversión y sube el riesgo de reclamos."
        cue = "Valida nombres, descripciones, modificadores y claridad del producto recibido."
        cls = "health-yellow" if (purchasing or 0) >= 0.75 else "health-orange"

    elif kind == "menu_pdf":
        main = "📄 PDF · Reactualización del algoritmo"
        argument = "El menú PDF está desactualizado respecto al algoritmo actual. Un PDF desactualizado afecta la indexación y visibilidad del catálogo en la plataforma."
        cue = "Solicitá al aliado la reactualización del PDF del menú para alinear con el algoritmo vigente."
        cls = "health-yellow"

    else:
        # kind == "menu_catalog": solo mostrar issues con datos reales confirmados
        issues = []
        if photos is not None and photos < 0.90:
            issues.append(f"📸 Fotos {fmt_percent0(photos)} — por debajo del 90%")
        if missing is not None and missing > 1:
            issues.append(f"📦 Missing products: {fmt_number(missing)} productos faltantes")
        if purchasing is not None and purchasing < 0.90:
            issues.append(f"🛒 Purchasing experience {fmt_percent0(purchasing)} — por debajo del 90%")
        main = " · ".join(issues) if issues else "🍔 Ajustar catálogo antes de activar presión comercial"
        argument = "Corregir estas métricas antes de escalar tráfico o activar pauta — si la base no convierte, el gasto es ineficiente." if issues else "Validar catálogo directamente con el aliado — no hay lectura cruzada de Perfect Store para este brand."
        cue = "Pide ajustar productos top/hero antes de activar una presión comercial más fuerte."
        cls = "health-yellow"

    return _make_tactical_item("🍔 " + metric, main, argument, cue, cls, metric=metric)


def _build_md_tactical_card(record, md_current, campaign_design, signals):
    metric = clean(record.get("metric"), "Markdown signal")
    roi = to_number(md_current.get("roi"), 0)
    active = bool(md_current.get("active", False))
    expired = int(signals.get("promo_vencida") or 0) if signals and signals.get("found") else 0
    expiring = int(signals.get("promo_por_vencer") or 0) if signals and signals.get("found") else 0
    reco = clean(campaign_design.get("md_reco"), "") if campaign_design else ""
    descr = clean(record.get("description"), "")

    if active:
        main = f"🏷️ MD activo · ROI {fmt_roi(roi)}"
        action_note = "Continuidad: defendé arquitectura, no solo el descuento."
        if roi < 2.0:
            action_note = "ROI bajo — corregir producto, profundidad o duración antes de renovar."
        elif roi > 4.0:
            action_note = "Buen ROI — propón upselling simbólico de arquitectura (+5% PRO o combo)."
    else:
        main = "🏷️ MD Acquisition"
        action_note = "Propón prueba controlada: producto correcto, descuento mínimo viable, ventana acotada."

    cue_parts = []
    if descr:
        cue_parts.append(f"SP: {descr}.")
    if reco:
        cue_parts.append(reco)
    else:
        cue_parts.append(action_note)
    if expired:
        cue_parts.append(f"⚠️ Promo vencida: {expired}.")
    if expiring:
        cue_parts.append(f"⏰ Por vencer: {expiring}.")
    cue = " ".join(cue_parts)
    return _make_tactical_item("🏷️ " + metric, main, "", cue, "health-yellow", metric=metric)


def _build_ads_tactical_card(record, ads_current, campaign_design):
    metric = clean(record.get("metric"), "Ads signal")
    roi = to_number(ads_current.get("roi"), 0)
    active = bool(ads_current.get("active", False))
    bookings = to_number(ads_current.get("bookings_usd"), 0)
    bookings_ars = bookings * ARS_PER_USD
    reco = clean(campaign_design.get("ads_action"), "") if campaign_design else ""
    descr = clean(record.get("description"), "")

    if active:
        if "upselling" in norm_text(descr) or "Upselling" in descr:
            main = f"🚀 Ads Upselling · ROI {fmt_roi(roi)}"
            action_note = f"Presupuesto consumido antes de 7d — propón subida: ARS {fmt_number(_round_budget_up_ars(bookings_ars * 1.25))} semanal estimado."
        else:
            main = f"🚀 Ads activo · ROI {fmt_roi(roi)}"
            action_note = f"Booking semanal aprox. {fmt_usd(bookings)}. Sostener si ROI ≥2.5 · ajustar CPC si bajo."
    else:
        main = "🚀 Ads Acquisition"
        budget_hint = clean(campaign_design.get("ads_budget_range"), "") if campaign_design else ""
        action_note = f"Test controlado. Budget sugerido: {budget_hint or 'ARS 15k–25k semanal'}. Medir ROI semana 1 antes de escalar."

    cue_parts = []
    if descr:
        cue_parts.append(f"SP: {descr}.")
    if reco:
        cue_parts.append(reco)
    else:
        cue_parts.append(action_note)
    cue = " ".join(cue_parts)
    cls = "health-green" if active and roi >= 2.5 else "health-yellow"
    return _make_tactical_item("🚀 " + metric, main, "", cue, cls, metric=metric)


def _build_pro_tactical_card(record, pro, campaign_design):
    metric = clean(record.get("metric"), "PRO signal")
    pro_extra = campaign_design.get("pro_extra", 5) if campaign_design else 5
    pro_reason = clean(campaign_design.get("pro_reason"), "") if campaign_design else ""
    descr = clean(record.get("description"), "")
    main = f"👑 PRO · Actual {fmt_percent0(pro)}"
    cue_parts = []
    if descr:
        cue_parts.append(f"SP: {descr}.")
    if pro_reason:
        cue_parts.append(pro_reason)
    else:
        cue_parts.append(f"Propón +{int(pro_extra)}% PRO sobre producto top. Upselling simbólico — no regalar a todo el tráfico.")
    cue = " ".join(cue_parts)
    return _make_tactical_item("👑 " + metric, main, "", cue, "health-yellow", metric=metric)


def _build_general_tactical_card(record):
    metric = clean(record.get("metric"), "Priority lever")
    descr = clean(record.get("description"), "")
    cue = f"SP: {descr}" if descr else "Pregunta de validación: ¿qué ajuste concreto dejamos definido hoy con esta palanca?"
    return _make_tactical_item(
        "🎯 " + metric,
        "🟡 Palanca SP detectada",
        "",
        cue,
        "health-yellow",
        metric=metric,
    )


def build_tactical_flow(brand_id, name, row, category, current, ads_current, md_current, md_pro_current, booster, actions, campaign_design):
    """Turns Smart Priorities into a call guide.

    Mission: help Sabas adhere to Priority Signals without losing speech quality.
    It creates one card per priority lever only. It does not force OPS/Menu/MD/Ads cards
    when Priority did not ask for them.
    """
    signals = get_priority_signals_for_brand(brand_id, name)
    records = _ordered_priority_records(_priority_lever_records(signals))

    cr = _normalize_rate_value(get_from_row(row, ["cr %", "conversion rate", "conversion"], 0))
    pro = _normalize_rate_value(get_from_row(row, ["pro users %", "pro %", "pro users", "prime users %"], 0))
    aov_ars = to_number(current.get("aov_ars") if current else 0, 0)
    orders_monthly = to_number(current.get("orders") if current else 0, 0) * 4
    current_gmv_ars = to_number(current.get("gmv_ars") if current else 0, 0)
    menu_metrics = get_menu_metrics_for_brand(name)
    ops_metrics = get_ops_metrics_for_brand(name)

    items = []
    for idx, record in enumerate(records, start=1):
        kind = record.get("kind", "general")
        if kind.startswith("ops"):
            item = _build_ops_tactical_card(record, name, ops_metrics, menu_metrics, ads_current, md_current, aov_ars=aov_ars, orders_monthly=orders_monthly, current_gmv_ars=current_gmv_ars)
        elif kind.startswith("menu"):
            item = _build_menu_tactical_card(record, name, menu_metrics, campaign_design)
        elif kind == "md":
            item = _build_md_tactical_card(record, md_current, campaign_design, signals)
        elif kind == "ads":
            item = _build_ads_tactical_card(record, ads_current, campaign_design)
        elif kind == "pro":
            item = _build_pro_tactical_card(record, pro, campaign_design)
        else:
            item = _build_general_tactical_card(record)
        item["order"] = idx
        item["score"] = record.get("score")
        item["lever_class"] = _lever_class_from_kind(kind)
        items.append(item)

    return {
        "found_priority": bool(signals.get("found")),
        "priority_score": signals.get("score") if signals.get("found") else None,
        "lever_texts": [clean(r.get("metric"), "-") for r in records],
        "conversation_order": _conversation_order_text(records),
        "rank": signals.get("rank") if signals.get("found") else None,
        "last_contact": signals.get("last_contact") if signals.get("found") else "-",
        "coinversion": signals.get("coinversion") if signals.get("found") else "No",
        "promo_vencida": signals.get("promo_vencida") if signals.get("found") else 0,
        "promo_por_vencer": signals.get("promo_por_vencer") if signals.get("found") else 0,
        "items": items,
    }


def render_tactical_flow_html(flow):
    if not flow:
        return ""
    levers = flow.get("lever_texts", [])
    if levers:
        lever_html = "".join([f"<span class='priority-chip'>{html.escape(clean(x, '-'))}</span>" for x in levers[:10]])
    else:
        lever_html = "<span class='priority-chip'>✅ No priority lever forcing the call</span>"

    order = flow.get("conversation_order", [])
    if order:
        order_html = "".join([f"<span class='priority-chip'><strong>{html.escape(clean(x, '-'))}</strong></span>" for x in order[:10]])
    else:
        order_html = "<span class='priority-chip'>No suggested order because no Priority levers were detected.</span>"

    cards_html = "".join([
        f"<div class='tactical-card {html.escape(clean(item.get('class'), 'health-green'))} {html.escape(clean(item.get('lever_class'), ''))}'>"
        f"<div class='tactical-area'>{html.escape(str(item.get('order', '-')))} · {html.escape(clean(item.get('area'), '-'))}</div>"
        f"<div class='tactical-main'>{html.escape(clean(item.get('main'), '✅ Following'))}</div>"
        f"<div class='tactical-cue'>{html.escape(clean(item.get('cue') or item.get('argument'), ''))}</div>"
        f"</div>"
        for item in flow.get("items", [])
    ])
    if not cards_html:
        cards_html = (
            "<div class='tactical-card health-green'>"
            "<div class='tactical-area'>✅ No Priority Levers</div>"
            "<div class='tactical-main'>Following</div>"
            "<div class='tactical-cue'>Usa 360 Action y Campaign Designer si decides abrir otro frente.</div>"
            "</div>"
        )

    score_text = _priority_score_display(flow.get("priority_score")) if flow.get("found_priority") else "No Priority Data"
    return (
        f"<div class='wide-info-card tactical-flow-card'>"
        f"<div class='wide-info-title'>🎯 Tactical Flow</div>"
        f"<div class='priority-top-grid'>"
        f"<div><div class='info-mini-label'>Priority Score</div><div class='info-mini-value'>{html.escape(score_text)}</div></div>"
        f"<div><div class='info-mini-label'># Contact Order</div><div class='info-mini-value'>{html.escape('#' + str(int(flow.get('rank'))) if flow.get('rank') not in [None, '', '-'] and not pd.isna(flow.get('rank')) else '-')}</div></div>"
        f"<div><div class='info-mini-label'>Last Contact</div><div class='info-mini-value'>{html.escape(clean(flow.get('last_contact'), '-'))}</div></div>"
        f"<div><div class='info-mini-label'>Coinversion MD</div><div class='info-mini-value'>{html.escape(clean(flow.get('coinversion'), 'No'))}</div></div>"
        f"<div><div class='info-mini-label'>Expired / Expiring MD</div><div class='info-mini-value'>{int(flow.get('promo_vencida') or 0)} / {int(flow.get('promo_por_vencer') or 0)}</div></div>"
        f"<div class='priority-levers'>{lever_html}</div>"
        f"<div class='priority-note'><strong>Suggested conversation order:</strong></div>"
        f"<div class='priority-levers'>{order_html}</div>"
        f"</div>"
        f"<div class='tactical-grid'>{cards_html}</div>"
        f"<div class='priority-note'>Tactical Flow se adhiere únicamente a Smart Priorities: una tarjeta por palanca detectada, con una micro-instrucción de llamada.</div>"
        f"</div>"
    )

def _merge_growth_manual_status(row, ads_current, md_current, md_pro_current):
    """Lets manual Activate/OFF updates in Growth OS reflect immediately in Finder Business Information.
    Ads Bookings rule:
    - Growth OS manual value is treated as confirmed WEEKLY booking.
    - Current ADS value is treated as accumulated MTD and normalized to weekly approximation upstream.
    - Revenue remains accumulated MTD.
    """
    ads = dict(ads_current)
    md = dict(md_current)
    mdp = dict(md_pro_current)

    ads_status = clean(get_from_row(row, ["ads"], ""))
    md_status = clean(get_from_row(row, ["md", "md status"], ""))

    if _is_active_status(ads_status):
        ads["active"] = True
        row_bookings = to_number(get_from_row(row, ["ads bookings", "ad bookings", "ads bookings ars", "booking ads"], 0), 0)
        row_roi = to_number(get_from_row(row, ["ads roi", "ad roi"], 0), 0)
        if row_bookings:
            # Manual value is Sabas-confirmed weekly booking in ARS.
            ads["bookings_usd"] = row_bookings / ARS_PER_USD
            ads["bookings_source"] = "growth_os_manual"
            ads["bookings_is_approx"] = False
        elif ads.get("bookings_usd", 0):
            ads.setdefault("bookings_source", "current_ads_approx")
            ads.setdefault("bookings_is_approx", True)
        if ads.get("roi", 0) == 0 and row_roi:
            ads["roi"] = row_roi
    elif "off" in ads_status.lower() or "inactive" in ads_status.lower() or "💤" in ads_status or "😴" in ads_status:
        ads.update({"active": False, "bookings_usd": 0, "bookings_accum_usd": 0, "bookings_source": "none", "bookings_is_approx": False, "revenue_usd": 0, "sales_usd": 0, "roi": 0})

    if _is_active_status(md_status):
        md["active"] = True
        # MD manual value is a discount/promo label, not money.
        # Do not convert it into sales_usd. Current MD sheet remains the source for MD sales.
        row_roi = to_number(get_from_row(row, ["md roi"], 0), 0)
        if md.get("roi", 0) == 0 and row_roi:
            md["roi"] = row_roi
    elif "off" in md_status.lower() or "inactive" in md_status.lower() or "💤" in md_status or "😴" in md_status:
        md.update({"active": False, "sales_usd": 0, "gmv_usd": 0, "campaigns": 0, "orders": 0, "roi": 0})

    return ads, md, mdp



def _action_area_lever_class(area):
    text = norm_text(area)
    if "ops" in text or "oper" in text:
        return "lever-ops"
    if "menu" in text or "catalog" in text:
        return "lever-menu"
    if "md pro" in text or "pro" in text:
        return "lever-pro"
    if "md" in text or "markdown" in text:
        return "lever-md"
    if "ads" in text:
        return "lever-ads"
    return ""


def _roi_chip(roi_value, benchmark=3.2):
    roi = to_number(roi_value, 0)
    if roi == 0:
        return ""

    # Gauge parameters: arc goes from -180° to 0° (left to right)
    # We cap the display at 2x benchmark
    cap = benchmark * 2
    ratio = min(roi / cap, 1.0)

    # Needle angle: -180deg = 0 ROI, 0deg = cap
    import math as _math
    needle_angle_deg = -180 + ratio * 180
    needle_rad = _math.radians(needle_angle_deg)
    cx, cy, r = 54, 42, 36
    nx = cx + (r - 6) * _math.cos(needle_rad)
    ny = cy + (r - 6) * _math.sin(needle_rad)

    # Color zones (fill arcs via path)
    # Red zone: 0 → 60% of benchmark
    # Orange zone: 60% → 100% of benchmark
    # Green zone: benchmark → cap
    def arc_path(start_ratio, end_ratio):
        sa = _math.radians(-180 + start_ratio * 180)
        ea = _math.radians(-180 + end_ratio * 180)
        x1 = cx + r * _math.cos(sa)
        y1 = cy + r * _math.sin(sa)
        x2 = cx + r * _math.cos(ea)
        y2 = cy + r * _math.sin(ea)
        xi1 = cx + (r - 9) * _math.cos(sa)
        yi1 = cy + (r - 9) * _math.sin(sa)
        xi2 = cx + (r - 9) * _math.cos(ea)
        yi2 = cy + (r - 9) * _math.sin(ea)
        return f"M {x1:.1f} {y1:.1f} A {r} {r} 0 0 1 {x2:.1f} {y2:.1f} L {xi2:.1f} {yi2:.1f} A {r-9} {r-9} 0 0 0 {xi1:.1f} {yi1:.1f} Z"

    red_end   = (benchmark * 0.6) / cap
    orange_end = benchmark / cap

    if roi >= benchmark:
        label_color, label_text = "#7ED321", f"ROI {fmt_roi(roi_value)} ✓"
    elif roi >= benchmark * 0.6:
        label_color, label_text = "#D95A10", f"ROI {fmt_roi(roi_value)} ~"
    else:
        label_color, label_text = "#FF4D2E", f"ROI {fmt_roi(roi_value)} ↓"

    gauge_svg = (
        f'<svg width="108" height="54" viewBox="0 0 108 54" xmlns="http://www.w3.org/2000/svg">'
        # Red arc
        f'<path d="{arc_path(0, red_end)}" fill="#FF4D2E" opacity="0.85"/>'
        # Orange arc
        f'<path d="{arc_path(red_end, orange_end)}" fill="#FF7124" opacity="0.85"/>'
        # Green arc
        f'<path d="{arc_path(orange_end, 1.0)}" fill="#7ED321" opacity="0.85"/>'
        # Needle
        f'<line x1="{cx}" y1="{cy}" x2="{nx:.1f}" y2="{ny:.1f}" stroke="#FFFFFF" stroke-width="2.5" stroke-linecap="round"/>'
        # Center dot
        f'<circle cx="{cx}" cy="{cy}" r="3.5" fill="#FFFFFF"/>'
        # Benchmark tick
        f'<line x1="{cx + r * _math.cos(_math.radians(-180 + orange_end * 180)):.1f}" y1="{cy + r * _math.sin(_math.radians(-180 + orange_end * 180)):.1f}" x2="{cx + (r + 4) * _math.cos(_math.radians(-180 + orange_end * 180)):.1f}" y2="{cy + (r + 4) * _math.sin(_math.radians(-180 + orange_end * 180)):.1f}" stroke="#FFFFFF" stroke-width="1.5"/>'
        f'<text x="{cx}" y="52" text-anchor="middle" font-size="7" fill="#6B7280" font-weight="700">/{benchmark}x</text>'
        '</svg>'
    )

    return (
        f'<div style="margin-top:10px;display:flex;flex-direction:column;align-items:flex-start;gap:2px;">'
        f'{gauge_svg}'
        f'<span style="font-size:11px;font-weight:900;color:{label_color};">{label_text}</span>'
        f'</div>'
    )


def _business_card(label, value, copy="", lever_class="", chip=""):
    chip_html = f"<div class='card-chipline'><span class='card-chip'>{html.escape(clean(chip, ''))}</span></div>" if clean(chip, "") else ""
    return (
        f"<div class='business-mini-card {html.escape(clean(lever_class, ''))}'>"
        f"<div class='card-label'>{html.escape(clean(label, '-'))}</div>"
        f"<div class='card-value'>{value}</div>"
        f"<div class='card-copy'>{html.escape(clean(copy, ''))}</div>"
        f"{chip_html}</div>"
    )


def render_business_cards_html(ads_current, md_current, md_pro_current, campaign_names, ads_booking_display, pro_users_display, conversion_display, commission_display, pro_users_raw=0, conversion_raw=0, commission_raw=0, cvr_weekly=None, cvr_source='Sin datos', cvr_bench=None):
    import math as _math

    def _waffle_icons(filled_count, total=10, filled_color="#7ED321", empty_color="rgba(255,255,255,0.15)"):
        person_path = '<circle cx="12" cy="7" r="4"/><path d="M4 21c0-4.418 3.582-8 8-8s8 3.582 8 8"/>'
        icons = []
        for i in range(total):
            color = filled_color if i < filled_count else empty_color
            icons.append(
                f'<svg width="20" height="20" viewBox="0 0 24 24" fill="{color}" '
                f'xmlns="http://www.w3.org/2000/svg" style="flex-shrink:0;">'
                f'{person_path}</svg>'
            )
        return '<div style="display:flex;gap:2px;flex-wrap:wrap;margin-top:6px;">' + "".join(icons) + '</div>'

    def _commission_gauge(commission_val, benchmark=0.22):
        """Commission bullet chart — horizontal bar with zone tints and benchmark tick."""
        cap = benchmark * 2.5
        if cap == 0:
            return ""
        width, height = 68, 26
        W, H = width, height
        pad_l, pad_r = 3, 3
        bar_w = W - pad_l - pad_r
        bar_h = 7
        bar_y = H - bar_h - 10
        rx    = 3

        # Commission zones: green (0→benchmark), orange (benchmark→1.25x), red (1.25x→cap)
        green_end  = benchmark / cap
        orange_end = min(benchmark * 1.25 / cap, 1.0)
        fill_ratio = min(commission_val / cap, 1.0)
        fill_w     = fill_ratio * bar_w
        bm_x       = pad_l + green_end * bar_w

        # Color: green if under benchmark, orange approaching, red over
        if commission_val >= benchmark:
            nc = "#FF4D2E"
        elif commission_val >= benchmark * 0.6:
            nc = "#FF7124"
        else:
            nc = "#7ED321"

        green_w  = green_end * bar_w
        orange_w = (orange_end - green_end) * bar_w
        red_w    = bar_w - green_w - orange_w
        orange_x = pad_l + green_w
        red_x    = orange_x + orange_w
        label_y  = bar_y - 3

        svg_parts = [
            f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" style="overflow:visible;">',
            f'<rect x="{pad_l}" y="{bar_y}" width="{bar_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(255,255,255,0.95)"/>',
            f'<rect x="{pad_l:.1f}" y="{bar_y}" width="{green_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(191,255,0,0.28)"/>',
            f'<rect x="{orange_x:.1f}" y="{bar_y}" width="{orange_w:.1f}" height="{bar_h}" fill="rgba(255,138,61,0.22)"/>',
            f'<rect x="{red_x:.1f}" y="{bar_y}" width="{red_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(255,92,122,0.25)"/>',
            f'<rect x="{pad_l}" y="{bar_y + 1}" width="{fill_w:.1f}" height="{bar_h - 2}" rx="{rx - 1}" fill="{nc}" opacity="0.90"/>',
            f'<rect x="{bm_x - 0.8:.1f}" y="{bar_y - 2}" width="1.6" height="{bar_h + 4}" rx="0.8" fill="#FFFFFF" opacity="0.40"/>',
            f'<text x="{pad_l}" y="{label_y}" font-size="6.5" font-weight="900" fill="{nc}">{fmt_percent0(commission_val)}</text>',
            f'<text x="{W - pad_r}" y="{label_y}" text-anchor="end" font-size="5.5" fill="rgba(107,114,128,0.60)" font-weight="700">/{fmt_percent0(benchmark)}</text>',
            '</svg>',
        ]
        return f'<div style="position:absolute;bottom:8px;right:8px;opacity:0.95;">{"".join(svg_parts)}</div>'

    def _needle_gauge_svg_bf(roi_val, bmark=3.2, width=68, height=26):
        """Compact bullet chart for Brand Finder cards — horizontal bar, bottom-right corner."""
        if not roi_val or roi_val == 0:
            return ""
        max_val    = max(bmark * 2.5, roi_val * 1.2 + 0.1)
        W, H       = width, height
        pad_l, pad_r = 3, 3
        bar_w      = W - pad_l - pad_r
        bar_h      = 7
        bar_y      = H - bar_h - 10
        rx         = 3

        red_end    = bmark * 0.6 / max_val
        orange_end = bmark       / max_val
        fill_ratio = min(max(roi_val / max_val, 0), 1)
        fill_w     = fill_ratio * bar_w
        bm_x       = pad_l + (bmark / max_val) * bar_w
        nc = "#7ED321" if roi_val >= bmark else ("#FF7124" if roi_val >= bmark * 0.6 else "#FF4D2E")

        red_w    = (bmark * 0.6 / max_val) * bar_w
        orange_w = ((bmark - bmark * 0.6) / max_val) * bar_w
        green_w  = bar_w - red_w - orange_w
        red_x    = pad_l
        orange_x = red_x + red_w
        green_x  = orange_x + orange_w

        label_y = bar_y - 3

        svg_parts = [
            f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" style="overflow:visible;">',
            # Track
            f'<rect x="{pad_l}" y="{bar_y}" width="{bar_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(255,255,255,0.95)"/>',
            # Zone tints
            f'<rect x="{red_x:.1f}" y="{bar_y}" width="{red_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(255,92,122,0.28)"/>',
            f'<rect x="{orange_x:.1f}" y="{bar_y}" width="{orange_w:.1f}" height="{bar_h}" fill="rgba(255,138,61,0.22)"/>',
            f'<rect x="{green_x:.1f}" y="{bar_y}" width="{green_w:.1f}" height="{bar_h}" rx="{rx}" fill="rgba(191,255,0,0.25)"/>',
            # Filled bar
            f'<rect x="{pad_l}" y="{bar_y + 1}" width="{fill_w:.1f}" height="{bar_h - 2}" rx="{rx - 1}" fill="{nc}" opacity="0.90"/>',
            # Benchmark tick
            f'<rect x="{bm_x - 0.8:.1f}" y="{bar_y - 2}" width="1.6" height="{bar_h + 4}" rx="0.8" fill="#FFFFFF" opacity="0.40"/>',
            # Value label
            f'<text x="{pad_l}" y="{label_y}" font-size="6.5" font-weight="900" fill="{nc}">{roi_val:.1f}x</text>',
            # Benchmark label
            f'<text x="{W - pad_r}" y="{label_y}" text-anchor="end" font-size="5.5" fill="rgba(107,114,128,0.60)" font-weight="700">/{bmark}x</text>',
            '</svg>',
        ]
        return f'<div style="position:absolute;top:8px;right:8px;opacity:0.95;">{"".join(svg_parts)}</div>'

    def _card_with_roi(label, value, copy, lever_class, roi_value, benchmark=3.2):
        roi_html = _needle_gauge_svg_bf(to_number(roi_value, 0), bmark=benchmark)
        return (
            f"<div class='business-mini-card {html.escape(clean(lever_class, ''))}' style='position:relative;overflow:hidden;'>"
            f"<div class='card-label'>{html.escape(clean(label, '-'))}</div>"
            f"<div class='card-value'>{value}</div>"
            f"<div class='card-copy'>{html.escape(clean(copy, ''))}</div>"
            f"{roi_html}</div>"
        )

    # PRO Users card with waffle
    pro_pct = round(pro_users_raw * 100) if pro_users_raw <= 1 else round(pro_users_raw)
    pro_icons = max(0, min(10, round(pro_pct / 10)))
    pro_waffle_html = _waffle_icons(pro_icons, filled_color="#7ED321", empty_color="rgba(255,255,255,0.15)")
    pro_card = (
        f"<div class='business-mini-card lever-pro'>"
        f"<div class='card-label'>PRO Users</div>"
        f"<div class='card-value'>{html.escape(clean(pro_users_display, '-'))}</div>"
        f"<div class='card-copy'>User mix / Prime opportunity</div>"
        f"{pro_waffle_html}"
        f'<div style="font-size:10px;color:#6B7280;margin-top:4px;">{pro_icons} de 10 · 1 figura = 10%</div>'
        f"</div>"
    )

    # Conversion Rate card — CVR semanal real + benchmark por categoría
    _cvr_val = cvr_weekly if cvr_weekly is not None else conversion_raw
    _cvr_norm = (_cvr_val if _cvr_val <= 1 else _cvr_val / 100) if _cvr_val else 0
    cr_pct = round(_cvr_norm * 100)
    cr_icons = max(0, min(10, round(cr_pct / 10)))
    cr_waffle_html = _waffle_icons(cr_icons, filled_color="#1B3F8B", empty_color="rgba(255,255,255,0.15)")
    _cvr_main = fmt_percent0(_cvr_norm) if _cvr_val is not None else clean(conversion_display, '-')
    _cvr_src_html = f"<span style='font-size:9px;color:rgba(107,114,128,0.60);margin-left:4px;'>({html.escape(cvr_source)})</span>"
    _bench_html = ""
    if cvr_bench is not None:
        _bn = cvr_bench if cvr_bench <= 1 else cvr_bench / 100
        _delta = _cvr_norm - _bn
        _dc = "#7ED321" if _delta >= 0 else "#FF4D2E"
        _ds = "+" if _delta >= 0 else ""
        _bench_html = (
            f'<div style="font-size:10px;color:#6B7280;margin-top:4px;">'
            f'Benchmark categ.: {fmt_percent0(_bn)} '
            f'<span style="color:{_dc};font-weight:700;">{_ds}{fmt_percent0(_delta)}</span>'
            f'</div>'
        )
    cr_card = (
        f"<div class='business-mini-card lever-menu'>"
        f"<div class='card-label'>Conversion Rate {_cvr_src_html}</div>"
        f"<div class='card-value'>{_cvr_main}</div>"
        f"<div class='card-copy'>Últimas 4 semanas · vs benchmark categoría</div>"
        f"{cr_waffle_html}"
        f"{_bench_html}"
        f'<div style="font-size:10px;color:#6B7280;margin-top:2px;">{cr_icons} de 10 · 1 figura = 10%</div>'
        f"</div>"
    )

    # Commission Rate card with gauge (benchmark 22%)
    comm_gauge_html = _commission_gauge(commission_raw, benchmark=0.22)
    comm_card = (
        f"<div class='business-mini-card lever-md' style='position:relative;overflow:hidden;'>"
        f"<div class='card-label'>Commission Rate</div>"
        f"<div class='card-value'>{html.escape(clean(commission_display, '-'))}</div>"
        f"<div class='card-copy'>Margin guardrail for promo pressure</div>"
        f"{comm_gauge_html}"
        f"</div>"
    )

    cards = []
    cards.append(_card_with_roi("Ads", html.escape(status_from_active(ads_current.get('active', False))), f"Booking {clean(ads_booking_display, '—')} · Revenue {fmt_ars(to_number(ads_current.get('revenue_usd'),0) * ARS_PER_USD)}", "lever-ads", ads_current.get('roi'), benchmark=4.5))
    cards.append(_card_with_roi("Markdown", html.escape(status_from_active(md_current.get('active', False))), f"Campaign {clean(campaign_names.get('md'), '-')}", "lever-md", md_current.get('roi'), benchmark=3.2))
    cards.append(_card_with_roi("Markdown PRO", html.escape(status_from_active(md_pro_current.get('active', False))), f"Campaign {clean(campaign_names.get('md_pro'), '-')}", "lever-pro", md_pro_current.get('roi'), benchmark=3.2))
    cards.append(pro_card)
    cards.append(cr_card)
    cards.append(comm_card)
    return f"<div class='wide-info-card'><div class='wide-info-title'>Business Information + Portfolio Metrics</div><div class='business-card-grid'>{''.join(cards)}</div></div>"

@st.cache_data(ttl=600, show_spinner=False)
def _build_pareto_hub_data():
    """
    Construye los datos completos de todas las marcas Tier A (80% de GMV)
    para el Pareto Hub: cruza Growth OS, Current GMV/ADS/MD/MD PRO/Churn,
    Perfect Store, CVR%/Traffic, y Priority Data (requerimiento de PDF).
    Devuelve lista de dicts, uno por marca, ya con la clasificación de salud.
    """
    growth_df = load_growth_data()
    if growth_df.empty:
        return []

    id_col = get_id_column_name(growth_df)
    if not id_col:
        return []

    tiers_map = get_pareto_tiers_map()
    tier_a_ids = {bid for bid, t in tiers_map.items() if t == "A"}
    if not tier_a_ids:
        return []

    # Restringir al portafolio vigente (Asignacion Junio) por si Current GMV
    # todavía trae marcas que ya fueron reasignadas a otro Farmer.
    try:
        _aj_pareto = load_asignacion_activa()
        if not _aj_pareto.empty:
            _aj_pareto_ids = set(_aj_pareto["brand_id"].dropna().astype(str))
            _aj_pareto_ids.discard("")
            if _aj_pareto_ids:
                tier_a_ids = tier_a_ids & _aj_pareto_ids
    except Exception:
        pass
    if not tier_a_ids:
        return []

    prod_map = get_productivity_last_contact_map(EXCEL_FILE)
    meta_map = get_last_comment_meta_map(limit=1)
    priority_df = load_priority_data()

    # Set de brand_ids que tienen una fila "PDF Menu" pendiente en Priority Data
    _pdf_required_ids = set()
    if not priority_df.empty and "_metric_norm" in priority_df.columns:
        _pdf_rows = priority_df[priority_df["_metric_norm"] == norm_text("PDF Menu")]
        _pdf_required_ids = set(_pdf_rows["_id"].apply(normalize_brand_id))

    rows = []
    for _, row in growth_df.iterrows():
        bid = normalize_brand_id(row.get(id_col))
        if bid not in tier_a_ids:
            continue

        name = clean(get_from_row(row, ["name", "brand name", "restaurant name"]), "-")
        category = clean(get_from_row(row, ["category"]), "-")
        category_main, _ = _split_category_and_stickers(category)

        ads_m    = get_current_ads_metrics(bid)
        md_m     = get_current_md_metrics(bid, pro=False)
        mdpro_m  = get_current_md_metrics(bid, pro=True)
        churn_lbl = get_churn_status(bid)

        menu_metrics = get_menu_metrics_for_brand(name)
        perfect_store_pct = round(menu_metrics.get("health_score", 0)) if menu_metrics.get("found") else None
        requires_pdf = bid in _pdf_required_ids

        cvr_raw, _ = get_cvr_for_brand(name, cr_fallback=get_from_row(row, ["cr %", "conversion rate"], 0))
        cvr_bench  = get_cvr_category_benchmark(category_main)
        traffic_raw   = get_traffic_for_brand(name)
        traffic_bench = get_traffic_category_benchmark(category_main)

        last_dt = get_last_contact_dt(bid, name, prod_map=prod_map, meta_map=meta_map)
        days_since = _days_since_timestamp(last_dt)

        ads_active   = bool(ads_m.get("active", False))
        md_active    = bool(md_m.get("active", False))
        mdpro_active = bool(mdpro_m.get("active", False))
        ads_roi   = to_number(ads_m.get("roi"), 0)
        md_roi    = to_number(md_m.get("roi"), 0)
        mdpro_roi = to_number(mdpro_m.get("roi"), 0)

        # ── Clasificación de salud (verde / azul / review / tangerine) ──────────
        # Regla: Acquisition = le falta activar al menos UNA de las 3 palancas
        # (0, 1 o 2 activas). Si tiene las TRES activas:
        #   - alguna con ROI/ROAS por debajo de 3.5x -> Review (color distinto, no Acquisition)
        #   - las tres por encima de 3.5x -> Sana (verde) si hay contacto reciente
        #     y Perfect Store ok, sino Upselling (azul) como reconocimiento del buen ROI
        #     sin las otras condiciones de salud cumplidas todavía
        _is_recent_contact = (days_since is not None and days_since <= 21)
        _perfect_store_ok  = (perfect_store_pct is not None and perfect_store_pct > 90 and not requires_pdf)
        _has_all_three = ads_active and md_active and mdpro_active
        _needs_acquisition = not _has_all_three

        if _needs_acquisition:
            health = "tangerine"
        else:
            _all_rois = [ads_roi, md_roi, mdpro_roi]
            _below_review_threshold = any(r < 3.5 for r in _all_rois)

            if _below_review_threshold:
                health = "review"
            elif _is_recent_contact and _perfect_store_ok:
                health = "green"
            else:
                health = "blue"

        _acq_missing = []
        if not ads_active:
            _acq_missing.append("Ads")
        if not md_active:
            _acq_missing.append("MD")
        if not mdpro_active:
            _acq_missing.append("MD PRO")

        rows.append({
            "brand_id":    bid,
            "name":        name,
            "category":    category_main,
            "last_contact_days": days_since,
            "ads_active":  ads_active,  "ads_roi":  ads_roi,
            "md_active":   md_active,   "md_roi":   md_roi,
            "mdpro_active": mdpro_active, "mdpro_roi": mdpro_roi,
            "perfect_store_pct": perfect_store_pct,
            "requires_pdf": requires_pdf,
            "churn_label": churn_lbl,
            "cvr_brand":   cvr_raw,
            "cvr_bench":   cvr_bench,
            "traffic_brand": traffic_raw,
            "traffic_bench": traffic_bench,
            "health":      health,
            "acq_missing": _acq_missing,
        })

    return rows


def page_pareto_hub():
    render_header("Pareto Hub", "Marcas que representan el 80% del GMV total · Tier A")

    data = _build_pareto_hub_data()
    if not data:
        st.info("No se pudo construir el Pareto Hub — verificá que Current GMV y Growth OS tengan datos cargados.")
        return

    _HEALTH_STYLE = {
        "green":     {"border": "#7ED321", "bg": "rgba(126,211,33,0.14)",  "label": "🟢 Sana"},
        "blue":      {"border": "#1B3F8B", "bg": "rgba(46,107,255,0.14)",  "label": "🔵 Upselling"},
        "review":    {"border": "#D9A300", "bg": "rgba(255,196,0,0.16)",   "label": "🟡 Review"},
        "tangerine": {"border": "#FF7124", "bg": "rgba(255,113,36,0.14)",  "label": "🟠 Acquisition"},
    }

    _n_green  = sum(1 for d in data if d["health"] == "green")
    _n_blue   = sum(1 for d in data if d["health"] == "blue")
    _n_review = sum(1 for d in data if d["health"] == "review")
    _n_tang   = sum(1 for d in data if d["health"] == "tangerine")

    if "_pareto_health_filter" not in st.session_state:
        st.session_state["_pareto_health_filter"] = "all"

    _filter_cols = st.columns(5)
    _filter_specs = [
        ("all",       f"📊 Todas ({len(data)})",        _filter_cols[0]),
        ("green",     f"🟢 Sanas ({_n_green})",          _filter_cols[1]),
        ("blue",      f"🔵 Upselling ({_n_blue})",       _filter_cols[2]),
        ("review",    f"🟡 Review ({_n_review})",        _filter_cols[3]),
        ("tangerine", f"🟠 Acquisition ({_n_tang})",     _filter_cols[4]),
    ]
    for _health_key, _btn_label, _col in _filter_specs:
        with _col:
            _is_selected = st.session_state["_pareto_health_filter"] == _health_key
            if st.button(_btn_label, key=f"pareto_filter_{_health_key}", use_container_width=True,
                         type="primary" if _is_selected else "secondary"):
                st.session_state["_pareto_health_filter"] = _health_key
                st.rerun()

    _active_filter = st.session_state["_pareto_health_filter"]
    if _active_filter != "all":
        data = [d for d in data if d["health"] == _active_filter]
        if not data:
            st.info(f"No hay marcas en el filtro seleccionado. Volviendo a 'Todas'.")
            st.session_state["_pareto_health_filter"] = "all"
            st.rerun()

    st.markdown("""
    <style>
    .pareto-scroll {
        max-height: 760px;
        overflow-y: auto;
        padding-right: 6px;
    }
    .pareto-grid {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 14px;
        margin-bottom: 14px;
    }
    .pareto-card {
        border-radius: 14px;
        padding: 14px 16px;
        transition: transform .22s cubic-bezier(.34,1.56,.64,1), box-shadow .2s ease;
        cursor: pointer;
    }
    .pareto-card:hover {
        transform: translateY(-4px) scale(1.03);
        box-shadow: 0 12px 30px rgba(0,0,0,0.12);
    }
    .pareto-name { font-size: 14px; font-weight: 800; color: #1A1A2E; line-height: 1.2; }
    .pareto-meta { font-size: 11px; color: #6B7280; margin-top: 2px; margin-bottom: 8px; }
    .pareto-row { display: flex; justify-content: space-between; font-size: 11px; padding: 2px 0; }
    .pareto-row-label { color: #6B7280; }
    .pareto-row-value { font-weight: 700; color: #1A1A2E; }
    .pareto-badge {
        display: inline-block; font-size: 9px; font-weight: 800; letter-spacing: .04em;
        text-transform: uppercase; padding: 2px 8px; border-radius: 10px; margin-top: 8px;
    }
    </style>
    """, unsafe_allow_html=True)

    def _fmt_roi_cell(active, roi):
        if not active:
            return '<span style="color:#FF7124;">No</span>'
        return f'<span style="color:#7ED321;">Sí ({fmt_ratio(roi)})</span>'

    def _fmt_cvr_cell(brand, bench):
        if not brand or brand <= 0:
            return '<span style="color:#aaa;">s/d</span>'
        brand_pct = brand if brand <= 1 else brand / 100
        bench_pct = bench if bench and bench > 0 else None
        if bench_pct:
            color = "#7ED321" if brand_pct >= bench_pct else "#FF4D2E"
            return f'<span style="color:{color};">{round(brand_pct*100,1)}% (bench {round(bench_pct*100,1)}%)</span>'
        return f'{round(brand_pct*100,1)}%'

    def _fmt_traffic_cell(brand, bench):
        if not brand or brand <= 0:
            return '<span style="color:#aaa;">s/d</span>'
        if bench and bench > 0:
            color = "#7ED321" if brand >= bench else "#FF4D2E"
            return f'<span style="color:{color};">{round(brand):,}/sem (bench {round(bench):,})</span>'.replace(",", ".")
        return f'{round(brand):,}/sem'.replace(",", ".")

    # ── Render en filas de 4 cards con scroll ──────────────────────────────────
    st.markdown('<div class="pareto-scroll">', unsafe_allow_html=True)

    _sorted_data = sorted(data, key=lambda d: (d["health"] != "tangerine", d["health"] != "review", d["health"] != "blue", d["name"]))

    def _build_pareto_card_html(d):
        style = _HEALTH_STYLE[d["health"]]
        _days_lbl = f"{d['last_contact_days']}d" if d["last_contact_days"] is not None else "Sin contacto"
        _ps_lbl = (
            f'{d["perfect_store_pct"]}%' + (' · requiere PDF' if d["requires_pdf"] else '')
            if d["perfect_store_pct"] is not None else "s/d"
        )
        _acq_note = (
            f'<div class="pareto-badge" style="background:{style["bg"]};color:{style["border"]};">'
            f'Falta: {", ".join(d["acq_missing"])}</div>'
        ) if d["health"] == "tangerine" and d["acq_missing"] else ""

        # NOTA: cada línea sin indentación inicial — un f-string multilínea con
        # 4+ espacios al comienzo de línea es interpretado por el parser de
        # Markdown de Streamlit como bloque de código, mostrando el HTML crudo
        # en vez de renderizarlo. Por eso este builder concatena con join() en
        # una sola línea lógica por fragmento, sin sangría.
        parts = [
            f'<div class="pareto-card" style="background:{style["bg"]};border:2px solid {style["border"]};">',
            f'<div class="pareto-name">{html.escape(d["name"])}</div>',
            f'<div class="pareto-meta">AR-{d["brand_id"]} · {html.escape(d["category"])}</div>',
            f'<div class="pareto-row"><span class="pareto-row-label">Last Contact</span><span class="pareto-row-value">{_days_lbl}</span></div>',
            f'<div class="pareto-row"><span class="pareto-row-label">Ads</span><span class="pareto-row-value">{_fmt_roi_cell(d["ads_active"], d["ads_roi"])}</span></div>',
            f'<div class="pareto-row"><span class="pareto-row-label">MD</span><span class="pareto-row-value">{_fmt_roi_cell(d["md_active"], d["md_roi"])}</span></div>',
            f'<div class="pareto-row"><span class="pareto-row-label">MD PRO</span><span class="pareto-row-value">{_fmt_roi_cell(d["mdpro_active"], d["mdpro_roi"])}</span></div>',
            f'<div class="pareto-row"><span class="pareto-row-label">Perfect Store</span><span class="pareto-row-value">{_ps_lbl}</span></div>',
            f'<div class="pareto-row"><span class="pareto-row-label">Churn</span><span class="pareto-row-value">{html.escape(d["churn_label"])}</span></div>',
            f'<div class="pareto-row"><span class="pareto-row-label">CVR vs bench</span><span class="pareto-row-value">{_fmt_cvr_cell(d["cvr_brand"], d["cvr_bench"])}</span></div>',
            f'<div class="pareto-row"><span class="pareto-row-label">Traffic vs bench</span><span class="pareto-row-value">{_fmt_traffic_cell(d["traffic_brand"], d["traffic_bench"])}</span></div>',
            f'<div class="pareto-badge" style="background:{style["bg"]};color:{style["border"]};">{style["label"]}</div>',
            _acq_note,
            '</div>',
        ]
        return "".join(parts)

    for i in range(0, len(_sorted_data), 4):
        chunk = _sorted_data[i:i+4]
        cards_html = "".join(_build_pareto_card_html(d) for d in chunk)

        st.markdown(f'<div class="pareto-grid">{cards_html}</div>', unsafe_allow_html=True)

        # Botones reales de Streamlit para navegar al Brand Finder (debajo de cada fila,
        # ya que el onclick de arriba es solo decorativo — Streamlit no puede recibir
        # postMessage sin un listener adicional, así que usamos botones nativos).
        btn_cols = st.columns(4)
        for ci, d in enumerate(chunk):
            with btn_cols[ci]:
                if st.button("Ver ficha →", key=f"pareto_goto_{i+ci}_{d['brand_id']}", use_container_width=True):
                    st.session_state["_bf_goto_brand_id"] = d["brand_id"]
                    st.session_state["active_page"] = "Brand Finder"
                    st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)


def render_brand_profile(row, brand_id):
    name = clean(get_from_row(row, ["name", "brand name", "restaurant name"]))
    ltor = clean(get_from_row(row, ["ltor tier", "ltor"]))
    churn = get_churn_status(brand_id)  # Source of truth: Current Churn sheet (On if not listed)
    category_raw = clean(get_from_row(row, ["category"]))
    category, stickers = _split_category_and_stickers(category_raw)
    current = get_current_brand_metrics(brand_id)
    # Ranking: ordenado por GMV mayor a menor, usando el cruce que ya hace Current GMV/Detalle CABA.
    ranking = current["caba_rank"] if current else "-"

    ads_current_raw = get_current_ads_metrics(brand_id)
    md_current_raw = get_current_md_metrics(brand_id, pro=False)
    md_pro_current_raw = get_current_md_metrics(brand_id, pro=True)
    ads_current, md_current, md_pro_current = _merge_growth_manual_status(row, ads_current_raw, md_current_raw, md_pro_current_raw)

    # Pareto Tier badge belongs beside the brand name — replaces the old Mundialista badge slot.
    mundialista_name_badge_html = render_pareto_badge_html(brand_id)

    # booster/actions se siguen calculando aquí porque el Campaign Designer más abajo los reutiliza
    current_gmv_ars_badge = current["gmv_ars"] if current else 0
    current_aov_ars_badge = current["aov_ars"] if current else 0
    cr_for_badge = get_from_row(row, ["cr %", "conversion rate", "conversion"], 0)
    booster_for_badge = recommend_booster_for_brand(
        category,
        current_gmv_ars_badge,
        current_aov_ars_badge,
        cr_for_badge,
        get_from_row(row, ["pro users %", "pro %"], 0),
        ads_current,
        md_current,
    )
    actions_for_badge = build_360_actions(name, category, ads_current, md_current, md_pro_current, booster_for_badge)

    turbo_badge_html = (
        "<span class='hero-mundialista-badge' style='background:linear-gradient(145deg,#FF7124,#D95A10);margin-left:8px;'>⚡ STORE TURBO</span>"
        if get_turbo_info(brand_id) else ""
    )

    badges = []
    ads_roi = to_number(ads_current.get("roi"), 0)
    ads_bookings = to_number(ads_current.get("bookings_usd"), 0)
    ads_revenue = to_number(ads_current.get("revenue_usd"), 0)
    ads_consumption = (ads_revenue / ads_bookings) if ads_bookings else 0

    if not ads_current.get("active", False):
        badges.append("⚽ Ads Acquisition")
    elif ads_roi > 4.5:
        badges.append("⚡ Ads Upselling")
    if ads_current.get("active", False):
        if ads_roi > 2.0 and ads_consumption < 0.80:
            badges.append("⚡ Raise CPC")
        elif ads_roi < 2.0 and ads_consumption > 0.80:
            badges.append("🔻 Lower CPC")

    md_roi = to_number(md_current.get("roi"), 0)
    if not md_current.get("active", False):
        badges.append("⚽ MD Acquisition")
    elif md_roi > 3.2:
        badges.append("⚡ MD Upselling")

    md_pro_roi = to_number(md_pro_current.get("roi"), 0)
    if not md_pro_current.get("active", False):
        badges.append("⚽ MD PRO Acquisition")
    elif md_pro_roi > 3.2:
        badges.append("⚡ MD PRO Upselling")

    signals_html = "".join([f"<div class='signal-pill'>{b}</div>" for b in badges[:5]]) or "<div class='signal-pill'>✅ No critical commercial signal</div>"

    # Coinversion sticker
    coin_info = get_coinversion_info(brand_id, name)
    extra_stickers_html = ""
    if coin_info.get("found"):
        tier_color = "linear-gradient(145deg,#1B3F8B,#FFFFFF)" if "GOLDEN" in (coin_info.get("tier") or "") else "linear-gradient(145deg,#7ED321,#7ED321)"
        extra_stickers_html += f"<span class='hero-mundialista-badge' style='background:{tier_color};margin-left:8px'>{coin_info['sticker']}</span>"

    # General Information fields — computed here so they can be embedded in the hero-card
    sticker_html = "".join([f"<span class='category-chip'>{html.escape(str(s))}</span>" for s in stickers]) or "<span class='category-chip'>-</span>"
    multibrand_html = render_multibrand_html(row, brand_id)
    manager = clean(get_from_row(row, ["manager", "restaurant manager", "account manager"]))
    assistant = clean(get_from_row(row, ["assistant"]))

    # ── Saved link (computed before hero-card so it can be embedded) ────────
    finder_contact_number = fmt_contact_number(get_from_row(row, ["contact number", "phone", "contact"]))
    finder_saved_link = _get_saved_brand_link(brand_id)
    finder_search_url = _build_google_search_url(name, category, finder_contact_number)
    _saved_link_sticker = (
        f"<a href='{html.escape(finder_saved_link)}' target='_blank' rel='noopener noreferrer' "
        f"style='color:#7ED321;font-size:13px;font-weight:600;text-decoration:none;'>📍 Local</a>"
        if finder_saved_link else
        f"<a href='{html.escape(finder_search_url)}' target='_blank' rel='noopener noreferrer' "
        f"style='color:rgba(169,187,255,0.6);font-size:13px;text-decoration:none;'>🔎 Buscar</a>"
    )

    st.markdown(f"""
<div class="hero-card">
    <div class="hero-grid">
        <div>
            <div class="hero-name-row"><div class="hero-name">{name}</div>{mundialista_name_badge_html}{turbo_badge_html}{extra_stickers_html}</div>
            <div class="hero-id">AR-{normalize_brand_id(brand_id)}</div>
        </div>
        <div class="signal-stack">{signals_html}</div>
        <div class="sticker-grid">
            <div class="sticker"><div class="sticker-label">LTOR Tier</div><div class="sticker-value">{ltor}</div></div>
            <div class="sticker"><div class="sticker-label">Churn Status</div><div class="sticker-value">{churn}</div></div>
            <div class="sticker"><div class="sticker-label">Ranking</div><div class="sticker-value">{ranking}</div></div>
        </div>
    </div>
    <hr class="hero-divider"/>
    <div class="hero-info-grid">
        <div class="hero-info-item">
            <div class="hero-info-label">Category</div>
            <div class="hero-info-value">{html.escape(str(category))}</div>
        </div>
        <div class="hero-info-item">
            <div class="hero-info-label">Stickers</div>
            <div class="chip-line" style="margin-top:4px;">{sticker_html}</div>
        </div>
        <div class="hero-info-item">
            <div class="hero-info-label">Contact</div>
            <div class="hero-info-value" id="contact-num-{normalize_brand_id(brand_id)}">{html.escape(fmt_contact_number(get_from_row(row, ["contact number", "phone", "contact"])))}&nbsp;<button onclick="(function(){{var t=document.getElementById('contact-num-{normalize_brand_id(brand_id)}').innerText.trim();navigator.clipboard.writeText(t).then(function(){{var b=document.getElementById('copy-btn-{normalize_brand_id(brand_id)}');b.textContent='✅';setTimeout(function(){{b.textContent='📋';}},1800);}}).catch(function(){{var ta=document.createElement('textarea');ta.value=t;document.body.appendChild(ta);ta.select();try{{document.execCommand('copy');}}catch(e){{}}document.body.removeChild(ta);var b=document.getElementById('copy-btn-{normalize_brand_id(brand_id)}');b.textContent='✅';setTimeout(function(){{b.textContent='📋';}},1800);}});}})();" id="copy-btn-{normalize_brand_id(brand_id)}" title="Copiar número" style="background:rgba(59,72,131,0.28);border:1px solid rgba(59,72,131,0.55);border-radius:5px;padding:1px 6px;font-size:11px;cursor:pointer;color:#1B3F8B;vertical-align:middle;margin-left:2px;">📋</button></div>
        </div>
        <div class="hero-info-item">
            <div class="hero-info-label">Email</div>
            <div class="hero-info-value">{html.escape(clean(get_from_row(row, ["email"])))}</div>
        </div>
        <div class="hero-info-item">
            <div class="hero-info-label">Country</div>
            <div class="hero-info-value">{html.escape(clean(get_from_row(row, ["country"])))}</div>
        </div>
        <div class="hero-info-item">
            <div class="hero-info-label">Manager</div>
            <div class="hero-info-value">{html.escape(str(manager))}</div>
        </div>
        <div class="hero-info-item">
            <div class="hero-info-label">Assistant</div>
            <div class="hero-info-value">{html.escape(str(assistant))}</div>
        </div>
        <div class="hero-info-item">
            <div class="hero-info-label">Local</div>
            <div class="hero-info-value">{_saved_link_sticker}</div>
        </div>
    </div>
    {f'<div style="margin-top:14px;"></div>' if multibrand_html else ''}
</div>
""", unsafe_allow_html=True)

    # ── JS para botón copiar teléfono (st_components ejecuta JS real) ─────────
    _phone_val = fmt_contact_number(get_from_row(row, ["contact number", "phone", "contact"]))
    _bid_safe  = normalize_brand_id(brand_id)
    import json as _json_phone
    st_components.html(f"""
<script>
(function() {{
  var phone = {_json_phone.dumps(_phone_val)};
  // Busca el botón en el documento padre (el iframe de st_components puede acceder al padre en Streamlit)
  function findBtn() {{
    try {{
      var btn = window.parent.document.getElementById('copy-btn-{_bid_safe}');
      if (!btn) return;
      btn.onclick = function() {{
        navigator.clipboard.writeText(phone).then(function() {{
          btn.textContent = '✅';
          setTimeout(function() {{ btn.textContent = '📋'; }}, 1800);
        }}).catch(function() {{
          var ta = window.parent.document.createElement('textarea');
          ta.value = phone;
          window.parent.document.body.appendChild(ta);
          ta.select();
          try {{ window.parent.document.execCommand('copy'); }} catch(e) {{}}
          window.parent.document.body.removeChild(ta);
          btn.textContent = '✅';
          setTimeout(function() {{ btn.textContent = '📋'; }}, 1800);
        }});
      }};
    }} catch(e) {{}}
  }}
  // Intentar inmediatamente y luego con delay por si el DOM aún no está listo
  findBtn();
  setTimeout(findBtn, 400);
  setTimeout(findBtn, 900);
}})();
</script>
""", height=0, scrolling=False)

    # ── Editor inline — reemplaza la página separada "Brand Update" ──────────
    # Guardado inmediato: abre el Excel solo para esta marca, escribe, guarda y
    # cierra en el mismo click — pensado para completar marcas nuevas que llegan
    # al portafolio sin tener que crear nada manualmente en Growth OS.
    with st.expander("✏️ Editar ficha (guardado inmediato)", expanded=False):
        _edit_category_current = clean(get_from_row(row, ["category"]), "")
        _edit_manager_current  = clean(get_from_row(row, ["manager", "restaurant manager", "account manager"]), "")
        _edit_assistant_current = clean(get_from_row(row, ["assistant"]), "")
        _edit_email_current    = clean(get_from_row(row, ["email", "mail"]), "")
        _edit_phone_current    = fmt_contact_number(get_from_row(row, ["contact number", "phone", "contact"]))
        _edit_commission_current = to_number(get_from_row(row, ["comm. rate", "commission rate", "commission"], 0), 0)
        _edit_pro_current      = to_number(get_from_row(row, ["pro users %", "pro %"], 0), 0)
        # Normalizar a 0-100 para el number_input — el dato puede venir como 0.27 o 27
        _edit_commission_pct = round(_edit_commission_current * 100, 1) if _edit_commission_current <= 1 else round(_edit_commission_current, 1)
        _edit_pro_pct         = round(_edit_pro_current * 100, 1) if _edit_pro_current <= 1 else round(_edit_pro_current, 1)

        with st.form(f"inline_brand_edit_{brand_id}"):
            ie1, ie2 = st.columns(2)
            with ie1:
                _new_category   = st.text_input("Categoría / Stickers", value=_edit_category_current, help="Formato: 'Categoría Principal | Sticker1 | Sticker2'")
                _new_manager    = st.text_input("Manager", value=_edit_manager_current)
                _new_assistant  = st.text_input("Assistant", value=_edit_assistant_current)
            with ie2:
                _new_email      = st.text_input("Email", value=_edit_email_current)
                _new_phone      = st.text_input("Teléfono / Contact Number", value=_edit_phone_current)
                _ie_c1, _ie_c2  = st.columns(2)
                with _ie_c1:
                    _new_commission_pct = st.number_input("Comisión %", value=float(_edit_commission_pct), min_value=0.0, max_value=100.0, step=0.5)
                with _ie_c2:
                    _new_pro_pct = st.number_input("Usuarios PRO %", value=float(_edit_pro_pct), min_value=0.0, max_value=100.0, step=0.5)

            _inline_submitted = st.form_submit_button("💾 Guardar ahora")

        if _inline_submitted:
            _inline_updates = {
                "category":        _new_category.strip(),
                "manager":         _new_manager.strip(),
                "assistant":       _new_assistant.strip(),
                "email":           _new_email.strip(),
                "contact_number":  _new_phone.strip(),
                "commission_rate": round(_new_commission_pct / 100, 4),
                "pro_users_pct":   round(_new_pro_pct / 100, 4),
            }
            _old_row_for_changelog = row
            _ok_inline, _msg_inline, _updated_inline, _locked_inline, _missing_inline, _backup_inline = update_brand_in_excel(brand_id, _inline_updates)
            if _ok_inline:
                try:
                    save_brand_changelog(brand_id, name, _inline_updates, _old_row_for_changelog)
                except Exception:
                    pass
                st.success("✅ Ficha actualizada — guardado inmediato.")
                if _locked_inline:
                    st.caption("Algunos campos protegidos por fórmula no se actualizaron: " + ", ".join(_locked_inline))
                st.rerun()
            else:
                st.error(_msg_inline)

    # ── Sticker "New Acquisition" — visible si esta marca cerró un deal reciente ──
    _acq_tracker_df = None
    try:
        if os.path.exists(ACQUISITION_TRACKER_FILE):
            _acq_tracker_df = pd.read_csv(ACQUISITION_TRACKER_FILE, dtype=str).fillna("")
    except Exception:
        _acq_tracker_df = None

    if _acq_tracker_df is not None and not _acq_tracker_df.empty:
        _acq_brand_rows = _acq_tracker_df[
            (_acq_tracker_df["brand_id"].apply(normalize_brand_id) == normalize_brand_id(brand_id))
            & (_acq_tracker_df["movement"] == "Acquisition")
            & (_acq_tracker_df["pipeline_stage"] == "Closed")
        ]
        if not _acq_brand_rows.empty:
            _acq_latest = _acq_brand_rows.sort_values("datetime", ascending=False).iloc[0]
            _acq_type = clean(_acq_latest.get("type", ""), "—")
            _acq_ads_ars = to_number(_acq_latest.get("ads_booking_ars"), 0)
            _acq_md_disc = clean(_acq_latest.get("md_discount", ""), "")
            _acq_date = clean(_acq_latest.get("date", ""), "")
            _acq_detail = (
                f"{fmt_ars(_acq_ads_ars)}" if _acq_type == "Ads" and _acq_ads_ars
                else _acq_md_disc if _acq_md_disc
                else "—"
            )
            st.markdown(
                f'<div style="display:inline-flex;align-items:center;gap:8px;background:rgba(126,211,33,0.10);'
                f'border:1px solid #7ED321;border-radius:20px;padding:6px 14px;margin-bottom:14px;font-size:12px;font-weight:700;color:#5A9E00;">'
                f'🆕 NEW ACQUISITION · {html.escape(_acq_type)} · {html.escape(_acq_detail)} · {html.escape(_acq_date)}'
                f'</div>',
                unsafe_allow_html=True,
            )

    # ── Multibrand: chips interactivos → navegan al Brand Finder de esa marca ──
    _mb_matches = get_multibrand_matches(row, brand_id)
    if _mb_matches:
        _mb_total      = len(_mb_matches)
        _mb_high_count = sum(1 for m in _mb_matches if m["confidence"] == "High")
        _mb_title      = "🏢 Multibrand detected" if _mb_high_count else "🏢 Possible multibrand"
        _mb_summary    = f"{_mb_total} linked account{'s' if _mb_total != 1 else ''} · {_mb_high_count} high confidence"

        st.markdown(
            f"<div class='multibrand-box'>"
            f"<div class='info-mini-label'>{_mb_title}</div>",
            unsafe_allow_html=True,
        )
        _mb_cols = st.columns(min(_mb_total, 3))
        for _mb_i, _mb_match in enumerate(_mb_matches):
            _mb_conf_icon = "✅" if _mb_match["confidence"] == "High" else "⚠️"
            _mb_reason    = "/".join(_mb_match["reasons"])
            _mb_label     = (
                f"{_mb_conf_icon} AR-{_mb_match['id']} · "
                f"{clean(_mb_match['name'], '-')} · {_mb_reason}"
            )
            with _mb_cols[_mb_i % len(_mb_cols)]:
                if st.button(_mb_label, key=f"mb_goto_{brand_id}_{_mb_match['id']}"):
                    st.session_state["_bf_goto_brand_id"] = str(_mb_match["id"])
                    st.session_state["active_page"] = "Brand Finder"
                    st.rerun()
        st.markdown(
            f"<div class='multibrand-summary'>{_mb_summary}</div></div>",
            unsafe_allow_html=True,
        )

    # ── Historia del aliado (changelog) ──────────────────────────────────────
    try:
        if os.path.exists(CHANGELOG_FILE):
            _cl = pd.read_csv(CHANGELOG_FILE, encoding="utf-8-sig")
            _cl_brand = _cl[_cl["brand_id"].astype(str).apply(normalize_brand_id) == normalize_brand_id(brand_id)].copy()
            if not _cl_brand.empty:
                _cl_brand = _cl_brand.sort_values("datetime", ascending=False).head(5)
                _hist_rows = "".join([
                    f"<div style='display:flex;gap:12px;padding:7px 0;border-bottom:1px solid rgba(0,0,0,0.06);align-items:baseline;'>"
                    f"<span style='font-size:11px;color:rgba(107,114,128,0.60);white-space:nowrap;min-width:110px;'>{html.escape(str(r.get('datetime',''))[:16])}</span>"
                    f"<span style='font-size:12px;color:#6B7280;min-width:120px;'>{html.escape(str(r.get('field','')))}</span>"
                    f"<span style='font-size:12px;color:rgba(107,114,128,0.60);text-decoration:line-through;margin-right:6px;'>{html.escape(str(r.get('old_value','')))}</span>"
                    f"<span style='font-size:12px;font-weight:700;color:#1A1A2E;'>{html.escape(str(r.get('new_value','')))}</span>"
                    f"</div>"
                    for _, r in _cl_brand.iterrows()
                ])
                st.markdown(
                    f"<div class='campaign-mini-card lever-ops' style='padding:16px 20px;margin-bottom:14px;'>"
                    f"<div class='card-label'>🕓 Historia del aliado · últimos cambios registrados</div>"
                    f"<div style='margin-top:10px;'>{_hist_rows}</div>"
                    f"</div>",
                    unsafe_allow_html=True
                )
    except Exception:
        pass  # Si el changelog no existe o falla, no romper el perfil

    # ── Last contact + opportunity score + cadence badge ─────────────────────
    _bf_prod_map  = get_productivity_last_contact_map(EXCEL_FILE)
    _bf_meta_map  = get_last_comment_meta_map(limit=1)
    _bf_brand_key = name.strip().lower()
    _bf_meta      = _bf_meta_map.get(normalize_brand_id(brand_id), {})

    # Resolve last contact: Productivity first, comments CSV fallback
    _bf_last_dt = _bf_prod_map.get(_bf_brand_key)
    if _bf_last_dt is None:
        _bf_last_dt = _bf_meta.get("last_dt")

    _bf_days = _days_since_timestamp(_bf_last_dt)
    if _bf_days is None:
        _bf_days_label = "Sin contacto registrado"
        _bf_days_color = "#FF4D2E"
    elif _bf_days == 0:
        _bf_days_label = "Contactado hoy"
        _bf_days_color = "#7ED321"
    elif _bf_days <= 7:
        _bf_days_label = f"Hace {_bf_days}d · en ciclo activo"
        _bf_days_color = "#7ED321"
    elif _bf_days <= 14:
        _bf_days_label = f"Hace {_bf_days}d · en cadencia"
        _bf_days_color = "#7ED321"
    elif _bf_days <= 21:
        _bf_days_label = f"Hace {_bf_days}d · zona de alerta"
        _bf_days_color = "#FF7124"
    else:
        _bf_days_label = f"Hace {_bf_days}d · marca fría ❄️"
        _bf_days_color = "#FF4D2E"

    # Temperature badge
    if _bf_days is None or _bf_days > 21:
        _bf_temp_badge = "<span style='background:rgba(229,51,42,0.10);color:#FF4D2E;font-size:11px;font-weight:700;padding:2px 10px;border-radius:20px;'>❄️ Fría</span>"
    elif _bf_days > 14:
        _bf_temp_badge = "<span style='background:rgba(255,113,36,0.10);color:#D95A10;font-size:11px;font-weight:700;padding:2px 10px;border-radius:20px;'>🟠 Alerta</span>"
    elif _bf_days > 7:
        _bf_temp_badge = "<span style='background:rgba(255,113,36,0.08);color:#D95A10;font-size:11px;font-weight:700;padding:2px 10px;border-radius:20px;'>🟡 Cadencia</span>"
    else:
        _bf_temp_badge = "<span style='background:rgba(111,242,75,0.08);color:#7ED321;font-size:11px;font-weight:700;padding:2px 10px;border-radius:20px;'>🟢 Activa</span>"

    # Opportunity score from scored data
    _bf_scored_df = _prepare_growth_scored_data()
    _bf_opp_score = None
    _bf_opp_rank  = None
    if not _bf_scored_df.empty:
        _bf_id_col = get_id_column_name(_bf_scored_df)
        if _bf_id_col:
            _bf_match = _bf_scored_df[_bf_scored_df[_bf_id_col].apply(normalize_brand_id) == normalize_brand_id(brand_id)]
            if not _bf_match.empty:
                _bf_opp_score = round(float(_bf_match.iloc[0].get("_opportunity_score", 0)), 1)
                # Rank = position in descending order of opportunity score
                _bf_ranked = _bf_scored_df.copy()
                _bf_ranked = _bf_ranked.dropna(subset=["_opportunity_score"])
                _bf_ranked = _bf_ranked.sort_values("_opportunity_score", ascending=False).reset_index(drop=True)
                _bf_rank_match = _bf_ranked[_bf_ranked[_bf_id_col].apply(normalize_brand_id) == normalize_brand_id(brand_id)]
                if not _bf_rank_match.empty:
                    _bf_opp_rank = int(_bf_rank_match.index[0]) + 1

    _bf_score_html = (
        f"<span style='font-size:20px;font-weight:900;color:#1B3F8B;'>{_bf_opp_score}</span>"
        f"<span style='font-size:11px;color:#6B7280;margin-left:4px;'>opp. score</span>"
        f"<span style='font-size:11px;color:#6B7280;margin-left:6px;'>· #{_bf_opp_rank} en portafolio</span>"
        if _bf_opp_score is not None else
        "<span style='font-size:12px;color:#aaa;'>Score no disponible</span>"
    )

    _bf_last_note = _bf_meta.get("notes", "-")

    # ── Última nota: la entrada más reciente (no el historial completo) ───────
    _excel_comments_bf = clean(get_from_row(row, ["comments", "comment"], ""))
    _saved_comments_df_bf = _load_comments_df()
    _last_saved_comment = ""
    if not _saved_comments_df_bf.empty and "brand_id" in _saved_comments_df_bf.columns:
        _brand_rows_bf = _saved_comments_df_bf[
            _saved_comments_df_bf["brand_id"] == normalize_brand_id(brand_id)
        ].copy()
        if not _brand_rows_bf.empty:
            _brand_rows_bf = _brand_rows_bf.sort_values("_dt", ascending=False, na_position="last")
            _most_recent = _brand_rows_bf.iloc[0]
            _last_saved_comment = clean(_most_recent.get("comment", ""), "")

    # ── Priority logic ────────────────────────────────────────────────────────
    # 1. [Auto] transcript summary   → always wins (most informative)
    # 2. Productivity sheet levers   → if no transcript comment this month
    # 3. Regular CSV comment         → next fallback
    # 4. Excel comments / meta notes → last resort
    #
    # _nota_source tracks where we got the note from, for the badge label.
    _nota_source = "meta"
    _prod_levers = None  # populated if we fall into branch 2

    _is_transcript_comment = (
        _last_saved_comment.strip().startswith("[Auto]")
        if _last_saved_comment else False
    )

    if _is_transcript_comment:
        # Branch 1: transcript-based comment from Brand Update — highest fidelity
        _display_last_note = _last_saved_comment.strip()
        _nota_source = "transcript"
    else:
        # Branch 2: try Productivity sheet for current month
        _brand_name_for_prod = clean(get_from_row(row, ["brand", "name", "brand name", "nombre"], ""))
        _prod_levers = get_productivity_levers_for_brand(EXCEL_FILE, _brand_name_for_prod)
        if _prod_levers and _prod_levers.get("levers"):
            _display_last_note = "[Productivity] " + _prod_levers["nota_generada"]
            _nota_source = "productivity"
        elif _last_saved_comment.strip():
            # Branch 3: regular (non-transcript) CSV comment
            _display_last_note = _last_saved_comment.strip()
            _nota_source = "csv"
        elif _excel_comments_bf not in ["", "-"]:
            # Branch 4a: Excel comment column
            _display_last_note = _excel_comments_bf.strip()
            _nota_source = "excel"
        else:
            # Branch 4b: meta notes
            _display_last_note = _bf_last_note
            _nota_source = "meta"

    # ── Retomar desde Productivity (palancas reales) ──────────────────────────
    def _build_retomar_from_levers(levers_data):
        """
        Generates a call re-entry suggestion based on real levers logged
        in the Productivity sheet for this brand this month.
        Much more specific than text-analysis because we have exact lever names.
        """
        levers  = levers_data.get("levers", [])
        churn   = levers_data.get("churn", False)
        on_hold = levers_data.get("on_hold", False)
        ads_ok  = "Ads" in levers
        md_ok   = "Markdown" in levers
        accepted_md = levers_data.get("accepted_md", False)
        ads_tipo    = levers_data.get("ads_tipo", "")
        ajustes     = levers_data.get("ajustes", [])
        fase        = levers_data.get("fase", "")
        rc          = levers_data.get("row_count", 0)
        latest      = levers_data.get("latest_str", "")

        # Determine priority lever for the opener
        if churn:
            opener = "⚠️ Retomá priorizando retención — Churn registrado este mes. Abrí con el valor generado vs. costo de baja y una propuesta concreta."
            color  = PALETTE["burning_orange"]
        elif on_hold:
            opener = "🔴 Aliado en On Hold — necesitás destrabar el bloqueo antes de cualquier pitch. Abrí preguntando qué cambió y qué necesita para reactivar."
            color  = PALETTE["burning_orange"]
        elif ads_ok and md_ok:
            md_note = " (MD aceptado ✓)" if accepted_md else " (MD ofrecido, sin cierre)"
            ads_note = f" — tipo: {ads_tipo}" if ads_tipo else ""
            opener = f"Retomá combinando ADS{ads_note} + Markdown{md_note}. Presentá el paquete integrado: visibilidad + conversión en la misma propuesta."
            color  = PALETTE["blue_estate"]
        elif ads_ok:
            ads_note = f" ({ads_tipo})" if ads_tipo else ""
            opener = f"Retomá desde ADS{ads_note} — fue la palanca del mes. Abrí con el benchmark de la categoría y un budget concreto propuesto."
            color  = PALETTE["blue_estate"]
        elif md_ok:
            if accepted_md:
                opener = "MD aceptado este mes ✓ — retomá verificando la activación y proponiendo la siguiente promo con fecha. El momentum está."
                color  = PALETTE["laser_green"]
            else:
                opener = "MD ofrecido pero sin cierre — retomá con el descuento pendiente, la fecha de activación y una comparación de GMV con/sin promo."
                color  = PALETTE["laser_green"]
        elif ajustes:
            opener = f"Retomá desde catálogo — se trabajó: {', '.join(ajustes[:3])}. Mostrá el impacto en conversión de los cambios hechos."
            color  = PALETTE["cinnamon_ice"]
        elif "Conectividad" in levers:
            opener = "Retomá desde conectividad — verificá que el issue esté resuelto antes de cualquier pitch comercial. Luego aprovechá para anclar una propuesta."
            color  = PALETTE["blue_glow"]
        elif levers:
            opener = f"Retomá desde {levers[0]} — la palanca trabajada este mes. Abrí con el estado actual y el próximo paso concreto."
            color  = PALETTE["cinnamon_ice"]
        else:
            opener = "Sin palancas específicas registradas este mes — abrí con contexto de rendimiento general."
            color  = PALETTE["cinnamon_ice"]

        # Sidebar: fase + contacts
        context_parts = []
        if fase and fase.lower() not in ["nan", ""]:
            context_parts.append(f"Fase: {fase}")
        if rc:
            context_parts.append(f"{rc} contacto{'s' if rc != 1 else ''} este mes")
        if latest:
            context_parts.append(f"último: {latest}")

        context_html = ""
        if context_parts:
            context_html = (
                f'<div style="margin-top:7px;font-size:10px;color:rgba(219,187,167,.6);">'
                + " · ".join(context_parts)
                + "</div>"
            )

        return (
            f'<div style="font-size:13px;font-weight:600;color:{color};line-height:1.4;">{opener}</div>'
            + context_html
        )

    # ── Retomar generado por Claude (lectura directa, sin re-adivinar) ────────
    def _extract_claude_retomar(note_text):
        """
        Busca una línea 'Retomar: ...' dentro de una nota [Auto] con análisis
        completo (generada en el chat de Claude, no por el analizador local).
        Devuelve solo el enfoque de la próxima llamada, sin pasos a seguir.
        Si no existe (nota vieja del analizador local, sin este campo), devuelve
        None y se cae al scoring por keywords de siempre (_build_retomar_html).
        """
        if not note_text:
            return None
        matches = re.findall(r"(?im)^\s*retomar:\s*(.+?)\s*$", note_text)
        return matches[-1].strip() if matches else None

    # ── Retomar desde texto (fallback cuando no hay datos de Productivity) ────
    def _build_retomar_html(note_text):
        """Pure-Python analysis of last note to suggest call re-entry point."""
        if not note_text or note_text.strip() in ["-", ""]:
            return '<span style="font-size:11px;color:#aaa;">Sin nota previa para analizar</span>'

        low = note_text.lower()

        lever_scores = {
            "ADS":               sum(1 for k in ["ads", "publicidad", "banner", "campaña", "sponsored", "visibilidad paga", "investment"] if k in low),
            "Markdown":          sum(1 for k in ["descuento", "promo", "markdown", "porcentaje", "%", "oferta"] if k in low),
            "Top Restaurant":    sum(1 for k in ["top restaurant", "destacado", "posicionamiento", "ranking", "orgánica"] if k in low),
            "Menú / Assortment": sum(1 for k in ["menú", "menu", "catálogo", "fotos", "productos", "carta", "assortment"] if k in low),
            "Churn":             sum(1 for k in ["cancelar", "baja", "churn", "retiro", "no quiero seguir", "cerrar cuenta"] if k in low),
        }
        top_lever = max(lever_scores, key=lever_scores.get)
        top_score = lever_scores[top_lever]

        pending_signals = []
        if any(k in low for k in ["lo pienso", "lo consulto", "voy a ver", "te llamo", "la próxima"]):
            pending_signals.append("el aliado quedó en pensar")
        if any(k in low for k in ["enviar", "mandar", "propuesta", "plantilla", "mail"]):
            pending_signals.append("pendiente envío de propuesta")
        if any(k in low for k in ["negociando", "pendiente", "negotiation", "esperando"]):
            pending_signals.append("hay una negociación abierta")
        if any(k in low for k in ["rechazó", "no le interesa", "rejected", "no quiere"]):
            pending_signals.append("la última interacción fue un rechazo")

        if top_score == 0:
            opener = "Arrancá con una apertura de contexto general — no hay palanca clara en la nota anterior."
            color = PALETTE["cinnamon_ice"]
        elif top_lever == "Churn":
            opener = "⚠️ Retomá priorizando retención — hay señales de riesgo de baja. Abrí con datos de valor y propuesta concreta."
            color = PALETTE["burning_orange"]
        elif top_lever == "ADS":
            opener = "Retomá desde ADS — fue la palanca dominante. Abrí con el ROI de la categoría y un budget concreto."
            color = PALETTE["blue_estate"]
        elif top_lever == "Markdown":
            opener = "Retomá desde la promo — fue lo que se estaba trabajando. Abrí con el descuento pendiente y la fecha de activación."
            color = PALETTE["laser_green"]
        elif top_lever == "Top Restaurant":
            opener = "Retomá desde posicionamiento — hablaron de visibilidad. Abrí con el ranking actual y qué cambiaría activando."
            color = PALETTE["blue_glow"]
        else:
            opener = "Retomá desde menú / catálogo — hablaron de productos. Abrí con fotos o la lista de top products de la categoría."
            color = PALETTE["cinnamon_ice"]

        pending_html = ""
        if pending_signals:
            items = "".join(f'<li style="margin-bottom:3px;">{s.capitalize()}</li>' for s in pending_signals)
            pending_html = f'<ul style="margin:6px 0 0 0;padding-left:16px;font-size:11px;color:#6B7280;">{items}</ul>'

        return (
            f'<div style="font-size:13px;font-weight:600;color:{color};line-height:1.4;">{opener}</div>'
            + pending_html
        )

    # ── Choose retomar renderer based on source ───────────────────────────────
    _claude_retomar_text = (
        _extract_claude_retomar(_display_last_note) if _nota_source == "transcript" else None
    )
    if _claude_retomar_text:
        # Fuente: análisis completo de Claude → mostrar el enfoque tal cual, sin re-adivinar
        _retomar_html = (
            f'<div style="font-size:13px;font-weight:600;color:{PALETTE["blue_glow"]};line-height:1.4;">'
            f'{html.escape(_claude_retomar_text)}</div>'
        )
    elif _nota_source == "productivity" and _prod_levers:
        _retomar_html = _build_retomar_from_levers(_prod_levers)
    else:
        _retomar_html = _build_retomar_html(_display_last_note)

    # ── Render: última nota display ───────────────────────────────────────────
    def _extract_claude_resumen(note_text):
        """
        Busca el párrafo 'Resumen: ...' dentro de una nota [Auto] con análisis
        completo (generada en el chat de Claude). Devuelve solo ese párrafo
        condensado para mostrar en la carta Última Nota — no la nota completa.
        Si no existe (nota vieja sin este campo), retorna None y se muestra
        el texto completo como fallback.
        """
        if not note_text:
            return None
        m = re.search(r"(?i)resumen:\s*(.+?)(?:\n\s*\n|\Z)", note_text, re.DOTALL)
        return m.group(1).strip() if m else None

    _note_display_clean = _display_last_note
    # Strip well-known prefixes for cleaner display
    for _pfx in ["[Auto] ·", "[Auto]", "[Productivity]"]:
        if _note_display_clean.startswith(_pfx):
            _note_display_clean = _note_display_clean[len(_pfx):].strip()

    # Si la nota viene de Claude (transcript) y tiene el párrafo Resumen:,
    # usarlo en vez del texto completo — la carta muestra solo el resumen.
    if _nota_source == "transcript":
        _claude_resumen_text = _extract_claude_resumen(_display_last_note)
        if _claude_resumen_text:
            _note_display_clean = _claude_resumen_text

    # Source badge (tiny label below the note)
    _source_badge_map = {
        "transcript":  ("#7ED321", "📋 Transcripción"),
        "productivity": (PALETTE["blue_glow"], "📊 Productivity mes"),
        "csv":         (PALETTE["cinnamon_ice"], "💬 Último contacto"),
        "excel":       ("#aaa", "📄 Excel"),
        "meta":        ("#aaa", "📝 Meta"),
    }
    _src_color, _src_label = _source_badge_map.get(_nota_source, ("#aaa", ""))
    _source_badge_html = (
        f'<div style="margin-top:5px;font-size:9px;font-weight:700;'
        f'color:{_src_color};text-transform:uppercase;letter-spacing:.5px;">'
        f'{_src_label}</div>'
    )

    _bf_last_note_html = (
        f'<span style="font-size:11px;color:#6B7280;font-style:italic;white-space:pre-wrap;line-height:1.5;">'
        f'{html.escape(_note_display_clean)}</span>'
        f'{_source_badge_html}'
        if _display_last_note and _display_last_note.strip() not in ["-", ""] else
        '<span style="font-size:11px;color:#aaa;">Sin nota reciente</span>'
    )

    st.markdown(f"""
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:16px;">
      <div style="background:rgba(255,255,255,0.90);border-radius:12px;padding:14px 16px;">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#aaa;margin-bottom:6px;">Último contacto</div>
        <div style="font-size:13px;font-weight:800;color:{_bf_days_color};">{_bf_days_label}</div>
        <div style="margin-top:6px;">{_bf_temp_badge}</div>
      </div>
      <div style="background:rgba(255,255,255,0.90);border-radius:12px;padding:14px 16px;max-height:160px;overflow-y:auto;">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#aaa;margin-bottom:6px;">Última nota</div>
        <div style="margin-top:2px;">{_bf_last_note_html}</div>
      </div>
      <div style="background:rgba(59,72,131,0.18);border:1px solid rgba(27,63,139,0.12);border-radius:12px;padding:14px 16px;max-height:160px;overflow-y:auto;">
        <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:{PALETTE['blue_glow']};margin-bottom:8px;">🎯 Retomar llamada</div>
        <div style="margin-top:2px;">{_retomar_html}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    growth_gmv_ars = to_number(get_from_row(row, ["last gmv ars", "gmv ars", "last gmv local", "last gmv"], 0), 0)
    growth_gmv_usd = to_number(get_from_row(row, ["last gmv usd", "gmv usd"], 0), 0) or (growth_gmv_ars / ARS_PER_USD if growth_gmv_ars else 0)
    growth_aov_ars = to_number(get_from_row(row, ["last aov ars", "aov ars", "last aov local"], 0), 0)
    growth_aov_usd = to_number(get_from_row(row, ["last aov usd", "aov usd"], 0), 0) or (growth_aov_ars / ARS_PER_USD if growth_aov_ars else 0)

    current_gmv_ars = current["gmv_ars"] if current else 0
    current_gmv_usd = current["gmv_usd"] if current else 0
    current_aov_ars = current["aov_ars"] if current else 0
    current_aov_usd = current["aov_usd"] if current else 0
    gmv_progress_ars = safe_ratio(current_gmv_ars, growth_gmv_ars)
    aov_change_ars = safe_ratio(current_aov_ars - growth_aov_ars, growth_aov_ars)

    pro_users_display = fmt_percent0(get_from_row(row, ["pro users %", "pro users", "pro user %", "pro %", "prime users %"]))
    conversion_display = fmt_percent0(get_from_row(row, ["cr %", "conversion rate", "conversion"]))
    commission_display = fmt_percent0(get_from_row(row, ["comm. rate", "commission rate", "commission"]))
    pro_users_raw = _normalize_rate_value(get_from_row(row, ["pro users %", "pro users", "pro user %", "pro %", "prime users %"], 0))
    conversion_raw = _normalize_rate_value(get_from_row(row, ["cr %", "conversion rate", "conversion"], 0))
    commission_raw = _normalize_rate_value(get_from_row(row, ["comm. rate", "commission rate", "commission"], 0))
    def _dot_line_chart_card(label, val_current, val_may, val_abril, fmt_fn, sub_current, orders_inline=None):
        """
        Gráfico de línea con puntos mostrando hasta 3 meses:
          Abril (Growth OS) → Mayo (MAY GMV) → Actual (Current GMV)
        - Si los 3 tienen valor > 0, se muestran los 3.
        - Si falta Abril pero hay Mayo y Actual, se muestran 2.
        - Si falta Mayo pero hay Abril y Actual, se muestran los 3 igualmente
          con Mayo en 0 (para que el gráfico refleje la caída/suba).
        - Si solo hay Actual, se muestra solo el punto.
        Cada punto muestra el % de cambio respecto al punto anterior
        (excepto el primero, que no tiene comparación previa).
        """
        points_raw = []
        has_abril = val_abril and val_abril > 0
        has_may   = val_may   and val_may   > 0

        if has_abril:
            points_raw.append(("Abr", val_abril))
        if has_may:
            points_raw.append(("May", val_may))
        elif has_abril:
            # Mayo no tiene dato pero sí hay Abril: incluir Mayo con 0
            # para que el eje X muestre los 3 meses correctamente
            points_raw.append(("May", 0))
        points_raw.append(("Actual", val_current or 0))

        # Si solo queda 1 punto, agregar un punto ficticio para dibujar la línea
        if len(points_raw) < 2:
            points_raw = [("May", val_may or 0), ("Actual", val_current or 0)]

        n = len(points_raw)
        vals = [p[1] for p in points_raw]
        v_max = max(vals + [1])
        v_min = min(vals + [0])
        v_range = (v_max - v_min) or 1

        W, H = 150, 72
        ML, MR, MT, MB = 10, 10, 20, 18
        PW = W - ML - MR
        PH = H - MT - MB

        xs = [ML + (i / (n - 1)) * PW if n > 1 else ML + PW / 2 for i in range(n)]
        ys = [MT + PH - ((v - v_min) / v_range) * PH for v in vals]

        # Línea conectando los puntos
        line_pts = " ".join(f"{x:.1f},{y:.1f}" for x, y in zip(xs, ys))

        svg_parts = [
            f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg">',
            # Línea base (eje X)
            f'<line x1="{ML}" y1="{MT+PH}" x2="{W-MR}" y2="{MT+PH}" stroke="rgba(0,0,0,0.07)" stroke-width="1"/>',
            # Línea de tendencia
            f'<polyline points="{line_pts}" fill="none" stroke="#FF7124" stroke-width="1.8"/>',
        ]

        def _fmt_compact(v):
            """Formats value as compact: 1.2M, 560k, etc."""
            if not v or v == 0:
                return ""
            if v >= 1_000_000:
                n = v / 1_000_000
                return f"{n:.1f}M".replace(".0M", "M")
            if v >= 1_000:
                n = v / 1_000
                # If round number (e.g. 560000 -> 560k), no decimal
                if n == int(n):
                    return f"{int(n)}k"
                return f"{n:.0f}k"
            return str(int(v))

        for i, ((lbl, v), x, y) in enumerate(zip(points_raw, xs, ys)):
            # Valor compacto del GMV en el punto
            val_txt = _fmt_compact(v) if v and v > 0 else ""

            svg_parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="#FF7124"/>')
            if val_txt:
                svg_parts.append(
                    f'<text x="{x:.1f}" y="{y-7:.1f}" text-anchor="middle" '
                    f'font-size="7" font-weight="800" fill="#FF7124">{val_txt}</text>'
                )
            # Etiqueta del mes en el eje X
            svg_parts.append(
                f'<text x="{x:.1f}" y="{MT+PH+11}" text-anchor="middle" '
                f'font-size="6.5" fill="rgba(107,114,128,0.70)" font-weight="700">{lbl}</text>'
            )

        svg_parts.append('</svg>')
        svg = "".join(svg_parts)

        # Cambio total mostrado en el header (último vs anterior inmediato)
        if len(points_raw) >= 2 and points_raw[-2][1] != 0:
            change_pct = points_raw[-1][1] / points_raw[-2][1] - 1
        else:
            change_pct = None
        _change_color = "#7ED321" if (change_pct or 0) >= 0 else "#FF4D2E"
        _change_sign  = "+" if (change_pct or 0) >= 0 else ""
        _change_text  = f"{_change_sign}{fmt_percent0(change_pct)}" if change_pct is not None else "-"

        _orders_inline_html = (
            f'<span style="color:rgba(107,114,128,0.65);font-weight:600;margin-left:6px;">📦 {int(orders_inline):,}</span>'.replace(",", ".")
            if orders_inline is not None else ""
        )

        return (
            '<div class="stack-card" style="position:relative;overflow:hidden;min-height:140px;padding:22px 24px 18px;">'
            f'<div class="stack-label" style="margin-bottom:6px;">{label}</div>'
            f'<div style="font-size:30px;font-weight:900;color:#1A1A2E;letter-spacing:-0.02em;line-height:1.1;">{fmt_fn(val_current)}</div>'
            f'<div style="font-size:12px;font-weight:600;color:#6B7280;margin-top:5px;">{sub_current}{_orders_inline_html}</div>'
            f'<div style="margin-top:8px;font-size:11px;font-weight:800;color:{_change_color};">{_change_text} vs mes anterior</div>'
            f'<div style="position:absolute;bottom:12px;right:12px;opacity:0.95">{svg}</div>'
            '</div>'
        )

    # ── Datos de mayo (MAY GMV) y abril (Growth OS) para el gráfico de 3 meses ──
    may_metrics = get_may_brand_metrics(brand_id, brand_name=name)
    may_gmv_ars = may_metrics["gmv_ars"] if may_metrics else 0
    may_aov_ars = may_metrics["aov_ars"] if may_metrics else 0

    # Abril viene de Growth OS (Last GMV/AOV) — puede no existir
    abril_gmv_ars = growth_gmv_ars
    abril_aov_ars = growth_aov_ars

    gmv_col, aov_col = st.columns(2)
    with gmv_col:
        _orders_from_caba = get_orders_from_detalle_caba(brand_id, brand_name=name)
        _gmv_card = _dot_line_chart_card(
            "📈 GMV · Este mes vs Anterior",
            current_gmv_ars, may_gmv_ars, abril_gmv_ars, fmt_ars,
            f"{fmt_usd(current_gmv_usd)} · {fmt_cop(current_gmv_usd * COP_PER_USD)}",
            orders_inline=_orders_from_caba,
        )
        st.markdown(_gmv_card, unsafe_allow_html=True)
    with aov_col:
        _aov_card = _dot_line_chart_card(
            "🛒 AOV · Este mes vs Anterior",
            current_aov_ars, may_aov_ars, abril_aov_ars, fmt_ars,
            f"{fmt_usd(current_aov_usd)} · {fmt_cop(current_aov_usd * COP_PER_USD)}",
        )
        st.markdown(_aov_card, unsafe_allow_html=True)

    # ── Calculadora Consultiva · 4 métricas financieras automáticas ──────────
    # Estas tarjetas cruzan los datos ya disponibles del aliado para generar
    # los 4 argumentos financieros listos para usar en la reunión con el dueño.

    _comm_rate   = commission_raw if commission_raw and commission_raw > 0 else 0.27
    _aov         = current_aov_ars if current_aov_ars and current_aov_ars > 0 else growth_aov_ars
    _orders      = to_number(current["orders"], 0) if current else 0
    _gmv         = current_gmv_ars if current_gmv_ars and current_gmv_ars > 0 else growth_gmv_ars
    # CVR real de la hoja CVR% (promedio 4 semanas). Fallback: CR% del Growth OS.
    _cvr_from_sheet, _ = get_cvr_for_brand(name, cr_fallback=conversion_raw)
    _cr_raw      = _cvr_from_sheet if _cvr_from_sheet and _cvr_from_sheet > 0 else (conversion_raw if conversion_raw and conversion_raw > 0 else 0)

    # Food cost estimado por categoría (% sobre precio de venta)
    _category_fc_map = {
        "hamburguesa": 0.42, "burger": 0.42,
        "pizza": 0.38, "pizzeria": 0.38,
        "sushi": 0.45, "japonesa": 0.45,
        "pollo": 0.40, "chicken": 0.40,
        "helado": 0.35, "heladeria": 0.35,
        "cafe": 0.30, "cafeteria": 0.30,
        "empanada": 0.36, "empanadas": 0.36,
        "medialunas": 0.33, "panaderia": 0.33,
        "wrap": 0.40, "saludable": 0.38,
    }
    _cat_key = category.lower() if category else ""
    _food_cost_rate = next(
        (v for k, v in _category_fc_map.items() if k in _cat_key),
        0.40  # default 40%
    )

    # ── 1. Margen neto estimado por orden ────────────────────────────────────
    _margin_per_order = _aov * (1 - _comm_rate) * (1 - _food_cost_rate) if _aov > 0 else 0
    _margin_pct_display = round((1 - _comm_rate) * (1 - _food_cost_rate) * 100, 1)

    # ── 1b. GMV neto total potencial a día de hoy ────────────────────────────
    # Margen por orden × órdenes del mes = margen bruto total del período.
    # Si hay campaña de Ads activa, se descuenta el presupuesto — los bookings
    # de Current ADS ya vienen normalizados a base semanal, por eso se
    # mensualizan ×4 antes de restar.
    _margin_total_bruto = _margin_per_order * _orders if _margin_per_order > 0 and _orders > 0 else 0
    _ads_is_active = bool(ads_current.get("active", False))
    _ads_weekly_budget = to_number(ads_current.get("bookings_usd"), 0) if _ads_is_active else 0
    _ads_monthly_budget_usd = _ads_weekly_budget * 4
    _ads_monthly_budget_ars = _ads_monthly_budget_usd * ARS_PER_USD
    _margin_total_neto = max(_margin_total_bruto - _ads_monthly_budget_ars, 0) if _ads_is_active else _margin_total_bruto

    # ── 2. Punto de equilibrio MD ────────────────────────────────────────────
    # Lógica basada en AOV: el costo del descuento se calcula por orden,
    # no sobre el GMV total histórico (que mezcla coberturas pasadas).
    # Costo del descuento por orden = AOV × 20%
    # Órdenes extra necesarias = costo por orden ÷ margen por orden
    _md_discount_rate      = 0.20
    _promo_cost_per_order  = _aov * _md_discount_rate if _aov > 0 else 0
    _be_orders             = math.ceil(_promo_cost_per_order / _margin_per_order) if _margin_per_order > 0 else 0
    _be_pct_over_current   = round(_be_orders / _orders * 100, 1) if _orders > 0 else 0

    # ── 3. GMV incremental con Traffic real x CVR benchmark ──────────────────
    # Traffic mensual = promedio semanal x 4 (mas robusto que ultima semana sola)
    _cr_bench_cat      = get_cvr_category_benchmark(category)
    _cr_current_norm   = _cr_raw if _cr_raw <= 1 else _cr_raw / 100
    _cr_benchmark_norm = _cr_bench_cat if _cr_bench_cat and _cr_bench_cat > 0 else 0.045
    _traffic_weekly    = get_traffic_for_brand(name)
    _traffic_monthly   = (_traffic_weekly * 4) if _traffic_weekly and _traffic_weekly > 0 else 0
    _cr_above_bench    = _cr_current_norm > 0 and _cr_benchmark_norm > 0 and _cr_current_norm >= _cr_benchmark_norm
    if _traffic_monthly > 0 and _aov > 0:
        _gmv_at_benchmark = _traffic_monthly * _cr_benchmark_norm * _aov
        _gmv_incremental  = max(_gmv_at_benchmark - _gmv, 0) if not _cr_above_bench else 0
    else:
        _impressions_est  = round(_orders / _cr_current_norm) if _cr_current_norm > 0 else 0
        _gmv_at_benchmark = _impressions_est * _cr_benchmark_norm * _aov if _impressions_est > 0 and _aov > 0 else 0
        _gmv_incremental  = max(_gmv_at_benchmark - _gmv, 0) if not _cr_above_bench else 0
    _traffic_source    = "Traffic real" if _traffic_monthly > 0 else "est. por inversa"



    def _consultive_card(emoji, title, main_value, main_sub, pitch_label, pitch_text, color="#1B3F8B"):
        return f"""
<div style="background:rgba(255,255,255,0.90);border-radius:12px;padding:14px 16px;min-height:160px;">
  <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#aaa;letter-spacing:.05em;margin-bottom:8px;">{emoji} {title}</div>
  <div style="font-size:24px;font-weight:900;color:{color};line-height:1.1;">{main_value}</div>
  <div style="font-size:12px;color:#6B7280;margin-top:3px;margin-bottom:10px;">{main_sub}</div>
  <div style="border-top:1px solid rgba(78,99,217,0.12);padding-top:8px;">
    <div style="font-size:10px;font-weight:700;text-transform:uppercase;color:#1B3F8B;margin-bottom:4px;">Cómo decírselo al dueño</div>
    <div style="font-size:12px;color:#6B7280;line-height:1.45;font-style:italic;">"{pitch_text}"</div>
  </div>
</div>"""

    # ── Compute card values for all 4 analytics cards ───────────────────────
    _be_color = "#FF4D2E" if _be_pct_over_current > 40 else "#FF7124" if _be_pct_over_current > 20 else "#7ED321"
    _aov_rounded        = round(_aov / 1000) * 1000
    _promo_cost_rounded = round(_promo_cost_per_order / 100) * 100
    _promos_per_order   = round(_margin_per_order / _promo_cost_per_order, 1) if _promo_cost_per_order > 0 else 0
    _coverage_line      = f" · 1 orden limpia cubre {_promos_per_order}x promos" if _promos_per_order > 0 else ""
    _be_pitch = (
        f"Con un descuento del 20% sobre un ticket de {fmt_ars(_aov_rounded)}, cada orden con promo te cuesta {fmt_ars(_promo_cost_rounded)}. "
        + (f"Pero tu margen por orden es {fmt_ars(round(_margin_per_order))} — eso alcanza para cubrir {_promos_per_order} promos con un solo pedido limpio. " if _promos_per_order >= 1 else "")
        + f"Con {_be_orders} pedido{'s' if _be_orders != 1 else ''} extra ya cubrís ese costo — sin tocar tu estructura."
        + (" Es una meta razonable con tráfico de temporada." if _be_pct_over_current <= 30 else " Es exigente pero alcanzable si hay un evento de alto tráfico." if _be_pct_over_current <= 60 else " Es muy exigente — evaluá acotar los productos en promo para bajar el umbral.")
    )
    _cr_display    = f"{round(_cr_current_norm*100,1)}%" if _cr_current_norm > 0 else "s/d"
    _bench_display = f"{round(_cr_benchmark_norm*100,1)}%" if _cr_benchmark_norm > 0 else "s/d"
    _bench_source  = "real categ." if _cr_bench_cat and _cr_bench_cat > 0 else "ref. general"
    _traffic_disp  = f"{round(_traffic_monthly):,}".replace(",", ".") if _traffic_monthly > 0 else None
    if _cr_current_norm <= 0:
        _inc_color = "#aaa"
        _c3_main   = "Sin dato de CR"
        _c3_sub    = f"Benchmark {_bench_display} ({_bench_source}) · activá ads para medir tráfico"
        _c3_pitch  = "No hay tasa de conversión registrada. Con ads activos medimos el tráfico real y de ahí calculamos el potencial exacto."
    elif _cr_above_bench:
        _delta_pp  = round((_cr_current_norm - _cr_benchmark_norm) * 100, 1)
        _inc_color = "#1A1A2E"
        _c3_main   = f"+{_delta_pp}pp sobre benchmark"
        _c3_sub    = f"CR {_cr_display} vs benchmark {_bench_display} ({_bench_source}) · ya convertís mejor que el promedio"
        _c3_pitch  = (
            f"Tu CR ({_cr_display}) ya está {_delta_pp}pp por encima del promedio de tu categoría ({_bench_display}). "
            + (f"Con {_traffic_disp} visitas mensuales reales, el problema no es la tienda — es el volumen de tráfico." if _traffic_disp else "El problema no es la tienda — es el volumen de tráfico.")
        )
    else:
        _inc_color = "#7ED321" if _gmv_incremental > 0 else "#aaa"
        _c3_main   = fmt_ars(round(_gmv_incremental)) if _gmv_incremental > 0 else "Sin dato suficiente"
        _c3_sub    = f"CR {_cr_display} → benchmark {_bench_display} ({_bench_source}) · {_traffic_source}"
        _c3_pitch  = (
            f"Tu tienda convierte al {_cr_display}, el promedio de tu categoría está en {_bench_display}. "
            + (f"Con tus {_traffic_disp} visitas mensuales, si llegás al benchmark sumás {fmt_ars(round(_gmv_incremental))} por mes — sin invertir más en pauta." if _traffic_disp and _gmv_incremental > 0 else f"Si llegás al benchmark, sumarías {fmt_ars(round(_gmv_incremental))} por mes con el mismo tráfico que ya tenés." if _gmv_incremental > 0 else "Activá ads para empezar a generar tráfico medible.")
        )

    _t_bench  = get_traffic_category_benchmark(category)
    _traffic_ok  = (_traffic_weekly is not None and _t_bench is not None and _traffic_weekly >= _t_bench * 0.85)
    _cvr_ok      = (_cr_current_norm > 0 and _cr_benchmark_norm > 0 and _cr_current_norm >= _cr_benchmark_norm * 0.85)
    _has_traffic = _traffic_weekly is not None and _traffic_weekly > 0
    _has_cvr     = _cr_current_norm > 0
    if not _has_traffic and not _has_cvr:
        _d4_color = "#aaa"; _d4_main = "Sin datos"
        _d4_sub   = "No hay Traffic ni CVR registrado esta semana"
        _d4_pitch = "Activá ads para empezar a generar tráfico y CVR medibles."
    else:
        if not _traffic_ok and not _cvr_ok:
            _d4_color = "#FF4D2E"; _d4_diag = "Problema doble"; _d4_diag_sub = "Tráfico bajo y conversión baja"
        elif not _traffic_ok:
            _d4_color = "#FF7124"; _d4_diag = "Problema: Tráfico"; _d4_diag_sub = "La tienda convierte bien · falta visibilidad"
        elif not _cvr_ok:
            _d4_color = "#FF7124"; _d4_diag = "Problema: Conversión"; _d4_diag_sub = "Hay visitas · la tienda no convierte"
        else:
            _d4_color = "#7ED321"; _d4_diag = "Ambas métricas OK"; _d4_diag_sub = "Traffic y CVR sobre benchmark de categoría"
        _lost_orders = round(_traffic_weekly * max(_cr_benchmark_norm - _cr_current_norm, 0)) if (_has_traffic and _has_cvr and not _cvr_ok) else 0
        _d4_main = _d4_diag
        _tw_disp = f"{round(_traffic_weekly):,}".replace(",", ".") if _has_traffic else "s/d"
        _tb_disp = f"{round(_t_bench):,}".replace(",", ".") if _t_bench else "s/d"
        _d4_sub  = f"{_d4_diag_sub} · " + (f"{_lost_orders} ords/sem perdidas por CVR bajo" if _lost_orders > 0 else f"Traffic {_tw_disp} vs bench {_tb_disp}/sem")
        if _d4_diag == "Problema: Tráfico":
            _traffic_gap = max((_t_bench - _traffic_weekly), 0) if _t_bench and _traffic_weekly else 0
            _step_visits = round(_traffic_gap * 0.30) if _traffic_gap > 0 else 0
            _step_cost_ars = _step_visits * 650
            _step_orders = round(_step_visits * _cr_current_norm, 1) if _cr_current_norm > 0 else 0
            _step_gmv = round(_step_orders * _aov) if _aov > 0 else 0
            _step_cost_disp = f"ARS {_step_cost_ars:,.0f}".replace(",", ".")
            _step_gmv_disp  = fmt_ars(round(_step_gmv / 1000) * 1000) if _step_gmv > 0 else ""
            _traffic_proj = (f" Para arrancar: compramos {_step_visits} visitas más por semana — son {_step_cost_disp}/sem. Con tu conversión del {_cr_display} eso son {_step_orders} pedidos extra · {_step_gmv_disp} GMV/sem." if _step_visits > 0 and _step_gmv > 0 else "")
            _d4_pitch = f"Tu tienda convierte al {_cr_display} — está por encima del promedio de tu categoría. El problema es que ves {_tw_disp} visitas por semana contra un benchmark de {_tb_disp}. Más tráfico con esta tasa de conversión se convierte directo en pedidos." + _traffic_proj
        elif _d4_diag == "Problema: Conversión":
            _d4_pitch = f"Tenés {_tw_disp} visitas por semana — el tráfico no es el problema. Pero tu tienda convierte al {_cr_display} cuando el promedio de tu categoría es {_bench_display}. " + (f"Eso son {_lost_orders} pedidos por semana que te estás perdiendo sin gastar un peso más en pauta." if _lost_orders > 0 else "Con mejoras en menú y fotos ese CVR sube sin invertir más en pauta.")
        elif _d4_diag == "Problema doble":
            _d4_pitch = f"Dos frentes abiertos: traffic de {_tw_disp} vs benchmark {_tb_disp} y CVR de {_cr_display} vs {_bench_display}. " + (f"Combinados, perdés {_lost_orders} pedidos por semana. " if _lost_orders > 0 else "") + "La prioridad es primero limpiar la tienda y después escalar tráfico — al revés es tirar plata."
        else:
            _d4_pitch = f"Traffic en {_tw_disp}/sem y CVR en {_cr_display} — ambas métricas sobre el benchmark de tu categoría. Estás en condiciones de escalar: más presupuesto en ads se convierte directo en GMV."



    booster = booster_for_badge
    actions = actions_for_badge

    campaign_design = design_campaign_for_brand(
        name,
        category,
        current_gmv_ars,
        current_aov_ars,
        get_from_row(row, ["cr %", "conversion rate", "conversion"], 0),
        get_from_row(row, ["pro users %", "pro %", "pro users", "prime users %"], 0),
        get_from_row(row, ["comm. rate", "commission rate", "commission"], 0),
        ads_current,
        md_current,
        md_pro_current,
        booster,
        actions,
        brand_id=brand_id,
    )

    # MD action must show both the recommended booster and the promo suggested by Campaign Designer.
    for _a in actions:
        if "MD" in clean(_a.get("area"), ""):
            _secondary = list(_a.get("secondary", []))
            _promo = clean(campaign_design.get("md_reco"), "") if campaign_design else ""
            _booster = clean(booster.get("event"), "-") if isinstance(booster, dict) else "-"
            if _promo and _promo not in ["", "-"]:
                _secondary.append(f"Promo { _promo }")
            if _booster not in ["", "-"]:
                _secondary.append(f"Booster { _booster }")
            _a["secondary"] = list(dict.fromkeys([clean(x, "") for x in _secondary if clean(x, "")]))

    tactical_flow = build_tactical_flow(
        brand_id,
        name,
        row,
        category,
        current,
        ads_current,
        md_current,
        md_pro_current,
        booster,
        actions,
        campaign_design,
    )

    # ── Build lever→tactical mapping so each 360 card gets its priority content ──
    # Map lever_class → list of tactical items from priority
    # Items with no lever_class (e.g. "general") fall into OPS as a catch-all.
    _lever_tactical_map = {}
    for _item in tactical_flow.get("items", []):
        _lc = clean(_item.get("lever_class"), "") or "lever-ops"
        _lever_tactical_map.setdefault(_lc, []).append(_item)

    fill_colors_map = {
        'health-green':  '#7ED321',
        'health-yellow': '#1B3F8B',
        'health-orange': '#FF7124',
        'health-red':    '#FF4D2E',
    }
    pct_colors_map = {
        'health-green':  '#7ED321',
        'health-yellow': '#D95A10',
        'health-orange': '#D95A10',
        'health-red':    '#FF4D2E',
    }
    area_emojis_map = {
        'lever-ops':  '⚙️',
        'lever-menu': '🍔',
        'lever-md':   '🏷️',
        'lever-pro':  '👑',
        'lever-ads':  '🚀',
    }

    def _merged_action_card(a):
        area_raw      = clean(a.get('area'), '')
        lever_cls     = _action_area_lever_class(area_raw)
        health_cls    = clean(a.get('health_class'), 'health-green')
        action_text   = clean(a.get('action'), 'Following')
        reason_text   = clean(a.get('reason'), '')
        secondary_text = ' · '.join([clean(x, '') for x in a.get('secondary', []) if clean(x, '')])

        badge_raw   = clean(a.get('health_badge'), '100')
        score_match = re.search(r'(\d+(?:\.\d+)?)', badge_raw)
        score_pct   = float(score_match.group(1)) if score_match else 100.0
        score_pct   = max(0.0, min(100.0, score_pct))

        ring_color = '#FFFFFF' if lever_cls == 'lever-menu' else '#7ED321'
        pct_color  = pct_colors_map.get(health_cls, '#7ED321')
        emoji      = area_emojis_map.get(lever_cls, '📊')

        r     = 22
        circ  = 2 * 3.14159 * r
        filled = round(circ * score_pct / 100, 2)
        gap    = round(circ - filled, 2)

        ring_svg = (
            f'<svg width="60" height="60" viewBox="0 0 60 60" xmlns="http://www.w3.org/2000/svg">'
            f'<circle cx="30" cy="30" r="{r}" fill="none" stroke="rgba(0,0,0,0.08)" stroke-width="5"/>'
            f'<circle cx="30" cy="30" r="{r}" fill="none" stroke="{ring_color}" stroke-width="5" '
            f'stroke-linecap="round" stroke-dasharray="{filled} {gap}" transform="rotate(-90 30 30)"/>'
            f'<text x="30" y="35" text-anchor="middle" font-size="18">{emoji}</text>'
            f'</svg>'
        )

        # ── Top half: ring + health data ──────────────────────────────────────
        top_html = (
            f"<div style='display:flex;align-items:center;gap:10px;'>"
            f"{ring_svg}"
            f"<div style='display:flex;flex-direction:column;gap:1px;'>"
            f"<span style='font-size:28px;font-weight:900;color:{pct_color};line-height:1;'>{score_pct:.0f}%</span>"
            f"<span class='action-area'>{html.escape(area_raw)}</span>"
            f"</div>"
            f"</div>"
            f"<div class='action-main'>{html.escape(action_text)}</div>"
            + (f"<div class='action-reason'>{html.escape(reason_text)}</div>" if reason_text else "")
            + (f"<div class='action-secondary'>{html.escape(secondary_text)}</div>" if secondary_text else "")
        )

        # ── Bottom half: priority tactical items for this lever ───────────────
        tactical_items = _lever_tactical_map.get(lever_cls, [])
        if tactical_items:
            divider = (
                "<div style='margin:14px 0 10px;border-top:1px solid rgba(78,99,217,0.18);'></div>"
                "<div style='font-size:10px;font-weight:900;text-transform:uppercase;"
                "letter-spacing:.06em;color:#1B3F8B;margin-bottom:8px;'>🎯 Priority Signal</div>"
            )
            items_html = ""
            for ti in tactical_items:
                t_main = clean(ti.get('main'), '')
                t_cue  = clean(ti.get('cue') or ti.get('argument'), '')
                t_cls  = clean(ti.get('class'), 'health-yellow')
                t_color_map = {
                    'health-green': '#7ED321', 'health-yellow': '#D95A10',
                    'health-orange': '#D95A10', 'health-red': '#FF4D2E',
                }
                t_col = t_color_map.get(t_cls, '#D95A10')
                items_html += (
                    f"<div style='margin-bottom:8px;'>"
                    f"<div style='font-size:13px;font-weight:900;color:{t_col};line-height:1.15;'>{html.escape(t_main)}</div>"
                    + (f"<div style='font-size:11px;color:#6B7280;margin-top:3px;line-height:1.35;font-weight:700;'>{html.escape(t_cue)}</div>" if t_cue else "")
                    + "</div>"
                )
            bottom_html = divider + items_html
        else:
            bottom_html = ""

        return (
            f"<div class='action-card {html.escape(health_cls)} {html.escape(lever_cls)}'>"
            + top_html
            + bottom_html
            + "</div>"
        )

    # ── Priority metadata header ──────────────────────────────────────────────
    flow = tactical_flow
    score_text = _priority_score_display(flow.get("priority_score")) if flow.get("found_priority") else "—"
    rank_text  = '#' + str(int(flow.get('rank'))) if flow.get('rank') not in [None, '', '-'] and not pd.isna(flow.get('rank')) else '—'
    last_contact_text = clean(flow.get('last_contact'), '—')
    coinv_text = clean(flow.get('coinversion'), 'No')
    expired_n  = int(flow.get('promo_vencida') or 0)
    expiring_n = int(flow.get('promo_por_vencer') or 0)
    levers     = flow.get('lever_texts', [])
    lever_chips = "".join([f"<span class='priority-chip'>{html.escape(clean(x,'-'))}</span>" for x in levers[:10]]) or "<span class='priority-chip'>✅ Sin palancas priority activas</span>"

    meta_html = (
        f"<div class='priority-top-grid' style='margin-bottom:14px;'>"
        f"<div><div class='info-mini-label'>🔥 Priority Score</div><div class='info-mini-value'>{html.escape(score_text)}</div></div>"
        f"<div><div class='info-mini-label'># Contacto</div><div class='info-mini-value'>{html.escape(rank_text)}</div></div>"
        f"<div><div class='info-mini-label'>Último Contacto</div><div class='info-mini-value'>{html.escape(last_contact_text)}</div></div>"
        f"<div><div class='info-mini-label'>Coinversión MD</div><div class='info-mini-value'>{html.escape(coinv_text)}</div></div>"
        f"<div><div class='info-mini-label'>Vencida / Por vencer</div><div class='info-mini-value'>{expired_n} / {expiring_n}</div></div>"
        f"<div class='priority-levers' style='grid-column:1/-1;'>{lever_chips}</div>"
        f"</div>"
    )

    campaign_names = get_md_campaign_names_for_brand(name)
    ads_booking_display, _ads_booking_note = _ads_booking_display_parts(ads_current)

    _cvr_weekly_val, _cvr_source = get_cvr_for_brand(name, cr_fallback=conversion_raw)
    _cvr_bench = get_cvr_category_benchmark(category)
    st.markdown(render_business_cards_html(ads_current, md_current, md_pro_current, campaign_names, ads_booking_display, pro_users_display, conversion_display, commission_display, pro_users_raw, conversion_raw, commission_raw, cvr_weekly=_cvr_weekly_val, cvr_source=_cvr_source, cvr_bench=_cvr_bench), unsafe_allow_html=True)

    actions_html = "".join([_merged_action_card(a) for a in actions])
    st.markdown(f"""
<div class="wide-info-card tactical-flow-card">
    <div class="wide-info-title">360° Action</div>
    {meta_html}
    <div class="action-grid">{actions_html}</div>
</div>
""", unsafe_allow_html=True)


    # ── Pitch calculation (must happen before Analytics render) ──────────────
    _pi_category = category.split("·")[0].strip() if "·" in category else category.strip()
    _pi_lever = "Ads" if md_current.get("active", False) else "MD"
    _pi_gmv = current_gmv_ars
    _pi_aov = current_aov_ars
    _pi_orders = current["orders"] if current else 0
    mctx = get_market_context(_pi_category, _pi_lever, _pi_gmv, _pi_orders)

    # ── Reasoning paragraph (same logic as render_campaign_designer_html) ──
    _cd = campaign_design  # shorthand
    _rp_strategy  = clean(_cd.get("strategy"), "")
    _rp_focus     = clean(_cd.get("focus"), "")
    _rp_ads       = clean(_cd.get("ads_action"), "")
    _rp_md        = clean(_cd.get("md_reco"), "")
    _rp_promo     = clean(_cd.get("promo_action"), "")
    _rp_event     = clean(_cd.get("event"), "")
    _rp_cross     = clean(_cd.get("cross_sell_reco"), "")
    _rp_pro_extra = int(to_number(_cd.get("pro_extra"), 0))
    _rp_impact_low  = int(_cd.get("impact_low", 0))
    _rp_impact_high = int(_cd.get("impact_high", 0))
    _rp_risk      = clean(_cd.get("risk"), "Medium")
    _rp_pressure  = _cd.get("partner_pressure", 0)
    _rp_budget    = _format_budget_range(_cd.get("budget_low_ars", 0), _cd.get("budget_high_ars", 0))
    _rp_raw_reasons = list(_cd.get("reasons", []))
    if _cd.get("cross_sell_reason") not in ["", "-"]:
        _rp_raw_reasons.append(_cd.get("cross_sell_reason"))
    if _cd.get("pro_reason") not in ["", "-"]:
        _rp_raw_reasons.append(_cd.get("pro_reason"))

    _rp_parts = []
    if _rp_strategy and _rp_strategy != "-":
        _rp_lead = f"Estrategia <strong>{html.escape(_rp_strategy)}</strong>"
        if _rp_focus and _rp_focus != "-":
            _rp_lead += f" con foco en <em>{html.escape(_rp_focus)}</em>"
        _rp_parts.append(_rp_lead)
    _rp_levers = []
    if _rp_ads and _rp_ads != "-":
        _rp_levers.append(f"Ads → {html.escape(_rp_ads)} ({_rp_budget})")
    if _rp_md and _rp_md != "-":
        _rp_md_detail = f"{html.escape(_rp_promo)}" if _rp_promo and _rp_promo != "-" else ""
        _rp_pro_detail = f" +{_rp_pro_extra}% PRO" if _rp_pro_extra > 0 else ""
        _rp_levers.append(f"MD → {html.escape(_rp_md)}{(' · ' + _rp_md_detail) if _rp_md_detail else ''}{_rp_pro_detail}")
    if _rp_cross and _rp_cross not in ["-", ""]:
        _rp_levers.append(f"Cross-sell: {html.escape(_rp_cross)}")
    if _rp_levers:
        _rp_parts.append(". ".join(_rp_levers))
    if _rp_event and _rp_event not in ["-", "", "No seasonal event priority"]:
        _rp_parts.append(f"Booster estacional disponible: <strong>{html.escape(_rp_event)}</strong>")
    if _rp_raw_reasons:
        _rp_signals = "; ".join([clean(r, "") for r in _rp_raw_reasons[:3] if clean(r, "")])
        if _rp_signals:
            _rp_parts.append(f"Señales clave: {html.escape(_rp_signals)}")
    _rp_parts.append(
        f"Impacto proyectado <strong>+{_rp_impact_low}%–+{_rp_impact_high}% GMV</strong> · "
        f"Riesgo <strong>{html.escape(_rp_risk)}</strong> · Presión aliado {int(_rp_pressure * 100)}%."
    )
    reasoning_paragraph = ". ".join(_rp_parts) + "." if _rp_parts else ""

    _bullets = []

    # 1. Posición en la categoría
    if mctx.get("brand_percentile") and mctx["brand_percentile"] != "N/D":
        pct_val = mctx["brand_percentile"].replace("%", "").strip()
        try:
            pct_num = float(pct_val)
            if pct_num >= 75:
                _bullets.append(f"Esta marca está en el <strong>percentil {mctx['brand_percentile']}</strong> de {_pi_category} en CABA — ya es de las que más venden en su categoría.")
            elif pct_num >= 50:
                _bullets.append(f"La marca está en el <strong>percentil {mctx['brand_percentile']}</strong> de {_pi_category} — por encima de la mitad de la categoría, con espacio real para subir.")
            else:
                _bullets.append(f"La marca está en el <strong>percentil {mctx['brand_percentile']}</strong> de {_pi_category} — hay marcas similares vendiendo mucho más en la misma categoría.")
        except Exception:
            pass

    # 2. GMV vs promedio de categoría
    if _pi_gmv > 0 and mctx.get("market_gmv_avg"):
        try:
            avg_raw = mctx["market_gmv_avg"].replace("ARS", "").replace("$", "").replace(".", "").replace(",", ".").strip()
            avg_num = float(avg_raw)
            if avg_num > 0:
                ratio = _pi_gmv / avg_num
                if ratio >= 1.5:
                    _bullets.append(f"Su GMV actual ({fmt_ars(_pi_gmv)}) es <strong>{ratio:.1f}x el promedio</strong> de la categoría ({mctx['market_gmv_avg']}) — argumento sólido para escalar inversión.")
                elif ratio >= 0.8:
                    _bullets.append(f"Su GMV ({fmt_ars(_pi_gmv)}) está cerca del promedio de la categoría ({mctx['market_gmv_avg']}) — ya tiene la base, le falta el empujón.")
                else:
                    _bullets.append(f"Su GMV ({fmt_ars(_pi_gmv)}) está por debajo del promedio de la categoría ({mctx['market_gmv_avg']}) — hay un gap concreto para trabajar con {_pi_lever}.")
        except Exception:
            pass

    # 3. AOV vs promedio de categoría
    if _pi_aov > 0 and mctx.get("market_aov_avg"):
        try:
            aov_avg_raw = mctx["market_aov_avg"].replace("ARS", "").replace("$", "").replace(".", "").replace(",", ".").strip()
            aov_avg_num = float(aov_avg_raw)
            if aov_avg_num > 0:
                aov_ratio = _pi_aov / aov_avg_num
                if aov_ratio >= 1.2:
                    _bullets.append(f"Su ticket promedio ({fmt_ars(_pi_aov)}) está <strong>{((aov_ratio-1)*100):.0f}% por encima</strong> del AOV de la categoría ({mctx['market_aov_avg']}) — cliente de mayor valor, más razón para darle visibilidad.")
                elif aov_ratio < 0.85:
                    _bullets.append(f"Su ticket promedio ({fmt_ars(_pi_aov)}) está por debajo del AOV de la categoría ({mctx['market_aov_avg']}) — MD puede ayudar a mover volumen y compensar el ticket bajo.")
        except Exception:
            pass

    # 4. Palanca activa / inactiva
    if _pi_lever == "Ads":
        if not ads_current.get("active", False):
            _bullets.append(f"No tiene Ads activo — en {_pi_category}, las marcas con Ads capturan tráfico que esta marca hoy está regalando a la competencia.")
        elif ads_roi > 0:
            _bullets.append(f"Ads activo con ROI de <strong>{ads_roi:.1f}x</strong> — ya está probado que funciona, el argumento es escalar, no empezar.")
    else:
        if not md_current.get("active", False):
            _bullets.append(f"Sin MD activo — en {_pi_category}, el markdown es la palanca más directa para aumentar frecuencia de pedido y subir en el ranking.")
        elif md_roi > 0:
            _bullets.append(f"MD activo con ROI de <strong>{md_roi:.1f}x</strong> — base para proponer un upgrade de descuento o ampliar el alcance.")

    # 5. Top de la categoría
    if mctx.get("market_top_brand") and mctx["market_top_brand"] not in ["-", "N/D"]:
        _bullets.append(f"El top de {_pi_category} en CABA es <strong>{html.escape(mctx['market_top_brand'])}</strong> con {mctx.get('market_top_gmv','N/D')} — ese es el benchmark real de la categoría.")


    # ── Analytics — single wide-info-card with 4 inner cards + pitch ────────
    # Build pitch html first (pure Python, no st calls)
    _items_html = ""
    _reasoning_html = ""
    if _bullets or reasoning_paragraph:
        _items_html = "".join(
            f"<div style='display:flex;gap:10px;margin-bottom:9px;'>"
            f"<span style='color:#7ED321;font-size:14px;line-height:1.5;flex-shrink:0;'>›</span>"
            f"<span style='font-size:13px;line-height:1.6;'>{b}</span>"
            f"</div>"
            for b in _bullets
        )
        if reasoning_paragraph:
            _reasoning_html = (
                f"<div style='font-size:13px;color:#6B7280;line-height:1.65;margin-bottom:0;'>{reasoning_paragraph}</div>"
                f"<hr style='border:none;border-top:1px solid rgba(255,255,255,0.95);margin:14px 0;'>"
            )

    _pitch_block = ""
    if _bullets or reasoning_paragraph:
        _pitch_block = (
            f"<div style='background:rgba(27,63,139,0.03);border:1px solid rgba(255,255,255,0.97);"
            f"border-radius:14px;padding:20px 22px;margin-top:16px;'>"
            f"<div style='font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.60);"
            f"letter-spacing:.06em;margin-bottom:12px;'>📋 Datos para el pitch · {html.escape(_pi_lever)} · {html.escape(_pi_category)}</div>"
            f"<div>{_reasoning_html}{_items_html}</div>"
            f"</div>"
        )

    # Build cheat sheet html
    _cs_lines = []
    _cs_lines.append(f"🗣️ <strong>Apertura:</strong> \"Hola, soy Sabas de Rappi. Te llamo porque vi que {name} tiene una oportunidad concreta de mejorar su posición en {_pi_category} esta semana.\"")
    if mctx.get("brand_percentile") and mctx["brand_percentile"] != "N/D":
        _cs_lines.append(f"📊 <strong>Dato ancla:</strong> \"Estás en el percentil {mctx['brand_percentile']} de {_pi_category} en CABA. Hay marcas similares a la tuya que están vendiendo significativamente más con la palanca correcta.\"")
    if mctx.get("market_top_brand") and mctx["market_top_brand"] not in ["-", "N/D"]:
        _cs_lines.append(f"🏆 <strong>Benchmark:</strong> \"El líder de {_pi_category} en CABA es {html.escape(mctx['market_top_brand'])} con {mctx.get('market_top_gmv','N/D')}. Eso es lo que podés apuntar con el stack correcto.\"")
    if _pi_lever == "MD" and not md_current.get("active", False):
        _cs_lines.append(f"💡 <strong>Pitch {_pi_lever}:</strong> \"Sin MD activo, estás perdiendo frecuencia de pedido. En {_pi_category}, el markdown es la palanca más directa para subir en el ranking. ¿Arrancamos con {campaign_design.get('discount', 20)}% esta semana?\"")
    elif _pi_lever == "Ads" and not ads_current.get("active", False):
        _cs_lines.append(f"💡 <strong>Pitch {_pi_lever}:</strong> \"Sin Ads activo, el tráfico que genera Rappi en {_pi_category} va directo a tu competencia. Con el presupuesto inicial te asegurás visibilidad inmediata.\"")
    elif _pi_lever == "Ads" and ads_roi > 0:
        _cs_lines.append(f"💡 <strong>Pitch {_pi_lever}:</strong> \"Tus Ads ya tienen ROI de {ads_roi:.1f}x. Eso significa que ya probaste que funciona. El paso lógico es escalar, no mantener el mismo presupuesto.\"")
    _cs_lines.append(f"✅ <strong>Cierre:</strong> \"Entonces quedamos en activar {campaign_design.get('ads_action','la palanca')} esta semana. ¿El martes a las 10 te va bien para confirmar que quedó activo?\"")

    # ── Pitch Facts: 3 cards — Dato Ancla · Benchmark · Pitch Lever ────────────
    # Card 1: Dato Ancla — posición percentil de la marca
    _pf_ancla_label = "📊 Dato Ancla"
    if mctx.get("brand_percentile") and mctx["brand_percentile"] != "N/D":
        _pf_ancla_main = f"Percentil {mctx['brand_percentile']}"
        _pf_ancla_body = f"Estás en el percentil {mctx['brand_percentile']} de {_pi_category} en CABA."
        try:
            _pf_pct_num = float(mctx["brand_percentile"].replace("%","").strip())
            if _pf_pct_num >= 75:
                _pf_ancla_body += " Ya sos de las marcas que más venden en tu categoría."
            elif _pf_pct_num >= 50:
                _pf_ancla_body += " Estás por encima de la mitad — hay espacio real para subir."
            else:
                _pf_ancla_body += " Hay marcas similares vendiendo mucho más con la palanca correcta."
        except Exception:
            pass
        _pf_ancla_color = "#7ED321" if (_pf_pct_num if "brand_percentile" in mctx else 0) >= 75 else "#FF7124"
    elif _pi_gmv > 0 and mctx.get("market_gmv_avg"):
        try:
            _avg_n = float(mctx["market_gmv_avg"].replace("ARS","").replace("$","").replace(".","").replace(",",".").strip())
            _ratio = _pi_gmv / _avg_n if _avg_n > 0 else 0
            _pf_ancla_main = f"{_ratio:.1f}x el promedio"
            _pf_ancla_body = f"Su GMV ({fmt_ars(_pi_gmv)}) es {_ratio:.1f}x el promedio de la categoría ({mctx['market_gmv_avg']})."
            _pf_ancla_color = "#7ED321" if _ratio >= 1.5 else "#FF7124"
        except Exception:
            _pf_ancla_main = "N/D"
            _pf_ancla_body = "Sin datos de posición en la categoría."
            _pf_ancla_color = "#6B7280"
    else:
        _pf_ancla_main = "N/D"
        _pf_ancla_body = "Sin datos de posición en la categoría."
        _pf_ancla_color = "#6B7280"

    # Card 2: Benchmark — top de la categoría
    _pf_bench_label = "🏆 Benchmark"
    if mctx.get("market_top_brand") and mctx["market_top_brand"] not in ["-", "N/D"]:
        _pf_bench_main = mctx.get("market_top_gmv", "N/D")
        _pf_bench_body = f"El líder de {_pi_category} en CABA es {mctx['market_top_brand']} con {mctx.get('market_top_gmv','N/D')}. Ese es el benchmark real."
        _pf_bench_color = "#FF7124"
    else:
        _pf_bench_main = "N/D"
        _pf_bench_body = "Sin datos del top de la categoría."
        _pf_bench_color = "#6B7280"

    # Card 3: Pitch Lever — argumento directo para la palanca
    _pf_pitch_label = f"💡 Pitch {_pi_lever}"
    if _pi_lever == "Ads":
        if not ads_current.get("active", False):
            _pf_pitch_main = "Sin Ads activo"
            _pf_pitch_body = f"El tráfico que genera Rappi en {_pi_category} va directo a tu competencia. Con el presupuesto inicial te asegurás visibilidad inmediata."
            _pf_pitch_color = "#FF4D2E"
        elif ads_roi > 0:
            _pf_pitch_main = f"ROI {ads_roi:.1f}x"
            _pf_pitch_body = f"Tus Ads ya tienen ROI de {ads_roi:.1f}x. Ya probaste que funciona — el paso lógico es escalar, no mantener el mismo presupuesto."
            _pf_pitch_color = "#7ED321"
        else:
            _pf_pitch_main = "Ads activo"
            _pf_pitch_body = "Ads activo. Revisá el ROI para definir si mantener o escalar."
            _pf_pitch_color = "#FF7124"
    else:
        if not md_current.get("active", False):
            _pf_pitch_main = "Sin MD activo"
            _pf_pitch_body = f"Sin MD activo estás perdiendo frecuencia de pedido. En {_pi_category} el markdown es la palanca más directa para subir en el ranking."
            _pf_pitch_color = "#FF4D2E"
        elif md_roi > 0:
            _pf_pitch_main = f"ROI {md_roi:.1f}x"
            _pf_pitch_body = f"MD activo con ROI de {md_roi:.1f}x — base para proponer un upgrade de descuento o ampliar el alcance."
            _pf_pitch_color = "#7ED321"
        else:
            _pf_pitch_main = "MD activo"
            _pf_pitch_body = "MD activo. Revisá el ROI para definir la próxima acción."
            _pf_pitch_color = "#FF7124"

    _pitch_facts_block = "".join([
        '<div style="margin-top:16px;">',
        f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.60);letter-spacing:.06em;margin-bottom:10px;">📋 Pitch Facts · {html.escape(_pi_lever)} · {html.escape(_pi_category)}</div>',
        '<div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;">',
        '<div style="background:rgba(255,255,255,0.90);border:1px solid rgba(255,255,255,0.97);border-radius:12px;padding:16px 18px;">',
        f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.60);letter-spacing:.06em;margin-bottom:8px;">{_pf_ancla_label}</div>',
        f'<div style="font-size:22px;font-weight:900;color:{_pf_ancla_color};line-height:1.1;margin-bottom:8px;">{_pf_ancla_main}</div>',
        f'<div style="font-size:12px;color:#6B7280;line-height:1.55;">{_pf_ancla_body}</div>',
        '</div>',
        '<div style="background:rgba(255,255,255,0.90);border:1px solid rgba(255,255,255,0.97);border-radius:12px;padding:16px 18px;">',
        f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.60);letter-spacing:.06em;margin-bottom:8px;">{_pf_bench_label}</div>',
        f'<div style="font-size:22px;font-weight:900;color:{_pf_bench_color};line-height:1.1;margin-bottom:8px;">{_pf_bench_main}</div>',
        f'<div style="font-size:12px;color:#6B7280;line-height:1.55;">{_pf_bench_body}</div>',
        '</div>',
        '<div style="background:rgba(255,255,255,0.90);border:1px solid rgba(255,255,255,0.97);border-radius:12px;padding:16px 18px;">',
        f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.60);letter-spacing:.06em;margin-bottom:8px;">{_pf_pitch_label}</div>',
        f'<div style="font-size:22px;font-weight:900;color:{_pf_pitch_color};line-height:1.1;margin-bottom:8px;">{_pf_pitch_main}</div>',
        f'<div style="font-size:12px;color:#6B7280;line-height:1.55;">{_pf_pitch_body}</div>',
        '</div>',
        '</div>',
        '</div>',
    ])

    # ── Card unificada: Funnel Traffic/CVR + GMV incremental + diagnóstico ──────
    # Reemplaza las dos cards separadas (GMV incremental / Diagnóstico Traffic&CVR)
    # por un único funnel de 3 niveles: benchmark traffic → traffic marca →
    # conversión marca (con benchmark de conversión al lado). Cada nivel sale
    # según la data disponible — si falta, queda "s/d" en vez de inventar un valor.

    _fn_bench_traffic_val = _t_bench if _t_bench and _t_bench > 0 else None
    _fn_brand_traffic_val = _traffic_weekly if _traffic_weekly and _traffic_weekly > 0 else None
    _fn_brand_cvr_val     = _cr_current_norm if _cr_current_norm > 0 else None
    _fn_bench_cvr_val     = _cr_benchmark_norm if _cr_benchmark_norm > 0 else None

    _fn_bench_traffic_disp = f"{round(_fn_bench_traffic_val):,}/sem".replace(",", ".") if _fn_bench_traffic_val else "s/d"
    _fn_brand_traffic_disp = f"{round(_fn_brand_traffic_val):,}/sem".replace(",", ".") if _fn_brand_traffic_val else "s/d"
    _fn_brand_cvr_disp     = f"{round(_fn_brand_cvr_val*100,1)}%" if _fn_brand_cvr_val else "s/d"
    _fn_bench_cvr_disp     = f"{round(_fn_bench_cvr_val*100,1)}%" if _fn_bench_cvr_val else "s/d"

    # ── Anchos del funnel (en % del ancho disponible del SVG, 0-100) ──────────
    # Regla pedida: las dos barras de tráfico (benchmark y marca) son siempre
    # más anchas que la de conversión. El ancho del nivel 1 (benchmark) es fijo
    # como referencia visual; el nivel 2 (marca) escala proporcional al
    # benchmark y SÍ puede superarlo (desbordar hacia afuera, no solo angostar).
    # El nivel 3 (conversión) se calcula como fracción del ancho del nivel 2
    # (tráfico de marca) — no de un techo externo — para que la forma de
    # embudo sea consistente con "la conversión se angosta sobre el tráfico
    # que la marca realmente tiene", tal como se pidió.
    _FN_MAX_W = 100.0   # ancho máximo absoluto permitido (referencia del benchmark)
    _FN_MIN_W = 22.0    # ancho mínimo visual para que una barra nunca desaparezca

    if _fn_bench_traffic_val and _fn_brand_traffic_val:
        _fn_w1 = _FN_MAX_W
        _fn_w2 = round(min(max(_fn_brand_traffic_val / _fn_bench_traffic_val * _FN_MAX_W, _FN_MIN_W), _FN_MAX_W * 1.35), 1)
    elif _fn_brand_traffic_val and not _fn_bench_traffic_val:
        # Solo hay traffic de marca, sin benchmark — el nivel 1 queda en s/d (mínimo)
        # y el nivel 2 toma el ancho de referencia para no aplastar el embudo.
        _fn_w1 = _FN_MIN_W
        _fn_w2 = _FN_MAX_W
    elif _fn_bench_traffic_val and not _fn_brand_traffic_val:
        _fn_w1 = _FN_MAX_W
        _fn_w2 = _FN_MIN_W
    else:
        _fn_w1 = _FN_MIN_W
        _fn_w2 = _FN_MIN_W

    # Conversión: SIEMPRE más angosta que el tráfico de marca (nivel 2), escalada
    # como fracción del propio nivel 2 según qué tan bien convierte vs su benchmark.
    # Si convierte igual al benchmark → ~55% del ancho del nivel 2 (referencia visual
    # de embudo). Si convierte mejor → un poco más ancho; si peor → más angosto.
    if _fn_brand_cvr_val:
        _fn_cvr_ratio = (_fn_brand_cvr_val / _fn_bench_cvr_val) if _fn_bench_cvr_val else 1.0
        _fn_cvr_ratio = max(min(_fn_cvr_ratio, 1.8), 0.35)  # clamp para que no se desborde ni desaparezca
        _fn_w3 = round(min(_fn_w2 * 0.55 * _fn_cvr_ratio, _fn_w2 * 0.92), 1)  # nunca >92% del nivel 2
        _fn_w3 = max(_fn_w3, _FN_MIN_W * 0.7)
    else:
        _fn_w3 = _FN_MIN_W * 0.5  # s/d → barra mínima casi invisible, sin inventar dato

    # ── Estimado de órdenes según visitas de la marca × su CVR real ──────────
    # Solo se calcula si hay AMBOS datos reales (traffic de marca y CVR) —
    # nunca se infiere a partir de un s/d.
    _fn_orders_est = None
    if _fn_brand_traffic_val and _fn_brand_cvr_val:
        _fn_orders_est = round(_fn_brand_traffic_val * _fn_brand_cvr_val, 1)

    _fn_traffic_brand_above = bool(_fn_bench_traffic_val and _fn_brand_traffic_val and _fn_brand_traffic_val > _fn_bench_traffic_val)
    _fn_cvr_brand_above     = bool(_fn_bench_cvr_val and _fn_brand_cvr_val and _fn_brand_cvr_val >= _fn_bench_cvr_val)

    # Paleta del funnel: 3 tonos distintos y legibles en light y dark mode —
    # naranja (benchmark, referencia neutra), azul/verde (traffic de marca),
    # y verde/rojo condicional para CVR según esté sobre o bajo benchmark.
    _fn_c1 = "#FF7124"  # benchmark traffic — naranja de marca, siempre neutro
    _fn_c2 = "#7ED321" if _fn_traffic_brand_above else "#1B6FE0"  # traffic marca: verde si supera benchmark, azul si no
    _fn_c3 = "#7ED321" if _fn_cvr_brand_above else ("#FF4D2E" if _fn_brand_cvr_val else "#8C93AC")

    # ── Construcción del SVG tipo embudo (trapecios apilados, sin sangría de
    # línea para evitar el bug de bloque-de-código de Markdown). ──────────────
    _FN_SVG_W, _FN_SVG_H = 520, 168
    _FN_LEVEL_H = 46
    _FN_GAP = 5
    _fn_cx = _FN_SVG_W / 2

    def _fn_trapezoid_points(top_w_pct, bot_w_pct, y_top, level_h, max_shape_w):
        top_w = (top_w_pct / 100) * max_shape_w
        bot_w = (bot_w_pct / 100) * max_shape_w
        x1, x2 = _fn_cx - top_w / 2, _fn_cx + top_w / 2
        x3, x4 = _fn_cx + bot_w / 2, _fn_cx - bot_w / 2
        y_bot = y_top + level_h
        return f"{x1:.1f},{y_top:.1f} {x2:.1f},{y_top:.1f} {x3:.1f},{y_bot:.1f} {x4:.1f},{y_bot:.1f}"

    _fn_y1 = 0
    _fn_y2 = _fn_y1 + _FN_LEVEL_H + _FN_GAP
    _fn_y3 = _fn_y2 + _FN_LEVEL_H + _FN_GAP

    # El embudo ocupa solo la mitad izquierda del SVG — la mitad derecha queda
    # libre para las etiquetas, que ya no van adentro de la forma (eso era lo
    # que cortaba el texto en niveles angostos como el de tráfico de marca).
    _FN_SHAPE_MAX_W = _FN_SVG_W * 0.46
    _fn_cx = _FN_SHAPE_MAX_W / 2 + 6

    _fn_pts1 = _fn_trapezoid_points(_fn_w1, _fn_w2, _fn_y1, _FN_LEVEL_H, _FN_SHAPE_MAX_W)
    _fn_pts2 = _fn_trapezoid_points(_fn_w2, _fn_w3, _fn_y2, _FN_LEVEL_H, _FN_SHAPE_MAX_W)
    _fn_pts3 = _fn_trapezoid_points(_fn_w3, max(_fn_w3 * 0.78, _FN_MIN_W * 0.4), _fn_y3, _FN_LEVEL_H, _FN_SHAPE_MAX_W)

    _fn_orders_badge = f" · ~{_fn_orders_est} ped/sem" if _fn_orders_est is not None else ""

    _fn_label1_top = "Traffic benchmark categoría"
    _fn_label1_val = _fn_bench_traffic_disp
    _fn_label2_top = "Traffic de la marca" + (" ▲ sobre benchmark" if _fn_traffic_brand_above else "")
    _fn_label2_val = _fn_brand_traffic_disp
    _fn_label3_top = "Conversión de la marca"
    _fn_label3_bench_note = f"bench {_fn_bench_cvr_disp}"
    _fn_label3_val = f"{_fn_brand_cvr_disp}{_fn_orders_badge}"

    def _fn_level_center_y(y_top, level_h):
        return y_top + level_h / 2

    # ── Etiquetas FUERA de la forma: columna de texto a la derecha + línea
    # conectora desde el centro del trapecio. Así el ancho de texto nunca
    # depende de qué tan angosto sea el nivel, eliminando el corte de texto. ──
    _fn_label_x = _FN_SHAPE_MAX_W + 22
    _fn_line_x_end = _fn_label_x - 6

    def _fn_label_block(y_top, level_h, color, top_text, val_text, val_size=13, note_text=None):
        _cy = _fn_level_center_y(y_top, level_h)
        _note_svg = (
            f'<text x="{_fn_label_x:.1f}" y="{_cy+22:.1f}" font-size="9" font-weight="600" fill="currentColor" opacity="0.5">{html.escape(note_text)}</text>'
            if note_text else ""
        )
        return "".join([
            f'<line x1="{_fn_cx:.1f}" y1="{_cy:.1f}" x2="{_fn_line_x_end:.1f}" y2="{_cy:.1f}" stroke="{color}" stroke-width="1.5" stroke-dasharray="2,3" opacity="0.55"></line>',
            f'<circle cx="{_fn_line_x_end:.1f}" cy="{_cy:.1f}" r="3" fill="{color}"></circle>',
            f'<text x="{_fn_label_x:.1f}" y="{_cy-7:.1f}" font-size="10" font-weight="700" fill="currentColor" opacity="0.65">{html.escape(top_text)}</text>',
            f'<text x="{_fn_label_x:.1f}" y="{_cy+9:.1f}" font-size="{val_size}" font-weight="900" fill="{color}">{html.escape(val_text)}</text>',
            _note_svg,
        ])

    _fn_labels_svg = (
        _fn_label_block(_fn_y1, _FN_LEVEL_H, _fn_c1, _fn_label1_top, _fn_label1_val)
        + _fn_label_block(_fn_y2, _FN_LEVEL_H, _fn_c2, _fn_label2_top, _fn_label2_val)
        + _fn_label_block(_fn_y3, _FN_LEVEL_H, _fn_c3, _fn_label3_top, _fn_label3_val, val_size=12, note_text=_fn_label3_bench_note)
    )

    _funnel_svg = "".join([
        f'<svg viewBox="0 0 {_FN_SVG_W} {_fn_y3 + _FN_LEVEL_H + 4}" width="100%" height="auto" xmlns="http://www.w3.org/2000/svg" style="color:{"#E4E7F1" if DARK_MODE else "#1A1A2E"};">',
        f'<polygon points="{_fn_pts1}" fill="{_fn_c1}" opacity="0.92"></polygon>',
        f'<polygon points="{_fn_pts2}" fill="{_fn_c2}" opacity="0.92"></polygon>',
        f'<polygon points="{_fn_pts3}" fill="{_fn_c3}" opacity="0.92"></polygon>',
        _fn_labels_svg,
        '</svg>',
    ])

    _funnel_html = f'<div style="display:flex;justify-content:center;padding:4px 0;">{_funnel_svg}</div>'

    # ── Párrafo combinado: funde el diagnóstico de GMV incremental (card 3)
    # con el diagnóstico de traffic/CVR (card 4 vieja) en una sola lectura. ──
    if not _has_traffic and not _has_cvr:
        _fn_combined_text = "No hay traffic ni conversión registrados esta semana. Activá ads para empezar a generar ambas métricas de forma medible."
        _fn_pitch = "Activá ads para empezar a generar tráfico y CVR medibles — sin eso no podemos calcular dónde está la oportunidad real."
        _fn_headline = "Sin datos suficientes"
        _fn_headline_color = "#8C93AC"
    elif _cr_above_bench and not _has_traffic:
        _fn_combined_text = f"Tu CR ({_fn_brand_cvr_disp}) ya está sobre el benchmark de tu categoría ({_fn_bench_cvr_disp}), pero no hay traffic registrado esta semana para medir el volumen."
        _fn_pitch = f"Tu tienda convierte mejor que el promedio ({_fn_brand_cvr_disp} vs {_fn_bench_cvr_disp}). El problema no es la tienda — necesitamos activar ads para medir y escalar el tráfico real."
        _fn_headline = "Conversión fuerte, falta tráfico medible"
        _fn_headline_color = "#FF7124"
    elif _d4_diag == "Problema doble":
        _fn_combined_text = f"Dos frentes abiertos: traffic de {_fn_brand_traffic_disp} vs benchmark {_fn_bench_traffic_disp}, y conversión de {_fn_brand_cvr_disp} vs {_fn_bench_cvr_disp}." + (f" Si llegaras al benchmark de conversión con el mismo tráfico, sumarías {fmt_ars(round(_gmv_incremental))}/mes." if _gmv_incremental > 0 else "")
        _fn_pitch = f"Dos frentes abiertos: traffic de {_fn_brand_traffic_disp} vs benchmark {_fn_bench_traffic_disp} y CVR de {_fn_brand_cvr_disp} vs {_fn_bench_cvr_disp}. " + (f"Combinados, perdés {_lost_orders} pedidos por semana. " if _lost_orders > 0 else "") + "La prioridad es primero limpiar la tienda y después escalar tráfico — al revés es tirar plata."
        _fn_headline = "Problema doble"
        _fn_headline_color = "#FF4D2E"
    elif _d4_diag == "Problema: Tráfico":
        _fn_combined_text = f"Tu conversión ({_fn_brand_cvr_disp}) está sobre el benchmark ({_fn_bench_cvr_disp}), pero el tráfico ({_fn_brand_traffic_disp}) está por debajo del benchmark de categoría ({_fn_bench_traffic_disp})." + (f" Si alcanzaras el benchmark de CVR con más tráfico, el incremental estimado sería {fmt_ars(round(_gmv_incremental))}/mes." if _gmv_incremental > 0 else "")
        _fn_pitch = _d4_pitch
        _fn_headline = "Problema: Tráfico"
        _fn_headline_color = "#FF7124"
    elif _d4_diag == "Problema: Conversión":
        _fn_combined_text = f"Tu tráfico ({_fn_brand_traffic_disp}) está en línea con el benchmark ({_fn_bench_traffic_disp}), pero tu conversión ({_fn_brand_cvr_disp}) está por debajo del promedio de categoría ({_fn_bench_cvr_disp})." + (f" Si llegás al benchmark, sumás {fmt_ars(round(_gmv_incremental))}/mes con el mismo tráfico." if _gmv_incremental > 0 else "")
        _fn_pitch = _d4_pitch
        _fn_headline = "Problema: Conversión"
        _fn_headline_color = "#FF7124"
    else:
        _fn_combined_text = f"Traffic ({_fn_brand_traffic_disp}) y conversión ({_fn_brand_cvr_disp}) están alineados o por encima del benchmark de categoría ({_fn_bench_traffic_disp} / {_fn_bench_cvr_disp})."
        _fn_pitch = _d4_pitch
        _fn_headline = "Ambas métricas OK"
        _fn_headline_color = "#7ED321"

    # NOTA: mismo fix — fragmentos sin sangría de línea, unidos con join().
    _funnel_card_html = "".join([
        '<div style="background:rgba(255,255,255,0.90);border:1px solid rgba(0,0,0,0.07);border-radius:14px;padding:16px 18px;display:flex;flex-direction:column;">',
        '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.60);letter-spacing:.06em;margin-bottom:10px;">🔍 Funnel Traffic &amp; Conversión vs Benchmark</div>',
        f'<div style="font-size:15px;font-weight:900;color:{_fn_headline_color};margin-bottom:10px;">{_fn_headline}</div>',
        f'<div style="margin-bottom:12px;">{_funnel_html}</div>',
        '<div style="border-top:1px solid rgba(255,255,255,0.95);padding-top:10px;margin-top:auto;">',
        f'<div style="font-size:11px;color:#6B7280;line-height:1.5;margin-bottom:10px;">{_fn_combined_text}</div>',
        '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.55);margin-bottom:4px;">Cómo decírselo al dueño</div>',
        f'<div style="font-size:11px;color:#6B7280;line-height:1.5;font-style:italic;">"{_fn_pitch}"</div>',
        '</div>',
        '</div>',
    ])

    # NOTA: mismo fix aplicado a todo el bloque Analytics — sin sangría de línea,
    # construido con join() de fragmentos en vez de f-string multilínea indentado.
    _analytics_html = "".join([
        '<div class="wide-info-card">',
        '<div class="wide-info-title">Analytics</div>',
        '<div style="display:grid;grid-template-columns:1fr 1fr 1.4fr;gap:14px;margin-top:4px;">',
        '<div style="background:rgba(255,255,255,0.90);border:1px solid rgba(0,0,0,0.07);border-radius:14px;padding:16px 18px;">',
        '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.60);letter-spacing:.06em;margin-bottom:8px;">💰 Margen neto / orden</div>',
        f'<div style="font-size:26px;font-weight:900;color:#7ED321;line-height:1.1;">{fmt_ars(round(_margin_per_order))}</div>',
        f'<div style="font-size:12px;color:#6B7280;margin-top:4px;margin-bottom:8px;">{_margin_pct_display}% del ticket · food cost {round(_food_cost_rate*100)}% + comisión {round(_comm_rate*100)}%</div>',
        '<div style="background:rgba(126,211,33,0.08);border-radius:8px;padding:8px 10px;margin-bottom:12px;">',
        '<div style="font-size:10px;font-weight:700;color:#5A9E00;text-transform:uppercase;margin-bottom:2px;">GMV neto total · este mes</div>',
        f'<div style="font-size:16px;font-weight:900;color:#1A1A2E;">{fmt_ars(round(_margin_total_neto))}</div>',
        f'<div style="font-size:10px;color:#6B7280;margin-top:2px;">{f"Bruto {fmt_ars(round(_margin_total_bruto))} − Ads {fmt_ars(round(_ads_monthly_budget_ars))}/mes (booking semanal ×4)" if _ads_is_active else "Sin descuento de Ads · campaña no activa"}</div>',
        '</div>',
        '<div style="border-top:1px solid rgba(255,255,255,0.95);padding-top:10px;">',
        '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.55);margin-bottom:4px;">Cómo decírselo al dueño</div>',
        f'<div style="font-size:11px;color:#6B7280;line-height:1.5;font-style:italic;">"Por cada pedido de {fmt_ars(round(_aov))} que te entra, después de la comisión ({round(_comm_rate*100)}%) y el costo del producto ({round(_food_cost_rate*100)}%), quedan {fmt_ars(round(_margin_per_order))} para cubrir fijos. Con tu volumen del mes, eso son {fmt_ars(round(_margin_total_neto))} de margen real{" después de descontar tu inversión en Ads" if _ads_is_active else ""}."</div>',
        '</div>',
        '</div>',
        '<div style="background:rgba(255,255,255,0.90);border:1px solid rgba(0,0,0,0.07);border-radius:14px;padding:16px 18px;">',
        '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.60);letter-spacing:.06em;margin-bottom:8px;">⚖️ Punto de equilibrio MD 20%</div>',
        f'<div style="font-size:26px;font-weight:900;color:{_be_color};line-height:1.1;">+{_be_orders} orden{"es" if _be_orders != 1 else ""}</div>',
        f'<div style="font-size:12px;color:#6B7280;margin-top:4px;margin-bottom:12px;">Cada orden con promo te cuesta {fmt_ars(round(_promo_cost_per_order))}{_coverage_line}</div>',
        '<div style="border-top:1px solid rgba(255,255,255,0.95);padding-top:10px;">',
        '<div style="font-size:10px;font-weight:700;text-transform:uppercase;color:rgba(107,114,128,0.55);margin-bottom:4px;">Cómo decírselo al dueño</div>',
        f'<div style="font-size:11px;color:#6B7280;line-height:1.5;font-style:italic;">"{_be_pitch}"</div>',
        '</div>',
        '</div>',
        _funnel_card_html,
        '</div>',
        _pitch_facts_block,
        '</div>',
    ])
    st.markdown(_analytics_html, unsafe_allow_html=True)

    # ── Campaign Designer (after Analytics) ───────────────────────────────────
    st.markdown(render_campaign_designer_html(campaign_design), unsafe_allow_html=True)

    # ── Card "Brand vs Brand": evolución GMV/AOV de 3 meses, renderizada inline ──
    # Reemplaza el flujo anterior de prompt → Gemini. Reutiliza _dot_line_chart_card,
    # la misma función que ya pinta las cards de arriba — sin redirect externo.
    _mctx = get_market_context(category, "GMV", brand_gmv=current_gmv_ars or growth_gmv_ars)
    _percentil  = _mctx.get("brand_percentile", "N/D")
    _cvr_brand_norm = _cr_current_norm if _cr_current_norm and _cr_current_norm > 0 else 0
    _cvr_bench_norm = _cr_benchmark_norm if _cr_benchmark_norm and _cr_benchmark_norm > 0 else 0
    _cvr_is_below = (_cvr_brand_norm > 0 and _cvr_bench_norm > 0 and _cvr_brand_norm < _cvr_bench_norm)

    _traffic_bench = get_traffic_category_benchmark(category)
    _traffic_brand_weekly = _traffic_weekly if _traffic_weekly and _traffic_weekly > 0 else 0
    _traffic_is_below = (_traffic_brand_weekly > 0 and _traffic_bench and _traffic_bench > 0 and _traffic_brand_weekly < _traffic_bench)

    # GMV incremental si se alcanza el benchmark de la métrica que está fallando
    _gmv_incremental_bvc = 0
    if _cvr_is_below and _traffic_brand_weekly > 0:
        _gmv_incremental_bvc = max((_traffic_brand_weekly * 4) * _cvr_bench_norm * (current_aov_ars or growth_aov_ars) - (current_gmv_ars or growth_gmv_ars), 0)
    elif _traffic_is_below and _cvr_brand_norm > 0:
        _gmv_incremental_bvc = max((_traffic_bench * 4) * _cvr_brand_norm * (current_aov_ars or growth_aov_ars) - (current_gmv_ars or growth_gmv_ars), 0)

    def _funnel_step_html(label, value_display, is_below, benchmark_display):
        _step_color = "#FF4D2E" if is_below else "#7ED321"
        _step_icon  = "▼" if is_below else "▲"
        return f"""
        <div style="display:flex;align-items:center;justify-content:space-between;background:rgba(255,255,255,0.92);
            border-left:3px solid {_step_color};border-radius:8px;padding:10px 14px;margin-bottom:8px;">
          <div>
            <div style="font-size:10px;font-weight:700;color:#6B7280;text-transform:uppercase;letter-spacing:.04em;">{label}</div>
            <div style="font-size:18px;font-weight:900;color:#1A1A2E;margin-top:2px;">{value_display}</div>
          </div>
          <div style="text-align:right;color:{_step_color};font-weight:800;font-size:12px;">
            {_step_icon} vs bench {benchmark_display}
          </div>
        </div>"""

    _funnel_html = (
        _funnel_step_html("Traffic benchmark categoría", f"{round(_traffic_bench):,}/sem".replace(",", ".") if _traffic_bench else "s/d", False, "—")
        + _funnel_step_html("Traffic de la marca", f"{round(_traffic_brand_weekly):,}/sem".replace(",", ".") if _traffic_brand_weekly else "s/d", _traffic_is_below, f"{round(_traffic_bench):,}/sem".replace(",", ".") if _traffic_bench else "s/d")
        + _funnel_step_html("CVR de la marca", f"{round(_cvr_brand_norm*100,1)}%" if _cvr_brand_norm else "s/d", _cvr_is_below, f"{round(_cvr_bench_norm*100,1)}%" if _cvr_bench_norm else "s/d")
    )

    _bvc_summary = (
        f"Tu tienda está en el percentil {_percentil} de GMV en su categoría. "
        + (f"El CVR ({round(_cvr_brand_norm*100,1)}%) está por debajo del benchmark ({round(_cvr_bench_norm*100,1)}%) — " if _cvr_is_below else "")
        + (f"el tráfico ({round(_traffic_brand_weekly):,}/sem) está por debajo del benchmark ({round(_traffic_bench):,}/sem) — ".replace(",", ".") if _traffic_is_below else "")
        + (f"si se alcanza el benchmark de la métrica más débil, el incremental estimado es {fmt_ars(round(_gmv_incremental_bvc))}/mes." if _gmv_incremental_bvc > 0 else "ambas métricas están alineadas o por encima del benchmark de categoría.")
    )

    st.markdown(f"""
<style>
  .inf-card {{
    border-radius: 14px;
    padding: 20px 22px 18px;
    display: flex;
    flex-direction: column;
    gap: 10px;
    margin-bottom: 14px;
  }}
  .inf-card.blue {{
    background: linear-gradient(135deg, rgba(59,72,131,.08), rgba(59,72,131,.14));
    border: 1.5px solid #1B3F8B;
  }}
  .inf-card.orange {{
    background: linear-gradient(135deg, rgba(255,113,36,.06), rgba(255,113,36,.12));
    border: 1.5px solid #FF7124;
  }}
  .inf-title {{ font-size: 14px; font-weight: 800; color: #1A1A2E; }}
  .inf-desc  {{ font-size: 11px; color: #6B7280; line-height: 1.5; }}
  .inf-chart-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
</style>

<div class="inf-card blue">
  <div class="inf-title">📈 Brand vs Brand</div>
  <div class="inf-desc">Evolución de GMV y AOV en los últimos 3 meses.</div>
  <div class="inf-chart-row">
    {_dot_line_chart_card("GMV", current_gmv_ars, may_gmv_ars, abril_gmv_ars, fmt_ars, fmt_usd(current_gmv_usd))}
    {_dot_line_chart_card("AOV", current_aov_ars, may_aov_ars, abril_aov_ars, fmt_ars, fmt_usd(current_aov_usd))}
  </div>
</div>

<div class="inf-card orange">
  <div class="inf-title">🏪 Brand vs Benchmark</div>
  <div class="inf-desc">{_bvc_summary}</div>
  <div>{_funnel_html}</div>
</div>
""", unsafe_allow_html=True)

    return name



def _priority_bucket_from_kind(kind):
    if kind.startswith("ops"):
        return "ops"
    if kind.startswith("menu"):
        return "menu"
    if kind == "md":
        return "md"
    if kind == "pro":
        return "pro"
    if kind == "ads":
        return "ads"
    return "other"


def get_priority_overview_counts():
    dfp = load_priority_data()
    buckets = {"ops": set(), "menu": set(), "md": set(), "pro": set(), "ads": set()}
    if dfp.empty:
        return {k: 0 for k in buckets}
    lever_rows = dfp[dfp["_metric_norm"] != "total"].copy()
    for _, row in lever_rows.iterrows():
        bid = normalize_brand_id(row.get("_id"))
        metric = clean(row.get("_metric"), "")
        if not bid or not metric:
            continue
        bucket = _priority_bucket_from_kind(_classify_priority_lever(metric))
        if bucket in buckets:
            buckets[bucket].add(bid)
    return {k: len(v) for k, v in buckets.items()}


def render_priority_overview_html():
    counts = get_priority_overview_counts()
    specs = [
        ("ops", "⚙️ OPS Priority", "Operation, wait time, claims, availability", "lever-ops"),
        ("menu", "🍔 Menu Priority", "Photos, catalog, purchase experience", "lever-menu"),
        ("md", "🏷️ Markdown Priority", "Promos, discounts, MD continuity", "lever-md"),
        ("pro", "👑 MD PRO Priority", "PRO capture and coinvestment", "lever-pro"),
        ("ads", "🚀 Ads Priority", "CPC, booking, revenue and visibility", "lever-ads"),
    ]
    cards = []
    for key, title, copy, cls in specs:
        cards.append(
            f"<div class='priority-overview-card {cls}'>"
            f"<div class='card-label'>{html.escape(title)}</div>"
            f"<div class='card-value'>{fmt_number(counts.get(key,0))} accounts</div>"
            f"<div class='card-copy'>{html.escape(copy)}</div>"
            f"<div class='card-chipline'><span class='card-chip'>Priority count by lever</span></div>"
            f"</div>"
        )
    return (
        "<div class='wide-info-card'>"
        "<div class='wide-info-title'>🎯 Priority Overview</div>"
        "<div class='priority-note'>Counts can overlap: if one brand is priority in OPS and Ads, it counts once in each card.</div>"
        f"<div class='priority-overview-grid'>{''.join(cards)}</div>"
        "</div>"
    )


def _priority_overview_specs():
    return [
        ("ops", "⚙️ OPS", "OPS Priority"),
        ("menu", "🍔 Menu", "Menu Priority"),
        ("md", "🏷️ Markdown", "Markdown Priority"),
        ("pro", "👑 MD PRO", "MD PRO Priority"),
        ("ads", "🚀 Ads", "Ads Priority"),
    ]


def _brand_name_map_from_growth():
    growth_df = load_growth_data()
    if growth_df.empty:
        return {}
    id_col = get_id_column_name(growth_df)
    if not id_col:
        return {}
    name_col = _first_existing_col(growth_df, ["name", "brand name", "restaurant name"])
    if not name_col:
        return {}
    result = {}
    for _, row in growth_df.iterrows():
        bid = normalize_brand_id(row.get(id_col))
        if bid:
            result[bid] = clean(row.get(name_col), "-")
    return result


def get_priority_overview_detail(bucket):
    """Returns one row per brand that has at least one Smart Priority lever in the requested bucket.
    Brands can appear in multiple buckets because this is a lever-level view, not a unique portfolio view.
    """
    dfp = load_priority_data()
    if dfp.empty:
        return pd.DataFrame(columns=["Contact Order", "Brand ID", "Brand", "Priority Score", "Main Signal", "Signals", "Last Contact"])

    bucket = clean(bucket, "").lower()
    if bucket not in ["ops", "menu", "md", "pro", "ads"]:
        return pd.DataFrame(columns=["Contact Order", "Brand ID", "Brand", "Priority Score", "Main Signal", "Signals", "Last Contact"])

    work = dfp.copy()
    work["_bucket"] = work["_metric"].apply(lambda m: _priority_bucket_from_kind(_classify_priority_lever(m)))
    lever_rows = work[(work["_metric_norm"] != "total") & (work["_bucket"] == bucket)].copy()
    if lever_rows.empty:
        return pd.DataFrame(columns=["Contact Order", "Brand ID", "Brand", "Priority Score", "Main Signal", "Signals", "Last Contact"])

    names = _brand_name_map_from_growth()
    total_rows = work[work["_metric_norm"] == "total"].copy()
    total_score_map = {}
    total_rank_map = {}
    for _, row in total_rows.iterrows():
        bid = normalize_brand_id(row.get("_id"))
        if not bid:
            continue
        total_score_map.setdefault(bid, to_number(row.get("_score"), 0))
        rank = row.get("_total_rank")
        if rank not in [None, "", "-"] and not pd.isna(rank):
            total_rank_map.setdefault(bid, int(rank))

    contact_col = _first_existing_col(work, ["último contacto", "ultimo contacto", "last contact"])
    rows = []
    for bid, group in lever_rows.groupby("_id", sort=False):
        bid = normalize_brand_id(bid)
        if not bid:
            continue
        group = group.sort_values(by=["_score", "_row_order"], ascending=[False, True])
        main_signal = clean(group.iloc[0].get("_metric"), "-")
        signal_items = []
        for _, r in group.iterrows():
            metric = clean(r.get("_metric"), "-")
            score = _priority_score_display(r.get("_score"))
            label = f"{metric} ({score})" if score != "-" else metric
            if label not in signal_items:
                signal_items.append(label)
        brand_name = names.get(bid, clean(group.iloc[0].get("_brand_col"), "-"))
        all_brand_rows = work[work["_id"].astype(str) == str(bid)].copy()
        last_contact = "-"
        if contact_col and contact_col in all_brand_rows.columns:
            non_empty = all_brand_rows[contact_col].dropna()
            if not non_empty.empty:
                last_contact = _format_priority_date(non_empty.iloc[0])
        rows.append({
            "Contact Order": f"#{total_rank_map.get(bid)}" if bid in total_rank_map else "-",
            "Brand ID": f"AR-{bid}",
            "Brand": brand_name,
            "Priority Score": _priority_score_display(total_score_map.get(bid, 0)),
            "Main Signal": main_signal,
            "Signals": " | ".join(signal_items[:4]),
            "Last Contact": last_contact,
            "_sort_rank": total_rank_map.get(bid, 999999),
            "_sort_score": to_number(total_score_map.get(bid, 0), 0),
        })

    result = pd.DataFrame(rows)
    if result.empty:
        return pd.DataFrame(columns=["Contact Order", "Brand ID", "Brand", "Priority Score", "Main Signal", "Signals", "Last Contact"])
    result = result.sort_values(by=["_sort_rank", "_sort_score"], ascending=[True, False])
    return result[["Contact Order", "Brand ID", "Brand", "Priority Score", "Main Signal", "Signals", "Last Contact"]]


def render_priority_overview_controls():
    counts = get_priority_overview_counts()
    specs = _priority_overview_specs()
    if "priority_overview_bucket" not in st.session_state:
        st.session_state["priority_overview_bucket"] = None

    cols = st.columns(len(specs))
    for col, (key, short_label, full_label) in zip(cols, specs):
        with col:
            if st.button(f"{short_label}\n{fmt_number(counts.get(key, 0))} accounts", key=f"priority_overview_{key}", use_container_width=True):
                st.session_state["priority_overview_bucket"] = key

    selected = st.session_state.get("priority_overview_bucket")
    if selected:
        label_map = {k: full for k, _, full in specs}
        df_detail = get_priority_overview_detail(selected)
        st.markdown(
            f"<div class='wide-info-card'><div class='wide-info-title'>{html.escape(label_map.get(selected, 'Priority Detail'))} Detail</div>"
            f"<div class='priority-note'>Copy a Brand ID from this list and paste it into the search bar to open the full Finder profile.</div></div>",
            unsafe_allow_html=True,
        )
        if df_detail.empty:
            st.info("No brands found for this priority lever.")
        else:
            _render_html_table(df_detail)


def _campaign_period_label(baseline=False):
    if baseline:
        return "Q2 W20 2026"
    return APP_PERIOD


@st.cache_data(ttl=3000, show_spinner=False)
def _load_cpc_supervisor_data(excel_path):
    """Load CPC sheet from main Excel. Returns only STATUS=OK rows (active campaigns)."""
    try:
        df = pd.read_excel(excel_path, sheet_name="CPC", header=0)
        df.columns = [c.strip() for c in df.columns]
        df["BRAND_ID"] = df["BRAND_ID"].apply(normalize_brand_id)
        df["DELIVERY RATE"] = pd.to_numeric(df["DELIVERY RATE"], errors="coerce")
        df["%BUDGET/SALES"] = pd.to_numeric(df["%BUDGET/SALES"], errors="coerce")
        df["%SUG_BUDGET/SALES"] = pd.to_numeric(df["%SUG_BUDGET/SALES"], errors="coerce")
        df["REVENUE TOTAL"] = pd.to_numeric(df["REVENUE TOTAL"], errors="coerce")
        df["BUDGET CONFIGURADO USD"] = pd.to_numeric(df["BUDGET CONFIGURADO USD"], errors="coerce")
        df["END_DATE"] = pd.to_datetime(df["END_DATE"], errors="coerce")
        return df[df["STATUS"] == "OK"].copy()
    except Exception:
        return pd.DataFrame()


def _parse_accionables(text):
    """Parse ACCIONABLES string into a structured dict of flags."""
    if not text or pd.isna(text):
        return {"budget_alto": False, "momentos": False, "usuarios": False, "segmentos": False}
    t = str(text).lower()
    return {
        "budget_alto": "budget alto" in t,
        "momentos":    "momentos de consumo" in t,
        "usuarios":    "tipos de usuarios" in t,
        "segmentos":   "segmentos" in t,
    }


def _ads_cpc_recommendation_from_row(row):
    """
    CPC recommendation combining:
    1. Original ROI + consumption logic (raise/lower/maintain)
    2. Supervisor ACCIONABLES: momentos de consumo, tipos de usuarios, segmentos
    3. Delivery Rate: if <50% with good ROI, reach is the problem not price
    4. Budget gap: %SUG_BUDGET/SALES vs %BUDGET/SALES
    """
    roi         = to_number(row.get("roi"), 0)
    bookings    = to_number(row.get("bookings_usd"), 0)
    revenue     = to_number(row.get("revenue_usd"), 0)
    consumption = revenue / bookings if bookings else 0

    acc         = row.get("_accionables_parsed") or {}
    delivery    = to_number(row.get("_delivery_rate"), None)
    pct_budget  = to_number(row.get("_pct_budget_sales"), None)
    pct_sug     = to_number(row.get("_pct_sug_budget_sales"), None)
    budget_gap  = (pct_sug - pct_budget) if (pct_sug is not None and pct_budget is not None) else None

    parts = []

    # 1. Base CPC signal
    if roi > 2.0 and consumption < 0.80:
        base = "Raise CPC ⚡"
    elif roi < 2.0 and consumption > 0.80:
        base = "Lower CPC 🔻"
    elif roi >= 2.0:
        base = "Maintain CPC ✅"
    else:
        base = "Review CPC 🟡"

    # 2. Delivery rate override: DR < 50% with good ROI = reach problem, not price
    if delivery is not None and delivery < 0.50 and roi >= 2.0:
        base = "Maintain CPC ✅"
        parts.append("DR bajo (<50%) — subir CPC no ayuda si el presupuesto no se consume")

    # 3. ACCIONABLES from supervisor
    if acc.get("momentos"):
        parts.append("⏰ Ampliar momentos de consumo")
    if acc.get("usuarios"):
        parts.append("👥 Agregar tipos de usuario")
    if acc.get("segmentos"):
        parts.append("🎯 Ampliar segmentos de audiencia")
    if acc.get("budget_alto") and not acc.get("momentos") and not acc.get("usuarios") and not acc.get("segmentos"):
        if delivery is not None and delivery < 0.70:
            parts.append("💰 Budget OK — revisar horarios o CPC real")

    # 4. Budget gap signal
    if budget_gap is not None and budget_gap > 0.05:
        parts.append(f"📈 Budget sugerido +{budget_gap*100:.0f}pp vs actual")

    if parts:
        return base + " · " + " / ".join(parts)
    return base



def _pct_change(new, old):
    new = to_number(new, 0)
    old = to_number(old, 0)
    if old == 0:
        return None
    return (new - old) / old




def _campaign_response_speed_from_week(change):
    if change is None:
        return "No baseline"
    if change >= 0.10:
        return "Fast Response 🟢"
    if change > 0:
        return "Normal Response 🟡"
    return "Slow / No lift 🔴"


def _promo_fatigue_score(change, roi):
    roi = to_number(roi, 0)
    if change is None:
        return "No baseline"
    if roi >= 3.2 and change < -0.05:
        return "Fatigue risk ⚠️"
    if change < -0.10:
        return "High fatigue 🔴"
    if change > 0:
        return "Fresh / responding 🟢"
    return "Stable"


def _escalability_score(roi, change):
    roi = to_number(roi, 0)
    if roi >= 4 and (change is not None and change > 0):
        return "High Scale 🟢"
    if roi >= 3.2:
        return "Medium Scale 🟡"
    if roi > 0:
        return "Low Scale ⚪"
    return "Unknown"


def _pressure_stability_ads(roi, consumption):
    roi = to_number(roi, 0)
    consumption = to_number(consumption, 0)
    if roi < 2 and consumption > 0.80:
        return "Overpressure 🔴"
    if roi >= 2 and consumption <= 0.80:
        return "Room to push 🟢"
    if roi >= 2:
        return "Stable ✅"
    return "Watch 🟡"

def _md_weekly_intelligence(change, roi):
    """Promo architecture recommendation for Campaign Weekly Tracker."""
    roi = to_number(roi, 0)
    if change is None:
        return "No baseline yet"
    if change <= -0.15:
        return "Critical: change product or promo architecture"
    if change <= -0.10:
        return "Urgent: rotate HQ / test stronger product"
    if change <= -0.05:
        return "Review: check fatigue and product fit"
    if change > 0.10 and roi >= 3.2:
        return "Scale: add storewide min-ticket or combo"
    if change > 0 and roi >= 3.2:
        return "Maintain or extend winning promo"
    if roi < 3.2:
        return "Optimize: ROI below benchmark"
    return "Stable: maintain"

def _md_alert_from_change(change):
    if change is None:
        return "No baseline yet"
    if change <= -0.15:
        return "Critical 🔴"
    if change <= -0.10:
        return "Urgent 🟠"
    if change <= -0.05:
        return "Review 🟡"
    if change > 0:
        return "Growing 🟢"
    return "Stable ✅"


def _current_campaign_snapshot_rows(period_label, baseline=False):
    rows = []
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    ads = load_current_ads_data(portfolio_only=True)
    if not ads.empty:
        for _, r in ads.iterrows():
            if any(to_number(r.get(c), 0) > 0 for c in ["bookings net", "revenue net", "sales ads usd"]):
                rows.append({
                    "snapshot_datetime": ts,
                    "period": period_label,
                    "channel": "Ads",
                    "brand_id": normalize_brand_id(r.get("_id")),
                    # Use MTD booking with MTD revenue so consumption/CPC decisions compare the same period.
                    "bookings_usd": to_number(r.get("bookings net"), 0),
                    "revenue_usd": to_number(r.get("revenue net"), 0),
                    "sales_usd": to_number(r.get("sales ads usd"), 0),
                    "roi": to_number(r.get("roi"), 0),
                    "gmv_usd": 0,
                    "orders": 0,
                    "campaigns": 0,
                })
    for pro_flag, channel_name in [(False, "Markdown"), (True, "Markdown PRO")]:
        md = load_current_md_data(portfolio_only=True, pro=pro_flag)
        if not md.empty:
            grouped = md.groupby("_id", as_index=False).agg({"_sales_usd":"sum", "_gmv_usd":"sum", "_orders":"sum", "_campaigns":"sum", "_roi_raw":"mean"})
            for _, r in grouped.iterrows():
                sales = to_number(r.get("_sales_usd"), 0)
                gmv = to_number(r.get("_gmv_usd"), 0)
                roi = (gmv / sales) if sales else to_number(r.get("_roi_raw"), 0)
                rows.append({
                    "snapshot_datetime": ts,
                    "period": period_label,
                    "channel": channel_name,
                    "brand_id": normalize_brand_id(r.get("_id")),
                    "bookings_usd": 0,
                    "revenue_usd": 0,
                    "sales_usd": sales,
                    "roi": roi,
                    "gmv_usd": gmv,
                    "orders": to_number(r.get("_orders"), 0),
                    "campaigns": to_number(r.get("_campaigns"), 0),
                })
    return rows


def _load_campaign_weekly_tracker_df():
    if not os.path.exists(CAMPAIGN_WEEKLY_TRACKER_FILE):
        rows = _current_campaign_snapshot_rows(_campaign_period_label(baseline=True), baseline=True)
        df = pd.DataFrame(rows)
        if df.empty:
            df = pd.DataFrame(columns=["snapshot_datetime","period","channel","brand_id","bookings_usd","revenue_usd","sales_usd","roi","gmv_usd","orders","campaigns"])
        df.to_csv(CAMPAIGN_WEEKLY_TRACKER_FILE, index=False, encoding="utf-8-sig")
        return df
    try:
        df = pd.read_csv(CAMPAIGN_WEEKLY_TRACKER_FILE)
    except Exception:
        return pd.DataFrame()
    return df


def _save_campaign_snapshot(period_label):
    df = _load_campaign_weekly_tracker_df()
    rows = _current_campaign_snapshot_rows(period_label)
    new_df = pd.DataFrame(rows)
    if df.empty:
        final = new_df
    else:
        df = df[df["period"].astype(str) != str(period_label)].copy() if "period" in df.columns else df
        final = pd.concat([df, new_df], ignore_index=True)
    final.to_csv(CAMPAIGN_WEEKLY_TRACKER_FILE, index=False, encoding="utf-8-sig")
    return len(new_df)


def _brand_name_map():
    df = load_growth_data()
    id_col = get_id_column_name(df) if not df.empty else None
    if not id_col:
        return {}
    result = {}
    for _, row in df.iterrows():
        bid = normalize_brand_id(row.get(id_col))
        if bid:
            result[bid] = clean(get_from_row(row, ["name", "brand name", "restaurant name"]), "-")
    return result


def _last_four_periods(df):
    if df.empty or "period" not in df.columns:
        return []
    periods = list(dict.fromkeys(df["period"].astype(str).tolist()))
    return periods[-4:]


def page_campaign_weekly_tracker():
    render_header("Campaign Weekly Tracker", "Last 4 weeks · Ads CPC and Markdown performance monitor")

    # ── Reset histórico: si el CSV existe con datos pre-junio 2026, borrarlo ──
    # El nuevo ciclo empieza desde el primer snapshot manual del próximo domingo.
    _reset_col, _capture_col = st.columns([2, 1])
    with _reset_col:
        st.caption(
            f"Histórico reiniciado — el nuevo ciclo empieza con el primer snapshot manual. "
            f"Capturá cada domingo después de exportar Current ADS y Current MD."
        )
    with _capture_col:
        if st.button("📸 Capture current week snapshot"):
            # Wipe CSV completely before saving so old periods don't persist
            if os.path.exists(CAMPAIGN_WEEKLY_TRACKER_FILE):
                os.remove(CAMPAIGN_WEEKLY_TRACKER_FILE)
            saved = _save_campaign_snapshot(_campaign_period_label())
            st.success(f"Snapshot guardado para {_campaign_period_label()}: {saved} filas activas.")
            st.rerun()

    # ── Targets desde Earnings ───────────────────────────────────────────────
    raw_earnings = load_earnings_data()
    _ads_target_usd = to_number(cell(raw_earnings, 2, 1)) if not raw_earnings.empty else ADS_REVENUE_TARGET_USD
    _ads_result_usd = to_number(cell(raw_earnings, 2, 2)) if not raw_earnings.empty else 0
    _md_target_usd  = to_number(cell(raw_earnings, 2, 5)) if not raw_earnings.empty else 0
    _md_result_usd  = to_number(cell(raw_earnings, 2, 6)) if not raw_earnings.empty else 0
    _ads_target_usd = _ads_target_usd if _ads_target_usd > 0 else ADS_REVENUE_TARGET_USD

    # ── KPIs superiores: todos en vivo desde el Excel ────────────────────────
    live_coverage = get_live_campaign_coverage_counts()

    # Ads Revenue en vivo: suma revenue_net de Current ADS (portfolio)
    _live_ads_df = load_current_ads_data(portfolio_only=True)
    _live_ads_revenue = (
        pd.to_numeric(_live_ads_df["revenue net"], errors="coerce").fillna(0).sum()
        if not _live_ads_df.empty and "revenue net" in _live_ads_df.columns
        else 0.0
    )

    # MD GMV en vivo: suma _gmv_usd de Current MD + Current MD PRO (portfolio)
    _live_md_df     = load_current_md_data(portfolio_only=True, pro=False)
    _live_md_pro_df = load_current_md_data(portfolio_only=True, pro=True)
    _live_md_gmv = (
        pd.to_numeric(_live_md_df["_gmv_usd"], errors="coerce").fillna(0).sum()
        if not _live_md_df.empty else 0.0
    ) + (
        pd.to_numeric(_live_md_pro_df["_gmv_usd"], errors="coerce").fillna(0).sum()
        if not _live_md_pro_df.empty else 0.0
    )

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Active Ads",         fmt_number(live_coverage["ads"]))
    c2.metric("Active MD / MD PRO", fmt_number(live_coverage["md"]))
    c3.metric("Ads Revenue (vivo)",  fmt_usd(_live_ads_revenue))
    c4.metric("MD GMV (vivo)",       fmt_usd(_live_md_gmv))

    # ── Tabla Ads CPC Monitor (desde snapshots históricos) ───────────────────
    df = _load_campaign_weekly_tracker_df()
    names = _brand_name_map()
    if df.empty:
        st.markdown("### Ads CPC Monitor")
        st.info("Sin historial de snapshots todavía. Capturá el primer snapshot este domingo para empezar a ver la tabla.")
        return
    periods = _last_four_periods(df)
    work = df[df["period"].astype(str).isin(periods)].copy()
    for c in ["bookings_usd","revenue_usd","sales_usd","roi","gmv_usd","orders","campaigns"]:
        if c in work.columns:
            work[c] = pd.to_numeric(work[c], errors="coerce").fillna(0)
    latest_period = periods[-1] if periods else "-"
    latest = work[work["period"].astype(str) == latest_period].copy()
    ads_latest = latest[latest["channel"] == "Ads"].copy()
    md_latest  = latest[latest["channel"].isin(["Markdown", "Markdown PRO"])].copy()

    st.markdown("### Ads CPC Monitor")
    if ads_latest.empty:
        ads_view = pd.DataFrame(columns=["Period","Brand ID","Brand","Bookings USD","Revenue USD","ROI","ROI Trend","Consumption","Pressure Stability","False ROI Check","CPC Recommendation","Strategic Note"])
    else:
        ads_latest["Brand"] = ads_latest["brand_id"].apply(lambda x: names.get(normalize_brand_id(x), "-"))
        ads_latest["Consumption"] = ads_latest.apply(lambda r: (to_number(r.get("revenue_usd"),0) / to_number(r.get("bookings_usd"),0)) if to_number(r.get("bookings_usd"),0) else 0, axis=1)

        # ── Enrich with CPC supervisor sheet data ────────────────────────────
        _cpc_sup = _load_cpc_supervisor_data(EXCEL_FILE)
        if not _cpc_sup.empty:
            # Aggregate per brand (one brand can have multiple OK campaigns — take latest by END_DATE)
            _cpc_sup["END_DATE"] = pd.to_datetime(_cpc_sup["END_DATE"], errors="coerce")
            _cpc_sup_sorted = _cpc_sup.sort_values("END_DATE", ascending=False, na_position="last")
            _cpc_latest = _cpc_sup_sorted.drop_duplicates(subset="BRAND_ID", keep="first")
            _cpc_map = {}
            for _, _cr in _cpc_latest.iterrows():
                _bid = normalize_brand_id(_cr.get("BRAND_ID"))
                _cpc_map[_bid] = {
                    "_delivery_rate":        _cr.get("DELIVERY RATE"),
                    "_pct_budget_sales":     _cr.get("%BUDGET/SALES"),
                    "_pct_sug_budget_sales": _cr.get("%SUG_BUDGET/SALES"),
                    "_accionables_raw":      _cr.get("ACCIONABLES"),
                    "_accionables_parsed":   _parse_accionables(_cr.get("ACCIONABLES")),
                    "_bidding_method":       _cr.get("BIDDING_METHOD"),
                    "_revenue_supervisor":   _cr.get("REVENUE TOTAL"),
                }
            for col, default in [
                ("_delivery_rate", None), ("_pct_budget_sales", None),
                ("_pct_sug_budget_sales", None), ("_accionables_raw", None),
                ("_accionables_parsed", {}), ("_bidding_method", None),
                ("_revenue_supervisor", None),
            ]:
                ads_latest[col] = ads_latest["brand_id"].apply(
                    lambda x, c=col, d=default: _cpc_map.get(normalize_brand_id(x), {}).get(c, d)
                )
        else:
            for col in ["_delivery_rate", "_pct_budget_sales", "_pct_sug_budget_sales",
                        "_accionables_raw", "_accionables_parsed", "_bidding_method", "_revenue_supervisor"]:
                ads_latest[col] = None

        ads_latest["CPC Recommendation"] = ads_latest.apply(_ads_cpc_recommendation_from_row, axis=1)
        ads_latest["Delivery Rate"] = ads_latest["_delivery_rate"].apply(
            lambda v: f"{v*100:.0f}%" if pd.notna(v) else "—"
        )
        ads_latest["Accionables"] = ads_latest["_accionables_raw"].apply(
            lambda v: str(v) if pd.notna(v) else "—"
        )
        ads_latest["False ROI Check"] = ads_latest.apply(lambda r: "Validate before scaling ⚠️" if to_number(r.get("roi"),0) >= 4 and to_number(r.get("bookings_usd"),0) < 10 else "OK", axis=1)
        ads_latest["Pressure Stability"] = ads_latest.apply(lambda r: _pressure_stability_ads(r.get("roi"), r.get("Consumption")), axis=1)
        ads_latest["ROI"] = ads_latest["roi"].apply(fmt_roi2)

        # ROI Trend — inline SVG sparkline (dots + line, colored by direction)
        def _roi_trend_for_brand(brand_id_val):
            bid = normalize_brand_id(brand_id_val)
            roi_vals = []
            for p in periods:
                subset = work[(work["period"].astype(str) == p) & (work["channel"] == "Ads") & (work["brand_id"].apply(normalize_brand_id) == bid)]
                if not subset.empty:
                    roi_vals.append(to_number(subset.iloc[0].get("roi", 0), 0))
            if not roi_vals:
                return "-"
            # Build mini SVG sparkline 90x22
            if len(roi_vals) == 1:
                return fmt_roi2(roi_vals[0])
            W, H, PAD = 90, 22, 4
            _mn, _mx = min(roi_vals), max(roi_vals)
            _rng = max(_mx - _mn, 0.1)
            def _sx(i): return PAD + i * (W - 2*PAD) / max(len(roi_vals)-1, 1)
            def _sy(v): return H - PAD - (v - _mn) / _rng * (H - 2*PAD)
            line_color = "#7ED321" if roi_vals[-1] >= roi_vals[0] else "#FF4D2E"
            pts = " ".join(f"{_sx(i):.1f},{_sy(v):.1f}" for i, v in enumerate(roi_vals))
            dots = "".join(
                f'<circle cx="{_sx(i):.1f}" cy="{_sy(v):.1f}" r="2.8" fill="{line_color}" title="{fmt_roi2(v)}"/>' 
                for i, v in enumerate(roi_vals)
            )
            label_first = fmt_roi2(roi_vals[0])
            label_last  = fmt_roi2(roi_vals[-1])
            svg = (
                f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" style="vertical-align:middle;" title="{label_first} → {label_last}">' 
                f'<polyline points="{pts}" fill="none" stroke="{line_color}" stroke-width="1.8" stroke-linejoin="round"/>' 
                + dots + f'</svg> <small style="color:{line_color};font-weight:700;">{label_last}</small>'
            )
            return svg

        ads_latest["ROI Trend"] = ads_latest["brand_id"].apply(_roi_trend_for_brand)

        # Strategic Note
        def _ads_strategic_note(row):
            roi = to_number(row.get("roi"), 0)
            consumption = to_number(row.get("Consumption"), 0)
            menu_ok = True  # placeholder — could cross with Perfect Store data
            if roi < 2.0 and consumption > 0.80:
                return "ROI bajo + presupuesto consumido — revisar CPC real externo antes de renovar."
            if roi < 2.0:
                return "ROI bajo · Maintain CPC, probar Markdown antes de subir tráfico."
            if roi >= 4.5 and consumption < 0.60:
                return "Buen ROI trend — posible upselling controlado de presupuesto."
            if consumption >= 0.95:
                return "Presupuesto consumido antes de 7d — revisar subida de presupuesto."
            if roi >= 2.5:
                return "ROI saludable — mantener y monitorear conversión semanal."
            return "Sin fricción visible — revisar CPC real externo si ROI no sube."

        ads_latest["Strategic Note"] = ads_latest.apply(_ads_strategic_note, axis=1)

        # ── ROI drop detector: flag brands that dropped ≥1.0x WoW ────────────
        def _roi_drop_alert(brand_id_val):
            bid = normalize_brand_id(brand_id_val)
            roi_series = []
            for p in periods:
                subset = work[(work["period"].astype(str) == p) & (work["channel"] == "Ads") & (work["brand_id"].apply(normalize_brand_id) == bid)]
                if not subset.empty:
                    roi_series.append(to_number(subset.iloc[0].get("roi", 0), 0))
            if len(roi_series) >= 2:
                drop = roi_series[-2] - roi_series[-1]   # prev - latest
                if drop >= 1.0:
                    return f"🔻 -{drop:.1f}x WoW"
                if drop >= 0.5:
                    return f"⚠️ -{drop:.1f}x WoW"
            return ""

        ads_latest["ROI Alert"] = ads_latest["brand_id"].apply(_roi_drop_alert)

        # Revenue at Risk: revenue USD en juego si la cuenta se pierde o deteriora
        def _ads_revenue_at_risk(row):
            roi = to_number(row.get("roi"), 0)
            consumption = to_number(row.get("Consumption"), 0)
            rev = to_number(row.get("revenue_usd"), 0)
            target = _ads_target_usd if _ads_target_usd > 0 else 1
            if roi < 2.0 or consumption >= 0.95:
                # Alto riesgo: toda la revenue está en riesgo
                pct = round((rev / target) * 100, 1) if target > 0 else 0.0
                return f"🔴 {fmt_usd(rev)} ({pct}% target)"
            elif roi < 2.5 or consumption >= 0.80:
                # Riesgo medio
                at_risk = rev * 0.5
                pct = round((at_risk / target) * 100, 1) if target > 0 else 0.0
                return f"🟡 {fmt_usd(at_risk)} ({pct}% target)"
            else:
                return "🟢 Estable"

        ads_latest["Revenue at Risk"] = ads_latest.apply(_ads_revenue_at_risk, axis=1)
        ads_view = ads_latest[["period","brand_id","Brand","bookings_usd","revenue_usd","ROI","ROI Alert","ROI Trend","Consumption","Pressure Stability","False ROI Check","CPC Recommendation","Accionables","Delivery Rate","Revenue at Risk","Strategic Note"]].rename(columns={"period":"Period","brand_id":"Brand ID","bookings_usd":"Bookings USD","revenue_usd":"Revenue USD"})
    # ── ROI drop alert banner ─────────────────────────────────────────────────
    if not ads_latest.empty and "ROI Alert" in ads_latest.columns:
        _drop_brands = ads_latest[ads_latest["ROI Alert"].str.startswith("🔻", na=False)][["Brand", "ROI Alert", "ROI Trend"]].copy()
        _warn_brands = ads_latest[ads_latest["ROI Alert"].str.startswith("⚠️", na=False)][["Brand", "ROI Alert", "ROI Trend"]].copy()
        if not _drop_brands.empty:
            _drop_items = " &nbsp;·&nbsp; ".join(
                f"<b>{r['Brand']}</b> {r['ROI Alert']} ({r['ROI Trend']})"
                for _, r in _drop_brands.iterrows()
            )
            st.markdown(
                f'<div style="background:rgba(229,51,42,0.10);border-left:4px solid #FF4D2E;border-radius:0 8px 8px 0;'
                f'padding:10px 16px;margin-bottom:10px;font-size:12px;color:#FF4D2E;">'
                f'🔻 <b>Caída de ROI crítica esta semana:</b> {_drop_items}</div>',
                unsafe_allow_html=True,
            )
        if not _warn_brands.empty:
            _warn_items = " &nbsp;·&nbsp; ".join(
                f"<b>{r['Brand']}</b> {r['ROI Alert']} ({r['ROI Trend']})"
                for _, r in _warn_brands.iterrows()
            )
            st.markdown(
                f'<div style="background:rgba(255,113,36,0.10);border-left:4px solid #D95A10;border-radius:0 8px 8px 0;'
                f'padding:10px 16px;margin-bottom:10px;font-size:12px;color:#D95A10;">'
                f'⚠️ <b>Caída de ROI moderada:</b> {_warn_items}</div>',
                unsafe_allow_html=True,
            )

    # Sort by Bookings USD descending so highest-spend brands appear first
    if not ads_view.empty and "Bookings USD" in ads_view.columns:
        ads_view = ads_view.sort_values("Bookings USD", ascending=False).reset_index(drop=True)
        ads_view.index = ads_view.index + 1  # N. starts at 1

    _render_html_table(ads_view)

    # Separate MD Normal and MD PRO sections
    md_normal_latest = latest[latest["channel"] == "Markdown"].copy()
    md_pro_latest = latest[latest["channel"] == "Markdown PRO"].copy()

    # ── Penetration maps: GMV brand y GMV MD desde Current MD (en vivo) ────────
    # Necesitamos el GMV total del brand (Last GMV ARS) para calcular penetración.
    # Fuente: Growth OS sheet (mismo origen que Opportunity List).
    _growth_df = load_growth_data()
    _growth_gmv_map = {}  # brand_id → gmv_brand_usd
    if not _growth_df.empty:
        _id_col_g = get_id_column_name(_growth_df)
        if _id_col_g:
            for _, _gr in _growth_df.iterrows():
                _bid = normalize_brand_id(_gr.get(_id_col_g))
                _gmv_ars = to_number(get_from_row(_gr, ["last gmv ars", "gmv ars", "last gmv local"]), 0)
                if _bid and _gmv_ars > 0:
                    _growth_gmv_map[_bid] = _gmv_ars / ARS_PER_USD

    # GMV MD en vivo desde Current MD (normal) y Current MD PRO
    def _build_live_md_gmv_map(pro_flag):
        """Devuelve {brand_id: gmv_md_usd} leyendo Current MD en tiempo real."""
        _md_live = load_current_md_data(portfolio_only=False, pro=pro_flag)
        result = {}
        if _md_live.empty:
            return result
        grouped = _md_live.groupby("_id", as_index=False).agg({"_gmv_usd": "sum"})
        for _, _r in grouped.iterrows():
            _bid = normalize_brand_id(_r.get("_id"))
            if _bid:
                result[_bid] = to_number(_r.get("_gmv_usd"), 0)
        return result

    _live_md_gmv_map      = _build_live_md_gmv_map(pro_flag=False)
    _live_md_pro_gmv_map  = _build_live_md_gmv_map(pro_flag=True)

    MD_PENE_TARGET = 0.10   # 10% — piso de penetración sana

    def _penetration_block(brand_id_val, is_pro):
        """
        Calcula penetración real del brand usando Current MD en vivo.
        Devuelve dict con: actual_pct, gap_usd, target_usd, status, label.
        """
        bid = normalize_brand_id(brand_id_val)
        gmv_brand_usd = _growth_gmv_map.get(bid, 0)
        gmv_md_usd    = (_live_md_pro_gmv_map if is_pro else _live_md_gmv_map).get(bid, 0)

        if gmv_brand_usd <= 0:
            return {"actual_pct": None, "gap_usd": 0, "target_usd": 0, "status": "sin_gmv", "label": "Sin GMV base"}

        target_usd  = gmv_brand_usd * MD_PENE_TARGET
        actual_pct  = gmv_md_usd / gmv_brand_usd if gmv_md_usd > 0 else 0.0
        gap_usd     = max(target_usd - gmv_md_usd, 0)

        if actual_pct >= MD_PENE_TARGET:
            status = "sano"
            label  = f"✅ {actual_pct*100:.1f}% (≥10%)"
        elif actual_pct > 0:
            status = "bajo"
            label  = f"⚠️ {actual_pct*100:.1f}% (< 10%)"
        else:
            status = "cero"
            label  = "🔴 0% · sin penetración"
        return {"actual_pct": actual_pct, "gap_usd": gap_usd, "target_usd": target_usd, "status": status, "label": label}

    def _renegotiation_from_penetration(pene, roi, change):
        """
        Lógica de renegociación 100% basada en penetración real.
        Solo recomienda renegociar si pene < 10%.
        Devuelve (renegotiation_status, recommendation).
        """
        status = pene["status"]
        gap    = pene["gap_usd"]
        actual = pene["actual_pct"] or 0.0

        if status == "sin_gmv":
            return "—", "Sin GMV base · no se puede calcular penetración."

        if status == "sano":
            return "✅ Penetración OK", "Mantener promo activa y monitorear ROI semana a semana."

        # Penetración < 10% — determinar causa y paso siguiente
        if status == "cero":
            return (
                "🔴 Renegociar · 0% penetración",
                f"El brand tiene promo activa pero 0% de penetración. "
                f"Revisar si la promo está visible y bien configurada. "
                f"Gap al 10%: {fmt_usd(gap)}. "
                f"Paso: llamar y revisar configuración de campaña antes de proponer nueva promo."
            )

        # actual > 0 pero < 10%
        if roi < 2.0:
            rec = (
                f"Penetración baja ({actual*100:.1f}%) + ROI bajo ({roi:.1f}x). "
                f"La promo no está convirtiendo. Gap al 10%: {fmt_usd(gap)}. "
                f"Paso: renegociar el descuento o cambiar producto destacado antes de renovar."
            )
        elif roi >= 3.5 and (change or 0) >= 0:
            rec = (
                f"Penetración baja ({actual*100:.1f}%) pero ROI saludable ({roi:.1f}x). "
                f"El problema es alcance, no conversión. Gap al 10%: {fmt_usd(gap)}. "
                f"Paso: proponer ampliar cobertura de promo (más productos o mayor descuento visible)."
            )
        elif (change or 0) < -0.10:
            rec = (
                f"Penetración baja ({actual*100:.1f}%) + caída WoW de {fmt_signed_percent(change)}. "
                f"Tendencia negativa. Gap al 10%: {fmt_usd(gap)}. "
                f"Paso: renegociar promo esta semana antes de que el brand pierda confianza en el canal."
            )
        else:
            rec = (
                f"Penetración baja ({actual*100:.1f}%). Gap al 10%: {fmt_usd(gap)}. "
                f"Paso: revisar con el brand si hay fricción operativa (stock, visibilidad) "
                f"o proponer un ajuste de descuento para empujar penetración."
            )

        return "🟡 Renegociar", rec

    def _revenue_at_risk_md(pene, gmv_md_usd):
        """Revenue at Risk = GMV que falta para llegar al 10% de penetración."""
        _md_target = _md_target_usd if _md_target_usd > 0 else 1
        gap = pene["gap_usd"]
        status = pene["status"]
        if status == "sin_gmv":
            return "—"
        if status == "sano":
            return "🟢 Penetración OK"
        # En riesgo: lo que falta para llegar al 10% y lo que ya está pero sin penetrar
        at_risk = gmv_md_usd  # el GMV activo está en riesgo si no se renegocia
        pct_target = round((at_risk / _md_target) * 100, 1) if _md_target > 0 else 0.0
        return f"🔴 {fmt_usd(at_risk)} activo + gap {fmt_usd(gap)} al 10% ({pct_target}% target)"

    def _build_md_monitor_view(md_subset, label, is_pro):
        st.markdown(f"### {label}")
        if md_subset.empty:
            st.info(f"No active {label} campaigns this period.")
            return
        prev_period = periods[-2] if len(periods) >= 2 else None
        prev = work[work["period"].astype(str) == str(prev_period)].copy() if prev_period else pd.DataFrame()
        prev_map = {}
        if not prev.empty:
            for _, r in prev.iterrows():
                prev_map[(clean(r.get("channel")), normalize_brand_id(r.get("brand_id")))] = to_number(r.get("gmv_usd"), 0)

        # GMV Trend from last 4 periods
        def _gmv_trend_for_brand(brand_id_val, channel_val):
            bid = normalize_brand_id(brand_id_val)
            trend_vals = []
            for p in periods:
                subset = work[(work["period"].astype(str) == p) & (work["channel"] == channel_val) & (work["brand_id"].apply(normalize_brand_id) == bid)]
                if not subset.empty:
                    gmv = to_number(subset.iloc[0].get("gmv_usd", 0), 0)
                    trend_vals.append(f"{gmv:.0f}k" if gmv >= 1000 else f"{gmv:.1f}")
            return " → ".join(trend_vals) if trend_vals else "-"

        rows = []
        for _, r in md_subset.iterrows():
            channel = clean(r.get("channel"))
            key = (channel, normalize_brand_id(r.get("brand_id")))
            change = _pct_change(r.get("gmv_usd"), prev_map.get(key, 0))
            gmv    = to_number(r.get("gmv_usd"), 0)
            roi    = to_number(r.get("roi"), 0)

            # Penetración real desde Current MD en vivo
            pene = _penetration_block(r.get("brand_id"), is_pro=is_pro)

            # Renegociación y recomendación basadas en penetración
            renegotiation_status, recommendation = _renegotiation_from_penetration(pene, roi, change)

            # Revenue at Risk basado en gap de penetración
            revenue_at_risk = _revenue_at_risk_md(pene, gmv)

            # Promo Intelligence (se mantiene como contexto adicional)
            pi = _md_weekly_intelligence(change, roi)

            rows.append({
                "Brand ID":        normalize_brand_id(r.get("brand_id")),
                "Brand":           names.get(normalize_brand_id(r.get("brand_id")), "-"),
                "GMV USD":         f"{gmv:,.0f}",
                "GMV Trend":       _gmv_trend_for_brand(r.get("brand_id"), channel),
                "Sales USD":       f"{to_number(r.get('sales_usd'), 0):,.0f}",
                "Orders":          fmt_number(to_number(r.get("orders"), 0)),
                "ROI":             fmt_roi2(roi),
                "WoW GMV":         fmt_signed_percent(change) if change is not None else "-",
                "Penetración MD":  pene["label"],
                "Gap al 10%":      fmt_usd(pene["gap_usd"]) if pene["gap_usd"] > 0 else "—",
                "Revenue at Risk": revenue_at_risk,
                "Renegotiación":   renegotiation_status,
                "Recomendación":   recommendation,
            })
        md_view_out = pd.DataFrame(rows)
        _render_html_table(md_view_out)

    _build_md_monitor_view(md_normal_latest, "Markdown Normal Monitor", is_pro=False)
    _build_md_monitor_view(md_pro_latest, "Markdown PRO Monitor", is_pro=True)

def page_brand_finder():
    render_header("Brand Finder", "Search and review brand information")

    df = load_growth_data()

    if df.empty:
        st.error("Growth OS sheet not found.")
        return

    # ── Pre-fill from navigation (Day Queue / Multibrand) ────────────────────
    # We can't set bf_brand_id_input directly (widget key conflict), so callers
    # set _bf_goto_brand_id and we transfer it here before the widget renders.
    if "_bf_goto_brand_id" in st.session_state and st.session_state["_bf_goto_brand_id"]:
        st.session_state["bf_brand_id_input"] = st.session_state.pop("_bf_goto_brand_id")

    brand_id_input = st.text_input("Search Brand ID", key="bf_brand_id_input")
    brand_id = normalize_brand_id(brand_id_input)

    if not brand_id_input:
        st.info("Type or paste a Brand ID to load the full brand profile. Example: AR65184 - Multistorefull")
        st.markdown(render_priority_overview_html(), unsafe_allow_html=True)
        render_priority_overview_controls()
        return

    id_col = get_id_column_name(df)
    if not id_col:
        st.error("ID column not found in Growth OS sheet.")
        st.caption("Detected columns: " + ", ".join([str(c) for c in df.columns[:25]]))
        return

    result = df[df[id_col].apply(normalize_brand_id) == brand_id]

    if result.empty:
        # Fallback: buscar en Asignacion Junio → construir fila sintética enriquecida
        aj = load_asignacion_activa()
        aj_match = aj[aj["brand_id"] == brand_id] if not aj.empty else pd.DataFrame()
        if not aj_match.empty:
            aj_row = aj_match.iloc[0]
            bname = str(aj_row["brand_name"])

            # Badge de marca nueva
            new_badge_html = (
                "<div style='display:inline-flex;align-items:center;gap:8px;background:linear-gradient(135deg,#FF8A3D22,#FF8A3D44);"
                "border:1.5px solid #FF7124;border-radius:8px;padding:6px 14px;margin-bottom:12px;font-size:13px;font-weight:700;color:#FF7124;'>"
                "🆕 Marca nueva · Asignacion Junio · Sin ficha en Growth OS todavía</div>"
            )
            st.markdown(new_badge_html, unsafe_allow_html=True)

            # Category from Detalle CABA (by brand_id)
            cat_syn = ""
            try:
                _dc_bf = load_detalle_caba()
                if not _dc_bf.empty and "brand_id" in _dc_bf.columns:
                    cat_col_bf = next((c for c in _dc_bf.columns if "categor" in c), None)
                    if cat_col_bf:
                        _dc_match = _dc_bf[_dc_bf["brand_id"] == brand_id]
                        if not _dc_match.empty:
                            _cats = _dc_match[cat_col_bf].dropna().astype(str)
                            cat_syn = _cats.mode().iloc[0] if not _cats.mode().empty else (_cats.iloc[0] if not _cats.empty else "")
            except Exception:
                pass

            # CVR from CVR% sheet (by brand name)
            cvr_syn = ""
            try:
                _cvr_bf = load_cvr_data()
                _cvr_val = _cvr_bf.get(bname.strip().lower(), None)
                if _cvr_val and isinstance(_cvr_val, float) and _cvr_val > 0:
                    cvr_syn = str(round(_cvr_val * 100, 1))
            except Exception:
                pass

            # Portfolio ranking: compare GMV from Detalle CABA vs Growth OS brands
            ranking_syn = ""
            try:
                _dc_bf2 = load_detalle_caba()
                _gmv_col_bf = "_gmv" if "_gmv" in _dc_bf2.columns else next((c for c in _dc_bf2.columns if "gmv" in c), None)
                if not _dc_bf2.empty and _gmv_col_bf and "brand_id" in _dc_bf2.columns:
                    _dc_brand = _dc_bf2[_dc_bf2["brand_id"] == brand_id]
                    _gmv_syn = float(pd.to_numeric(_dc_brand[_gmv_col_bf], errors="coerce").fillna(0).sum()) if not _dc_brand.empty else 0.0
                    if _gmv_syn > 0:
                        _gos_df = load_growth_data()
                        _gos_id_c = get_id_column_name(_gos_df) if not _gos_df.empty else None
                        _gos_gmv_c = next((c for c in (_gos_df.columns if not _gos_df.empty else []) if c in ["last gmv ars", "gmv ars"]), None)
                        if _gos_id_c and _gos_gmv_c:
                            _higher = sum(1 for _, _r in _gos_df.iterrows() if to_number(_r.get(_gos_gmv_c, 0), 0) > _gmv_syn)
                            ranking_syn = f"#{_higher + 1}"
            except Exception:
                pass

            # Build synthetic row
            synthetic = {
                "id":              brand_id,
                "name":            bname,
                "country":         PORTFOLIO_COUNTRY or "-",
                "ltor tier":       "No Priorizado",
                "churn":           "",
                "churn status":    "",
                "category":        cat_syn,
                "ads":             "",
                "ads bookings":    "",
                "ads roi":         "",
                "md":              "",
                "md status":       "",
                "md bookings":     "",
                "md roi":          "",
                "last gmv ars":    "",
                "gmv ars":         "",
                "last aov ars":    "",
                "aov ars":         "",
                "cr %":            cvr_syn,
                "conversion rate": cvr_syn,
                "pro users %":     "",
                "pro %":           "",
                "comm. rate":      "",
                "manager":         "",
                "assistant":       "",
                "email":           "",
                "contact number":  "",
                "ranking":         ranking_syn,
                "comments":        "",
            }
            synthetic_df = pd.DataFrame([synthetic])
            synthetic_df.columns = [normalize(c) for c in synthetic_df.columns]
            result = synthetic_df
        else:
            st.error("Brand not found.")
            st.caption(f"Typed ID: {brand_id}")
            return

    row = result.iloc[0]
    name = render_brand_profile(row, brand_id)

    _render_followup_form(row, brand_id, name)


@st.fragment
def _render_followup_form(row, brand_id, name):
    st.markdown("<div class='form-card-static'>", unsafe_allow_html=True)
    st.markdown("<div class='wide-info-title'>Comments History</div>", unsafe_allow_html=True)

    st.markdown("<div class='wide-info-title' style='margin-top:20px;'>Follow-up Update</div>", unsafe_allow_html=True)

    fu1, fu2 = st.columns(2)
    with fu1:
        contact_channel = st.selectbox(
            "Contact Channel",
            ["Call", "WhatsApp", "Email", "Meet", "Other"],
            index=0,
            key=f"comment_channel_{brand_id}"
        )
    with fu2:
        opportunity_status = st.selectbox(
            "Status",
            [
                "📅 Campaign Follow Up",
                "📅 Campaign Negotiation",
                "📅 Contractual Changes",
                "Deal Closed 🏆",
                "── No Answer ──",
                "📵 No Answer / Bad Number",
                "⏰ No Answer / No time — Call me later",
                "🙅 No Answer / Not interested in meeting",
            ],
            index=0,
            key=f"comment_status_{brand_id}"
        )
    template_type = "None"

    # ── Detectar si es un No Answer para simplificar el formulario ────────────
    _is_no_answer = opportunity_status.startswith("📵") or opportunity_status.startswith("⏰") or opportunity_status.startswith("🙅")
    _is_separator  = opportunity_status == "── No Answer ──"

    # ── Transcripción / nota de contacto ─────────────────────────────────────
    transcript_label = "📋 Transcripción de la llamada" if contact_channel == "Call" else "📝 Nota del contacto (WhatsApp / Email / Meet)"
    transcript_placeholder = (
        "Pegá acá el resumen que te da Claude — se guarda tal cual, sin análisis ni cambios automáticos."
        if contact_channel == "Call"
        else "Escribí o pegá el resumen del contacto — se guarda tal cual."
    )
    call_transcript = st.text_area(
        transcript_label,
        placeholder=transcript_placeholder,
        height=160,
        key=f"call_transcript_{brand_id}",
    )

    # ── Transcripción / resumen: se pega tal cual el resumen ya elaborado por
    # Claude — no hay análisis local en vivo ni auto-detección de palancas.
    # Sí se parsea el bloque "Calendario:" embebido para prellenar el cuadro
    # de accionable más abajo (fecha, canal, prioridad, tema). ────────────────
    transcript_analysis = None
    _claude_parsed = _parse_claude_note_fields(call_transcript)

    new_comment = ""  # kept for calendar default_notes compatibility below

    comment_auto = ""
    followup_type = ""
    rejection_reason = ""
    negotiation_type = ""
    negotiation_ads_ars = 0.0
    negotiation_md_discount = ""
    negotiation_action = "No commercial change"
    commercial_action = "No commercial change"
    ad_budget_input = 0.0
    md_discount_input = ""
    event_required = False
    event_data = None

    # ── Agendamiento automático ────────────────────────────────────────────────
    # No Answer → próximo contacto en 7 días
    # Contestó (cualquier otro status) → en 14 días
    # El usuario puede cambiar la fecha manualmente en el campo que aparece abajo.
    _auto_days = 7 if _is_no_answer else 14
    _auto_next_date = date.today() + timedelta(days=_auto_days)
    _auto_label = (
        f"📵 No contestó — próximo intento en 7 días ({_auto_next_date.strftime('%d/%m/%Y')})"
        if _is_no_answer else
        f"✅ Contacto registrado — próximo seguimiento en 14 días ({_auto_next_date.strftime('%d/%m/%Y')})"
    )
    st.markdown(
        f"<div style='background:rgba(27,63,139,0.03);border-radius:8px;padding:8px 14px;"
        f"margin:10px 0 6px 0;font-size:12px;color:#6B7280;'>📅 {_auto_label}</div>",
        unsafe_allow_html=True
    )
    # Manual override: allow changing the scheduled date
    _override_key = f"_override_date_{brand_id}"
    if _override_key not in st.session_state:
        st.session_state[_override_key] = False
    if st.checkbox("Cambiar fecha de seguimiento", key=f"override_chk_{brand_id}", value=st.session_state[_override_key]):
        st.session_state[_override_key] = True
        _auto_next_date = st.date_input(
            "Próximo contacto",
            value=_auto_next_date,
            min_value=date.today(),
            key=f"manual_next_date_{brand_id}"
        )
    else:
        st.session_state[_override_key] = False

    def _render_calendar_fields(suffix, default_task="Follow-up", default_notes="", parsed_cal=None):
        """
        Cuadro de accionable: se auto-completa con lo que Claude dejó en el
        bloque '📅 Calendario:' de la transcripción pegada (fecha, canal,
        prioridad, tema). Por default se muestra solo como resumen — el
        formulario editable (Date/Time/Channel/Priority/Notes) solo aparece
        si se tilda "✏️ Editar accionable".
        """
        parsed_cal = parsed_cal or {}

        # ── Color por tipo de task ────────────────────────────────────────────
        task_colors = {
            "Campaign Follow Up":  "#1B3F8B",
            "Campaign Negotiation": "#FF7124",
            "Contractual Changes":  "#1D9E75",
        }
        task_color = next((v for k, v in task_colors.items() if k.lower() in default_task.lower()), "#1B3F8B")

        # ── Defaults: primero lo que Claude dejó parseado, si no hay, la
        # lógica automática de siempre (7/14 días) y el canal del contacto. ──
        _default_date = _auto_next_date
        if parsed_cal.get("cal_fecha"):
            try:
                _default_date = datetime.strptime(parsed_cal["cal_fecha"], "%Y-%m-%d").date()
            except Exception:
                pass
        _channel_options = ["Call", "WhatsApp", "Email", "Meet", "Other"]
        _default_channel = parsed_cal.get("cal_canal") if parsed_cal.get("cal_canal") in _channel_options else contact_channel
        if _default_channel not in _channel_options:
            _default_channel = "Call"
        _priority_options = ["High", "Mid", "Low"]
        _default_priority = parsed_cal.get("cal_prioridad") if parsed_cal.get("cal_prioridad") in _priority_options else "Mid"
        _default_tema = parsed_cal.get("cal_tema") or default_notes

        st.markdown(f"""
        <div style="
            border-left: 4px solid {task_color};
            background: rgba(27,63,131,0.03);
            border-radius: 0 10px 10px 0;
            padding: 14px 18px 12px 16px;
            margin: 16px 0 10px 0;
        ">
            <div style="font-size:11px;font-weight:700;letter-spacing:.08em;color:{task_color};margin-bottom:6px;">
                📅 ACCIONABLE — {default_task}
            </div>
            <div style="display:grid;grid-template-columns:1fr 1fr 1fr 2fr;gap:12px;font-size:13px;">
                <div><div style="font-size:10px;color:#6B7280;">FECHA</div><div style="font-weight:700;">{_default_date.strftime('%d/%m/%Y')}</div></div>
                <div><div style="font-size:10px;color:#6B7280;">CANAL</div><div style="font-weight:600;">{html.escape(_default_channel)}</div></div>
                <div><div style="font-size:10px;color:#6B7280;">PRIORIDAD</div><div style="font-weight:600;">{html.escape(_default_priority)}</div></div>
                <div><div style="font-size:10px;color:#6B7280;">TEMA</div><div style="font-weight:600;">{html.escape(_default_tema) if _default_tema else '—'}</div></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if not st.checkbox("✏️ Editar accionable", key=f"edit_accionable_{suffix}_{brand_id}"):
            return {
                "date": _default_date,
                "time": time(9, 0).strftime("%I:%M %p").lstrip("0"),
                "id": brand_id,
                "name": name,
                "task": default_task,
                "channel": _default_channel,
                "priority": _default_priority,
                "status": default_task,
                "notes": _default_tema.strip(),
            }

        e1, e2, e3 = st.columns([1.2, 1, 1.5])
        with e1:
            _event_date = st.date_input("Date", value=_default_date, key=f"event_date_{suffix}_{brand_id}")
            _event_time = st.time_input("Time", value=time(9, 0), key=f"event_time_{suffix}_{brand_id}")
        with e2:
            _event_channel = st.selectbox(
                "Channel",
                _channel_options,
                index=_channel_options.index(_default_channel),
                key=f"event_channel_{suffix}_{brand_id}"
            )
            _event_priority = st.selectbox(
                "Priority",
                _priority_options,
                index=_priority_options.index(_default_priority),
                key=f"event_priority_{suffix}_{brand_id}"
            )
            _event_status = st.selectbox(
                "Task Status",
                ["Campaign Follow Up", "Campaign Negotiation", "Contractual Changes"],
                index=["Campaign Follow Up", "Campaign Negotiation", "Contractual Changes"].index(default_task)
                    if default_task in ["Campaign Follow Up", "Campaign Negotiation", "Contractual Changes"] else 0,
                key=f"event_status_{suffix}_{brand_id}"
            )
        with e3:
            _event_notes = st.text_area(
                "Notes",
                value=_default_tema,
                placeholder="Próximos pasos, acuerdos pendientes...",
                height=112,
                key=f"event_notes_{suffix}_{brand_id}"
            )
        return {
            "date": _event_date,
            "time": _event_time.strftime("%I:%M %p").lstrip("0"),
            "id": brand_id,
            "name": name,
            "task": default_task,
            "channel": _event_channel,
            "priority": _event_priority,
            "status": _event_status,
            "notes": _event_notes.strip(),
        }

    def _render_markdown_activation_fields(prefix):
        top_products = get_definitive_top_products_for_brand(brand_id)
        product_options = []
        for i, product in enumerate(top_products[:3], start=1):
            pname = clean(product.get("name"), "-")
            if pname not in ["", "-"]:
                product_options.append(pname)
        while len(product_options) < 3:
            product_options.append(f"Top Product {len(product_options) + 1}")
        product_options.append("Custom")

        c1, c2, c3 = st.columns(3)
        with c1:
            full_discount = st.selectbox(
                "Full Users Discount / Promo Type",
                ["5%", "10%", "15%", "20%", "25%", "30%", "35%", "40%", "2x1", "Free Shipping", "Storewide Discount"],
                index=3,
                key=f"{prefix}_full_discount_{brand_id}"
            )
            storewide_discount = ""
            if full_discount == "Storewide Discount":
                storewide_discount = st.selectbox(
                    "Storewide Discount %",
                    ["5%", "10%", "15%", "20%", "25%", "30%", "35%", "40%", "Custom"],
                    index=1,
                    key=f"{prefix}_storewide_discount_{brand_id}"
                )
                if storewide_discount == "Custom":
                    storewide_discount = st.text_input("Custom Storewide Discount", value="", placeholder="Ej: 12%, 18%", key=f"{prefix}_storewide_custom_{brand_id}")
        with c2:
            pro_extra = st.selectbox(
                "Extra PRO Discount",
                ["+0%", "+5%", "+10%", "+15%"],
                index=0,
                key=f"{prefix}_pro_extra_{brand_id}"
            )
        with c3:
            if full_discount == "Storewide Discount":
                product_name = "Toda la tienda"
                st.info("Storewide applies to the full store.")
            else:
                product_choice = st.selectbox(
                    "Product",
                    product_options,
                    index=0,
                    key=f"{prefix}_product_{brand_id}"
                )
                if product_choice == "Custom":
                    product_name = st.text_input("Custom Product Name", value="", placeholder="Write product name", key=f"{prefix}_product_custom_{brand_id}").strip()
                else:
                    product_name = product_choice

        if full_discount == "Storewide Discount":
            promo_name = f"Storewide {clean(storewide_discount, '-') } {pro_extra} PRO · Toda la tienda"
        elif full_discount == "Free Shipping":
            promo_name = f"Free Shipping {pro_extra} PRO · {clean(product_name, 'Custom Product')}"
        elif full_discount == "2x1":
            promo_name = f"2x1 {pro_extra} PRO · {clean(product_name, 'Custom Product')}"
        else:
            promo_name = f"{full_discount} {pro_extra} PRO · {clean(product_name, 'Custom Product')}"
        st.caption(f"Promo name: {promo_name}")
        return promo_name

    if _is_no_answer or _is_separator:
        # No Answer: skip calendar, skip commercial fields, skip validations
        # The auto-schedule logic above already handled date calculation.
        comment_auto = opportunity_status  # save the exact No Answer label as the note
    elif opportunity_status == "📅 Campaign Follow Up":
        event_required = True
        event_data = _render_calendar_fields("camp_followup", default_task="Campaign Follow Up", parsed_cal=_claude_parsed)
        comment_auto = "📅 Campaign Follow Up"

    elif opportunity_status == "📅 Campaign Negotiation":
        event_required = True
        event_data = _render_calendar_fields("camp_negotiation", default_task="Campaign Negotiation", parsed_cal=_claude_parsed)
        comment_auto = "📅 Campaign Negotiation"

    elif opportunity_status == "📅 Contractual Changes":
        event_required = True
        event_data = _render_calendar_fields("contractual", default_task="Contractual Changes", parsed_cal=_claude_parsed)
        comment_auto = "📅 Contractual Changes"

    elif opportunity_status == "Deal Closed 🏆":
        st.markdown(
            "<div class='wide-info-title' style='margin-top:14px;'>🏆 Negociación cerrada — ¿qué palanca se activó?</div>",
            unsafe_allow_html=True
        )
        commercial_action_type = st.radio(
            "Tipo de cierre",
            ["Ads", "Markdown"],
            index=0,
            horizontal=True,
            key=f"deal_closed_type_{brand_id}"
        )
        if commercial_action_type == "Ads":
            ad_budget_input = st.number_input(
                "Valor de campaña semanal cerrado (ARS)",
                min_value=0.0,
                value=0.0,
                step=1000.0,
                format="%.0f",
                key=f"deal_closed_ads_ars_{brand_id}"
            )
            commercial_action = f"Ads · {fmt_ars(ad_budget_input)}/semana"
        else:
            md_discount_input = _render_markdown_activation_fields("deal_closed_md")
            commercial_action = f"Markdown · {md_discount_input}"
        comment_auto = f"🏆 Deal Closed — {commercial_action}"

    # ── Auto-router desde transcripción: desactivado. El resumen ya viene
    # armado por Claude con las palancas identificadas — no hace falta
    # re-detectarlas acá con keywords. ─────────────────────────────────────────

    if st.button("Save Follow-up"):
        # Build final comment: se guarda tal cual el contenido del campo de
        # transcripción (el resumen ya armado por Claude) — no se regenera nada.
        if _is_no_answer or _is_separator:
            # No Answer: comment is the status label itself — no transcript needed
            final_comment = opportunity_status
        elif call_transcript.strip():
            final_comment = call_transcript.strip()
        else:
            final_comment = comment_auto.strip()

        # Skip validations for No Answer statuses
        if not (_is_no_answer or _is_separator):
            if opportunity_status in ["Negotiation ⏳", "Deal Closed 🏆"] and not final_comment:
                st.warning("Pegá la transcripción o escribí una nota antes de guardar este status.")
                st.stop()
            if opportunity_status == "Follow-up ✅" and followup_type in ["Cambios contractuales", "Revisiones específicas"] and not call_transcript.strip():
                st.warning("Pegá la transcripción antes de guardar este tipo de follow-up.")
                st.stop()
            if event_required and (not event_data or not event_data.get("task")):
                st.warning("Write a task before saving the Weekly Calendar event.")
                st.stop()

        comment_commercial_action = commercial_action
        if commercial_action == "No commercial change" and opportunity_status == "Negotiation ⏳":
            comment_commercial_action = negotiation_action
        if opportunity_status == "Rejected ❌":
            comment_commercial_action = f"Rejected · {rejection_reason} ❌"

        save_comment_csv(
            brand_id,
            name,
            final_comment,
            contact_channel=contact_channel,
            opportunity_status=opportunity_status,
            commercial_action=comment_commercial_action,
        )

        # ── Histórico para analytics (sentimiento, palancas recurrentes, etc.) ────
        save_call_history_row(
            brand_id,
            name,
            final_comment,
            contact_channel=contact_channel,
            opportunity_status=opportunity_status,
        )

        # ── Backup async: corre en background, no bloquea el guardado principal ──
        import threading as _threading
        _threading.Thread(target=make_backup, args=(EXCEL_FILE,), daemon=True).start()

        # ── Apertura única del Excel para TODO el flujo de guardado, incluido
        # Call Detail — evita abrir el workbook completo dos veces en el mismo click. ──
        commercial_ok, commercial_msg = True, "No commercial change selected."
        tracker_ok, tracker_msg = True, "No commercial action, negotiation or rejection to track."
        st.toast("Guardando...", icon="💾")
        _wb_save = openpyxl.load_workbook(EXCEL_FILE)

        # ── Evaluación IA de transcripción en Call Detail (silenciosa) ─────────
        if contact_channel == "Call" and call_transcript.strip():
            try:
                evaluate_and_save_call_detail(
                    transcript=call_transcript.strip(),
                    brand_id=brand_id,
                    brand_name=name,
                    farmer_email="sabas.ramirez@rappi.com",
                    call_date=date.today(),
                    wb=_wb_save,
                )
            except Exception:
                pass  # Falla silenciosa — el follow-up ya se guardó

            # ── Detección automática de objeciones para Role Play Trainer ──────
            try:
                detect_and_save_objection_from_transcript(call_transcript.strip())
            except Exception:
                pass  # Falla silenciosa — no bloquea el guardado del follow-up

        ok, msg = _update_agenda_notes_inner(_wb_save, brand_id, final_comment, append=True)
        follow_ok, follow_msg = _update_contact_followup_fields_inner(
            _wb_save,
            brand_id,
            contact_channel=contact_channel,
            opportunity_status=opportunity_status,
            comment_text=final_comment,
        )

        event_ok, event_msg = True, "No calendar event requested."
        if event_required and event_data:
            event_ok, event_msg = _add_event_to_agenda_inner(_wb_save, event_data)

        if opportunity_status == "Deal Closed 🏆" and commercial_action != "No commercial change":
            updates = {}
            if "Ads" in commercial_action:
                updates["ads"] = "Active 🚀"
                updates["ads_bookings"] = float(ad_budget_input or 0)
            if "Markdown" in commercial_action:
                updates["md"] = "Active 🚀"
                updates["md_bookings"] = normalize_markdown_discount(md_discount_input)
            if updates:
                commercial_ok, commercial_msg, _, _, _ = _update_brand_in_excel_inner(_wb_save, brand_id, updates)
                if commercial_ok and "comments" in updates:
                    _update_agenda_notes_inner(_wb_save, brand_id, updates["comments"], append=False)
            tracker_ok, tracker_msg = save_acquisition_tracker_event(
                brand_id,
                name,
                commercial_action,
                ads_budget_ars=float(ad_budget_input or 0),
                md_discount=normalize_markdown_discount(md_discount_input),
                opportunity_status=opportunity_status,
                comment=final_comment,
                pipeline_stage="Closed",
                negotiation_type="",
            )
        elif opportunity_status == "Negotiation ⏳":
            tracker_ok, tracker_msg = save_acquisition_tracker_event(
                brand_id,
                name,
                negotiation_action,
                ads_budget_ars=float(negotiation_ads_ars or 0),
                md_discount=normalize_markdown_discount(negotiation_md_discount),
                opportunity_status=opportunity_status,
                comment=final_comment,
                pipeline_stage="Negotiation",
                negotiation_type=negotiation_type,
            )
        elif opportunity_status == "Rejected ❌":
            tracker_ok, tracker_msg = save_acquisition_tracker_event(
                brand_id,
                name,
                f"Rejected · {rejection_reason} ❌",
                ads_budget_ars=0,
                md_discount="",
                opportunity_status=opportunity_status,
                comment=final_comment,
                pipeline_stage="Rejected",
                negotiation_type="",
                rejection_reason=rejection_reason,
            )

        # ── Auto-calendar: schedule next contact based on 14d / 7d rule ─────────
        _auto_event_data = {
            "date":     _auto_next_date,
            "time":     "09:00 AM",
            "id":       brand_id,
            "name":     name,
            "task":     "No Answer — Retry" if (_is_no_answer or _is_separator) else "Follow-up",
            "channel":  contact_channel,
            "priority": "High" if (_is_no_answer or _is_separator) else "Mid",
            "status":   "Campaign Follow Up",
            "notes":    f"Auto-agendado desde Brand Finder · {final_comment[:80] if final_comment else opportunity_status}",
        }
        # Only auto-schedule if no manual calendar event was already added
        auto_event_ok = True
        auto_event_msg = ""
        if not event_required:
            auto_event_ok, auto_event_msg = _add_event_to_agenda_inner(_wb_save, _auto_event_data)

        # ── Guardado único + cierre + invalidación selectiva de caché ───────────
        # Solo Growth OS y Agenda se modificaron en este flujo — invalidamos
        # exclusivamente esas dos funciones cacheadas en vez de st.cache_data.clear()
        # completo, que forzaba releer las 39 funciones cacheadas (incluye Detalle CABA,
        # Asignación Junio, CVR%, Current GMV/Ads/MD/Churn, etc. que NO cambiaron aquí).
        try:
            _wb_save.calculation.fullCalcOnLoad = True
            _wb_save.calculation.forceFullCalc = True
        except Exception:
            pass
        try:
            _wb_save.save(EXCEL_FILE)
            _wb_save.close()
            load_growth_data.clear()
            load_agenda_data.clear()
            _read_growth_summary_values.clear()
        except PermissionError:
            _wb_save.close()
            ok = False
            msg = f"'{os.path.basename(EXCEL_FILE)}' está abierto en Excel. Cerralo y reintentá el guardado."

        if ok and follow_ok and event_ok and commercial_ok and tracker_ok:
            _days_label = "7 días (No Answer)" if (_is_no_answer or _is_separator) else "14 días"
            success_msg = f"Follow-up guardado · próximo contacto agendado en {_days_label} ({_auto_next_date.strftime('%d/%m/%Y')})."
            if event_required:
                success_msg += " Evento de calendario manual también añadido."
            st.success(success_msg)
            st.rerun()
        else:
            st.warning(f"Saved with warnings. Agenda: {msg}. Follow-up: {follow_msg}. Event: {event_msg}. Commercial: {commercial_msg}. Tracker: {tracker_msg}.")

    st.markdown("</div>", unsafe_allow_html=True)


# =========================
# WEEKLY CALENDAR
# =========================


def mark_agenda_row_done(excel_path, excel_row):
    """
    Marks one Agenda row as Done. We do not delete the row, so history stays in Excel.
    Weekly Calendar hides rows with status Done/Completed.
    """
    wb = openpyxl.load_workbook(excel_path)

    if AGENDA_SHEET not in wb.sheetnames:
        wb.close()
        return False, "Agenda sheet not found."

    ws = wb[AGENDA_SHEET]

    header_row = None
    headers = {}

    for r in range(1, min(25, ws.max_row) + 1):
        row_headers = {}
        for c in range(1, ws.max_column + 1):
            value = ws.cell(r, c).value
            if value is not None:
                row_headers[normalize(value)] = c

        if "id" in row_headers and "name" in row_headers and "status" in row_headers:
            header_row = r
            headers = row_headers
            break

    if not header_row:
        wb.close()
        return False, "Agenda headers not found."

    status_col = headers.get("status")

    try:
        excel_row = int(excel_row)
    except Exception:
        wb.close()
        return False, "Invalid Excel row."

    if excel_row <= header_row or excel_row > ws.max_row:
        wb.close()
        return False, "Agenda row out of range."

    ws.cell(excel_row, status_col).value = "Done"

    try:
        wb.save(excel_path)
        wb.close()
        # Invalidate only Agenda cache — Weekly Calendar reflects Done status
        # immediately without forcing every other cached sheet to re-read.
        try:
            load_agenda_data.clear()
        except Exception:
            st.cache_data.clear()  # fallback de seguridad
        return True, "Task marked as Done."
    except PermissionError:
        wb.close()
        return False, f"'{os.path.basename(EXCEL_FILE)}' está abierto en Excel. Cerralo para poder marcar la tarea como Done."



def parse_agenda_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    except Exception:
        return None


def parse_agenda_time(value):
    if isinstance(value, datetime):
        return value.strftime("%I:%M %p").lstrip("0")
    if isinstance(value, time):
        return value.strftime("%I:%M %p").lstrip("0")
    if value is None or pd.isna(value):
        return "-"
    return str(value)


def priority_class(priority):
    p = normalize(priority)
    if "high" in p:
        return "high"
    if "low" in p:
        return "low"
    if "complete" in p or "done" in p:
        return "done"
    return "mid"


def priority_dot(priority):
    pclass = priority_class(priority)
    if pclass == "high":
        return "🟠"
    if pclass == "mid":
        return "🟡"
    if pclass == "low":
        return "🔵"
    return "⚪"


def page_weekly_calendar():
    render_header("Weekly Calendar", "Your scheduled contacts and follow-up tasks")

    agenda = load_agenda_data()
    today = date.today()

    # ── Semana: lunes de la semana actual ─────────────────────────────────────
    week_start = today - timedelta(days=today.weekday())  # lunes
    days = [week_start + timedelta(days=i) for i in range(7)]

    # ── Task colors ───────────────────────────────────────────────────────────
    TASK_COLORS = {
        "campaign follow up":   {"bg": "rgba(59,72,131,0.85)",  "border": "#1B3F8B", "text": "#FFFFFF"},
        "campaign negotiation": {"bg": "rgba(255,113,36,0.85)", "border": "#FF7124", "text": "#FFFFFF"},
        "contractual changes":  {"bg": "rgba(29,158,117,0.85)", "border": "#1D9E75", "text": "#FFFFFF"},
    }
    PRIORITY_COLORS = {"high": "#FF4D2E", "mid": "#FF7124", "low": "#1B3F8B"}

    def _task_color(task_str, priority_str):
        tl = task_str.lower()
        for key, val in TASK_COLORS.items():
            if key in tl:
                return val
        p = priority_str.lower()
        c = PRIORITY_COLORS.get(p, "#1B3F8B")
        return {"bg": c, "border": c, "text": "#FFFFFF"}

    st.markdown("### This Week")

    # ── Contadores ────────────────────────────────────────────────────────────
    if not agenda.empty:
        agenda["_parsed_date"] = get_col(agenda, ["date", "data"]).apply(parse_agenda_date)
        agenda["_time_display"] = get_col(agenda, ["time"]).apply(parse_agenda_time)
        status_text = get_col(agenda, ["status"]).astype(str).str.strip().str.lower()
        done_mask = (
            status_text.eq("done")
            | status_text.eq("completed")
            | status_text.eq("complete")
            | status_text.str.contains("done", na=False)
            | status_text.str.contains("completed", na=False)
        )
        active_agenda = agenda[~done_mask].copy()
        active_agenda["_sort_date"] = active_agenda["_parsed_date"].apply(lambda x: x or date.max)
        active_agenda = active_agenda.sort_values(by=["_sort_date", "_time_display"], ascending=True)
    else:
        active_agenda = pd.DataFrame()

    _today_count    = int((active_agenda["_parsed_date"] == today).sum()) if not active_agenda.empty else 0
    _tomorrow_count = int((active_agenda["_parsed_date"] == today + timedelta(days=1)).sum()) if not active_agenda.empty else 0
    _overdue_count  = int((active_agenda["_parsed_date"] < today).sum()) if not active_agenda.empty else 0
    _today_color    = "#FF4D2E" if _today_count >= 5 else ("#D95A10" if _today_count >= 3 else "#7ED321")
    _tmrw_color     = "#D95A10" if _tomorrow_count >= 5 else "#6B7280"
    _overdue_color  = "#FF4D2E" if _overdue_count > 0 else "#6B7280"

    st.markdown(
        f'''<div style="display:flex;gap:12px;margin-bottom:16px;flex-wrap:wrap;">
        <div style="background:rgba(27,63,139,0.03);border:1px solid rgba(0,0,0,0.07);border-radius:8px;padding:8px 16px;font-size:13px;">
            📅 <b>Hoy:</b> <span style="color:{_today_color};font-weight:700;">{_today_count} tareas</span>
        </div>
        <div style="background:rgba(27,63,139,0.03);border:1px solid rgba(0,0,0,0.07);border-radius:8px;padding:8px 16px;font-size:13px;">
            📆 <b>Mañana:</b> <span style="color:{_tmrw_color};font-weight:700;">{_tomorrow_count} tareas</span>
        </div>
        <div style="background:rgba(27,63,139,0.03);border:1px solid rgba(0,0,0,0.07);border-radius:8px;padding:8px 16px;font-size:13px;">
            ⚠️ <b>Vencidas:</b> <span style="color:{_overdue_color};font-weight:700;">{_overdue_count} tareas</span>
        </div>
        </div>''',
        unsafe_allow_html=True,
    )

    # ── Leyenda ───────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="display:flex;gap:16px;align-items:center;margin-bottom:18px;flex-wrap:wrap;">
        <span style="font-size:11px;font-weight:700;letter-spacing:.06em;color:#6B7280;">TASK TYPES</span>
        <span style="font-size:12px;color:#1B3F8B;">● Campaign Follow Up</span>
        <span style="font-size:12px;color:#FF7124;">● Campaign Negotiation</span>
        <span style="font-size:12px;color:#1D9E75;">● Contractual Changes</span>
        <span style="font-size:12px;color:#FF4D2E;margin-left:12px;">● Overdue</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Cabecera de días ──────────────────────────────────────────────────────
    header_html = '<div style="display:grid;grid-template-columns:52px repeat(7,1fr);gap:0;margin-bottom:0;">'
    header_html += '<div></div>'  # spacer para columna de horas
    for d in days:
        ddate = d if isinstance(d, date) else d.date()
        is_today = ddate == today
        day_count = int((active_agenda["_parsed_date"] == ddate).sum()) if not active_agenda.empty else 0
        bg = "rgba(27,63,139,0.12)" if is_today else "rgba(27,63,139,0.03)"
        border_b = "2px solid #1B3F8B" if is_today else "1px solid rgba(255,255,255,0.95)"
        num_color = "#FFFFFF" if is_today else "#1A1A2E"
        num_bg = "#1B3F8B" if is_today else "transparent"
        count_color = "#7ED321" if day_count > 0 else "#6B7280"
        header_html += f'''
        <div style="background:{bg};border-bottom:{border_b};padding:8px 4px 8px 4px;text-align:center;border-right:1px solid rgba(255,255,255,0.92);">
            <div style="font-size:10px;font-weight:700;letter-spacing:.06em;color:#6B7280;">{d.strftime("%a").upper()}</div>
            <div style="display:inline-block;background:{num_bg};border-radius:50%;width:28px;height:28px;line-height:28px;font-size:16px;font-weight:700;color:{num_color};margin:2px 0;">{d.strftime("%d")}</div>
            <div style="font-size:10px;color:{count_color};font-weight:600;">{day_count} task{"s" if day_count != 1 else ""}</div>
        </div>'''
    header_html += '</div>'
    st.markdown(header_html, unsafe_allow_html=True)

    # ── Malla horaria tipo Google Calendar ───────────────────────────────────
    _render_calendar_events(active_agenda, today, days, _task_color, PRIORITY_COLORS)


@st.fragment
def _render_calendar_events(active_agenda, today, days, _task_color, PRIORITY_COLORS):
    done_rows = st.session_state.setdefault("_wc_done_rows", set())

    # ── Parsear hora numérica de cada evento ──────────────────────────────────
    def _parse_hour(time_val):
        """Devuelve float (ej: 9.5 para 9:30). None si no se puede parsear."""
        if isinstance(time_val, (time, datetime)):
            t = time_val if isinstance(time_val, time) else time_val.time()
            return t.hour + t.minute / 60
        if time_val is None:
            return None
        try:
            if pd.isna(time_val):
                return None
        except Exception:
            pass
        s = str(time_val).strip()
        # "9:00 AM", "09:00", "9:30 PM"
        for fmt in ("%I:%M %p", "%H:%M", "%I %p"):
            try:
                t = datetime.strptime(s, fmt)
                return t.hour + t.minute / 60
            except Exception:
                pass
        # fallback: primer número
        m = re.search(r"(\d+)", s)
        if m:
            return float(m.group(1))
        return None

    if not active_agenda.empty:
        active_agenda = active_agenda.copy()
        active_agenda["_hour"] = get_col(active_agenda, ["time"]).apply(_parse_hour)

    # ── Configuración de la malla horaria ─────────────────────────────────────
    HOUR_START = 7     # 7 AM
    HOUR_END   = 22    # 10 PM
    SLOT_HEIGHT = 48   # px por hora

    # Construir index de eventos por (día, hora_slot)
    # Un slot es la hora entera (7, 8, …, 21)
    def _events_for_day(ddate):
        if active_agenda.empty:
            return []
        mask = active_agenda["_parsed_date"] == ddate
        return active_agenda[mask].to_dict("records")

    # ── HTML de la malla completa ─────────────────────────────────────────────
    # Estructura: columna de horas + 7 columnas de días, filas = horas
    # Usamos position:relative en cada celda para superponer tarjetas.

    hour_labels_html = ""
    for h in range(HOUR_START, HOUR_END):
        top = (h - HOUR_START) * SLOT_HEIGHT
        if h == 0:
            label = "12:00 AM"
        elif h < 12:
            label = f"{h}:00 AM"
        elif h == 12:
            label = "12:00 PM"
        else:
            label = f"{h-12}:00 PM"
        hour_labels_html += f'<div style="position:absolute;top:{top}px;left:0;width:48px;font-size:10px;color:#6B7280;font-weight:600;text-align:right;padding-right:6px;line-height:1;">{label}</div>'

    total_height = (HOUR_END - HOUR_START) * SLOT_HEIGHT

    # ── Líneas horizontales ───────────────────────────────────────────────────
    grid_lines_html = ""
    for h in range(HOUR_START, HOUR_END + 1):
        top = (h - HOUR_START) * SLOT_HEIGHT
        color = "rgba(0,0,0,0.07)" if h % 2 == 0 else "rgba(27,63,139,0.03)"
        grid_lines_html += f'<div class="grid-line" style="top:{top}px;background:{color};"></div>'

    # ── Línea de "ahora" ──────────────────────────────────────────────────────
    now_line_html = ""
    if today in [d if isinstance(d, date) else d.date() for d in days]:
        now_h = datetime.now().hour + datetime.now().minute / 60
        if HOUR_START <= now_h < HOUR_END:
            now_top = int((now_h - HOUR_START) * SLOT_HEIGHT)
            now_line_html = f"""<div class="now-line" style="top:{now_top}px;">
                <div class="now-line-bar"></div>
                <div class="now-dot"></div>
            </div>"""

    # ── Columnas de días ──────────────────────────────────────────────────────
    day_cols_html = ""
    done_buttons = []

    for d in days:
        ddate = d if isinstance(d, date) else d.date()
        is_today = ddate == today
        bg = "rgba(59,72,131,0.08)" if is_today else "transparent"

        events = _events_for_day(ddate)
        cards_html = ""

        for row in events:
            excel_row = row.get("_excel_row")
            if excel_row in done_rows:
                continue

            task     = clean(get_from_row(row, ["task"], "Task"))
            name_ev  = clean(get_from_row(row, ["name"], "—"))
            channel  = clean(get_from_row(row, ["channel"], ""))
            priority = clean(get_from_row(row, ["priority"], "Mid"))
            hour_val = row.get("_hour")
            is_overdue = ddate < today

            colors = _task_color(task, priority)
            time_raw = get_from_row(row, ["time"], "")
            time_str = parse_agenda_time(time_raw)

            short_name = (name_ev[:18] + "…") if len(name_ev) > 18 else name_ev
            short_task = (task[:22] + "…") if len(task) > 22 else task

            if hour_val is not None and HOUR_START <= hour_val < HOUR_END:
                top = int((hour_val - HOUR_START) * SLOT_HEIGHT)
                h_px = max(SLOT_HEIGHT - 4, 36)
                pos = f"top:{top}px;height:{h_px}px;"
            else:
                pos = "top:0px;height:44px;"

            overdue_border = "border-top:2px solid #FF4D2E;" if is_overdue else ""
            overdue_badge = "<span style='font-size:9px;font-weight:700;color:#FF4D2E;'>⚠ OVERDUE</span>" if is_overdue else ""
            ch_str = f" · {html.escape(channel)}" if channel else ""

            cards_html += f"""<div class="evt" style="{pos}background:{colors['bg']};border-left:3px solid {colors['border']};{overdue_border}"
                title="{html.escape(name_ev)} · {html.escape(task)}">
                <div class="evt-task" style="color:{colors['text']};">{html.escape(short_task)}</div>
                <div class="evt-name">{html.escape(short_name)}</div>
                <div class="evt-time">{html.escape(time_str)}{ch_str} {overdue_badge}</div>
            </div>"""

            done_buttons.append((excel_row, excel_row, name_ev, task))

        col_now = now_line_html if is_today else ""

        day_cols_html += f"""<div class="day-col" style="background:{bg};">
            {grid_lines_html}
            {col_now}
            {cards_html}
        </div>"""

    st_components.html(f"""
    <!DOCTYPE html>
    <html>
    <head><meta charset="utf-8">
    <style>
      * {{ box-sizing: border-box; margin: 0; padding: 0; }}
      body {{ background: transparent; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; }}
      .cal-grid {{
        display: grid;
        grid-template-columns: 52px repeat(7, 1fr);
        border: 1px solid rgba(255,255,255,0.95);
        border-radius: 10px;
        overflow: hidden;
      }}
      .hour-col {{
        position: relative;
        height: {total_height}px;
        background: rgba(27,63,139,0.02);
        border-right: 1px solid rgba(255,255,255,0.95);
      }}
      .hour-label {{
        position: absolute;
        left: 0; width: 48px;
        font-size: 10px;
        color: #6B7280;
        font-weight: 600;
        text-align: right;
        padding-right: 6px;
        line-height: 1;
        transform: translateY(-6px);
      }}
      .day-col {{
        position: relative;
        height: {total_height}px;
        border-right: 1px solid rgba(255,255,255,0.92);
        min-width: 0;
      }}
      .grid-line {{
        position: absolute;
        left: 0; right: 0;
        height: 1px;
      }}
      .now-line {{
        position: absolute;
        left: 0; right: 0;
        z-index: 10;
        pointer-events: none;
      }}
      .now-line-bar {{ height: 2px; background: #FF4D2E; opacity: .85; }}
      .now-dot {{
        position: absolute;
        top: -4px; left: -4px;
        width: 8px; height: 8px;
        border-radius: 50%;
        background: #FF4D2E;
      }}
      .evt {{
        position: absolute;
        left: 2px; right: 2px;
        border-radius: 0 6px 6px 0;
        padding: 4px 6px;
        overflow: hidden;
        cursor: default;
        z-index: 5;
      }}
      .evt-task {{
        font-size: 10px; font-weight: 700;
        white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
      }}
      .evt-name {{
        font-size: 10px; color: rgba(255,255,255,.85);
        white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
      }}
      .evt-time {{
        font-size: 9px; color: rgba(255,255,255,.5);
      }}
    </style>
    </head>
    <body>
    <div class="cal-grid">
      <div class="hour-col">
        {hour_labels_html}
      </div>
      {day_cols_html}
    </div>
    </body>
    </html>
    """, height=total_height + 24, scrolling=False)

    # ── Botones Done (Streamlit nativo, fuera del HTML) ───────────────────────
    if done_buttons:
        st.markdown(
            "<div style='font-size:11px;font-weight:700;letter-spacing:.06em;color:#6B7280;margin:8px 0 6px 0;'>MARK AS DONE</div>",
            unsafe_allow_html=True
        )
        btn_cols = st.columns(min(len(done_buttons), 4))
        for i, (excel_row, idx, name_ev, task) in enumerate(done_buttons):
            with btn_cols[i % len(btn_cols)]:
                label = f"✓ {name_ev[:16]}…" if len(name_ev) > 16 else f"✓ {name_ev}"
                if st.button(label, key=f"done_{excel_row}_{idx}", use_container_width=True, help=task):
                    ok, msg = mark_agenda_row_done(EXCEL_FILE, excel_row)
                    if ok:
                        done_rows.add(excel_row)
                        st.success(f"✅ {name_ev}")
                    else:
                        st.error(msg)

    # ── Eventos sin fecha asignada ────────────────────────────────────────────
    if not active_agenda.empty:
        undated = active_agenda[active_agenda["_parsed_date"].isna()]
        if not undated.empty:
            st.markdown("---")
            st.markdown("<div style='font-size:11px;font-weight:700;letter-spacing:.06em;color:#6B7280;margin-bottom:10px;'>WITHOUT DATE</div>", unsafe_allow_html=True)
            for idx, row in undated.iterrows():
                excel_row = row.get("_excel_row", None)
                if excel_row in done_rows:
                    continue
                task  = clean(get_from_row(row, ["task"], "Task"))
                name_ev = clean(get_from_row(row, ["name"], "—"))
                priority = clean(get_from_row(row, ["priority"], "Mid"))
                colors = _task_color(task, priority)
                st.markdown(f"""
                <div style="background:{colors['bg']};border-left:3px solid {colors['border']};border-radius:0 8px 8px 0;padding:8px 10px;margin-bottom:6px;">
                    <div style="font-size:11px;font-weight:700;color:{colors['text']};">{task}</div>
                    <div style="font-size:12px;color:#1A1A2E;">{name_ev}</div>
                </div>
                """, unsafe_allow_html=True)
                if st.button("✓ Done", key=f"done_nd_{excel_row}_{idx}", use_container_width=True):
                    ok, msg = mark_agenda_row_done(EXCEL_FILE, excel_row)
                    if ok:
                        done_rows.add(excel_row)
                        st.success("✅")
                    else:
                        st.error(msg)


# =========================
# BRAND UPDATE
# =========================

def page_brand_update():
    render_header("Brand Update", "Search and edit brand information")

    df = load_growth_data()

    if df.empty:
        st.error("Growth OS sheet not found.")
        return

    brand_id_input = st.text_input("Search by Brand ID")
    brand_id = normalize_brand_id(brand_id_input)

    if not brand_id_input:
        st.info("Type a Brand ID to open the editable brand form.")
        return

    id_col = get_id_column_name(df)
    if not id_col:
        st.error("ID column not found in Growth OS sheet.")
        st.caption("Detected columns: " + ", ".join([str(c) for c in df.columns[:25]]))
        return

    result = df[df[id_col].apply(normalize_brand_id) == brand_id]

    if result.empty:
        st.error("Brand not found.")
        st.caption(f"Typed ID: {brand_id}")
        return

    row = result.iloc[0]

    name_current = clean(get_from_row(row, ["name", "brand name", "restaurant name"], ""))
    ltor_current = clean(get_from_row(row, ["ltor tier", "ltor"], ""))
    churn_current = get_churn_status(brand_id)  # Source of truth: Current Churn sheet
    gmv_ars_current = to_number(get_from_row(row, ["last gmv ars", "gmv ars"], 0))
    aov_ars_current = to_number(get_from_row(row, ["last aov ars", "aov ars"], 0))

    st.markdown(f"""
    <div class="update-card">
        <div style="display:flex; align-items:center; justify-content:space-between;">
            <div>
                <div class="small-muted">Brand found</div>
                <div style="font-size:28px; font-weight:800; color:{COLORS['text']};">{name_current}</div>
                <div class="small-muted">ID: {brand_id}</div>
            </div>
            <div class="priority-pill low">Editable Mode</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""
        <div class="update-card">
            <div class="small-muted">LAST GMV ARS</div>
            <div class="metric-value ars">{fmt_ars(gmv_ars_current)}</div>
        </div>
        """, unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <div class="update-card">
            <div class="small-muted">AOV ARS</div>
            <div class="metric-value ars">{fmt_ars(aov_ars_current)}</div>
        </div>
        """, unsafe_allow_html=True)
    with c3:
        st.markdown(f"""
        <div class="update-card">
            <div class="small-muted">BRAND LTOR</div>
            <div class="metric-value cop">{ltor_current}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("### Editable Fields")

    with st.form("brand_update_form"):
        sec1, sec2 = st.columns(2)

        with sec1:
            st.markdown('<div class="update-card"><div class="update-title">Brand Identity</div>', unsafe_allow_html=True)
            name_new = st.text_input("Brand Name", value=name_current)
            ltor_options = ["Prioritized 🔥", "Non Prioritized ❄️"]
            ltor_new = st.selectbox("LTOR Tier", ltor_options, index=option_index(ltor_options, ltor_current))
            churn_options = ["✅On", "⚠️W1", "🚨W2", "🆘W3", "😴Off", "☠️Off"]
            churn_new = st.selectbox("Churn Status", churn_options, index=option_index(churn_options, churn_current))
            st.markdown("</div>", unsafe_allow_html=True)

        with sec2:
            st.markdown('<div class="update-card"><div class="update-title">Commercial Performance</div>', unsafe_allow_html=True)
            gmv_ars_new = st.number_input("Last GMV ARS", value=float(gmv_ars_current), step=1000.0)
            aov_ars_new = st.number_input("AOV ARS", value=float(aov_ars_current), step=100.0)
            st.caption("USD fields are protected because they are calculated by formula.")
            st.markdown("</div>", unsafe_allow_html=True)

        ads_col, md_col = st.columns(2)

        with ads_col:
            st.markdown('<div class="update-card"><div class="update-title">ADS</div>', unsafe_allow_html=True)
            ads_current = clean(get_from_row(row, ["ads"], "Inactive 💤"))
            ads_options = ["Active 🚀", "Inactive 💤", "OFF 😴"]
            ads_new = st.selectbox("Ads Status", ads_options, index=option_index(ads_options, ads_current))

            ads_bookings_current = to_number(get_from_row(row, ["ads bookings", "ad bookings"], 0))
            ads_bookings_new = st.number_input("Ads Bookings ARS", value=float(ads_bookings_current), step=1000.0)

            ads_roi_current = clean(get_from_row(row, ["ads roi", "ad roi"], ""))
            ads_roi_new = st.text_input("Ads ROI", value=ads_roi_current)
            st.caption("Ads USD is protected because it is calculated automatically.")
            st.markdown("</div>", unsafe_allow_html=True)

        with md_col:
            st.markdown('<div class="update-card"><div class="update-title">Merchant Development (MD)</div>', unsafe_allow_html=True)
            md_current = clean(get_from_row(row, ["md", "md status"], "Inactive 💤"))
            md_options = ["Active 🚀", "Inactive 💤", "OFF 😴"]
            md_new = st.selectbox("MD Status", md_options, index=option_index(md_options, md_current))

            md_discount_current = normalize_markdown_discount(get_from_row(row, ["md discount", "md promo", "markdown discount", "markdown promo", "md bookings", "md bookings ars", "md booking"], ""))
            md_bookings_new = st.text_input("MD Discount / Promo", value=md_discount_current, placeholder="Ej: 15%, 20%, 2x1")

            md_roi_current = clean(get_from_row(row, ["md roi"], ""))
            md_roi_new = st.text_input("MD ROI", value=md_roi_current)
            st.caption("MD USD is protected because it is calculated automatically.")
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="update-card"><div class="update-title">Account & Ownership</div>', unsafe_allow_html=True)
        acc1, acc2, acc3 = st.columns(3)
        with acc1:
            manager_new = st.text_input("Manager / Editor", value=clean(get_from_row(row, ["manager", "restaurant manager", "account manager"], "")))
        with acc2:
            assistant_new = st.text_input("Assistant", value=clean(get_from_row(row, ["assistant"], "")))
        with acc3:
            email_new = st.text_input("Mail", value=clean(get_from_row(row, ["email", "mail"], "")))
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="update-card"><div class="update-title">Comments / Notes</div>', unsafe_allow_html=True)
        notes_current = clean(get_from_row(row, ["comments", "comment"], ""))
        notes_new = st.text_area("Notes", value="" if notes_current == "-" else notes_current, height=120)
        st.caption("Notes are saved in the brand record and mirrored to Weekly Calendar when the Brand ID exists in Agenda.")
        st.markdown("</div>", unsafe_allow_html=True)

        submitted = st.form_submit_button("Save Changes")

    # ── Badge: último guardado exitoso ────────────────────────────────────────
    _last_saved_key = f"brand_update_last_saved_{brand_id}"
    if st.session_state.get(_last_saved_key):
        _elapsed = (datetime.now() - st.session_state[_last_saved_key]).seconds
        if _elapsed < 60:
            _elapsed_label = "hace menos de 1 min"
        elif _elapsed < 3600:
            _elapsed_label = f"hace {_elapsed // 60} min"
        else:
            _elapsed_label = st.session_state[_last_saved_key].strftime("%H:%M")
        st.markdown(
            f'''<div style="background:rgba(111,242,75,0.08);border:1px solid rgba(111,242,75,0.25);border-radius:8px;
            padding:8px 14px;font-size:12px;color:#7ED321;font-weight:600;margin-bottom:10px;">
            ✓ Último guardado exitoso: {_elapsed_label}</div>''',
            unsafe_allow_html=True,
        )

    if submitted:
        updates = {
            "name": name_new,
            "last_gmv_ars": float(gmv_ars_new),
            "last_aov_ars": float(aov_ars_new),
            "ltor": ltor_new,
            "churn": churn_new,
            "ads": ads_new,
            "ads_bookings": float(ads_bookings_new),
            "ads_roi": ads_roi_new,
            "md": md_new,
            "md_bookings": normalize_markdown_discount(md_bookings_new),
            "md_roi": md_roi_new,
            "manager": manager_new,
            "assistant": assistant_new,
            "email": email_new,
            "comments": notes_new,
        }

        # Capture old row BEFORE writing, for the changelog
        _old_row = row  # 'row' is the pandas Series loaded at top of this page

        ok, msg, updated, locked, missing, backup_path = update_brand_in_excel(brand_id, updates)

        if ok:
            # Write changelog (only changed fields)
            try:
                save_brand_changelog(brand_id, name_current, updates, _old_row)
            except Exception:
                pass
            # Persist badge timestamp in session state
            st.session_state[_last_saved_key] = datetime.now()
            st.success(msg)
            st.caption("📦 Backup guardándose en segundo plano.")
            if locked:
                st.warning("Some formula-protected fields were not updated: " + ", ".join(locked))
            if missing:
                st.caption("Some optional columns were not found and were skipped: " + ", ".join(missing))
            st.rerun()
        else:
            st.error(msg)



# =========================
# ROLE PLAY TRAINER — EVALUADOR LOCAL (sin API)
# =========================

def _evaluate_objection_response_locally(user_response, ideal_response, objection_text, lever):
    """
    Evalúa la respuesta del usuario a una objeción usando reglas de palabras clave
    100% locales — sin llamadas a ninguna API externa. Mide las mismas 4
    dimensiones que antes evaluaba un modelo externo:
      1. Anclaje en datos concretos (números, %, montos, benchmarks)
      2. Acción propuesta con fecha específica
      3. Manejo de la resistencia sin perder el pitch
      4. Cierre con próximo paso claro
    Devuelve el mismo formato de dict que el flujo anterior, para no romper
    el render ni el guardado de historial.
    """
    text = norm_text(user_response)
    ideal_norm = norm_text(ideal_response)

    # ── Dimensión 1: Anclaje en datos concretos ───────────────────────────────
    _data_patterns = [
        r"\d+\s*%", r"\$\s*\d+", r"\d+\s*(usd|ars|cop|pesos)",
        r"benchmark", r"promedio", r"categor[ií]a", r"top\s*\d", r"ranking",
    ]
    _data_hits = sum(1 for p in _data_patterns if re.search(p, text))
    score_datos = 5 if _data_hits >= 3 else 4 if _data_hits == 2 else 2 if _data_hits == 1 else 1

    # ── Dimensión 2: Acción propuesta con fecha específica ────────────────────
    _action_verbs = ["activamos", "arrancamos", "proponemos", "te propongo", "vamos a", "podemos"]
    _date_markers = [
        "hoy", "mañana", "esta semana", "el lunes", "el martes", "el miércoles",
        "el jueves", "el viernes", "la próxima semana", "este mes", "en 48 horas",
        "en 24 horas",
    ]
    _has_action = any(v in text for v in _action_verbs)
    _has_date = any(d in text for d in _date_markers)
    if _has_action and _has_date:
        score_accion = 5
    elif _has_action or _has_date:
        score_accion = 3
    else:
        score_accion = 1

    # ── Dimensión 3: Manejo de resistencia sin perder el pitch ────────────────
    _resistance_acknowledge = ["entiendo", "comprendo", "tiene sentido", "es válido", "te escucho"]
    _pitch_recovery = ["pero", "sin embargo", "aun así", "igual te cuento", "de todas formas", "lo que te propongo"]
    _gives_up = ["está bien", "no hay problema", "como quieras", "ok entonces no", "te dejo tranquilo"]
    _ack_hit = any(a in text for a in _resistance_acknowledge)
    _recovery_hit = any(r in text for r in _pitch_recovery)
    _giveup_hit = any(g in text for g in _gives_up) and not _recovery_hit
    if _giveup_hit:
        score_manejo = 1
    elif _ack_hit and _recovery_hit:
        score_manejo = 5
    elif _ack_hit or _recovery_hit:
        score_manejo = 3
    else:
        score_manejo = 2

    # ── Dimensión 4: Cierre con próximo paso claro ─────────────────────────────
    _close_patterns = [
        "¿te parece?", "¿lo hacemos?", "¿avanzamos?", "¿confirmamos?", "¿qué te parece?",
        "quedamos en", "te confirmo", "te escribo", "te llamo", "coordinamos",
    ]
    _close_hits = sum(1 for c in _close_patterns if c in text)
    score_cierre = 5 if _close_hits >= 2 else 4 if _close_hits == 1 else 1

    # ── Similitud textual con la respuesta ideal (señal de apoyo, no determinante) ──
    _ideal_words = set(ideal_norm.split())
    _user_words = set(text.split())
    _overlap = len(_ideal_words & _user_words) / len(_ideal_words) if _ideal_words else 0

    # ── Texto de feedback generado localmente ─────────────────────────────────
    _bien_parts = []
    if score_datos >= 4:
        _bien_parts.append("ancló la respuesta en datos concretos")
    if score_accion >= 4:
        _bien_parts.append("propuso una acción con fecha específica")
    if score_manejo >= 4:
        _bien_parts.append("manejó la resistencia sin perder el pitch")
    if score_cierre >= 4:
        _bien_parts.append("cerró con un próximo paso claro")
    que_hizo_bien = (
        "La respuesta " + ", ".join(_bien_parts) + "." if _bien_parts
        else "La respuesta tocó la objeción pero sin un punto fuerte claro todavía."
    )

    _falta_parts = []
    if score_datos < 4:
        _falta_parts.append("anclar en un dato concreto (%, monto, benchmark de categoría)")
    if score_accion < 4:
        _falta_parts.append("proponer una acción con fecha específica")
    if score_manejo < 4:
        _falta_parts.append("reconocer la objeción del aliado antes de retomar el pitch")
    if score_cierre < 4:
        _falta_parts.append("cerrar con una pregunta concreta que empuje al siguiente paso")
    que_falto = (
        "Para subir el puntaje: " + "; ".join(_falta_parts) + "." if _falta_parts
        else "Muy completa — esta respuesta está cerca de tu versión ideal."
    )

    # ── Frase sugerida: usa la respuesta ideal cargada por Sabas como referencia ──
    if _overlap < 0.3 and ideal_response.strip():
        _primera_frase_ideal = ideal_response.strip().split(".")[0].strip()
        frase_sugerida = _primera_frase_ideal if _primera_frase_ideal else ideal_response.strip()[:120]
    else:
        frase_sugerida = "Estás cerca de tu respuesta ideal — seguí anclando en datos y cerrando con fecha."

    return {
        "score_datos":  score_datos,
        "score_accion": score_accion,
        "score_manejo": score_manejo,
        "score_cierre": score_cierre,
        "que_hizo_bien": que_hizo_bien,
        "que_falto": que_falto,
        "frase_que_faltó": frase_sugerida,
    }


# =========================
# ROLE PLAY TRAINER
# =========================

def page_role_play_trainer():
    render_header("Role Play Trainer", f"Entrenamiento y práctica de manejo de objeciones · {FARMER_NAME}")

    # ── Helper: load/save objections CSV ─────────────────────────────────────
    def _load_objections():
        if not os.path.exists(ROLEPLAY_OBJECTIONS_FILE):
            return pd.DataFrame(columns=["objection_id", "datetime", "objection_text", "lever", "category", "ideal_response", "tags"])
        try:
            return pd.read_csv(ROLEPLAY_OBJECTIONS_FILE, dtype=str).fillna("")
        except Exception:
            return pd.DataFrame(columns=["objection_id", "datetime", "objection_text", "lever", "category", "ideal_response", "tags"])

    def _save_objection(row_dict):
        df_existing = _load_objections()
        new_row = pd.DataFrame([row_dict])
        combined = pd.concat([df_existing, new_row], ignore_index=True)
        combined.to_csv(ROLEPLAY_OBJECTIONS_FILE, index=False, encoding="utf-8-sig")

    def _delete_objection(obj_id):
        df_existing = _load_objections()
        df_existing = df_existing[df_existing["objection_id"] != obj_id]
        df_existing.to_csv(ROLEPLAY_OBJECTIONS_FILE, index=False, encoding="utf-8-sig")

    def _save_history(row_dict):
        if os.path.exists(ROLEPLAY_HISTORY_FILE):
            try:
                existing = pd.read_csv(ROLEPLAY_HISTORY_FILE)
                combined = pd.concat([existing, pd.DataFrame([row_dict])], ignore_index=True)
            except Exception:
                combined = pd.DataFrame([row_dict])
        else:
            combined = pd.DataFrame([row_dict])
        combined.to_csv(ROLEPLAY_HISTORY_FILE, index=False, encoding="utf-8-sig")

    # ── Dynamic category list from Detalle CABA ───────────────────────────────
    def _get_categories():
        detalle = load_detalle_caba()
        if not detalle.empty:
            cat_col = next((c for c in detalle.columns if "categor" in c), None)
            if cat_col:
                cats = sorted(detalle[cat_col].dropna().unique().tolist())
                return [str(c) for c in cats if str(c).strip()]
        return ["General", "Restaurantes", "Supermercados", "Farmacia", "Mascotas"]

    LEVERS = ["Ads", "MD", "Churn", "General"]

    tab1, tab2 = st.tabs(["🏋️ Entrenamiento", "🎯 Práctica"])

    # ── TAB 1: ENTRENAMIENTO ──────────────────────────────────────────────────
    with tab1:
        st.markdown(f"""
        <div class="update-card">
            <div class="update-title">➕ Cargar nueva objeción al banco</div>
        </div>
        """, unsafe_allow_html=True)

        categories = _get_categories()

        with st.form("add_objection_form"):
            objection_text = st.text_area("Texto exacto de la objeción", height=100,
                placeholder="Ej: No me interesa Ads, ya probé y no me funcionó...")
            col1, col2 = st.columns(2)
            with col1:
                lever = st.selectbox("Palanca", LEVERS)
            with col2:
                category = st.selectbox("Categoría", categories)
            ideal_response = st.text_area("Tu respuesta ideal", height=120,
                placeholder="Escribí la respuesta que darías cuando estás en tu mejor versión...")
            tags = st.text_input("Tags (opcional, separados por coma)", placeholder="Ej: precio, repartidores, experiencia previa")
            submitted = st.form_submit_button("💾 Guardar en banco de objeciones")

        if submitted:
            if not objection_text.strip() or not ideal_response.strip():
                st.warning("La objeción y la respuesta ideal son obligatorias.")
            else:
                _save_objection({
                    "objection_id": str(uuid.uuid4()),
                    "datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "objection_text": objection_text.strip(),
                    "lever": lever,
                    "category": category,
                    "ideal_response": ideal_response.strip(),
                    "tags": tags.strip(),
                })
                st.success("✅ Objeción guardada en el banco.")
                st.rerun()

        # ── Banco actual ──────────────────────────────────────────────────────
        obj_df = _load_objections()
        _n_auto = int((obj_df.get("tags", pd.Series(dtype=str)) == "auto-detectada").sum()) if not obj_df.empty else 0
        with st.expander(f"📚 Banco de objeciones ({len(obj_df)} registros · {_n_auto} auto-detectadas)", expanded=False):
            if obj_df.empty:
                st.info("El banco está vacío. Cargá tu primera objeción arriba.")
            else:
                if _n_auto > 0:
                    st.caption(f"🤖 {_n_auto} objeción(es) detectadas automáticamente desde transcripciones de llamadas — revisalas y completá la respuesta ideal si quedó vacía.")
                st.dataframe(obj_df[["datetime", "lever", "category", "objection_text", "ideal_response", "tags"]],
                             use_container_width=True, hide_index=True)

                _pending_ideal = obj_df[(obj_df["tags"] == "auto-detectada") & (obj_df["ideal_response"].str.strip() == "")]
                if not _pending_ideal.empty:
                    st.markdown("**Completar respuesta ideal de una objeción auto-detectada:**")
                    _pending_options = {f"{r['objection_text'][:60]}...": r["objection_id"] for _, r in _pending_ideal.iterrows()}
                    _chosen_label = st.selectbox("Objeción pendiente", list(_pending_options.keys()), key="rp_pending_ideal_select")
                    _ideal_fill = st.text_area("Respuesta ideal", height=100, key="rp_pending_ideal_text")
                    if st.button("💾 Guardar respuesta ideal", key="rp_save_pending_ideal") and _ideal_fill.strip():
                        _target_id = _pending_options[_chosen_label]
                        obj_df.loc[obj_df["objection_id"] == _target_id, "ideal_response"] = _ideal_fill.strip()
                        obj_df.to_csv(ROLEPLAY_OBJECTIONS_FILE, index=False, encoding="utf-8-sig")
                        st.success("Respuesta ideal guardada.")
                        st.rerun()

                del_id = st.text_input("ID de objeción a eliminar (pegá el objection_id)", key="del_obj_id")
                if st.button("🗑️ Eliminar objeción") and del_id.strip():
                    _delete_objection(del_id.strip())
                    st.success("Objeción eliminada.")
                    st.rerun()

    # ── TAB 2: PRÁCTICA ───────────────────────────────────────────────────────
    with tab2:
        st.markdown(f"""
        <div class="update-card">
            <div class="update-title">🎯 Práctica de objeciones</div>
        </div>
        """, unsafe_allow_html=True)

        categories = _get_categories()
        obj_df2 = _load_objections()

        f1, f2 = st.columns(2)
        with f1:
            filter_lever = st.selectbox("Filtrar por palanca", ["Todas"] + LEVERS, key="prac_lever")
        with f2:
            filter_cat = st.selectbox("Filtrar por categoría", ["Todas"] + categories, key="prac_cat")

        filtered = obj_df2.copy()
        if filter_lever != "Todas":
            filtered = filtered[filtered["lever"] == filter_lever]
        if filter_cat != "Todas":
            filtered = filtered[filtered["category"] == filter_cat]

        if st.button("🎲 Cargar objeción aleatoria") and not filtered.empty:
            chosen = filtered.sample(1).iloc[0].to_dict()
            st.session_state["rp_current_obj"] = chosen

        current_obj = st.session_state.get("rp_current_obj")

        if current_obj:
            st.markdown(f"""
            <div class="update-card" style="border-left:4px solid {PALETTE['neon_tangerine']};margin-top:16px;">
                <div class="small-muted">PALANCA: {html.escape(current_obj.get('lever',''))} · CATEGORÍA: {html.escape(current_obj.get('category',''))}</div>
                <div style="font-size:20px;font-weight:700;margin-top:8px;color:{PALETTE['neon_tangerine']};">"{html.escape(current_obj.get('objection_text',''))}"</div>
            </div>
            """, unsafe_allow_html=True)

            user_response = st.text_area("Tu respuesta", height=140, placeholder="Escribí tu respuesta a esta objeción...", key="rp_user_resp")

            if st.button("🤖 Evaluar respuesta"):
                if not user_response.strip():
                    st.warning("Escribí una respuesta antes de evaluar.")
                else:
                    result = _evaluate_objection_response_locally(
                        user_response,
                        current_obj.get("ideal_response", ""),
                        current_obj.get("objection_text", ""),
                        current_obj.get("lever", ""),
                    )
                    st.session_state["rp_eval_result"] = result
                    st.session_state["rp_eval_response"] = user_response.strip()

            eval_result = st.session_state.get("rp_eval_result")
            if eval_result:
                def _score_color(s):
                    s = int(s)
                    if s >= 4:
                        return PALETTE["laser_green"]
                    if s == 3:
                        return PALETTE["neon_tangerine"]
                    return PALETTE["tangerine_dark"]

                def _score_bar(s):
                    s = int(s)
                    pct = s / 5 * 100
                    color = _score_color(s)
                    return f'<div style="background:rgba(255,255,255,.1);border-radius:6px;height:10px;margin-top:4px;"><div style="background:{color};width:{pct}%;height:10px;border-radius:6px;"></div></div>'

                st.markdown(f"""
                <div class="update-card" style="margin-top:20px;">
                    <div class="update-title">📊 Resultado de la evaluación</div>
                    <div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:12px;margin-top:14px;">
                        <div>
                            <div class="small-muted">Datos concretos</div>
                            <div style="font-size:28px;font-weight:800;color:{_score_color(eval_result['score_datos'])};">{eval_result['score_datos']}/5</div>
                            {_score_bar(eval_result['score_datos'])}
                        </div>
                        <div>
                            <div class="small-muted">Acción con fecha</div>
                            <div style="font-size:28px;font-weight:800;color:{_score_color(eval_result['score_accion'])};">{eval_result['score_accion']}/5</div>
                            {_score_bar(eval_result['score_accion'])}
                        </div>
                        <div>
                            <div class="small-muted">Manejo resistencia</div>
                            <div style="font-size:28px;font-weight:800;color:{_score_color(eval_result['score_manejo'])};">{eval_result['score_manejo']}/5</div>
                            {_score_bar(eval_result['score_manejo'])}
                        </div>
                        <div>
                            <div class="small-muted">Cierre con paso</div>
                            <div style="font-size:28px;font-weight:800;color:{_score_color(eval_result['score_cierre'])};">{eval_result['score_cierre']}/5</div>
                            {_score_bar(eval_result['score_cierre'])}
                        </div>
                    </div>
                    <hr style="border-color:rgba(255,255,255,.15);margin:16px 0;">
                    <div style="margin-bottom:10px;"><span class="small-muted">✅ Qué hizo bien:</span><br>{html.escape(str(eval_result.get('que_hizo_bien','')))} </div>
                    <div style="margin-bottom:10px;"><span class="small-muted">⚠️ Qué faltó:</span><br>{html.escape(str(eval_result.get('que_falto','')))} </div>
                    <div style="background:{PALETTE['space_indigo']};border-radius:8px;padding:12px;margin-top:8px;">
                        <div class="small-muted">💬 Frase que te faltó decir:</div>
                        <div style="font-style:italic;margin-top:4px;">"{html.escape(str(eval_result.get('frase_que_faltó','')))} "</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                if st.button("💾 Guardar intento"):
                    _save_history({
                        "datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "objection_id": current_obj.get("objection_id", ""),
                        "objection_text": current_obj.get("objection_text", ""),
                        "lever": current_obj.get("lever", ""),
                        "category": current_obj.get("category", ""),
                        "user_response": st.session_state.get("rp_eval_response", ""),
                        "score_datos": eval_result["score_datos"],
                        "score_accion": eval_result["score_accion"],
                        "score_manejo": eval_result["score_manejo"],
                        "score_cierre": eval_result["score_cierre"],
                        "que_hizo_bien": eval_result.get("que_hizo_bien", ""),
                        "que_falto": eval_result.get("que_falto", ""),
                        "frase_faltante": eval_result.get("frase_que_faltó", ""),
                    })
                    st.success("✅ Intento guardado en el historial.")
        elif obj_df2.empty:
            st.info("No hay objeciones en el banco. Cargá algunas en la pestaña Entrenamiento.")
        else:
            st.info("Aplicá filtros y hacé click en 'Cargar objeción aleatoria' para empezar a practicar.")


# =========================
# CACHE WARM-UP (silencioso — evita spinners en primera carga)
# =========================
if "_cache_warmed" not in st.session_state:
    load_cvr_data()
    load_detalle_caba()
    st.session_state["_cache_warmed"] = True

# =========================
# ROUTER
# =========================

if page == "Management Dashboard":
    page_management_dashboard()
elif page == "Opportunity List":
    page_opportunity_list()
elif page == "Follow-Up List":
    page_follow_up_list()
elif page == "Brand Finder":
    page_brand_finder()
elif page == "Day Queue":
    page_day_queue()
elif page == "Pareto Hub":
    page_pareto_hub()
elif page == "Acquisition Tracker":
    page_acquisition_tracker()
elif page == "Campaign Weekly Tracker":
    page_campaign_weekly_tracker()
elif page == "Weekly Calendar":
    page_weekly_calendar()
elif page == "Brand Update":
    page_brand_update()
elif page == "Earnings Calculator":
    page_earnings_calculator()
elif page == "Productivity HeatMap":
    page_productivity_heatmap()
elif page == "Call Quality Trainer":
    page_call_quality_trainer()
elif page == "Role Play Trainer":
    page_role_play_trainer()
